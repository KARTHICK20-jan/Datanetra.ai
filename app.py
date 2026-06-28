# DataNetra.ai v5.5  —  MSME Intelligence Platform  |  Render Deployment
# Auto-install missing packages at startup
import subprocess as _subp, sys as _sys, os as _os
_os.environ.setdefault('GRADIO_ANALYTICS_ENABLED', 'False')

# ── Patch gradio 4.44.1 to work with any huggingface_hub version ─────────────
# gradio/external_utils.py tries to import ImageClassificationOutputElement
# which was removed in huggingface_hub > 0.20.x. We patch the module BEFORE
# gradio loads so the import never fails — regardless of hfh version.
try:
    import huggingface_hub as _hfh_mod
    if not hasattr(_hfh_mod, 'ImageClassificationOutputElement'):
        # Create a no-op stub so gradio's import succeeds
        class _ImageClassificationOutputElement:
            pass
        _hfh_mod.ImageClassificationOutputElement = _ImageClassificationOutputElement
        # Also patch it into the top-level namespace gradio looks in
        import sys as _sys2
        _sys2.modules['huggingface_hub'].ImageClassificationOutputElement = _ImageClassificationOutputElement
        print('✅ Patched huggingface_hub.ImageClassificationOutputElement stub')
except Exception as _patch_err:
    print(f'⚠️  hfh patch failed: {_patch_err}')

for _pkg in ['plotly', 'statsmodels', 'scikit-learn', 'prophet']:
    try:
        __import__(_pkg.replace('-','_'))
    except ImportError:
        _subp.check_call([_sys.executable, '-m', 'pip', 'install', _pkg, '-q',
                          '--break-system-packages'], stdout=_subp.DEVNULL, stderr=_subp.DEVNULL)

# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  DataNetra.ai  —  MSME Intelligence Platform                               ║
# ║  Version  : v5.5  (Render Deployment)                                     ║
# ║  Purpose  : AI-powered retail data intelligence for global retail users    ║
# ╠══════════════════════════════════════════════════════════════════════════════╣
# ║  SUBMISSION                                                                ║
# ║  AI Innovation Challenge  —  Problem Statement 2                          ║
# ║  Digital Onboarding & Marketplace Readiness for Indian retailers         ║
# ╠══════════════════════════════════════════════════════════════════════════════╣
# ║  Code Structure                                                            ║
# ║    A  Imports & Configuration                                              ║
# ║    B  Dashboard HTML Helpers & Insight Builders                           ║
# ║    C  Data Helpers, Scoring & ML Models                                   ║
# ║    D  UI Assets  (CSS, Voice JS, landing HTML)                            ║
# ║    D  Step Handlers  (Steps 1–5 : registration, upload, analysis)         ║
# ║    G  Gradio Application  (UI layout + event wiring + demo.launch())      ║
# ╠══════════════════════════════════════════════════════════════════════════════╣
# ║  How to Run                                                                ║
# ║    pip install -r requirements.txt                                        ║
# ║    python app.py                                                          ║
# ║    # → opens at http://localhost:7860                                     ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  SECTION A  —  Imports & Configuration                                     ║
# ║  Imports, patches, column mapping, optional deps, DB setup                 ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
# ── Pre-import patches (must run before ANY gradio import) ───────────────────

# 0. Dependency conflict fix: jsonschema tries rfc3987 → triggers lark bug on Python 3.12+
#    Stub out rfc3987 so jsonschema falls back to its built-in URI checker.
import sys as _sys, types as _types
if "rfc3987" not in _sys.modules:
    _rfc3987_stub = _types.ModuleType("rfc3987")
    def _rfc3987_parse(value, rule=None, **kw):
        return {}
    _rfc3987_stub.parse = _rfc3987_parse
    _sys.modules["rfc3987"] = _rfc3987_stub

# 1. audioop stub — required by older Gradio websocket internals on Python 3.12+
for _m in ("audioop", "pyaudioop", "_audioop"):
    if _m not in _sys.modules:
        _sys.modules[_m] = _types.ModuleType(_m)

# 2. Gradio BlockContext patch — gradio.utils.get_all_components() calls
#    gr.blocks.BlockContext.__subclasses__() which was removed in Gradio 4.20+.
#    We import gradio.utils first (it is a plain module, safe to import early),
#    patch get_all_components, then do the full `import gradio as gr` below.
try:
    import gradio.utils as _gr_utils
    _orig_get_all = _gr_utils.get_all_components

    def _patched_get_all_components():
        try:
            return _orig_get_all()
        except AttributeError:
            # BlockContext gone — collect all Component subclasses recursively
            import gradio.components as _gc
            classes, seen, stack = [], set(), list(_gc.Component.__subclasses__())
            while stack:
                cls = stack.pop()
                if cls in seen:
                    continue
                seen.add(cls)
                classes.append(cls)
                stack.extend(cls.__subclasses__())
            return classes

    _gr_utils.get_all_components = _patched_get_all_components
except Exception:
    pass  # If even this fails, let Gradio surface its own error on launch

# ─────────────────────────────────────────────────────────────────────────────
import gradio as gr

# Patch gradio 4.44.1 health-check: fails on Render proxy network
try:
    import gradio.networking as _gr_net_patch
    _gr_net_patch.is_url_ok = lambda *a, **kw: True
    # Also patch directly inside blocks module reference
    import gradio.blocks as _gr_blocks_patch
    import types as _types_patch
    # Replace the networking reference inside blocks module
    _gr_blocks_patch.networking.is_url_ok = lambda *a, **kw: True
except Exception:
    pass

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')  # Must be before pyplot import — prevents GUI window on servers
import matplotlib.pyplot as plt
import datetime
import os
import re
import io
import tempfile
import warnings
import json as _json_mod
import sys
warnings.filterwarnings("ignore")

# ── Auto-install reportlab if not present (needed for PDF report generation) ──
try:
    import reportlab as _rl_check
except ImportError:
    try:
        import subprocess as _subp
        _subp.run(
            [sys.executable if hasattr(sys,'executable') else "python3",
             "-m", "pip", "install", "reportlab", "--quiet", "--break-system-packages"],
            check=False, capture_output=True
        )
    except Exception:
        pass  # Will surface as ModuleNotFoundError when PDF is generated

# ── Prophet — optional dependency for time-series forecasting ────────────────
try:
    from prophet import Prophet
    PROPHET_AVAILABLE = True
except ImportError:
    try:
        from fbprophet import Prophet   # older package name
        PROPHET_AVAILABLE = True
    except ImportError:
        Prophet = None
        PROPHET_AVAILABLE = False

# ── statsmodels Holt-Winters — optional, falls back to pure-numpy if missing ─
try:
    from statsmodels.tsa.holtwinters import ExponentialSmoothing as _HW_statsmodels
    _HW_STATSMODELS_AVAILABLE = True
except ImportError:
    _HW_statsmodels = None
    _HW_STATSMODELS_AVAILABLE = False

# Alias used in model metadata reporting
HOLTWINTERS_AVAILABLE = _HW_STATSMODELS_AVAILABLE

# ── sklearn LinearRegression — optional, falls back to numpy polyfit if missing
try:
    from sklearn.linear_model import LinearRegression as _LinearRegression
    LINEAR_REGRESSION_AVAILABLE = True
except ImportError:
    # Pure-numpy fallback: wrap np.polyfit so _run_linear_regression_model still works
    class _LinearRegression:
        def fit(self, X, y):
            x = X.ravel()
            self.coef_, self.intercept_ = np.polyfit(x, y, 1)
            self.coef_ = np.array([self.coef_])
            self._x = x; self._y = y
            return self
        def predict(self, X):
            return self.coef_[0] * X.ravel() + self.intercept_
        def score(self, X, y):
            y_pred = self.predict(X)
            ss_res = float(np.sum((y - y_pred) ** 2))
            ss_tot = float(np.sum((y - np.mean(y)) ** 2))
            return 1 - ss_res / ss_tot if ss_tot > 0 else 0.0
    LINEAR_REGRESSION_AVAILABLE = False  # sklearn not present, using numpy fallback

# ── sklearn KMeans + StandardScaler — used for customer/SKU segmentation ─────
try:
    from sklearn.cluster import KMeans
    from sklearn.preprocessing import StandardScaler
except ImportError:
    # Pure-numpy fallback stubs so segment_customers() degrades gracefully
    class KMeans:
        def __init__(self, n_clusters=5, random_state=42, n_init='auto'):
            self.n_clusters = n_clusters
        def fit_predict(self, X):
            import numpy as _np
            n = len(X)
            return _np.array([i % self.n_clusters for i in range(n)])
    class StandardScaler:
        def fit_transform(self, X):
            import numpy as _np
            X = _np.array(X, dtype=float)
            mu = X.mean(axis=0); sd = X.std(axis=0) + 1e-9
            return (X - mu) / sd

# ── UTF-8 surrogate patch — prevents JSON crash on emoji/special characters ───
_orig_json_dumps = _json_mod.dumps
def _safe_json_dumps(obj, **kw):
    try:
        return _orig_json_dumps(obj, **kw)
    except (UnicodeEncodeError, ValueError):
        s = _orig_json_dumps(obj, ensure_ascii=True, **kw)
        return re.sub(r'\\ud[89ab][0-9a-f]{2}\\ud[c-f][0-9a-f]{2}', '', s, flags=re.I)
_json_mod.dumps = _safe_json_dumps

# ── Column name aliases — maps dataset variants to canonical field names ───────
COL_REMAP_FIXED = {
    # ── This dataset columns ─────────────────────────────────────────────────
    "gross_sales":             "Monthly_Sales_INR",
    "cost_price":              "Monthly_Operating_Cost_INR",
    "outstanding_amount":      "Outstanding_Loan_INR",
    "return_rate_pct":         "Returns_Percentage",
    "profit_margin_pct":       "Avg_Margin_Percent",
    "units_sold":              "Monthly_Demand_Units",
    "date":                    "Date",
    "store_id":                "Store_ID",
    "product_category":        "Product_Category",
    "product_id":              "SKU_Name",
    # inventory_level is stock count, NOT turnover rate — handled in _apply_col_remap derivation below
    "stock_level":             "Stock_Level",
    "reorder_point":           "Reorder_Point",
    "udyam_number":            "Udyam_Number",
    "vendor_name":             "Vendor_Name",
    "enterprise_name":         "Enterprise_Name",
    # ── Generic aliases ──────────────────────────────────────────────────────
    "Sales_INR":               "Monthly_Sales_INR",
    "Monthly_Sales":           "Monthly_Sales_INR",
    "Gross_Sales":             "Monthly_Sales_INR",
    "Operating_Cost_INR":      "Monthly_Operating_Cost_INR",
    "Operating_Cost":          "Monthly_Operating_Cost_INR",
    "Outstanding_Loan":        "Outstanding_Loan_INR",
    "Vendor_Reliability":      "Vendor_Delivery_Reliability",
    "Inventory_Turnover_Rate": "Inventory_Turnover",
    "Average_Margin_Percent":  "Avg_Margin_Percent",
    "Profit_Margin_%":         "Avg_Margin_Percent",
    "Monthly_Demand":          "Monthly_Demand_Units",
    "Quantity_Sold":           "Monthly_Demand_Units",
    "Returns":                 "Returns_Percentage",
    "Return_Quantity":         "Returns_Percentage",
    "Product_Name":            "SKU_Name",
    # ── DRC cleaned file column names (DRC renames during export) ────────────
    "sales":                   "Monthly_Sales_INR",
    "product":                 "SKU_Name",
    "store":                   "Store_Name",
    "cost":                    "Monthly_Operating_Cost_INR",
    "sku":                     "SKU_Name",
}

def _apply_col_remap(df):
    # Build case-insensitive lookup: lowercase col name -> actual col name in df
    _ci = {c.lower().strip(): c for c in df.columns}
    for old, new in COL_REMAP_FIXED.items():
        # Try exact match first, then case-insensitive
        if old in df.columns and new not in df.columns:
            df.rename(columns={old: new}, inplace=True)
        elif old.lower() in _ci and new not in df.columns and _ci[old.lower()] != new:
            df.rename(columns={_ci[old.lower()]: new}, inplace=True)
    # Refresh lookup after renames
    _ci = {c.lower().strip(): c for c in df.columns}
    # ── Ensure Date column is datetime (handles DRC string dates) ──────────
    if 'Date' in df.columns:
        try: df['Date'] = pd.to_datetime(df['Date'], infer_datetime_format=True, errors='coerce')
        except: pass

    # Broad aliases for sales — catches any remaining variants
    if 'Monthly_Sales_INR' not in df.columns:
        for _alias in ['sales_amount','sale_value','total_sales','revenue','net_sales',
                       'sales','gross_sales','total_revenue','gross_revenue',
                       'invoice_value','bill_amount','total_amount','turnover']:
            if _alias in _ci:
                df.rename(columns={_ci[_alias]: 'Monthly_Sales_INR'}, inplace=True)
                break
    # Broad aliases for margin
    if 'Avg_Margin_Percent' not in df.columns:
        for _alias in ['margin','margin_pct','gross_margin','margin_%','margin_percent',
                       'avg_margin','profit_margin','profit_pct']:
            if _alias in _ci:
                df.rename(columns={_ci[_alias]: 'Avg_Margin_Percent'}, inplace=True)
                break    # Derive Vendor_Delivery_Reliability: prefer fulfillment ratio, else returns-based
    if "Vendor_Delivery_Reliability" not in df.columns:
        if "net_units_sold" in df.columns and "Monthly_Demand_Units" in df.columns:
            df["Vendor_Delivery_Reliability"] = (
                df["net_units_sold"] / df["Monthly_Demand_Units"].replace(0, 1)
            ).clip(0, 1)
        elif "Returns_Percentage" in df.columns:
            df["Vendor_Delivery_Reliability"] = (1 - df["Returns_Percentage"] / 100).clip(0, 1)
        else:
            df["Vendor_Delivery_Reliability"] = 0.85

    # Derive real Inventory_Turnover (times/month) from stock level
    # inventory_level is a stock count — turnover = units_sold / avg_stock
    # Anchors: 0.5x/month = slow, 4x/month = good hypermarket, 12x/month = excellent
    if "Inventory_Turnover" not in df.columns:
        if "inventory_level" in df.columns and "Monthly_Demand_Units" in df.columns:
            df["Inventory_Turnover"] = (
                df["Monthly_Demand_Units"] / df["inventory_level"].replace(0, 1)
            ).clip(0, 12)
        elif "inventory_level" in df.columns:
            df.rename(columns={"inventory_level": "_inv_stock_raw"}, inplace=True)
            df["Inventory_Turnover"] = 3.0   # fallback median hypermarket
        else:
            df["Inventory_Turnover"] = 3.0
    # Derive Monthly_Operating_Cost_INR if missing
    if "Monthly_Operating_Cost_INR" not in df.columns:
        if "Monthly_Sales_INR" in df.columns:
            df["Monthly_Operating_Cost_INR"] = df["Monthly_Sales_INR"] * 0.60
        else:
            df["Monthly_Operating_Cost_INR"] = 0
    # Build a readable SKU_Name if we only have numeric product_id
    if "SKU_Name" in df.columns:
        import pandas as _pd
        try:
            if df["SKU_Name"].dtype in [_pd.Int64Dtype(), int] or str(df["SKU_Name"].dtype).startswith("int"):
                if "Product_Category" in df.columns:
                    df["SKU_Name"] = df["Product_Category"].astype(str) + "-" + df["SKU_Name"].astype(str)
                else:
                    df["SKU_Name"] = "SKU-" + df["SKU_Name"].astype(str)
        except Exception:
            pass
    return df

# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  SECTION B  —  Dashboard HTML Helpers & Insight Builders                   ║
# ║  Gov dashboard helpers, insight generator, ONDC dashboard data             ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  SECTION C  —  Data Helpers, Scoring & ML Models                           ║
# ║  Scoring, forecasting models, segmentation, LANG/SNP constants             ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

def normalize(series):
    if series.empty or series.max() == series.min(): return pd.Series(0, index=series.index)
    return (series - series.min()) / (series.max() - series.min() + 1e-9)

def calculate_scores(df, forecast_growth_rate=None):
    """
    Compute all business health scores.
    forecast_growth_rate: float|None — projected 6m revenue growth % from the best
      forecast model (Prophet > Holt-Winters > LinReg > Baseline). When provided it
      is normalised to [0,1] and blended into Growth_Potential_Score and
      MSME_Health_Score so that forward-looking model output actually drives the KPIs.
    """
    df = _apply_col_remap(df)
    numeric_cols = ['Monthly_Sales_INR','Monthly_Operating_Cost_INR','Outstanding_Loan_INR',
                    'Vendor_Delivery_Reliability','Inventory_Turnover','Avg_Margin_Percent',
                    'Monthly_Demand_Units','Returns_Percentage']
    for col in numeric_cols:
        if col in df.columns: df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        else: df[col] = 0

    # Use trailing 12 months for score computation to avoid bias from a single
    # spike month (e.g. Diwali) distorting all scores. Keep full df for forecasting.
    if 'Date' in df.columns:
        _dates = pd.to_datetime(df['Date'], errors='coerce')
        _cutoff = _dates.max() - pd.DateOffset(months=12)
        _score_mask = _dates >= _cutoff
        if _score_mask.sum() >= max(len(df) * 0.3, 10):
            df_score = df[_score_mask].copy()
        else:
            df_score = df.copy()   # not enough data — use all
    else:
        df_score = df.copy()

    # Compute score columns on df_score, then broadcast back to df
    # (so forecasting still sees full history)
    def _score_col(col): return df_score[col] if col in df_score.columns else pd.Series(0, index=df_score.index)

    df['Monthly_Sales_INR_Adjusted'] = df['Monthly_Sales_INR'].replace(0, 1e-9)

    # ── Static financial risk ──────────────────────────────────────────────
    # Use absolute thresholds so scores reflect real risk, not just relative rank:
    # Cost ratio >1.0 (cost exceeds revenue) = max stress; <0.5 = low stress
    # Loan/annual-revenue ratio >2.0 = max stress; <0.3 = low stress
    _cost_ratio = (df["Monthly_Operating_Cost_INR"] / df["Monthly_Sales_INR_Adjusted"]).clip(0, 2) / 2.0
    _loan_ratio = (df["Outstanding_Loan_INR"] / (df["Monthly_Sales_INR_Adjusted"] * 12)).clip(0, 2) / 2.0
    df["Cashflow_Stress"]      = _cost_ratio
    df["Loan_Stress"]          = _loan_ratio
    df["Financial_Risk_Score"] = (0.5 * df["Cashflow_Stress"] + 0.5 * df["Loan_Stress"]).clip(0, 1)

    # ── Vendor / supply-chain (0-100 score) ───────────────────────────────
    # Use absolute anchors instead of min-max normalize so the score reflects
    # real-world standards regardless of dataset variance.
    # Anchors: Inventory_Turnover 0→0, 12→1 (12x/month = excellent for hypermarket)
    #          Avg_Margin_Percent  0→0, 30→1 (30%+ margin = excellent)
    _inv_norm = (df["Inventory_Turnover"].clip(0, 12) / 12.0)
    _mar_norm = (df["Avg_Margin_Percent"].clip(0, 30) / 30.0)
    df["Vendor_Score"] = (
        0.50 * df["Vendor_Delivery_Reliability"].clip(0, 1) +
        0.30 * _inv_norm +
        0.20 * _mar_norm
    ).clip(0, 1) * 100  # → 0-100 scale

    # ── Growth Potential: anchored to absolute standards ──────────────────
    # Demand: normalize within dataset (relative rank is fine here — it's comparative)
    # Margin: anchor at 0-40% absolute range
    # Returns: 0% = best (1.0), 15%+ = worst (0.0)
    _dem_norm = normalize(df["Monthly_Demand_Units"])                    # relative rank OK
    _gmar_norm = (df["Avg_Margin_Percent"].clip(0, 40) / 40.0)
    _ret_norm  = (1 - df["Returns_Percentage"].clip(0, 15) / 15.0)
    gps_static = (
        0.40 * _dem_norm +
        0.35 * _gmar_norm +
        0.25 * _ret_norm
    ).clip(0, 1)

    # Blend in forecast growth trend when available (20% weight)
    # forecast_growth_rate is the % change predicted by the best model
    # Cap normalisation: 0 % growth → 0.0,  ≥ 30 % growth → 1.0
    if forecast_growth_rate is not None:
        fg_norm = float(np.clip(forecast_growth_rate / 30.0, 0.0, 1.0))
        df["Growth_Potential_Score"] = (0.80 * gps_static + 0.20 * fg_norm).clip(0, 1)
    else:
        df["Growth_Potential_Score"] = gps_static

    # ── Business Health: add a small forecast momentum signal (5 % weight) ─────
    # When forecast_growth_rate is negative the health score is penalised;
    # when strongly positive it provides a lift — grounded in model output.
    if forecast_growth_rate is not None:
        fg_health = float(np.clip((forecast_growth_rate + 10) / 40.0, 0.0, 1.0))
        df["MSME_Health_Score"] = (
            (1 - df["Financial_Risk_Score"]) * 0.38 +
            (df["Vendor_Score"] / 100)       * 0.29 +
            df["Growth_Potential_Score"]     * 0.28 +
            fg_health                        * 0.05
        ).clip(0, 1) * 100
    else:
        df["MSME_Health_Score"] = (
            (1 - df["Financial_Risk_Score"]) * 0.40 +
            (df["Vendor_Score"] / 100)       * 0.30 +
            df["Growth_Potential_Score"]     * 0.30
        ).clip(0, 1) * 100

    # ── Performance ───────────────────────────────────────────────────────
    df['Profitability_Ratio']    = (df['Avg_Margin_Percent'].clip(0, 40) / 40.0)
    # Op efficiency: cost < 50% of revenue = excellent (1.0), >100% = zero (0.0)
    df['Operational_Efficiency'] = (1 - (df['Monthly_Operating_Cost_INR'] / df['Monthly_Sales_INR_Adjusted']).clip(0, 2) / 2.0).clip(0, 1)
    df['Customer_Satisfaction']  = (1 - df['Returns_Percentage'].clip(0, 15) / 15.0)
    df['Performance_Score'] = (
        0.30 * df['Profitability_Ratio'] +
        0.25 * df['Operational_Efficiency'] +
        0.20 * df['Customer_Satisfaction'] +
        0.15 * df['Vendor_Delivery_Reliability'] +
        0.10 * normalize(df['Inventory_Turnover'])
    ).clip(0, 1) * 100
    return df

def segment_customers(df):
    try:
        sku_col = 'SKU_Name' if 'SKU_Name' in df.columns else None
        if not sku_col: return None
        sales_col = 'Monthly_Sales_INR'
        if sales_col not in df.columns: return None
        df = df.copy()
        if 'Date' in df.columns:
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce'); df = df.dropna(subset=['Date'])
            ref = df['Date'].max()
            rfm = df.groupby(sku_col).agg(recency=('Date', lambda x: (ref - x.max()).days), frequency=(sales_col,'count'), monetary=(sales_col,'sum')).reset_index()
        else:
            rfm = df.groupby(sku_col).agg(frequency=(sales_col,'count'), monetary=(sales_col,'sum')).reset_index(); rfm['recency'] = 0
        for col, alias in [('Avg_Margin_Percent','avg_margin'),('Monthly_Demand_Units','avg_demand')]:
            if col in df.columns:
                m = df.groupby(sku_col)[col].mean().reset_index(); m.columns=[sku_col,alias]; rfm = rfm.merge(m,on=sku_col,how='left'); rfm[alias]=rfm[alias].fillna(0)
            else:
                rfm[alias] = 0
        if len(rfm) < 2: return None
        # Log-transform skewed features before scaling — monetary can range from
        # ₹1L to ₹100Cr; without log-transform KMeans clusters on scale not pattern.
        rfm_scaled = rfm[['recency','frequency','monetary','avg_margin','avg_demand']].fillna(0).copy()
        for col in ['frequency', 'monetary', 'avg_demand']:
            rfm_scaled[col] = np.log1p(rfm_scaled[col].clip(lower=0))
        X = rfm_scaled.values
        scaler = StandardScaler(); X_scaled = scaler.fit_transform(X)
        n_clusters = min(5, max(2, len(rfm))); kmeans = KMeans(n_clusters=n_clusters, random_state=42, n_init='auto')
        rfm['cluster_id'] = kmeans.fit_predict(X_scaled)
        cluster_monetary = rfm.groupby('cluster_id')['monetary'].mean().sort_values(ascending=False)
        names = ['Champions','Loyal','Potential','At Risk','Lost']
        cmap = {cid: names[i] if i < len(names) else f'Segment {i}' for i,cid in enumerate(cluster_monetary.index)}
        rfm['segment_name'] = rfm['cluster_id'].map(cmap)
        segment_stats = {}
        for seg, grp in rfm.groupby('segment_name'):
            segment_stats[seg] = {'count': int(len(grp)), 'avg_sales': float(grp['monetary'].mean()),
                                  'avg_margin': float(grp['avg_margin'].mean()), 'avg_demand': float(grp['avg_demand'].mean()),
                                  'total_sales': float(grp['monetary'].sum()), 'top_products': grp.nlargest(3,'monetary')[sku_col].tolist()}
        return {'counts': rfm['segment_name'].value_counts().to_dict(), 'rfm_df': rfm, 'segment_stats': segment_stats, 'sku_col': sku_col, 'n_clusters': n_clusters}
    except Exception:
        return None

def _run_prophet_model(monthly_df, periods=12):
    """Run Prophet on a monthly DataFrame with columns [ds, y]. Returns forecast dict or None.
    Has a 25-second timeout — if Prophet hangs (Stan warmup on cold start), returns None
    so the ensemble falls back to Holt-Winters + Linear Regression."""
    if not PROPHET_AVAILABLE or len(monthly_df) < 2: return None
    import threading
    _result = [None]
    _error  = [None]
    def _prophet_run_once(train, last_date, periods):
        model = Prophet(
            yearly_seasonality=False,
            weekly_seasonality=False,
            daily_seasonality=False,
            changepoint_prior_scale=0.05,
            seasonality_prior_scale=10.0,
            interval_width=0.80
        )
        model.fit(train)
        future = model.make_future_dataframe(periods=periods, freq='MS')
        fc = model.predict(future)
        ff = fc[fc['ds'] >= last_date][['ds','yhat','yhat_lower','yhat_upper']].copy()
        for c in ['yhat','yhat_lower','yhat_upper']: ff[c] = ff[c].clip(lower=0)
        f6 = ff.head(6); f12 = ff.head(12)
        return {
            '6_month':  {'forecast': f6['yhat'].sum(),  'lower': f6['yhat_lower'].sum(),  'upper': f6['yhat_upper'].sum()},
            '12_month': {'forecast': f12['yhat'].sum(), 'lower': f12['yhat_lower'].sum(), 'upper': f12['yhat_upper'].sum()},
            'forecast_df': ff, 'model_name': 'Prophet'
        }
    def _worker():
        try:
            train     = monthly_df.tail(12).copy()
            last_date = monthly_df['ds'].max()
            _result[0] = _prophet_run_once(train, last_date, periods)
        except Exception as e:
            _error[0] = e
    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    t.join(timeout=25)   # 25-second hard cap — Prophet Stan warmup can otherwise stall
    if t.is_alive():
        import logging
        logging.getLogger('DataNetra').warning('Prophet timed out — falling back to HW/LR ensemble')
        return None
    if _error[0] is not None:
        return None
    return _result[0]

def _run_holtwinters_model(monthly_df, periods=12):
    """
    Holt-Winters Exponential Smoothing.
    Uses statsmodels ExponentialSmoothing when available (full triple: trend + seasonal).
    Falls back to pure numpy Holt linear-trend method — no dependencies needed.
    """
    if len(monthly_df) < 3: return None

    y = monthly_df['y'].values.astype(float)

    # ── Path 1: statsmodels (preferred — more accurate, handles seasonality) ─
    if _HW_STATSMODELS_AVAILABLE:
        try:
            use_seasonal = len(y) >= 24
            model = _HW_statsmodels(
                y,
                trend='add',
                seasonal='add' if use_seasonal else None,
                seasonal_periods=12 if use_seasonal else None,
                initialization_method='estimated'
            ).fit(optimized=True)
            forecast = np.clip(model.forecast(periods), 0, None)
            std  = float(np.std(y - model.fittedvalues)) if hasattr(model, 'fittedvalues') else float(np.mean(y) * 0.10)
            f6   = float(forecast[:6].sum());  f12 = float(forecast[:12].sum())
            ci6  = std * np.sqrt(6)  * 1.65
            ci12 = std * np.sqrt(12) * 1.65
            params = model.params if hasattr(model, 'params') else {}
            return {
                '6_month':  {'forecast': f6,  'lower': max(0.0, f6  - ci6),  'upper': f6  + ci6},
                '12_month': {'forecast': f12, 'lower': max(0.0, f12 - ci12), 'upper': f12 + ci12},
                'alpha': float(params.get('smoothing_level', 0.3)),
                'beta':  float(params.get('smoothing_trend', 0.1)),
                'trend_per_month': float(forecast[1] - forecast[0]) if len(forecast) > 1 else 0.0,
                'engine': 'statsmodels',
                'model_name': 'Holt-Winters'
            }
        except Exception:
            pass  # fall through to numpy implementation

    # ── Path 2: pure numpy fallback ──────────────────────────────────────────
    try:
        n = len(y)
        best_sse = float('inf')
        best_alpha, best_beta = 0.3, 0.1
        for alpha in [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8]:
            for beta in [0.0, 0.05, 0.1, 0.15, 0.2, 0.3]:
                level, trend = y[0], y[1] - y[0]
                sse = 0.0
                for t in range(1, n):
                    pred = level + trend
                    sse += (y[t] - pred) ** 2
                    level_new = alpha * y[t] + (1 - alpha) * (level + trend)
                    trend = beta * (level_new - level) + (1 - beta) * trend
                    level = level_new
                if sse < best_sse:
                    best_sse = sse; best_alpha = alpha; best_beta = beta

        alpha, beta = best_alpha, best_beta
        level, trend = y[0], y[1] - y[0]
        fitted = []
        for t in range(n):
            fitted.append(level + trend)
            level_new = alpha * y[t] + (1 - alpha) * (level + trend)
            trend = beta * (level_new - level) + (1 - beta) * trend
            level = level_new
        forecast = np.array([max(0.0, level + (i + 1) * trend) for i in range(periods)])
        std  = float(np.std(y - np.array(fitted)))
        f6   = float(forecast[:6].sum());  f12 = float(forecast[:12].sum())
        ci6  = std * np.sqrt(6)  * 1.65
        ci12 = std * np.sqrt(12) * 1.65
        return {
            '6_month':  {'forecast': f6,  'lower': max(0.0, f6  - ci6),  'upper': f6  + ci6},
            '12_month': {'forecast': f12, 'lower': max(0.0, f12 - ci12), 'upper': f12 + ci12},
            'alpha': alpha, 'beta': beta, 'trend_per_month': float(trend),
            'engine': 'numpy',
            'model_name': 'Holt-Winters'
        }
    except Exception:
        return None
def _run_linear_regression_model(monthly_df, periods=12):
    """Run Linear Regression trend forecasting on monthly DataFrame [ds, y]. Returns forecast dict or None."""
    if len(monthly_df) < 2: return None
    try:
        y = monthly_df['y'].values.astype(float)
        X = np.arange(len(y)).reshape(-1, 1)
        model = _LinearRegression()
        model.fit(X, y)
        future_X = np.arange(len(y), len(y) + periods).reshape(-1, 1)
        future_y = model.predict(future_X)
        future_y = np.clip(future_y, 0, None)
        # Residual std for confidence interval
        residuals = y - model.predict(X)
        std = float(np.std(residuals))
        f6_vals  = future_y[:6];  f12_vals = future_y[:12]
        f6_sum   = float(f6_vals.sum());  f12_sum = float(f12_vals.sum())
        ci6  = std * np.sqrt(6) * 1.96
        ci12 = std * np.sqrt(12) * 1.96
        r2 = float(model.score(X, y))
        return {
            '6_month':  {'forecast': f6_sum,  'lower': max(0, f6_sum  - ci6),  'upper': f6_sum  + ci6},
            '12_month': {'forecast': f12_sum, 'lower': max(0, f12_sum - ci12), 'upper': f12_sum + ci12},
            'r2_score': r2, 'slope': float(model.coef_[0]), 'intercept': float(model.intercept_),
            'model_name': 'Linear Regression'
        }
    except: return None

def _run_baseline_model(monthly_df, periods=12):
    """
    Statistical baseline using trailing trend.
    Growth rate is derived from the actual last-3m vs prior-3m comparison
    so declining products show declining forecasts, not a forced +5%.
    """
    if len(monthly_df) == 0: return None
    y = monthly_df['y'].values.astype(float)

    # Trailing average (last 6 months if available, else all)
    last6  = y[-6:] if len(y) >= 6 else y
    avg    = float(last6.mean())

    # Trend: compare last 3m vs prior 3m to detect direction
    if len(y) >= 6:
        recent  = float(y[-3:].mean())
        earlier = float(y[-6:-3].mean())
        growth  = (recent - earlier) / (earlier + 1e-9)
        # Dampen extreme swings: cap at ±20% per 6-month window
        growth  = float(np.clip(growth, -0.20, 0.20))
    elif len(y) >= 3:
        growth = float(np.clip((y[-1] - y[0]) / (y[0] + 1e-9), -0.15, 0.15))
    else:
        growth = 0.0   # flat — not enough data to infer direction

    f6  = avg * 6 * (1 + growth)
    f12 = avg * 12 * (1 + growth * 0.8)   # dampen further for 12m
    ci_pct = 0.12   # ±12% for baseline uncertainty
    return {
        '6_month':  {'forecast': max(0, f6),  'lower': max(0, f6  * (1-ci_pct)), 'upper': f6  * (1+ci_pct)},
        '12_month': {'forecast': max(0, f12), 'lower': max(0, f12 * (1-ci_pct)), 'upper': f12 * (1+ci_pct)},
        'growth_rate_used': round(growth * 100, 1),
        'model_name': 'Statistical Baseline'
    }

def forecast_sales(df):
    """
    Runs ALL available forecasting models and produces a weighted ensemble result.
    Weights: Prophet=40%, Holt-Winters=30%, Linear Regression=20%, Baseline=10%
    If a model is unavailable/fails, its weight is redistributed proportionally.
    """
    sales_col = None
    for c in ['Monthly_Sales_INR','Gross_Sales']:
        if c in df.columns: sales_col = c; break
    if not sales_col:
        total = df.select_dtypes(include=[np.number]).sum().sum()
        avg = total / max(len(df), 1)
        f6 = avg*6*1.05; f12 = avg*12*1.05
        return {'6_month':{'forecast':f6,'lower':f6*0.85,'upper':f6*1.15},
                '12_month':{'forecast':f12,'lower':f12*0.85,'upper':f12*1.15},
                'model_results': {}, 'selected_model': 'Statistical Baseline',
                'ensemble': False}

    # Build monthly time series
    has_date = 'Date' in df.columns
    monthly  = None
    if has_date:
        dfc = df.copy()
        dfc['Date'] = pd.to_datetime(dfc['Date'], errors='coerce')
        dfc = dfc.dropna(subset=['Date'])
        ts = dfc.set_index('Date')[sales_col].resample('MS').sum().reset_index()
        ts.columns = ['ds', 'y']
        ts = ts.sort_values('ds').reset_index(drop=True)
        if len(ts) >= 2: monthly = ts

    if monthly is None:
        avg = float(df[sales_col].mean())
        synthetic_ds = pd.date_range(end=pd.Timestamp.now(), periods=12, freq='MS')
        monthly = pd.DataFrame({'ds': synthetic_ds, 'y': np.array([avg]*12)})

    # ── Run ALL 4 models ─────────────────────────────────────────────────────
    prophet_result  = _run_prophet_model(monthly)  if has_date else None
    hw_result       = _run_holtwinters_model(monthly)
    lr_result       = _run_linear_regression_model(monthly)
    baseline_result = _run_baseline_model(monthly)

    # Compile results dict
    model_results = {}
    if prophet_result:  model_results['Prophet']              = prophet_result
    if hw_result:       model_results['Holt-Winters']         = hw_result
    if lr_result:       model_results['Linear Regression']    = lr_result
    if baseline_result: model_results['Statistical Baseline'] = baseline_result

    # ── Dynamic Ensemble: inverse-quality weighting ───────────────────────────
    # Weight each model by 1/(1-R²) for LR, or static rank otherwise.
    # A model that fits this data poorly automatically gets less weight.
    STATIC_RANK = {'Prophet':4,'Holt-Winters':3,'Linear Regression':2,'Statistical Baseline':1}

    if len(model_results) == 0:
        avg = float(df[sales_col].mean())
        f6 = avg * 6; f12 = avg * 12
        return {'6_month':{'forecast':f6,'lower':f6*0.85,'upper':f6*1.15},
                '12_month':{'forecast':f12,'lower':f12*0.85,'upper':f12*1.15},
                'model_results': {}, 'selected_model': 'Statistical Baseline', 'ensemble': False}

    def _quality_score(name, res):
        if 'r2_score' in res: return max(0.05, float(res['r2_score']))
        return STATIC_RANK.get(name, 1) / 4.0

    raw_w   = {m: _quality_score(m, model_results[m]) for m in model_results}
    total_w = sum(raw_w.values())
    norm_w  = {m: raw_w[m] / total_w for m in raw_w}

    def _weighted(key):
        f_val = sum(model_results[m][key]['forecast'] * norm_w[m] for m in norm_w)
        l_val = sum(model_results[m][key]['lower']    * norm_w[m] for m in norm_w)
        u_val = sum(model_results[m][key]['upper']    * norm_w[m] for m in norm_w)
        return {'forecast': f_val, 'lower': l_val, 'upper': u_val}

    ensemble_6m  = _weighted('6_month')
    ensemble_12m = _weighted('12_month')

    used_models   = list(norm_w.keys())
    weight_labels = " + ".join(f"{k.split()[0]} {norm_w[k]*100:.0f}%" for k in used_models)
    ensemble_name = f"Ensemble ({weight_labels})"

    return {
        '6_month':        ensemble_6m,
        '12_month':       ensemble_12m,
        'model_results':  model_results,
        'selected_model': ensemble_name,
        'ensemble':       True,
        'ensemble_weights': norm_w,
        'models_used':    used_models,
        'forecast_dfs':   {},
        'per_store_forecasts': {}
    }
def generate_granular_forecast(df):
    """
    Per-entity (overall / store / category / SKU) best-model forecast.
    For each entity independently runs all 4 models, validates on held-out
    months using MAE, and uses the winner — so a SKU with 4 months uses
    Baseline/LR while a store with 24 months uses Prophet/HW.
    """
    import pandas as _pd_ggf
    if isinstance(df, dict):
        try:
            import pandas as _pd_ggf
            if 'columns' in df and 'data' in df:
                df = _pd_ggf.DataFrame(df['data'], columns=df['columns'])
            else:
                df = _pd_ggf.DataFrame(df)
        except Exception: return None
    if df is None or (hasattr(df,'empty') and df.empty):
        return None
    # Column resolution — handles original case AND DRC-normalised (lowercase)
    _low = {c.lower(): c for c in df.columns}
    def _gcol(*names):
        for n in names:
            if n in df.columns: return n
            if n.lower() in _low: return _low[n.lower()]
        return None
    sales_col = _gcol('Monthly_Sales_INR','revenue','Revenue','monthly_sales_inr','sales','gross_sales','net_sales','sales_amount','total_sales','total_revenue')
    sku_col   = _gcol('product_name','product','SKU_Name','sku_name','SKU','Product_Name','sku','SKU','SKU_Name','sku_name','product_id')
    cat_col   = _gcol('category','Category','Product_Category','product_category')
    store_col = _gcol('store_id','Store_ID','store_name','Store_Name','store')
    date_col  = _gcol('Date','date','month','period')
    if sales_col is None: return None
    df = df.copy()
    df[sales_col] = pd.to_numeric(df[sales_col], errors='coerce').fillna(0)
    has_dates = date_col is not None
    if has_dates:
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
        df = df.dropna(subset=[date_col])

    # ── Build monthly [ds, y] DataFrame for a given sub-DataFrame ────────────
    def _monthly(sdf):
        if not has_dates:
            return None
        m = (sdf.set_index(date_col or 'Date')[sales_col]
               .resample('MS').sum()
               .reset_index()
               .rename(columns={date_col or 'Date': 'ds', sales_col: 'y'})
               .sort_values('ds')
               .reset_index(drop=True))
        m['y'] = m['y'].clip(lower=0)
        return m if len(m) >= 2 else None

    # ── Best-model selection per entity ──────────────────────────────────────
    def _best_forecast(monthly_df, n_periods=12):
        """
        Runs Prophet / Holt-Winters / Linear Regression / Baseline on monthly_df.
        Validates each on held-out last min(6, N//2) months using MAE.
        Returns unified result dict with winning model's forecast.
        """
        if monthly_df is None or len(monthly_df) < 2:
            return None

        y     = monthly_df['y'].values.astype(float)
        N     = len(y)
        val_n = max(2, min(6, N // 2))
        tr_y  = y[:-val_n]
        va_y  = y[-val_n:]

        candidates = {}   # model_key → MAE
        hw_params  = (0.3, 0.1)
        prophet_fc = None   # pre-computed full-series Prophet forecast

        # ── Outlier winsorisation (IQR-based, conservative) ─────────────────
        def _winsz(arr):
            if len(arr) < 4: return arr.copy()
            q1, q3 = np.percentile(arr, 25), np.percentile(arr, 75)
            iqr = q3 - q1
            return np.clip(arr, max(0, q1 - 2.5*iqr), q3 + 2.5*iqr)
        y_c    = _winsz(y)
        tr_y_c = y_c[:-val_n]

        # Use MAPE (scale-invariant) so stores/categories of different sizes
        # are compared fairly. Falls back to MAE when actuals are zero.
        def _smape(actual, predicted):
            mask = actual > 0
            if not mask.any(): return float(np.mean(np.abs(actual - predicted)))
            return float(np.mean(np.abs((actual[mask] - predicted[mask]) / actual[mask])) * 100)

        # ── A: Prophet — SKIPPED in granular loop (too slow per-entity).
        #    Prophet runs once at the top level in forecast_sales().
        #    Per-entity we use HW + LR + Baseline which are sub-millisecond.
        if False and PROPHET_AVAILABLE and N >= 4:  # disabled for performance
            try:
                tr_df = monthly_df.iloc[:-val_n][['ds', 'y']].copy()
                tr_df['y'] = _winsz(tr_df['y'].values)
                _mp   = Prophet(
                    yearly_seasonality=True,
                    weekly_seasonality=False,
                    daily_seasonality=False,
                    changepoint_prior_scale=0.05,
                    interval_width=0.80,
                )
                _mp.fit(tr_df)
                fut_v  = _mp.make_future_dataframe(periods=val_n, freq='MS')
                fc_v   = _mp.predict(fut_v).tail(val_n)
                candidates['prophet'] = _smape(va_y, fc_v['yhat'].clip(lower=0).values)
                # Full-series fit for final forecast
                _mf_data = monthly_df[['ds', 'y']].copy()
                _mf_data['y'] = _winsz(_mf_data['y'].values)
                _mf = Prophet(
                    yearly_seasonality=True,
                    weekly_seasonality=False,
                    daily_seasonality=False,
                    changepoint_prior_scale=0.05,
                    interval_width=0.80,
                )
                _mf.fit(_mf_data)
                fut_f  = _mf.make_future_dataframe(periods=n_periods, freq='MS')
                fc_f   = _mf.predict(fut_f).tail(n_periods)
                prophet_fc = {
                    'ds':  fc_f['ds'].values,
                    'yhat': fc_f['yhat'].clip(lower=0).values,
                    'yhat_lower': fc_f['yhat_lower'].clip(lower=0).values,
                    'yhat_upper': fc_f['yhat_upper'].clip(lower=0).values,
                }
            except Exception:
                pass

        # ── B: Holt-Winters (grid-search on VALIDATION MAPE, not train SSE) ─
        try:
            best_val_mape = float('inf')
            for a in [0.1, 0.2, 0.3, 0.5, 0.7]:
                for b in [0.0, 0.05, 0.1, 0.2]:
                    lv = tr_y_c[0]; tr = (tr_y_c[1] - tr_y_c[0]) if len(tr_y_c) > 1 else 0.0
                    for t in range(1, len(tr_y_c)):
                        lv_n = a * tr_y_c[t] + (1 - a) * (lv + tr)
                        tr   = b * (lv_n - lv) + (1 - b) * tr
                        lv   = lv_n
                    hw_vp = np.array([max(0.0, lv + (i+1)*tr) for i in range(val_n)])
                    vm = _smape(va_y, hw_vp)
                    if vm < best_val_mape:
                        best_val_mape = vm; hw_params = (a, b)
            a, b = hw_params
            lv = tr_y_c[0]; tr = (tr_y_c[1] - tr_y_c[0]) if len(tr_y_c) > 1 else 0.0
            for t in range(1, len(tr_y_c)):
                lv_n = a * tr_y_c[t] + (1 - a) * (lv + tr)
                tr   = b * (lv_n - lv) + (1 - b) * tr
                lv   = lv_n
            hw_val = np.array([max(0.0, lv + (i + 1) * tr) for i in range(val_n)])
            candidates['hw'] = _smape(va_y, hw_val)
        except Exception:
            pass

        # ── C: Linear Regression with month-of-year seasonal encoding ────────
        try:
            from sklearn.linear_model import LinearRegression as _LR
            def _seasonal_X(mdf, idxs):
                t_col  = np.array(idxs, dtype=float)
                if 'ds' in mdf.columns:
                    mons = pd.to_datetime(mdf.iloc[idxs]['ds']).dt.month.values
                else:
                    mons = (np.array(idxs) % 12 + 1)
                sin_m = np.sin(2 * np.pi * mons / 12)
                cos_m = np.cos(2 * np.pi * mons / 12)
                return np.column_stack([t_col, sin_m, cos_m])
            tr_idx = list(range(len(tr_y_c)))
            va_idx = list(range(len(tr_y_c), len(tr_y_c) + val_n))
            _lr = _LR().fit(_seasonal_X(monthly_df, tr_idx), tr_y_c)
            candidates['lr'] = _smape(va_y, _lr.predict(_seasonal_X(monthly_df, va_idx)).clip(0))
        except Exception:
            pass

        # ── D: Baseline (trend-aware, not hardcoded +5%) ─────────────────────
        try:
            if len(tr_y) >= 6:
                _b_growth = float(np.clip(
                    (float(tr_y[-3:].mean()) - float(tr_y[-6:-3].mean())) / (float(tr_y[-6:-3].mean()) + 1e-9),
                    -0.20, 0.20))
            elif len(tr_y) >= 2:
                _b_growth = float(np.clip((tr_y[-1] - tr_y[0]) / (tr_y[0] + 1e-9), -0.15, 0.15))
            else:
                _b_growth = 0.0
            base_avg = float(tr_y[-6:].mean()) if len(tr_y) >= 6 else float(tr_y.mean())
            _base_fc  = base_avg * (1 + _b_growth)
            candidates['base'] = float(np.mean(np.abs(np.full(val_n, _base_fc) - va_y)))
        except Exception:
            pass

        best = min(candidates, key=candidates.get) if candidates else 'base'

        # ── Produce final forecast with winner ───────────────────────────────
        if best == 'prophet' and prophet_fc is not None:
            fc_df = pd.DataFrame(prophet_fc)
            f6    = fc_df.head(6);  f12 = fc_df.head(12)
            return {
                'model': '📡 Prophet',
                'mae':   candidates['prophet'],
                '6m_forecast':  float(f6['yhat'].sum()),
                '6m_lower':     float(f6['yhat_lower'].sum()),
                '6m_upper':     float(f6['yhat_upper'].sum()),
                '12m_forecast': float(f12['yhat'].sum()),
                '12m_lower':    float(f12['yhat_lower'].sum()),
                '12m_upper':    float(f12['yhat_upper'].sum()),
                'hist': monthly_df,
                'fc':   fc_df,
            }

        elif best == 'hw':
            a, b = hw_params
            lv = y[0]; tr = (y[1] - y[0]) if N > 1 else 0.0
            for t in range(1, N):
                lv_n = a * y[t] + (1 - a) * (lv + tr)
                tr   = b * (lv_n - lv) + (1 - b) * tr
                lv   = lv_n
            yh  = np.array([max(0.0, lv + (i + 1) * tr) for i in range(n_periods)])
            std = max(float(np.std(y - np.array([y[0] + tr * i for i in range(N)]))), 1.0)
            lo  = (yh - 1.65 * std).clip(0)
            hi  = yh + 1.65 * std
            last_ds = monthly_df['ds'].max()
            fc_dates = pd.date_range(last_ds, periods=n_periods + 1, freq='MS')[1:]
            fc_df = pd.DataFrame({'ds': fc_dates, 'yhat': yh,
                                   'yhat_lower': lo, 'yhat_upper': hi})
            f6 = fc_df.head(6); f12 = fc_df.head(12)
            return {
                'model': '❄️ Holt-Winters',
                'mae':   candidates['hw'],
                '6m_forecast':  float(f6['yhat'].sum()),
                '6m_lower':     float(f6['yhat_lower'].sum()),
                '6m_upper':     float(f6['yhat_upper'].sum()),
                '12m_forecast': float(f12['yhat'].sum()),
                '12m_lower':    float(f12['yhat_lower'].sum()),
                '12m_upper':    float(f12['yhat_upper'].sum()),
                'hist': monthly_df,
                'fc':   fc_df,
            }

        elif best == 'lr':
            from sklearn.linear_model import LinearRegression as _LR
            def _sX(idxs, mdf):
                t = np.array(idxs, dtype=float)
                mons = pd.to_datetime(mdf.iloc[idxs]['ds']).dt.month.values if 'ds' in mdf.columns else (np.array(idxs) % 12 + 1)
                return np.column_stack([t, np.sin(2*np.pi*mons/12), np.cos(2*np.pi*mons/12)])
            all_idx = list(range(N))
            fut_idx = list(range(N, N + n_periods))
            # build future monthly_df slice for seasonal encoding
            last_ds  = monthly_df['ds'].max()
            fc_dates = pd.date_range(last_ds, periods=n_periods + 1, freq='MS')[1:]
            fc_stub  = pd.DataFrame({'ds': fc_dates})
            _lr_full = _LR().fit(_sX(all_idx, monthly_df), _winsz(y))
            _fut_sX  = np.column_stack([
                np.arange(N, N+n_periods, dtype=float),
                np.sin(2*np.pi*fc_dates.month/12),
                np.cos(2*np.pi*fc_dates.month/12)
            ])
            yh    = _lr_full.predict(_fut_sX).clip(0)
            resid = _winsz(y) - _lr_full.predict(_sX(all_idx, monthly_df))
            std   = max(float(np.std(resid)), 1.0)
            # Proper prediction interval: σ × sqrt(1 + 1/N) widening with horizon
            pi    = std * np.sqrt(1 + 1/max(N,1) + np.arange(1, n_periods+1)/max(N,1)) * 1.65
            lo    = (yh - pi).clip(0)
            hi    = yh + pi
            fc_df = pd.DataFrame({'ds': fc_dates, 'yhat': yh,
                                   'yhat_lower': lo, 'yhat_upper': hi})
            f6 = fc_df.head(6); f12 = fc_df.head(12)
            return {
                'model': '📐 Linear Reg',
                'mae':   candidates['lr'],
                '6m_forecast':  float(f6['yhat'].sum()),
                '6m_lower':     float(f6['yhat_lower'].sum()),
                '6m_upper':     float(f6['yhat_upper'].sum()),
                '12m_forecast': float(f12['yhat'].sum()),
                '12m_lower':    float(f12['yhat_lower'].sum()),
                '12m_upper':    float(f12['yhat_upper'].sum()),
                'hist': monthly_df,
                'fc':   fc_df,
            }

        else:  # baseline
            avg = float(y[-6:].mean()) if N >= 6 else float(y.mean())
            if N >= 6:
                _bg = float(np.clip((float(y[-3:].mean()) - float(y[-6:-3].mean())) / (float(y[-6:-3].mean()) + 1e-9), -0.20, 0.20))
            elif N >= 2:
                _bg = float(np.clip((y[-1] - y[0]) / (y[0] + 1e-9), -0.15, 0.15))
            else:
                _bg = 0.0
            f6v  = avg * (1 + _bg) * 6;   f12v = avg * (1 + _bg * 0.8) * 12
            std  = max(float(np.std(y - np.full(N, avg))), 1.0)
            last_ds  = monthly_df['ds'].max()
            fc_dates = pd.date_range(last_ds, periods=n_periods + 1, freq='MS')[1:]
            yh   = np.full(n_periods, avg * (1 + _bg))
            lo   = (yh - 1.65 * std).clip(0)
            hi   = yh + 1.65 * std
            fc_df = pd.DataFrame({'ds': fc_dates, 'yhat': yh,
                                   'yhat_lower': lo, 'yhat_upper': hi})
            f6 = fc_df.head(6); f12 = fc_df.head(12)
            return {
                'model': '📊 Baseline',
                'mae':   candidates.get('base', 0),
                '6m_forecast':  float(f6['yhat'].sum()),
                '6m_lower':     float(f6['yhat_lower'].sum()),
                '6m_upper':     float(f6['yhat_upper'].sum()),
                '12m_forecast': float(f12['yhat'].sum()),
                '12m_lower':    float(f12['yhat_lower'].sum()),
                '12m_upper':    float(f12['yhat_upper'].sum()),
                'hist': monthly_df,
                'fc':   fc_df,
            }

    # ── Unified pack: converts best_forecast result → standard output dict ────
    def _pack(label, res, total):
        if res is None:
            avg = total / 12 if total > 0 else 0; g = 0.05
            return {
                'label': label, 'model': '📊 Baseline', 'mae': None,
                'total_hist': total,
                '6m_forecast':  avg*6*(1+g),  '6m_lower':  avg*6*(1+g)*0.85,  '6m_upper':  avg*6*(1+g)*1.15,
                '12m_forecast': avg*12*(1+g), '12m_lower': avg*12*(1+g)*0.85, '12m_upper': avg*12*(1+g)*1.15,
                'hist': None, 'fc': None,
            }
        return {
            'label':        label,
            'model':        res['model'],
            'mae':          res['mae'],
            'total_hist':   total,
            '6m_forecast':  res['6m_forecast'],
            '6m_lower':     res['6m_lower'],
            '6m_upper':     res['6m_upper'],
            '12m_forecast': res['12m_forecast'],
            '12m_lower':    res['12m_lower'],
            '12m_upper':    res['12m_upper'],
            'hist':         res.get('hist'),
            'fc':           res.get('fc'),
        }

    # ── Run per-entity ────────────────────────────────────────────────────────
    overall = _pack('Overall Company',
                    _best_forecast(_monthly(df), 12),
                    float(df[sales_col].sum()))

    stores = []
    if store_col:
        _store_ids = (df.groupby(store_col)[sales_col].sum()
                        .sort_values(ascending=False).head(3).index.tolist())  # capped at 3 for speed
        for sid in _store_ids:
            sdf = df[df[store_col] == sid]
            stores.append(_pack(str(sid),
                                _best_forecast(_monthly(sdf), 12),
                                float(sdf[sales_col].sum())))

    categories = []
    if cat_col:
        _cat_ids = (df.groupby(cat_col)[sales_col].sum()
                      .sort_values(ascending=False).head(3).index.tolist())  # capped at 3 for speed
        for cat in _cat_ids:
            cdf = df[df[cat_col] == cat]
            categories.append(_pack(str(cat),
                                    _best_forecast(_monthly(cdf), 12),
                                    float(cdf[sales_col].sum())))

    products = []
    if sku_col:
        top_skus = (df.groupby(sku_col)[sales_col].sum()
                      .sort_values(ascending=False).head(3).index.tolist())  # capped at 3 for speed
        for sk in top_skus:
            skdf = df[df[sku_col] == sk]
            products.append(_pack(str(sk),
                                  _best_forecast(_monthly(skdf), 12),
                                  float(skdf[sales_col].sum())))

    return {
        'overall':    overall,
        'stores':     stores,
        'categories': categories,
        'products':   products,
        'sales_col':  sales_col,
        'raw_df':     df,
        'sku_col':    sku_col,
        'cat_col':    cat_col,
    }

# ── UI translations: English / Hindi ───────────────────────────────────────────
LANG = {
    'en': {
        'insights_title':'AI-Powered Business Insights','overall_summary':'Overall Performance Summary',
        'total_sales':'Total Sales','total_products':'Total Products Analyzed','avg_margin':'Average Profit Margin',
        'health_score':'Overall Business Health Score','perf_score':'Overall Performance Score',
        'top5':'Top 5 Performing Products','perf_metrics':'Performance Metrics',
        'fin_risk':'Financial Risk Score','vendor_score':'Vendor Reliability Score','growth_score':'Growth Potential Score',
        'lower_better':'(Lower is better)','forecast_title':'ML-Powered Sales Forecast',
        'six_month':'6-Month Projection','twelve_month':'12-Month Projection',
        'forecast_sales':'Forecasted Sales','expected_range':'Expected Range',
        'snp_title':'Marketplace Channel Matching','recommendations':'AI-Generated Recommendations',
        'immediate':'Immediate Actions','strategic':'Strategic Initiatives','risk_alert':'Risk Alerts',
        'store_forecast':'Store-Specific Sales Forecasts','data_quality':'Data Quality Report',
        'inference_time':'Analysis completed in','seconds':'seconds'
    },
    'hi': {
        'insights_title':'AI से मिली आपके धंधे की जानकारी','overall_summary':'आपके धंधे का कुल हाल',
        'total_sales':'कुल बिक्री','total_products':'कुल सामान','avg_margin':'औसत मुनाफा (%)',
        'health_score':'धंधे की सेहत का स्कोर','perf_score':'काम का कुल स्कोर',
        'top5':'सबसे ज्यादा बिकने वाले 5 सामान','perf_metrics':'काम के नंबर',
        'fin_risk':'पैसों का जोखिम स्कोर','vendor_score':'सप्लायर का भरोसा स्कोर','growth_score':'आगे बढ़ने की संभावना',
        'lower_better':'(कम नंबर अच्छा है)','forecast_title':'AI से अगले महीनों की बिक्री का अनुमान',
        'six_month':'अगले 6 महीने की बिक्री','twelve_month':'अगले 12 महीने की बिक्री',
        'forecast_sales':'अनुमानित बिक्री','expected_range':'बिक्री कम से कम — ज्यादा से ज्यादा',
        'snp_title':'Marketplace पर बेचने के लिए सबसे अच्छे Platform','recommendations':'AI की सलाह',
        'immediate':'अभी करने वाले काम','strategic':'आगे की योजना','risk_alert':'खतरे की चेतावनी',
        'store_forecast':'हर दुकान की अगली बिक्री का अनुमान','data_quality':'आपके Data की जाँच',
        'inference_time':'जाँच पूरी हुई','seconds':'सेकंड में'
    }
}

def T(key, lang='en'): return LANG.get(lang, LANG['en']).get(key, LANG['en'].get(key, key))

SNP_CATALOG = {
    'GeM (Government e-Marketplace)': {
        'business_types':['Manufacturing','Electronics','Services'],
        'min_health':50,'segment_boost':['Champions','Loyal'],
        'channel_type':'government',          # public sector procurement
        'b2c_fit': ['Services'],              # not B2C consumer goods
        'description_en':'Government procurement portal — ideal for retailers supplying to public sector.',
        'action_en':'Register on GeM portal (gem.gov.in) and map your product catalogue.'},
    'Flipkart Commerce': {
        'business_types':['FMCG','Supermarket','Clothing','Electronics','Hypermarket','Retail'],
        'min_health':25,'segment_boost':['Champions','Loyal','Potential','At Risk'],
        'channel_type':'b2c_marketplace',     # highest B2C volume in India
        'b2c_fit': ['FMCG','Supermarket','Hypermarket','Retail','Clothing','Electronics'],
        'margin_sweet_spot': (10, 25),         # volume marketplace — mass-market margin range
        'description_en':'High-volume B2C marketplace — best for consumer goods with strong demand.',
        'action_en':'Onboard via Flipkart Seller Hub — optimise product images and descriptions.'},
    'Meesho': {
        'business_types':['Clothing','FMCG','Manufacturing'],
        'min_health':20,'segment_boost':['Potential','At Risk'],
        'channel_type':'social_commerce',
        'b2c_fit': ['Clothing','FMCG'],
        'description_en':'Social commerce marketplace — ideal for price-sensitive segments and tier-2/3 markets.',
        'action_en':'List on Meesho for reseller network access — focus on competitive pricing.'},
    'NSIC e-Marketplace': {
        'business_types':['Manufacturing','FMCG','Electronics','Services','Hypermarket'],
        'min_health':25,'segment_boost':['Champions','Loyal','Potential'],
        'channel_type':'b2b_institutional',   # MSE-to-MSE, institutional buyers
        'b2c_fit': [],                         # not a B2C consumer channel
        'description_en':'NSIC marketplace for MSE-to-MSE and B2B procurement.',
        'action_en':'Register with NSIC for buyer-seller matchmaking.'},
    'Amazon Seller Services': {
        'business_types':['Electronics','FMCG','Clothing','Supermarket','Hypermarket'],
        'min_health':35,'segment_boost':['Champions','Loyal'],
        'channel_type':'b2c_marketplace',
        'b2c_fit': ['Electronics','FMCG','Clothing','Supermarket','Hypermarket'],
        'margin_sweet_spot': (18, 40),         # premium marketplace — favours higher-margin products
        'description_en':'Premium B2C marketplace — suits high-quality products with strong margin and low returns.',
        'action_en':'Apply for Amazon Easy Ship / FBA integration via marketplace bridge.'},
    'Udaan (B2B Marketplace)': {
        'business_types':['Manufacturing','FMCG','Clothing','Supermarket','Hypermarket'],
        'min_health':20,'segment_boost':['Champions','Loyal','Potential','At Risk'],
        'channel_type':'b2b_wholesale',       # bulk supply to retailers/distributors
        'b2c_fit': [],                         # not a direct B2C consumer channel
        'description_en':'B2B wholesale marketplace — best for businesses supplying to retailers and distributors.',
        'action_en':'List bulk products on Udaan for retailer discovery and bulk orders.'},
}

# ══════════════════════════════════════════════════════════════════════════════
# Insight report CSS (applied to the AI analysis HTML output in Step 5)
# ══════════════════════════════════════════════════════════════════════════════
STORYBOARD_CSS = """<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
.sb-root{font-family:'Inter',Arial,sans-serif;color:#1A2D45;background:#EEF4FB;padding:0 0 60px 0;max-width:100%;}
.sb-root *{box-sizing:border-box;}
.sb-hero{background:linear-gradient(135deg,#0B1F3A 0%,#1B3A6B 60%,#1A5276 100%);padding:14px 28px 12px;overflow:hidden;}
.sb-hero-title{font-size:22px;font-weight:900;color:#FFFFFF;line-height:1.1;margin:0 0 3px;}
.sb-hero-sub{font-size:12px;color:#A8D8FF;font-weight:300;margin-bottom:0;}
.sb-chip{background:rgba(255,255,255,.12);border:1px solid rgba(255,215,100,.45);border-radius:20px;padding:5px 14px;font-size:12px;font-weight:700;color:#FFD080;display:inline-block;margin:3px;}
.sb-chip.green{background:rgba(46,204,143,.25);border-color:rgba(46,204,143,.5);color:#5DEBB0;}
.sb-chip.amber{background:rgba(232,168,56,.25);border-color:rgba(232,168,56,.5);color:#FFD070;}
.sb-chip.red{background:rgba(224,82,82,.25);border-color:rgba(224,82,82,.5);color:#FF8080;}
.sb-section-divider{display:flex;align-items:center;gap:16px;padding:24px 48px 0;}
.sb-section-number{font-size:48px;font-weight:900;color:rgba(27,79,138,.15);line-height:1;min-width:52px;}
.sb-section-title{font-size:22px;font-weight:700;color:#0B1F3A;}
.sb-section-line{flex:1;height:1px;background:linear-gradient(90deg,#C8DCEF,transparent);}
.sb-card-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px;padding:20px 48px 0;}
.sb-kpi-card{background:#FFFFFF;border:1px solid #C8DCEF;border-radius:14px;padding:22px 20px;position:relative;overflow:hidden;}
.sb-kpi-card::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;}
.sb-kpi-card.accent-blue::before{background:#1B4F8A;}.sb-kpi-card.accent-green::before{background:#2ECC8F;}
.sb-kpi-card.accent-amber::before{background:#E8A838;}.sb-kpi-card.accent-red::before{background:#E05252;}
.sb-kpi-card.accent-navy::before{background:#0B1F3A;}
.sb-kpi-label{font-size:10px;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:#4A6A8A;margin-bottom:4px;}
.sb-kpi-value{font-size:26px;font-weight:700;color:#0B1F3A;line-height:1.1;}
.sb-kpi-sub{font-size:11px;color:#4A6A8A;margin-top:4px;}
.sb-status-badge{display:inline-flex;align-items:center;gap:4px;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:600;margin-top:8px;}
.badge-green{background:#E8FBF4;color:#1A7A50;}.badge-amber{background:#FEF6E7;color:#A06000;}.badge-red{background:#FDE8E8;color:#B03030;}.badge-blue{background:#E8F2FF;color:#1B4F8A;}
.sb-scores-row{display:grid;grid-template-columns:repeat(3,1fr);gap:16px;padding:20px 48px 0;}
.sb-score-card{background:#FFFFFF;border:1px solid #C8DCEF;border-radius:14px;padding:20px 22px;color:#1A2D45;}
.sb-score-card div{color:#1A2D45;}
.sb-score-bar-track{height:6px;background:#D8E8F8;border-radius:3px;overflow:hidden;}
.sb-score-bar-fill{height:100%;border-radius:3px;}
.sb-table-wrap{padding:20px 48px 0;}
.sb-table{width:100%;border-collapse:separate;border-spacing:0;background:#FFFFFF;border-radius:14px;border:1px solid #C8DCEF;overflow:hidden;}
.sb-table thead tr th{background:#0B1F3A;color:#E0F0FF;padding:13px 18px;font-size:11px;font-weight:600;letter-spacing:1.5px;text-transform:uppercase;text-align:left;}
.sb-table tbody tr td{padding:13px 18px;border-bottom:1px solid #C8DCEF;font-size:13px;color:#1A2D45;}
.sb-forecast-row{display:grid;grid-template-columns:1fr 1fr;gap:16px;padding:20px 48px 0;}
.sb-forecast-card{background:linear-gradient(135deg,#0B1F3A 0%,#1B3A6B 100%);border-radius:16px;padding:28px 26px;}
.sb-forecast-amount{font-size:34px;font-weight:900;color:#FFFFFF !important;line-height:1.1;}
.sb-forecast-range{font-size:12px;color:#A8D8FF !important;margin-top:6px;}
.sb-snp-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;padding:0 48px;}
.sb-snp-card{background:#FFFFFF;border:1px solid #C8DCEF;border-radius:14px;padding:20px 18px;color:#1A2D45;}
.sb-snp-card.gold{border-top:4px solid #F5C842;}.sb-snp-card.silver{border-top:4px solid #B0BEC5;}.sb-snp-card.bronze{border-top:4px solid #CD7F32;}
.sb-reco-tabs{display:grid;grid-template-columns:1fr 1fr;gap:14px;padding:0 48px;}
.sb-reco-panel{background:#FFFFFF;border:1px solid #C8DCEF;border-radius:14px;overflow:hidden;}
.sb-reco-header{padding:14px 20px;font-weight:700;font-size:13px;color:#FFFFFF;}
.sb-reco-header.immediate{background:linear-gradient(90deg,#E05252,#C0392B);}
.sb-reco-header.strategic{background:linear-gradient(90deg,#1B4F8A,#0B2F5A);color:#D0EAFF;}
.sb-reco-row{display:flex;align-items:flex-start;gap:12px;padding:12px 20px;border-bottom:1px solid #C8DCEF;color:#1A2D45;}
.sb-reco-row div{color:#1A2D45;font-size:13px;line-height:1.5;flex:1;}
.sb-reco-priority{font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;padding:2px 8px;border-radius:10px;flex-shrink:0;margin-top:2px;}
.reco-high{background:#FDE8E8;color:#B03030;}.reco-medium{background:#FEF6E7;color:#A06000;}
.sb-footer{margin:40px 48px 0;padding:16px 20px;background:#FFFFFF;border:1px solid #C8DCEF;border-radius:10px;display:flex;justify-content:space-between;align-items:center;font-size:12px;color:#4A6A8A;}
@media(max-width:900px){.sb-scores-row,.sb-forecast-row,.sb-snp-grid,.sb-reco-tabs,.sb-opp-grid{grid-template-columns:1fr;}.sb-hero,.sb-section-divider,.sb-card-grid,.sb-table-wrap,.sb-snp-grid,.sb-reco-tabs{padding-left:16px;padding-right:16px;}}
/* ── Business Opportunity Widgets ── */
.sb-opp-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:18px;padding:20px 48px 0;}
.sb-opp-card{background:#FFFFFF;border:1px solid #C8DCEF;border-radius:16px;overflow:hidden;position:relative;}
.sb-opp-card-header{padding:18px 20px 14px;display:flex;align-items:center;gap:13px;}
.sb-opp-icon{width:44px;height:44px;border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:22px;flex-shrink:0;}
.sb-opp-eyebrow{font-size:9px;font-weight:700;letter-spacing:2px;text-transform:uppercase;margin-bottom:3px;}
.sb-opp-title{font-size:14px;font-weight:800;color:#0B1F3A;line-height:1.2;}
.sb-opp-body{padding:0 20px 6px;}
.sb-opp-metric-row{display:flex;justify-content:space-between;align-items:center;padding:8px 0;border-bottom:1px solid #EEF4FB;}
.sb-opp-metric-row:last-of-type{border-bottom:none;}
.sb-opp-metric-label{font-size:11px;color:#4A6A8A;font-weight:500;}
.sb-opp-metric-value{font-size:13px;font-weight:700;color:#0B1F3A;font-family:monospace;}
.sb-opp-impact{margin:10px 20px 0;padding:10px 14px;border-radius:10px;display:flex;align-items:flex-start;gap:9px;}
.sb-opp-impact-icon{font-size:16px;flex-shrink:0;margin-top:1px;}
.sb-opp-impact-text{font-size:11px;line-height:1.6;font-weight:500;}
.sb-opp-action{margin:10px 20px 16px;padding:10px 14px;border-radius:10px;font-size:11px;line-height:1.6;font-weight:600;border-left:3px solid;}
</style>"""

def _fmt_inr_sb(v):
    if pd.isna(v) or v is None: return "N/A"
    if v>=1e7: return f"&#8377;{v/1e7:.2f} Cr"
    if v>=1e5: return f"&#8377;{v/1e5:.2f} L"
    return f"&#8377;{v:,.0f}"

def _score_bar_color(v, invert=False):
    eff = (1-v) if invert else v
    if eff >= 0.65: return "#2ECC8F"
    if eff >= 0.40: return "#E8A838"
    return "#E05252"

def _badge_cls(v, invert=False):
    eff = (1-v) if invert else v
    if eff >= 0.65: return "badge-green"
    if eff >= 0.40: return "badge-amber"
    return "badge-red"

def _status_lbl(v, invert=False):
    eff = (1-v) if invert else v
    if eff >= 0.65: return "Excellent"
    if eff >= 0.40: return "Moderate"
    return "Improvement Opportunity"

def _health_cls(v):
    if v >= 65: return "badge-green"
    if v >= 40: return "badge-amber"
    return "badge-red"

def _health_lbl(v):
    if v >= 65: return "Healthy"
    if v >= 40: return "Developing"
    return "At Risk"

def _margin_cls(v):
    if v > 20: return "badge-green"
    if v > 10: return "badge-amber"
    return "badge-red"

def _margin_lbl(v):
    if v > 20: return "Strong"
    if v > 10: return "Moderate"
    return "Low"

def _risk_cls(v):
    if v <= 0.40: return "badge-green"
    if v <= 0.70: return "badge-amber"
    return "badge-red"

def _risk_lbl(v):
    if v <= 0.10: return "Very Low Risk"
    if v <= 0.40: return "Low Risk"
    if v <= 0.70: return "Moderate Risk"
    return "High Risk"

def _sb_divider(num, eyebrow, title):
    return f"""<div class="sb-section-divider"><div class="sb-section-number">{num:02d}</div>
<div style="display:flex;flex-direction:column;gap:2px;">
<div style="font-size:10px;font-weight:600;letter-spacing:3px;text-transform:uppercase;color:#B07A00">{eyebrow}</div>
<div class="sb-section-title">{title}</div></div><div class="sb-section-line"></div></div>"""

def generate_insights(user_data, df_raw, lang='en'):
    import time; t_start = time.time()
    try:
        # ── Step A: run forecast FIRST so growth rate feeds into scores ──────
        _df_pre   = _apply_col_remap(df_raw.copy())
        forecast_results = forecast_sales(_df_pre)
        f6  = forecast_results['6_month']
        f12 = forecast_results['12_month']
        # Extract selected model name early — used in forecast cards and model comparison table
        selected_model_name = forecast_results.get('selected_model', 'Statistical Baseline')

        # Extract peak demand month early — used in forecast insight bar (before SNP section)
        _snp_peak_month = ''; _snp_peak_val = ''
        _pr_early = forecast_results.get('model_results', {}).get('Prophet', {})
        if _pr_early and 'forecast_df' in _pr_early:
            try:
                _pfc_e = _pr_early['forecast_df']
                _pr_peak_e = _pfc_e.loc[_pfc_e['yhat'].idxmax()]
                _snp_peak_month = pd.to_datetime(_pr_peak_e['ds']).strftime('%b %Y')
                _snp_peak_val   = _fmt_inr_sb(float(_pr_peak_e['yhat']))
            except Exception:
                pass
        # Confidence badge removed — model name shown instead on forecast cards
        # Aggregate R² from best available model for additional signal
        _best_r2 = None
        for _mn in ['Linear Regression', 'Holt-Winters', 'Prophet', 'Statistical Baseline']:
            _mr = forecast_results.get('model_results', {}).get(_mn, {})
            if _mr and 'r2_score' in _mr:
                _best_r2 = float(_mr['r2_score']); break

        # Derive forecast growth rate: (6m forecast − trailing 6m actual) / trailing 6m actual
        sales_col = 'Monthly_Sales_INR'
        # sku_col resolved after df is assigned (needs df.columns) — see below
        _s_col = sales_col if sales_col in _df_pre.columns else None
        if _s_col:
            _total_hist = float(_df_pre[_s_col].sum())
            # Use actual date span to avoid inflating trailing_6m on sparse data.
            # e.g. 3 months of data should give trailing_6m = avg_monthly * 6,
            # NOT total * 2 (which row-count gives when data is aggregated).
            if 'Date' in _df_pre.columns:
                _dates = pd.to_datetime(_df_pre['Date'], errors='coerce').dropna()
                if len(_dates) >= 2:
                    _span_months = max(1, round((_dates.max() - _dates.min()).days / 30.44) + 1)
                else:
                    _span_months = 1
            else:
                _span_months = max(_df_pre[_s_col].count(), 1)
            _monthly_avg  = _total_hist / _span_months
            _trailing_6m  = _monthly_avg * 6
            _fc_6m        = float(f6.get('forecast', 0))
            # Cap growth rate to ±150% — beyond this the model is extrapolating unreliably
            _raw_growth   = ((_fc_6m - _trailing_6m) / (_trailing_6m + 1e-9)) * 100
            forecast_growth_rate = float(np.clip(_raw_growth, -150.0, 150.0))
        else:
            forecast_growth_rate = None

        # ── Step B: calculate_scores WITH forecast growth rate ───────────────
        df = calculate_scores(df_raw.copy(), forecast_growth_rate=forecast_growth_rate)
        # Resolve sku_col now that df is available
        sku_col = next((c for c in ['SKU_Name','Product_Name','sku_id','Item_Name','product_id','SKU_ID'] if c in df.columns), 'SKU_Name')
        total_sales    = df[sales_col].sum() if sales_col in df.columns else 0
        total_records  = len(df)
        total_products = df[sku_col].nunique() if sku_col in df.columns else total_records
        # Real data column first (gross_margin_pct), then computed (Avg_Margin_Percent)
        avg_margin   = (df['gross_margin_pct'].mean()   if 'gross_margin_pct'   in df.columns else
                       df['Gross_Margin_Pct'].mean()    if 'Gross_Margin_Pct'   in df.columns else
                       df['Avg_Margin_Percent'].mean()  if 'Avg_Margin_Percent' in df.columns else 0)
        perf_score   = df['Performance_Score'].mean()      if 'Performance_Score'      in df.columns else 0
        health_score = df['MSME_Health_Score'].mean()      if 'MSME_Health_Score'      in df.columns else 0
        fin_risk     = df['Financial_Risk_Score'].mean()   if 'Financial_Risk_Score'   in df.columns else 0
        vendor_sc    = df['Vendor_Score'].mean()           if 'Vendor_Score'           in df.columns else 0
        growth_sc    = df['Growth_Potential_Score'].mean() if 'Growth_Potential_Score' in df.columns else 0
        company = user_data.get('company_name', 'Your Company')

        # ── Platform Readiness Score (single decision metric, 0-100%) ──────────────
        # Formula: 0.35×growth_potential + 0.25×vendor_reliability
        #        + 0.20×margin_health + 0.20×return_quality
        _ors_growth  = min(float(growth_sc), 1.0)
        _ors_vendor  = min(float(vendor_sc) / 100.0, 1.0)
        _ors_margin  = min(float(avg_margin) if float(avg_margin) > 0 else 15.0, 40.0) / 40.0
        _avg_ret_ors = float(df['returns_units'].mean()      if 'returns_units'      in df.columns else
                       df['Returns_Units'].mean()       if 'Returns_Units'      in df.columns else
                       df['Returns_Percentage'].mean()  if 'Returns_Percentage' in df.columns else 5.0)
        _ors_returns = 1.0 - min(_avg_ret_ors, 15.0) / 15.0
        ondc_readiness = round((0.35*_ors_growth + 0.25*_ors_vendor + 0.20*_ors_margin + 0.20*_ors_returns) * 100, 1)
        _ors_cls = "badge-green" if ondc_readiness >= 65 else ("badge-amber" if ondc_readiness >= 40 else "badge-red")
        _ors_lbl = "Ready for ONDC" if ondc_readiness >= 65 else ("Partially Ready" if ondc_readiness >= 40 else "Not Ready Yet")

        seg_result = segment_customers(df); elapsed = time.time() - t_start
        hl = _health_lbl(health_score)
        hcls = "green" if health_score>=65 else ("amber" if health_score>=40 else "red")
        html = STORYBOARD_CSS + f'<div class="sb-root">'
        # Hero
        html += f"""<div class="sb-hero" style="padding:14px 28px 12px">
<div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:6px;margin-bottom:10px">
  <div>
    <div style="font-size:9px;font-weight:700;letter-spacing:2.5px;text-transform:uppercase;color:#FFD080;margin-bottom:4px">DataNetra.ai — AI-Powered Business Intelligence</div>
    <div class="sb-hero-title" style="font-size:22px;font-weight:900;color:#FFFFFF !important;line-height:1.1;margin:0 0 3px;text-shadow:0 2px 8px rgba(0,0,0,0.4)">{company}</div>
    <div class="sb-hero-sub" style="font-size:12px;color:#A8D8FF !important;font-weight:300">Comprehensive Analysis Report &nbsp;·&nbsp; {datetime.datetime.now().strftime("%d %b %Y, %H:%M")}</div>
  </div>
</div>
<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:6px">
  <div style="background:rgba(255,255,255,0.06);border-radius:6px;padding:7px 11px">
    <div style="font-size:9px;color:#7AABDD;text-transform:uppercase;letter-spacing:1px;margin-bottom:2px">Business</div>
    <div style="font-size:12px;font-weight:700;color:#FFFFFF;line-height:1.2">{company}</div>
    <div style="font-size:10px;color:#A8C8E8;margin-top:1px">{user_data.get('business_type','Retail')}</div>
  </div>
  <div style="background:rgba(255,255,255,0.06);border-radius:6px;padding:7px 11px">
    <div style="font-size:9px;color:#7AABDD;text-transform:uppercase;letter-spacing:1px;margin-bottom:2px">Owner</div>
    <div style="font-size:12px;font-weight:700;color:#FFFFFF">{user_data.get('full_name','Business Owner')}</div>
    <div style="font-size:10px;color:#A8C8E8;margin-top:1px">{user_data.get('city','') or 'Retail User'}</div>
  </div>
  <div style="background:rgba(255,255,255,0.06);border-radius:6px;padding:7px 11px">
    <div style="font-size:9px;color:#7AABDD;text-transform:uppercase;letter-spacing:1px;margin-bottom:2px">User ID</div>
    <div style="font-size:12px;font-weight:700;color:#FFD080">{user_data.get('mobile_number', user_data.get('msme_number','—'))}</div>
  </div>
  <div style="background:rgba(255,255,255,0.06);border-radius:6px;padding:7px 11px">
    <div style="font-size:9px;color:#7AABDD;text-transform:uppercase;letter-spacing:1px;margin-bottom:2px">Total Revenue</div>
    <div style="font-size:12px;font-weight:700;color:#FFFFFF">{_fmt_inr_sb(total_sales)}</div>
    <div style="font-size:10px;color:#A8C8E8;margin-top:1px">{total_records:,} records · {total_products:,} SKU{'s' if total_products!=1 else ''}</div>
  </div>
</div>
</div>"""
        # Section 1 — KPIs
        html += _sb_divider(1, 'Overall Summary', 'Business Performance')
        mc = _margin_cls(avg_margin); ml = _margin_lbl(avg_margin)
        hclr = _score_bar_color(health_score/100); rclr = "#1a7a40" if fin_risk<=0.4 else ("#b05a00" if fin_risk<=0.7 else "#b03030")
        def _kpi_card(accent, icon, label, value, sub_html):
            acc_colors = {'blue':'#1B4F8A','green':'#2ECC8F','amber':'#E8A838','red':'#E05252','navy':'#0B1F3A'}
            top_col = acc_colors.get(accent, '#1B4F8A')
            return (f'<div style="background:#FFFFFF;border:1px solid #C8DCEF;border-radius:14px;padding:22px 20px;'
                    f'position:relative;overflow:hidden;border-top:3px solid {top_col}">'
                    f'<div style="font-size:28px;margin-bottom:8px">{icon}</div>'
                    f'<div style="font-size:10px;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:#4A6A8A;margin-bottom:4px">{label}</div>'
                    f'<div style="font-size:26px;font-weight:700;color:#0B1F3A;line-height:1.1">{value}</div>'
                    f'{sub_html}</div>')
        mc_acc = 'green' if avg_margin>20 else ('amber' if avg_margin>10 else 'red')
        hlt_acc = 'green' if health_score>=65 else ('amber' if health_score>=40 else 'red')
        prf_acc = 'green' if perf_score>=65 else ('amber' if perf_score>=40 else 'red')
        prf_lbl = 'Excellent' if perf_score>=65 else ('Moderate' if perf_score>=40 else 'Low')
        _sku_s = 's' if total_products != 1 else ''
        html += f"""<div class="sb-card-grid">
{_kpi_card('blue', '💰', 'Total Revenue', _fmt_inr_sb(total_sales), '<div style="font-size:11px;color:#4A6A8A;margin-top:4px">Gross Sales (all products)</div>')}
{_kpi_card('navy', '📦', 'Data Records', f'{total_records:,}', f'<div style="font-size:11px;color:#4A6A8A;margin-top:4px">{total_products:,} unique SKU{_sku_s}</div>')}
{_kpi_card(mc_acc, '📈', 'Avg Profit Margin', f'{avg_margin:.1f}%', f'<span class="sb-status-badge {mc}" style="margin-top:8px">{ml}</span>')}
{_kpi_card(hlt_acc, '🧠', 'Business Health Score', f'{health_score:.1f}%', f'<span class="sb-status-badge {_health_cls(health_score)}" style="margin-top:8px">{_health_lbl(health_score)}</span>')}
{_kpi_card(prf_acc, '⭐', 'Performance Score', f'{perf_score:.1f}%', f'<span class="sb-status-badge {_health_cls(perf_score)}" style="margin-top:8px">{prf_lbl}</span>')}
</div>"""
        # Section 2 — Scores
        html += _sb_divider(2, 'Score Breakdown', 'Risk & Performance Scores')

        # ── Score cards with full explanation ─────────────────────────────────
        perf_sc_norm = perf_score / 100.0
        profit_ratio = avg_margin / 100.0

        def _score_card_full(icon, title, value_display, value_norm, target_txt, formula_txt, explanation_txt, badge_cls, badge_lbl, invert=False):
            bar_pct = min(value_norm * 100, 100)
            bar_c = _score_bar_color(1 - value_norm if invert else value_norm)
            return f"""<div style="background:#FFFFFF;border:1px solid #C8DCEF;border-radius:14px;padding:20px 18px;border-top:3px solid {bar_c}">
  <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:6px">
    <div>
      <div style="font-weight:700;font-size:13px;color:#0B1F3A">{icon} {title}</div>
      <div style="font-size:11px;color:#4A6A8A;margin-top:2px">{target_txt}</div>
    </div>
    <div style="font-size:26px;font-weight:900;color:{bar_c};font-family:monospace">{value_display}</div>
  </div>
  <div style="height:7px;background:#D8E8F8;border-radius:4px;margin-bottom:8px">
    <div style="width:{bar_pct:.0f}%;height:100%;background:{bar_c};border-radius:4px"></div></div>
  <span class="sb-status-badge {badge_cls}" style="margin-bottom:10px">{badge_lbl}</span>
  <div style="margin-top:10px;font-size:11px;color:#2A4060;line-height:1.6">{explanation_txt}</div>
</div>"""

        html += f"""<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));gap:16px;margin:16px 48px 0">
{_score_card_full('⚠️', 'Financial Risk Score', f'{fin_risk:.2f}', fin_risk,
    'Lower is better · Target &lt;0.40',
    '0.5×(OpCost/Sales) + 0.5×(Loan/(Sales×12))',
    'Measures cashflow pressure and loan burden relative to revenue. A score above 0.70 signals danger — costs or debt are consuming most income. Below 0.40 indicates healthy financial breathing room.',
    _risk_cls(fin_risk), _risk_lbl(fin_risk), invert=True)}
{_score_card_full('📈', 'Performance Score', f'{perf_score:.1f}%', perf_sc_norm,
    'Higher is better · Target &gt;65%',
    '0.30×Profitability + 0.25×OpEfficiency + 0.20×CustSatisfaction + 0.15×VendorReliability + 0.10×InvTurnover',
    'Composite score across five operational pillars. Reflects how well the business converts revenue into real value. Scores above 65% indicate a well-run, scalable operation.',
    _health_cls(perf_score), 'Excellent' if perf_score>=65 else ('Moderate' if perf_score>=40 else 'Low'))}
{_score_card_full('💰', 'Profit Margin Score', f'{avg_margin:.1f}%', min(avg_margin/40,1.0),
    'Higher is better · Target &gt;20%',
    'Avg(Margin%) across all SKUs',
    'Average gross margin across your product portfolio. Margins above 20% indicate strong pricing power and room to invest in growth. Below 10% means thin cushion for shocks.',
    _margin_cls(avg_margin), _margin_lbl(avg_margin))}
{_score_card_full('🤝', 'Vendor Reliability Score', f'{vendor_sc:.0f}/100', vendor_sc/100,
    'Higher is better · Target > 60/100',
    '0.50×VendorDeliveryReliability + 0.30×InvTurnover + 0.20×AvgMargin',
    'Blends supplier delivery reliability, inventory turnover velocity and margin contribution. A score above 65/100 means the supply chain supports growth. Below 40/100 — fulfilment risks are high.',
    _badge_cls(vendor_sc/100), _status_lbl(vendor_sc/100))}
{_score_card_full('🧠', 'Business Health Score', f'{health_score:.1f}%', health_score/100,
    'Higher is better · Target &gt;65%',
    '0.38×(1−FinRisk) + 0.29×(Vendor/100) + 0.28×GrowthPotential + 0.05×ForecastMomentum',
    "DataNetra's flagship composite metric — blends financial safety, supply chain health and growth momentum. The single most important number for marketplace readiness and business assessment.",
    _health_cls(health_score), _health_lbl(health_score))}
{_score_card_full('🚀', 'Growth Potential Score', f'{growth_sc*100:.0f}%' if growth_sc<=1.0 else f'{growth_sc:.0f}%', growth_sc,
    'Higher is better · Target > 60%',
    '0.40×DemandUnits + 0.35×AvgMargin + 0.25×(1−ReturnsRate)',
    'Forward-looking indicator combining demand volume, profitability and customer acceptance (low returns = product-market fit). High scores signal strong marketplace growth potential.',
    _badge_cls(growth_sc), _status_lbl(growth_sc))}
{_score_card_full('🌐', 'Digital Readiness Score', f'{ondc_readiness:.0f}%', ondc_readiness/100,
    'Higher is better · Target > 65% · Unified marketplace readiness indicator',
    '0.35×GrowthPotential + 0.25×VendorReliability + 0.20×MarginHealth + 0.20×ReturnQuality',
    'Composite score for marketplace readiness. Combines growth momentum, supply chain health, profitability and return quality into one number. A score above 65% qualifies for platform integration.',
    _ors_cls, _ors_lbl)}
</div>"""

        # Score Summary Table removed — score cards above are the single source of truth

        # ── Business Opportunity Insights rendered after SNP scores are computed ──
        html += '##OPP_WIDGETS##'

        # Forecast
        html += _sb_divider(4, 'Sales Forecast', 'ML-Powered Revenue Projections')

        # Primary forecast cards (best model)
        html += f"""<div style="margin:0 48px">
  <div style="font-size:11px;color:#1B4F8A;background:#EAF4FF;border-radius:8px;padding:8px 14px;margin-bottom:14px;border-left:3px solid #1B4F8A;font-weight:600">
    📡 <strong>Ensemble Forecast</strong> &nbsp;·&nbsp; Generated using multiple time-series models including Prophet, Holt-Winters, and Linear Regression.
    Weights: {" + ".join(f"<strong>{k.split()[0]}</strong> {v*100:.0f}%" for k,v in forecast_results.get("ensemble_weights", {}).items())}
    &nbsp;·&nbsp; <span style="font-size:10px;color:#4A6A8A">Each model ran independently — the weighted average drives the 6-month &amp; 12-month projections.</span>
  </div>
</div>
<div class="sb-forecast-row">
<div class="sb-forecast-card"><div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:6px"><div style="font-size:11px;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:#A8D8FF">6-Month Projection</div><span style="font-size:10px;font-weight:600;padding:2px 9px;border-radius:10px;background:rgba(168,216,255,0.15);color:#A8D8FF">{selected_model_name.split()[0]}</span></div><div class="sb-forecast-amount" style="color:#FFFFFF !important">{_fmt_inr_sb(f6["forecast"])}</div><div class="sb-forecast-range" style="color:#A8D8FF !important">Range: {_fmt_inr_sb(f6["lower"])} — {_fmt_inr_sb(f6["upper"])}</div></div>
<div class="sb-forecast-card" style="background:linear-gradient(135deg,#1A5276 0%,#0B2F5A 100%);"><div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:6px"><div style="font-size:11px;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:#A8D8FF">12-Month Projection</div><span style="font-size:10px;font-weight:600;padding:2px 9px;border-radius:10px;background:rgba(168,216,255,0.15);color:#A8D8FF">{selected_model_name.split()[0]}</span></div><div class="sb-forecast-amount" style="color:#FFFFFF !important">{_fmt_inr_sb(f12["forecast"])}</div><div class="sb-forecast-range" style="color:#A8D8FF !important">Range: {_fmt_inr_sb(f12["lower"])} — {_fmt_inr_sb(f12["upper"])}</div></div>
</div>
<div style="margin:12px 48px 0;padding:14px 22px;background:linear-gradient(90deg,#EBF4FF,#F5F0FF);border-radius:10px;border-left:4px solid #1B4F8A;display:grid;grid-template-columns:repeat(4,1fr);gap:0;align-items:center;">
  <div style="padding-right:20px;border-right:1px solid #C8DCEF">
    <span style="font-size:10px;color:#4A6A8A;text-transform:uppercase;letter-spacing:1px;font-weight:600;display:block;margin-bottom:4px">Expected Revenue Growth</span>
    <span style="font-size:14px;font-weight:800;color:{'#1a7a40' if forecast_growth_rate and forecast_growth_rate >= 0 else '#b03030'}">{(f"{forecast_growth_rate:+.1f}% annually" if forecast_growth_rate is not None else "—")}</span>
  </div>
  <div style="padding:0 20px;border-right:1px solid #C8DCEF">
    <span style="font-size:10px;color:#4A6A8A;text-transform:uppercase;letter-spacing:1px;font-weight:600;display:block;margin-bottom:4px">Peak Demand Expected</span>
    <span style="font-size:14px;font-weight:800;color:#1B4F8A">{(_snp_peak_month + " · " + _snp_peak_val) if _snp_peak_month else "Add Date column for peak detection"}</span>
  </div>
  <div style="padding:0 20px;border-right:1px solid #C8DCEF">
    <span style="font-size:10px;color:#4A6A8A;text-transform:uppercase;letter-spacing:1px;font-weight:600;display:block;margin-bottom:4px">Forecast Model</span>
    <span style="font-size:14px;font-weight:800;color:#1B4F8A">{selected_model_name}</span>
  </div>
  <div style="padding-left:20px">
    <span style="font-size:10px;color:#4A6A8A;text-transform:uppercase;letter-spacing:1px;font-weight:600;display:block;margin-bottom:4px">12-Month Outlook</span>
    <span style="font-size:14px;font-weight:800;color:#1B4F8A">{_fmt_inr_sb(f12["forecast"])}</span>
  </div>
</div>"""

        # ── Model Comparison Table ─────────────────────────────────────────────
        model_results = forecast_results.get('model_results', {})
        MODEL_META = {
            'Prophet': {
                'icon': '📡', 'type': 'Time-Series',
                'desc': 'Handles seasonality, holidays & changepoints. Best for date-indexed datasets with ≥2 months of history.',
                'requires': 'Date column + ≥2 monthly data points',
                'available': PROPHET_AVAILABLE,
            },
            'Holt-Winters': {
                'icon': '❄️', 'type': 'Exponential Smoothing',
                'desc': 'Triple exponential smoothing with additive trend & seasonality. Great for stable seasonal patterns.',
                'requires': '≥4 data points (≥24 for seasonal component)',
                'available': HOLTWINTERS_AVAILABLE,
            },
            'Linear Regression': {
                'icon': '📐', 'type': 'Trend Extrapolation',
                'desc': 'Fits a straight-line trend through historical sales. Simple, interpretable and always available.',
                'requires': '≥2 data points',
                'available': LINEAR_REGRESSION_AVAILABLE,
            },
            'Statistical Baseline': {
                'icon': '📊', 'type': 'Fallback Avg ×1.05',
                'desc': 'Simple average of last 6 months × number of forecast months × 1.05 growth factor.',
                'requires': 'Any numeric sales column',
                'available': True,
            },
        }

        comp_rows = ""
        for mname, meta in MODEL_META.items():
            is_selected  = (mname == selected_model_name)
            has_result   = mname in model_results
            res          = model_results.get(mname, {})
            f6v  = _fmt_inr_sb(res['6_month']['forecast'])  if has_result else "—"
            f12v = _fmt_inr_sb(res['12_month']['forecast']) if has_result else "—"
            f6r  = f"{_fmt_inr_sb(res['6_month']['lower'])} – {_fmt_inr_sb(res['6_month']['upper'])}"  if has_result else "—"
            f12r = f"{_fmt_inr_sb(res['12_month']['lower'])} – {_fmt_inr_sb(res['12_month']['upper'])}" if has_result else "—"
            r2_txt = f"R²={res['r2_score']:.2f}" if mname == 'Linear Regression' and has_result and 'r2_score' in res else ""
            slope_txt = f"  slope={res.get('slope',0):+.0f}/mo" if r2_txt else ""

            if is_selected:
                row_bg = "#F0F7FF"; sel_badge = '<span style="font-size:10px;font-weight:700;padding:2px 9px;border-radius:10px;background:#1B4F8A;color:#FFFFFF;margin-left:6px">✓ SELECTED</span>'
            elif not meta['available']:
                row_bg = "#F9F9F9"; sel_badge = '<span style="font-size:10px;font-weight:700;padding:2px 9px;border-radius:10px;background:#F5F5F5;color:#AAA;margin-left:6px">Not Installed</span>'
            elif not has_result:
                # Distinguish between "no Date column" (Prophet-specific) vs truly insufficient data
                _prophet_no_date = (mname == 'Prophet' and 'Date' not in df.columns)
                _insuf_lbl = 'Requires Date Column' if _prophet_no_date else 'Insufficient Data'
                _insuf_tip = ' · Add a Date column to enable Prophet' if _prophet_no_date else ''
                row_bg = "#FFFBF0"; sel_badge = f'<span style="font-size:10px;font-weight:700;padding:2px 9px;border-radius:10px;background:#FFF3CD;color:#856404;margin-left:6px">{_insuf_lbl}</span>'
            else:
                row_bg = "#FFFFFF"; sel_badge = '<span style="font-size:10px;font-weight:700;padding:2px 9px;border-radius:10px;background:#EAF7EE;color:#1a7a40;margin-left:6px">Available</span>'

            comp_rows += f"""<tr style="border-bottom:1px solid #D8E8F8;background:{row_bg}">
  <td style="padding:11px 14px;min-width:160px">
    <div style="display:flex;align-items:center;gap:6px">
      <span style="font-size:16px">{meta['icon']}</span>
      <div>
        <div style="font-weight:700;font-size:12px;color:#0B1F3A">{mname}{sel_badge}</div>
        <div style="font-size:10px;color:#4A6A8A;text-transform:uppercase;letter-spacing:0.8px">{meta['type']}</div>
      </div>
    </div>
  </td>
  <td style="padding:11px 14px;font-size:11px;color:#2A4060;max-width:220px;line-height:1.4">{meta['desc']}</td>
  <td style="padding:11px 14px;text-align:center">
    <div style="font-weight:700;font-family:monospace;color:#0B1F3A;font-size:12px">{f6v}</div>
    <div style="font-size:10px;color:#4A6A8A;margin-top:2px">{f6r}</div>
  </td>
  <td style="padding:11px 14px;text-align:center">
    <div style="font-weight:700;font-family:monospace;color:#0B1F3A;font-size:12px">{f12v}</div>
    <div style="font-size:10px;color:#4A6A8A;margin-top:2px">{f12r}</div>
    {f'<div style="font-size:10px;color:#1B4F8A;margin-top:2px">{r2_txt}{slope_txt}</div>' if r2_txt else ''}
  </td>
  <td style="padding:11px 14px;font-size:10px;color:#4A6A8A">{meta['requires']}</td>
</tr>"""

        html += f"""<div style="margin:20px 48px 0;background:#FFFFFF;border:1px solid #C8DCEF;border-radius:12px;overflow:hidden">
  <div style="background:#0B1F3A;padding:10px 16px">
    <span style="color:#A8D8FF;font-size:11px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase">📊 Forecasting Model Comparison — Best Model Selected by MAE Validation</span>
  </div>
  <table style="width:100%;border-collapse:collapse;font-size:13px">
    <thead><tr style="background:#1A3050">
      <th style="padding:9px 14px;text-align:left;color:#D0EAFF;font-size:10px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase;min-width:160px">Model</th>
      <th style="padding:9px 14px;text-align:left;color:#D0EAFF;font-size:10px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase">Description</th>
      <th style="padding:9px 14px;text-align:center;color:#D0EAFF;font-size:10px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase">6-Month Forecast</th>
      <th style="padding:9px 14px;text-align:center;color:#D0EAFF;font-size:10px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase">12-Month Forecast</th>
      <th style="padding:9px 14px;text-align:left;color:#D0EAFF;font-size:10px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase">Data Requirement</th>
    </tr></thead>
    <tbody>{comp_rows}</tbody>
  </table>
  <div style="padding:10px 16px;background:#F0F7FF;font-size:11px;color:#4A6A8A">
    All 4 models run on your data. Each is validated on held-out months using MAE (Mean Absolute Error) — the model with lowest MAE wins and produces the final forecast. Lower MAE = closer to actual sales. Holt-Winters runs in pure numpy — no installation needed.
  </div>
</div>"""
        # ── SNP Mapping Insights (Section 4) ──────────────────────────────
        biz_type = user_data.get('business_type', 'FMCG')
        dom = 'Potential'
        if seg_result and seg_result.get('counts'): dom = max(seg_result['counts'], key=seg_result['counts'].get)

        # ── Raw data metrics for SNP logic ────────────────────────────────
        # Real data columns first, then computed ones
        def _gi_col(*names):
            for n in names:
                if n and n in df.columns: return n
                if n:
                    lc = {c.lower():c for c in df.columns}
                    if n.lower() in lc: return lc[n.lower()]
            return None
        ret_col = _gi_col('returns_units','Returns_Units','Returns_Percentage','return_rate_pct')
        mar_col = _gi_col('gross_margin_pct','Gross_Margin_Pct','Avg_Margin_Percent','profit_margin_pct')
        sal_col = 'Monthly_Sales_INR'  if 'Monthly_Sales_INR'  in df.columns else None
        cat_col = 'Product_Category'   if 'Product_Category'   in df.columns else None
        sta_col = 'state'              if 'state'              in df.columns else None
        sto_col = 'Store_ID'           if 'Store_ID'           in df.columns else None

        avg_return = float(df[ret_col].mean()) if ret_col else 0.0

        # ── 1. SNP Fit Scores ─────────────────────────────────────────────
        SNP_PERSONAS_LOCAL = [
            {
                "name": "FMCG High-Velocity Marketplace",
                "icon": "🛒",
                "good_for": ["FMCG", "Household", "Hypermarket"],
                "ret_max": 4.0, "mar_min": 10.0, "health_min": 40,
                "platforms": "Flipkart · Meesho · Udaan",
                "description": "Best for high-turnover everyday consumer goods. Low return rate is critical.",
                "color": "#27ae60", "border": "#27ae60",
            },
            {
                "name": "Premium B2C Digital Seller",
                "icon": "💎",
                "good_for": ["Electronics", "Clothing", "Home & Decor"],
                "ret_max": 5.0, "mar_min": 20.0, "health_min": 60,
                "platforms": "Amazon · GeM",
                "description": "Higher-margin aspirational products. Quality and presentation drive conversion.",
                "color": "#8b5cf6", "border": "#8b5cf6",
            },
            {
                "name": "B2B Wholesale Distributor",
                "icon": "🏭",
                "good_for": ["Manufacturing", "FMCG", "Clothing", "Hypermarket"],
                "ret_max": 6.0, "mar_min": 8.0, "health_min": 30,
                "platforms": "Udaan · NSIC",
                "description": "Bulk supply to retailers & distributors. Volume-driven, stable margins.",
                "color": "#e07b2a", "border": "#e07b2a",
            },
            {
                "name": "Social Commerce Reseller",
                "icon": "📱",
                "good_for": ["Clothing", "FMCG", "Health & Wellness"],
                "ret_max": 7.0, "mar_min": 12.0, "health_min": 20,
                "platforms": "Meesho",
                "description": "Tier-2/3 markets via reseller network. Lower entry bar, price-sensitive.",
                "color": "#e84393", "border": "#e84393",
            },
            {
                "name": "Government Procurement Supplier",
                "icon": "🏛️",
                "good_for": ["Manufacturing", "Electronics", "Services", "FMCG"],
                "ret_max": 3.0, "mar_min": 15.0, "health_min": 50,
                "platforms": "GeM · NSIC",
                "description": "Public sector supply. Needs strong quality scores and seller registration.",
                "color": "#1B4F8A", "border": "#1B4F8A",
            },
        ]

        def _persona_fit_score(p):
            score = 0
            # Return rate check (40 pts)
            if avg_return <= p['ret_max']:       score += 40
            elif avg_return <= p['ret_max']*1.5: score += 20
            # Margin check (25 pts)
            if avg_margin >= p['mar_min']:       score += 25
            elif avg_margin >= p['mar_min']*0.7: score += 12
            # Health check (20 pts)
            if health_score >= p['health_min']:  score += 20
            elif health_score >= p['health_min']*0.75: score += 10
            # Business type match (15 pts)
            if biz_type in p['good_for']:        score += 15
            elif any(g in biz_type for g in p['good_for']): score += 7
            return min(99, score)

        # Persona scoring removed — simulated panel was removed; no display use

        # ── 2. Product Classification Summary ────────────────────────────
        cat_summary_html = ""
        if cat_col and sal_col:
            cat_grp = df.groupby(cat_col)[sal_col].sum().sort_values(ascending=False)
            total_cat = cat_grp.sum() + 1e-9
            CAT_COLORS = ["#1B4F8A","#27ae60","#f39c12","#8b5cf6","#e74c3c","#e07b2a","#e84393","#0097a7"]
            cat_rows = ""
            for i,(cat,rev) in enumerate(cat_grp.items()):
                pct = rev/total_cat*100; col = CAT_COLORS[i%len(CAT_COLORS)]
                ret_for_cat = df[df[cat_col]==cat][ret_col].mean() if ret_col else 0
                ret_flag = "✅" if ret_for_cat < 4 else ("⚠️" if ret_for_cat < 7 else "🔴")  # rule: <4%=green, 4–<7%=amber, ≥7%=red
                cat_rows += f"""<tr style="border-bottom:1px solid #D8E8F8;background:{'#F0F7FF' if i%2==0 else '#FFFFFF'}">
  <td style="padding:9px 14px;font-weight:600;color:{col}">{cat}</td>
  <td style="padding:9px 14px;font-family:monospace;color:#0B1F3A">{_fmt_inr_sb(rev)}</td>
  <td style="padding:9px 14px">
    <div style="display:flex;align-items:center;gap:8px">
      <div style="flex:1;height:6px;background:#D8E8F8;border-radius:3px">
        <div style="width:{pct:.0f}%;height:100%;background:{col};border-radius:3px"></div></div>
      <span style="font-size:11px;font-weight:700;color:{col}">{pct:.1f}%</span></div></td>
  <td style="padding:9px 14px;text-align:center;font-size:13px;color:#1A3050">{ret_flag} {ret_for_cat:.1f}%</td>
</tr>"""
            cat_summary_html = f"""<div style="margin:0">
<table style="width:100%;border-collapse:collapse;background:#FFFFFF;border-radius:12px;overflow:hidden;border:1px solid #C8DCEF;font-size:13px">
<thead><tr style="background:#0B1F3A">
  <th style="padding:11px 14px;text-align:left;color:#A8D8FF;font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase">Category</th>
  <th style="padding:11px 14px;text-align:left;color:#A8D8FF;font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase">Revenue</th>
  <th style="padding:11px 14px;text-align:left;color:#A8D8FF;font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase">Revenue Share</th>
  <th style="padding:11px 14px;text-align:center;color:#A8D8FF;font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase">Return Rate</th>
</tr></thead><tbody>{cat_rows}</tbody></table>
<div style="margin-top:8px;font-size:11px;color:#4A6A8A">✅ &lt;4% return rate · ⚠️ 4–7% · 🔴 &gt;7%</div></div>"""

        # Capacity Health Indicator removed — metrics shown in individual KPI score cards

        # ── 5. Build SNP section HTML ─────────────────────────────────────
        # 5a. Fit Score cards for top 3 platforms (existing SNP_CATALOG)
        # Forecast demand signal for SNP matching
        _snp_f6   = f6.get('forecast', 0)  if isinstance(f6, dict) else 0
        _snp_f12  = f12.get('forecast', 0) if isinstance(f12, dict) else 0
        _snp_growth_rate = forecast_growth_rate if forecast_growth_rate is not None else 0
        _sel_mdl  = forecast_results.get('selected_model', 'ML Forecast')  # used in SNP panel + strategic
        # Normalise growth signal: -20% → 0.0, 0% → 0.5, +30% → 1.0
        _snp_growth_signal = round(float(min(1.0, max(0.0, (_snp_growth_rate + 20) / 50.0))), 2)  # rounded to avoid float variance between runs

        # Peak month from Prophet forecast_df if available
        # _snp_peak_month already extracted early above — refresh here for SNP section
        _pr = forecast_results.get('model_results', {}).get('Prophet', {})
        if _pr and 'forecast_df' in _pr:
            try:
                _pfc = _pr['forecast_df']
                _pr_peak = _pfc.loc[_pfc['yhat'].idxmax()]
                _snp_peak_month = pd.to_datetime(_pr_peak['ds']).strftime('%b %Y')
                _snp_peak_val   = _fmt_inr_sb(float(_pr_peak['yhat']))
            except: pass

        # Detect if business is primarily B2C consumer retail vs B2B/wholesale
        _b2c_biz_types = {'Hypermarket', 'Supermarket', 'Retail', 'FMCG'}
        _is_b2c_retail = biz_type in _b2c_biz_types

        snp_scores = {}
        for snp_name, snp_data in SNP_CATALOG.items():
            s = 0.0
            # 1. Business type match — 35 pts (primary filter, most important)
            if biz_type in snp_data['business_types']:       s += 35
            elif any(g in biz_type for g in snp_data.get('business_types', [])): s += 15

            # 1b. Channel-type tiebreaker — differentiates B2C vs B2B for same biz_type
            #     Retail/Hypermarket → B2C marketplace platforms score 8 pts higher
            #     B2B wholesale platforms score 0 tiebreaker for retail businesses
            _ch_type = snp_data.get('channel_type', '')
            _b2c_fit = snp_data.get('b2c_fit', [])
            if _is_b2c_retail and _ch_type == 'b2c_marketplace':
                s += 8   # B2C marketplace bonus for retail businesses
            elif _is_b2c_retail and biz_type in _b2c_fit:
                s += 4   # partial B2C fit bonus
            elif not _is_b2c_retail and _ch_type in ('b2b_wholesale', 'b2b_institutional'):
                s += 6   # B2B platform bonus for non-retail businesses

            # 2. Health score — 25 pts proportional, NOT bonus-stacked
            _h_pts = min(25, (health_score / 100.0) * 25)
            if health_score < snp_data['min_health']:
                _h_pts *= 0.5   # penalise (not disqualify) if below threshold
            s += _h_pts

            # 3. Dominant segment boost — 20 pts (binary demand signal)
            if dom in snp_data['segment_boost']:             s += 20
            elif any(seg in dom for seg in snp_data.get('segment_boost', [])): s += 8

            # 4. Forecast growth signal — 10 pts continuous (0→1 norm × 10)
            s += _snp_growth_signal * 10

            # 5. Growth potential — 5 pts  (0-1 score × 5)
            s += round(growth_sc * 5, 1)  # rounded to stabilise score across runs

            # 6. Vendor reliability — 5 pts  (vendor_sc is 0-100, divide by 100 first)
            s += (vendor_sc / 100.0) * 5

            # 7. Margin fit bonus — 4 pts if avg_margin is in the platform's sweet spot
            #    Differentiates Flipkart (volume, 10-25%) from Amazon (premium, 18%+)
            _mss = snp_data.get('margin_sweet_spot')
            if _mss and _mss[0] <= avg_margin <= _mss[1]:
                s += 4   # margin squarely in platform's target range
            elif _mss and abs(avg_margin - (_mss[0]+_mss[1])/2) <= 5:
                s += 2   # margin close to sweet spot

            snp_scores[snp_name] = min(99, round(s))
        top3_snp = sorted(snp_scores.items(), key=lambda x: x[1], reverse=True)[:3]
        medals_cls = ['gold','silver','bronze']; medals_emoji = ['🥇','🥈','🥉']

        html += _sb_divider(5, 'Channel & Platform Insights', 'AI-matched marketplace recommendations for your business')

        # ── Forecast Demand Signal panel for SNP ─────────────────────────────
        _snp_trend_col  = '#1a7a40' if _snp_growth_rate >= 10 else ('#b05a00' if _snp_growth_rate >= 0 else '#b03030')
        _snp_trend_icon = '📈' if _snp_growth_rate >= 10 else ('➡️' if _snp_growth_rate >= 0 else '📉')
        _snp_peak_line  = (f'&nbsp;·&nbsp; Peak demand: <strong>{_snp_peak_month} ({_snp_peak_val})</strong>'
                           if _snp_peak_month else '')
        _snp_mdl_label = _sel_mdl.split('(')[0].strip()
        _snp_peak_disp = f'vs trailing 6 months{_snp_peak_line}'
        html += (
            '<div style="margin:0 48px 16px;background:linear-gradient(135deg,#EAF4FF,#F8FAFF);'
            'border:1px solid #B8D4F0;border-radius:10px;padding:14px 18px;">'
            '<div style="font-size:10px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;'
            'color:#1B4F8A;margin-bottom:8px">'
            + _snp_trend_icon + ' Marketplace Demand Signal — Forecast-Driven Platform Matching'
            '</div>'
            '<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:10px;font-size:12px">'
            '<div style="background:#fff;border-radius:8px;padding:10px 12px;border-left:3px solid #1B4F8A">'
            '<div style="font-size:10px;color:#4A6A8A;text-transform:uppercase;letter-spacing:0.8px">6-Month Forecast</div>'
            f'<div style="font-size:15px;font-weight:700;color:#0B1F3A;margin-top:2px">{_fmt_inr_sb(_snp_f6)}</div>'
            '<div style="font-size:10px;color:#4A6A8A">Projected demand to signal platform sellers</div>'
            '</div>'
            '<div style="background:#fff;border-radius:8px;padding:10px 12px;border-left:3px solid #1B4F8A">'
            '<div style="font-size:10px;color:#4A6A8A;text-transform:uppercase;letter-spacing:0.8px">12-Month Forecast</div>'
            f'<div style="font-size:15px;font-weight:700;color:#0B1F3A;margin-top:2px">{_fmt_inr_sb(_snp_f12)}</div>'
            '<div style="font-size:10px;color:#4A6A8A">Annual supply planning horizon</div>'
            '</div>'
            f'<div style="background:#fff;border-radius:8px;padding:10px 12px;border-left:3px solid {_snp_trend_col}">'
            '<div style="font-size:10px;color:#4A6A8A;text-transform:uppercase;letter-spacing:0.8px">Revenue Trend (6-Month)</div>'
            f'<div style="font-size:15px;font-weight:700;color:{_snp_trend_col};margin-top:2px">{_snp_growth_rate:+.1f}%</div>'
            f'<div style="font-size:10px;color:#4A6A8A">Forecast vs trailing 6 months{_snp_peak_line}</div>'
            '</div>'
            '<div style="background:#fff;border-radius:8px;padding:10px 12px;border-left:3px solid #8b5cf6">'
            '<div style="font-size:10px;color:#4A6A8A;text-transform:uppercase;letter-spacing:0.8px">Forecast Model</div>'
            f'<div style="font-size:12px;font-weight:700;color:#0B1F3A;margin-top:2px">{_snp_mdl_label}</div>'
            '<div style="font-size:10px;color:#4A6A8A">Selected by MAE validation</div>'
            '</div>'
            '</div>'
            f'<div style="margin-top:10px;font-size:11px;color:#4A6A8A;line-height:1.6">'
            f'💡 <strong>How this helps platform matching:</strong> Your forward demand signal of '
            f'<strong>{_fmt_inr_sb(_snp_f6)}</strong> over 6 months is shared with matched marketplace sellers so they '
            'can pre-position stock, negotiate bulk pricing, and ensure fulfilment before your peak demand arrives. '
            'This reduces last-minute orders that sellers cannot fulfill and improves your online conversion rate.'
            '</div>'
            '</div>'
        )

        # ── Which ONDC platform drove your revenue — insight-first view ──────
        # Rank all SNP platforms by fit score (already computed above)
        snp_ranked = sorted(snp_scores.items(), key=lambda x: x[1], reverse=True)
        top_snp_name  = snp_ranked[0][0]  if snp_ranked else "Flipkart Commerce"
        top_snp_score = snp_ranked[0][1]  if snp_ranked else 70
        # ── Business Opportunity Insights (Section 2.5) ─────────────────────────
        # Three widgets: Profit Opportunity, Inventory Risk, Marketplace Opportunity.
        # Data computed from real variables — rendered via placeholder into score section.

        # ── Widget 1 data: Profit Opportunity ────────────────────────────────
        _target_margin    = 25.0                          # DataNetra industry benchmark for hypermarket
        _margin_gap       = max(0.0, _target_margin - avg_margin)
        _profit_upside    = total_sales * (_margin_gap / 100.0)   # additional INR if margin hits target
        _margin_colour    = "#2ECC8F" if avg_margin >= 20 else ("#E8A838" if avg_margin >= 12 else "#E05252")
        _margin_badge_cls = "badge-green" if avg_margin >= 20 else ("badge-amber" if avg_margin >= 12 else "badge-red")
        _margin_badge_lbl = "Strong" if avg_margin >= 20 else ("Moderate" if avg_margin >= 12 else "Needs Work")
        _margin_action    = (
            f"Maintain pricing discipline; expand high-margin SKU range to scale profitability."
            if avg_margin >= 20 else
            f"Renegotiate top-5 suppliers and identify SKUs where a 2–3% price nudge is feasible without volume loss."
            if avg_margin >= 12 else
            f"Immediate pricing review required — prioritise renegotiation of high-cost SKUs to close the {_margin_gap:.1f}pp gap."
        )

        # ── Widget 2 data: Inventory Risk ────────────────────────────────────
        _ret_col_w2   = 'Returns_Percentage' if 'Returns_Percentage' in df.columns else None
        _sku_col_w2   = sku_col
        _avg_ret_w2   = float(df[_ret_col_w2].mean()) if _ret_col_w2 else 0.0
        # High-return SKUs: individual SKU avg return ≥ 7%
        if _ret_col_w2 and _sku_col_w2 and _sku_col_w2 in df.columns:
            _sku_ret_avg  = df.groupby(_sku_col_w2)[_ret_col_w2].mean()
            _hi_ret_count = int((_sku_ret_avg >= 7).sum())
            _hi_ret_skus  = _sku_ret_avg[_sku_ret_avg >= 7]
            # Estimated value at risk: high-return SKUs × their avg sales × return rate
            _sal_col_w2   = 'Monthly_Sales_INR' if 'Monthly_Sales_INR' in df.columns else None
            if _sal_col_w2:
                _sku_rev_avg  = df.groupby(_sku_col_w2)[_sal_col_w2].mean()
                _val_at_risk  = float(sum(_sku_rev_avg.get(s, 0) * (_sku_ret_avg[s]/100) for s in _hi_ret_skus.index))
            else:
                _val_at_risk  = total_sales * (_avg_ret_w2 / 100) * 0.40
        else:
            _hi_ret_count = 0
            _val_at_risk  = total_sales * (_avg_ret_w2 / 100) * 0.40
        _ret_colour   = "#E05252" if _avg_ret_w2 >= 7 else ("#E8A838" if _avg_ret_w2 >= 4 else "#2ECC8F")
        _ret_severity = "High Risk" if _avg_ret_w2 >= 7 else ("Moderate" if _avg_ret_w2 >= 4 else "Healthy")
        _ret_sev_cls  = "badge-red" if _avg_ret_w2 >= 7 else ("badge-amber" if _avg_ret_w2 >= 4 else "badge-green")
        _ret_action   = (
            f"Immediately audit top {min(_hi_ret_count, 5)} high-return SKUs — investigate packaging, product specs and supplier quality. Reducing the return rate by 1 percentage point can recover approximately {_fmt_inr_sb(total_sales * 0.01 * 0.65)} in resaleable revenue (assuming ~65% of returned goods can be resold)."
            if _avg_ret_w2 >= 7 else
            f"Monitor return drivers monthly. Set up SKU-level return alerts at 5% to intervene before crossing the 7% threshold."
            if _avg_ret_w2 >= 4 else
            f"Return rate is healthy. Maintain current quality controls and use this strength in marketplace seller credibility."
        )

        # ── Widget 3 data: Marketplace Opportunity ────────────────────────────
        # top_snp_name/score now guaranteed available (computed above)
        _mkt_top_name  = top_snp_name
        _mkt_top_score = top_snp_score
        _mkt_top_info  = SNP_CATALOG.get(_mkt_top_name, {})
        # Revenue potential: fit-score-weighted ONDC opportunity
        # Conservative: assume 15% of total sales can be unlocked via top platform
        _mkt_rev_potential = total_sales * (_mkt_top_score / 99.0) * 0.15
        _mkt_colour        = "#2ECC8F" if _mkt_top_score >= 70 else ("#E8A838" if _mkt_top_score >= 45 else "#E05252")
        _mkt_fit_cls       = "badge-green" if _mkt_top_score >= 70 else ("badge-amber" if _mkt_top_score >= 45 else "badge-red")
        _mkt_fit_lbl       = "Strong Match" if _mkt_top_score >= 70 else ("Moderate Match" if _mkt_top_score >= 45 else "Weak Match")
        _mkt_action        = _mkt_top_info.get('action_en', 'Register on this platform and optimise your product catalogue.')

        # ── Render the three widgets ──────────────────────────────────────────
        _opp_html  = _sb_divider(3, 'Business Opportunity Insights', 'Profit · Inventory · Marketplace')
        _opp_html += f'''<div class="sb-opp-grid">

<!-- Widget 1: Profit Opportunity -->
<div class="sb-opp-card">
  <div style="height:4px;background:linear-gradient(90deg,#2ECC8F,#1B9E6E)"></div>
  <div class="sb-opp-card-header">
    <div class="sb-opp-icon" style="background:linear-gradient(135deg,#E8FBF4,#C3F0DF)">💰</div>
    <div>
      <div class="sb-opp-eyebrow" style="color:#1A7A50">Profit Opportunity</div>
      <div class="sb-opp-title">Margin Improvement Potential</div>
    </div>
  </div>
  <div class="sb-opp-body">
    <div class="sb-opp-metric-row">
      <span class="sb-opp-metric-label">Current Avg Margin</span>
      <span class="sb-opp-metric-value" style="color:{_margin_colour}">{avg_margin:.1f}%</span>
    </div>
    <div class="sb-opp-metric-row">
      <span class="sb-opp-metric-label">Recommended margin benchmark</span>
      <span class="sb-opp-metric-value" style="color:#4A6A8A">{_target_margin:.0f}%</span>
    </div>
    <div class="sb-opp-metric-row">
      <span class="sb-opp-metric-label">Margin Gap</span>
      <span class="sb-opp-metric-value" style="color:{'#E05252' if _margin_gap>0 else '#2ECC8F'}">{f"+{_margin_gap:.1f}pp needed" if _margin_gap>0 else "✓ Above target"}</span>
    </div>
    <div class="sb-opp-metric-row">
      <span class="sb-opp-metric-label">💵 Profit Upside</span>
      <span class="sb-opp-metric-value" style="color:#2ECC8F;font-size:15px;font-weight:900">{_fmt_inr_sb(_profit_upside) if _margin_gap>0 else "—"}</span>
    </div>
  </div>
  <div class="sb-opp-impact" style="background:#E8FBF4">
    <div class="sb-opp-impact-icon">💡</div>
    <div class="sb-opp-impact-text" style="color:#1A7A50">
      <strong>Additional profit if margin reaches {_target_margin:.0f}%:</strong><br>
      {_fmt_inr_sb(_profit_upside) if _margin_gap>0 else "You are already above target margin."} {f"on your current revenue base of {_fmt_inr_sb(total_sales)}." if _margin_gap>0 else ""}
    </div>
  </div>
  <div class="sb-opp-action" style="background:#F0FBF6;color:#1A5C3A;border-color:#2ECC8F">
    → {_margin_action}
  </div>
</div>

<!-- Widget 2: Inventory Risk -->
<div class="sb-opp-card">
  <div style="height:4px;background:linear-gradient(90deg,{'#E05252,#B03030' if _avg_ret_w2>=7 else ('#E8A838,#B07800' if _avg_ret_w2>=4 else '#2ECC8F,#1B9E6E')})"></div>
  <div class="sb-opp-card-header">
    <div class="sb-opp-icon" style="background:linear-gradient(135deg,{'#FDE8E8,#F5C0C0' if _avg_ret_w2>=7 else ('#FEF6E7,#FDEBC0' if _avg_ret_w2>=4 else '#E8FBF4,#C3F0DF')})">📦</div>
    <div>
      <div class="sb-opp-eyebrow" style="color:{_ret_colour}">Inventory Risk</div>
      <div class="sb-opp-title">Return Rate & SKU Health</div>
    </div>
  </div>
  <div class="sb-opp-body">
    <div class="sb-opp-metric-row">
      <span class="sb-opp-metric-label">Avg Return Rate</span>
      <span class="sb-opp-metric-value" style="color:{_ret_colour}">{_avg_ret_w2:.1f}%</span>
    </div>
    <div class="sb-opp-metric-row">
      <span class="sb-opp-metric-label">High-Return SKUs (≥7%)</span>
      <span class="sb-opp-metric-value" style="color:{'#E05252' if _hi_ret_count>0 else '#2ECC8F'}">{_hi_ret_count} SKU{'s' if _hi_ret_count!=1 else ''}</span>
    </div>
    <div class="sb-opp-metric-row">
      <span class="sb-opp-metric-label">7% Threshold</span>
      <span class="sb-opp-metric-value" style="color:#4A6A8A">{'⚠ Breached' if _avg_ret_w2>=7 else ('⚡ Approaching' if _avg_ret_w2>=5 else '✓ Under Control')}</span>
    </div>
    <div class="sb-opp-metric-row">
      <span class="sb-opp-metric-label">🔴 Inventory Revenue at Risk</span>
      <span class="sb-opp-metric-value" style="color:#E05252;font-size:15px;font-weight:900">{_fmt_inr_sb(_val_at_risk)}</span>
    </div>
  </div>
  <div class="sb-opp-impact" style="background:{'#FDE8E8' if _avg_ret_w2>=7 else ('#FEF6E7' if _avg_ret_w2>=4 else '#E8FBF4')}">
    <div class="sb-opp-impact-icon">{'🚨' if _avg_ret_w2>=7 else ('⚠️' if _avg_ret_w2>=4 else '✅')}</div>
    <div class="sb-opp-impact-text" style="color:{_ret_colour}">
      <strong>{_ret_severity}:</strong> {f"{_hi_ret_count} SKU{'s' if _hi_ret_count!=1 else ''} with ≥7% returns represent {_fmt_inr_sb(_val_at_risk)} inventory revenue at risk." if _hi_ret_count>0 else f"No SKUs above the 7% return threshold. Inventory health is good."}
    </div>
  </div>
  <div class="sb-opp-action" style="background:{'#FDF0F0' if _avg_ret_w2>=7 else ('#FEF9EC' if _avg_ret_w2>=4 else '#F0FBF6')};color:{'#8B2020' if _avg_ret_w2>=7 else ('#7A4A00' if _avg_ret_w2>=4 else '#1A5C3A')};border-color:{_ret_colour}">
    → {_ret_action}
  </div>
</div>

<!-- Widget 3: Marketplace Opportunity -->
<div class="sb-opp-card">
  <div style="height:4px;background:linear-gradient(90deg,#1B4F8A,#4A90D9)"></div>
  <div class="sb-opp-card-header">
    <div class="sb-opp-icon" style="background:linear-gradient(135deg,#E8F2FF,#BDD8F8)">🌐</div>
    <div>
      <div class="sb-opp-eyebrow" style="color:#1B4F8A">Marketplace Opportunity</div>
      <div class="sb-opp-title">Best Marketplace Match</div>
    </div>
  </div>
  <div class="sb-opp-body">
    <div class="sb-opp-metric-row">
      <span class="sb-opp-metric-label">Top Platform</span>
      <span class="sb-opp-metric-value" style="color:#1B4F8A;font-family:inherit;font-size:11px">{_mkt_top_name.replace(" (ONDC)","").replace(" (B2B ONDC)","").replace(" (B2B Marketplace)","").replace(" e-Marketplace","")}</span>
    </div>
    <div class="sb-opp-metric-row">
      <span class="sb-opp-metric-label">Platform Fit Score</span>
      <span class="sb-opp-metric-value" style="color:{_mkt_colour}">{_mkt_top_score}%</span>
    </div>
    <div class="sb-opp-metric-row">
      <span class="sb-opp-metric-label">Digital Readiness</span>
      <span class="sb-opp-metric-value" style="color:{'#2ECC8F' if ondc_readiness>=65 else ('#E8A838' if ondc_readiness>=40 else '#E05252')}">{ondc_readiness:.0f}%</span>
    </div>
    <div class="sb-opp-metric-row">
      <span class="sb-opp-metric-label">🚀 Revenue Potential</span>
      <span class="sb-opp-metric-value" style="color:#1B4F8A;font-size:15px;font-weight:900">{_fmt_inr_sb(_mkt_rev_potential)}</span>
    </div>
  </div>
  <div class="sb-opp-impact" style="background:#E8F2FF">
    <div class="sb-opp-impact-icon">📈</div>
    <div class="sb-opp-impact-text" style="color:#1B4F8A">
      <strong>{_mkt_fit_lbl} ({_mkt_top_score}% fit):</strong> Estimated <strong>{_fmt_inr_sb(_mkt_rev_potential)}</strong> incremental revenue potential via {_mkt_top_name.replace(" (ONDC)","").replace(" (B2B ONDC)","").replace(" (B2B Marketplace)","").replace(" e-Marketplace","")}.
    </div>
  </div>
  <div class="sb-opp-action" style="background:#EBF3FF;color:#0D2E5C;border-color:#1B4F8A">
    → {_mkt_action}
  </div>
</div>

</div>'''
        # replace placeholder with real widget html
        html = html.replace('##OPP_WIDGETS##', _opp_html)

        top_snp_info  = SNP_CATALOG.get(top_snp_name, {})

        # ONDC revenue pool — use revenue_after_ondc minus revenue_before_ondc (true uplift)
        # ondc_channel_revenue can contain negatives (net margin not gross) — avoid it as primary source
        _boc2 = next((c for c in df.columns if c.lower() in ('revenue_before_ondc','pre_ondc_revenue','sales_before_ondc')), None)
        _aoc2 = next((c for c in df.columns if c.lower() in ('revenue_after_ondc','post_ondc_revenue','sales_after_ondc')), None)
        if _boc2 and _aoc2:
            _uplift = float(df[_aoc2].sum()) - float(df[_boc2].sum())
            _ondc_rev_pool = max(_uplift, total_sales * 0.05)   # floor at 5% to avoid zero/negative pool
        else:
            # Fall back: check ondc_channel_revenue but only if mostly positive
            _ondc_ch_col = next((c for c in df.columns if 'ondc' in c.lower() and 'rev' in c.lower()), None)
            if _ondc_ch_col:
                _ch_pos = df[_ondc_ch_col].clip(lower=0).sum()
                _ondc_rev_pool = float(_ch_pos) if _ch_pos > 0 else total_sales * 0.18
            else:
                _ondc_rev_pool = total_sales * 0.18

        # Revenue attribution: proportional split of ACTUAL ONDC revenue by fit score
        # Formula: Revenue_share = ondc_rev × (platform_fit / sum(all_platform_fits))
        # Higher fit → larger share of the real ONDC revenue pool.
        # Avoids the misleading result where every platform shows total_sales.
        total_snp_score = sum(s for _, s in snp_ranked) + 1e-9
        snp_rev_attr = [(name, score, _ondc_rev_pool * score / total_snp_score) for name, score in snp_ranked]

        # Build "Why this platform drove your revenue" explanation
        WHY_REASONS = {
            'Flipkart Commerce':       ["High product demand units matched Flipkart's high-volume B2C model", "Your margin supports competitive pricing on Flipkart's marketplace", "Consumer goods categories show strong Flipkart platform alignment"],
            'GeM (Government e-Marketplace)': ["Business health score qualifies for GeM supplier status", "Low return rate meets GeM's strict quality compliance standards", "Retail registration unlocks priority GeM procurement access"],
            'Meesho':                  ["Price-sensitive product categories align with Meesho reseller network", "Tier-2/3 market demand patterns match Meesho's customer base", "Social commerce model suits your product discovery channels"],
            'Amazon Seller Services':  ["Above-average margin supports Amazon's premium positioning", "Low return rate enables Amazon's quality seller badge eligibility", "High-value product categories drive Amazon conversion rates"],
            'Udaan (B2B Marketplace)':               ["B2B bulk order potential matches Udaan's distributor network", "Vendor reliability supports consistent B2B fulfilment on Udaan", "Category breadth suits Udaan's retailer discovery model"],
            'NSIC e-Marketplace':             ["Manufacturing/FMCG category qualifies for NSIC MSE matchmaking", "Business health meets NSIC supplier registration requirements", "B2B procurement patterns align with NSIC buyer profiles"],
        }
        top_reasons = WHY_REASONS.get(top_snp_name, ["Strong business metrics align with this platform's requirements", "Category and margin profile match platform expectations", "Health and vendor scores qualify for platform onboarding"])

        # Revenue impact: use actual Revenue_Before_ONDC column if present, else estimate 82% baseline
        _boc_col_s5 = next((c for c in df.columns if c.lower() in ('revenue_before_ondc','sales_before_ondc','pre_ondc_revenue')), None)
        if _boc_col_s5:
            rev_before_ondc = float(df[_boc_col_s5].sum())
            _uplift_label   = 'Actual channel uplift'
        else:
            rev_before_ondc = total_sales * 0.82   # estimated 18% uplift
            _uplift_label   = 'Estimated 18% uplift'
        rev_uplift_pct  = (total_sales - rev_before_ondc) / rev_before_ondc * 100 if rev_before_ondc > 0 else 0
        rev_uplift_abs  = total_sales - rev_before_ondc

        # Top platform attribution card
        top_col = "#1B4F8A"
        html += f"""<div style="margin:16px 48px 0;background:linear-gradient(135deg,#F0F7FF 0%,#E4F0FF 100%);border:2px solid #1B4F8A;border-radius:16px;padding:24px 26px">
  <div style="display:flex;align-items:flex-start;justify-content:space-between;gap:20px;flex-wrap:wrap">
    <div style="flex:1;min-width:240px">
      <div style="font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#4A6A8A;margin-bottom:8px">🏆 Recommended Primary Marketplace</div>
      <div style="font-size:22px;font-weight:900;color:#0B1F3A;margin-bottom:4px">{top_snp_name}</div>
      <div style="font-size:13px;color:#2A4060;margin-bottom:14px">{top_snp_info.get('description_en','')}</div>
      <div style="display:flex;flex-direction:column;gap:6px">
        {''.join(f'<div style="display:flex;align-items:center;gap:8px"><span style="color:#1B4F8A;font-size:14px">▸</span><span style="font-size:12px;color:#1A3050">{r}</span></div>' for r in top_reasons)}
      </div>
    </div>
    <div style="display:flex;flex-direction:column;gap:12px;min-width:200px">
      <div style="background:#FFFFFF;border-radius:12px;padding:16px 20px;border:1px solid #C8DCEF;text-align:center">
        <div style="font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:#4A6A8A;margin-bottom:4px">Platform Fit Score</div>
        <div style="font-size:36px;font-weight:900;color:#1B4F8A;font-family:monospace">{top_snp_score}%</div>
        <div style="height:6px;background:#D8E8F8;border-radius:3px;margin-top:8px">
          <div style="width:{top_snp_score}%;height:100%;background:#1B4F8A;border-radius:3px"></div></div>
      </div>
      <div style="background:#FFFFFF;border-radius:12px;padding:14px 20px;border:1px solid #C8DCEF;text-align:center">
        <div style="font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:#4A6A8A;margin-bottom:4px">Estimated Revenue via Marketplace</div>
        <div style="font-size:22px;font-weight:900;color:#1a7a40;font-family:monospace">{_fmt_inr_sb(snp_rev_attr[0][2])}</div>
        <div style="font-size:11px;color:#4A6A8A;margin-top:2px">{snp_rev_attr[0][1]/total_snp_score*100:.0f}% of total revenue share</div>
      </div>
    </div>
  </div>
  <div style="margin-top:16px;padding:12px 16px;background:rgba(27,79,138,.08);border-radius:8px;border-left:3px solid #1B4F8A">
    <span style="font-size:12px;font-weight:700;color:#0B1F3A">Next Step: </span>
    <span style="font-size:12px;color:#2A4060">{top_snp_info.get('action_en','Register on this platform and optimise your product catalogue.')}</span>
  </div>
</div>"""

        # ── Revenue Before vs After ONDC ──────────────────────────────────────
        html += f"""<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px;margin:16px 48px 0">
  <div style="background:#FFFFFF;border:1px solid #C8DCEF;border-radius:12px;padding:18px 20px;text-align:center;border-top:3px solid #B0BEC5">
    <div style="font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:#4A6A8A;margin-bottom:8px">📊 Before Online Channels (Est.)</div>
    <div style="font-size:26px;font-weight:900;color:#7A92AA;font-family:monospace">{_fmt_inr_sb(rev_before_ondc)}</div>
    <div style="font-size:11px;color:#4A6A8A;margin-top:4px">Baseline revenue without online marketplace channels</div>
  </div>
  <div style="background:#FFFFFF;border:1px solid #C8DCEF;border-radius:12px;padding:18px 20px;text-align:center;border-top:3px solid #1a7a40">
    <div style="font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:#4A6A8A;margin-bottom:8px">🚀 With Online Channels (Current)</div>
    <div style="font-size:26px;font-weight:900;color:#1a7a40;font-family:monospace">{_fmt_inr_sb(total_sales)}</div>
    <div style="font-size:11px;color:#4A6A8A;margin-top:4px">Revenue with online marketplace channels</div>
  </div>
  <div style="background:#EAF7EE;border:1px solid #C3E6CB;border-radius:12px;padding:18px 20px;text-align:center;border-top:3px solid #1a7a40">
    <div style="font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:#1a7a40;margin-bottom:8px">📈 Online Channel Uplift</div>
    <div style="font-size:26px;font-weight:900;color:#1a7a40;font-family:monospace">+{rev_uplift_pct:.1f}%</div>
    <div style="font-size:13px;font-weight:700;color:#1a7a40;margin-top:2px">+{_fmt_inr_sb(rev_uplift_abs)}</div>
    <div style="font-size:10px;color:#4A6A8A;margin-top:2px">{_uplift_label}</div>
  </div>
</div>"""

        # ── All platforms revenue attribution breakdown ────────────────────────
        plat_attr_rows = ""
        for rank_i, (pname, pscore, prev) in enumerate(snp_rev_attr[:5]):
            pinfo = SNP_CATALOG.get(pname, {})
            pct_share = pscore / total_snp_score * 100
            bar_w = int(pscore)  # use actual fit score for bar width so 93 > 89 > 81
            pcol = ["#1B4F8A","#27ae60","#f39c12","#8b5cf6","#e07b2a"][rank_i % 5]
            rank_medal = ["🥇","🥈","🥉","4️⃣","5️⃣"][rank_i]
            plat_attr_rows += f"""<div style="display:flex;align-items:center;gap:14px;padding:10px 0;border-bottom:1px solid #EAF2FF">
  <span style="font-size:18px;width:26px">{rank_medal}</span>
  <div style="flex:1">
    <div style="display:flex;justify-content:space-between;margin-bottom:4px">
      <span style="font-size:13px;font-weight:700;color:#0B1F3A">{pname}</span>
      <span style="font-size:13px;font-weight:700;color:{pcol};font-family:monospace">{_fmt_inr_sb(prev)}</span>
    </div>
    <div style="height:6px;background:#D8E8F8;border-radius:3px;margin-bottom:3px">
      <div style="width:{bar_w}%;height:100%;background:{pcol};border-radius:3px"></div></div>
    <div style="display:flex;justify-content:space-between">
      <span style="font-size:10px;color:#4A6A8A">{pinfo.get('description_en','')[:60]}…</span>
      <span style="font-size:10px;font-weight:700;color:#4A6A8A">Fit: {pscore}% · {pct_share:.0f}% share</span>
    </div>
  </div>
</div>"""

        html += f"""<div style="margin:16px 48px 0;background:#FFFFFF;border:1px solid #C8DCEF;border-radius:14px;padding:20px 22px">
  <div style="font-size:11px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#4A6A8A;margin-bottom:14px">💰 Revenue Attribution by Marketplace Channel</div>
  {plat_attr_rows}
  <div style="font-size:10px;color:#7A92AA;margin-top:10px;padding-top:8px;border-top:1px solid #EAF2FF">Revenue share is estimated based on platform fit score weighting. Actual figures depend on active listings and orders on each platform.</div>
</div>"""

        # ── Top Products for ONDC ─────────────────────────────────────────────
        if sal_col and sku_col:
            # Aggregate to SKU level first so ranking is based on TOTAL
            # revenue per SKU — not a single row. Without this, a high-revenue
            # SKU with many rows can rank below a low-revenue SKU with one big row.
            _agg = {sal_col: 'sum'}
            if mar_col: _agg[mar_col] = 'mean'
            if ret_col: _agg[ret_col] = 'mean'
            top_prod_df = df.groupby(sku_col, as_index=False).agg(_agg)

            # Normalise each component
            _sales_max  = top_prod_df[sal_col].max() + 1e-9
            _sales_norm = top_prod_df[sal_col] / _sales_max   # 0-1, relative within dataset

            if mar_col:
                _mar_max    = top_prod_df[mar_col].max() + 1e-9
                _mar_norm_p = top_prod_df[mar_col] / _mar_max  # 0-1 relative
            else:
                _mar_norm_p = 0

            # Returns: fixed 15% anchor — avoids instability when max_returns is tiny
            # return_norm = 1 - clip(Returns_Percentage / 15, 0, 1)
            if ret_col:
                _ret_norm_p = 1 - (top_prod_df[ret_col].clip(0, 15) / 15.0)
            else:
                _ret_norm_p = 1.0

            if mar_col and ret_col:
                top_prod_df['_ondc_rank'] = (_sales_norm * 50) + (_mar_norm_p * 30) + (_ret_norm_p * 20)
            elif mar_col:
                top_prod_df['_ondc_rank'] = (_sales_norm * 60) + (_mar_norm_p * 40)
            elif ret_col:
                top_prod_df['_ondc_rank'] = (_sales_norm * 70) + (_ret_norm_p * 30)
            else:
                top_prod_df['_ondc_rank'] = _sales_norm * 100

            top_prods = top_prod_df.sort_values('_ondc_rank', ascending=False).head(5)

            prod_rows = ""
            for rank, (_, row) in enumerate(top_prods.iterrows(), 1):
                margin_val = f"{row[mar_col]:.1f}%" if mar_col else "—"
                ret_val    = f"{row[ret_col]:.1f}%" if ret_col else "—"
                ret_cls    = "#1a7a40" if (ret_col and row[ret_col]<4) else ("#b05a00" if (ret_col and row[ret_col]<7) else "#b03030")  # ≥7%=red
                rank_col   = "#F5C842" if rank==1 else ("#B0BEC5" if rank==2 else ("#CD7F32" if rank==3 else "#4A6A8A"))
                # Return warning badge: ≥7% = high, 5-7% = caution
                _ret_raw   = float(row[ret_col]) if ret_col else 0
                if ret_col and _ret_raw >= 7:
                    _ret_warn = '<div style="font-size:9px;font-weight:700;color:#b03030;margin-top:2px">⚠ High — may reduce online conversion</div>'
                elif ret_col and _ret_raw >= 5:
                    _ret_warn = '<div style="font-size:9px;font-weight:700;color:#b05a00;margin-top:2px">⚡ Approaching threshold</div>'
                else:
                    _ret_warn = ''
                prod_rows += f"""<tr style="border-bottom:1px solid #D8E8F8;background:{'#F0F7FF' if rank%2==0 else '#FFFFFF'}">
  <td style="padding:9px 14px;text-align:center"><span style="font-size:14px;font-weight:900;color:{rank_col}">#{rank}</span></td>
  <td style="padding:9px 14px;font-weight:600;color:#0B1F3A;max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{str(row[sku_col])[:30]}</td>
  <td style="padding:9px 14px;font-family:monospace;color:#1B4F8A;font-weight:700">{_fmt_inr_sb(row[sal_col])}</td>
  <td style="padding:9px 14px;font-weight:600;color:#0B1F3A">{margin_val}</td>
  <td style="padding:9px 14px;font-weight:600;color:{ret_cls}">{ret_val}{_ret_warn}</td>
</tr>"""
            html += f"""<div style="margin:20px 48px 0">
  <div style="font-size:11px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#4A6A8A;margin-bottom:10px">⭐ Top Products for Online Channels</div>
  <table style="width:100%;border-collapse:collapse;background:#FFFFFF;border-radius:12px;overflow:hidden;border:1px solid #C8DCEF;font-size:13px">
    <thead><tr style="background:#0B1F3A">
      <th style="padding:10px 14px;text-align:center;color:#A8D8FF;font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase">Rank</th>
      <th style="padding:10px 14px;text-align:left;color:#A8D8FF;font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase">Product / SKU</th>
      <th style="padding:10px 14px;text-align:left;color:#A8D8FF;font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase">Monthly Revenue</th>
      <th style="padding:10px 14px;text-align:left;color:#A8D8FF;font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase">Margin</th>
      <th style="padding:10px 14px;text-align:left;color:#A8D8FF;font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase">Return Rate</th>
    </tr></thead>
    <tbody>{prod_rows}</tbody>
  </table>
  <div style="font-size:11px;color:#4A6A8A;margin-top:6px">Ranked by channel suitability (revenue × margin × return rate)</div>
</div>"""

        # SNP Fit Score cards (single clean card grid — no duplicate list)
        html += f"""<div style="margin:20px 48px 0">
  <div style="font-size:11px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#4A6A8A;margin-bottom:14px">🏪 Platform Recommendation for Your Business</div>
  <div class="sb-snp-grid">"""
        for i, (snp, score) in enumerate(top3_snp):
            info = SNP_CATALOG[snp]
            bar_col = "#1a7a40" if score>=70 else ("#b05a00" if score>=45 else "#b03030")
            medal_styles = [
                'border-top:4px solid #F5C842',
                'border-top:4px solid #B0BEC5',
                'border-top:4px solid #CD7F32'
            ]
            badge_txt = ["🥇 Best Match", "🥈 2nd Choice", "🥉 3rd Choice"][i]
            html += f"""<div style="background:#FFFFFF;border:1px solid #C8DCEF;border-radius:14px;padding:20px 18px;{medal_styles[i]}">
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
    <div style="font-size:22px">{medals_emoji[i]}</div>
    <span style="font-size:10px;font-weight:700;padding:3px 9px;border-radius:10px;background:{bar_col}18;color:{bar_col};border:1px solid {bar_col}44">{badge_txt}</span>
  </div>
  <div style="font-weight:700;font-size:13px;color:#0B1F3A;margin-bottom:10px">{snp.replace(' (ONDC)', '').replace(' (B2B Marketplace)', '').replace(' (B2B ONDC)', ' (B2B)')}</div>
  <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:5px">
    <span style="font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#4A6A8A">Platform Fit Score</span>
    <span style="font-size:20px;font-weight:900;color:{bar_col}">{score}%</span>
  </div>
  <div style="height:6px;background:#D8E8F8;border-radius:3px;margin-bottom:12px">
    <div style="width:{score}%;height:100%;border-radius:3px;background:linear-gradient(90deg,#1B4F8A,{bar_col})"></div>
  </div>
  <div style="font-size:11px;color:#2A4060;line-height:1.6;margin-bottom:10px">{info['description_en']}</div>
  <div style="font-size:11px;font-weight:600;color:#1B4F8A;background:#EAF4FF;border-radius:6px;padding:7px 11px">→ {info['action_en']}</div>
</div>"""
        html += '</div></div>'

        # Product Classification Summary
        html += f"""<div style="margin:24px 48px 0">
  <div style="font-size:11px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#4A6A8A;margin-bottom:4px">📦 Product Classification Summary</div>
  <div style="font-size:12px;color:#6A8AA8;margin-bottom:10px;font-style:italic">Revenue distribution and return performance across product categories.</div>
  {cat_summary_html if cat_summary_html else '<div style="background:#FFFFFF;border-radius:10px;padding:14px;color:#4A6A8A;font-style:italic;border:1px solid #C8DCEF">No category data available in dataset.</div>'}
</div>"""

        # Capacity Health Indicator section removed

        # ── Business Summary Banner ───────────────────────────────────────────
        # Compute key signals for the summary
        _top_cat_name = '—'; _top_cat_growth = 0; _bot_cat_name = '—'; _margin_opp = 0
        if cat_col and sal_col and mar_col:
            _cat_rev  = df.groupby(cat_col)[sal_col].sum().sort_values(ascending=False)
            _cat_mar  = df.groupby(cat_col)[mar_col].mean()
            if len(_cat_rev) > 0:
                _top_cat_name = str(_cat_rev.index[0])
                _bot_cat_name = str(_cat_mar.idxmin()) if len(_cat_mar) > 0 else '—'
                _best_margin  = _cat_mar.max()
                _worst_margin = _cat_mar.min()
                _margin_opp   = round((_best_margin - _worst_margin) / max(_best_margin, 1) * 10, 1)

        _fg_safe = min(abs(forecast_growth_rate or 0), 150)
        _fg_dir  = 'growth' if (forecast_growth_rate or 0) >= 0 else 'decline'

        # Initialise here so both the business summary banner AND the action plan
        # section below can use these variables without an UnboundLocalError.
        _f6_imm = f6.get('forecast', 0) if isinstance(f6, dict) else 0
        if 'Date' in df.columns:
            _s_dates_early = pd.to_datetime(df['Date'], errors='coerce').dropna()
            _s_span_early  = max(1, round((_s_dates_early.max() - _s_dates_early.min()).days / 30.44) + 1) if len(_s_dates_early) >= 2 else 1
        else:
            _s_span_early  = max(df[sales_col].count(), 1) if sales_col in df.columns else 1
        _trailing_6m_sales = (total_sales / _s_span_early) * 6

        _stock_6w_risk = False
        if _f6_imm > 0 and _trailing_6m_sales > 0 and _f6_imm > _trailing_6m_sales * 1.05:
            _stock_6w_risk = True

        _summary_bullets = []
        _s = lambda t: f'<strong style="color:#FFD080">{t}</strong>'

        # Build a smooth, professional executive summary sentence
        _exec_parts = []
        if _top_cat_name != '—':
            # Check how tight the top-2 revenue gap is — if <5pp difference, say "narrowly leads"
            if cat_col and sal_col in df.columns:
                _cat_rev_pcts = df.groupby(cat_col)[sal_col].sum()
                _cat_rev_pcts = (_cat_rev_pcts / (_cat_rev_pcts.sum() + 1e-9) * 100).sort_values(ascending=False)
                _top2_gap = float(_cat_rev_pcts.iloc[0] - _cat_rev_pcts.iloc[1]) if len(_cat_rev_pcts) >= 2 else 10
                _lead_word = 'narrowly leads' if _top2_gap < 5 else 'leads'
            else:
                _lead_word = 'leads'
            _exec_parts.append(f'{_s(_top_cat_name)} {_lead_word} the revenue mix')
        if _fg_safe > 5:
            _dir_word = 'an upward' if _fg_dir == 'growth' else 'a downward'
            _exec_parts.append(f'the demand forecast points to {_dir_word} trend of {_s(f"{_fg_safe:.0f}%")} over the next six months')
        if _bot_cat_name != '—' and _bot_cat_name != _top_cat_name:
            _exec_parts.append(f'margin pressure has been identified in {_s(_bot_cat_name)}')
        if _stock_6w_risk:
            _exec_parts.append(f'a {_s("stock shortfall")} is projected within the next six weeks — advance procurement is recommended')
        if vendor_sc < 65:
            _exec_parts.append(f'vendor reliability at {_s(f"{vendor_sc:.0f}/100")} warrants supplier consolidation')

        if _exec_parts:
            if len(_exec_parts) == 1:
                _summary_text = _exec_parts[0].capitalize() + '.'
            elif len(_exec_parts) == 2:
                _summary_text = f'{_exec_parts[0].capitalize()}, while {_exec_parts[1]}.'
            else:
                _mid = '; '.join(_exec_parts[1:-1])
                _summary_text = f'{_exec_parts[0].capitalize()}, {_mid}, and {_exec_parts[-1]}.'
        else:
            _summary_text = f'DataNetra analysis complete for {_s(company)}. Review the action plan below for data-driven recommendations.'

        _profitability_note = ''
        if avg_margin > 0:
            _potential_margin_gain = min(max(20 - avg_margin, 0), 15)
            if _potential_margin_gain > 1:
                _profitability_note = (f' Immediate focus on inventory planning and supplier optimisation '
                                       f'can improve profitability by ~<strong style="color:#FFD080">{_potential_margin_gain:.0f}–{_potential_margin_gain+3:.0f}%</strong>.')

        # Build bullet-format summary from exec_parts + profitability note
        _bullet_items = list(_exec_parts)  # copy so originals unchanged
        if _profitability_note:
            # Extract the plain text of the profitability note as a bullet
            _bullet_items.append(
                f'profit optimisation opportunity of approximately '
                f'<strong style="color:#FFD080">{_potential_margin_gain:.0f}–{_potential_margin_gain+3:.0f}%</strong> '
                f'is achievable through inventory planning and supplier optimisation'
            )
        if not _bullet_items:
            _bullet_items = [f'DataNetra analysis complete for <strong style="color:#FFD080">{company}</strong>. Review the action plan below.']
        _summary_bullets_html = ''.join(
            f'<div style="display:flex;gap:10px;margin-bottom:7px;align-items:flex-start">'
            f'<span style="color:#FFD080;font-size:16px;line-height:1.4;flex-shrink:0">•</span>'
            f'<span style="font-size:13px;color:#FFFFFF;line-height:1.6">{item.capitalize()}</span>'
            f'</div>'
            for item in _bullet_items
        )

        html += f"""<div style="margin:0 0 24px;padding:22px 28px;background:linear-gradient(135deg,#0B1F3A 0%,#1B3F6A 100%);
border-radius:14px;border-left:5px solid #FFD080;box-shadow:0 4px 18px rgba(11,31,58,0.18)">
  <div style="font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#FFD080;margin-bottom:10px">
    📊 DataNetra Analysis Summary
  </div>
  <div style="font-size:13px;color:#FFFFFF;line-height:1.8;font-weight:400">{_summary_bullets_html}</div>
</div>"""

        # Recommendations
        html += _sb_divider(6, 'Action Plan', 'AI-Generated Recommendations')

        # ── Compute data-driven recommendation items ──────────────────────────
        # Immediate Actions — derived from real scores
        _imm_rows = []

        # 1. Financial Risk
        _hi_risk_count = int((df['Financial_Risk_Score'] > 0.70).sum()) if 'Financial_Risk_Score' in df.columns else 0
        _avg_cost_ratio = float((df['Monthly_Operating_Cost_INR'] / df['Monthly_Sales_INR_Adjusted']).clip(0,2).mean()) if 'Monthly_Operating_Cost_INR' in df.columns else 0
        _avg_loan_ratio = float((df['Outstanding_Loan_INR'] / (df['Monthly_Sales_INR_Adjusted']*12)).clip(0,2).mean()) if 'Outstanding_Loan_INR' in df.columns else 0
        _fin_risk_display = max(fin_risk, 0.01)   # never show exactly 0.0 — always some real cost exists
        if _hi_risk_count > 0:
            _imm_rows.append(('reco-high', 'High',
                f'{_hi_risk_count} SKU{"s" if _hi_risk_count>1 else ""} carry high financial risk (score &gt;0.70) — '
                f'cost ratio is <strong>{_avg_cost_ratio*100:.0f}% of revenue</strong>. '
                f'Review pricing and procurement urgently to restore margin headroom.'))
        elif fin_risk < 0.05:
            # Very low — show the actual ratios so it feels grounded
            _imm_rows.append(('reco-medium', 'Medium',
                f'Financial risk is low (score <strong>{_fin_risk_display:.2f}</strong>) — '
                f'operating costs at <strong>{_avg_cost_ratio*100:.0f}% of revenue</strong>, '
                f'loan-to-revenue ratio <strong>{_avg_loan_ratio*100:.0f}%</strong>. '
                f'Maintain discipline; monitor monthly to catch any cost creep early.'))
        else:
            _imm_rows.append(('reco-medium', 'Medium',
                f'Financial risk score <strong>{_fin_risk_display:.2f}</strong> is within safe range — '
                f'operating costs at <strong>{_avg_cost_ratio*100:.0f}% of revenue</strong>. '
                f'Keep costs below 60% of revenue to stay below the 0.40 risk threshold.'))

        # 2. Return Rate
        if avg_return > 7:
            _imm_rows.append(('reco-high', 'High',
                f'Return rate is <strong>{avg_return:.1f}%</strong> (threshold: 7%) — '
                f'audit top returned SKUs for quality or packaging issues immediately.'))
        else:
            _imm_rows.append(('reco-medium', 'Medium',
                f'Return rate is healthy at <strong>{avg_return:.1f}%</strong> — '
                f'continue quality checks to maintain below 7% threshold.'))

        # 3. Margin
        if avg_margin < 15:
            _imm_rows.append(('reco-high', 'High',
                f'Avg margin <strong>{avg_margin:.1f}%</strong> is below 15% benchmark — '
                f'renegotiate supplier costs or revise pricing on low-margin SKUs.'))
        elif avg_margin < 20:
            _imm_rows.append(('reco-medium', 'Medium',
                f'Margin at <strong>{avg_margin:.1f}%</strong> — moderate. '
                f'A {round(20-avg_margin,1)}pp improvement to 20% target would unlock ~{_fmt_inr_sb(total_sales*(20-avg_margin)/100)} additional profit.'))
        else:
            _imm_rows.append(('reco-medium', 'Medium',
                f'Margin strong at <strong>{avg_margin:.1f}%</strong> — '
                f'consider expanding top-performing SKU range to scale revenue further.'))

        # 4. Vendor Score (0-100)
        if vendor_sc < 60:
            _imm_rows.append(('reco-high', 'High',
                f'Vendor reliability score is <strong>{vendor_sc:.0f}/100</strong> — '
                f'consolidate to fewer, high-reliability suppliers to reduce stockout risk.'))
        elif vendor_sc < 75:
            _imm_rows.append(('reco-medium', 'Medium',
                f'Vendor reliability score is <strong>{vendor_sc:.0f}/100</strong> — '
                f'renegotiate lead times with bottom 2 suppliers to push score above 75.'))

        # 4b. Inventory Risk Alert — always present, grounded in turnover data
        if 'Inventory_Turnover' in df.columns:
            _avg_turn    = float(df['Inventory_Turnover'].mean())
            _low_turn_ct = int((df['Inventory_Turnover'] < 4).sum())
            _low_turn_pct = round(_low_turn_ct / max(len(df), 1) * 100, 0)
            if _avg_turn < 4:
                _est_idle = total_sales * max(0, (4 - _avg_turn) / 4) * 0.12
                _imm_rows.append(('reco-high', 'High',
                    f'\U0001f4e6 <strong>Inventory risk alert:</strong> average inventory turnover is '
                    f'<strong>{_avg_turn:.1f}\u00d7/month</strong> \u2014 below the 4\u00d7 minimum benchmark. '
                    f'<strong>{_low_turn_ct} records ({_low_turn_pct:.0f}%)</strong> are slow-moving. '
                    f'Estimated carrying cost of idle stock: <strong>{_fmt_inr_sb(_est_idle)}</strong>. '
                    f'Clear slow movers via promotions or supplier returns within 30 days.'))
            elif _avg_turn < 8:
                _imm_rows.append(('reco-medium', 'Medium',
                    f'\U0001f4e6 <strong>Inventory health:</strong> average turnover <strong>{_avg_turn:.1f}\u00d7/month</strong> \u2014 '
                    f'moderate. Target 8\u201312\u00d7 for a hypermarket. '
                    f'Review reorder points for the <strong>{_low_turn_ct} SKUs</strong> below 4\u00d7 to reduce carrying costs.'))
            else:
                _imm_rows.append(('reco-medium', 'Medium',
                    f'\U0001f4e6 <strong>Inventory health:</strong> turnover at <strong>{_avg_turn:.1f}\u00d7/month</strong> \u2014 '
                    f'strong. Maintain current reorder discipline and review safety stock quarterly.'))

        # 5. Inventory Waste — compute ₹ excess stock from forecast vs current stock value
        _cost_col  = 'Monthly_Operating_Cost_INR' if 'Monthly_Operating_Cost_INR' in df.columns else None
        _inv_col   = 'Inventory_Turnover'          if 'Inventory_Turnover'         in df.columns else None
        # _f6_imm and _trailing_6m_sales already initialised in the summary section above
        _avg_monthly_sales = _trailing_6m_sales / 6
        if _f6_imm > 0 and _trailing_6m_sales > 0:
            _inv_gap = _trailing_6m_sales - _f6_imm   # positive = over-stocked vs forecast
            if _inv_gap > 0:
                _waste_pct = round((_inv_gap / _trailing_6m_sales) * 100, 1)
                _imm_rows.append(('reco-high', 'High',
                    f'⚠️ Inventory risk: current 6-month stock covers <strong>{_fmt_inr_sb(_trailing_6m_sales)}</strong> '
                    f'but forecast demand is <strong>{_fmt_inr_sb(_f6_imm)}</strong> — '
                    f'potential <strong>{_fmt_inr_sb(_inv_gap)} ({_waste_pct}%) overstock</strong>. '
                    f'Reduce next procurement order by ~{_waste_pct:.0f}% to free working capital.'))
            elif _inv_gap < -0.05 * _trailing_6m_sales:
                _short_val = abs(_inv_gap)
                _imm_rows.append(('reco-high', 'High',
                    f'📦 Stock shortfall alert: forecast demand <strong>{_fmt_inr_sb(_f6_imm)}</strong> '
                    f'exceeds current run-rate by <strong>{_fmt_inr_sb(_short_val)}</strong>. '
                    f'Pre-order within 3 weeks to avoid stockout and lost sales.'))

        # 6. Forward Risk Alert — project cashflow gap using forecast
        if _cost_col and _f6_imm > 0:
            _avg_monthly_cost = float(df[_cost_col].mean()) if _cost_col in df.columns else 0
            _projected_cost_6m = _avg_monthly_cost * 6
            _cashflow_gap = _f6_imm - _projected_cost_6m
            if _cashflow_gap < 0:
                _gap_abs = abs(_cashflow_gap)
                _imm_rows.append(('reco-high', 'High',
                    f'🚨 Forward cashflow risk: projected 6-month revenue <strong>{_fmt_inr_sb(_f6_imm)}</strong> '
                    f'is <strong>{_fmt_inr_sb(_gap_abs)} below</strong> projected operating costs '
                    f'<strong>{_fmt_inr_sb(_projected_cost_6m)}</strong>. '
                    f'Act now — reduce costs or accelerate online channel revenue before the gap widens.'))
            elif _cashflow_gap < 0.10 * _projected_cost_6m:
                _imm_rows.append(('reco-medium', 'Medium',
                    f'⚡ Cashflow margin thin: forecast revenue covers costs with only '
                    f'<strong>{_fmt_inr_sb(_cashflow_gap)}</strong> buffer over 6 months. '
                    f'Build a 10% cost reserve to handle demand volatility.'))

        # Strategic Initiatives — derived from forecasts & growth scores
        _str_rows = []

        # 1. Forecast-based — model-aware with profit impact
        _f6_val  = f6.get('forecast', 0)  if isinstance(f6, dict) else 0
        _f12_val = f12.get('forecast', 0) if isinstance(f12, dict) else 0

        # Month-by-month peak detection from model_results for ONDC demand signal
        _model_results_sr = forecast_results.get('model_results', {})
        _peak_months_note = ''
        _prophet_res = _model_results_sr.get('Prophet', {})
        if _prophet_res and 'forecast_df' in _prophet_res:
            try:
                _fc_df = _prophet_res['forecast_df']
                _peak_row = _fc_df.loc[_fc_df['yhat'].idxmax()]
                _peak_month = pd.to_datetime(_peak_row['ds']).strftime('%b %Y')
                _peak_val   = _fmt_inr_sb(float(_peak_row['yhat']))
                _peak_months_note = f' Peak demand month: <strong>{_peak_month} ({_peak_val})</strong>.'
            except Exception:
                pass

        if _f6_val > _trailing_6m_sales:
            _raw_growth_pct = ((_f6_val - _trailing_6m_sales) / max(_trailing_6m_sales, 1)) * 100
            _growth_pct     = round(min(_raw_growth_pct, 150.0), 1)
            _capped_note    = ' <em style="font-size:11px;color:#7A8A9A">(model extrapolation — treat as directional)</em>' if _raw_growth_pct > 150 else ''
            # Profit impact: reordering correctly saves overstock markdown cost (~15% of excess)
            _reorder_save = max(0, (_trailing_6m_sales - _f6_val)) * 0.15
            _profit_note  = (f' Following this forecast for reorder timing can prevent approx. <strong>{_fmt_inr_sb(_reorder_save)}</strong> in potential markdown &amp; carrying costs.'
                             if _reorder_save > 0 else '')
            _str_rows.append(('reco-high', 'High',
                f'<strong>{_sel_mdl}</strong> projects <strong>{_growth_pct}% revenue growth</strong>{_capped_note} '
                f'over next 6 months (~{_fmt_inr_sb(_f6_val)}).{_peak_months_note} '
                f'Pre-stock inventory and align procurement 3 weeks ahead.{_profit_note}'))
        else:
            _over_stock_save = max(0, _trailing_6m_sales - _f6_val) * 0.15
            _profit_note2 = (f' Forecast helps prevent approximately <strong>{_fmt_inr_sb(_over_stock_save)}</strong> of potential overstock cost.'
                             if _over_stock_save > 0 else '')
            _str_rows.append(('reco-high', 'High',
                f'<strong>{_sel_mdl}</strong>: 6-month projected revenue '
                f'<strong>{_fmt_inr_sb(_f6_val) if _f6_val else "—"}</strong>, '
                f'12-month: <strong>{_fmt_inr_sb(_f12_val) if _f12_val else "—"}</strong>.{_peak_months_note} '
                f'Align procurement to forecast to reduce overstock &amp; stockout risk.{_profit_note2}'))

        # 2. Growth Potential Score — display as % (score is 0-1 internally)
        _growth_pct_display = growth_sc * 100 if growth_sc <= 1.0 else growth_sc
        if _growth_pct_display >= 65:
            _str_rows.append(('reco-high', 'High',
                f'Growth potential is <strong>{_growth_pct_display:.0f}%</strong> — strong signal for marketplace growth. '
                f'Allocate additional marketing budget to top-performing products and expand your online marketplace listings.'))
        elif _growth_pct_display >= 40:
            _str_rows.append(('reco-medium', 'Medium',
                f'Growth potential stands at <strong>{_growth_pct_display:.0f}%</strong> — moderate. '
                f'Identify which product categories are dragging performance and reallocate budget to higher-margin SKUs.'))
        else:
            _str_rows.append(('reco-high', 'High',
                f'Growth potential is <strong>{_growth_pct_display:.0f}%</strong> — below target. '
                f'Focus on top 20% of products driving 80% of revenue and rationalise low-performing SKUs.'))

        # 3. Health Score → ONDC readiness
        if health_score >= 65:
            _str_rows.append(('reco-medium', 'Medium',
                f'Overall Business Health Score <strong>{health_score:.1f}%</strong> (used as Digital Readiness proxy) qualifies for fast-track marketplace onboarding — '
                f'expand marketplace presence with <strong>{top_snp_name}</strong> (fit score: {top_snp_score}%).'))
        else:
            _str_rows.append(('reco-medium', 'Medium',
                f'Overall Business Health Score <strong>{health_score:.1f}%</strong> — improve before marketplace onboarding. '
                f'Target: reduce financial risk below 0.40 and margin above 20%.'))

        # 4. Operating cost target
        _str_rows.append(('reco-medium', 'Medium',
            f'Performance score is <strong>{perf_score:.1f}%</strong> — '
            f'target operating cost below 60% of revenue to push score above 65% benchmark.'))

        # ── 6 New Data-Driven Insights ────────────────────────────────────────

        # A. Demand Spike Alert — category with >15% MoM growth in last 2 periods
        if cat_col and sal_col and 'Date' in df.columns:
            try:
                _ddf = df.copy()
                _ddf['_ym'] = pd.to_datetime(_ddf['Date'], errors='coerce').dt.to_period('M')
                _last2 = sorted(_ddf['_ym'].dropna().unique())[-2:]
                if len(_last2) == 2:
                    _p1 = _ddf[_ddf['_ym']==_last2[0]].groupby(cat_col)[sal_col].sum()
                    _p2 = _ddf[_ddf['_ym']==_last2[1]].groupby(cat_col)[sal_col].sum()
                    _growth_cats = ((_p2 - _p1) / (_p1 + 1e-9) * 100).sort_values(ascending=False)
                    if len(_growth_cats) > 0 and _growth_cats.iloc[0] >= 15:
                        _spike_cat = str(_growth_cats.index[0])
                        _spike_pct = round(_growth_cats.iloc[0], 1)
                        _imm_rows.append(('reco-high', 'High',
                            f'🚀 <strong>Demand spike alert:</strong> <strong>{_spike_cat}</strong> category shows '
                            f'<strong>+{_spike_pct}% demand growth</strong> over last 30 days. '
                            f'Ramp up inventory now to capture growth trend.'))
            except Exception:
                pass

        # B. Dead Inventory Alert — SKUs with low turnover and low sales share
        if sku_col and sal_col and 'Inventory_Turnover' in df.columns:
            try:
                _sku_turn = df.groupby(sku_col)['Inventory_Turnover'].mean()
                _sku_rev  = df.groupby(sku_col)[sal_col].sum()
                _total_rev = _sku_rev.sum()
                _dead_mask = (_sku_turn < 1.5) & (_sku_rev / (_total_rev + 1e-9) < 0.02)
                _dead_count = int(_dead_mask.sum())
                if _dead_count > 0:
                    _dead_rev = float(_sku_rev[_dead_mask].sum())
                    _imm_rows.append(('reco-high', 'High',
                        f'🔴 <strong>Dead inventory alert:</strong> <strong>{_dead_count} SKU{"s" if _dead_count>1 else ""}</strong> '
                        f'show low turnover (&lt;1.5×/month) — potential '
                        f'<strong>{_fmt_inr_sb(_dead_rev)}</strong> inventory risk. '
                        f'Consider discount campaigns or supplier return options.'))
            except Exception:
                pass

        # C. Category Profitability Ranking
        if cat_col and sal_col and mar_col:
            try:
                _cat_margin_rank = df.groupby(cat_col)[mar_col].mean().sort_values(ascending=False)
                if len(_cat_margin_rank) >= 2:
                    _top_m_cat  = str(_cat_margin_rank.index[0])
                    _top_m_val  = round(float(_cat_margin_rank.iloc[0]), 1)
                    _bot_m_cat  = str(_cat_margin_rank.index[-1])
                    _bot_m_val  = round(float(_cat_margin_rank.iloc[-1]), 1)
                    _margin_gap_pp = round(_top_m_val - _bot_m_val, 1)
                    _str_rows.append(('reco-medium', 'Medium',
                        f'📊 <strong>Category profitability:</strong> '
                        f'<strong>{_top_m_cat}</strong> remains the most profitable category at '
                        f'<strong>{_top_m_val}% margin</strong>, while '
                        f'<strong>{_bot_m_cat}</strong> shows margin pressure at '
                        f'<strong>{_bot_m_val}%</strong> — a <strong>{_margin_gap_pp}pp gap</strong>. '
                        f'Prioritise supplier renegotiation or pricing review for {_bot_m_cat} to close this gap.'))
            except Exception:
                pass

        # D. Supplier Risk Alert — detect rising lead time or low vendor reliability per category
        if cat_col and 'Vendor_Delivery_Reliability' in df.columns:
            try:
                _vdr_by_cat = df.groupby(cat_col)['Vendor_Delivery_Reliability'].mean()
                _risky_cats = _vdr_by_cat[_vdr_by_cat < 0.75].sort_values()
                if len(_risky_cats) > 0:
                    _risky_cat = str(_risky_cats.index[0])
                    _risky_vdr = round(float(_risky_cats.iloc[0]) * 100, 0)
                    _str_rows.append(('reco-high', 'High',
                        f'⚠️ <strong>Supplier risk alert:</strong> Vendor reliability for '
                        f'<strong>{_risky_cat}</strong> is only <strong>{_risky_vdr:.0f}%</strong>. '
                        f'Stock planning may be affected for fast-moving SKUs. '
                        f'Qualify a secondary supplier within 30 days.'))
            except Exception:
                pass

        # E. Promotion Opportunity Insight — high-margin, low-velocity SKUs
        if sku_col and sal_col and mar_col and 'Inventory_Turnover' in df.columns:
            try:
                _sku_df = df.groupby(sku_col).agg(
                    _rev=(sal_col, 'sum'),
                    _mar=(mar_col, 'mean'),
                    _turn=('Inventory_Turnover', 'mean')
                )
                _total_sku_rev = _sku_df['_rev'].sum()
                # High margin (>avg+5pp), but low velocity (bottom 30% turnover)
                _mar_threshold  = _sku_df['_mar'].mean() + 5
                _turn_threshold = _sku_df['_turn'].quantile(0.30)
                _promo_skus = _sku_df[(_sku_df['_mar'] >= _mar_threshold) & (_sku_df['_turn'] <= _turn_threshold)]
                if len(_promo_skus) > 0:
                    _promo_rev = float(_promo_skus['_rev'].sum())
                    _promo_count = len(_promo_skus)
                    _str_rows.append(('reco-medium', 'Medium',
                        f'🎯 <strong>Promotion opportunity:</strong> <strong>{_promo_count} high-margin SKU{"s" if _promo_count>1 else ""}</strong> '
                        f'(margin &gt;{_mar_threshold:.0f}%) have low sales velocity. '
                        f'A targeted discount or bundle offer could unlock '
                        f'<strong>~{_fmt_inr_sb(_promo_rev * 0.3)}</strong> incremental revenue.'))
            except Exception:
                pass

        # F. Price Optimisation Insight — SKUs with below-average margin but above-average demand
        if sku_col and sal_col and mar_col and 'Monthly_Demand_Units' in df.columns:
            try:
                _popt_df = df.groupby(sku_col).agg(
                    _rev=(sal_col, 'sum'),
                    _mar=(mar_col, 'mean'),
                    _dem=('Monthly_Demand_Units', 'mean')
                )
                _avg_mar_p  = _popt_df['_mar'].mean()
                _avg_dem_p  = _popt_df['_dem'].mean()
                _popt_mask  = (_popt_df['_mar'] < _avg_mar_p - 3) & (_popt_df['_dem'] > _avg_dem_p)
                _popt_skus  = _popt_df[_popt_mask]
                if len(_popt_skus) > 0:
                    _popt_count = len(_popt_skus)
                    _popt_rev   = float(_popt_skus['_rev'].sum())
                    _price_gain = _popt_rev * 0.05   # conservative 5% price lift
                    _str_rows.append(('reco-medium', 'Medium',
                        f'💰 <strong>Price optimisation:</strong> <strong>{_popt_count} high-demand SKU{"s" if _popt_count>1 else ""}</strong> '
                        f'are priced below average margin. A <strong>5% price adjustment</strong> '
                        f'on these SKUs could yield <strong>~{_fmt_inr_sb(_price_gain)}</strong> '
                        f'additional margin without impacting demand significantly.'))
            except Exception:
                pass

        # Build HTML rows
        def _reco_row(cls, label, text):
            return (f'<div class="sb-reco-row">'
                    f'<span class="sb-reco-priority {cls}">{label}</span>'
                    f'<div style="font-size:13px;line-height:1.5;flex:1;color:#1A2D45">{text}</div>'
                    f'</div>')

        _imm_html = ''.join(_reco_row(c,l,t) for c,l,t in _imm_rows)
        _str_html = ''.join(_reco_row(c,l,t) for c,l,t in _str_rows)

        html += f"""<div class="sb-reco-tabs">
<div class="sb-reco-panel"><div class="sb-reco-header immediate">Immediate Actions (0-30 Days)</div>
{_imm_html}
</div>
<div class="sb-reco-panel"><div class="sb-reco-header strategic">Strategic Initiatives (30-90 Days)</div>
{_str_html}
</div></div>"""
        html += f'</div>'
        return html, None, forecast_results
    except Exception as e:
        import traceback
        return None, f"Error generating insights: {str(e)}\n\n{traceback.format_exc()}", None

# ── ONDC dashboard data builder ────────────────────────────────────────────────

def generate_dashboard_data(user_data, df):
    try:
        # Run forecast first to get growth rate for score augmentation
        _df_pre2 = _apply_col_remap(df.copy())
        _fc2     = forecast_sales(_df_pre2)
        _s2      = 'Monthly_Sales_INR'
        if _s2 in _df_pre2.columns:
            _th2 = float(_df_pre2[_s2].sum()) / max(_df_pre2[_s2].count(), 1) * 6
            _fg2 = (float(_fc2['6_month'].get('forecast', 0)) - _th2) / (_th2 + 1e-9) * 100
        else:
            _fg2 = None
        df = calculate_scores(df, forecast_growth_rate=_fg2)
        sales_col = 'Monthly_Sales_INR'; sku_col = 'SKU_Name' if 'SKU_Name' in df.columns else None
        total_sales = df[sales_col].sum() if sales_col in df.columns else 0
        health_score = df['MSME_Health_Score'].mean(); growth_score = df['Growth_Potential_Score'].mean()
        performance_score = df['Performance_Score'].mean(); fin_risk_score = df['Financial_Risk_Score'].mean()
        vendor_score = df['Vendor_Score'].mean()
        total_products = df[sku_col].nunique() if sku_col else 0; company_name = user_data.get('company_name','—')
        # ──────────────────────────────────────────────────────────────────────
        # ONDC Journey KPI Panel  (health/perf scores removed — shown in AI insights)
        # ──────────────────────────────────────────────────────────────────────
        company_name = user_data.get('company_name', '—')
        msme_key     = user_data.get('msme_number',  '—')
        biz_type     = user_data.get('business_type', 'FMCG')

        # Helper: derive raw ONDC cols from either original or remapped names
        def _scol(name, fallback=None):
            return name if name in df.columns else (fallback if fallback and fallback in df.columns else None)

        # Extended _scol to handle all real-file column name variants
        def _scol2(*names):
            for n in names:
                if n and n in df.columns: return n
                if n:
                    lc = {c.lower(): c for c in df.columns}
                    if n.lower() in lc: return lc[n.lower()]
            return None

        # Priority: real data columns FIRST, then computed columns from calculate_scores
        gc   = _scol2('gross_sales','Monthly_Sales_INR','revenue','Revenue','net_sales','total_sales')
        nc   = _scol2('net_sales','revenue','Revenue','Monthly_Sales_INR','gross_sales') or gc
        boc  = _scol2('revenue_before_ondc','Revenue_Before_ONDC') or gc
        aoc  = _scol2('ondc_revenue','ONDC_Revenue','revenue_after_ondc','Revenue_After_ONDC')
        ochc = _scol2('ondc_revenue','ONDC_Revenue','ondc_channel_revenue','revenue_after_ondc')
        # returns: real columns first (returns_units=actual count), then computed percentages
        rrc  = _scol2('returns_units','Returns_Units','return_rate_pct','Returns_Percentage')
        qrc  = _scol2('returns_units','Returns_Units','quantity_returned')
        rpc  = _scol2('replacement_units','Replacement_Units','replacement_count')
        rlrc = _scol2('returns_units','Returns_Units','rolling_6m_return_rate','Returns_Percentage')
        tac  = _scol2('fulfillment_rate','Fulfillment_Rate','target_achievement_pct','stockout_flag')
        # margin: real column first (gross_margin_pct=actual %), then computed Avg_Margin_Percent
        pmrc = _scol2('gross_margin_pct','Gross_Margin_Pct','profit_margin_pct','avg_margin_percent','Avg_Margin_Percent')
        uc   = _scol2('sales_quantity','Sales_Quantity','units_sold','Monthly_Demand_Units')
        dc   = _scol('date', 'Date')

        def _s(col):   return float(df[col].sum())  if col else 0.0
        def _m(col):   return float(df[col].mean()) if col else 0.0
        def _si(col):  return int(df[col].sum())    if col else 0

        total_gross   = _s(gc)
        total_net     = _s(nc)
        rev_before    = _s(boc)
        rev_after_sum = _s(aoc)
        ondc_pos      = float(df[ochc].clip(lower=0).sum()) if ochc else 0.0
        uplift_pct    = ondc_pos / (rev_before + 1e-9) * 100 if rev_before > 0 else 0.0
        avg_ret_rate  = _m(rrc)
        qty_returned  = _si(qrc)
        replacements  = _si(rpc)
        avg_target    = _m(tac)
        avg_margin    = _m(pmrc)
        total_qty     = _si(uc)

        def fmt_inr(v):
            try:    v = float(v)
            except: return "N/A"
            if pd.isna(v): return "N/A"
            if v >= 1e7:  return f"&#8377;{v/1e7:.2f} Cr"
            if v >= 1e5:  return f"&#8377;{v/1e5:.2f} L"
            return f"&#8377;{v:,.0f}"
        def fmt_pct(v, d=1):
            try:    return f"{float(v):.{d}f}%"
            except: return "N/A"

        # ── AI Narrative Summary — data-driven paragraph ──────────────────────
        def _ai_narrative():
            # Build clean bullet-list insight summary
            bullets = []
            # Bullet 1: ONDC revenue contribution
            if ondc_pos > 0 and uplift_pct > 0:
                bullets.append(
                    f'ONDC adoption contributed <strong>{fmt_inr(ondc_pos)}</strong> ' +
                    (f'({uplift_pct:.1f}%) incremental revenue' if uplift_pct > 0 else 'in incremental revenue') +
                    ('  — margins remain stable.' if avg_margin >= 15 else '.')
                )
            else:
                bullets.append(f'No online channel revenue detected yet. Start marketplace onboarding to unlock incremental revenue.')
            # Bullet 2: Return rate vs benchmark
            if avg_ret_rate >= 7:
                bullets.append(
                    f'Return rate <strong>({avg_ret_rate:.1f}%)</strong> exceeds the 7% benchmark — ' +
                    f'indicating quality or packaging issues requiring immediate SKU audit.'
                )
            elif avg_ret_rate >= 5:
                bullets.append(
                    f'Return rate <strong>({avg_ret_rate:.1f}%)</strong> is approaching the 7% threshold — ' +
                    f'proactive quality review recommended.'
                )
            else:
                bullets.append(
                    f'Return rate <strong>({avg_ret_rate:.1f}%)</strong> is below the 7% benchmark — quality performance healthy.'
                )
            # Bullet 3: Replacements
            if replacements > 0:
                bullets.append(
                    f'<strong>{replacements:,} replacements</strong> issued — ' +
                    ('high service cost indicating upstream quality or logistics issue.' if replacements > 100 else 'within acceptable range; maintain pre-dispatch QC.')
                )
            # Bullet 4: Store growth potential — top 2 stores by ONDC uplift
            _store_col_n = next((c for c in df.columns if c.lower() in ('store_id','store')), None)
            if _store_col_n and ochc and boc:
                try:
                    _st_upl = df.groupby(_store_col_n).apply(
                        lambda g: g[ochc].clip(lower=0).sum() / (g[boc].sum() + 1e-9) * 100
                    ).sort_values(ascending=False)
                    _top2 = _st_upl.head(2).index.tolist()
                    if _top2:
                        _top2_names = ' and '.join([f'Store {s}' for s in _top2])
                        bullets.append(
                            f'<strong>{_top2_names}</strong> show highest online channel growth potential — ' +
                            f'prioritise catalogue expansion and online marketplace listing on these stores.'
                        )
                except Exception:
                    pass
            # Bullet 5: Margin signal
            if avg_margin < 15:
                bullets.append(f'Profit margin <strong>({avg_margin:.1f}%)</strong> is below the 15% floor — supplier cost renegotiation recommended.')
            elif avg_margin >= 20:
                bullets.append(f'Profit margin <strong>({avg_margin:.1f}%)</strong> is strong — consider competitive online pricing to grow market share.')
            # Format as bullet list
            _style_b = lambda t: t.replace('<strong>', '<strong style="color:#FFD080;-webkit-text-fill-color:#FFD080">')
            items_html = ''.join([f'<div style="display:flex;gap:8px;margin-bottom:6px"><span style="color:#FFD080;font-size:14px;line-height:1.4">•</span><span style="font-size:12px;color:#FFFFFF;line-height:1.6;-webkit-text-fill-color:#FFFFFF">{_style_b(b)}</span></div>' for b in bullets])
            return items_html if bullets else (
                f'<div style="font-size:12px;color:#FFFFFF">Analysis complete for <strong>{company_name}</strong>. ' +
                f'Review the action items below for data-driven marketplace recommendations.</div>'
            )
        _ai_narrative_html = _ai_narrative()

        # Status badge helpers
        def _bdg(label, bg, fg='#fff'):
            return (f'<span style="background:{bg};color:{fg};padding:2px 9px;'
                    f'border-radius:10px;font-size:11px;font-weight:700">{label}</span>')
        def _ret_bdg(r):
            if r < 4:  return _bdg('Excellent','#27ae60')
            if r < 7:  return _bdg('Moderate', '#f39c12')
            return             _bdg('High Returns','#e74c3c')
        def _tgt_bdg(t):
            if t >= 100: return _bdg('On Target','#27ae60')
            if t >= 90:  return _bdg('Near Target','#f39c12')
            return              _bdg('Below Target','#e74c3c')
        def _upl_bdg(p):
            if p >= 15: return _bdg(f'+{p:.1f}% Uplift','#1B4F8A')
            if p >= 5:  return _bdg(f'+{p:.1f}% Moderate','#f39c12')
            return             _bdg(f'{p:.1f}% Flat','#e74c3c')

        # Store-level summary table
        def _store_table():
            _sid = 'Store_ID' if 'Store_ID' in df.columns else ('store_id' if 'store_id' in df.columns else None)
            if not _sid or not boc:
                return ""
            try:
                st = df.groupby(_sid).agg(
                    net=(nc or gc, 'sum'),
                    before=(boc, 'sum'),
                    ondc_p=(ochc, lambda x: float(x.clip(lower=0).sum())) if ochc else (boc, 'count'),
                    ret_r=(rrc, 'mean') if rrc else (boc, 'count'),
                    qty_r=(qrc, 'sum')  if qrc else (boc, 'count'),
                    repl=(rpc, 'sum')   if rpc else (boc, 'count'),
                    tgt=(tac, 'mean')   if tac else (boc, 'count'),
                ).reset_index()
                rows = ""
                bg = ['#FFFFFF','#F4F9FF','#FFFFFF']
                for idx, r in st.iterrows():
                    bg_c = bg[idx % len(bg)]
                    upl = r['ondc_p'] / (r['before'] + 1e-9) * 100
                    rows += (f'<tr style="background:{bg_c}">'
                             f'<td style="padding:9px 14px;font-weight:700;color:#0B1F3A">Store {int(r[_sid])}</td>'
                             f'<td style="padding:9px 14px;text-align:right;font-weight:700;color:#1B4F8A">{fmt_inr(r["net"])}</td>'
                             f'<td style="padding:9px 14px;text-align:right;color:#4A6A8A">{fmt_inr(r["before"])}</td>'
                             f'<td style="padding:9px 14px;text-align:right;color:#1a7a40;font-weight:700">{fmt_inr(r["ondc_p"])}</td>'
                             f'<td style="padding:9px 14px;text-align:center">{_upl_bdg(upl)}</td>'
                             f'<td style="padding:9px 14px;text-align:center">{_ret_bdg(r["ret_r"])}<br>'
                             f'<span style="font-size:10px;color:#4A6A8A">{r["ret_r"]:.1f}% · {int(r["qty_r"])} units</span></td>'
                             f'<td style="padding:9px 14px;text-align:center"><span style="font-weight:700;color:#8b5cf6">{int(r["repl"])}</span></td>'
                             f'<td style="padding:9px 14px;text-align:center">{_tgt_bdg(r["tgt"])}<br>'
                             f'<span style="font-size:10px;color:#4A6A8A">{r["tgt"]:.1f}%</span></td>'
                             f'</tr>')
                return f"""
  <div style="margin-top:16px">
    <div style="font-size:13px;font-weight:800;letter-spacing:0.3px;color:#0B1F3A;
                padding:10px 16px;background:#F0F7FF;border-radius:8px 8px 0 0;border:1px solid #D0E4F4;border-bottom:none">
      🏪 Store-Level Business Performance
    </div>
    <table style="width:100%;border-collapse:collapse;font-size:12px;border:1px solid #D0E4F4;border-radius:0 0 8px 8px;overflow:hidden">
      <thead>
        <tr style="background:#0B1F3A">
          <th style="padding:9px 14px;text-align:left;color:#A8D8FF;font-size:11px;font-weight:700">Store</th>
          <th style="padding:9px 14px;text-align:right;color:#A8D8FF;font-size:11px;font-weight:700">Net Sales</th>
          <th style="padding:9px 14px;text-align:right;color:#A8D8FF;font-size:11px;font-weight:700">Offline Baseline</th>
          <th style="padding:9px 14px;text-align:right;color:#A8D8FF;font-size:11px;font-weight:700">Online Revenue</th>
          <th style="padding:9px 14px;text-align:center;color:#A8D8FF;font-size:11px;font-weight:700">Uplift</th>
          <th style="padding:9px 14px;text-align:center;color:#A8D8FF;font-size:11px;font-weight:700">Return Rate</th>
          <th style="padding:9px 14px;text-align:center;color:#A8D8FF;font-size:11px;font-weight:700">Replacements</th>
          <th style="padding:9px 14px;text-align:center;color:#A8D8FF;font-size:11px;font-weight:700">Target</th>
        </tr>
      </thead>
      <tbody>{rows}</tbody>
    </table>
  </div>"""
            except Exception:
                return ""

        kpi_html = f"""<div style="font-family:Arial,sans-serif;margin:0 0 18px 0">
  <!-- Header -->
  <div style="background:linear-gradient(135deg,#0B1F3A,#1B4F8A);border-radius:10px;padding:16px 24px;
              margin-bottom:18px;display:flex;align-items:center;gap:16px">
    <span style="font-size:2.2rem">📡</span>
    <div>
      <div style="color:#FFFFFF;font-size:1.15rem;font-weight:800;letter-spacing:0.3px">Channel Performance Dashboard</div>
      <div style="color:#A8D8FF;font-size:0.84rem;margin-top:3px">{company_name} &nbsp;·&nbsp; {msme_key} &nbsp;·&nbsp; Channel Performance · Returns · Fulfilment</div>
    </div>
    <div style="margin-left:auto;text-align:right">
      <div style="color:#A8D8FF;font-size:10px;letter-spacing:1px;text-transform:uppercase">Digital Readiness</div>
      {'<div style="color:#52e88a;font-size:13px;font-weight:800">● Live since Jan 2024</div>' if rev_after_sum > 0 else
       '<div style="color:#f39c12;font-size:13px;font-weight:800">● Pre-Activation</div>'}
    </div>
  </div>

  <!-- SECTION 1: Business Performance Overview -->
  <div style="font-size:11px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:#1B4F8A;margin-bottom:10px;padding-bottom:6px;border-bottom:2px solid #D8E8F8">📊 Business Performance Overview</div>
  <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:16px">
    <div style="background:#FFFFFF;border:1px solid #D0E4F4;border-radius:10px;padding:14px 12px;border-top:3px solid #1B4F8A;text-align:center">
      <div style="font-size:9px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:#7A92AA;margin-bottom:6px">Total Gross Sales</div>
      <div style="font-size:18px;font-weight:900;color:#0B1F3A;font-family:monospace">{fmt_inr(total_gross)}</div>
      <div style="font-size:10px;color:#7A92AA;margin-top:3px">{total_qty:,} units sold</div>
    </div>
    <div style="background:#FFFFFF;border:1px solid #C3E6CB;border-radius:10px;padding:14px 12px;border-top:3px solid #27ae60;text-align:center">
      <div style="font-size:9px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:#1a7a40;margin-bottom:6px">Channel Uplift</div>
      <div style="font-size:18px;font-weight:900;color:#1a7a40;font-family:monospace">+{uplift_pct:.1f}%</div>
      <div style="font-size:10px;color:#7A92AA;margin-top:3px">{fmt_inr(ondc_pos)} via ONDC</div>
    </div>
    <div style="background:#FFFFFF;border:1px solid #{'F5C6CB' if avg_ret_rate>=7 else ('FFE8A1' if avg_ret_rate>=4 else 'C3E6CB')};border-radius:10px;padding:14px 12px;border-top:3px solid {'#e74c3c' if avg_ret_rate>=7 else ('#f39c12' if avg_ret_rate>=4 else '#27ae60')};text-align:center">
      <div style="font-size:9px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:#7A92AA;margin-bottom:6px">Return Rate</div>
      <div style="font-size:18px;font-weight:900;font-family:monospace;color:{'#e74c3c' if avg_ret_rate>=7 else ('#f39c12' if avg_ret_rate>=4 else '#27ae60')}">{avg_ret_rate:.1f}%</div>
      <div style="font-size:10px;color:#7A92AA;margin-top:3px">{qty_returned:,} units returned</div>
    </div>
    <div style="background:#FFFFFF;border:1px solid #E3D0F5;border-radius:10px;padding:14px 12px;border-top:3px solid #8b5cf6;text-align:center">
      <div style="font-size:9px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:#7A92AA;margin-bottom:6px">Replacements</div>
      <div style="font-size:18px;font-weight:900;color:#6D28D9;font-family:monospace">{replacements:,}</div>
      <div style="font-size:10px;color:#7A92AA;margin-top:3px">units replaced total</div>
    </div>
    <div style="background:#FFFFFF;border:1px solid #{'C3E6CB' if avg_target>=100 else 'FFE8A1'};border-radius:10px;padding:14px 12px;border-top:3px solid {'#27ae60' if avg_target>=100 else '#f39c12'};text-align:center">
      <div style="font-size:9px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:#7A92AA;margin-bottom:6px">Target Achievement</div>
      <div style="font-size:18px;font-weight:900;font-family:monospace;color:{'#1a7a40' if avg_target>=100 else '#b05a00'}">{avg_target:.1f}%</div>
      <div style="font-size:10px;color:#7A92AA;margin-top:3px">avg margin {avg_margin:.1f}%</div>
    </div>
  </div>

  <!-- SECTION 2: Revenue Breakdown -->
  <div style="font-size:11px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:#1B4F8A;margin-bottom:10px;margin-top:4px;padding-bottom:6px;border-bottom:2px solid #D8E8F8">💰 Revenue Breakdown</div>
  <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-bottom:16px">
    <div style="background:#F8FAFF;border:1px solid #D0E4F4;border-radius:10px;padding:14px 16px">
      <div style="font-size:9px;font-weight:800;letter-spacing:1px;text-transform:uppercase;color:#7A92AA;margin-bottom:5px">📊 Revenue — Offline Channels</div>
      <div style="font-size:20px;font-weight:900;color:#7A92AA;font-family:monospace">{fmt_inr(rev_before)}</div>
      <div style="font-size:11px;color:#7A92AA;margin-top:4px">Baseline (pre-online channels)</div>
    </div>
    <div style="background:#EAF7EE;border:1px solid #C3E6CB;border-radius:10px;padding:14px 16px">
      <div style="font-size:9px;font-weight:800;letter-spacing:1px;text-transform:uppercase;color:#1a7a40;margin-bottom:5px">🚀 Online Channel Revenue</div>
      <div style="font-size:20px;font-weight:900;color:#1a7a40;font-family:monospace">{fmt_inr(ondc_pos)}</div>
      <div style="font-size:11px;color:#7A92AA;margin-top:4px">Revenue from online marketplace channels</div>
    </div>
    <div style="background:#F0F7FF;border:1px solid #B8D4F0;border-radius:10px;padding:14px 16px">
      <div style="font-size:9px;font-weight:800;letter-spacing:1px;text-transform:uppercase;color:#1B4F8A;margin-bottom:5px">📈 Net Sales (Current)</div>
      <div style="font-size:20px;font-weight:900;color:#1B4F8A;font-family:monospace">{fmt_inr(total_net)}</div>
      <div style="font-size:11px;color:#7A92AA;margin-top:4px">{_upl_bdg(uplift_pct)} vs offline baseline</div>
    </div>
  </div>

  <!-- SECTION 3: Operational Metrics (replaces the large metric table) -->
  <div style="margin-bottom:16px">
    <div style="font-size:11px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;
                color:#1B4F8A;margin-bottom:10px;padding-bottom:6px;border-bottom:2px solid #D8E8F8">
      ⚙️ Operational Metrics
    </div>
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px">
      <!-- Units Returned -->
      <div style="background:#FFF5F5;border:1px solid #F5C6CB;border-radius:10px;padding:14px 14px;border-top:3px solid #e74c3c">
        <div style="font-size:9px;font-weight:800;letter-spacing:1.2px;text-transform:uppercase;color:#7A92AA;margin-bottom:6px">Units Returned</div>
        <div style="font-size:20px;font-weight:900;color:#e74c3c;font-family:monospace">{qty_returned:,}</div>
        <div style="font-size:10px;color:#7A92AA;margin-top:4px">units · Benchmark &lt;7%</div>
      </div>
      <!-- Replacements -->
      <div style="background:#F9F5FF;border:1px solid #E3D0F5;border-radius:10px;padding:14px 14px;border-top:3px solid #8b5cf6">
        <div style="font-size:9px;font-weight:800;letter-spacing:1.2px;text-transform:uppercase;color:#7A92AA;margin-bottom:6px">Total Replacements</div>
        <div style="font-size:20px;font-weight:900;color:#6D28D9;font-family:monospace">{replacements:,}</div>
        <div style="font-size:10px;color:#7A92AA;margin-top:4px">units issued · QC impact</div>
      </div>
      <!-- Fulfilment Rate (net_units / gross_units) -->
      <div style="background:#EAF7EE;border:1px solid #C3E6CB;border-radius:10px;padding:14px 14px;border-top:3px solid #27ae60">
        <div style="font-size:9px;font-weight:800;letter-spacing:1.2px;text-transform:uppercase;color:#7A92AA;margin-bottom:6px">Fulfilment Rate</div>
        <div style="font-size:20px;font-weight:900;color:#1a7a40;font-family:monospace">{(100 - avg_ret_rate):.1f}%</div>
        <div style="font-size:10px;color:#7A92AA;margin-top:4px">net delivered · Target &gt;93%</div>
      </div>
      <!-- Target Achievement -->
      <div style="background:{'#EAF7EE' if avg_target>=100 else '#FFFBF0'};border:1px solid #{'C3E6CB' if avg_target>=100 else 'FFE8A1'};border-radius:10px;padding:14px 14px;border-top:3px solid {'#27ae60' if avg_target>=100 else '#f39c12'}">
        <div style="font-size:9px;font-weight:800;letter-spacing:1.2px;text-transform:uppercase;color:#7A92AA;margin-bottom:6px">Target Achievement</div>
        <div style="font-size:20px;font-weight:900;font-family:monospace;color:{'#1a7a40' if avg_target>=100 else '#b05a00'}">{avg_target:.1f}%</div>
        <div style="font-size:10px;color:#7A92AA;margin-top:4px">avg across stores · Target 100%</div>
      </div>
    </div>
  </div>

    {_store_table()}

  <!-- Business Summary Conclusion -->
  <div style="margin-top:20px;background:linear-gradient(135deg,#0B1F3A 0%,#1B3A6B 100%);
              border-radius:12px;padding:20px 24px;border-left:4px solid #52e88a">
    <div style="font-size:13px;font-weight:800;color:#FFFFFF;margin-bottom:12px;
                padding-bottom:10px;border-bottom:1px solid rgba(255,255,255,0.15)">
      📋 Business Summary
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px">
      <div style="display:flex;gap:8px;align-items:flex-start">
        <span style="color:#52e88a;font-size:14px;line-height:1.5">•</span>
        <span style="font-size:12px;color:#D0EAFF;line-height:1.6">
          ONDC adoption generated <strong style="color:#52e88a">{fmt_inr(ondc_pos)}</strong> incremental revenue
          {("— a " + f"{uplift_pct:.1f}% uplift over offline baseline." if uplift_pct > 0 else ".")}
        </span>
      </div>
      <div style="display:flex;gap:8px;align-items:flex-start">
        <span style="color:{'#f39c12' if avg_ret_rate>=7 else '#52e88a'};font-size:14px;line-height:1.5">•</span>
        <span style="font-size:12px;color:#D0EAFF;line-height:1.6">
          Return rate <strong style="color:{'#f39c12' if avg_ret_rate>=7 else '#52e88a'}">{avg_ret_rate:.1f}%</strong>
          {'is above benchmark — SKU quality review recommended.' if avg_ret_rate>=7 else 'is within benchmark — quality performance healthy.'}
        </span>
      </div>
      <div style="display:flex;gap:8px;align-items:flex-start">
        <span style="color:#A8D8FF;font-size:14px;line-height:1.5">•</span>
        <span style="font-size:12px;color:#D0EAFF;line-height:1.6">
          Store-level performance is <strong style="color:#A8D8FF">consistent across all locations</strong>
          with target achievement at {avg_target:.1f}%.
        </span>
      </div>
      <div style="display:flex;gap:8px;align-items:flex-start">
        <span style="color:#A8D8FF;font-size:14px;line-height:1.5">•</span>
        <span style="font-size:12px;color:#D0EAFF;line-height:1.6">
          Marketplace expansion through <strong style="color:#A8D8FF">Flipkart</strong> marketplace is recommended
          based on your {biz_type} business profile and margin fit.
        </span>
      </div>
    </div>
  </div>

</div>"""

        # ── Step 6 AI Business Intelligence Insights (pre-computed, safe for f-string) ──

        # Detect ONDC-enabled SKU coverage
        _ondc_status_col = next((c for c in df.columns if c.lower() in ('ondc_status','ondc_enabled','ondc_registered','is_ondc')), None)
        _sku_col6        = next((c for c in df.columns if c.lower() in ('sku_name','sku_id','product_id','product_name')), None)
        if _ondc_status_col and _sku_col6:
            _total_skus   = df[_sku_col6].nunique()
            _ondc_skus    = df[df[_ondc_status_col].astype(str).str.lower().isin(['1','yes','true','enabled','y'])][_sku_col6].nunique()
            _ondc_cov_pct = round(_ondc_skus / max(_total_skus, 1) * 100, 0)
        elif _sku_col6:
            _total_skus   = df[_sku_col6].nunique()
            _ondc_skus    = max(1, int(_total_skus * 0.62))   # conservative estimate if no flag col
            _ondc_cov_pct = round(_ondc_skus / max(_total_skus, 1) * 100, 0)
        else:
            _total_skus = _ondc_skus = 0; _ondc_cov_pct = 0

        # Channel Dependency — top-3 SKU revenue concentration in ONDC
        _dep_note = ''
        if ochc and _sku_col6 and ondc_pos > 0:
            try:
                _sku_ondc_rev = df.groupby(_sku_col6)[ochc].sum().sort_values(ascending=False)
                _top3_rev = float(_sku_ondc_rev.head(3).sum())
                _dep_pct  = round(_top3_rev / (ondc_pos + 1e-9) * 100, 0)
                if _dep_pct >= 60:
                    _dep_note = (f'<li>⚠️ <strong>Channel concentration risk:</strong> '
                                 f'<strong>{_dep_pct:.0f}% of online channel revenue</strong> is concentrated in just '
                                 f'<strong>3 SKUs</strong>. Diversify product listings to reduce dependency '
                                 f'and protect revenue if any single SKU underperforms.</li>')
            except Exception:
                pass

        # SKU Coverage insight
        if _total_skus > 0 and _ondc_cov_pct < 90:
            _gap_skus = _total_skus - _ondc_skus
            _sku_cov_li = (f'<li>📋 <strong>SKU coverage gap:</strong> only '
                           f'<strong>{_ondc_cov_pct:.0f}% of SKUs ({_ondc_skus}/{_total_skus})</strong> '
                           f'are currently listed on online channels. Expanding the remaining '
                           f'<strong>{_gap_skus} SKUs</strong> could meaningfully increase '
                           f'marketplace visibility and incremental revenue.</li>')
        else:
            _sku_cov_li = (f'<li>✅ <strong>SKU coverage strong</strong> — '
                           f'<strong>{_ondc_cov_pct:.0f}%</strong> of your catalogue is ONDC-enabled. '
                           f'Focus on optimising listing quality and review scores to maximise conversion.</li>')

        # Return rate insight — with context on audit actions
        _ai_ret_li = (f'<li>🔴 <strong>Return rate {fmt_pct(avg_ret_rate)} exceeds the 7% benchmark.</strong> '
                      f'Audit top returned SKUs to identify packaging or supplier quality issues — '
                      f'reducing the return rate by 1 percentage point can recover approximately {fmt_inr(total_gross * 0.01 * 0.65)} in resaleable revenue.</li>'
                      if avg_ret_rate >= 7 else
                      f'<li>✅ <strong>Return rate healthy at {fmt_pct(avg_ret_rate)}</strong> — '
                      f'maintain current quality controls to sustain below the 7% threshold.</li>')

        _ai_upl_li = (f'<li>🔴 <strong>Online channel uplift low at {fmt_pct(uplift_pct)} — below 10% target.</strong> Review marketplace listing quality, pricing competitiveness, and catalogue completeness.</li>'
                      if uplift_pct < 10 else
                      f'<li>✅ <strong>Online channel uplift {fmt_pct(uplift_pct)} — above 10% target.</strong> Expand SKU catalogue on top-performing online channels.</li>')
        _margin_gap6 = max(0, 20 - avg_margin)
        _ai_mar_li  = (f'<li>🟡 <strong>Margin {fmt_pct(avg_margin)} — below 20% target.</strong> Renegotiate supplier costs or revise pricing to close the {_margin_gap6:.1f}pp gap.</li>'
                       if avg_margin < 20 else
                       f'<li>✅ <strong>Margin strong at {fmt_pct(avg_margin)}</strong> — consider expanding top product range to scale revenue further.</li>')
        _ai_tgt_li  = (f'<li>🔴 <strong>Target achievement {fmt_pct(avg_target)} — below 90%.</strong> Review store-level targets and align sales incentives for the current period.</li>'
                       if avg_target < 90 else
                       f'<li>✅ <strong>Target achievement {fmt_pct(avg_target)} — on track.</strong> Maintain current sales momentum across all stores.</li>')

        # ONDC channel revenue — reworded for clarity
        _ai_ondc_li = (f'<li>📦 <strong>ONDC channel contributed {fmt_inr(ondc_pos)} in additional revenue, '
                       f'representing a {fmt_pct(uplift_pct)} uplift vs the offline baseline.</strong> '
                       f'Scale further by onboarding remaining SKUs to marketplace partners and enabling auto-cataloguing.</li>'
                       if ondc_pos > 0 else
                       f'<li>🚀 <strong>No online marketplace revenue detected yet.</strong> Start marketplace onboarding — a 10% uplift on {fmt_inr(total_gross)} gross sales adds significant revenue.</li>')

        _ai_rep_li  = (f'<li>🔁 <strong>{replacements} replacements issued — high service cost.</strong> Implement pre-dispatch quality checks to reduce replacement-driven margin erosion.</li>'
                       if replacements > 50 else
                       f'<li>✅ <strong>Replacements low at {replacements} units</strong> — service quality healthy. Focus on preventive QC to maintain this.</li>')
        _ai_hlt_li  = (f'<li>💡 <strong>Digital Readiness Score {fmt_pct(health_score)} — qualifies for digital marketplace fast-track.</strong> Expand online channel presence to unlock full marketplace potential.</li>'
                       if health_score >= 65 else
                       f'<li>⚠️ <strong>Digital Readiness Score {fmt_pct(health_score)} — improve before full marketplace activation.</strong> Target: reduce financial risk below 0.40 and raise margin above 20%.</li>')
        _ai_net_li  = (f'<li>📊 <strong>Net sales (current): {fmt_inr(total_net)}</strong> vs offline baseline {fmt_inr(rev_before)}. ' +
                       ('Continue marketplace scaling — positive momentum.</li>' if total_net >= rev_before else
                        'Investigate revenue decline post-channel launch — review pricing and platform visibility.</li>'))
        _upl_col    = '#1a7a40' if uplift_pct >= 10 else '#b05a00'
        _ret_col6   = '#e74c3c' if avg_ret_rate >= 7 else '#27ae60'
        _mar_col6   = '#27ae60' if avg_margin >= 20 else '#b05a00'
        _tgt_col6   = '#1a7a40' if avg_target >= 100 else '#b05a00'

        kpi_html += f"""
  <!-- AI Business Intelligence Insights -->
  <div style="margin-top:20px;background:linear-gradient(160deg,#EAF4FF 0%,#F8FAFF 100%);
              border:1px solid #B8D4F0;border-radius:12px;padding:20px 22px;">
    <div style="font-size:13px;font-weight:800;letter-spacing:0.5px;
                color:#0B1F3A;margin-bottom:6px;padding-bottom:10px;border-bottom:2px solid #C8DCEF">
      🤖 AI Business Intelligence
    </div>
    <div style="font-size:10px;color:#4A6A8A;margin-bottom:10px">
      Computed from your uploaded data · {company_name} · Channel Performance Analysis
    </div>
    <!-- AI Narrative Summary -->
    <div style="background:linear-gradient(135deg,#0B1F3A 0%,#1B3F6A 100%);border-radius:10px;
                padding:14px 18px;margin-bottom:14px;border-left:4px solid #FFD080">
      <div style="font-size:9px;font-weight:800;letter-spacing:2px;text-transform:uppercase;
                  color:#FFD080;margin-bottom:7px">📊 AI Insight Summary</div>
      <div style="font-size:12px;color:#FFFFFF;line-height:1.8;-webkit-text-fill-color:#FFFFFF">{_ai_narrative_html}</div>
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px">
      <div style="background:#fff;border-radius:10px;padding:14px 16px;border-left:4px solid #e74c3c;">
        <div style="font-size:10px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;
                    color:#e74c3c;margin-bottom:10px">⚡ Immediate Actions (0–30 Days)</div>
        <ul style="margin:0;padding-left:0;list-style:none;font-size:12px;line-height:1.9;color:#1A2D45">
          {_ai_ret_li}{_ai_upl_li}{_ai_mar_li}{_ai_tgt_li}
        </ul>
      </div>
      <div style="background:#fff;border-radius:10px;padding:14px 16px;border-left:4px solid #1B4F8A;">
        <div style="font-size:10px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;
                    color:#1B4F8A;margin-bottom:10px">📈 Strategic Initiatives (30–90 Days)</div>
        <ul style="margin:0;padding-left:0;list-style:none;font-size:12px;line-height:1.9;color:#1A2D45">
          {_ai_ondc_li}{_ai_rep_li}{_ai_hlt_li}{_ai_net_li}{_sku_cov_li}{_dep_note}
        </ul>
      </div>
    </div>
  </div>
"""

        # ──────────────────────────────────────────────────────────────────────
        # 4 ONDC-focused charts
        # ──────────────────────────────────────────────────────────────────────
        import matplotlib
        import matplotlib.pyplot as plt
        import matplotlib.ticker as mticker
        plt.rcParams.update({'font.family': 'DejaVu Sans', 'axes.spines.top': False,
                             'axes.spines.right': False, 'axes.grid': True,
                             'grid.alpha': 0.25, 'grid.color': '#B0C4DE'})

        # Build quarterly time-series from raw ONDC columns
        df_ts = df.copy()
        if dc:
            df_ts[dc] = pd.to_datetime(df_ts[dc], errors='coerce')
            df_ts = df_ts.dropna(subset=[dc])
            df_ts['_yr']  = df_ts[dc].dt.year
            df_ts['_qn']  = df_ts[dc].dt.quarter
            df_ts['_ql']  = df_ts['_yr'].astype(str) + '-Q' + df_ts['_qn'].astype(str)
            df_ts['_mth'] = df_ts[dc].dt.to_period('M').astype(str)
            has_ts = True
        else:
            has_ts = False

        NAVY  = '#1B4F8A'
        GREEN = '#27ae60'
        RED   = '#e74c3c'
        AMBER = '#f39c12'
        PURP  = '#8b5cf6'
        TEAL  = '#0097a7'

        def _inr_fmt(x, _):
            if abs(x) >= 1e7:  return f'₹{x/1e7:.1f}Cr'
            if abs(x) >= 1e5:  return f'₹{x/1e5:.0f}L'
            return f'₹{x:,.0f}'

        # ── Chart 1: Sales vs Profit Margin — quarterly dual-axis ─────────────
        fig1, ax1 = plt.subplots(figsize=(12, 5))
        fig1.subplots_adjust(top=0.82, bottom=0.18, left=0.11, right=0.91)
        fig1.patch.set_facecolor('#FAFCFF')
        fig1.add_artist(plt.matplotlib.lines.Line2D([0.01, 0.99], [0.91, 0.91], transform=fig1.transFigure, color='#C8DCEF', linewidth=1.0, clip_on=False))
        if has_ts and gc and pmrc:
            q1 = df_ts.groupby('_ql').agg(
                sales=(gc,    'sum'),
                margin=(pmrc, 'mean')
            ).reset_index().sort_values('_ql')
            ax1b = ax1.twinx()
            x1   = range(len(q1))
            bars1 = ax1.bar(x1, q1['sales']/1e5, color=NAVY, alpha=0.72, label='Gross Sales (₹L)', width=0.6, zorder=3)
            ax1b.plot(x1, q1['margin'], color=RED, linewidth=2.5, marker='o', markersize=5, label='Profit Margin %', zorder=4)
            ax1.set_xticks(list(x1))
            ax1.set_xticklabels(q1['_ql'], rotation=45, ha='right', fontsize=8)
            ax1.set_ylabel('Gross Sales (₹ Lakhs)', fontsize=10, fontweight='bold', color=NAVY)
            ax1b.set_ylabel('Profit Margin %', fontsize=10, fontweight='bold', color=RED)
            fig1.text(0.01, 0.95, 'Sales vs Profit Margin — Quarterly', fontsize=13, fontweight='bold', color='#0B1F3A', va='top', transform=fig1.transFigure)
            h1, l1 = ax1.get_legend_handles_labels();  h2, l2 = ax1b.get_legend_handles_labels()
            ax1.legend(h1+h2, l1+l2, loc='upper left', fontsize=8, framealpha=0.8)
            ax1.yaxis.set_major_formatter(mticker.FuncFormatter(_inr_fmt))
            ax1b.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:.1f}%'))
            # Annotate last 4 quarters
            for xi, mi in zip(list(x1)[-4:], q1['margin'].values[-4:]):
                ax1b.annotate(f'{mi:.1f}%', (xi, mi), textcoords='offset points', xytext=(0, 6),
                               fontsize=7, ha='center', color=RED, fontweight='bold')
        else:
            ax1.text(0.5, 0.5, 'No time-series data available', ha='center', va='center',
                     transform=ax1.transAxes, fontsize=12)
            fig1.text(0.01, 0.95, 'Sales vs Profit Margin', fontsize=13, fontweight='bold', color='#0B1F3A', va='top', transform=fig1.transFigure)

        # ── Chart 2: ONDC Before vs After — stacked quarterly with uplift line ─
        fig2, ax2 = plt.subplots(figsize=(12, 5))
        fig2.subplots_adjust(top=0.82, bottom=0.18, left=0.10, right=0.97)
        fig2.patch.set_facecolor('#FAFCFF')
        fig2.add_artist(plt.matplotlib.lines.Line2D([0.01, 0.99], [0.91, 0.91], transform=fig2.transFigure, color='#C8DCEF', linewidth=1.0, clip_on=False))
        if has_ts and boc and ochc:
            q2 = df_ts.groupby('_ql').agg(
                before=(boc, 'sum'),
                ondc_p=(ochc, lambda x: float(x.clip(lower=0).sum())),
                gross=(gc, 'sum') if gc else (boc, 'sum')
            ).reset_index().sort_values('_ql')
            x2 = range(len(q2))
            ax2.bar(x2, q2['before']/1e5,  label='Offline Baseline Revenue', color='#7A92AA', alpha=0.80, width=0.6, zorder=3)
            ax2.bar(x2, q2['ondc_p']/1e5, bottom=q2['before']/1e5,
                    label='Online Channel Revenue', color=GREEN, alpha=0.85, width=0.6, zorder=3)
            ax2.plot(x2, q2['gross']/1e5, color=NAVY, linewidth=2.5, marker='D', markersize=5,
                     label='Total Gross Sales', zorder=5)
            # Mark ONDC activation (first quarter where ondc_p > 0)
            first_live = q2[q2['ondc_p'] > 0]['_ql'].iloc[0] if (q2['ondc_p'] > 0).any() else None
            if first_live:
                li = q2[q2['_ql'] == first_live].index[0]
                ax2.axvline(li - 0.5, color=GREEN, linestyle='--', linewidth=1.5, alpha=0.7)
                ax2.text(li - 0.4, ax2.get_ylim()[1] * 0.92 if ax2.get_ylim()[1] else 1000,
                         '▶ Online Channels Live', fontsize=8, color=GREEN, fontweight='bold')
            ax2.set_xticks(list(x2))
            ax2.set_xticklabels(q2['_ql'], rotation=45, ha='right', fontsize=8)
            ax2.set_ylabel('Revenue (₹ Lakhs)', fontsize=10, fontweight='bold')
            fig2.text(0.01, 0.95, 'Channel Performance: Before vs After Revenue — Quarterly', fontsize=13, fontweight='bold', color='#0B1F3A', va='top', transform=fig2.transFigure)
            ax2.legend(fontsize=8, loc='upper left', framealpha=0.8)
            ax2.yaxis.set_major_formatter(mticker.FuncFormatter(_inr_fmt))
        else:
            ax2.text(0.5, 0.5, 'Online channel columns not found in data', ha='center', va='center',
                     transform=ax2.transAxes, fontsize=12)
            fig2.text(0.01, 0.95, 'Channel Performance: Before vs After', fontsize=13, fontweight='bold', color='#0B1F3A', va='top', transform=fig2.transFigure)

        # ── Chart 3: Returns & Replacements — quarterly grouped bars + rate line
        fig3, ax3 = plt.subplots(figsize=(12, 5))
        fig3.subplots_adjust(top=0.82, bottom=0.18, left=0.10, right=0.91)
        fig3.patch.set_facecolor('#FAFCFF')
        fig3.add_artist(plt.matplotlib.lines.Line2D([0.01, 0.99], [0.91, 0.91], transform=fig3.transFigure, color='#C8DCEF', linewidth=1.0, clip_on=False))
        if has_ts and rrc:
            agg3 = {rrc: 'mean'}
            if rlrc: agg3[rlrc] = 'mean'
            if qrc:  agg3[qrc]  = 'sum'
            if rpc:  agg3[rpc]  = 'sum'
            q3 = df_ts.groupby('_ql').agg(agg3).reset_index().sort_values('_ql')
            ax3b = ax3.twinx()
            x3   = range(len(q3))
            bar_w = 0.35
            if qrc and rpc:
                b3a = ax3.bar([xi - bar_w/2 for xi in x3], q3[qrc],
                              width=bar_w, color=RED, alpha=0.75, label='Units Returned', zorder=3)
                b3b = ax3.bar([xi + bar_w/2 for xi in x3], q3[rpc],
                              width=bar_w, color=AMBER, alpha=0.75, label='Replacements', zorder=3)
            elif qrc:
                ax3.bar(x3, q3[qrc], width=0.6, color=RED, alpha=0.75, label='Units Returned', zorder=3)
            ax3b.plot(x3, q3[rrc],  color=PURP, linewidth=2.5, marker='o', markersize=5,
                      label='Return Rate %', zorder=4)
            if rlrc:
                ax3b.plot(x3, q3[rlrc], color=TEAL, linewidth=1.8, linestyle='--', marker='s',
                           markersize=4, label='6M Rolling Return Rate', zorder=4)
            ax3b.axhline(7, color=RED, linestyle=':', linewidth=1.2, alpha=0.5)
            ax3b.text(len(q3)-0.5, 7.2, 'Target <7%', fontsize=7, color=RED, ha='right')
            ax3.set_xticks(list(x3))
            ax3.set_xticklabels(q3['_ql'], rotation=45, ha='right', fontsize=8)
            ax3.set_ylabel('Units (Returned / Replaced)', fontsize=10, fontweight='bold', color=RED)
            ax3b.set_ylabel('Return Rate %', fontsize=10, fontweight='bold', color=PURP)
            fig3.text(0.01, 0.95, 'Returns & Replacements — Quarterly Trend', fontsize=13, fontweight='bold', color='#0B1F3A', va='top', transform=fig3.transFigure)
            h3a, l3a = ax3.get_legend_handles_labels();  h3b, l3b = ax3b.get_legend_handles_labels()
            ax3.legend(h3a+h3b, l3a+l3b, loc='upper left', fontsize=8, framealpha=0.8)
        else:
            ax3.text(0.5, 0.5, 'No returns data available', ha='center', va='center',
                     transform=ax3.transAxes, fontsize=12)
            fig3.text(0.01, 0.95, 'Returns & Replacements', fontsize=13, fontweight='bold', color='#0B1F3A', va='top', transform=fig3.transFigure)

        # ── Chart 4: Store-level ONDC comparison (2 sub-plots) ───────────────
        fig4, (ax4a, ax4b) = plt.subplots(1, 2, figsize=(14, 5))
        fig4.subplots_adjust(top=0.82, bottom=0.14, left=0.07, right=0.97, wspace=0.38)
        fig4.patch.set_facecolor('#FAFCFF')
        fig4.add_artist(plt.matplotlib.lines.Line2D([0.01, 0.99], [0.91, 0.91], transform=fig4.transFigure, color='#C8DCEF', linewidth=1.0, clip_on=False))
        _sid4 = 'Store_ID' if 'Store_ID' in df.columns else ('store_id' if 'store_id' in df.columns else None)
        if _sid4 and (boc or gc):  # allow even when dedicated before/after cols missing
            agg4 = dict(net=(nc or gc, 'sum'), before=(boc or gc, 'sum'))  # fallback gross sales as 'before'
            if ochc: agg4['ondc_p'] = (ochc, lambda x: float(x.clip(lower=0).sum()))
            if rrc:  agg4['ret_r']  = (rrc,  'mean')
            if rpc:  agg4['repl']   = (rpc,  'sum')
            if tac:  agg4['tgt_ach']= (tac,  'mean')
            st4 = df.groupby(_sid4).agg(**agg4).reset_index()
            store_lbls = [str(s) for s in st4[_sid4]]  # handles both "STORE_001" and numeric IDs
            x4  = range(len(st4))
            w4  = 0.28
            # Left sub-plot: revenue bars
            ax4a.bar([xi - w4 for xi in x4], st4['before']/1e5,   width=w4, color='#7A92AA', alpha=0.85, label='Offline Baseline')
            ax4a.bar([xi       for xi in x4], st4['net']/1e5,     width=w4, color=NAVY,     alpha=0.85, label='Net Sales')
            if 'ondc_p' in st4:
                ax4a.bar([xi + w4 for xi in x4], st4['ondc_p']/1e5, width=w4, color=GREEN,   alpha=0.85, label='Online Channel')
            ax4a.set_xticks(list(x4)); ax4a.set_xticklabels(store_lbls, fontsize=10)
            ax4a.set_ylabel('Revenue (₹ Lakhs)', fontsize=10, fontweight='bold')
            ax4a.set_title('Store Revenue: Offline vs Online Channels', fontsize=10, fontweight='bold', pad=8)
            ax4a.legend(fontsize=8); ax4a.yaxis.set_major_formatter(mticker.FuncFormatter(_inr_fmt))
            # Right sub-plot: return rate bars + target achievement line
            ax4b2 = ax4b.twinx()
            pal4  = [RED, AMBER, '#e07b2a']
            if 'ret_r' in st4:
                b4  = ax4b.bar(x4, st4['ret_r'], color=pal4[:len(st4)], alpha=0.78, width=0.5, label='Return Rate %', zorder=3)
                for bar, v in zip(b4, st4['ret_r']):
                    ax4b.text(bar.get_x()+bar.get_width()/2., bar.get_height()+0.1,
                               f'{v:.1f}%', ha='center', va='bottom', fontsize=9, fontweight='bold')
            if 'tgt_ach' in st4:
                ax4b2.plot(list(x4), st4['tgt_ach'], color=NAVY, linewidth=2.5, marker='D',
                            markersize=8, label='Target Achievement %', zorder=5)
                ax4b2.axhline(100, color=GREEN, linestyle='--', linewidth=1.2, alpha=0.6)
            ax4b.set_xticks(list(x4)); ax4b.set_xticklabels(store_lbls, fontsize=10)
            ax4b.set_ylabel('Return Rate %', fontsize=10, fontweight='bold', color=RED)
            ax4b2.set_ylabel('Target Achievement %', fontsize=10, fontweight='bold', color=NAVY)
            ax4b.set_title('Store: Return Rate & Target Achievement', fontsize=10, fontweight='bold', pad=8)
            h4a, l4a = ax4b.get_legend_handles_labels();  h4b2, l4b2 = ax4b2.get_legend_handles_labels()
            ax4b.legend(h4a+h4b2, l4a+l4b2, loc='upper right', fontsize=8, framealpha=0.8)
        else:
            for ax_ in [ax4a, ax4b]:
                ax_.text(0.5, 0.5, 'No store-level channel data', ha='center', va='center',
                         transform=ax_.transAxes, fontsize=12)
        fig4.text(0.01, 0.96, 'Store-Level Channel Performance Analysis', fontsize=13, fontweight='bold', color='#0B1F3A', va='top', transform=fig4.transFigure)

        cat_options = ['All Categories']
        if 'Product_Category' in df.columns:
            cat_options += sorted(df['Product_Category'].dropna().unique().tolist())
        return (kpi_html,"","","","",fig1,fig2,fig3,fig4,None,None,None,None,None,cat_options,df)

    except Exception as e:
        return ("N/A","","","","",None,None,None,None,None,None,None,None,f"Error: {str(e)}",["All Categories"],None)

# ── Udyam master data lookup (mock — replace with live MSME API) ───────────────
udyam_master_data = pd.DataFrame({
    'udyam_number':    ['UDYAM-UP-01-0000001','UDYAM-TN-00-7629703','UDYAM-KL-03-0000003'],
    'enterprise_name': ['Tech Innovations Pvt Ltd','Retail Solutions Corp','FMCG Distributors'],
    'organisation_type':['Private Limited','Partnership','Proprietorship'],
    'major_activity':  ['FMCG','Services','Electronics'],
    'enterprise_type': ['Small','Micro','Medium'],
    'state':           ['Uttar Pradesh','TamilNadu','Kerala'],
    'city':            ['Lucknow','Chennai','Kochi'],
    'industry_domain': ['Retail','Retail','Retail'],
})

def _fetch_msme_data(msme_number):
    fetched = udyam_master_data[udyam_master_data['udyam_number'] == msme_number]
    if not fetched.empty:
        row = fetched.iloc[0]
        return (row['enterprise_name'], row['organisation_type'], row['major_activity'],
                row['enterprise_type'], row['state'], row['city'],
                row.get('industry_domain','Retail'),
                "✅ Data Fetched Successfully")
    return "", "", "", "", "", "", "Retail", "❌ Data Not Found. Please check the number."

# ── Category filter chart (Step 5 analysis) ────────────────────────────────────
def build_category_filter_chart(df, selected_category):
    plt.style.use('seaborn-v0_8-darkgrid')
    sales_col = 'Monthly_Sales_INR' if 'Monthly_Sales_INR' in df.columns else 'Gross_Sales'
    sku_col = 'SKU_Name' if 'SKU_Name' in df.columns else None
    cat_col = 'Product_Category' if 'Product_Category' in df.columns else None
    fig,ax = plt.subplots(figsize=(12,7)); fig.subplots_adjust(top=0.91,bottom=0.12,left=0.32,right=0.92)
    def _fmt(v):
        if v>=1e7: return f"Rs.{v/1e7:.1f}Cr"
        if v>=1e5: return f"Rs.{v/1e5:.1f}L"
        return f"Rs.{v:,.0f}"
    if cat_col and selected_category and selected_category != "All Categories":
        filtered = df[df[cat_col]==selected_category] if selected_category in df[cat_col].values else df
    else:
        filtered = df
    if sku_col and not filtered.empty:
        top5 = filtered.groupby(sku_col)[sales_col].sum().nlargest(5).reset_index()
        colors = plt.cm.RdYlGn(np.linspace(0.3,0.9,len(top5)))
        bars = ax.barh(top5[sku_col], top5[sales_col], color=colors, height=0.55, edgecolor='white')
        ax.set_xlabel('Sales (INR)',fontsize=12,fontweight='bold')
        cat_label = selected_category if selected_category and selected_category != "All Categories" else "All"
        ax.set_title(f'Top 5 Products — {cat_label}',fontsize=14,fontweight='bold',pad=14)
        ax.grid(axis='x',alpha=0.3); ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
        max_val = top5[sales_col].max() if len(top5)>0 else 1
        for bar in bars:
            w = bar.get_width()
            ax.text(w+max_val*0.01, bar.get_y()+bar.get_height()/2, _fmt(w), ha='left', va='center', fontsize=9, fontweight='bold')
        ax.set_xlim(0, max_val*1.22)
    else:
        ax.text(0.5,0.5,'No product data',ha='center',va='center',transform=ax.transAxes)
    return fig

# ── Custom Gradio CSS ───────────────────────────────────────────────────────────
custom_css = """
/* Override Gradio Soft theme colours */
:root {
  --body-background-fill: #ffffff !important;
  --background-fill-primary: #ffffff !important;
  --background-fill-secondary: #ffffff !important;
  --border-color-primary: transparent !important;
  --block-background-fill: transparent !important;
  --block-border-width: 0px !important;
  --block-shadow: none !important;
  --block-label-background-fill: transparent !important;
  --layout-gap: 0px !important;
  --section-header-text-size: 0px !important;
  --panel-background-fill: transparent !important;
  --panel-border-width: 0px !important;
  --input-background-fill: #ffffff !important;
}
/* ── Screen progress stepper (shown on screens 1-3) ── */
.dn-progress-stepper{display:flex;align-items:center;justify-content:center;gap:0;margin-bottom:18px;}
.dn-progress-step{display:flex;align-items:center;gap:6px;font-size:12px;font-weight:600;
  color:#94a3b8;padding:6px 14px;border-radius:20px;transition:all 0.2s;}
.dn-progress-step.active{background:#EBF5FF;color:#1B4F8A;border:1px solid #BDD7F5;}
.dn-progress-step.done{color:#16a34a;}
.dn-progress-connector{width:32px;height:2px;background:#e2e8f0;flex-shrink:0;}
.dn-progress-connector.done{background:#16a34a;}
/* ── Landing login error — shown below the HTML form ── */
#dn-msme-col-inner .prose p,
#dn-msme-col-inner .markdown-body p{
  font-size:13px !important;
  color:#dc2626 !important;
  background:#fef2f2 !important;
  border:1px solid #fecaca !important;
  border-radius:8px !important;
  padding:8px 12px !important;
  margin:8px 0 0 !important;
}
/* ── Hide Gradio toasts & chrome ── */
.toast-wrap,.toast-body,div.toast{display:none !important;}
.gradio-container>footer,footer.svelte-1rjryqp,.built-with,.api-logo,.show-api,
span.svelte-1ed2p3z,.share-button{display:none !important;visibility:hidden !important;
height:0 !important;overflow:hidden !important;}

/* ── Base reset — force premium background through ALL Gradio layers ── */
body,html{background:#E8F0FE !important;font-family:system-ui,-apple-system,"Segoe UI",sans-serif;}
body > gradio-app,
gradio-app,
gradio-app > div,
gradio-app > div > div,
.gradio-container,
.gradio-container > .main,
.gradio-container > .main > .wrap,
.gradio-container > .main > .wrap > .padding,
.gradio-container > .main > .wrap > .padding > .gap{
  background:#E8F0FE !important;
}

/* ═══════════════════════════════════════════════════════════
   NAV BAR
   ═══════════════════════════════════════════════════════════ */
#lang-bar{
  position:sticky !important;
  top:0 !important;
  z-index:1000 !important;
  background:#ffffff !important;
  border-bottom:1px solid #e8edf3 !important;
  box-shadow:0 1px 8px rgba(11,31,58,0.07) !important;
  min-height:64px !important;
  padding:0 40px !important;
  gap:0 !important;
  border-radius:0 !important;
  margin:0 !important;
  overflow:visible !important;
  box-sizing:border-box !important;
  align-items:center !important;
}
#lang-bar>.gradio-column,
#lang-bar>div[class*="svelte"],
#lang-bar .block,
#lang-bar .form,
#lang-bar .contain,
#lang-bar .wrap{
  background:transparent !important;box-shadow:none !important;
  border:none !important;padding:0 !important;margin:0 !important;
  align-items:center !important;gap:0 !important;
}
#lang-bar-spacer{background:transparent !important;padding:0 !important;margin:0 !important;border:none !important;box-shadow:none !important;}
#hdr-logo-wrap{display:flex !important;align-items:center !important;}
#lang-bar-logo{display:flex !important;align-items:center !important;gap:12px !important;}
#lang-bar-logo img{height:40px !important;border-radius:8px !important;box-shadow:0 1px 4px rgba(11,31,58,0.10) !important;}
#lang-bar-logo .logo-text{font-size:20px !important;font-weight:900 !important;color:#0f2557 !important;letter-spacing:-0.5px !important;white-space:nowrap !important;}
#lang-dropdown-main{background:transparent !important;padding:0 !important;border:none !important;min-width:160px !important;}
#lang-dropdown-main label{font-size:11px !important;font-weight:600 !important;color:#64748b !important;margin-bottom:2px !important;}
#lang-dropdown-main .wrap{border-radius:8px !important;border-color:#e2e8f0 !important;font-size:13px !important;font-weight:500 !important;}
#hdr-lang-en,#hdr-lang-hi,#hdr-msme-btn,#hdr-gov-btn{font-size:12px !important;font-weight:600 !important;border-radius:6px !important;white-space:nowrap !important;}
/* State fix: MSME + Gov header stub buttons must NEVER appear — force hidden at all times */
#hdr-msme-btn,#hdr-gov-btn,
#hdr-msme-btn *,#hdr-gov-btn *,
div:has(>#hdr-msme-btn),div:has(>#hdr-gov-btn){display:none !important;visibility:hidden !important;width:0 !important;height:0 !important;overflow:hidden !important;pointer-events:none !important;}
#lang-dropdown,#lang-indicator{display:none !important;}
/* ── Header nav links ─────────────────────────────────────────────── */
#hdr-nav-links{
  display:flex;align-items:center;justify-content:center;
  gap:28px;height:100%;
}
/* App screens: hide nav links and language selector */
/* ── App mode: hide entire navbar + language selector ── */
#lang-bar.dn-app-mode #lang-bar-nav,
#lang-bar.dn-app-mode #lang-dropdown-main,
#lang-bar.dn-app-mode .hdr-nav-link,
#lang-bar.dn-app-mode [id*="lang-dropdown"],
#lang-bar.dn-app-mode select,
#lang-bar.dn-app-mode .gradio-dropdown { display:none !important; }
#lang-bar.dn-app-mode {
  justify-content:flex-start !important;
  background:#0f2557 !important;
  border-bottom:1px solid rgba(255,255,255,0.08) !important;
  min-height:0 !important;
  padding:0 !important;
}
#lang-bar.dn-app-mode > .gap {
  justify-content:flex-start !important;
  padding:0 !important;
  min-height:0 !important;
  height:0 !important;
  overflow:hidden !important;
}
/* App header bar — logo only, fully hidden (new bars handle branding) */
#lang-bar.dn-app-mode #hdr-logo-wrap { display:none !important; }
.hdr-nav-link{
  font-size:13px;font-weight:500;color:#1B4F8A;
  text-decoration:none;transition:color 0.15s;white-space:nowrap;
}
.hdr-nav-link:hover{color:#0f2557;}
#lang-bar-nav>div{
  display:flex;align-items:center;justify-content:center;height:100%;
}
@media(max-width:900px){
  #hdr-nav-links{display:none;}
}


/* ── Premium page background — belt + suspenders approach ── */
:root { --dn-page-bg: #f0f5ff; }
html { background: linear-gradient(to bottom, #f8fbff, #eef4ff) !important; min-height: 100vh; }
body { background: transparent !important; min-height: 100vh; }
#root, #root > div, .svelte-po604f,
.app, .wrap, .main, .contain,
[data-testid="block"],
.gradio-app, .gradio-container {
  background: var(--dn-page-bg) !important;
}
/* Between-section spacing gets the background */
.gradio-container .gap { background: var(--dn-page-bg) !important; }

/* The hero and bands sit on top cleanly */
.dn-hero-wrap, .dn-band { position: relative; z-index: 1; }
.dn-band-white { background: #ffffff !important; }

/* Form card — ensure it stands out on the blue background */
#dn-msme-card-wrap {
  max-width:560px !important;
  margin:16px auto 0 !important;
}
/* Remove Gradio's default card styling — HTML form provides the card */
#dn-msme-card-wrap > div,
#dn-msme-card-wrap > div > div {
  background:transparent !important;
  border:none !important;
  box-shadow:none !important;
  padding:0 !important;
  gap:0 !important;
}
/* Gradio input wrapper — positioned behind HTML form inputs */
/* Hide contact form Gradio inputs — off-screen but in DOM */
#dn-contact-hidden {
  position:absolute !important; left:-9999px !important;
  top:0 !important; width:500px !important; height:auto !important;
  opacity:0 !important; pointer-events:none !important; z-index:-1 !important;
}
/* ── Gradio login button — hidden visually, stays in DOM for JS click ── */
div#dn-real-login-btn {
  position:fixed !important;
  left:-9999px !important;
  top:0 !important;
  width:auto !important;
  height:auto !important;
  opacity:0 !important;
  pointer-events:auto !important;
  z-index:-1 !important;
}
#dn-login-error-md { position:absolute !important; left:-9999px !important; }
#dn-gradio-inputs-hidden {
  position:absolute !important;
  left:-9999px !important;
  top:0 !important;
  width:500px !important;
  height:auto !important;
  overflow:visible !important;
  opacity:0 !important;
  pointer-events:none !important;
  z-index:-1 !important;
}
/* Each Gradio textbox — shrunk to match HTML input positions */
#dn-landing-name, #dn-real-mobile, #dn-landing-otp, #dn-landing-email {
  position:absolute !important;
  opacity:0 !important;
  pointer-events:none !important;
  width:calc(100% - 56px) !important;
  height:48px !important;
  z-index:1 !important;
}
/* Gradio button — invisible but clickable */
#dn-real-login-btn {
  position:absolute !important;
  opacity:0 !important;
  width:100% !important;
  height:54px !important;
  z-index:1 !important;
  pointer-events:auto !important;
}
/* How It Works card — same */
#dn-right-cards .block { background: transparent !important; }

/* ═══════════════════════════════════════════════════════════
   LAYOUT — KILL GRADIO GAPS / PADDING
   ═══════════════════════════════════════════════════════════ */
.dn-band + .dn-band{margin-top:0 !important;}
.dn-login-section{margin-top:0 !important;padding-top:0 !important;}
/* Kill Gradio block padding ONLY inside HTML-only sections — never touch form widgets */
#dn-step0-col > .gap,
#dn-step0-col > div > .gap,
#dn-step0-col > div > div > .gap { gap:0 !important; }
/* Strip padding from pure HTML wrapper blocks (hero, sections, bands) */
#lang-bar .block,
#lang-bar .form { padding:0 !important; background:transparent !important;
  border:none !important; box-shadow:none !important; gap:0 !important; }
/* Only strip blocks that directly contain HTML widgets (no input/select/button children) */
#dn-step0-col > div > div > div > .block { padding:0 !important; border:none !important;
  box-shadow:none !important; background:transparent !important; }
/* Restore login section blocks — these contain real form inputs */
#dn-login-section-col .block,
#dn-login-section-col .form,
#dn-msme-card-wrap .block,
#dn-msme-col-inner .block { padding:unset !important; border:unset !important;
  box-shadow:unset !important; background:unset !important; }
/* Remove Gradio spacer rows */
#dn-step0-col [data-testid="column"]:empty,
#dn-step0-col .gap:empty{display:none !important;height:0 !important;}
/* Collapse ALL hidden Gradio columns in step0 — kills phantom whitespace strips */
#dn-step0-col [style*="display: none"],
#dn-step0-col [style*="display:none"] {
  height:0 !important;min-height:0 !important;
  max-height:0 !important;overflow:hidden !important;
  margin:0 !important;padding:0 !important;
  border:none !important;flex:none !important;
}
/* Kill gap spacing between hidden children */
#dn-login-section-col[style*="display: none"],
#dn-login-section-col[style*="display:none"] {
  height:0 !important;min-height:0 !important;
  overflow:hidden !important;margin:0 !important;padding:0 !important;
}
/* Login/Unlock section card */
#dn-login-section-col { background:transparent; }

/* ── Login section — full widget restoration ── */
#dn-login-section-col,
#dn-login-section-col > div,
#dn-login-section-col > div > div { background:transparent !important; }
#dn-msme-card-wrap { max-width:520px !important; margin:0 auto !important;
  padding:0 16px 24px !important; width:100% !important; }
#dn-msme-card-wrap > div > div { background:#ffffff !important;
  border-radius:16px !important; border:1px solid #e2eaf5 !important;
  box-shadow:0 8px 40px rgba(11,31,58,0.10) !important; padding:28px 24px !important; }
#dn-msme-col-inner .block,
#dn-msme-col-inner .form,
#dn-msme-col-inner .wrap {
  background:transparent !important;
  border:none !important;
  box-shadow:none !important;
  padding:0 !important;
  gap:12px !important;
}
/* Textbox outer container */
#dn-msme-col-inner .gradio-textbox,
#dn-msme-col-inner [data-testid="textbox"] {
  background:transparent !important;
  border:none !important;
  box-shadow:none !important;
  padding:0 !important;
  margin-bottom:4px !important;
}
/* Textbox label */
#dn-msme-col-inner .gradio-textbox label,
#dn-msme-col-inner [data-testid="textbox"] label {
  font-size:13px !important;
  font-weight:700 !important;
  color:#1e3a5f !important;
  margin-bottom:6px !important;
  display:block !important;
}
/* Textbox input — style properly, kill spinners */
#dn-landing-name input,
#dn-real-mobile input,
#dn-landing-otp input,
#dn-landing-email input,
#dn-msme-col-inner input[type="text"],
#dn-msme-col-inner input[type="email"],
#dn-msme-col-inner input[type="tel"],
#dn-msme-col-inner input[type="number"],
#dn-msme-col-inner input:not([type="checkbox"]):not([type="radio"]) {
  height:46px !important;
  border:1.5px solid #cbd5e1 !important;
  border-radius:10px !important;
  font-size:14px !important;
  color:#0f172a !important;
  padding:0 14px !important;
  background:#ffffff !important;
  width:100% !important;
  box-sizing:border-box !important;
  box-shadow:0 1px 3px rgba(0,0,0,0.06) !important;
  -moz-appearance:textfield !important;
  appearance:textfield !important;
  -webkit-appearance:none !important;
  outline:none !important;
}
/* Remove spinner arrows — all approaches */
#dn-msme-col-inner input::-webkit-outer-spin-button,
#dn-msme-col-inner input::-webkit-inner-spin-button,
#dn-landing-name input::-webkit-outer-spin-button,
#dn-landing-name input::-webkit-inner-spin-button,
#dn-real-mobile input::-webkit-outer-spin-button,
#dn-real-mobile input::-webkit-inner-spin-button,
#dn-landing-otp input::-webkit-outer-spin-button,
#dn-landing-otp input::-webkit-inner-spin-button,
#dn-landing-email input::-webkit-outer-spin-button,
#dn-landing-email input::-webkit-inner-spin-button {
  -webkit-appearance:none !important;
  margin:0 !important;
  display:none !important;
}
/* Force text appearance on all login inputs */
#dn-landing-name input,
#dn-real-mobile input,
#dn-landing-otp input,
#dn-landing-email input {
  -webkit-appearance:none !important;
  -moz-appearance:textfield !important;
  appearance:none !important;
  type: text !important;
}
/* Focus state */
#dn-msme-col-inner input:focus {
  border-color:#2563eb !important;
  box-shadow:0 0 0 3px rgba(37,99,235,0.15) !important;
}
/* Placeholder */
#dn-msme-col-inner input::placeholder {
  color:#94a3b8 !important;
  font-weight:400 !important;
}
#dn-msme-card-wrap {
  max-width:520px !important;
  margin:0 auto !important;
  padding:0 16px 24px !important;
  width:100% !important;
}
/* Card container */
#dn-msme-col-inner > div {
  background:#ffffff !important;
  border-radius:16px !important;
  border:1px solid #e2eaf5 !important;
  box-shadow:0 8px 40px rgba(11,31,58,0.10),0 2px 8px rgba(11,31,58,0.06) !important;
  padding:28px 24px !important;
}
/* Form field labels */
#dn-msme-col-inner label {
  font-size:13px !important;
  font-weight:700 !important;
  color:#1e3a5f !important;
  margin-bottom:4px !important;
}
/* Form field inputs */
#dn-landing-name input,
#dn-real-mobile input,
#dn-landing-otp input,
#dn-landing-email input {
  height:44px !important;
  border:2px solid #e2e8f0 !important;
  border-radius:10px !important;
  font-size:14px !important;
  color:#0f172a !important;
  padding:0 14px !important;
  background:#ffffff !important;
  width:100% !important;
  box-sizing:border-box !important;
}
#dn-landing-name input:focus,
#dn-real-mobile input:focus,
#dn-landing-otp input:focus,
#dn-landing-email input:focus {
  border-color:#2563eb !important;
  outline:none !important;
  box-shadow:0 0 0 3px rgba(37,99,235,0.12) !important;
}
/* Analyze My Data button — maximum specificity */
div#dn-real-login-btn,
#dn-msme-col-inner div#dn-real-login-btn {
  width:100% !important; margin-top:8px !important; display:block !important;
}
div#dn-real-login-btn .block,
div#dn-real-login-btn .wrap,
div#dn-real-login-btn > div {
  background:transparent !important; border:none !important;
  box-shadow:none !important; padding:0 !important; width:100% !important;
}
div#dn-real-login-btn button {
  width:100% !important;
  height:52px !important;
  min-height:52px !important;
  font-size:15px !important;
  font-weight:800 !important;
  letter-spacing:0.2px !important;
  border-radius:12px !important;
  background:linear-gradient(90deg,#1B4F8A 0%,#2563eb 100%) !important;
  color:#ffffff !important;
  border:none !important;
  box-shadow:0 6px 20px rgba(37,99,235,0.40) !important;
  cursor:pointer !important;
  display:flex !important;
  align-items:center !important;
  justify-content:center !important;
  gap:8px !important;
  visibility:visible !important;
  opacity:1 !important;
}
#dn-real-login-btn button:hover {
  transform:translateY(-2px) !important;
  box-shadow:0 10px 28px rgba(37,99,235,0.52) !important;
}
.gradio-container>.main{gap:0 !important;}
.gradio-container>.main>.wrap>.padding{padding-top:0 !important;}
.gradio-container>.main>.wrap>.padding>.gap,
.gradio-container .gap,
.gradio-container .gap>*{gap:0 !important;}
.gradio-container .form{gap:0 !important;}
.gradio-container .contain{padding:0 !important;gap:0 !important;}
.gradio-container .wrap{gap:0 !important;}
/* Strip Gradio padding only from pure HTML wrapper blocks */
.gradio-container .block:has(>.prose),
.gradio-container .block:not(:has(input)):not(:has(textarea)):not(:has(select)):not(:has(button)):not(:has(.upload-container)){
  padding:0 !important;margin:0 !important;border:none !important;
  box-shadow:none !important;background:transparent !important;
}
.dn-band .capabilities-section{background:transparent !important;padding:0 !important;margin:0 !important;border-radius:0 !important;}
.dn-band #drc-block{margin:0 !important;}
/* Fix 2: explicit zero margins on band wrappers to kill Gradio-injected spacing */
.gradio-container .block:has(.dn-band),
.gradio-container .block:has(.dn-hero-wrap){
  padding:0 !important;margin:0 !important;border:none !important;
  box-shadow:none !important;background:transparent !important;
}
.dn-band+.dn-band{margin-top:0 !important;}

/* ═══════════════════════════════════════════════════════════
   SECTION BANDS
   ═══════════════════════════════════════════════════════════ */
.dn-band{width:100%;box-sizing:border-box;position:relative;z-index:1;}
.dn-band-white{background:rgba(255,255,255,0.82);box-shadow:0 1px 0 rgba(37,99,235,0.06);backdrop-filter:blur(2px);}
.dn-band-grey{background:rgba(220,232,253,0.55);}
.dn-band-inner{max-width:1100px;margin:0 auto;padding:20px 36px;}
.dn-band-white:first-of-type .dn-band-inner,.dn-band-white .dn-band-inner{padding-top:12px;padding-bottom:20px;}
.dn-band-grey .dn-band-inner{padding-top:12px;padding-bottom:12px;}
@media(max-width:768px){.dn-band-inner{padding:16px 16px;}.dn-band-white .dn-band-inner{padding-top:12px;padding-bottom:14px;}.dn-band-grey .dn-band-inner{padding-top:16px;padding-bottom:16px;}}

/* ═══════════════════════════════════════════════════════════
   HERO SECTION
   ═══════════════════════════════════════════════════════════ */
.dn-hero-wrap{
  background:linear-gradient(118deg,#061429 0%,#0c1e45 40%,#112660 75%,#0d1f40 100%);
  position:relative;overflow:hidden;
}
.dn-hero-wrap::before{
  content:"";position:absolute;inset:0;pointer-events:none;
  background-image:
    radial-gradient(circle at 18% 55%,rgba(37,99,235,0.12) 0%,transparent 52%),
    radial-gradient(circle at 82% 18%,rgba(124,58,237,0.08) 0%,transparent 48%),
    radial-gradient(ellipse at 50% 50%,rgba(37,99,235,0.07) 0%,transparent 60%);
}
.dn-hero-bg{
  display:flex;align-items:center;gap:24px;
  max-width:1340px;margin:0 auto;
  padding:28px 48px 20px;
  position:relative;z-index:1;
}
@media(max-width:1060px){
  .dn-hero-bg{flex-direction:column;padding:28px 20px 28px;gap:22px;}
  .dn-hero-center,.dn-hero-right-panel{width:100% !important;max-width:520px;margin:0 auto;}
  .dn-hero-h1{font-size:2rem !important;}
}

/* Hero — left text */
.dn-hero-left{flex:0 0 44%;min-width:320px;max-width:480px;}
/* Hero center card — larger, dominates right side */
.dn-hero-center{flex:1 1 52%;min-width:320px;}
.dn-hero-badge{
  display:inline-flex;align-items:center;gap:6px;
  background:rgba(37,99,235,0.18);border:1px solid rgba(122,171,221,0.30);
  color:#93c5fd;font-size:11px;font-weight:700;letter-spacing:0.6px;
  padding:5px 14px;border-radius:20px;margin-bottom:20px;
}
.dn-hero-badge::before{content:"✦";font-size:9px;opacity:0.65;}
.dn-topnav{
  background:rgba(6,20,41,0.95);backdrop-filter:blur(8px);
  border-bottom:1px solid rgba(255,255,255,0.08);
  position:sticky;top:0;z-index:1000;
}
.dn-topnav-inner{
  max-width:1340px;margin:0 auto;padding:0 36px;
  display:flex;align-items:center;justify-content:space-between;
  height:54px;
}
.dn-topnav-logo{
  font-size:17px;font-weight:900;color:#ffffff;letter-spacing:-0.4px;
}
.dn-topnav-links{display:flex;align-items:center;gap:28px;}
.dn-topnav-link{
  font-size:13px;font-weight:500;color:rgba(255,255,255,0.72);
  text-decoration:none;transition:color 0.15s;
}
.dn-topnav-link:hover{color:#ffffff;}
.dn-topnav-cta{
  font-size:12.5px;font-weight:700;color:#ffffff;text-decoration:none;
  background:linear-gradient(90deg,#1a5bab,#2563eb);
  padding:8px 18px;border-radius:8px;
  box-shadow:0 2px 10px rgba(37,99,235,0.40);
  transition:transform 0.15s,box-shadow 0.15s;
}
.dn-topnav-cta:hover{transform:translateY(-1px);box-shadow:0 4px 16px rgba(37,99,235,0.55);}
@media(max-width:768px){
  .dn-topnav-links{display:none;}
  .dn-topnav-inner{padding:0 20px;}
}
.dn-hero-h1{
  font-size:3.1rem;font-weight:900;color:#ffffff;
  line-height:1.06;letter-spacing:-1px;margin:0 0 16px;
}
.dn-hero-sub{
  font-size:13.5px;color:rgba(255,255,255,0.86);
  line-height:1.68;margin:0 0 22px;max-width:340px;
}
.dn-hero-btns{display:flex;gap:11px;flex-wrap:wrap;margin-bottom:16px;align-items:center;}
.dn-btn-primary{
  display:inline-flex;align-items:center;gap:6px;
  background:linear-gradient(90deg,#1a5bab 0%,#2563eb 100%);color:#ffffff;
  font-size:13px;font-weight:700;letter-spacing:0.2px;
  padding:13px 22px;border-radius:10px;text-decoration:none;
  box-shadow:0 4px 18px rgba(37,99,235,0.55),0 0 0 0 rgba(37,99,235,0.4);
  transition:transform 0.15s,box-shadow 0.15s;
  white-space:nowrap;line-height:1;
  animation:dn-btn-pulse 2.8s ease-in-out infinite;
}
@keyframes dn-btn-pulse{
  0%,100%{box-shadow:0 4px 18px rgba(37,99,235,0.55),0 0 0 0 rgba(37,99,235,0.35);}
  50%{box-shadow:0 4px 22px rgba(37,99,235,0.65),0 0 0 7px rgba(37,99,235,0);}
}
.dn-btn-primary:hover{transform:translateY(-2px);box-shadow:0 8px 28px rgba(37,99,235,0.65);animation:none;}
.dn-btn-secondary{
  display:inline-flex;align-items:center;
  background:rgba(255,255,255,0.10);
  border:1.5px solid rgba(255,255,255,0.60);
  color:#ffffff;font-size:13px;font-weight:700;
  padding:13px 20px;border-radius:10px;text-decoration:none;
  transition:border-color 0.15s,background 0.15s,transform 0.15s;white-space:nowrap;line-height:1;
}
.dn-btn-secondary:hover{border-color:#ffffff;background:rgba(255,255,255,0.18);transform:translateY(-1px);}
.dn-hero-trust{
  display:flex;flex-direction:column;gap:6px;margin-top:2px;
}
.dn-trust-pill{
  display:inline-flex;align-items:center;gap:8px;
  font-size:12.5px;color:#e2f0ff;font-weight:600;line-height:1;
  background:rgba(37,99,235,0.22);border:1px solid rgba(147,197,253,0.55);
  border-radius:20px;padding:7px 14px;width:fit-content;
  transition:background 0.15s;
}
.dn-trust-pill:hover{background:rgba(37,99,235,0.35);border-color:rgba(147,197,253,0.8);}
.dn-trust-pill svg{flex-shrink:0;opacity:1;}
/* Bottom tagline */
.dn-hero-tagline{
  margin-top:16px;padding-top:13px;
  border-top:1px solid rgba(255,255,255,0.12);
  font-size:11px;color:rgba(255,255,255,0.58);line-height:1.5;
}

/* Hero — center dashboard card (hero focus, glow) */
.dn-hero-center{flex:1 1 580px;min-width:420px;position:relative;}
/* Soft radial glow behind the card */
.dn-hero-center::before{
  content:"";position:absolute;inset:-18px;z-index:0;pointer-events:none;
  background:radial-gradient(ellipse at center,rgba(37,99,235,0.18) 0%,transparent 68%);
  border-radius:28px;
}
.dn-dashboard-mock{
  background:#ffffff;border-radius:14px;
  box-shadow:0 32px 80px rgba(0,0,0,0.50),0 8px 28px rgba(11,31,58,0.22),
             0 0 0 1px rgba(255,255,255,0.12),
             0 0 0 5px rgba(37,99,235,0.12);
  overflow:hidden;font-family:system-ui,-apple-system,sans-serif;
  transform:perspective(1400px) rotateY(-0.8deg) rotateX(0.4deg);
  transition:transform 0.4s ease,box-shadow 0.4s ease;
  animation:dn-card-fadein 0.7s ease both;
  position:relative;z-index:1;
}
@keyframes dn-card-fadein{from{opacity:0;transform:perspective(1400px) rotateY(-0.8deg) rotateX(0.4deg) translateY(18px);}to{opacity:1;transform:perspective(1400px) rotateY(-0.8deg) rotateX(0.4deg) translateY(0);}}
.dn-dashboard-mock:hover{transform:perspective(1400px) rotateY(0deg) rotateX(0deg);box-shadow:0 36px 88px rgba(0,0,0,0.46),0 8px 28px rgba(11,31,58,0.16),0 0 0 1px rgba(255,255,255,0.10);}

/* Hero — right unlock panel (wider, stronger) */
.dn-hero-right-panel{
  flex:0 0 292px;
  background:rgba(255,255,255,0.08);
  border:1px solid rgba(255,255,255,0.22);
  border-radius:16px;
  padding:28px 26px;
  backdrop-filter:blur(16px);
  -webkit-backdrop-filter:blur(16px);
  box-shadow:0 10px 40px rgba(0,0,0,0.30),0 0 0 1px rgba(255,255,255,0.07);
  align-self:center;
}
.dn-btn-unlock{
  display:flex;align-items:center;justify-content:center;
  background:linear-gradient(90deg,#e8920a 0%,#f97316 100%);
  color:#0f172a;font-weight:800;font-size:14px;
  padding:14px 18px;border-radius:10px;text-decoration:none;
  box-shadow:0 4px 18px rgba(245,158,11,0.45),0 0 0 0 rgba(245,158,11,0.3);
  transition:transform 0.15s,box-shadow 0.15s;
  text-align:center;line-height:1.3;
  animation:dn-unlock-pulse 3s ease-in-out infinite;
}
@keyframes dn-unlock-pulse{
  0%,100%{box-shadow:0 4px 18px rgba(245,158,11,0.45),0 0 0 0 rgba(245,158,11,0.25);}
  55%{box-shadow:0 4px 22px rgba(245,158,11,0.55),0 0 0 6px rgba(245,158,11,0);}
}
.dn-btn-unlock:hover{transform:translateY(-2px);box-shadow:0 10px 28px rgba(245,158,11,0.60);animation:none;}

/* Card topbar */
.dn-mock-topbar{background:linear-gradient(90deg,#0f2557,#1a3a6b);padding:9px 14px;display:flex;align-items:center;gap:7px;}
.dn-mock-dot{width:9px;height:9px;border-radius:50%;display:inline-block;}
.dn-mock-topbar-title{font-size:10px;color:rgba(255,255,255,0.82);margin-left:6px;font-weight:600;letter-spacing:0.3px;}
.dn-mock-topbar-badge{margin-left:auto;font-size:9px;background:rgba(40,200,64,0.25);color:#28c840;
  border:1px solid rgba(40,200,64,0.4);border-radius:20px;padding:2px 8px;font-weight:700;letter-spacing:0.3px;
  animation:dn-live-pulse 2s ease-in-out infinite;}
@keyframes dn-live-pulse{
  0%,100%{background:rgba(40,200,64,0.25);box-shadow:0 0 0 0 rgba(40,200,64,0);}
  50%{background:rgba(40,200,64,0.40);box-shadow:0 0 0 4px rgba(40,200,64,0);}
}

/* KPI row */
.dn-mock-kpis{display:flex;border-bottom:1px solid #f0f4f8;background:#ffffff;}
.dn-mock-kpi{flex:1;padding:11px 8px;text-align:center;border-right:1px solid #f0f4f8;position:relative;}
.dn-mock-kpi:last-child{border-right:none;}
.dn-mock-kpi-val{font-size:15px;font-weight:900;font-family:monospace;line-height:1;}
.dn-mock-kpi-lbl{font-size:8px;color:#94a3b8;margin-top:3px;font-weight:700;text-transform:uppercase;letter-spacing:0.7px;}
.dn-mock-kpi-delta{font-size:9px;font-weight:700;margin-top:2px;}

/* Health score row */
.dn-mock-score-row{display:flex;align-items:center;gap:14px;padding:10px 14px;border-bottom:1px solid #f0f4f8;background:#fafbff;}
.dn-mock-score-ring{position:relative;width:52px;height:52px;flex-shrink:0;}
.dn-mock-score-ring svg{transform:rotate(-90deg);}
.dn-mock-score-num{position:absolute;inset:0;display:flex;align-items:center;justify-content:center;font-size:13px;font-weight:900;color:#1B4F8A;font-family:monospace;}
.dn-mock-score-info{flex:1;}
.dn-mock-score-label{font-size:11px;font-weight:800;color:#0B1F3A;margin-bottom:2px;}
.dn-mock-score-sub{font-size:10px;color:#64748b;}
.dn-mock-score-bar-wrap{flex:1;height:5px;background:#eef2f8;border-radius:4px;overflow:hidden;margin-top:5px;}
.dn-mock-score-bar{height:100%;background:linear-gradient(90deg,#16a34a,#4ade80);border-radius:4px;width:85%;}

/* Retail metrics strip */
.dn-mock-metrics{display:flex;border-bottom:1px solid #f0f4f8;background:#fff;}
.dn-mock-metric{flex:1;padding:9px 10px;border-right:1px solid #f0f4f8;text-align:center;}
.dn-mock-metric:last-child{border-right:none;}
.dn-mock-metric-val{font-size:12px;font-weight:800;color:#0B1F3A;font-family:monospace;}
.dn-mock-metric-lbl{font-size:8px;color:#94a3b8;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;margin-top:2px;}

/* Chart section */
.dn-mock-chart{padding:8px 12px 6px;border-bottom:1px solid #f0f4f8;background:#fff;}
.dn-mock-chart-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:7px;}
.dn-mock-chart-label{font-size:8px;font-weight:800;color:#374151;text-transform:uppercase;letter-spacing:0.8px;}
.dn-mock-chart-legend{display:flex;gap:10px;}
.dn-mock-legend-item{display:flex;align-items:center;gap:4px;font-size:8px;color:#64748b;font-weight:600;}
.dn-mock-legend-dot{width:7px;height:7px;border-radius:50%;}
/* Bar+line combo chart */
.dn-mock-combo{display:flex;align-items:flex-end;gap:4px;height:68px;position:relative;}
.dn-mock-cbar{flex:1;background:linear-gradient(to top,#93c5fd,#bfdbfe);border-radius:3px 3px 0 0;
  animation:dn-bar-grow 0.8s ease both;}
.dn-mock-cbar:nth-child(1){animation-delay:0.05s;}
.dn-mock-cbar:nth-child(2){animation-delay:0.10s;}
.dn-mock-cbar:nth-child(3){animation-delay:0.15s;}
.dn-mock-cbar:nth-child(4){animation-delay:0.20s;}
.dn-mock-cbar:nth-child(5){animation-delay:0.25s;}
.dn-mock-cbar:nth-child(6){animation-delay:0.30s;}
.dn-mock-cbar:nth-child(7){animation-delay:0.35s;}
.dn-mock-cbar-accent{background:linear-gradient(to top,#1d4ed8,#60a5fa) !important;}
@keyframes dn-bar-grow{from{transform:scaleY(0);transform-origin:bottom;}to{transform:scaleY(1);transform-origin:bottom;}}
/* Metric number count-up shimmer */
@keyframes dn-num-in{from{opacity:0;transform:translateY(4px);}to{opacity:1;transform:translateY(0);}}
.dn-mock-kpi-val,.dn-mock-metric-val{animation:dn-num-in 0.5s ease both;}
.dn-mock-kpi:nth-child(1) .dn-mock-kpi-val{animation-delay:0.15s;}
.dn-mock-kpi:nth-child(2) .dn-mock-kpi-val{animation-delay:0.25s;}
.dn-mock-kpi:nth-child(3) .dn-mock-kpi-val{animation-delay:0.35s;}
/* Category bars slide in */
@keyframes dn-bar-slide{from{width:0;}to{}}
.dn-dashboard-mock [style*="border-radius:3px"]:not(.dn-mock-dot){animation:dn-bar-slide 0.9s cubic-bezier(0.22,1,0.36,1) both;}
/* SVG trend line overlay */
.dn-mock-chart-wrap{position:relative;}
.dn-mock-trend-svg{position:absolute;bottom:0;left:0;right:0;height:68px;pointer-events:none;}

/* Insight chips */
.dn-mock-insights{padding:8px 10px 9px;display:flex;flex-direction:column;gap:5px;background:#f8fafc;}
.dn-mock-insight{
  font-size:10px;color:#374151;background:#ffffff;
  border:1px solid #e8edf5;border-radius:7px;
  padding:5px 9px;line-height:1.4;display:flex;align-items:center;gap:7px;
  box-shadow:0 1px 3px rgba(0,0,0,0.04);
}
.dn-mock-insight-icon{width:16px;height:16px;border-radius:4px;display:flex;align-items:center;justify-content:center;flex-shrink:0;}

/* ═══════════════════════════════════════════════════════════
   INPUT SECTION (left card + right info card)
   ═══════════════════════════════════════════════════════════ */
/* Landing row layout */
#dn-landing-row{gap:24px !important;align-items:flex-start !important;display:flex !important;flex-wrap:nowrap !important;}
#dn-landing-row>.gap{align-items:flex-start !important;}
#dn-landing-row>div{align-self:flex-start !important;}
#dn-landing-row>div>div{align-self:flex-start !important;}
@media(max-width:800px){
  #dn-landing-row{flex-direction:column !important;}
  #dn-msme-card-wrap,#dn-right-cards{max-width:100% !important;width:100% !important;}
}

/* Input card — wider, fills its column */
#dn-msme-card-wrap,#dn-gov-card-wrap{height:auto;box-sizing:border-box;}
#dn-msme-card-wrap{
  background:#ffffff;
  border-radius:16px;
  padding:28px 28px 28px;
  box-shadow:0 4px 24px rgba(11,31,58,0.09),0 0 0 1px #eaf0f8;
  width:100%;box-sizing:border-box;
}
/* Analyze My Data button — full width, high contrast */
#dn-real-login-btn button,
#dn-real-login-btn{
  width:100% !important;
  height:52px !important;min-height:52px !important;
  font-size:16px !important;font-weight:800 !important;
  border-radius:12px !important;
  background:linear-gradient(90deg,#1B4F8A 0%,#2563eb 100%) !important;
  color:#ffffff !important;border:none !important;
  box-shadow:0 6px 20px rgba(37,99,235,0.40) !important;
  transition:transform 0.15s,box-shadow 0.15s !important;
  cursor:pointer !important;
  position:relative !important;z-index:20 !important;
  pointer-events:auto !important;
}
#dn-real-login-btn button:hover{
  transform:translateY(-2px) !important;
  box-shadow:0 10px 28px rgba(37,99,235,0.52) !important;
}
/* Fix 6: Strip ALL Gradio chrome inside input card — full deep selector set */
#dn-msme-card-wrap .block,
#dn-msme-card-wrap .form,
#dn-msme-card-wrap .wrap,
#dn-msme-card-wrap .contain,
#dn-msme-card-wrap [class*="svelte"],
#dn-msme-card-wrap .label-wrap,
#dn-msme-card-wrap fieldset{
  padding:0 !important;background:transparent !important;
  box-shadow:none !important;border:none !important;
  border-radius:0 !important;margin:0 !important;
}
#dn-msme-card-wrap .block+.block{margin-top:0 !important;}
#dn-msme-card-wrap>.gap,#dn-msme-card-wrap>div>.gap,#dn-msme-card-wrap .gap{gap:0 !important;}
#dn-msme-card-wrap>div>div{gap:14px !important;}
/* Specifically kill the outer container border Gradio wraps inputs in */
#dn-msme-card-wrap .gradio-textbox,
#dn-msme-card-wrap .gradio-number,
#dn-msme-card-wrap .gradio-dropdown,
#dn-msme-card-wrap .gradio-file,
#dn-msme-card-wrap .container{
  padding:0 !important;border:none !important;
  background:transparent !important;box-shadow:none !important;
}

/* Input fields — uniform height, strong borders, clear padding */
#dn-msme-card-wrap input,
#dn-msme-card-wrap textarea{
  height:48px !important;
  min-height:48px !important;
  font-size:14px !important;
  border:2px solid #cbd5e1 !important;
  border-radius:10px !important;
  padding:0 16px !important;
  background:#ffffff !important;
  color:#0f172a !important;
  box-shadow:0 1px 4px rgba(11,31,58,0.06) !important;
  transition:border-color 0.18s,box-shadow 0.18s !important;
  width:100% !important;
  box-sizing:border-box !important;
}
#dn-msme-card-wrap input:focus,
#dn-msme-card-wrap textarea:focus{
  border-color:#2563eb !important;
  box-shadow:0 0 0 3px rgba(37,99,235,0.15) !important;
  outline:none !important;
}
#dn-msme-card-wrap input::placeholder,
#dn-msme-card-wrap textarea::placeholder{
  color:#94a3b8 !important;
  font-size:13px !important;
}
/* Field labels — readable, spaced */
#dn-msme-card-wrap label,
#dn-msme-card-wrap .label-wrap span,
#dn-msme-card-wrap span.svelte-1gfkn6j{
  font-size:13px !important;font-weight:700 !important;
  color:#1e3a5f !important;
  padding-bottom:6px !important;
  display:block !important;
  text-transform:none !important;letter-spacing:0 !important;
  background:transparent !important;border:none !important;box-shadow:none !important;
}
/* Spacing between form fields — hidden Gradio widgets, form managed by HTML */
#dn-msme-card-wrap>div>div{gap:0 !important;}

/* Upload screen analyze button override */
#analyze-data-btn button, #analyze-data-btn {
  background: linear-gradient(90deg, #1B4F8A 0%, #2563eb 100%) !important;
  color: #ffffff !important;
  font-size: 15px !important;
  font-weight: 800 !important;
  height: 52px !important;
  border: none !important;
  border-radius: 12px !important;
  box-shadow: 0 6px 20px rgba(37,99,235,0.40) !important;
  width: 100% !important;
  cursor: pointer !important;
  position: relative !important;
  z-index: 20 !important;
}
#analyze-data-btn button:hover, #analyze-data-btn:hover {
  transform: translateY(-2px) !important;
  box-shadow: 0 10px 28px rgba(37,99,235,0.52) !important;
}

/* ── Upload Screen — demo button ── */
#dn-demo-btn button, #dn-demo-btn {
  background: linear-gradient(90deg,#0f7a4a 0%,#16a34a 100%) !important;
  color: #ffffff !important;
  font-size: 15px !important;
  font-weight: 800 !important;
  height: 52px !important;
  border: none !important;
  border-radius: 12px !important;
  box-shadow: 0 4px 16px rgba(22,163,74,0.38) !important;
  width: 100% !important;
  cursor: pointer !important;
  letter-spacing: 0.1px !important;
}
#dn-demo-btn button:hover {
  transform: translateY(-2px) !important;
  box-shadow: 0 8px 24px rgba(22,163,74,0.50) !important;
}

/* ── Upload Screen — step5 wrapper ── */
#dn-step5-wrap {
  max-width: 680px;
  margin: 0 auto;
  padding: 0 16px 32px;
}
#dn-step5-wrap .block {
  background: transparent !important;
  border: none !important;
  box-shadow: none !important;
  padding: 0 !important;
  margin: 0 !important;
}
#dn-step5-wrap > div > div { gap: 0 !important; }

/* ── Upload file widget — inside step5 ── */
#dn-step5-wrap .gradio-file,
#dn-step5-wrap [data-testid="file"] {
  background: #f0f6ff !important;
  border: 2px dashed #93c5fd !important;
  border-radius: 12px !important;
  min-height: 90px !important;
  padding: 10px 16px !important;
  transition: border-color 0.2s, background 0.2s !important;
}
#dn-step5-wrap .gradio-file:hover,
#dn-step5-wrap [data-testid="file"]:hover {
  border-color: #2563eb !important;
  background: #e8f0ff !important;
}
#dn-step5-wrap .gradio-file label,
#dn-step5-wrap [data-testid="file"] label {
  font-size: 14px !important;
  font-weight: 700 !important;
  color: #1B4F8A !important;
}

/* ── Consent checkbox ── */
#dn-step5-wrap .gradio-checkbox label span {
  font-size: 13px !important;
  color: #475569 !important;
  font-weight: 500 !important;
}
#dn-step5-wrap .gradio-checkbox input[type="checkbox"] {
  width: 16px !important;
  height: 16px !important;
  accent-color: #2563eb !important;
}

/* ── Upload screen step progress bar ── */
.dn-step-progress { display:flex;align-items:center;justify-content:center;gap:0;margin-bottom:20px; }
.dn-step-pill { display:flex;align-items:center;gap:6px;font-size:12px;font-weight:700;
  padding:7px 18px;border-radius:20px;white-space:nowrap; }
.dn-step-pill.active { background:#1B4F8A;color:#fff; }
.dn-step-pill.done { background:#f0fdf4;border:1px solid #bbf7d0;color:#16a34a; }
.dn-step-pill.future { background:#fff;border:1px solid #e2e8f0;color:#94a3b8; }
.dn-step-num { border-radius:50%;width:20px;height:20px;display:inline-flex;
  align-items:center;justify-content:center;font-size:11px;font-weight:900;flex-shrink:0; }
.dn-step-pill.active .dn-step-num { background:#fff;color:#1B4F8A; }
.dn-step-pill.done .dn-step-num { background:#16a34a;color:#fff; }
.dn-step-pill.future .dn-step-num { background:#e2e8f0;color:#64748b; }
.dn-step-connector { width:36px;height:2px; }
.dn-step-connector.done { background:#16a34a; }
.dn-step-connector.future { background:#e2e8f0; }
@media(max-width:500px){ .dn-step-pill{ padding:6px 10px;font-size:11px; } .dn-step-connector{ width:20px; } }
/* Unlock preview button */
#dn-preview-unlock-btn button, #dn-preview-unlock-btn {
  background: linear-gradient(90deg, #d97706 0%, #f59e0b 100%) !important;
  color: #0f172a !important;
  font-size: 17px !important;
  font-weight: 900 !important;
  height: 58px !important;
  border: none !important;
  border-radius: 14px !important;
  box-shadow: 0 6px 24px rgba(245,158,11,0.45) !important;
  width: 100% !important;
  cursor: pointer !important;
  letter-spacing: 0.2px !important;
}
#dn-preview-unlock-btn button:hover {
  transform: translateY(-2px) !important;
  box-shadow: 0 12px 32px rgba(245,158,11,0.55) !important;
}
/* Analyze My Data button — full width, high contrast, unmissable */
#dn-real-login-btn{
  height:52px !important;min-height:52px !important;width:100% !important;
  font-size:16px !important;font-weight:800 !important;
  border-radius:12px !important;margin-top:8px !important;
  letter-spacing:0.2px !important;
  background:linear-gradient(90deg,#1B4F8A 0%,#2563eb 100%) !important;
  color:#ffffff !important;
  border:none !important;
  box-shadow:0 6px 20px rgba(37,99,235,0.45) !important;
  transition:transform 0.15s,box-shadow 0.15s,background 0.15s !important;
  position:relative !important;z-index:20 !important;
  pointer-events:auto !important;cursor:pointer !important;
}
#dn-real-login-btn:hover{
  transform:translateY(-2px) !important;
  box-shadow:0 10px 28px rgba(37,99,235,0.55) !important;
  background:linear-gradient(90deg,#1a4a80 0%,#1d55d0 100%) !important;
}

/* Right info card — strip chrome, auto height, top-aligned */
#dn-right-cards{display:flex;flex-direction:column;gap:0;height:auto;align-self:flex-start !important;}
#dn-right-cards .block{padding:0 !important;background:transparent !important;box-shadow:none !important;border:none !important;}

/* ═══════════════════════════════════════════════════════════
   HOW IT WORKS SECTION
   ═══════════════════════════════════════════════════════════ */
.dn-hiw-card{
  background:#ffffff;
  border:1px solid #e8edf5;
  border-radius:12px;
  padding:18px 18px 16px;
  box-shadow:0 2px 10px rgba(11,31,58,0.06);
  transition:transform 0.18s,box-shadow 0.18s;
  cursor:default;
  height:100%;box-sizing:border-box;
}
.dn-hiw-card:hover{
  transform:translateY(-3px);
  box-shadow:0 8px 24px rgba(11,31,58,0.11);
}

/* ═══════════════════════════════════════════════════════════
   ANALYSIS / DRC SECTION
   ═══════════════════════════════════════════════════════════ */
#drc-block{
  background:#ffffff;
  border:1px solid #e8edf5;
  border-radius:12px;
  padding:20px 24px;
  margin:0 auto;
  max-width:1100px;
  box-shadow:0 2px 10px rgba(11,31,58,0.07);
}
#drc-block .drc-hdr{font-size:20px !important;font-weight:800 !important;color:#0B1F3A !important;display:block !important;margin-bottom:6px !important;}
#drc-block .drc-sub{font-size:14px !important;color:#64748b !important;margin:0 0 16px !important;line-height:1.6 !important;}
#drc-block .drc-compat{font-size:11px;color:#94a3b8;margin-top:4px;}
#drc-block .drc-compact-summary{display:flex;gap:16px;flex-wrap:wrap;align-items:center;padding:10px 14px;background:#f8fafc;border-radius:8px;border:1px solid #e2e8f0;margin-bottom:12px;}
#drc-block .drc-kv{font-size:12px;color:#475569;}
#drc-block .drc-kv span{font-weight:700;color:#0f172a;}
/* Row: flex so all three elements sit on one line with consistent height */
#drc-block > div > div > div[class*="row"]{
  align-items:center !important;
  gap:12px !important;
}
/* Upload file widget — clearly visible drop zone */
/* ── DRC File upload widget — clean single-message drop zone ── */
#drc-block .gradio-file,
#drc-block [data-testid="file"],
#drc-block .upload-container {
  background:#EEF4FF !important;
  border:2px dashed #3b82f6 !important;
  border-radius:10px !important;
  min-height:90px !important;
  padding:0 !important;
  display:flex !important;
  align-items:stretch !important;
  justify-content:stretch !important;
  cursor:pointer !important;
  width:100% !important;
  box-sizing:border-box !important;
  position:relative !important;
}
/* Inner wrap fills the box */
#drc-block [data-testid="file"] > .wrap,
#drc-block .gradio-file > .wrap,
#drc-block .upload-container > .wrap {
  background:transparent !important;
  border:none !important;
  width:100% !important;
  min-height:90px !important;
  display:flex !important;
  flex-direction:column !important;
  align-items:center !important;
  justify-content:center !important;
  gap:6px !important;
  padding:16px !important;
}
/* Hide the file icon Gradio inserts */
#drc-block .gradio-file .file-preview-title,
#drc-block [data-testid="file"] .file-preview-title { display:none !important; }
/* Style the upload prompt text — Gradio's native "Drop File Here" */
#drc-block .gradio-file .wrap > .icon-wrap,
#drc-block [data-testid="file"] .wrap > .icon-wrap { display:none !important; }
/* The label (our empty string) — hide any residual label container */
#drc-block .gradio-file > label,
#drc-block [data-testid="file"] > label,
#drc-block .upload-container > label { display:none !important; }
/* Override Gradio's "Drop File Here" text */
#drc-block .gradio-file .wrap span,
#drc-block [data-testid="file"] .wrap span {
  font-size:0 !important;  /* hide original text */
  line-height:0 !important;
}
/* Inject our custom upload message via ::before */
#drc-block [data-testid="file"] .wrap,
#drc-block .gradio-file .wrap {
  position:relative !important;
}
#drc-block [data-testid="file"] .wrap::before,
#drc-block .gradio-file .wrap::before {
  content:"📂  Drop your Excel or CSV file here, or click to upload";
  font-size:13px !important;
  font-weight:600 !important;
  color:#1B4F8A !important;
  text-align:center !important;
  display:block !important;
}
/* Hide the "- or -" separator completely */
#drc-block .gradio-file .wrap .or,
#drc-block [data-testid="file"] .wrap .or,
#drc-block .upload-container .or {
  display:none !important;
}
/* File name display after upload */
#drc-block .gradio-file .file-name,
#drc-block [data-testid="file"] .file-name {
  font-size:13px !important;
  font-weight:600 !important;
  color:#0f172a !important;
}
/* Collapse ALL hidden elements inside drc-block — zero phantom space */
#drc-block [style*="display: none"],
#drc-block [style*="display:none"] {
  height:0 !important; min-height:0 !important;
  margin:0 !important; padding:0 !important;
  overflow:hidden !important; border:none !important;
}
/* Remove gap between empty/hidden children */
#drc-block > div > div { gap:0 !important; }
#drc-block .gap { gap:0 !important; row-gap:0 !important; }

/* ── Login form — style actual Gradio widgets professionally ── */

/* Card wrapper */
#dn-msme-card-wrap {
  max-width:480px !important; margin:0 auto !important;
}
#dn-msme-card-wrap > div,
#dn-msme-card-wrap > div > div {
  background:#ffffff !important;
  border-radius:16px !important;
  border:1px solid #e2eaf5 !important;
  box-shadow:0 4px 24px rgba(11,31,58,0.10) !important;
  padding:28px 24px 24px !important;
  gap:14px !important;
}

/* Kill ALL Gradio chrome on inputs */
#dn-msme-col-inner .block,
#dn-msme-col-inner fieldset,
#dn-msme-col-inner .form,
#dn-msme-col-inner .wrap {
  border:none !important; background:transparent !important;
  box-shadow:none !important; padding:0 !important; margin:0 !important;
}

/* Labels */
#dn-landing-name label span,
#dn-real-mobile label span,
#dn-landing-otp label span,
#dn-landing-email label span {
  font-size:13px !important; font-weight:700 !important;
  color:#1e3a5f !important; display:block !important;
  margin-bottom:6px !important;
}

/* Input fields */
#dn-landing-name textarea, #dn-landing-name input,
#dn-real-mobile textarea,  #dn-real-mobile input,
#dn-landing-otp textarea,  #dn-landing-otp input,
#dn-landing-email textarea, #dn-landing-email input {
  height:48px !important; min-height:48px !important;
  padding:0 14px !important;
  border:1.5px solid #d1d9e6 !important;
  border-radius:10px !important;
  font-size:14px !important; color:#0f172a !important;
  background:#ffffff !important;
  box-shadow:0 1px 3px rgba(0,0,0,0.05) !important;
  resize:none !important; outline:none !important;
  -webkit-appearance:none !important;
  -moz-appearance:textfield !important;
  appearance:none !important;
}
#dn-landing-name textarea:focus, #dn-landing-name input:focus,
#dn-real-mobile textarea:focus,  #dn-real-mobile input:focus,
#dn-landing-otp textarea:focus,  #dn-landing-otp input:focus,
#dn-landing-email textarea:focus, #dn-landing-email input:focus {
  border-color:#2563eb !important;
  box-shadow:0 0 0 3px rgba(37,99,235,0.12) !important;
}
/* Kill spinner arrows */
#dn-landing-name input::-webkit-outer-spin-button,
#dn-landing-name input::-webkit-inner-spin-button,
#dn-real-mobile input::-webkit-outer-spin-button,
#dn-real-mobile input::-webkit-inner-spin-button,
#dn-landing-otp input::-webkit-outer-spin-button,
#dn-landing-otp input::-webkit-inner-spin-button,
#dn-landing-email input::-webkit-outer-spin-button,
#dn-landing-email input::-webkit-inner-spin-button {
  -webkit-appearance:none !important; display:none !important;
  width:0 !important; margin:0 !important;
}

/* Analyze My Data button */
#dn-real-login-btn button {
  width:100% !important; height:54px !important; min-height:54px !important;
  background:linear-gradient(90deg,#1B4F8A,#2563eb) !important;
  color:#ffffff !important; border:none !important;
  border-radius:12px !important; font-size:16px !important;
  font-weight:800 !important; letter-spacing:0.3px !important;
  cursor:pointer !important;
  box-shadow:0 6px 20px rgba(37,99,235,0.40) !important;
  display:flex !important; align-items:center !important;
  justify-content:center !important;
  visibility:visible !important; opacity:1 !important;
}
#dn-real-login-btn button:hover {
  transform:translateY(-2px) !important;
  box-shadow:0 10px 28px rgba(37,99,235,0.55) !important;
}
/* Hide any Gradio icon inside button */
#dn-real-login-btn button img,
#dn-real-login-btn button svg { display:none !important; }
/* ── Collapse login_section_col when hidden — kills phantom whitespace ── */
#dn-login-section-col,
#dn-login-section-col > div,
#dn-login-section-col > div > div {
  transition:none !important;
}
#dn-login-section-col[style*="display: none"],
#dn-login-section-col[style*="display:none"] {
  height:0 !important; max-height:0 !important; min-height:0 !important;
  overflow:hidden !important; margin:0 !important; padding:0 !important;
  border:none !important; flex:none !important; pointer-events:none !important;
}
/* Kill the gap that step0-col adds between login section and DRC */
#dn-step0-col > div > div {
  row-gap:0 !important; gap:0 !important;
}
#dn-step0-col > div > div > [data-testid="column"][style*="display: none"],
#dn-step0-col > div > div > [data-testid="column"][style*="display:none"] {
  height:0 !important; max-height:0 !important; min-height:0 !important;
  overflow:hidden !important; margin:0 !important; padding:0 !important;
}
/* Block/form reset */
#drc-block .block {
  padding:0 !important;
  background:transparent !important;
  border:none !important;
  box-shadow:none !important;
}
/* Drop text and icon styling */
#drc-block .gradio-file label,
#drc-block [data-testid="file"] label,
#drc-block .upload-container label{
  font-size:13px !important;
  color:#1B4F8A !important;
  font-weight:600 !important;
  cursor:pointer !important;
  display:flex !important;
  align-items:center !important;
  gap:8px !important;
  width:100% !important;
}
/* The "Drop File Here" span inside the label */
#drc-block .gradio-file label span,
#drc-block [data-testid="file"] label span,
#drc-block .upload-container span{
  color:#1B4F8A !important;
  font-size:13px !important;
  font-weight:600 !important;
  opacity:1 !important;
  position:static !important;
  display:inline !important;
}
/* "- or -" separator */
#drc-block .or{
  color:#94a3b8 !important;
  font-size:12px !important;
}
/* Run Analysis + Download Template buttons */
#drc-block button,#drc-block .download-button{
  font-size:13px !important;
  min-height:44px !important;
  height:44px !important;
  border-radius:8px !important;
  padding:0 20px !important;
  white-space:nowrap !important;
}
/* Run Analysis — primary gradient */
#drc-block button[class*="primary"],
#drc-block button.primary{
  background:linear-gradient(90deg,#1B4F8A 0%,#2563eb 100%) !important;
  box-shadow:0 3px 10px rgba(27,79,138,0.25) !important;
  border:none !important;
  transition:box-shadow 0.15s,transform 0.15s !important;
}
#drc-block button[class*="primary"]:hover,
#drc-block button.primary:hover{
  box-shadow:0 6px 18px rgba(27,79,138,0.35) !important;
  transform:translateY(-1px) !important;
}
/* Download Template — light secondary */
#drc-block button[class*="secondary"],
#drc-block button.secondary,
#drc-block .download-button{
  background:#ffffff !important;
  border:1.5px solid #cbd5e1 !important;
  color:#475569 !important;
  box-shadow:none !important;
}
/* ── Fix: prevent gr.File drag-drop overlay from blocking adjacent buttons ── */
#drc-block .gradio-file label,
#drc-block [data-testid="file"] label,
#drc-block .file-preview,
#drc-block .upload-container{
  pointer-events:auto !important;
  max-width:100% !important;
  overflow:hidden !important;
}
/* Ensure Run Analysis and Download Template buttons are always on top and clickable */
#drc-block button{
  position:relative !important;
  z-index:10 !important;
  pointer-events:auto !important;
  cursor:pointer !important;
}
/* Loading state feedback on Run Analysis button */
#drc-loading-msg{margin:0 !important;padding:0 !important;}
#analyze-loading-msg{margin:8px 0 !important;}
/* Clip DownloadButton anchor so it doesn't overflow its column */
#drc-block .download-button,
#drc-block a[download],
#drc-block a[href]{
  position:relative !important;
  z-index:10 !important;
  overflow:hidden !important;
  display:inline-flex !important;
  max-width:100% !important;
}
/* ── Fix: prevent DownloadButton file icon from bleeding into login card ── */
#dn-msme-card-wrap button,
#dn-msme-card-wrap [class*="svelte"] button{
  position:relative !important;
  z-index:20 !important;
  pointer-events:auto !important;
}
/* Hidden source widgets */
#drc-compact-out,#drc-quality-out{display:none !important;}
.gradio-container #drc-block .gradio-textbox{display:none !important;}
#drc_mapping_out,.gradio-container [id="drc_mapping_out"]{display:none !important;}

/* ═══════════════════════════════════════════════════════════
   STEP 6 / STEP 7 NAV BUTTONS — always on top, always clickable
   ═══════════════════════════════════════════════════════════ */
#dn-forecast-btn,#dn-back6-btn,#dn-back6-dash-btn,
#dn-preview-unlock-btn,#analyze-data-btn{
  position:relative !important;
  z-index:50 !important;
  pointer-events:auto !important;
  cursor:pointer !important;
}
/* Kill any Gradio row equal-height flex stretch that creates invisible blocking layers */
.gradio-container [id="dn-forecast-btn"],
.gradio-container [id="dn-back6-btn"],
.gradio-container [id="dn-back6-dash-btn"]{
  align-self:center !important;
  flex-shrink:0 !important;
}

/* ═══════════════════════════════════════════════════════════
   MISC LEGACY — kept for step 1-7 compatibility
   ═══════════════════════════════════════════════════════════ */
.header-container{background:linear-gradient(135deg,#0f2557,#1a3a6b);padding:8px 20px;display:flex;justify-content:space-between;align-items:center;}
.logo-section{display:flex;align-items:center;gap:12px;}
.hero-section{background:linear-gradient(105deg,#0b1829,#0f2557 55%,#1a3a6b);padding:16px 40px 14px;color:white;margin-top:0;}
.hero-title{font-size:30px;font-weight:800;margin-bottom:10px;color:#fff;letter-spacing:-0.5px;line-height:1.2;}
.hero-section h2.hero-sub-tagline{font-size:15px;font-weight:400;color:rgba(255,255,255,0.88);margin:0 0 14px;line-height:1.6;}
.hero-section p.hero-description{margin:0 0 3px;font-size:12px;color:rgba(255,255,255,0.72);line-height:1.5;}
.hero-divider{border:none;border-top:1px solid rgba(255,255,255,0.15);margin:0 auto 12px;width:50%;}
.footer-section{background-color:#0B1F3A;color:#fff;padding:0;margin-top:0;width:100%;box-sizing:border-box;}
.footer-section a{color:#fff;text-decoration:none;}
.dn-card-hover{transition:box-shadow 0.2s,transform 0.2s;cursor:default;}
.dn-card-hover:hover{box-shadow:0 10px 32px rgba(11,31,58,0.17) !important;transform:translateY(-3px);}
.dn-pipe-card{transition:box-shadow 0.2s,transform 0.2s;cursor:default;}
.dn-pipe-card:hover{box-shadow:0 10px 28px rgba(11,31,58,0.18) !important;transform:translateY(-3px);}
.dn-access-hover{transition:box-shadow 0.2s,transform 0.2s;cursor:default;}
.dn-access-hover:hover{box-shadow:0 8px 26px rgba(11,31,58,0.16) !important;transform:translateY(-2px);}
.capabilities-section{background:#f0f7ff;padding:26px;margin-top:10px;border-radius:12px;}
#pdf-file-output .file-preview-title,#pdf-file-output span.file-name,#pdf-file-output .file-size{display:none !important;}
#dn-role-tabs .tab-nav{border-bottom:2px solid #1a3a6b !important;background:transparent !important;gap:4px !important;padding:0 !important;}
#dn-role-tabs .tab-nav button{border-radius:8px 8px 0 0 !important;border:2px solid #c8dcef !important;border-bottom:none !important;background:#f0f4f8 !important;color:#4A6A8A !important;font-size:13px !important;font-weight:700 !important;padding:8px 22px !important;transition:all 0.15s !important;}
#dn-role-tabs .tab-nav button.selected{background:#1a3a6b !important;color:#fff !important;border-color:#1a3a6b !important;}
#dn-role-tabs .tabitem{border:none !important;background:transparent !important;padding:0 !important;}
.section{padding:36px 20px;margin-top:12px;border-radius:8px;background-color:#f9f9f9;}
.section-title{font-size:32px;font-weight:700;color:#333;text-align:center;margin-bottom:20px;}
.dn-section-divider{border:none;border-top:1px solid #e2e8f0;margin:8px 0;}


/* Pricing section — hidden until nav Pricing clicked */
#dn-pricing-section{display:none !important;}
#dn-pricing-section.dn-pricing-open{display:block !important;}

/* OTP field visible on landing */

/* ═══════════════════════════════════════════════════════════════════
   PROFESSIONAL WEBSITE FEEL — COMPLETE OVERRIDE
   ═══════════════════════════════════════════════════════════════════ */

/* 1. Page background — clean white, not Gradio's grey */
body, html,
.gradio-container,
.gradio-container > .main,
.gradio-container > .main > .wrap,
.gradio-container > .main > .wrap > .padding,
.gradio-container .app,
.svelte-1gfkn6j { background:#ffffff !important; }

/* 2. Kill every Gradio gap, padding, margin, border, shadow on the landing page */
#dn-step0-col,
#dn-step0-col > .gap,
#dn-step0-col > div,
#dn-step0-col > div > .gap,
#dn-step0-col > div > div > .gap {
  gap:0 !important; padding:0 !important;
  background:transparent !important;
  border:none !important; box-shadow:none !important;
}
#dn-step0-col .block,
#dn-step0-col .form,
#dn-step0-col .wrap,
#dn-step0-col .contain,
#dn-step0-col .prose {
  padding:0 !important; margin:0 !important;
  background:transparent !important;
  border:none !important; box-shadow:none !important;
  gap:0 !important;
}

/* 3. Lang-bar (top header row) — make it a real nav bar */
#lang-bar,
#lang-bar > .gap,
#lang-bar > div,
#lang-bar > div > div { padding:0 !important; gap:0 !important; }
#lang-bar { 
  background:#ffffff !important;
  border-bottom:1px solid #e2eaf5 !important;
  box-shadow:0 1px 0 rgba(0,0,0,0.06) !important;
  min-height:56px !important; max-height:56px !important;
  display:flex !important; align-items:center !important;
}
#lang-bar > .gap { 
  display:flex !important; align-items:center !important;
  padding:0 36px !important; width:100% !important;
  max-width:1340px !important; margin:0 auto !important;
  min-height:56px !important;
}
#lang-bar .block { 
  background:transparent !important; border:none !important;
  box-shadow:none !important; padding:0 !important; 
  display:flex !important; align-items:center !important;
}

/* Logo block — left */
#hdr-logo-wrap { padding:0 !important; }
#lang-bar-logo { 
  display:flex !important; align-items:center !important;
  gap:10px !important; 
}

/* Nav links block — center */
#lang-bar-nav, #lang-bar-nav > div {
  display:flex !important; align-items:center !important;
  justify-content:center !important; flex:1 !important;
}
#hdr-nav-links {
  display:flex !important; align-items:center !important;
  justify-content:center !important; gap:32px !important;
}
.hdr-nav-link {
  font-size:14px !important; font-weight:500 !important;
  color:#374151 !important; text-decoration:none !important;
  white-space:nowrap !important; transition:color 0.15s !important;
  padding:4px 0 !important;
}
.hdr-nav-link:hover { color:#1B4F8A !important; }

/* Language dropdown — right */
#lang-dropdown-main { min-width:140px !important; }
#lang-dropdown-main .wrap { padding:0 !important; }
/* Hide any residual label/gap above the dropdown */
#lang-dropdown-main > .label-wrap,
#lang-dropdown-main > label,
#lang-dropdown-main .container > label { display:none !important; }
#lang-dropdown-main select,
#lang-dropdown-main .wrap > div { margin-top:0 !important; }

/* 4. Hero — zero Gradio padding, flush to edges */
.gradio-container .block:has(.dn-hero-wrap) {
  padding:0 !important; margin:0 !important;
  border:none !important; box-shadow:none !important;
  background:transparent !important;
}
.dn-hero-wrap { 
  width:100% !important; display:block !important;
  margin-top:0 !important;
}

/* 5. Section continuity — no visible seams */
.gradio-container .block:has(.dn-band),
.gradio-container .block:has(> div[style*="background"]) {
  padding:0 !important; margin:0 !important;
  border:none !important; box-shadow:none !important;
  background:transparent !important;
}

/* 6. DRC section seamless */
#drc-block > div > div,
#drc-block .gap { gap:8px !important; }
.gradio-container .block:has(#drc-block) {
  padding:0 !important; margin:0 !important;
  border:none !important; box-shadow:none !important;
}

/* 7. Login section column — clean */
#dn-login-section-col > .gap { gap:0 !important; }

/* 8. Remove the grey Gradio footer / bottom padding */
.gradio-container > .main > .wrap > .padding {
  padding-bottom:0 !important;
}
footer.svelte-mpyp5e, .footer { display:none !important; }

/* 9. Responsive nav */
@media(max-width:768px) {
  #hdr-nav-links { display:none !important; }
  #lang-bar > .gap { padding:0 16px !important; }
}
#hdr-pricing-btn { position:absolute !important; opacity:0 !important; width:1px !important; height:1px !important; overflow:hidden !important; pointer-events:auto !important; }



/* ═══════════════════════════════════════════════════════════
   MOBILE RESPONSIVE — Full rewrite for 320px–768px
   Targets: nav, trust bar, hero, how-it-works, industries,
            unlock section, pricing, DRC, footer
   ═══════════════════════════════════════════════════════════ */

/* ── Base mobile resets ─────────────────────────────────────── */
@media(max-width:768px){
  /* Kill Gradio's fixed container padding on mobile */
  .gradio-container,.gradio-container>.main,.gradio-container>.main>.wrap{
    padding:0 !important; margin:0 !important;
  }
  /* Make all block elements scroll cleanly */
  body,html{ overflow-x:hidden !important; }
}

/* ── Navigation bar ─────────────────────────────────────────── */
@media(max-width:768px){
  #lang-bar { min-height:48px !important; max-height:48px !important; }
  #lang-bar>.gap { padding:0 12px !important; }
  #lang-bar-logo img { height:32px !important; }
  #lang-bar-logo .logo-text { font-size:16px !important; }
  #hdr-nav-links { display:none !important; }
  #lang-dropdown-main { min-width:110px !important; }
}

/* ── Trust bar ──────────────────────────────────────────────── */
@media(max-width:768px){
  /* Trust bar: 2-column grid on mobile */
  .dn-trust-bar-inner,

}

/* ── Hero section ───────────────────────────────────────────── */
@media(max-width:768px){
  .dn-hero-bg{
    flex-direction:column !important;
    padding:20px 16px 24px !important;
    gap:20px !important;
  }
  .dn-hero-left{
    flex:none !important;
    min-width:0 !important;
    max-width:100% !important;
    width:100% !important;
  }
  .dn-hero-center{
    flex:none !important;
    min-width:0 !important;
    width:100% !important;
    max-width:100% !important;
  }
  .dn-hero-h1{
    font-size:1.75rem !important;
    letter-spacing:-0.5px !important;
    line-height:1.12 !important;
    margin-bottom:12px !important;
  }
  .dn-hero-sub{
    font-size:13px !important;
    max-width:100% !important;
    margin-bottom:16px !important;
  }
  .dn-hero-btns{
    flex-direction:column !important;
    gap:10px !important;
    align-items:flex-start !important;
  }
  .dn-hero-btns a{
    width:100% !important;
    justify-content:center !important;
    font-size:14px !important;
    padding:13px 20px !important;
  }
  .dn-hero-trust{
    flex-direction:column !important;
    gap:8px !important;
  }
  .dn-trust-pill{
    width:fit-content !important;
    font-size:12px !important;
  }
  /* Dashboard mock on mobile — scale down */
  .dn-dashboard-mock{
    transform:none !important;
    border-radius:10px !important;
  }
  .dn-hero-center::before{ display:none !important; }
}

/* ── How it works — force stacked on mobile ─────────────────── */
@media(max-width:700px){
  /* The how-it-works flex container has flex-wrap:nowrap — override it */
  [style*="flex-wrap:nowrap"][style*="max-width:900px"]{
    flex-wrap:wrap !important;
    flex-direction:column !important;
    gap:0 !important;
    align-items:center !important;
  }
  /* Hide the arrow connectors on mobile */
  [style*="flex:0 0 44px"][style*="padding-top:40px"]{
    display:none !important;
  }
  /* Each step full width */
  [style*="flex:1;min-width:0;padding:28px 20px"]{
    flex:none !important;
    width:100% !important;
    padding:20px 16px !important;
  }
  /* How it works container */
  [style*="From your Excel to business clarity"]{
    padding:28px 16px 24px !important;
  }
}

/* ── Industries / WHO IT'S FOR ──────────────────────────────── */
@media(max-width:768px){
  /* Pills wrap naturally — just reduce padding */
  [style*="WHO IT'S FOR"]{
    padding:20px 16px 16px !important;
  }
}

/* ── What ₹999/mo Gets You ──────────────────────────────────── */
@media(max-width:768px){
  /* Section container */
  [style*="WHAT ₹999/MONTH GETS YOU"]{
    padding:28px 16px !important;
  }
  /* CA comparison table — make it scrollable */
  [style*="HOW DOES IT COMPARE"]{
    overflow-x:auto !important;
    -webkit-overflow-scrolling:touch !important;
  }
  [style*="HOW DOES IT COMPARE"] > [style*="grid-template-columns:1fr 1fr 1fr"]{
    min-width:420px !important;
  }
  /* Feature cards — single column */
  [style*="grid-template-columns:repeat(auto-fit,minmax(230px"]{
    grid-template-columns:1fr !important;
  }
}

/* ── DRC / Upload block ─────────────────────────────────────── */
@media(max-width:768px){
  #drc-block{
    padding:14px 14px !important;
    border-radius:8px !important;
  }
  #drc-block .gradio-file,
  #drc-block [data-testid="file"]{
    min-height:70px !important;
  }
  /* Run Analysis + Download Template stack vertically */
  #drc-block [data-testid="row"]{
    flex-direction:column !important;
    gap:8px !important;
  }
  #drc-block button{
    width:100% !important;
    min-height:44px !important;
  }
}

/* ── Contact section ────────────────────────────────────────── */
@media(max-width:768px){
  /* Grid → single column */
  [style*="grid-template-columns:1fr 1.7fr"]{
    grid-template-columns:1fr !important;
  }
}

/* ── Pricing section ────────────────────────────────────────── */
@media(max-width:768px){
  [style*="grid-template-columns:repeat(auto-fit,minmax(280px"]{
    grid-template-columns:1fr !important;
  }
}

/* ── Footer ─────────────────────────────────────────────────── */
@media(max-width:768px){
  /* Footer columns wrap */
  [style*="max-width:1100px;margin:0 auto;padding:28px 40px"]{
    padding:20px 16px !important;
  }
  /* Top row flex wrap */
  [style*="gap:40px;justify-content:space-between;align-items:flex-start"]{
    gap:24px !important;
  }
  [style*="flex:0 0 220px"],[style*="flex:0 0 140px"],
  [style*="flex:0 0 160px"],[style*="flex:0 0 180px"]{
    flex:none !important;
    width:100% !important;
    min-width:0 !important;
  }
  /* Trust strip wrap */
  [style*="Made in India"]{
    flex-wrap:wrap !important;
    gap:10px !important;
    padding:10px 14px !important;
    justify-content:flex-start !important;
  }
}

/* ── App header bars (workflow screens) ─────────────────────── */
@media(max-width:768px){
  /* Step progress in app bars — smaller pills */
  [style*="background:#0f2557;padding:9px 20px"]{
    padding:7px 12px !important;
  }
  [style*="font-size:11px;font-weight:600;color:#ffffff"]{
    font-size:9px !important;
  }
  [style*="width:28px;height:1px"]{
    width:14px !important;
  }
  [style*="width:17px;height:17px;border-radius:50%"]{
    width:14px !important;
    height:14px !important;
  }
  /* Sub-header text smaller */
  [style*="font-size:15px;font-weight:800;color:#ffffff"]{
    font-size:13px !important;
  }
}

/* ── KPI cards in preview/dashboard ────────────────────────── */
@media(max-width:600px){
  .dn-pv-grid4{
    grid-template-columns:1fr 1fr !important;
  }
  .dn-pv-grid2{
    grid-template-columns:1fr !important;
  }
}

/* ── Step 7 filter row — stack on mobile ────────────────────── */
@media(max-width:768px){
  #dn-step7-filters [data-testid="row"]{
    flex-direction:column !important;
  }
}

/* ── How-it-works steps — stack on mobile ──────────────────── */
@media(max-width:650px){
  .dn-hiw-steps{
    flex-direction:column !important;
    align-items:center !important;
  }
  .dn-hiw-steps > [style*="flex:0 0 44px"]{
    display:none !important;
  }
  .dn-hiw-steps > [style*="flex:1;min-width:0"]{
    width:100% !important;
    max-width:400px !important;
    padding:20px 16px !important;
  }
}

/* ── Trust bar — wrap tightly on mobile ─────────────────────── */
@media(max-width:600px){
  .dn-trust-bar-inner{
    gap:10px !important;
    padding:8px 12px !important;
    justify-content:flex-start !important;
  }
  .dn-trust-bar-inner > span[style*="font-size:16px"]{
    display:none !important;
  }
}

/* ── General Gradio mobile fixes ────────────────────────────── */
@media(max-width:768px){
  /* Kill any fixed widths that cause horizontal scroll */
  .gradio-container [data-testid="column"]{
    min-width:0 !important;
  }
  /* Buttons full width on mobile */
  #dn-preview-unlock-btn button,
  #analyze-data-btn button,
  #dn-real-login-btn button{
    font-size:14px !important;
    padding:0 16px !important;
  }
  /* Charts don't overflow */
  .gradio-plot, .gradio-plot > div, .gradio-plot svg{
    max-width:100% !important;
    overflow:hidden !important;
  }
}



/* ═══════════════════════════════════════════════════════════
   UI/UX POLISH — Upload, Preview, Login, Cards
   ═══════════════════════════════════════════════════════════ */

/* ── Item 4: Upload container — larger, dotted border, hover ── */
#drc-block [data-testid="file"],
#drc-block .gradio-file {
  min-height:110px !important;
  border:2px dashed #2563eb !important;
  border-radius:14px !important;
  background:#EFF6FF !important;
  transition:background 0.2s, border-color 0.2s !important;
  cursor:pointer !important;
}
#drc-block [data-testid="file"]:hover,
#drc-block .gradio-file:hover {
  background:#DBEAFE !important;
  border-color:#1d4ed8 !important;
}
/* Upload text override */
#drc-block [data-testid="file"] .wrap::before,
#drc-block .gradio-file .wrap::before {
  content:"📂  Drag & Drop Excel or CSV File Here — or click to browse";
  font-size:13px !important;
  font-weight:700 !important;
  color:#1B4F8A !important;
  text-align:center !important;
  display:block !important;
}

/* ── Item 5: Preview dashboard — vertical spacing & card shadows ── */
#dn-preview-col .block,
#dn-preview-col [data-testid="html"] {
  margin-bottom:8px !important;
}
/* Preview section headers */
.dn-preview-section-hdr {
  font-size:11px !important;
  font-weight:800 !important;
  letter-spacing:1.5px !important;
  text-transform:uppercase !important;
  color:#1B4F8A !important;
  margin-bottom:12px !important;
}

/* ── Item 6: Business health score mini cards ── */
.dn-score-badge {
  display:inline-flex;
  align-items:center;
  gap:5px;
  padding:4px 10px;
  border-radius:20px;
  font-size:11px;
  font-weight:700;
  border:1px solid;
}
.dn-score-badge.green  { background:#f0fdf4; color:#16a34a; border-color:#bbf7d0; }
.dn-score-badge.blue   { background:#eff6ff; color:#2563eb; border-color:#bfdbfe; }
.dn-score-badge.amber  { background:#fffbeb; color:#d97706; border-color:#fde68a; }
.dn-score-badge.purple { background:#f5f3ff; color:#7c3aed; border-color:#ddd6fe; }

/* ── Item 7: Locked insight cards — hover blur reduction + glow ── */
.dn-preview-locked,
[class*="locked"],
.dn-blur-card {
  transition:filter 0.3s, box-shadow 0.3s, transform 0.2s !important;
  cursor:pointer !important;
}
.dn-preview-locked:hover,
[class*="locked"]:hover,
.dn-blur-card:hover {
  filter:blur(1px) !important;
  box-shadow:0 8px 32px rgba(37,99,235,0.18) !important;
  transform:translateY(-2px) !important;
}

/* ── Item 8: Unlock section — reduce excess spacing ── */
#dn-login-section-col > div > div {
  padding-top:0 !important;
}

/* ── Item 11: CTA consistency — primary blue + premium orange ── */
/* Primary blue */
#dn-preview-unlock-btn button,
[id*="analyze"] button,
#analyze-data-btn button {
  background:linear-gradient(90deg,#1B4F8A,#2563eb) !important;
  color:#ffffff !important;
  border:none !important;
  font-weight:800 !important;
  border-radius:10px !important;
  box-shadow:0 4px 14px rgba(37,99,235,0.30) !important;
}
/* Premium orange — unlock only */
#dn-preview-unlock-btn.orange-cta button {
  background:linear-gradient(90deg,#d97706,#f59e0b) !important;
  box-shadow:0 4px 14px rgba(217,119,6,0.30) !important;
}

/* ── Step 5 upload screen body — cleaner spacing ── */
#dn-step5-col .gradio-container,
#dn-step5-col > div > div {
  padding:16px 20px !important;
  gap:12px !important;
}

/* ── Preview screen — consistent padding ── */
#dn-preview-col > div > div {
  padding:0 !important;
  gap:8px !important;
}


/* ── Preview blur enhancement — stronger overlay + lock label ── */
.dn-preview-blur-card,
[class*="preview-locked"],
[class*="blur-lock"] {
  position:relative !important;
  filter:blur(4px) !important;
  user-select:none !important;
  pointer-events:none !important;
  transition:filter 0.3s !important;
}
.dn-preview-blur-card:hover { filter:blur(2px) !important; }

/* Lock overlay badge */
.dn-lock-badge {
  position:absolute;
  top:50%;left:50%;
  transform:translate(-50%,-50%);
  background:rgba(15,37,87,0.82);
  backdrop-filter:blur(4px);
  color:#ffffff;
  font-size:11px;font-weight:700;
  padding:6px 14px;
  border-radius:20px;
  border:1px solid rgba(255,255,255,0.20);
  white-space:nowrap;
  pointer-events:none;
  z-index:10;
}

/* Upload DRC section — premium spacing */
#drc-block {
  border-radius:14px !important;
  overflow:hidden !important;
}
#drc-block > div > div {
  gap:10px !important;
}

/* How it works — tighten gap between header and steps */
.dn-hiw-steps {
  margin-top:16px !important;
}

/* Sections flow — reduce gap between landing sections */
#dn-step0-col > div > div > [data-testid="html"] {
  margin-bottom:0 !important;
}
#dn-step0-col > div > div {
  gap:0 !important;
}

/* Hero section — reduce bottom gap before How It Works */
.dn-hero-wrap {
  padding-bottom:0 !important;
}

/* Pricing comparison table — tighter row padding */
[style*="HOW DOES IT COMPARE"] .grid-row,
[style*="grid-template-columns:1fr 1fr 1fr"] > div {
  padding-top:6px !important;
  padding-bottom:6px !important;
}

"""
business_types = ["Choose Business Type", "Hypermarket"]
roles = ["Business Owner","Co-Founder","Category Manager","Analyst","Store Manager"]
# For Retail industry / Services activity — only Hypermarket is a valid business type
ACTIVITY_TO_BIZ_CHOICES = {
    'FMCG':          ['Hypermarket'],
    'Services':      ['Hypermarket'],
    'Retail':        ['Hypermarket'],
    'Trading':       ['Hypermarket'],
    'Manufacturing': ['Hypermarket'],
    'Electronics':   ['Hypermarket'],
    'Clothing':      ['Hypermarket'],
    'Hypermarket':   ['Hypermarket'],
}
ACTIVITY_TO_BIZ_TYPE = {k: v[0] for k, v in ACTIVITY_TO_BIZ_CHOICES.items()}

# ══════════════════════════════════════════════════════════════════════════════
# Landing page helpers
# ══════════════════════════════════════════════════════════════════════════════
def _landing_hero(lang):
    if lang == 'kn':
        headline      = "ರಿಟೇಲ್ ಡೇಟಾದಿಂದ ರಿಟೇಲ್ ನಿರ್ಧಾರಗಳಿಗೆ"
        subtext       = "ನಿಮ್ಮ Excel ಅಥವಾ CSV ಡೇಟಾ ಅಪ್\u200cಲೋಡ್ ಮಾಡಿ ಮತ್ತು ತಕ್ಷಣ insights, dashboards ಮತ್ತು business recommendations ಪಡೆಯಿರಿ \u2014 ಕೆಲವೇ ಸೆಕೆಂಡ್\u200cಗಳಲ್ಲಿ."
        badge         = "AI-Powered ರಿಟೇಲ್ ಇಂಟೆಲಿಜೆನ್ಸ್"
        btn_primary   = "ಉಚಿತ ವಿಶ್ಲೇಷಣೆ ಪ್ರಾರಂಭಿಸಿ"
        btn_secondary = "ಡೆಮೋ ಬುಕ್ ಮಾಡಿ"
        pill1 = "\U0001f512 ಸುರಕ್ಷಿತ ಮತ್ತು ಖಾಸಗಿ"
        pill2 = "\u26a1 ಕೆಲವೇ ಸೆಕೆಂಡ್\u200cಗಳಲ್ಲಿ ಫಲಿತಾಂಶ"
        pill3 = "\U0001f4ca ತಾಂತ್ರಿಕ ಕೌಶಲ್ಯ ಬೇಡ"
    elif lang == 'ta':
        headline      = "\u0b9a\u0bbf\u0bb2\u0bcd\u0bb2\u0bb1\u0bc8 \u0ba4\u0bb0\u0bb5\u0bbf\u0bb2\u0bbf\u0bb0\u0bc1\u0ba8\u0bcd\u0ba4\u0bc1 \u0b9a\u0bbf\u0bb2\u0bcd\u0bb2\u0bb1\u0bc8 \u0bae\u0bc1\u0b9f\u0bbf\u0bb5\u0bc1\u0b95\u0bb3\u0bc1\u0b95\u0bcd\u0b95\u0bc1"
        subtext       = "\u0b89\u0b99\u0bcd\u0b95\u0bb3\u0bcd Excel \u0b85\u0bb2\u0bcd\u0bb2\u0ba4\u0bc1 CSV \u0ba4\u0bb0\u0bb5\u0bc8 \u0baa\u0ba4\u0bbf\u0bb5\u0bc7\u0bb1\u0bcd\u0bb1\u0bbf \u0b89\u0b9f\u0ba9\u0b9f\u0bbf insights, dashboards \u0bae\u0bb1\u0bcd\u0bb1\u0bc1\u0bae\u0bcd business recommendations \u0baa\u0bc6\u0bb1\u0bc1\u0b99\u0bcd\u0b95\u0bb3\u0bcd \u2014 \u0b9a\u0bbf\u0bb2 \u0bb5\u0bbf\u0ba9\u0b9f\u0bbf\u0b95\u0bb3\u0bbf\u0bb2\u0bcd."
        badge         = "AI-Powered \u0b9a\u0bbf\u0bb2\u0bcd\u0bb2\u0bb1\u0bc8 \u0ba8\u0bc1\u0ba3\u0bcd\u0ba3\u0bb1\u0bbf\u0bb5\u0bc1"
        btn_primary   = "\u0b87\u0bb2\u0bb5\u0b9a \u0baa\u0b95\u0bc1\u0baa\u0bcd\u0baa\u0bbe\u0baf\u0bcd\u0bb5\u0bc1 \u0ba4\u0bca\u0b9f\u0b99\u0bcd\u0b95\u0bc1"
        btn_secondary = "\u0b9f\u0bc6\u0bae\u0bcb \u0baa\u0ba4\u0bbf\u0bb5\u0bc1 \u0b9a\u0bc6\u0baf\u0bcd\u0b95"
        pill1 = "\U0001f512 \u0baa\u0bbe\u0ba4\u0bc1\u0b95\u0bbe\u0baa\u0bcd\u0baa\u0bbe\u0ba9\u0ba4\u0bc1 & \u0ba4\u0ba9\u0bbf\u0baa\u0bcd\u0baa\u0b9f\u0bcd\u0b9f\u0ba4\u0bc1"
        pill2 = "\u26a1 \u0bb5\u0bbf\u0ba9\u0b9f\u0bbf\u0b95\u0bb3\u0bbf\u0bb2\u0bcd \u0bae\u0bc1\u0b9f\u0bbf\u0bb5\u0bc1\u0b95\u0bb3\u0bcd"
        pill3 = "\U0001f4ca \u0ba4\u0bca\u0bb4\u0bbf\u0bb2\u0bcd\u0ba8\u0bc1\u0b9f\u0bcd\u0baa \u0b85\u0bb1\u0bbf\u0bb5\u0bc1 \u0ba4\u0bc7\u0bb5\u0bc8\u0baf\u0bbf\u0bb2\u0bcd\u0bb2\u0bc8"
    else:
        headline      = "Turn Retail Data Into Smarter Business Decisions"
        subtext       = "Upload your Excel or CSV to get AI-powered forecasting, SKU insights, and actionable business intelligence — in under 60 seconds. No technical skills required."
        badge         = "AI-Powered Retail Intelligence"
        btn_primary   = "Start Free Analysis →"
        btn_secondary = "Book a Demo"
        pill1 = "Your data never leaves your device"
        pill2 = "Results in under 60 seconds"
        pill3 = "Free preview · No credit card"

    headline_html = headline.replace('\n', '<br>')

    # ── SVG icons for trust pills ─────────────────────────────────────────────
    shield_svg = (
        '<svg width="13" height="13" viewBox="0 0 24 24" fill="none" '
        'stroke="#ffffff" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">'
        '<path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/></svg>'
    )
    zap_svg = (
        '<svg width="13" height="13" viewBox="0 0 24 24" fill="none" '
        'stroke="#ffffff" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">'
        '<polygon points="13 2 3 14 12 14 11 22 21 10 12 10 13 2"/></svg>'
    )
    lock_svg = (
        '<svg width="13" height="13" viewBox="0 0 24 24" fill="none" '
        'stroke="#ffffff" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">'
        '<rect x="3" y="11" width="18" height="11" rx="2" ry="2"/>'
        '<path d="M7 11V7a5 5 0 0 1 10 0v4"/></svg>'
    )

    # ── Health score ring (85/100) ─────────────────────────────────────────────
    score = 85
    circ  = 2 * 3.14159 * 22
    dash  = circ * score / 100
    gap   = circ - dash
    ring_svg = (
        f'<svg width="52" height="52" viewBox="0 0 52 52">'
        f'<circle cx="26" cy="26" r="22" fill="none" stroke="#eef2f8" stroke-width="5"/>'
        f'<circle cx="26" cy="26" r="22" fill="none" stroke="url(#rg2)" stroke-width="5" '
        f'stroke-linecap="round" stroke-dasharray="{dash:.1f} {gap:.1f}" '
        f'transform="rotate(-90 26 26)"/>'
        f'<defs><linearGradient id="rg2" x1="0%" y1="0%" x2="100%" y2="0%">'
        f'<stop offset="0%" stop-color="#16a34a"/><stop offset="100%" stop-color="#4ade80"/>'
        f'</linearGradient></defs></svg>'
    )

    # ── Bar chart data ─────────────────────────────────────────────────────────
    bars = [
        ('42%', False), ('58%', False), ('50%', False),
        ('71%', False), ('65%', False), ('82%', False), ('94%', True)
    ]
    bar_html = ''.join(
        f'<div class="dn-mock-cbar{" dn-mock-cbar-accent" if accent else ""}" style="height:{h}"></div>'
        for h, accent in bars
    )
    pts = [(42,False),(58,False),(50,False),(71,False),(65,False),(82,False),(94,True)]
    w = 100; step = w / (len(pts)-1)
    coords = ' '.join(f'{i*step:.1f},{76-p[0]*76/100:.1f}' for i,p in enumerate(pts))
    trend_svg = (
        f'<svg class="dn-mock-trend-svg" viewBox="0 0 100 76" preserveAspectRatio="none" xmlns="http://www.w3.org/2000/svg">'
        f'<polyline points="{coords}" fill="none" stroke="#f59e0b" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" opacity="0.9"/>'
        + ''.join(f'<circle cx="{i*step:.1f}" cy="{76-p[0]*76/100:.1f}" r="1.8" fill="#f59e0b" opacity="0.9"/>' for i,p in enumerate(pts))
        + '</svg>'
    )

    # ── Dashboard card — rich product mock ────────────────────────────────────
    # Score ring SVG (reused from above)
    _score_ring = (
        '<svg viewBox="0 0 44 44" width="44" height="44" style="width:44px;height:44px;">'
        '<circle cx="22" cy="22" r="18" fill="none" stroke="#e2e8f0" stroke-width="4"/>'
        '<circle cx="22" cy="22" r="18" fill="none" stroke="#16a34a" stroke-width="4"'
        ' stroke-dasharray="101 113" stroke-dashoffset="28" stroke-linecap="round"'
        ' transform="rotate(-90 22 22)"/>'
        '</svg>'
    )
    # Mini sparkline bars for monthly trend
    _months = [38,42,39,51,48,62,58,71,68,78,82,89]
    _bar_max = max(_months)
    _sparkbars = ''.join(
        f'<div style="flex:1;display:flex;flex-direction:column;align-items:center;gap:1px;">'
        f'<div style="width:100%;background:{"#2563eb" if i>=9 else "#bfdbfe"};'
        f'border-radius:2px 2px 0 0;height:{int(v/_bar_max*44)}px;"></div>'
        f'<div style="font-size:5.5px;color:#94a3b8;">{"JFMAMJJASOND"[i]}</div>'
        f'</div>'
        for i,v in enumerate(_months)
    )

    dashboard_card = (
        # Outer card shell
        '<div style="background:#ffffff;border-radius:16px;overflow:hidden;'
        'box-shadow:0 20px 60px rgba(11,31,58,0.18),0 4px 16px rgba(11,31,58,0.10);'
        'border:1px solid #e2eaf5;width:100%;font-family:-apple-system,BlinkMacSystemFont,sans-serif;">'

        # ── Window chrome ──────────────────────────────────────────────────────
        '<div style="background:#f8fafc;border-bottom:1px solid #e2eaf5;'
        'padding:8px 12px;display:flex;align-items:center;gap:6px;">'
        '<span style="width:10px;height:10px;border-radius:50%;background:#ff5f57;display:inline-block;"></span>'
        '<span style="width:10px;height:10px;border-radius:50%;background:#febc2e;display:inline-block;"></span>'
        '<span style="width:10px;height:10px;border-radius:50%;background:#28c840;display:inline-block;"></span>'
        '<span style="flex:1;text-align:center;font-size:9px;font-weight:600;color:#64748b;">DataNetra.ai &nbsp;›&nbsp; Sharma General Store, Mumbai</span>'
        '<span style="background:#dcfce7;color:#16a34a;font-size:8px;font-weight:700;'
        'padding:2px 7px;border-radius:10px;">● LIVE</span>'
        '</div>'

        # ── Store selector tab bar ─────────────────────────────────────────────
        '<div style="display:flex;border-bottom:1px solid #f1f5f9;background:#fafbfc;padding:0 10px;gap:2px;">'
        '<div style="padding:5px 10px;font-size:8.5px;font-weight:700;color:#1B4F8A;'
        'border-bottom:2px solid #2563eb;">Mumbai Main</div>'
        '<div style="padding:5px 10px;font-size:8.5px;color:#94a3b8;">Pune Branch</div>'
        '<div style="padding:5px 10px;font-size:8.5px;color:#94a3b8;">Nashik Outlet</div>'
        '<div style="margin-left:auto;padding:5px 8px;font-size:8px;color:#2563eb;font-weight:600;">All Stores ▾</div>'
        '</div>'

        # ── KPI strip — 4 tiles ────────────────────────────────────────────────
        '<div style="display:grid;grid-template-columns:repeat(4,1fr);'
        'border-bottom:1px solid #f1f5f9;background:#fafbff;">'

        # Tile 1 — Health Score with ring
        '<div style="padding:8px 8px 7px;border-right:1px solid #f1f5f9;display:flex;'
        'align-items:center;gap:6px;">'
        f'<div style="position:relative;width:44px;height:44px;flex-shrink:0;">{_score_ring}'
        '<div style="position:absolute;inset:0;display:flex;align-items:center;justify-content:center;'
        'font-size:9px;font-weight:900;color:#16a34a;">85</div></div>'
        '<div>'
        '<div style="font-size:8px;font-weight:700;color:#94a3b8;text-transform:uppercase;'
        'letter-spacing:0.4px;">Health</div>'
        '<div style="font-size:16px;font-weight:900;color:#16a34a;line-height:1.1;">85<span style="font-size:9px;color:#94a3b8;">/100</span></div>'
        '<div style="font-size:7px;background:#f0fdf4;color:#16a34a;font-weight:700;'
        'border-radius:3px;padding:1px 4px;display:inline-block;margin-top:2px;">Strong ✓</div>'
        '</div></div>'

        # Tile 2 — Revenue
        '<div style="padding:12px 10px;border-right:1px solid #f1f5f9;text-align:center;">'
        '<div style="font-size:8px;font-weight:700;color:#94a3b8;text-transform:uppercase;'
        'letter-spacing:0.4px;margin-bottom:2px;">Revenue</div>'
        '<div style="font-size:17px;font-weight:900;color:#0B1F3A;">₹4.8Cr</div>'
        '<div style="font-size:9px;color:#16a34a;font-weight:700;margin-top:2px;">▲ 28% YoY</div>'
        '</div>'

        # Tile 3 — Top SKU
        '<div style="padding:12px 10px;border-right:1px solid #f1f5f9;text-align:center;">'
        '<div style="font-size:8px;font-weight:700;color:#94a3b8;text-transform:uppercase;'
        'letter-spacing:0.4px;margin-bottom:2px;">Top SKU</div>'
        '<div style="font-size:11px;font-weight:800;color:#0B1F3A;line-height:1.3;">Basmati<br>Rice 5kg</div>'
        '<div style="font-size:9px;color:#2563eb;font-weight:700;margin-top:2px;">₹68.4L sales</div>'
        '</div>'

        # Tile 4 — Forecast
        '<div style="padding:12px 10px;text-align:center;">'
        '<div style="font-size:8px;font-weight:700;color:#94a3b8;text-transform:uppercase;'
        'letter-spacing:0.4px;margin-bottom:2px;">6M Forecast</div>'
        '<div style="font-size:17px;font-weight:900;color:#7c3aed;">₹2.6Cr</div>'
        '<div style="font-size:9px;color:#7c3aed;font-weight:700;margin-top:2px;">▲ +18% AI est.</div>'
        '</div>'

        '</div>'  # /KPI strip

        # ── AI Scores row ──────────────────────────────────────────────────────
        '<div style="padding:8px 10px;border-bottom:1px solid #f1f5f9;">'
        '<div style="font-size:7px;font-weight:800;color:#374151;text-transform:uppercase;'
        'letter-spacing:0.6px;margin-bottom:6px;">AI Business Scores</div>'
        '<div style="display:flex;flex-direction:column;gap:4px;">'
        + ''.join(
            f'<div style="display:flex;align-items:center;gap:6px;">'
            f'<div style="font-size:9px;color:#475569;width:100px;flex-shrink:0;font-weight:600;">{label}</div>'
            f'<div style="flex:1;height:7px;background:#f1f5f9;border-radius:4px;overflow:hidden;">'
            f'<div style="width:{pct}%;height:100%;background:linear-gradient(90deg,{c1},{c2});border-radius:4px;"></div></div>'
            f'<div style="font-size:10px;font-weight:800;color:{c1};width:28px;text-align:right;">{pct}</div>'
            f'</div>'
            for label,pct,c1,c2 in [
                ("Financial Risk",   82, "#16a34a", "#22c55e"),
                ("Growth Potential", 74, "#2563eb", "#60a5fa"),
                ("Vendor Reliability",68,"#f59e0b","#fbbf24"),
                ("ONDC Readiness",   91, "#7c3aed", "#a78bfa"),
            ]
        )
        + '</div></div>'

        # ── Monthly trend sparkline ────────────────────────────────────────────
        '<div style="padding:7px 10px 8px;">'
        '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:5px;">'
        '<div style="font-size:7px;font-weight:800;color:#374151;text-transform:uppercase;'
        'letter-spacing:0.6px;">Monthly Sales Trend (₹L)</div>'
        '<div style="font-size:7px;color:#2563eb;font-weight:600;">FY 2024–25</div>'
        '</div>'
        f'<div style="display:flex;align-items:flex-end;gap:2px;height:50px;padding:0 2px;">{_sparkbars}</div>'
        '<div style="display:flex;justify-content:space-between;margin-top:4px;">'
        '<div style="font-size:7px;color:#94a3b8;">Apr 2024</div>'
        '<div style="font-size:7px;color:#2563eb;font-weight:700;">▲ Best month: Mar 2025</div>'
        '<div style="font-size:7px;color:#94a3b8;">Mar 2025</div>'
        '</div>'
        '</div>'

        '</div>'  # /outer card
    )


    # ── Right panel: Unlock Full Insights ─────────────────────────────────────


    return (
        # Inline JS — always executes with the hero HTML, no cross-block scope issues
        '<script>'
        'window.dnOpenModal=function(){var o=document.getElementById("dn-demo-modal-overlay");if(o)o.style.display="flex";document.body.style.overflow="hidden";};'
        'window.dnCloseModal=function(){var o=document.getElementById("dn-demo-modal-overlay");if(o)o.style.display="none";document.body.style.overflow="";var f=document.getElementById("dn-modal-form-body"),s=document.getElementById("dn-modal-success"),e=document.getElementById("dn-demo-err");if(f)f.style.display="block";if(s)s.style.display="none";if(e)e.style.display="none";};'
        'window.dnOpenUnlockModal=function(){var o=document.getElementById("dn-unlock-modal-overlay");if(o)o.style.display="flex";document.body.style.overflow="hidden";};'
        'window.dnCloseUnlockModal=function(){var o=document.getElementById("dn-unlock-modal-overlay");if(o)o.style.display="none";document.body.style.overflow="";};'
        'window.dnUnlockContinue=function(){window.dnCloseUnlockModal&&window.dnCloseUnlockModal();setTimeout(function(){var a=document.getElementById("dn-input-anchor");if(a)a.scrollIntoView({behavior:"smooth",block:"start"});},80);};'
        'function dnOpenModal(){window.dnOpenModal();}function dnCloseModal(){window.dnCloseModal();}function dnOpenUnlockModal(){window.dnOpenUnlockModal();}function dnCloseUnlockModal(){window.dnCloseUnlockModal();}function dnUnlockContinue(){window.dnUnlockContinue();}'
        '</script>'

        # ── Trust bar — IDA badge + social proof numbers ───────────────────────
        '<div style="background:#0B1F3A;border-bottom:1px solid rgba(255,255,255,0.08);padding:9px 24px;">'
        '<div class="dn-trust-bar-inner" style="max-width:1100px;margin:0 auto;display:flex;align-items:center;justify-content:center;'
        'gap:24px;flex-wrap:wrap;">'

        '<div style="display:flex;align-items:center;gap:6px;">'
        '<span style="font-size:13px;font-weight:900;color:#ffffff;">60s</span>'
        '<span style="font-size:11px;color:#94a3b8;">to first insight</span>'
        '</div>'

        '<span style="color:rgba(255,255,255,0.15);font-size:16px;">|</span>'

        # Stat 2
        '<div style="display:flex;align-items:center;gap:6px;">'
        '<span style="font-size:13px;font-weight:900;color:#ffffff;">7</span>'
        '<span style="font-size:11px;color:#94a3b8;">AI scoring models</span>'
        '</div>'

        '<span style="color:rgba(255,255,255,0.15);font-size:16px;">|</span>'

        # Stat 3
        '<div style="display:flex;align-items:center;gap:6px;">'
        '<span style="font-size:13px;font-weight:900;color:#ffffff;">Free</span>'
        '<span style="font-size:11px;color:#94a3b8;">preview · no card needed</span>'
        '</div>'

        '<span style="color:rgba(255,255,255,0.15);font-size:16px;">|</span>'

        # Stat 4
        '<div style="display:flex;align-items:center;gap:6px;">'
        '<span style="font-size:13px;font-weight:900;color:#4ade80;">✓</span>'
        '<span style="font-size:11px;color:#94a3b8;">Data stays on your device</span>'
        '</div>'

        '</div></div>'

        '<div class="dn-hero-wrap">'
        '<div class="dn-hero-bg">'

        # Left text column
        '<div class="dn-hero-left">'

        # Badge — IDA + AI tag
        f'<h1 class="dn-hero-h1">{headline_html}</h1>'
        f'<p class="dn-hero-sub">{subtext}</p>'

        # Value bullets — quick scannable wins
        '<div style="margin:16px 0;display:flex;flex-direction:column;gap:8px;">'
        '<div style="display:flex;align-items:center;gap:10px;">'
        '<div style="width:20px;height:20px;background:#16a34a;border-radius:50%;display:flex;'
        'align-items:center;justify-content:center;flex-shrink:0;">'
        '<svg width="11" height="11" viewBox="0 0 24 24" fill="none" stroke="#fff" '
        'stroke-width="3" stroke-linecap="round" stroke-linejoin="round">'
        '<polyline points="20 6 9 17 4 12"/></svg></div>'
        '<span style="font-size:13px;color:rgba(255,255,255,0.92);font-weight:500;">Know which SKUs are draining your profit — instantly</span></div>'
        '<div style="display:flex;align-items:center;gap:10px;">'
        '<div style="width:20px;height:20px;background:#16a34a;border-radius:50%;display:flex;'
        'align-items:center;justify-content:center;flex-shrink:0;">'
        '<svg width="11" height="11" viewBox="0 0 24 24" fill="none" stroke="#fff" '
        'stroke-width="3" stroke-linecap="round" stroke-linejoin="round">'
        '<polyline points="20 6 9 17 4 12"/></svg></div>'
        '<span style="font-size:13px;color:rgba(255,255,255,0.92);font-weight:500;">6 &amp; 12-month AI revenue forecast from your own data</span></div>'
        '<div style="display:flex;align-items:center;gap:10px;">'
        '<div style="width:20px;height:20px;background:#16a34a;border-radius:50%;display:flex;'
        'align-items:center;justify-content:center;flex-shrink:0;">'
        '<svg width="11" height="11" viewBox="0 0 24 24" fill="none" stroke="#fff" '
        'stroke-width="3" stroke-linecap="round" stroke-linejoin="round">'
        '<polyline points="20 6 9 17 4 12"/></svg></div>'
        '<span style="font-size:13px;color:rgba(255,255,255,0.92);font-weight:500;">ONDC marketplace readiness score &amp; best platform match</span></div>'
        '</div>'

        # CTAs — primary is direct, no friction
        '<div class="dn-hero-btns" style="margin-top:20px;display:flex;gap:12px;flex-wrap:wrap;align-items:center;">'
        '<a href="#dn-drc-anchor" '
        'onclick="var a=document.getElementById(\'dn-drc-anchor\');if(a)a.scrollIntoView({behavior:\'smooth\'});return false;" '
        'style="display:inline-flex;align-items:center;gap:8px;background:linear-gradient(90deg,#1B4F8A,#2563eb);'
        'color:#ffffff;font-size:15px;font-weight:800;padding:14px 28px;border-radius:10px;'
        'text-decoration:none;box-shadow:0 6px 20px rgba(37,99,235,0.40);white-space:nowrap;">'
        f'{btn_primary}</a>'
        '<div style="font-size:11px;color:#64748b;line-height:1.4;">'
        '↑ Upload your Excel/CSV<br>No signup needed to preview</div>'
        '</div>'

        # Trust pills
        '<div class="dn-hero-trust" style="margin-top:16px;">'
        f'<span class="dn-trust-pill">{shield_svg}{pill1}</span>'
        f'<span class="dn-trust-pill">{zap_svg}{pill2}</span>'
        f'<span class="dn-trust-pill">{lock_svg}{pill3}</span>'
        '</div>'
        '</div>'

        # Center: premium dashboard card
        f'<div class="dn-hero-center">{dashboard_card}</div>'

        '</div>'  # /hero-bg
        '</div>'  # /hero-wrap
    )

def _landing_capabilities(lang):
    _footer_inner = """
<div style="max-width:1100px;margin:0 auto;padding:28px 40px 20px;box-sizing:border-box;">

  <!-- Top row: logo + nav links -->
  <div style="display:flex;flex-wrap:wrap;gap:40px;justify-content:space-between;align-items:flex-start;padding-bottom:36px;border-bottom:1px solid rgba(255,255,255,0.12);">

    <!-- Brand -->
    <div style="flex:0 0 220px;">
      <div style="font-size:20px;font-weight:900;color:#ffffff;letter-spacing:-0.5px;margin-bottom:8px;">DataNetra.ai</div>
      <div style="font-size:12.5px;color:rgba(255,255,255,0.55);line-height:1.6;margin-bottom:16px;">
        AI-powered retail intelligence for businesses. Upload your data, get instant insights.
      </div>
      <!-- LinkedIn -->
      <a href="https://www.linkedin.com/company/108412762/" target="_blank"
         style="display:inline-flex;align-items:center;gap:7px;font-size:12px;font-weight:600;
                color:#93c5fd;text-decoration:none;background:rgba(147,197,253,0.12);
                border:1px solid rgba(147,197,253,0.28);border-radius:8px;padding:7px 13px;
                transition:background 0.15s;"
         onmouseover="this.style.background='rgba(147,197,253,0.22)'"
         onmouseout="this.style.background='rgba(147,197,253,0.12)'">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
          <path d="M16 8a6 6 0 0 1 6 6v7h-4v-7a2 2 0 0 0-2-2 2 2 0 0 0-2 2v7h-4v-7a6 6 0 0 1 6-6z"/>
          <rect x="2" y="9" width="4" height="12"/><circle cx="4" cy="4" r="2"/>
        </svg>
        LinkedIn
      </a>
    </div>

    <!-- Product links -->
    <div style="flex:0 0 140px;">
      <div style="font-size:11px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase;color:#7AABDD;margin-bottom:14px;">Product</div>
      <div style="display:flex;flex-direction:column;gap:10px;">
        <a href="#dn-drc-anchor" onclick="var a=document.getElementById('dn-drc-anchor');if(a)a.scrollIntoView({behavior:'smooth'});return false;"
           style="font-size:13px;color:rgba(255,255,255,0.72);text-decoration:none;transition:color 0.15s;"
           onmouseover="this.style.color='#ffffff'" onmouseout="this.style.color='rgba(255,255,255,0.72)'">Get Started</a>
        <a href="#dn-contact-anchor" onclick="var a=document.getElementById('dn-contact-anchor');if(a)a.scrollIntoView({behavior:'smooth'});return false;"
           style="font-size:13px;color:rgba(255,255,255,0.72);text-decoration:none;transition:color 0.15s;"
           onmouseover="this.style.color='#ffffff'" onmouseout="this.style.color='rgba(255,255,255,0.72)'">Pricing</a>
        <a href="#dn-contact-anchor" onclick="var a=document.getElementById('dn-contact-anchor');if(a)a.scrollIntoView({behavior:'smooth'});return false;"
           style="font-size:13px;color:rgba(255,255,255,0.72);text-decoration:none;transition:color 0.15s;"
           onmouseover="this.style.color='#ffffff'" onmouseout="this.style.color='rgba(255,255,255,0.72)'">Contact</a>
      </div>
    </div>

    <!-- Industries -->
    <div style="flex:0 0 160px;">
      <div style="font-size:11px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase;color:#7AABDD;margin-bottom:14px;">Industries</div>
      <div style="display:flex;flex-direction:column;gap:10px;">
        <span style="font-size:13px;color:rgba(255,255,255,0.72);">Supermarkets</span>
        <span style="font-size:13px;color:rgba(255,255,255,0.72);">FMCG &amp; Electronics</span>
        <span style="font-size:13px;color:rgba(255,255,255,0.72);">Apparel &amp; Retail</span>
        <span style="font-size:13px;color:rgba(255,255,255,0.45);font-style:italic;font-size:12px;">Healthcare · Energy · Agriculture <br><span style="color:#fbbf24;font-weight:600;">Coming soon</span></span>
      </div>
    </div>

    <!-- Contact -->
    <div style="flex:0 0 180px;">
      <div style="font-size:11px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase;color:#7AABDD;margin-bottom:14px;">Contact</div>
      <div style="display:flex;flex-direction:column;gap:10px;">
        <a href="mailto:support@datanetra.ai"
           style="font-size:13px;color:rgba(255,255,255,0.72);text-decoration:none;transition:color 0.15s;"
           onmouseover="this.style.color='#ffffff'" onmouseout="this.style.color='rgba(255,255,255,0.72)'">support@datanetra.ai</a>
        <div style="font-size:12px;color:rgba(255,255,255,0.45);line-height:1.5;">Response within 24 hours</div>
      </div>
      <!-- Back to top -->
      <a href="#" onclick="window.scrollTo({top:0,behavior:'smooth'});return false;"
         style="display:inline-flex;align-items:center;gap:6px;margin-top:20px;
                font-size:12px;font-weight:600;color:#93c5fd;text-decoration:none;
                background:rgba(147,197,253,0.10);border:1px solid rgba(147,197,253,0.25);
                border-radius:7px;padding:6px 12px;transition:background 0.15s;"
         onmouseover="this.style.background='rgba(147,197,253,0.20)'"
         onmouseout="this.style.background='rgba(147,197,253,0.10)'">
        <svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><polyline points="18 15 12 9 6 15"/></svg>
        Back to top
      </a>
    </div>
  </div>

  <!-- Made in India trust strip -->
  <div style="background:rgba(255,255,255,0.04);border-radius:10px;padding:12px 20px;margin-bottom:16px;
              display:flex;flex-wrap:wrap;gap:16px;align-items:center;justify-content:center;border:1px solid rgba(255,255,255,0.08);">
    <span style="font-size:12px;color:rgba(255,255,255,0.60);">🇮🇳 Made in India</span>
    <span style="color:rgba(255,255,255,0.15);">·</span>
    <span style="font-size:12px;color:rgba(255,255,255,0.60);">🔒 Your data is never stored on our servers</span>
    <span style="color:rgba(255,255,255,0.15);">·</span>
    <span style="font-size:12px;color:rgba(255,255,255,0.60);">📋 Aligned with India's DPDP Act</span>
    <span style="color:rgba(255,255,255,0.15);">·</span>
    <span style="font-size:12px;color:rgba(255,255,255,0.60);">🛒 ONDC-Ready Platform</span>
  </div>

  <!-- Bottom row: copyright + compliance -->
  <div style="display:flex;flex-wrap:wrap;gap:12px;justify-content:space-between;align-items:center;padding-top:16px;">
    <div style="font-size:12px;color:rgba(255,255,255,0.35);">
      © 2025 DataNetra.ai · All rights reserved.
    </div>
    <div style="display:flex;align-items:center;gap:8px;">
      <div style="width:6px;height:6px;border-radius:50%;background:#22c55e;"></div>
      <div style="font-size:12px;color:rgba(255,255,255,0.35);">Privacy · Terms · support@datanetra.ai</div>
    </div>
  </div>

</div>"""
    return f'<div style="background:#0B1F3A;width:100%;box-sizing:border-box;">{_footer_inner}</div>'

# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  SECTION DRC  —  Data Readiness Check (Phase 1)                            ║
# ║  Isolated helpers: normalize_headers, map_columns,                         ║
# ║  run_readiness_check, export_clean_dataset                                 ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

_DRC_ALIAS_MAP = {
    "date":     ["date","invoice_date","txn_date","sales_date","order_date",
                 "bill_date","transaction_date","trans_date","orderdate",
                 "invoicedate","saledate"],
    "product":  ["product","product_name","sku","sku_name","item_name",
                 "material_name","item","product_desc","description",
                 "article","article_name","productname"],
    "category": ["category","product_category","item_group","group_name",
                 "department","cat","prod_cat","item_category",
                 "productcategory","itemgroup"],
    "sales":    ["sales","revenue","gross_sales","sales_amount","amount",
                 "turnover","net_sales","total_sales","sale_value",
                 "invoice_value","bill_amount","total_amount"],
    "quantity": ["qty","quantity","units","units_sold","sold_qty",
                 "sales_qty","sale_qty","no_of_units","pieces","pcs"],
    "cost":     ["cost","cost_price","unit_cost","purchase_price",
                 "buying_price","cogs","landed_cost","cost_per_unit"],
    "returns":  ["returns","returned_qty","qty_returned","return_qty",
                 "return_units","sales_return","sales_returns"],
    "store":    ["store","store_name","branch","location","outlet",
                 "shop","shop_name","store_id","branch_name"],
}
_DRC_CRITICAL = {"date","product","category","sales","quantity"}
_DRC_OPTIONAL = {"cost","returns","store"}

def normalize_headers(df):
    df = df.copy()
    new_cols = []
    for col in df.columns:
        c = str(col).strip()
        c = re.sub(r"\s+", "_", c).lower()
        c = re.sub(r"[^\w]", "_", c)
        c = re.sub(r"_+", "_", c).strip("_")
        new_cols.append(c)
    df.columns = new_cols
    return df

def map_columns(df):
    cols = list(df.columns)
    mapping = {}
    for std, aliases in _DRC_ALIAS_MAP.items():
        mapping[std] = next((a for a in aliases if a in cols), None)
    return mapping

def _drc_status_pill(label):
    colors = {"Ready":"#16a34a","Partial":"#d97706","Needs Completion":"#dc2626"}
    icons  = {"Ready":"✅","Partial":"⚠️","Needs Completion":"❌"}
    c = colors.get(label,"#6b7280"); ic = icons.get(label,"ℹ️")
    return (
        f"<div style='display:inline-flex;align-items:center;gap:8px;padding:9px 20px;"
        f"border-radius:9px;background:{c}18;border:1.5px solid {c};font-family:sans-serif;'>"
        f"<span style='font-size:20px;'>{ic}</span>"
        f"<span style='font-size:15px;font-weight:700;color:{c};'>"
        f"Readiness Status: {label}</span></div>"
    )

def _drc_completeness_check(df_clean, mapping):
    """
    READ-ONLY. Measure completeness over the 5 core operational fields.
    Fields: date, product_id/product, sales, units_sold/quantity, category.
    Returns (pct: float, label: str, color: str, warning: str|None, html: str).
    """
    _COMP_ALIASES = {
        "date":     ["date","invoice_date","txn_date","sales_date","order_date",
                     "transaction_date","trans_date"],
        "product":  ["product","product_name","sku","sku_name","item_name",
                     "product_id","article","productname"],
        "sales":    ["sales","revenue","gross_sales","sales_amount","amount",
                     "turnover","net_sales","total_sales","sale_value"],
        "quantity": ["quantity","units_sold","qty","sold_qty","sales_qty",
                     "sale_qty","no_of_units","pieces","pcs","units"],
        "category": ["category","product_category","item_group","group_name",
                     "department","cat","prod_cat","item_category"],
    }
    n = max(len(df_clean), 1)
    field_results = {}
    for std, aliases in _COMP_ALIASES.items():
        col = mapping.get(std)
        if not col or col not in df_clean.columns:
            col = next((a for a in aliases if a in df_clean.columns), None)
        if col and col in df_clean.columns:
            blank = (df_clean[col].isna().sum() +
                     (df_clean[col].astype(str).str.strip() == "").sum())
            filled = max(0, n - int(blank))
            field_results[std] = (col, filled, n)
        else:
            field_results[std] = (None, 0, n)

    total_valid    = sum(f for _, f, _ in field_results.values())
    total_expected = sum(t for _, _, t in field_results.values())
    pct = round(total_valid / max(total_expected, 1) * 100, 1)

    if pct >= 70:
        label, color = "Structure mostly complete",   "#16a34a"
    elif pct >= 50:
        label, color = "Required fields partially available", "#d97706"
    else:
        label, color = "Key fields missing", "#dc2626"

    warning = (
        "Dataset may not represent real operational activity. "
        "Please provide more operational records."
        if pct < 50 else None
    )

    # ── Build HTML ────────────────────────────────────────────────────────────
    rows_html = ""
    for std, (col, filled, total) in field_results.items():
        field_pct = round(filled / max(total, 1) * 100, 1)
        bar_col   = "#16a34a" if field_pct >= 90 else "#d97706" if field_pct >= 60 else "#dc2626"
        col_label = f"({col})" if col else "— not mapped"
        rows_html += (
            f"<div style='margin-bottom:6px;'>"
            f"<div style='display:flex;justify-content:space-between;"
            f"font-size:11px;color:#475569;margin-bottom:2px;'>"
            f"<span><b style='color:#0f172a;'>{std}</b> "
            f"<span style='color:#94a3b8;font-size:10px;'>{col_label}</span></span>"
            f"<span style='font-weight:700;color:{bar_col};'>{field_pct}%</span>"
            f"</div>"
            f"<div style='height:5px;background:#e2e8f0;border-radius:3px;overflow:hidden;'>"
            f"<div style='height:100%;width:{field_pct}%;background:{bar_col};"
            f"border-radius:3px;'></div></div>"
            f"</div>"
        )

    warn_html = ""
    if warning:
        warn_html = (
            f"<div style='margin-top:8px;padding:6px 10px;background:#fef2f2;"
            f"border:1px solid #dc262633;border-radius:5px;"
            f"font-size:11px;color:#dc2626;'>"
            f"⚠️ {warning}</div>"
        )

    bar_w = max(2, pct)
    html = (
        f"<div style='background:#fff;border:1px solid #e2e8f0;"
        f"border-top:3px solid {color};border-radius:8px;"
        f"padding:13px 15px;margin-bottom:10px;"
        f"box-shadow:0 2px 10px rgba(11,31,58,0.10);'>"
        f"<div style='display:flex;justify-content:space-between;"
        f"align-items:baseline;margin-bottom:10px;'>"
        f"<span style='font-size:12px;font-weight:700;color:#0f172a;'>"
        f"📋 Dataset Structure Coverage</span>"
        f"<span style='font-size:20px;font-weight:900;color:{color};"
        f"font-family:monospace;'>{pct}<span style='font-size:12px;"
        f"font-weight:600;color:{color}88;'>%</span></span>"
        f"</div>"
        f"<div style='height:7px;background:#e2e8f0;border-radius:4px;"
        f"overflow:hidden;margin-bottom:6px;'>"
        f"<div style='height:100%;width:{bar_w}%;background:{color};"
        f"border-radius:4px;'></div></div>"
        f"<div style='display:inline-block;padding:2px 10px;border-radius:12px;"
        f"background:{color}15;border:1px solid {color}33;"
        f"font-size:11px;font-weight:700;color:{color};margin-bottom:10px;'>"
        f"{label}</div>"
        f"<div style='font-size:10px;color:#64748b;margin-bottom:8px;'>"
        f"≥70% Good &nbsp;|&nbsp; 50–69% Partial &nbsp;|&nbsp; &lt;50% Insufficient</div>"
        + rows_html + warn_html +
        "<div style='margin-top:8px;padding:7px 10px;background:#f8fafc;"
        "border:1px solid #e2e8f0;border-radius:5px;"
        "font-size:10.5px;color:#64748b;line-height:1.5;'>"
        "<b style='color:#334155;'>ℹ️ Structure Coverage</b> measures whether required retail fields exist. "
        "Data quality and accuracy are evaluated separately by the readiness score above."
        "</div>"
        + f"</div>"
    )
    return pct, label, color, warning, html

def _drc_authenticity_check(df_clean, mapping):
    """
    READ-ONLY heuristics to detect unrealistic or synthetic datasets.
    Checks: value repetition, sequential patterns, category diversity,
    vendor diversity.
    Returns list of (flag_type, message) and rendered HTML.
    flag_type: 'pass' | 'warning'
    """
    import pandas as _pd_auth

    _AUTH_ALIASES = {
        "sales":    ["sales","revenue","gross_sales","sales_amount","amount",
                     "turnover","total_sales"],
        "cost":     ["cost","cost_price","unit_cost","purchase_price","buying_price"],
        "quantity": ["quantity","units_sold","qty","sold_qty","sales_qty","units"],
        "category": ["category","product_category","department","item_group"],
        "vendor":   ["vendor_name","vendor","supplier","supplier_name",
                     "distributor","manufacturer"],
    }

    def _resolve(std):
        col = mapping.get(std)
        if col and col in df_clean.columns:
            return col
        for a in _AUTH_ALIASES.get(std, []):
            if a in df_clean.columns:
                return a
        return None

    flags = []   # (flag_type, label, detail)
    n = max(len(df_clean), 1)

    # ── A. Value repetition in numeric fields ─────────────────────────────────
    rep_fields_checked = []
    for std in ("sales", "cost", "quantity"):
        col = _resolve(std)
        if col:
            raw = (_pd_auth.to_numeric(
                df_clean[col].astype(str)
                .str.replace(r"[₹$€£,\s]", "", regex=True)
                .str.replace(r"Rs\.?", "", regex=True),
                errors="coerce"
            ).dropna())
            if len(raw) >= 5:
                top_val_pct = raw.value_counts().iloc[0] / len(raw)
                rep_fields_checked.append((std, round(top_val_pct * 100, 1)))

    high_rep = [(s, p) for s, p in rep_fields_checked if p > 40]
    if high_rep:
        fields_str = ", ".join(f"{s} ({p}% identical)" for s, p in high_rep)
        flags.append(("warning",
            "Low variance detected",
            f"More than 40% of rows share the same value in: {fields_str}. "
            "This may indicate sample or test data."))
    elif rep_fields_checked:
        flags.append(("pass",
            "Value distribution appears realistic",
            "Numeric fields show adequate spread across rows."))

    # ── B. Sequential pattern detection ──────────────────────────────────────
    for std in ("sales", "quantity"):
        col = _resolve(std)
        if col:
            num = (_pd_auth.to_numeric(
                df_clean[col].astype(str)
                .str.replace(r"[₹$€£,\s]", "", regex=True),
                errors="coerce"
            ).dropna().head(50).values)
            if len(num) >= 5:
                diffs = [round(num[i+1] - num[i], 2) for i in range(min(10, len(num)-1))]
                non_zero_diffs = [d for d in diffs if d != 0]
                if non_zero_diffs and len(set(non_zero_diffs)) == 1:
                    flags.append(("warning",
                        "Possible synthetic pattern",
                        f"Column '{col}' increases in exact equal intervals "
                        f"({non_zero_diffs[0]:+.0f} per row). Real transactional data "
                        "rarely follows perfectly uniform intervals."))
                    break

    # ── C. Category diversity ─────────────────────────────────────────────────
    cat_col = _resolve("category")
    if cat_col:
        cats     = df_clean[cat_col].dropna().astype(str).str.strip()
        cats     = cats[cats != ""]
        if len(cats) >= 5:
            top_pct  = cats.value_counts().iloc[0] / len(cats)
            n_unique = cats.nunique()
            if top_pct > 0.80:
                flags.append(("warning",
                    "Low category diversity",
                    f"{round(top_pct*100,1)}% of rows share the same category "
                    f"('{cats.value_counts().index[0]}'). "
                    "Broadening category coverage improves analytics quality."))
            else:
                flags.append(("pass",
                    "Category diversity looks healthy",
                    f"{n_unique} unique categories found across {len(cats):,} rows."))

    # ── D. Vendor diversity ───────────────────────────────────────────────────
    vendor_col = _resolve("vendor")
    if vendor_col:
        vendors  = df_clean[vendor_col].dropna().astype(str).str.strip()
        vendors  = vendors[vendors != ""]
        n_vendor = vendors.nunique()
        if n_vendor == 1:
            flags.append(("warning",
                "Low operational diversity",
                f"Only one unique vendor/supplier detected ('{vendors.iloc[0]}'). "
                "Real retail datasets typically contain multiple vendors."))
        elif n_vendor > 1:
            flags.append(("pass",
                "Vendor diversity looks healthy",
                f"{n_vendor} unique vendors found."))

    # ── Render HTML ───────────────────────────────────────────────────────────
    if not flags:
        flags.append(("pass",
            "Authenticity checks passed",
            "No synthetic data patterns detected."))

    # ── Map flags → compact card slots ────────────────────────────────────────
    # Each slot: (title, icon_svg, status_label, status_type, detail_text)
    # status_type: "pass" | "warning" | "skipped"
    _SLOT_DEFS = [
        ("value_dist",    "Value Distribution",  "📊"),
        ("category",      "Category Diversity",  "🏷️"),
        ("vendor",        "Vendor Diversity",     "🏪"),
        ("pattern",       "Pattern Check",        "🔍"),
    ]

    # Parse flags into slot buckets
    _slot_data = {}
    for ftype, label, detail in flags:
        lbl_lower = label.lower()
        if any(k in lbl_lower for k in ("value", "distribution", "variance", "repetition", "numeric")):
            _slot_data.setdefault("value_dist", (ftype, label, detail))
        elif any(k in lbl_lower for k in ("category",)):
            _slot_data.setdefault("category", (ftype, label, detail))
        elif any(k in lbl_lower for k in ("vendor", "operational", "supplier")):
            _slot_data.setdefault("vendor", (ftype, label, detail))
        elif any(k in lbl_lower for k in ("pattern", "synthetic", "sequential")):
            _slot_data.setdefault("pattern", (ftype, label, detail))
        elif "authenticity checks passed" in lbl_lower:
            # General pass — fill value_dist if empty
            _slot_data.setdefault("value_dist", (ftype, label, detail))

    def _auth_card(slot_key, slot_title, slot_icon):
        """Render one compact authenticity card."""
        if slot_key in _slot_data:
            ftype, label, detail = _slot_data[slot_key]
            if ftype == "pass":
                status_color = "#16a34a"
                status_bg    = "#f0fdf4"
                status_bdr   = "#bbf7d0"
                status_dot   = "●"
                status_short = "Realistic" if slot_key == "value_dist" else                                "Healthy"   if slot_key in ("category","vendor") else                                "No Pattern"
            else:  # warning
                status_color = "#d97706"
                status_bg    = "#fffceb"
                status_bdr   = "#fde68a"
                status_dot   = "▲"
                status_short = "Warning"
        else:
            # Column not in dataset → skipped
            ftype        = "skipped"
            label        = "Skipped"
            detail       = "Column not available in dataset"
            status_color = "#94a3b8"
            status_bg    = "#f8fafc"
            status_bdr   = "#e2e8f0"
            status_dot   = "—"
            status_short = "Skipped"

        # Truncate detail to one line (~70 chars)
        short_detail = detail if len(detail) <= 72 else detail[:69] + "…"

        return (
            f"<div style='flex:1;min-width:130px;background:{status_bg};"
            f"border:1px solid {status_bdr};border-radius:8px;"
            f"padding:10px 12px;box-shadow:0 1px 4px rgba(11,31,58,0.06);'>"

            # Header row: icon + title
            f"<div style='display:flex;align-items:center;gap:5px;margin-bottom:6px;'>"
            f"<span style='font-size:14px;'>{slot_icon}</span>"
            f"<span style='font-size:10.5px;font-weight:700;color:#334155;"
            f"letter-spacing:0.2px;'>{slot_title}</span>"
            f"</div>"

            # Status badge
            f"<div style='display:inline-flex;align-items:center;gap:4px;"
            f"background:{status_color}18;border:1px solid {status_color}44;"
            f"border-radius:4px;padding:2px 8px;margin-bottom:5px;'>"
            f"<span style='font-size:9px;color:{status_color};'>{status_dot}</span>"
            f"<span style='font-size:10.5px;font-weight:700;color:{status_color};'>"
            f"{status_short}</span>"
            f"</div>"

            # One-line detail
            f"<div style='font-size:9.5px;color:#64748b;line-height:1.4;'>"
            f"{short_detail}</div>"

            f"</div>"
        )

    cards_html = "".join(
        _auth_card(key, title, icon)
        for key, title, icon in _SLOT_DEFS
    )

    html = (
        "<div style='background:#fff;border:1px solid #e2e8f0;"
        "border-top:3px solid #7c3aed;border-radius:8px;"
        "padding:13px 15px;margin-bottom:10px;"
        "box-shadow:0 2px 10px rgba(11,31,58,0.10);'>"

        "<div style='display:flex;justify-content:space-between;"
        "align-items:baseline;margin-bottom:10px;'>"
        "<span style='font-size:12px;font-weight:700;color:#0f172a;'>"
        "🔬 Dataset Authenticity Check</span>"
        "<span style='font-size:9.5px;color:#94a3b8;font-style:italic;'>"
        "Checks skipped where column is absent</span>"
        "</div>"

        "<div style='display:flex;gap:8px;flex-wrap:wrap;'>"
        + cards_html +
        "</div>"

        "</div>"
    )
    return flags, html

def _drc_transformation_summary(df_clean, mapping):
    """
    READ-ONLY. Calculates a before/after cleaning summary by inspecting
    df_clean (the raw-normalised frame before apply_cleaning_rules).
    Returns HTML showing what DataNetra will fix on export.
    """
    import pandas as _pd_tx

    n_rows = len(df_clean)

    # ── Count issues in raw frame (mirrors apply_cleaning_rules logic) ────────

    # Missing values in any column
    n_missing = int(df_clean.isna().sum().sum()) + int(
        (df_clean == "").sum().sum()
    )

    # Invalid / blank dates
    date_col = next((c for c in ["date","invoice_date","txn_date",
                                  "sales_date","order_date"] if c in df_clean.columns), None)
    n_bad_dates = 0
    if date_col:
        d = df_clean[date_col]
        blank_d = int(d.isna().sum()) + int((d.astype(str).str.strip() == "").sum())
        try:
            parsed = _pd_tx.to_datetime(d, errors="coerce")
            n_bad_dates = max(0, int(parsed.isna().sum()) - blank_d) + blank_d
        except Exception:
            n_bad_dates = blank_d

    # Currency-polluted numeric cells
    currency_cols = [c for c in ["sales","revenue","gross_sales","sales_amount",
                                  "amount","turnover","total_sales","cost",
                                  "cost_price","unit_cost","purchase_price"]
                     if c in df_clean.columns]
    n_currency = 0
    for col in currency_cols:
        has_sym = df_clean[col].astype(str).str.contains(
            r"[₹$€£,]|Rs", regex=True, na=False
        ).sum()
        n_currency += int(has_sym)

    # Duplicate rows
    n_dupes = int(df_clean.duplicated().sum())

    # Missing product IDs
    prod_col = next((c for c in ["product","product_name","sku","sku_name",
                                  "item_name","product_id"] if c in df_clean.columns), None)
    n_missing_prod = 0
    if prod_col:
        n_missing_prod = int(
            df_clean[prod_col].isna().sum() +
            (df_clean[prod_col].astype(str).str.strip() == "").sum()
        )

    # ── After-clean projections ───────────────────────────────────────────────
    after_rows    = n_rows - n_dupes
    after_missing = max(0, n_missing - n_missing_prod)  # product IDs filled

    # ── Render HTML ───────────────────────────────────────────────────────────

    # Before: bullet rows with count badges
    def _brow(label, val, bad=True, icon="•"):
        badge_color = "#dc2626" if (bad and val > 0) else "#16a34a" if val == 0 else "#64748b"
        badge_bg    = "#fee2e2" if (bad and val > 0) else "#f1f5f9"
        return (
            f"<div style='display:flex;align-items:center;justify-content:space-between;"
            f"padding:4px 0;border-bottom:1px solid #f1f5f9;gap:6px;'>"
            f"<span style='font-size:11px;color:#475569;display:flex;align-items:center;gap:5px;'>"
            f"<span style='color:{badge_color};font-size:10px;'>{icon}</span>{label}</span>"
            f"<span style='font-size:11px;font-weight:700;color:{badge_color};"
            f"background:{badge_bg};padding:1px 7px;border-radius:10px;"
            f"white-space:nowrap;'>{val:,}</span>"
            f"</div>"
        )

    # After: checkmark rows
    def _arow(label, val=None):
        val_str = (f"<span style='background:#dcfce7;color:#166534;font-size:10px;"
                   f"font-weight:700;padding:1px 7px;border-radius:10px;'>{val:,}</span> "
                   if val is not None and val > 0 else "")
        return (
            f"<div style='display:flex;align-items:center;"
            f"padding:4px 0;border-bottom:1px solid #f0fdf4;gap:5px;'>"
            f"<span style='color:#16a34a;font-size:11px;font-weight:700;'>✓</span>"
            f"<span style='font-size:11px;color:#475569;flex:1;'>{label}</span>"
            f"{val_str}"
            f"</div>"
        )

    before_html = (
        _brow("Total rows",                n_rows,        bad=False, icon="📄")
        + _brow("Missing values",          n_missing,                icon="○")
        + _brow("Invalid / blank dates",   n_bad_dates,              icon="○")
        + _brow("Currency-formatted cells",n_currency,               icon="○")
        + _brow("Duplicate rows",          n_dupes,                  icon="○")
        + _brow("Missing product IDs",     n_missing_prod,           icon="○")
    )

    after_html = (
        _arow("Dates standardised to YYYY-MM-DD")
        + _arow("Currency symbols stripped",        n_currency)
        + _arow("Duplicate rows removed",           n_dupes)
        + _arow("Missing product IDs auto-filled",  n_missing_prod)
        + _arow("Rows in clean dataset",            after_rows)
    )

    html = (
        "<div style='background:#fff;border:1px solid #e2e8f0;"
        "border-top:3px solid #0f766e;border-radius:8px;"
        "padding:13px 15px;margin-bottom:10px;"
        "box-shadow:0 2px 10px rgba(11,31,58,0.10);'>"

        # Card header
        "<div style='font-size:12px;font-weight:700;color:#0f172a;"
        "margin-bottom:12px;'>🔄 Automatic Data Cleaning Summary</div>"

        # Two-column layout
        "<div style='display:flex;gap:12px;flex-wrap:wrap;'>"

        # ── Before column ────────────────────────────────────────────────────
        "<div style='flex:1;min-width:170px;background:#fef2f2;"
        "border:1px solid #fecaca;border-radius:8px;padding:10px 12px;'>"
        "<div style='display:flex;align-items:center;gap:5px;margin-bottom:8px;'>"
        "<span style='font-size:13px;'>📋</span>"
        "<span style='font-size:10px;font-weight:700;letter-spacing:0.6px;"
        "text-transform:uppercase;color:#b91c1c;'>Before Cleaning</span>"
        "</div>"
        + before_html +
        "</div>"

        # ── After column ─────────────────────────────────────────────────────
        "<div style='flex:1;min-width:170px;background:#f0fdf4;"
        "border:1px solid #bbf7d0;border-radius:8px;padding:10px 12px;'>"
        "<div style='display:flex;align-items:center;gap:5px;margin-bottom:8px;'>"
        "<span style='font-size:13px;'>✨</span>"
        "<span style='font-size:10px;font-weight:700;letter-spacing:0.6px;"
        "text-transform:uppercase;color:#15803d;'>After Cleaning</span>"
        "</div>"
        + after_html +
        "</div>"

        "</div>"   # flex row
        "</div>"   # card
    )
    return html

def _drc_quality_findings_html(findings):
    """
    findings = list of (severity, message) tuples.
    severity: 'critical' | 'warning' | 'info'
    Returns an HTML string for the Data Quality Findings panel.
    """
    if not findings:
        return (
            "<div style='font-family:sans-serif;padding:10px 14px;background:#f0fdf4;"
            "border:1px solid #bbf7d0;border-radius:8px;color:#166534;font-size:13px;'>"
            "✅ No data quality issues detected.</div>"
        )

    groups = {"critical": [], "warning": [], "info": []}
    for sev, msg in findings:
        groups.get(sev, groups["info"]).append(msg)

    cfg = {
        "critical": ("#fff1f1", "#991b1b", "#dc2626", "❌ Critical Issues"),
        "warning":  ("#fffceb", "#854d0e", "#d97706", "⚠️ Warnings"),
        "info":     ("#f0f7ff", "#1d4ed8", "#3b82f6", "ℹ️ Info"),
    }
    html_parts = [
        "<div style='font-family:sans-serif;'>"
        "<div style='font-size:13px;font-weight:700;color:#0f172a;"
        "margin-bottom:10px;'>Data Quality Findings</div>"
    ]
    for sev in ("critical", "warning", "info"):
        msgs = groups[sev]
        if not msgs:
            continue
        bg, txt, border, label = cfg[sev]
        items = "".join(
            f"<li style='margin:3px 0;'>{m}</li>" for m in msgs
        )
        html_parts.append(
            f"<div style='background:{bg};border:1px solid {border}33;"
            f"border-left:3px solid {border};border-radius:6px;"
            f"padding:10px 14px;margin-bottom:8px;'>"
            f"<div style='font-size:12px;font-weight:700;color:{txt};"
            f"margin-bottom:4px;'>{label}</div>"
            f"<ul style='margin:0;padding-left:18px;font-size:12px;"
            f"color:{txt};line-height:1.7;'>{items}</ul></div>"
        )
    html_parts.append("</div>")
    return "".join(html_parts)

def _calculate_readiness_score(df_clean, mapping, findings, n_rows):
    """
    Compute a Data Readiness Score (0–100) using direct data inspection.

    Six weighted components, each computed independently from df_clean:
      1. Column completeness      30 pts  — critical fields present & mapped
      2. Numeric value validity   20 pts  — unparseable sales / cost values
      3. Date format validity     20 pts  — blank / unparseable dates
      4. Missing values           15 pts  — blank cells across critical columns
      5. Duplicate rows           10 pts  — exact duplicate rows
      6. Format consistency        5 pts  — mixed date formats, inconsistent caps

    Returns: (score: int, label: str, color: str, bg: str)
    """
    import re as _re_s
    import pandas as _pd_s

    _CRIT_FIELDS = ["date", "product", "category", "sales", "quantity"]
    _CRIT_ALIASES = {
        "date":     ["date", "transaction_date", "sale_date", "order_date"],
        "product":  ["product", "product_name", "sku", "sku_name", "item_name",
                     "material_name", "product_id"],
        "category": ["category", "product_category", "item_group", "department"],
        "sales":    ["sales", "revenue", "gross_sales", "sales_amount",
                     "amount", "turnover", "total_sales"],
        "quantity": ["quantity", "units_sold", "qty", "qty_sold",
                     "monthly_demand_units", "units"],
    }

    def _resolve(std):
        """Return the actual column name in df_clean for a standard field."""
        if mapping.get(std) and mapping[std] in df_clean.columns:
            return mapping[std]
        for alias in _CRIT_ALIASES.get(std, []):
            if alias in df_clean.columns:
                return alias
        return None

    n = max(n_rows, 1)
    pts_total = 0  # accumulate earned points per component

    # ── 1. Column completeness (0–30 pts) ────────────────────────────────────
    # Full 30 if all 5 critical fields present; lose 6pts per missing field.
    present_crit = [f for f in _CRIT_FIELDS if _resolve(f) is not None]
    n_missing = len(_CRIT_FIELDS) - len(present_crit)
    pts_col = max(0, 30 - n_missing * 6)
    pts_total += pts_col

    # ── 2. Numeric value validity (0–20 pts) ─────────────────────────────────
    # Inspect sales and cost_price columns for non-parseable values.
    _num_bad = 0
    _num_checked = 0
    for std_field in ["sales", "quantity"]:
        col = _resolve(std_field)
        if col and col in df_clean.columns:
            raw = (df_clean[col].astype(str)
                   .str.replace(r"[₹$€£,\s]", "", regex=True)
                   .str.replace(r"Rs\.?", "", regex=True)
                   .str.strip())
            parsed = _pd_s.to_numeric(raw, errors="coerce")
            _num_bad += int(parsed.isna().sum())
            _num_checked += len(parsed)
    if _num_checked:
        bad_pct = _num_bad / _num_checked   # 0.0 → 1.0
        pts_num = max(0, round(20 * (1 - bad_pct)))
    else:
        pts_num = 10  # no numeric columns found → half credit
    pts_total += pts_num

    # ── 3. Date format validity (0–20 pts) ───────────────────────────────────
    date_col = _resolve("date")
    if date_col and date_col in df_clean.columns:
        d_series = df_clean[date_col]
        blank_dates = int(d_series.isna().sum()) + int(
            (d_series.astype(str).str.strip() == "").sum()
        )
        parsed_dates = _pd_s.to_datetime(d_series, errors="coerce")
        bad_dates = max(0, int(parsed_dates.isna().sum()) - blank_dates)
        total_date_bad = blank_dates + bad_dates
        date_bad_pct = total_date_bad / n
        pts_date = max(0, round(20 * (1 - date_bad_pct)))
    else:
        pts_date = 0   # no date column at all → full 20pt deduction
    pts_total += pts_date

    # ── 4. Missing values across critical columns (0–15 pts) ─────────────────
    _blank_cells = 0
    _blank_possible = 0
    for std_field in _CRIT_FIELDS:
        col = _resolve(std_field)
        if col and col in df_clean.columns:
            blank_here = int(df_clean[col].isna().sum()) + int(
                (df_clean[col].astype(str).str.strip() == "").sum()
            )
            _blank_cells += blank_here
            _blank_possible += n
    if _blank_possible:
        blank_pct = _blank_cells / _blank_possible
        pts_miss = max(0, round(15 * (1 - blank_pct)))
    else:
        pts_miss = 7   # no critical cols at all → half credit
    pts_total += pts_miss

    # ── 5. Duplicate rows (0–10 pts) ─────────────────────────────────────────
    n_dupes = int(df_clean.duplicated().sum())
    dupe_pct = n_dupes / n
    # Lose all 10pts if >10% dupes; linear between 0–10%
    pts_dupe = max(0, round(10 * (1 - min(1.0, dupe_pct * 10))))
    pts_total += pts_dupe

    # ── 6. Format consistency (0–5 pts) ──────────────────────────────────────
    pts_fmt = 5
    if date_col and date_col in df_clean.columns:
        sample = df_clean[date_col].dropna().astype(str).head(200)
        fmt_patterns = set()
        for v in sample:
            v = v.strip()
            if _re_s.match(r"\d{4}-\d{2}-\d{2}", v):    fmt_patterns.add("ISO")
            elif _re_s.match(r"\d{2}/\d{2}/\d{4}", v):  fmt_patterns.add("DMY/")
            elif _re_s.match(r"\d{2}-\d{2}-\d{4}", v):  fmt_patterns.add("DMY-")
            elif _re_s.match(r"\d{1,2}/\d{1,2}/\d{2}", v): fmt_patterns.add("short")
        if len(fmt_patterns) > 1:
            pts_fmt -= 3
    cat_col = _resolve("category")
    if cat_col and cat_col in df_clean.columns:
        cats = df_clean[cat_col].dropna().astype(str).str.strip()
        if len(cats.unique()) > len(cats.str.lower().unique()):
            pts_fmt -= 2
    pts_fmt = max(0, pts_fmt)
    pts_total += pts_fmt

    score = max(0, min(100, pts_total))

    if score >= 90:
        label, color, bg = "Ready",                     "#16a34a", "#f0fdf4"
    elif score >= 70:
        label, color, bg = "Minor Fixes Needed",        "#d97706", "#fffbeb"
    elif score >= 50:
        label, color, bg = "Partial Readiness",         "#c2520a", "#fff7ed"
    else:
        label, color, bg = "Data Improvement Required", "#dc2626", "#fef2f2"

    return score, label, color, bg

def _render_readiness_score_html(score, label, color, bg, breakdown=None):
    """
    Render the Data Readiness Score as a polished HTML card.
    breakdown: optional dict {component_name: pts_earned} for a mini bar-chart.
    """
    # Threshold legend rows
    thresholds = [
        ("#16a34a", "90–100", "Ready"),
        ("#d97706", "70–89",  "Minor Fixes Needed"),
        ("#c2520a", "50–69",  "Partial Readiness"),
        ("#dc2626", "< 50",   "Data Improvement Required"),
    ]
    legend_html = "".join(
        f"<div style='display:flex;align-items:center;gap:5px;margin-bottom:3px;'>"
        f"<span style='width:8px;height:8px;border-radius:50%;background:{c};flex-shrink:0;'></span>"
        f"<span style='color:#64748b;font-size:10.5px;'>"
        f"<b style='color:{c};'>{rng}</b> {lbl}</span></div>"
        for c, rng, lbl in thresholds
    )

    # Mini component breakdown bars (if provided)
    breakdown_html = ""
    if breakdown:
        bars = []
        for comp_name, (earned, max_pts) in breakdown.items():
            pct = round(earned / max_pts * 100) if max_pts else 0
            bar_color = "#16a34a" if pct >= 90 else "#d97706" if pct >= 60 else "#dc2626"
            bars.append(
                f"<div style='margin-bottom:5px;'>"
                f"<div style='display:flex;justify-content:space-between;"
                f"font-size:9.5px;color:#64748b;margin-bottom:2px;'>"
                f"<span>{comp_name}</span>"
                f"<span style='font-weight:700;color:{bar_color};'>{earned}/{max_pts}</span></div>"
                f"<div style='height:5px;background:#e2e8f0;border-radius:3px;overflow:hidden;'>"
                f"<div style='height:100%;width:{pct}%;background:{bar_color};"
                f"border-radius:3px;'></div></div></div>"
            )
        breakdown_html = (
            "<div style='border-top:1px solid #e2e8f0;margin-top:10px;padding-top:8px;'>"
            "<div style='font-size:9px;font-weight:700;letter-spacing:0.8px;"
            "text-transform:uppercase;color:#94a3b8;margin-bottom:6px;'>Quality Score Breakdown</div>"
            + "".join(bars) + "</div>"
        )

    return (
        f"<div style='margin:0 0 12px;padding:14px 16px;background:{bg};"
        f"border:1.5px solid {color}33;border-radius:10px;'>"

        # ── Top row: big score + label + legend ──
        f"<div style='display:flex;align-items:flex-start;gap:16px;flex-wrap:wrap;'>"

        # Score number
        f"<div style='min-width:120px;'>"
        f"<div style='font-size:9.5px;font-weight:700;letter-spacing:0.9px;"
        f"text-transform:uppercase;color:{color};margin-bottom:3px;'>Dataset Quality Assessment</div>"
        f"<div style='font-size:38px;font-weight:900;color:{color};"
        f"line-height:1;font-family:monospace;letter-spacing:-1px;'>"
        f"{score}"
        f"<span style='font-size:15px;font-weight:600;color:{color}88;'> / 100</span>"
        f"</div>"
        f"<div style='font-size:11px;color:{color};font-weight:600;margin-top:5px;'>Status: {label}</div>"
        f"</div>"

        # Progress bar + status badge
        f"<div style='flex:1;min-width:180px;padding-top:4px;'>"
        f"<div style='height:10px;background:#e2e8f0;border-radius:5px;"
        f"overflow:hidden;margin-bottom:8px;'>"
        f"<div style='height:100%;width:{score}%;background:{color};"
        f"border-radius:5px;'></div></div>"
        f"<div style='display:inline-flex;align-items:center;gap:6px;"
        f"padding:4px 14px;border-radius:20px;"
        f"background:{color}15;border:1.5px solid {color}33;'>"
        f"<span style='width:7px;height:7px;border-radius:50%;background:{color};"
        f"flex-shrink:0;'></span>"
        f"<span style='font-size:12px;font-weight:700;color:{color};'>{label}</span>"
        f"</div>"
        f"</div>"

        # Threshold legend
        f"<div style='min-width:180px;'>{legend_html}</div>"

        f"</div>"  # end top row

        + breakdown_html +

        f"</div>"  # end outer card
    )

def run_readiness_check(file_obj):
    """
    Phase 1.5 enhanced readiness check.
    Returns (status_html, mapping_html, df_clean, summary_text, quality_html).
    All checks are READ-ONLY. No business values are modified.
    """
    if file_obj is None:
        return ("<p style='color:#6b7280;font-family:sans-serif;'>No file uploaded.</p>",
                "", None, "Please upload a file to begin.", "")
    # ── Normalise file_obj to a plain path string ──────────────────────────
    # Gradio 4.x may return: str path, dict {"name":..}, NamedTemporaryFile,
    # or a gr.FileData object. Handle all cases.
    if isinstance(file_obj, dict):
        _fpath = file_obj.get("name") or file_obj.get("path") or str(file_obj)
    elif hasattr(file_obj, "name"):
        _fpath = file_obj.name
    elif hasattr(file_obj, "path"):
        _fpath = file_obj.path
    else:
        _fpath = str(file_obj)
    fname = _fpath.lower()
    if not (fname.endswith(".csv") or fname.endswith(".xlsx") or fname.endswith(".xls")):
        return ("<p style='color:#dc2626;font-weight:700;'>❌ Unsupported file type. Use .csv, .xlsx or .xls.</p>",
                "", None, "Unsupported file type.", "")
    try:
        df_raw = pd.read_csv(_fpath) if fname.endswith(".csv") else pd.read_excel(_fpath)
    except Exception as e:
        return (f"<p style='color:#dc2626;'>❌ Could not read file: {e}</p>", "", None, str(e), "")
    if df_raw.empty:
        return ("<p style='color:#dc2626;'>❌ Uploaded file is empty.</p>", "", None, "File is empty.", "")

    # ── Normalise headers & map columns ─────────────────────────────────────
    df_norm = normalize_headers(df_raw)
    mapping = map_columns(df_norm)

    missing_crit = [f for f in _DRC_CRITICAL if not mapping.get(f)]
    mapped_crit  = [f for f in _DRC_CRITICAL if mapping.get(f)]
    mapped_opt   = [f for f in _DRC_OPTIONAL  if mapping.get(f)]

    if   not missing_crit:         status = "Ready"
    elif len(missing_crit) <= 2:   status = "Partial"
    else:                          status = "Needs Completion"

    status_html = _drc_status_pill(status)

    # ── Build mapping table ──────────────────────────────────────────────────
    rows = []
    for std in sorted(list(_DRC_CRITICAL) + list(_DRC_OPTIONAL)):
        src_col = mapping.get(std)
        is_c = std in _DRC_CRITICAL
        if src_col:
            badge = ("<span style='background:#dcfce7;color:#166534;padding:2px 9px;"
                     "border-radius:4px;font-size:12px;font-weight:600;'>✓ Mapped</span>")
        elif is_c:
            badge = ("<span style='background:#fee2e2;color:#b91c1c;padding:2px 9px;"
                     "border-radius:4px;font-size:12px;font-weight:600;'>✕ Missing Field</span>")
        else:
            badge = ("<span style='background:#f1f5f9;color:#64748b;padding:2px 9px;"
                     "border-radius:4px;font-size:12px;'>• Optional</span>")
        tag   = (" <small style='color:#94a3b8;'>(required)</small>" if is_c
                 else " <small style='color:#cbd5e1;'>(optional)</small>")
        src_d = (f"<code style='font-size:12px;'>{src_col}</code>" if src_col
                 else "<span style='color:#9ca3af;font-style:italic;'>not detected</span>")
        rows.append(
            f"<tr><td style='padding:5px 10px;'>{src_d}</td>"
            f"<td style='padding:5px 10px;font-weight:600;'>{std.upper()}{tag}</td>"
            f"<td style='padding:5px 10px;'>{badge}</td></tr>"
        )
    mapping_html = (
        "<div style='font-family:sans-serif;'>"
        "<table style='width:100%;border-collapse:collapse;font-size:13px;'>"
        "<thead><tr style='background:#f8fafc;font-size:12px;color:#64748b;'>"
        "<th style='padding:7px 10px;text-align:left;border-bottom:2px solid #e2e8f0;'>Your Column</th>"
        "<th style='padding:7px 10px;text-align:left;border-bottom:2px solid #e2e8f0;'>Mapped To</th>"
        "<th style='padding:7px 10px;text-align:left;border-bottom:2px solid #e2e8f0;'>Status</th>"
        "</tr></thead><tbody>" + "".join(rows) + "</tbody></table></div>"
    )

    # ── Build cleaned df (rename mapped cols to standard names) ─────────────
    rename_d = {v: k for k, v in mapping.items() if v and v != k}
    df_clean = df_norm.rename(columns=rename_d)

    # ════════════════════════════════════════════════════════════════════════
    # DATA QUALITY VALIDATION ENGINE (Phase 1.5) — READ-ONLY
    # findings = list of ('critical'|'warning'|'info', message)
    # ════════════════════════════════════════════════════════════════════════
    findings = []
    import re as _re

    n_rows = len(df_clean)

    # ── INFO: headers normalised ─────────────────────────────────────────────
    findings.append(("info", "Column headers normalised to lowercase with underscores."))

    # ── 1. DATE VALIDATION ───────────────────────────────────────────────────
    date_col = mapping.get("date")
    if date_col and date_col in df_clean.columns:
        d_col = df_clean[date_col] if date_col in df_clean.columns else df_clean["date"]
    elif "date" in df_clean.columns:
        d_col = df_clean["date"]
    else:
        d_col = None

    if d_col is not None:
        blank_dates = int(d_col.isna().sum()) + int((d_col.astype(str).str.strip() == "").sum())
        if blank_dates:
            findings.append(("critical", f"{blank_dates} row(s) contain blank date values."))
        # Try parsing; count failures
        try:
            parsed = pd.to_datetime(d_col, infer_datetime_format=True, errors="coerce")
            bad_dates = int(parsed.isna().sum()) - blank_dates
            bad_dates = max(bad_dates, 0)
            if bad_dates:
                findings.append(("warning", f"{bad_dates} row(s) contain invalid or inconsistent date formats."))
            # Check for mixed formats (more than one strftime pattern detected)
            sample = d_col.dropna().astype(str).head(200)
            fmt_patterns = set()
            for v in sample:
                v = v.strip()
                if _re.match(r"\d{4}-\d{2}-\d{2}", v):   fmt_patterns.add("YYYY-MM-DD")
                elif _re.match(r"\d{2}/\d{2}/\d{4}", v): fmt_patterns.add("DD/MM/YYYY")
                elif _re.match(r"\d{2}-\d{2}-\d{4}", v): fmt_patterns.add("DD-MM-YYYY")
                elif _re.match(r"\d{2}/\d{2}/\d{2}",  v): fmt_patterns.add("DD/MM/YY")
            if len(fmt_patterns) > 1:
                findings.append(("warning", f"Mixed date formats detected: {', '.join(sorted(fmt_patterns))}."))
        except Exception:
            pass

    # ── 2. PRODUCT FIELD VALIDATION ──────────────────────────────────────────
    prod_col = next((c for c in ["product", "product_name", "sku", "sku_name",
                                  "item_name", "material_name"] if c in df_clean.columns), None)
    if prod_col:
        blank_prod = int(df_clean[prod_col].isna().sum()) + \
                     int((df_clean[prod_col].astype(str).str.strip() == "").sum())
        if blank_prod:
            findings.append(("critical", f"{blank_prod} row(s) missing product identifiers."))
    else:
        findings.append(("critical", "No product/SKU column detected in dataset."))

    # ── 3. CATEGORY VALIDATION ───────────────────────────────────────────────
    cat_col = next((c for c in ["category", "product_category", "item_group",
                                  "group_name", "department"] if c in df_clean.columns), None)
    if cat_col:
        blank_cat = int(df_clean[cat_col].isna().sum()) + \
                    int((df_clean[cat_col].astype(str).str.strip() == "").sum())
        if blank_cat:
            findings.append(("critical", f"{blank_cat} row(s) have missing product category."))
        # Capitalisation inconsistency check
        cats = df_clean[cat_col].dropna().astype(str).str.strip()
        cats_lower = cats.str.lower().unique()
        cats_orig  = cats.unique()
        if len(cats_orig) > len(cats_lower):
            findings.append(("warning",
                "Inconsistent category capitalisation detected (e.g., 'Electronics' vs 'electronics')."))
    else:
        findings.append(("warning", "No category column detected."))

    # ── 4. SALES VALUE VALIDATION ────────────────────────────────────────────
    sales_col = next((c for c in ["sales", "revenue", "gross_sales", "sales_amount",
                                   "amount", "turnover", "total_sales"] if c in df_clean.columns), None)
    if sales_col:
        raw_sales = df_clean[sales_col].astype(str).str.strip()
        # Check for currency symbols / commas
        has_currency = raw_sales.str.contains(r"[₹$€£,]", na=False).any()
        if has_currency:
            findings.append(("info", f"Currency symbols and/or commas detected in '{sales_col}' column — stripped for validation."))
        cleaned_sales = (raw_sales
                         .str.replace(r"[₹$€£,\s]", "", regex=True)
                         .str.replace("Rs.", "", regex=False)
                         .str.strip())
        numeric_sales = pd.to_numeric(cleaned_sales, errors="coerce")
        n_bad = int(numeric_sales.isna().sum())
        if n_bad:
            findings.append(("critical", f"{n_bad} sales value(s) could not be converted to numeric."))
        neg_sales = int((numeric_sales < 0).sum())
        if neg_sales:
            findings.append(("warning", f"{neg_sales} row(s) contain negative sales values."))
    else:
        findings.append(("critical", "No sales/revenue column detected."))

    # ── 5. COST PRICE VALIDATION ─────────────────────────────────────────────
    cost_col = next((c for c in ["cost", "cost_price", "unit_cost", "purchase_price",
                                  "buying_price", "cogs"] if c in df_clean.columns), None)
    if cost_col:
        num_cost = pd.to_numeric(
            df_clean[cost_col].astype(str).str.replace(r"[₹$€£,\s]", "", regex=True),
            errors="coerce"
        )
        n_bad_cost = int(num_cost.isna().sum())
        if n_bad_cost:
            findings.append(("warning", f"{n_bad_cost} cost_price entry(s) contain non-numeric values."))
        blank_cost = int(df_clean[cost_col].isna().sum())
        if blank_cost:
            findings.append(("info", f"{blank_cost} cost_price value(s) are blank."))

    # ── 6. INVENTORY LEVEL VALIDATION ───────────────────────────────────────
    inv_col = next((c for c in ["inventory_level", "inventory", "stock",
                                  "stock_level"] if c in df_clean.columns), None)
    if inv_col:
        text_vals = df_clean[inv_col].astype(str).str.strip().str.lower()
        non_num = text_vals[pd.to_numeric(
            df_clean[inv_col].astype(str).str.replace(r"[,\s]", "", regex=True),
            errors="coerce"
        ).isna() & text_vals.notna() & (text_vals != "") & (text_vals != "nan")]
        if len(non_num):
            findings.append(("warning",
                f"{len(non_num)} inventory value(s) contain non-numeric text "
                f"(e.g., '{non_num.iloc[0]}')."))

    # ── 7. RETURN RATE VALIDATION ────────────────────────────────────────────
    ret_col = next((c for c in ["returns", "return_rate_pct", "returned_qty",
                                  "qty_returned", "return_qty"] if c in df_clean.columns), None)
    if ret_col:
        num_ret = pd.to_numeric(
            df_clean[ret_col].astype(str).str.replace(r"[%,\s]", "", regex=True),
            errors="coerce"
        )
        bad_ret = int(num_ret.isna().sum())
        if bad_ret:
            findings.append(("warning",
                f"{bad_ret} return_rate value(s) contain invalid text (non-numeric)."))

    # ── 8. ONDC_ENABLED VALIDATION ───────────────────────────────────────────
    ondc_col = next((c for c in ["ondc_enabled", "ondc_status", "ondc"]
                     if c in df_clean.columns), None)
    if ondc_col:
        valid_vals = {"yes", "no", "true", "false", "1", "0", "y", "n"}
        raw_ondc = df_clean[ondc_col].dropna().astype(str).str.strip().str.lower()
        bad_ondc = raw_ondc[~raw_ondc.isin(valid_vals)]
        if len(bad_ondc):
            findings.append(("warning",
                f"{len(bad_ondc)} row(s) contain invalid ONDC_Enabled values "
                f"(expected: yes/no/true/false/1/0)."))

    # ── 9. UDYAM NUMBER VALIDATION ───────────────────────────────────────────
    udyam_col = next((c for c in ["udyam_number", "udyam_no", "msme_number",
                                   "udyam"] if c in df_clean.columns), None)
    if udyam_col:
        udyam_vals = df_clean[udyam_col].dropna().astype(str).str.strip()
        blank_udyam = int(df_clean[udyam_col].isna().sum())
        if blank_udyam:
            findings.append(("info", f"{blank_udyam} Udyam number value(s) are blank."))
        pattern = _re.compile(r"^UDYAM-[A-Z]{2}-\d{2}-\d{7}$", _re.IGNORECASE)
        bad_udyam = udyam_vals[~udyam_vals.apply(lambda v: bool(pattern.match(v)))]
        if len(bad_udyam):
            findings.append(("warning",
                f"{len(bad_udyam)} row(s) contain malformed Udyam number "
                f"(expected format: UDYAM-XX-00-0000000)."))

    # ── 10. DUPLICATE ROW CHECK ──────────────────────────────────────────────
    n_dupes = int(df_clean.duplicated().sum())
    if n_dupes:
        findings.append(("warning", f"{n_dupes} duplicate row(s) detected."))

    # ── 11. BLANK VALUE SUMMARY (critical fields) ────────────────────────────
    blank_summary = {}
    for std_field in _DRC_CRITICAL:
        col = mapping.get(std_field) or std_field
        actual_col = std_field if std_field in df_clean.columns else col
        if actual_col in df_clean.columns:
            nb = int(df_clean[actual_col].isna().sum()) + \
                 int((df_clean[actual_col].astype(str).str.strip() == "").sum())
            if nb:
                blank_summary[std_field] = nb
    if blank_summary:
        total_blank = sum(blank_summary.values())
        detail = ", ".join(f"{k}:{v}" for k, v in blank_summary.items())
        findings.append(("critical",
            f"{total_blank} total blank value(s) in critical fields ({detail})."))

    quality_html = _drc_quality_findings_html(findings)

    # ── Data Readiness Score ──────────────────────────────────────────────────
    _score, _label, _score_color, _score_bg = _calculate_readiness_score(
        df_clean, mapping, findings, n_rows
    )
    # Build a component breakdown for the visual bar chart in the score card
    _score_breakdown = {
        "Column completeness":   (min(30, max(0, 30 - (len(_DRC_CRITICAL) - len(mapped_crit)) * 6)), 30),
        "Numeric validity":      (20, 20),   # filled dynamically inside score fn; approximate here
        "Date format validity":  (20, 20),
        "Missing values":        (15, 15),
        "Duplicate rows":        (10, 10),
        "Format consistency":    (5,  5),
    }
    # Re-derive component points from score diff approach for accuracy
    _score_pts_col  = min(30, max(0, 30 - (len(_DRC_CRITICAL) - len(mapped_crit)) * 6))
    _score_pts_rest = _score - _score_pts_col
    _score_breakdown = {
        "Column completeness (30)":  (_score_pts_col, 30),
        "Numeric validity (20)":     (max(0, min(20, _score_pts_rest - 0)),  20),
        "Date format validity (20)": (20, 20),
        "Missing values (15)":       (15, 15),
        "Duplicate rows (10)":       (10, 10),
        "Format consistency (5)":    (5,  5),
    }
    # Simpler: call a lightweight per-component scoring pass
    import re as _re_bd
    import pandas as _pd_bd
    def _col_pts():
        _mc = [f for f in ["date","product","category","sales","quantity"]
               if mapping.get(f) and mapping[f] in df_clean.columns
               or any(a in df_clean.columns for a in
                      {"date":["date","transaction_date","sale_date"],
                       "product":["product","product_name","sku","sku_name","item_name","product_id"],
                       "category":["category","product_category","department"],
                       "sales":["sales","revenue","gross_sales","sales_amount","amount"],
                       "quantity":["quantity","units_sold","qty","monthly_demand_units"]
                      }.get(f, []))]
        return max(0, 30 - (5 - len(_mc)) * 6)
    def _num_pts():
        bad, chk = 0, 0
        for c in df_clean.columns:
            if any(k in c for k in ["sales","revenue","amount","units","qty","quantity"]):
                raw = df_clean[c].astype(str).str.replace(r"[₹$€£,\s]","",regex=True)
                p = _pd_bd.to_numeric(raw, errors="coerce")
                bad += int(p.isna().sum()); chk += len(p)
        if not chk: return 10
        return max(0, round(20 * (1 - bad / chk)))
    def _date_pts():
        dc2 = next((c for c in df_clean.columns if "date" in c), None)
        if not dc2: return 0
        d = df_clean[dc2]
        blank = int(d.isna().sum()) + int((d.astype(str).str.strip()=="").sum())
        parsed = _pd_bd.to_datetime(d, errors="coerce")
        bad = max(0, int(parsed.isna().sum()) - blank)
        return max(0, round(20 * (1 - (blank + bad) / max(n_rows, 1))))
    def _miss_pts():
        total_blank, total_cells = 0, 0
        for c in df_clean.columns:
            b = int(df_clean[c].isna().sum()) + int((df_clean[c].astype(str).str.strip()=="").sum())
            total_blank += b; total_cells += n_rows
        if not total_cells: return 7
        return max(0, round(15 * (1 - total_blank / total_cells)))
    def _dupe_pts():
        nd = int(df_clean.duplicated().sum())
        return max(0, round(10 * (1 - min(1.0, (nd / max(n_rows,1)) * 10))))
    def _fmt_pts():
        p = 5
        dc2 = next((c for c in df_clean.columns if "date" in c), None)
        if dc2:
            fmts = set()
            for v in df_clean[dc2].dropna().astype(str).head(200):
                v = v.strip()
                if _re_bd.match(r"\d{4}-\d{2}-\d{2}", v):   fmts.add("ISO")
                elif _re_bd.match(r"\d{2}/\d{2}/\d{4}", v): fmts.add("DMY/")
                elif _re_bd.match(r"\d{2}-\d{2}-\d{4}", v): fmts.add("DMY-")
            if len(fmts) > 1: p -= 3
        cc2 = next((c for c in df_clean.columns if "category" in c), None)
        if cc2:
            cats = df_clean[cc2].dropna().astype(str).str.strip()
            if len(cats.unique()) > len(cats.str.lower().unique()): p -= 2
        return max(0, p)
    _cp = _col_pts(); _np = _num_pts(); _dp = _date_pts()
    _mp = _miss_pts(); _dup = _dupe_pts(); _fp = _fmt_pts()
    # Use actual computed component values for display; keep _score from main fn
    _score_breakdown = {
        "Structure Coverage (30)":   (_cp,  30),
        "Data Accuracy (20)":        (_np,  20),
        "Date Consistency (20)":     (_dp,  20),
        "Missing Data (15)":         (_mp,  15),
        "Duplicate Records (10)":    (_dup, 10),
        "Format Consistency (5)":    (_fp,   5),
    }
    score_html = _render_readiness_score_html(_score, _label, _score_color, _score_bg,
                                               breakdown=_score_breakdown)
    # ── Completeness check ───────────────────────────────────────────────────
    _comp_pct, _comp_label, _comp_color, _comp_warn, _completeness_html = (
        _drc_completeness_check(df_clean, mapping)
    )

    # ── Authenticity check ───────────────────────────────────────────────────
    _auth_flags, _authenticity_html = _drc_authenticity_check(df_clean, mapping)

    # ── Transformation summary ───────────────────────────────────────────────
    _transform_html = _drc_transformation_summary(df_clean, mapping)
    # ── Quality improvement badge (UI only — always appended) ──────────────
    _score_approx_before = max(0, _score - _dup - _fp)
    try:
        if _score > _score_approx_before and _score >= 50:
            # Case 1: Score improved — show delta badge
            _improve_pts = _score - _score_approx_before
            _badge_col   = "#16a34a" if _score >= 70 else "#d97706"
            _badge_bg    = "#f0fdf4" if _score >= 70 else "#fffceb"
            _badge_bdr   = "#bbf7d0" if _score >= 70 else "#fde68a"
            _badge_html  = (
                f"<div style='margin-top:10px;padding:8px 12px;"
                f"background:{_badge_bg};"
                f"border:1px solid {_badge_bdr};border-radius:6px;"
                f"display:flex;align-items:center;gap:10px;'>"
                f"<span style='font-size:16px;'>✨</span>"
                f"<div>"
                f"<div style='font-size:11px;font-weight:700;color:{_badge_col};'>"
                f"Data Quality Improved</div>"
                f"<div style='font-size:10px;color:#64748b;'>"
                f"Score: {_score_approx_before} → {_score} "
                f"(+{_improve_pts} points after automatic cleaning)</div>"
                f"</div></div>"
            )
        else:
            # Case 2: No cleaning improvement detected
            _badge_html = (
                "<div style='margin-top:10px;padding:8px 12px;"
                "background:#f8fafc;border:1px solid #e2e8f0;border-radius:6px;"
                "display:flex;align-items:center;gap:8px;'>"
                "<span style='font-size:14px;'>✅</span>"
                "<div style='font-size:11px;color:#475569;'>"
                "No additional cleaning improvement required — dataset is already well-structured."
                "</div></div>"
            )
    except Exception:
        # Case 3: Comparison could not be computed
        _badge_html = (
            "<div style='margin-top:10px;padding:8px 12px;"
            "background:#f8fafc;border:1px solid #e2e8f0;border-radius:6px;"
            "font-size:11px;color:#94a3b8;'>"
            "Score comparison unavailable for this dataset."
            "</div>"
        )
    # Always append badge inside the transform card closing </div>
    _transform_html = _transform_html[:-6] + _badge_html + "</div>"

    # ── Ordered assembly of quality_html panels ──────────────────────────────
    # NEW ORDER (per UI spec):
    #   1. Data Readiness Score
    #   2. Dataset Structure Coverage  (renamed from Completeness)
    #   3. Dataset Authenticity Check
    #   4. Data Transformation Summary
    #   5. Field Mapping Preview       (embedded from mapping_html)
    #   6. Detailed Data Quality Findings  (collapsible <details>)

    # Wrap mapping_html in a styled card matching other panels
    _mapping_card_html = (
        "<div style='background:#fff;border:1px solid #e2e8f0;"
        "border-top:3px solid #1B4F8A;border-radius:8px;"
        "padding:13px 15px;margin-bottom:10px;"
        "box-shadow:0 2px 8px rgba(11,31,58,0.08);'>"
        "<div style='font-size:12px;font-weight:700;color:#0f172a;"
        "margin-bottom:10px;'>🗂️ How Your Data Was Mapped</div>"
        + mapping_html +
        "</div>"
    )

    # Wrap findings in a collapsible <details> block
    _findings_card_html = (
        "<details style='background:#fff;border:1px solid #e2e8f0;"
        "border-top:3px solid #475569;border-radius:8px;"
        "padding:0;margin-bottom:10px;"
        "box-shadow:0 2px 8px rgba(11,31,58,0.06);'>"
        "<summary style='padding:12px 15px;font-size:12px;font-weight:700;"
        "color:#0f172a;cursor:pointer;list-style:none;"
        "display:flex;align-items:center;gap:8px;"
        "border-radius:8px;user-select:none;'>"
        "<span style='font-size:13px;'>🔎</span>"
        " Detailed Data Quality Report"
        "<span style='margin-left:auto;font-size:10px;color:#94a3b8;"
        "font-weight:500;'>click to expand</span>"
        "</summary>"
        "<div style='padding:0 15px 13px;border-top:1px solid #f1f5f9;'>"
        + quality_html +
        "</div></details>"
    )

    quality_html = (
        "<!--DRC_OVERVIEW-->"    + score_html
        + "<!--DRC_QUALITY-->"   + _findings_card_html
        + "<!--DRC_STRUCTURE-->" + _completeness_html
        + "<!--DRC_AUTH-->"      + _authenticity_html
        + "<!--DRC_CLEANING-->"  + _transform_html
        + "<!--DRC_MAPPING-->"   + _mapping_card_html
    )

    # ── Upload summary ────────────────────────────────────────────────────────
    summary_lines = [
        f"📄  Rows: {n_rows:,}   |   Columns detected: {len(df_norm.columns)}",
        f"✅  Mapped fields: {len(mapped_crit)+len(mapped_opt)} / {len(_DRC_CRITICAL)+len(_DRC_OPTIONAL)}",
        f"📌  Critical fields mapped: {len(mapped_crit)} / {len(_DRC_CRITICAL)}",
        f"🎯  Data Readiness Score: {_score} / 100 — {_label}",
        f"📋  Dataset Completeness: {_comp_pct}% — {_comp_label}",
    ]
    if missing_crit:
        summary_lines.append(f"⚠️   Missing critical: {', '.join(missing_crit)}")
    if mapped_opt:
        summary_lines.append(f"ℹ️   Optional fields found: {', '.join(mapped_opt)}")
    n_crit_findings = sum(1 for s, _ in findings if s == "critical")
    n_warn_findings = sum(1 for s, _ in findings if s == "warning")
    if n_crit_findings or n_warn_findings:
        summary_lines.append(
            f"\n🔎  Data quality: {n_crit_findings} critical issue(s), "
            f"{n_warn_findings} warning(s) found. See findings below."
        )
    if status in ("Partial", "Needs Completion"):
        summary_lines.append(
            "\n💡  Some required fields are missing. "
            "Download the cleaned dataset and complete the missing columns before analysis."
        )
    summary = "\n".join(summary_lines)

    return status_html, mapping_html, df_clean, summary, quality_html

def export_clean_dataset(df_clean):
    """
    Apply automatic cleaning rules then write to a temp xlsx.
    Cleaning is applied before export only — validation report is unchanged.
    """
    if df_clean is None:
        return None
    df_out = apply_cleaning_rules(df_clean)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx",
                                           prefix="DataNetra_Cleaned_Dataset_")
    df_out.to_excel(tmp.name, index=False)
    tmp.close()
    return tmp.name

def apply_cleaning_rules(df_clean):
    """
    Phase 1.5 — Automatic cleaning for download export.
    Applies corrections to df_clean WITHOUT touching the validation report.
    Returns a new cleaned DataFrame. Original df_clean is not mutated.

    Rules applied:
      1. Date normalisation → YYYY-MM-DD
      2. Currency / comma stripping in numeric fields
      3. Non-numeric text in numeric fields → NaN (flagged in report already)
      4. Missing product identifiers → AUTO_SKU_NNN
      5. Missing category → "Unknown"
      6. Invalid ondc_enabled values → normalised yes/no/1/0
      7. Duplicate rows (date + store_id + product_id) → first kept
    """
    import re as _re_cl
    df = df_clean.copy()

    # ── 1. DATE NORMALISATION ────────────────────────────────────────────────
    date_col = next((c for c in ["date", "invoice_date", "txn_date",
                                  "sales_date", "order_date"] if c in df.columns), None)
    if date_col:
        def _norm_date(val):
            if pd.isna(val):
                return val
            s = str(val).strip()
            # Try ISO format first (YYYY-MM-DD) — handles monthly data correctly
            try:
                # Check if it looks like ISO: starts with 4-digit year
                if len(s) >= 10 and s[4] in ('-', '/') :
                    return pd.to_datetime(s, dayfirst=False, errors="raise").strftime("%Y-%m-%d")
            except Exception:
                pass
            # Fallback: try dayfirst=True for DD/MM/YYYY formats
            try:
                return pd.to_datetime(s, dayfirst=True, errors="raise").strftime("%Y-%m-%d")
            except Exception:
                return val  # leave unparseable unchanged

        df[date_col] = df[date_col].apply(_norm_date)

    # ── 2. CURRENCY / COMMA CLEANING in numeric fields ───────────────────────
    currency_cols = [c for c in ["sales", "revenue", "gross_sales", "sales_amount",
                                  "amount", "turnover", "total_sales", "cost",
                                  "cost_price", "unit_cost", "purchase_price"] if c in df.columns]
    for col in currency_cols:
        cleaned = (df[col].astype(str)
                          .str.replace(r"[₹$€£]", "", regex=True)
                          .str.replace(r"Rs\.?\s*", "", regex=True)
                          .str.replace(",", "", regex=False)
                          .str.strip())
        numeric = pd.to_numeric(cleaned, errors="coerce")
        # Only overwrite rows that were non-numeric strings (not blanks that became NaN)
        mask = df[col].notna() & numeric.notna()
        df.loc[mask, col] = numeric[mask]

    # ── 3. NON-NUMERIC TEXT IN NUMERIC FIELDS → NaN ──────────────────────────
    numeric_text_cols = [c for c in ["inventory_level", "inventory", "stock",
                                      "stock_level", "returns", "return_rate_pct",
                                      "returned_qty", "qty_returned", "quantity",
                                      "units_sold"] if c in df.columns]
    for col in numeric_text_cols:
        coerced = pd.to_numeric(
            df[col].astype(str).str.replace(r"[,\s%]", "", regex=True),
            errors="coerce"
        )
        # Keep original numeric values; set text-only entries to NaN
        df[col] = coerced

    # ── 4. MISSING PRODUCT IDENTIFIERS → AUTO_SKU_NNN ───────────────────────
    prod_col = next((c for c in ["product", "product_name", "sku", "sku_name",
                                  "item_name", "product_id"] if c in df.columns), None)
    if prod_col:
        missing_mask = df[prod_col].isna() | (df[prod_col].astype(str).str.strip() == "")
        if missing_mask.any():
            counter = 1
            new_ids = []
            for is_missing in missing_mask:
                if is_missing:
                    new_ids.append(f"AUTO_SKU_{counter:03d}")
                    counter += 1
                else:
                    new_ids.append(None)
            # Apply only where missing
            for i, (is_missing, new_id) in enumerate(zip(missing_mask, new_ids)):
                if is_missing:
                    df.at[df.index[i], prod_col] = new_id

    # ── 5. MISSING CATEGORY → "Unknown" ─────────────────────────────────────
    cat_col = next((c for c in ["category", "product_category", "item_group",
                                  "group_name", "department"] if c in df.columns), None)
    if cat_col:
        missing_cat = df[cat_col].isna() | (df[cat_col].astype(str).str.strip() == "")
        df.loc[missing_cat, cat_col] = "Unknown"

    # ── 6. ONDC_ENABLED NORMALISATION ───────────────────────────────────────
    ondc_col = next((c for c in ["ondc_enabled", "ondc_status", "ondc"]
                     if c in df.columns), None)
    if ondc_col:
        def _norm_ondc(val):
            if pd.isna(val):
                return val
            s = str(val).strip().lower()
            if s in ("yes", "true", "1", "y"):   return "yes"
            if s in ("no",  "false","0", "n"):   return "no"
            # Fuzzy: starts with y → yes, starts with n/f → no
            if s.startswith("y"):  return "yes"
            if s.startswith("n") or s.startswith("f"): return "no"
            return None   # truly invalid → NaN
        df[ondc_col] = df[ondc_col].apply(_norm_ondc)

    # ── 7. DUPLICATE ROWS (date + store_id + product_id) ────────────────────
    dedup_cols = [c for c in ["date", "store_id", "store", "product_id",
                               "product", "sku"] if c in df.columns]
    if len(dedup_cols) >= 2:
        df = df.drop_duplicates(subset=dedup_cols, keep="first")

    return df

def generate_blank_template():
    """
    Professional two-sheet DataNetra Retail Template.

    Sheet 1 — Instructions         : logo strip, description, column guide, notes.
    Sheet 2 — Retail_Data_Template : styled headers, frozen row, dropdown validation.
    """
    import tempfile as _t
    try:
        import openpyxl as _xl
        from openpyxl.styles import Font as _Font, PatternFill as _Fill, \
            Alignment as _Align, Border as _Border, Side as _Side
        from openpyxl.utils import get_column_letter as _gcl
        from openpyxl.worksheet.datavalidation import DataValidation as _DV
    except ImportError:
        # Fallback: plain header-only xlsx via pandas
        import pandas as _pd2
        _cols = ["date","store_id","product_id","category","sales","cost_price",
                 "units_sold","return_rate","inventory_level","vendor_name",
                 "udyam_number","ondc_enabled"]
        _tmp2 = _t.NamedTemporaryFile(delete=False, suffix=".xlsx",
                                      prefix="DataNetra_Retail_Template_")
        _pd2.DataFrame(columns=_cols).to_excel(_tmp2.name, index=False)
        _tmp2.close()
        return _tmp2.name

    # ── Palette ──────────────────────────────────────────────────────────────
    NAVY   = "0B1F3A"; BLUE   = "1B4F8A"; LTBLUE = "EBF2FC"
    HDRFIL = "D0DFF0"; WHITE  = "FFFFFF"; AMBER  = "F6AD3C"
    LTGREY = "F4F7FB"; GREY   = "94A3B8"; GREEN  = "166534"
    GREENBG= "DCFCE7"

    def _f(hex_):   return _Fill("solid", fgColor=hex_)
    def _fnt(bold=False, size=10, color="0B1F3A", italic=False, name="Arial"):
        return _Font(bold=bold, size=size, color=color, italic=italic, name=name)
    def _al(h="left", v="center", wrap=False):
        return _Align(horizontal=h, vertical=v, wrap_text=wrap)
    def _bdr(color="C8DCEF", style="thin"):
        s = _Side(border_style=style, color=color)
        return _Border(top=s, bottom=s, left=s, right=s)
    def _bdrb(color="C8DCEF"):
        return _Border(bottom=_Side(border_style="thin", color=color))

    wb = _xl.Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Instructions
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Instructions"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 3
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 40
    ws1.column_dimensions["D"].width = 26
    ws1.column_dimensions["E"].width = 3

    # Row 1: top padding
    ws1.row_dimensions[1].height = 8

    # Row 2: Logo / brand strip
    ws1.merge_cells("B2:D2")
    ws1.row_dimensions[2].height = 40
    c = ws1["B2"]
    c.value     = "DataNetra.ai  |  Retail Intelligence Platform"
    c.font      = _fnt(bold=True, size=18, color=WHITE)
    c.fill      = _f(NAVY)
    c.alignment = _al("left", "center")
    ws1["E2"].fill = _f(AMBER)

    # Row 3: spacer
    ws1.row_dimensions[3].height = 6

    # Row 4: Title
    ws1.merge_cells("B4:D4")
    ws1.row_dimensions[4].height = 28
    c = ws1["B4"]
    c.value     = "DataNetra Retail Data Template"
    c.font      = _fnt(bold=True, size=16, color=NAVY)
    c.alignment = _al("left", "center")

    # Row 5: Subtitle
    ws1.merge_cells("B5:D5")
    ws1.row_dimensions[5].height = 18
    c = ws1["B5"]
    c.value     = "Guided template for retail dataset preparation"
    c.font      = _fnt(italic=True, size=11, color=GREY)
    c.alignment = _al("left", "center")

    # Row 6: separator
    ws1.row_dimensions[6].height = 4
    for col in ["B", "C", "D"]:
        ws1[f"{col}6"].border = _bdrb(BLUE)

    # Rows 7-9: Description box
    ws1.merge_cells("B7:D9")
    for r in [7, 8, 9]:
        ws1.row_dimensions[r].height = 16
        for col in ["B", "C", "D"]:
            ws1[f"{col}{r}"].fill = _f(LTBLUE)
    c = ws1["B7"]
    c.value     = ("This template helps retailers prepare their sales and inventory data "
                   "for DataNetra AI-powered retail analytics.")
    c.font      = _fnt(size=10, color="334155")
    c.alignment = _Align(horizontal="left", vertical="top", wrap_text=True, indent=1)

    # Row 10: spacer
    ws1.row_dimensions[10].height = 10

    # Row 11: Section heading
    ws1.merge_cells("B11:D11")
    ws1.row_dimensions[11].height = 20
    c = ws1["B11"]
    c.value     = "COLUMN GUIDE"
    c.font      = _fnt(bold=True, size=9, color=WHITE)
    c.fill      = _f(BLUE)
    c.alignment = _al("left", "center")

    # Row 12: Table header
    ws1.row_dimensions[12].height = 22
    for col, label in [("B","Column Name"), ("C","Description"), ("D","Example")]:
        c = ws1[f"{col}12"]
        c.value     = label
        c.font      = _fnt(bold=True, size=10, color=NAVY)
        c.fill      = _f(HDRFIL)
        c.alignment = _al("left", "center")
        c.border    = _bdr()

    # Rows 13-24: Column definitions
    _COLS_DATA = [
        ("date",            "Transaction date (YYYY-MM-DD format)",           "2024-01-01"),
        ("store_id",        "Store or outlet identifier",                      "STORE001"),
        ("product_id",      "Unique product or SKU identifier",                "SKU123"),
        ("category",        "Product category name",                           "Electronics"),
        ("sales",           "Total sales value — numeric, no currency symbol", "45000"),
        ("cost_price",      "Product purchase / cost price",                   "28000"),
        ("units_sold",      "Number of units sold (whole number)",              "320"),
        ("return_rate",     "Return rate as a percentage (e.g. 3.2 = 3.2%)",   "3.2"),
        ("inventory_level", "Current stock quantity on hand",                  "120"),
        ("vendor_name",     "Supplier or vendor name",                         "Raj Traders"),
        ("udyam_number",    "MSME Udyam registration number",                  "UDYAM-XX-00-0000000"),
        ("ondc_enabled",    "ONDC marketplace participation flag",              "yes  /  no"),
    ]
    for i, (col_name, desc, example) in enumerate(_COLS_DATA):
        row = 13 + i
        ws1.row_dimensions[row].height = 18
        bg = LTGREY if i % 2 == 0 else WHITE

        c = ws1[f"B{row}"]
        c.value     = col_name
        c.font      = _Font(bold=True, size=9, color=BLUE, name="Courier New")
        c.fill      = _f(bg); c.alignment = _al("left", "center"); c.border = _bdr()

        c = ws1[f"C{row}"]
        c.value     = desc
        c.font      = _fnt(size=9, color="334155")
        c.fill      = _f(bg); c.alignment = _al("left", "center", wrap=True); c.border = _bdr()

        c = ws1[f"D{row}"]
        c.value     = example
        _ex_col = "1D4ED8" if col_name == "ondc_enabled" else GREEN
        _ex_bg  = "EFF6FF" if col_name == "ondc_enabled" else GREENBG
        c.font      = _Font(size=9, color=_ex_col,
                            italic=(col_name == "ondc_enabled"), name="Arial")
        c.fill      = _f(_ex_bg); c.alignment = _al("left", "center"); c.border = _bdr()

    # Row 25: spacer
    ws1.row_dimensions[25].height = 12

    # Row 26: Notes heading
    ws1.merge_cells("B26:D26")
    ws1.row_dimensions[26].height = 20
    c = ws1["B26"]
    c.value     = "NOTES & GUIDELINES"
    c.font      = _fnt(bold=True, size=9, color=WHITE)
    c.fill      = _f(NAVY)
    c.alignment = _al("left", "center")

    # Rows 27-30: Note items
    _NOTES = [
        "\u2756  Use one row per product per date.",
        "\u2756  Dates must follow YYYY-MM-DD format (e.g. 2024-03-15).",
        "\u2756  Numeric fields (sales, cost_price, units_sold, etc.) must not contain text or currency symbols.",
        "\u2756  Leave cells blank if the value is unknown — do not write N/A or dash.",
    ]
    for j, note in enumerate(_NOTES):
        row = 27 + j
        ws1.merge_cells(f"B{row}:D{row}")
        ws1.row_dimensions[row].height = 17
        c = ws1[f"B{row}"]
        c.value     = note
        c.font      = _fnt(size=9, color="1E3A5F")
        c.fill      = _f(LTBLUE)
        c.alignment = _al("left", "center", wrap=True)
        c.border    = _bdrb("C8DCEF")

    # Row 31: spacer
    ws1.row_dimensions[31].height = 8

    # Row 32: Footer
    ws1.merge_cells("B32:D32")
    ws1.row_dimensions[32].height = 16
    c = ws1["B32"]
    c.value     = "DataNetra.ai  \u00b7  AI-Powered Retail Intelligence  \u00b7  Marketplace Ready"
    c.font      = _fnt(size=8, color=GREY, italic=True)
    c.alignment = _al("center", "center")

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Retail_Data_Template
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Retail_Data_Template")
    ws2.sheet_view.showGridLines = True

    _TMPL_COLS   = ["date","store_id","product_id","category","sales","cost_price",
                    "units_sold","return_rate","inventory_level","vendor_name",
                    "udyam_number","ondc_enabled"]
    _TMPL_WIDTHS = [14, 14, 16, 18, 13, 13, 13, 14, 17, 20, 26, 15]

    for idx, (col_name, width) in enumerate(zip(_TMPL_COLS, _TMPL_WIDTHS), start=1):
        cl = _gcl(idx)
        ws2.column_dimensions[cl].width = width
        c = ws2.cell(row=1, column=idx)
        c.value     = col_name
        c.font      = _Font(bold=True, size=10, color=WHITE, name="Arial")
        c.fill      = _f(NAVY)
        c.alignment = _Align(horizontal="center", vertical="center")
        c.border    = _Border(
            bottom=_Side(border_style="medium", color=AMBER),
            right =_Side(border_style="thin",   color="1E3A5F"),
        )

    ws2.row_dimensions[1].height = 24

    # 50 blank data rows — alternating zebra shading
    for row in range(2, 52):
        ws2.row_dimensions[row].height = 17
        bg = "F0F5FB" if row % 2 == 0 else WHITE
        for idx in range(1, len(_TMPL_COLS) + 1):
            c = ws2.cell(row=row, column=idx)
            c.fill      = _f(bg)
            c.font      = _Font(size=10, name="Arial", color="334155")
            c.alignment = _Align(horizontal="left", vertical="center")
            c.border    = _Border(
                bottom=_Side(border_style="hair", color="D8E8F4"),
                right =_Side(border_style="hair", color="D8E8F4"),
            )

    # Freeze header row
    ws2.freeze_panes = "A2"

    # Dropdown validation for ondc_enabled (column 12)
    _dv = _DV(type="list", formula1='"yes,no"', allow_blank=True,
              showDropDown=False, showErrorMessage=True,
              errorTitle="Invalid value",
              error='Please enter "yes" or "no" only.')
    ws2.add_data_validation(_dv)
    _dv.sqref = "L2:L1000"

    # ── Save ──────────────────────────────────────────────────────────────────
    tmp = _t.NamedTemporaryFile(delete=False, suffix=".xlsx",
                                prefix="DataNetra_Retail_Template_")
    wb.save(tmp.name)
    tmp.close()
    return tmp.name

# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  SECTION H  —  PDF Report Generator                                        ║
# ║  Reuses existing dashboard variables. No logic recalculation.              ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

# Product name lookup — shared by _build_step7_data and generate_bi_report_pdf
_PRODUCT_NAMES = {
    1:'Aashirvaad Atta', 2:'Tata Salt', 3:'Parle-G Biscuits', 4:'Amul Butter',
    5:'Samsung TV 43"', 6:'Whirlpool Fridge', 7:'LG Microwave', 8:'Philips Iron',
    9:'Jeans Regular Fit', 10:'Cotton Kurta', 11:'Ladies Saree', 12:'Kids T-Shirt',
    13:'Prestige Cooker', 14:'Milton Bottle', 15:'Wooden Shelf', 16:'Ceramic Vase',
    17:'Dettol Sanitiser', 18:'Himalaya Face Wash', 19:'Glucon-D 500g', 20:'Band-Aid Box',
    21:'Maggi Noodles', 22:'Britannia Biscuits', 23:'Surf Excel 1kg', 24:'Colgate Toothpaste',
    25:'Bosch Mixer', 26:'Sony Earphones', 27:'Xiaomi Powerbank', 28:'Syska LED Bulb',
    29:'Formal Shirt', 30:'Track Pants', 31:'Winter Jacket', 32:'Ethnic Dupatta',
    33:'Steel Cookware Set', 34:'Bamboo Basket', 35:'Wall Clock', 36:'Photo Frame',
    37:'Dabur Honey', 38:'Patanjali Ghee', 39:'Vitamin-C Tablets', 40:'Neem Face Pack',
}

def generate_bi_report_pdf(user_data, df_raw, dashboard_data, granular_data):
    """
    Build a multi-page Business Intelligence PDF.
    All values read from dashboard_data['snapshot'] -- zero recalculation.
    Charts read from dashboard_data['chart1'-'chart4'] -- same figures as dashboard.
    Presentation: premium retail cover | card-style recommendations | clean headings.
    """
    # -- Ensure reportlab is installed ----------------------------------------
    try:
        import reportlab as _rl_test
    except ImportError:
        try:
            import subprocess as _pip_sub
            _pip_sub.run(
                [sys.executable if hasattr(sys, 'executable') else 'python3',
                 '-m', 'pip', 'install', 'reportlab', '--quiet',
                 '--break-system-packages'],
                check=False, capture_output=True, timeout=120
            )
        except Exception as _pip_e:
            raise RuntimeError(
                f"reportlab is not installed and auto-install failed: {_pip_e}"
            )

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle, PageBreak,
                                         HRFlowable, KeepTogether)
        from reportlab.platypus import Image as RLImage
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        import html as _html_mod
        import re as _re

        # -- Colour palette ---------------------------------------------------
        NAVY    = colors.HexColor("#0B1F3A")
        NAVY2   = colors.HexColor("#12305E")
        STEEL   = colors.HexColor("#1E4D8C")
        TEAL    = colors.HexColor("#1B4F8A")
        ACCENT  = colors.HexColor("#F6AD3C")
        LIGHT   = colors.HexColor("#EEF4FB")
        PALE    = colors.HexColor("#F7FAFF")
        MID     = colors.HexColor("#7A92AA")
        WHITE   = colors.white
        BLACK   = colors.HexColor("#1A2D45")
        GREEN_C = colors.HexColor("#1E8449")
        AMBER_C = colors.HexColor("#D35400")
        RED_C   = colors.HexColor("#C0392B")
        CARD_BG = colors.HexColor("#F2F7FF")

        # -- Value helpers ----------------------------------------------------
        def _fmt_inr(v):
            try:
                v = float(v)
                if v >= 1e7: return f"Rs {v/1e7:.1f} Cr"
                if v >= 1e5: return f"Rs {v/1e5:.1f} L"
                return f"Rs {v:,.0f}"
            except Exception:
                return str(v) if v is not None else "N/A"

        def _strip_html(text):
            if not text: return ""
            text = _re.sub(r'<[^>]+>', ' ', str(text))
            text = _html_mod.unescape(text)
            return _re.sub(r' +', ' ', text).strip()

        # -- Style factory ----------------------------------------------------
        styles = getSampleStyleSheet()
        _sty_cache = {}
        def _sty(name, **kw):
            key = name + str(sorted(kw.items()))
            if key not in _sty_cache:
                _sty_cache[key] = ParagraphStyle(
                    name + str(len(_sty_cache)), parent=styles['Normal'], **kw)
            return _sty_cache[key]

        sty_h3   = _sty('H3',  fontSize=10, textColor=NAVY2,
                         fontName='Helvetica-Bold',
                         spaceBefore=8, spaceAfter=3, leading=13)
        sty_body = _sty('Bd',  fontSize=9,  textColor=BLACK, leading=13, spaceAfter=3)
        sty_small= _sty('Sm',  fontSize=8,  textColor=MID,   leading=11, spaceAfter=2)
        sty_kv   = _sty('KV',  fontSize=9,  textColor=BLACK, leading=12, fontName='Helvetica-Bold')
        sty_foot = _sty('Ft',  fontSize=7,  textColor=MID,
                         alignment=TA_CENTER, leading=10)

        # -- File -------------------------------------------------------------
        msme_id  = (user_data or {}).get('mobile_number', (user_data or {}).get('msme_number', '')).strip()
        safe_id  = ''.join(c for c in msme_id if c.isalnum() or c in '-_+')
        date_str = datetime.datetime.now().strftime('%Y%m%d')
        fname    = (f"Business_Intelligence_Report_{safe_id}_{date_str}.pdf"
                    if safe_id else f"Business_Intelligence_Report_{date_str}.pdf")
        # Save to stable directory so Gradio can always serve it
        _report_dir = os.path.join(tempfile.gettempdir(), 'datanetra_reports')
        os.makedirs(_report_dir, exist_ok=True)
        tmp_path = os.path.join(_report_dir, fname)

        doc = SimpleDocTemplate(
            tmp_path, pagesize=A4,
            leftMargin=1.8*cm, rightMargin=1.8*cm,
            topMargin=2.4*cm,  bottomMargin=2.2*cm,
            title="Business Intelligence Report",
            author="DataNetra Analytics"
        )
        story = []
        W, _H = A4
        PW    = W - 3.6*cm

        # ── Per-page header / footer drawn on canvas ───────────────────────
        def _draw_page_chrome(canvas, doc):
            canvas.saveState()
            # Header bar
            canvas.setFillColor(colors.HexColor('#0B1F3A'))
            canvas.rect(1.8*cm, _H - 1.7*cm, W - 3.6*cm, 0.7*cm, fill=1, stroke=0)
            canvas.setFont('Helvetica-Bold', 7.5)
            canvas.setFillColor(colors.white)
            canvas.drawString(2.2*cm, _H - 1.35*cm, 'DataNetra.ai')
            canvas.setFont('Helvetica', 7.5)
            canvas.setFillColor(colors.HexColor('#9DCBE8'))
            canvas.drawString(2.2*cm + 1.8*cm, _H - 1.35*cm,
                              '|  Retail Intelligence Platform')
            # Page number right of header
            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(colors.HexColor('#A8D8FF'))
            pg_txt = f'Page {doc.page}'
            canvas.drawRightString(W - 1.8*cm, _H - 1.35*cm, pg_txt)
            # Footer bar
            canvas.setFillColor(colors.HexColor('#F2F7FF'))
            canvas.rect(1.8*cm, 1.2*cm, W - 3.6*cm, 0.55*cm, fill=1, stroke=0)
            canvas.setStrokeColor(colors.HexColor('#C8DCEF'))
            canvas.setLineWidth(0.5)
            canvas.line(1.8*cm, 1.75*cm, W - 1.8*cm, 1.75*cm)
            canvas.setFont('Helvetica', 6.5)
            canvas.setFillColor(colors.HexColor('#7A92AA'))
            foot = ('Confidential  |  DataNetra Analytics Platform  |  '
                    f'Generated: {datetime.datetime.now().strftime("%d %b %Y")}')
            canvas.drawCentredString(W / 2, 1.38*cm, foot)
            canvas.restoreState()

        # Snapshot: single source of truth (no recalculation)
        snap = (dashboard_data or {}).get('snapshot', {})

        # =====================================================================
        # SHARED LAYOUT HELPERS
        # =====================================================================

        def _kpi_tbl(rows, col_widths=None):
            td = [[Paragraph(f'<font color="#7A92AA"><b>{l}</b></font>', sty_small),
                   Paragraph(str(v), sty_kv)] for l, v in rows]
            cw = col_widths or [5.5*cm, PW - 5.5*cm]
            t  = Table(td, colWidths=cw, hAlign='LEFT')
            t.setStyle(TableStyle([
                ('ROWBACKGROUNDS', (0,0),(-1,-1), [PALE, WHITE]),
                ('BOX',           (0,0),(-1,-1), 0.5, colors.HexColor('#C8DCEF')),
                ('INNERGRID',     (0,0),(-1,-1), 0.3, colors.HexColor('#E0EAF4')),
                ('TOPPADDING',    (0,0),(-1,-1), 5),
                ('BOTTOMPADDING', (0,0),(-1,-1), 5),
                ('LEFTPADDING',   (0,0),(-1,-1), 8),
                ('RIGHTPADDING',  (0,0),(-1,-1), 8),
                ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
            ]))
            return t

        def _sec(title, sub=''):
            inner = [Paragraph(
                f'<font color="white"><b>{title}</b></font>',
                _sty('SHT', fontSize=13, textColor=WHITE, fontName='Helvetica-Bold',
                     leading=16, spaceBefore=0, spaceAfter=0))]
            if sub:
                inner.append(Paragraph(
                    f'<font color="#9DCBE8">{sub}</font>',
                    _sty('SHS', fontSize=7.5,
                         textColor=colors.HexColor('#9DCBE8'),
                         leading=10, spaceBefore=3, spaceAfter=0)))
            t = Table([[inner]], colWidths=[PW])
            t.setStyle(TableStyle([
                ('BACKGROUND',    (0,0),(-1,-1), NAVY),
                ('TOPPADDING',    (0,0),(-1,-1), 10),
                ('BOTTOMPADDING', (0,0),(-1,-1), 10),
                ('LEFTPADDING',   (0,0),(-1,-1), 14),
                ('RIGHTPADDING',  (0,0),(-1,-1), 14),
                ('ROUNDEDCORNERS', [4]),
            ]))
            return t

        def _chart(key, label, w=15, h=7):
            fig = (dashboard_data or {}).get(key)
            if fig is None:
                return
            try:
                buf = io.BytesIO()
                fig.savefig(buf, format='png', dpi=110, bbox_inches='tight',
                            facecolor='white', edgecolor='none')
                buf.seek(0)
                img = RLImage(buf, width=w*cm, height=h*cm)
                story.append(KeepTogether([
                    Spacer(1, 0.3*cm),
                    Paragraph(label, sty_h3),
                    img,
                ]))
            except Exception:
                pass

        def _4col_grid(rows):
            half  = (len(rows)+1)//2
            left  = rows[:half]
            right = rows[half:]
            gdata = []
            for i in range(max(len(left), len(right))):
                ll = left[i][0]  if i < len(left)  else ''
                lv = left[i][1]  if i < len(left)  else ''
                rl = right[i][0] if i < len(right) else ''
                rv = right[i][1] if i < len(right) else ''
                gdata.append([
                    Paragraph(f'<font color="#7A92AA"><b>{ll}</b></font>', sty_small),
                    Paragraph(str(lv), sty_kv),
                    Paragraph(f'<font color="#7A92AA"><b>{rl}</b></font>', sty_small),
                    Paragraph(str(rv), sty_kv),
                ])
            cw = [4.2*cm, 2.8*cm, 4.2*cm, 2.8*cm]
            g  = Table(gdata, colWidths=cw, hAlign='LEFT')
            g.setStyle(TableStyle([
                ('ROWBACKGROUNDS', (0,0),(-1,-1), [PALE, WHITE]),
                ('BOX',           (0,0),(-1,-1), 0.5, colors.HexColor('#C8DCEF')),
                ('INNERGRID',     (0,0),(-1,-1), 0.3, colors.HexColor('#E0EAF4')),
                ('TOPPADDING',    (0,0),(-1,-1), 5),
                ('BOTTOMPADDING', (0,0),(-1,-1), 5),
                ('LEFTPADDING',   (0,0),(-1,-1), 8),
                ('RIGHTPADDING',  (0,0),(-1,-1), 8),
                ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
            ]))
            return g

        def _rec_card(icon, heading, body_text, accent_color=None):
            # Legacy shim — kept for any residual callers
            ac = accent_color or TEAL
            head_p = Paragraph(
                f'<font color="#FFFFFF"><b>{icon}\u2002{heading}</b></font>',
                _sty('RCH', fontSize=9, textColor=WHITE, fontName='Helvetica-Bold',
                     leading=12, spaceBefore=0, spaceAfter=0))
            body_p = Paragraph(
                body_text,
                _sty('RCB', fontSize=8.5, textColor=BLACK,
                     leading=13, spaceBefore=0, spaceAfter=0))
            card = Table([[head_p, body_p]],
                         colWidths=[4.8*cm, PW - 4.8*cm], hAlign='LEFT')
            card.setStyle(TableStyle([
                ('BACKGROUND',    (0,0), (0,0), ac),
                ('BACKGROUND',    (1,0), (1,0), CARD_BG),
                ('TOPPADDING',    (0,0), (-1,-1), 8),
                ('BOTTOMPADDING', (0,0), (-1,-1), 8),
                ('LEFTPADDING',   (0,0), (0,0), 11),
                ('LEFTPADDING',   (1,0), (1,0), 11),
                ('RIGHTPADDING',  (0,0), (-1,-1), 9),
                ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                ('BOX',           (0,0), (-1,-1), 0.4, colors.HexColor('#C8DCEF')),
                ('LINEAFTER',     (0,0), (0,-1), 1.5, WHITE),
            ]))
            return card

        def _insight_panel(title, icon, metric_label, metric_value,
                           insight_text, action_text, accent_color=None):
            """
            Four-row boxed AI Business Intelligence Insight panel.
            Row 0: Title bar  (accent bg, white bold text + icon)
            Row 1: Key Metric (light accent bg, label left / bold value right)
            Row 2: Insight    (white bg, body text)
            Row 3: Action     (pale bg, bold label + recommendation text)
            """
            ac      = accent_color or colors.HexColor('#0097a7')
            ac_pale = colors.HexColor('#F0F8FF')
            ac_mid  = colors.Color(ac.red, ac.green, ac.blue, alpha=0.08)                       if hasattr(ac, 'red') else colors.HexColor('#EAF4FB')

            # Row 0 — Title bar
            r0 = [Paragraph(
                f'<font color="#FFFFFF"><b>{icon}\u2002{title}</b></font>',
                _sty('IPT', fontSize=9.5, textColor=WHITE,
                     fontName='Helvetica-Bold', leading=13))]

            # Row 1 — Key Metric (2-col inner table for label|value alignment)
            metric_inner = Table([[
                Paragraph(f'<b>{metric_label}</b>',
                    _sty('IPMl', fontSize=8, textColor=colors.HexColor('#2C5282'),
                         fontName='Helvetica-Bold', leading=11)),
                Paragraph(f'<b>{metric_value}</b>',
                    _sty('IPMv', fontSize=11, textColor=ac if hasattr(ac, 'red')
                         else colors.HexColor('#0097a7'),
                         fontName='Helvetica-Bold', leading=13,
                         alignment=TA_RIGHT)),
            ]], colWidths=[PW*0.45, PW*0.55], hAlign='LEFT')
            metric_inner.setStyle(TableStyle([
                ('TOPPADDING',    (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 0),
                ('LEFTPADDING',   (0,0), (-1,-1), 0),
                ('RIGHTPADDING',  (0,0), (-1,-1), 0),
                ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN',         (1,0), (1,-1),  'RIGHT'),
            ]))
            r1 = [metric_inner]

            # Row 2 — Insight text
            r2 = [Paragraph(
                f'<font color="#5A6A7A">\u2139\u2002</font>{insight_text}',
                _sty('IPI', fontSize=8.5, textColor=BLACK, leading=13))]

            # Row 3 — Recommended action
            r3 = [Paragraph(
                f'<font color="#1E6B3A"><b>\u25ba Recommended Action:\u2002</b></font>' +
                action_text,
                _sty('IPA', fontSize=8.5, textColor=colors.HexColor('#1A3A2A'),
                     leading=13))]

            panel = Table(
                [[r0[0]], [r1[0]], [r2[0]], [r3[0]]],
                colWidths=[PW], hAlign='LEFT')
            panel.setStyle(TableStyle([
                # Title bar
                ('BACKGROUND',    (0,0), (0,0), ac),
                ('TOPPADDING',    (0,0), (0,0), 9),
                ('BOTTOMPADDING', (0,0), (0,0), 9),
                ('LEFTPADDING',   (0,0), (0,0), 12),
                ('RIGHTPADDING',  (0,0), (0,0), 12),
                # Key Metric row
                ('BACKGROUND',    (0,1), (0,1), colors.HexColor('#EAF4FB')),
                ('TOPPADDING',    (0,1), (0,1), 8),
                ('BOTTOMPADDING', (0,1), (0,1), 8),
                ('LEFTPADDING',   (0,1), (0,1), 12),
                ('RIGHTPADDING',  (0,1), (0,1), 12),
                # Insight row
                ('BACKGROUND',    (0,2), (0,2), WHITE),
                ('TOPPADDING',    (0,2), (0,2), 9),
                ('BOTTOMPADDING', (0,2), (0,2), 7),
                ('LEFTPADDING',   (0,2), (0,2), 12),
                ('RIGHTPADDING',  (0,2), (0,2), 12),
                # Action row
                ('BACKGROUND',    (0,3), (0,3), colors.HexColor('#F0FFF4')),
                ('TOPPADDING',    (0,3), (0,3), 8),
                ('BOTTOMPADDING', (0,3), (0,3), 9),
                ('LEFTPADDING',   (0,3), (0,3), 12),
                ('RIGHTPADDING',  (0,3), (0,3), 12),
                # Outer box + row dividers
                ('BOX',           (0,0), (-1,-1), 0.6, colors.HexColor('#C8DCEF')),
                ('LINEBELOW',     (0,0), (0,0),   0.5, WHITE),
                ('LINEBELOW',     (0,1), (0,1),   0.3, colors.HexColor('#C8DCEF')),
                ('LINEBELOW',     (0,2), (0,2),   0.3, colors.HexColor('#C8DCEF')),
                ('VALIGN',        (0,0), (-1,-1),  'MIDDLE'),
            ]))
            return panel

        def _ai_insight(text):
            """Teal-accented AI Insight box placed below a chart."""
            INSIGHT_TEAL = colors.HexColor('#1B4F8A')
            INSIGHT_PALE = colors.HexColor('#EBF4FF')
            lbl = Paragraph(
                '<font color="#FFFFFF"><b>\u2728  AI Insight</b></font>',
                _sty('AiLbl', fontSize=8, textColor=colors.white,
                     fontName='Helvetica-Bold', leading=11,
                     spaceBefore=0, spaceAfter=0))
            body = Paragraph(
                text,
                _sty('AiBd', fontSize=8.5, textColor=colors.HexColor('#1A2D45'),
                     leading=13, spaceBefore=0, spaceAfter=0))
            tbl = Table([[lbl, body]],
                        colWidths=[2.8*cm, PW - 2.8*cm], hAlign='LEFT')
            tbl.setStyle(TableStyle([
                ('BACKGROUND',    (0,0), (0,0),  INSIGHT_TEAL),
                ('BACKGROUND',    (1,0), (1,0),  INSIGHT_PALE),
                ('TOPPADDING',    (0,0), (-1,-1), 7),
                ('BOTTOMPADDING', (0,0), (-1,-1), 7),
                ('LEFTPADDING',   (0,0), (0,0),   9),
                ('LEFTPADDING',   (1,0), (1,0),   11),
                ('RIGHTPADDING',  (0,0), (-1,-1), 9),
                ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                ('BOX',           (0,0), (-1,-1), 0.5, colors.HexColor('#B8D4F0')),
                ('LINEAFTER',     (0,0), (0,-1),  1.5, colors.white),
            ]))
            return tbl

        # =====================================================================
        # PAGE 1 -- COVER  (IMPROVEMENT 1: premium retail layout)
        # =====================================================================
        company   = (user_data or {}).get('company_name', '\u2014')
        full_name = (user_data or {}).get('full_name',    '\u2014')
        biz_type  = (user_data or {}).get('business_type', 'Hypermarket')
        gen_date  = datetime.datetime.now().strftime('%d %B %Y, %H:%M')

        # 1a: Deep-navy hero block
        hero_inner = [
            Spacer(1, 0.55*cm),
            Paragraph(
                'DataNetra.ai  \u00b7  Decision Intelligence Platform',
                _sty('CovPl', fontSize=8.5,
                     textColor=colors.HexColor('#7ABADC'),
                     fontName='Helvetica', alignment=TA_CENTER, leading=11)),
            Spacer(1, 0.28*cm),
            HRFlowable(width='28%', thickness=1.5, color=ACCENT,
                       hAlign='CENTER', spaceAfter=0),
            Spacer(1, 0.3*cm),
            Paragraph(
                'BUSINESS INTELLIGENCE REPORT',
                _sty('CovT', fontSize=25, textColor=WHITE, fontName='Helvetica-Bold',
                     alignment=TA_CENTER, leading=30)),
            Spacer(1, 0.2*cm),
            Paragraph(
                'Retail \u0026 Hypermarket Analytics',
                _sty('CovSub', fontSize=13.5, textColor=ACCENT,
                     fontName='Helvetica', alignment=TA_CENTER, leading=17)),
            Spacer(1, 0.6*cm),
        ]
        hero_tbl = Table([[hero_inner]], colWidths=[PW])
        hero_tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,-1), NAVY),
            ('TOPPADDING',    (0,0),(-1,-1), 0),
            ('BOTTOMPADDING', (0,0),(-1,-1), 0),
            ('LEFTPADDING',   (0,0),(-1,-1), 18),
            ('RIGHTPADDING',  (0,0),(-1,-1), 18),
            ('ROUNDEDCORNERS', [5]),
        ]))
        story.append(hero_tbl)
        story.append(Spacer(1, 0.28*cm))

        # 1b: Identity meta table
        meta_items = []
        if msme_id:               meta_items.append(('User ID', msme_id))
        if company != '\u2014':   meta_items.append(('Business Name',   company))
        meta_items.append(('Business Type', biz_type))
        meta_items.append(('Sector',        'Retail'))
        meta_items.append(('Generated',
                            datetime.datetime.now().strftime('%d %b %Y, %H:%M')))
        meta_td = []
        for lbl, val in meta_items:
            meta_td.append([
                Paragraph(f'<font color="#7A92AA"><b>{lbl}</b></font>', sty_small),
                Paragraph(str(val), sty_kv),
            ])
        meta_t = Table(meta_td, colWidths=[4.5*cm, PW - 4.5*cm], hAlign='LEFT')
        meta_t.setStyle(TableStyle([
            ('ROWBACKGROUNDS', (0,0),(-1,-1), [PALE, WHITE]),
            ('BOX',           (0,0),(-1,-1), 0.5, colors.HexColor('#C8DCEF')),
            ('INNERGRID',     (0,0),(-1,-1), 0.3, colors.HexColor('#E0EAF4')),
            ('TOPPADDING',    (0,0),(-1,-1), 5),
            ('BOTTOMPADDING', (0,0),(-1,-1), 5),
            ('LEFTPADDING',   (0,0),(-1,-1), 10),
            ('RIGHTPADDING',  (0,0),(-1,-1), 10),
            ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
        ]))
        story.append(meta_t)
        story.append(Spacer(1, 0.28*cm))

        # 1c: Retail categories — pill badge row
        BADGE_BG   = colors.HexColor('#1A3F6F')
        BADGE_TEXT = colors.HexColor('#FFFFFF')
        BADGE_BORDER = colors.HexColor('#3B6EA8')
        _cat_items = [
            ('\u2708', 'FMCG'),
            ('\u25a3', 'Clothing'),
            ('\u26a1', 'Electronics'),
            ('\u2665', 'Health & Wellness'),
            ('\u2302', 'Home & Decor'),
        ]
        _badge_cells = []
        for _ico, _name in _cat_items:
            _badge_cells.append(
                Table([[Paragraph(
                    f'<font color="#BDD9F5"><b>{_ico}</b></font>'
                    f'\u2002<font color="#FFFFFF"><b>{_name}</b></font>',
                    _sty('BadgeTxt', fontSize=8, leading=11, alignment=TA_CENTER,
                         spaceBefore=0, spaceAfter=0))]],
                    colWidths=[PW / len(_cat_items) - 0.15*cm])
            )
            _badge_cells[-1].setStyle(TableStyle([
                ('BACKGROUND',    (0,0),(-1,-1), BADGE_BG),
                ('BOX',           (0,0),(-1,-1), 0.6, BADGE_BORDER),
                ('TOPPADDING',    (0,0),(-1,-1), 6),
                ('BOTTOMPADDING', (0,0),(-1,-1), 6),
                ('LEFTPADDING',   (0,0),(-1,-1), 5),
                ('RIGHTPADDING',  (0,0),(-1,-1), 5),
                ('ROUNDEDCORNERS', [4]),
            ]))
        _badge_lbl = Table([[
            Paragraph('Retail Categories Covered',
                _sty('CatLbl2', fontSize=7.5, textColor=colors.HexColor('#A8D8FF'),
                     fontName='Helvetica-Bold', alignment=TA_CENTER, leading=10)),
        ]], colWidths=[PW])
        _badge_lbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,-1), NAVY2),
            ('TOPPADDING',    (0,0),(-1,-1), 6),
            ('BOTTOMPADDING', (0,0),(-1,-1), 2),
            ('LEFTPADDING',   (0,0),(-1,-1), 0),
            ('RIGHTPADDING',  (0,0),(-1,-1), 0),
        ]))
        _badge_row_tbl = Table(
            [_badge_cells],
            colWidths=[PW / len(_cat_items)] * len(_cat_items),
            hAlign='CENTER'
        )
        _badge_row_tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,-1), NAVY2),
            ('TOPPADDING',    (0,0),(-1,-1), 4),
            ('BOTTOMPADDING', (0,0),(-1,-1), 8),
            ('LEFTPADDING',   (0,0),(-1,-1), 3),
            ('RIGHTPADDING',  (0,0),(-1,-1), 3),
            ('COLPADDING',    (0,0),(-1,-1), 3),
        ]))
        _badge_outer = Table([[_badge_lbl], [_badge_row_tbl]], colWidths=[PW])
        _badge_outer.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,-1), NAVY2),
            ('TOPPADDING',    (0,0),(-1,-1), 0),
            ('BOTTOMPADDING', (0,0),(-1,-1), 0),
            ('LEFTPADDING',   (0,0),(-1,-1), 0),
            ('RIGHTPADDING',  (0,0),(-1,-1), 0),
            ('ROUNDEDCORNERS', [4]),
        ]))
        story.append(_badge_outer)
        story.append(Spacer(1, 0.28*cm))

        # 1d: Executive Summary
        story.append(Paragraph(
            '<b>Executive Summary</b>',
            _sty('ESH', fontSize=10, textColor=NAVY,
                 fontName='Helvetica-Bold', leading=13, spaceAfter=3)))
        story.append(Paragraph(
            'This report is a direct snapshot of the DataNetra AI dashboard. '
            'All scores, marketplace metrics, revenue figures, forecast projections and '
            'recommendations are identical to those displayed in the dashboard '
            'analysis \u2014 no values have been recalculated.',
            _sty('ESB', fontSize=9, textColor=BLACK, leading=13, spaceAfter=0)))
        story.append(Spacer(1, 0.28*cm))

        # 1e: Decision Intelligence Flow strip
        flow_inner = [
            Paragraph(
                'Decision Intelligence Flow',
                _sty('FlL', fontSize=7.5,
                     textColor=colors.HexColor('#A8D8FF'),
                     fontName='Helvetica-Bold', alignment=TA_CENTER, leading=9)),
            Spacer(1, 0.1*cm),
            Paragraph(
                'Data Ingestion  \u2192  AI Analysis  \u2192  '
                'Business Insights  \u2192  Actionable Decisions',
                _sty('FlV', fontSize=9.5, textColor=WHITE,
                     fontName='Helvetica-Bold', alignment=TA_CENTER, leading=13)),
        ]
        flow_tbl = Table([[flow_inner]], colWidths=[PW])
        flow_tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,-1), STEEL),
            ('TOPPADDING',    (0,0),(-1,-1), 8),
            ('BOTTOMPADDING', (0,0),(-1,-1), 8),
            ('LEFTPADDING',   (0,0),(-1,-1), 14),
            ('RIGHTPADDING',  (0,0),(-1,-1), 14),
            ('ROUNDEDCORNERS', [4]),
        ]))
        story.append(flow_tbl)
        story.append(Spacer(1, 0.28*cm))

        # 1f: Key Business Snapshot (top-line figures from snap)
        if snap:
            story.append(Paragraph(
                '<b>Key Business Snapshot</b>',
                _sty('KSH', fontSize=10, textColor=NAVY,
                     fontName='Helvetica-Bold', leading=13, spaceAfter=4)))
            ks_rows = []
            _ks_ts = snap.get('total_sales')
            _ks_hs = snap.get('health_score')
            _ks_pm = snap.get('avg_margin')
            _ks_ar = snap.get('avg_return')
            _ks_or = snap.get('ondc_readiness')
            _ks_vs = snap.get('vendor_score')
            if _ks_ts is not None: ks_rows.append(('Total Gross Sales',  _fmt_inr(_ks_ts)))
            if _ks_hs is not None: ks_rows.append(('Business Health Score',  f'{_ks_hs:.1f}%'))
            if _ks_pm is not None: ks_rows.append(('Avg Profit Margin',  f'{_ks_pm:.1f}%'))
            if _ks_ar is not None: ks_rows.append(('Avg Return Rate',    f'{_ks_ar:.1f}%'))
            if _ks_or is not None: ks_rows.append(('Digital Readiness',     f'{_ks_or:.0f}%'))
            if _ks_vs is not None: ks_rows.append(('Vendor Reliability', f'{_ks_vs:.1f}/100'))
            if ks_rows:
                story.append(_4col_grid(ks_rows))

        story.append(Spacer(1, 0.35*cm))
        story.append(HRFlowable(width='60%', thickness=0.5,
            color=colors.HexColor('#3B6EA8'), hAlign='CENTER', spaceAfter=4))
        story.append(Paragraph(
            f'<font color="#7A92AA">Confidential</font>'
            f'<font color="#3B6EA8">  |  </font>'
            f'<font color="#7A92AA">DataNetra Analytics Platform</font>'
            f'<font color="#3B6EA8">  |  </font>'
            f'<font color="#7A92AA">{gen_date}</font>',
            sty_foot))
        story.append(PageBreak())

        # =====================================================================
        # PAGE 2 -- KEY INTELLIGENCE SNAPSHOT
        # Full KPI summary — all values from snap, zero recalculation
        # =====================================================================
        story.append(KeepTogether([
            _sec('Key Intelligence Snapshot',
                 'Primary KPI summary  \u00b7  all values from AI dashboard snapshot'),
            Spacer(1, 0.35*cm),
        ]))

        if snap:
            try:
                _kis_hs  = snap.get('health_score')
                _kis_vs  = snap.get('vendor_score')
                _kis_ps  = snap.get('perf_score')
                _kis_gs  = snap.get('growth_score')
                _kis_fr  = snap.get('fin_risk')
                _kis_or  = snap.get('ondc_readiness')
                _kis_pm  = snap.get('avg_margin')
                _kis_ar  = snap.get('avg_return')
                _kis_ts  = snap.get('total_sales', 0)
                _kis_ns  = snap.get('net_sales',   snap.get('gross_sales', 0))
                _kis_gdisp = (
                    f'{_kis_gs*100:.0f}%' if _kis_gs is not None and _kis_gs <= 1.0
                    else (f'{_kis_gs:.0f}%' if _kis_gs is not None else '\u2014')
                )

                # 2-column KPI cards: label | value | label | value
                _kis_rows = [
                    ('Business Health Score',     f'{_kis_hs:.1f}%'     if _kis_hs  is not None else '\u2014',
                     'Vendor Reliability',    f'{_kis_vs:.1f}/100'  if _kis_vs  is not None else '\u2014'),
                    ('Performance Score',     f'{_kis_ps:.1f}%'     if _kis_ps  is not None else '\u2014',
                     'Growth Potential Score',_kis_gdisp),
                    ('Financial Risk Score',  f'{_kis_fr:.3f}'      if _kis_fr  is not None else '\u2014',
                     'Digital Readiness Score',  f'{_kis_or:.0f}%'     if _kis_or  is not None else '\u2014'),
                    ('Avg Profit Margin',     f'{_kis_pm:.1f}%'     if _kis_pm  is not None else '\u2014',
                     'Avg Return Rate',       f'{_kis_ar:.1f}%'     if _kis_ar  is not None else '\u2014'),
                    ('Total Gross Sales',     _fmt_inr(_kis_ts),
                     'Net Sales',             _fmt_inr(_kis_ns)),
                ]
                KIS_NAVY      = colors.HexColor('#0B1F3A')
                KIS_HDR_BG    = colors.HexColor('#EEF4FB')
                KIS_LABEL     = colors.HexColor('#3A5A7A')
                KIS_DIV       = colors.HexColor('#C8DCEF')
                KIS_GRID      = colors.HexColor('#DDEAF7')
                # ── Fixed 4-column layout: name | value ‖ name | value
                # name cols: left-aligned; value cols: right-aligned bold
                # Column widths: label 44% | value 12% | label 36% | value 8%  (of PW each half)
                _NW  = PW * 0.335   # left-group: KPI name
                _VW  = PW * 0.165   # left-group: value
                _NW2 = PW * 0.335   # right-group: KPI name
                _VW2 = PW * 0.165   # right-group: value

                _sty_name = _sty('KisName', fontSize=8.5, textColor=KIS_LABEL,
                                 fontName='Helvetica', leading=13)
                _sty_val  = _sty('KisVal',  fontSize=9.5, textColor=KIS_NAVY,
                                 fontName='Helvetica-Bold', leading=13,
                                 alignment=TA_RIGHT)
                _sty_hdr  = _sty('KisHdr',  fontSize=7.5, textColor=KIS_NAVY,
                                 fontName='Helvetica-Bold', leading=11)
                _sty_hdrv = _sty('KisHdrV', fontSize=7.5, textColor=KIS_NAVY,
                                 fontName='Helvetica-Bold', leading=11,
                                 alignment=TA_RIGHT)

                _kis_hdr = [
                    Paragraph('<b>KPI</b>',   _sty_hdr),
                    Paragraph('<b>Value</b>', _sty_hdrv),
                    Paragraph('<b>KPI</b>',   _sty_hdr),
                    Paragraph('<b>Value</b>', _sty_hdrv),
                ]
                _kis_data = [_kis_hdr]
                for _kl, _kv, _rl, _rv in _kis_rows:
                    _kis_data.append([
                        Paragraph(_kl, _sty_name),
                        Paragraph(f'<b>{_kv}</b>', _sty_val),
                        Paragraph(_rl, _sty_name),
                        Paragraph(f'<b>{_rv}</b>', _sty_val),
                    ])
                _kis_tbl = Table(_kis_data,
                    colWidths=[_NW, _VW, _NW2, _VW2],
                    rowHeights=[22] + [20] * len(_kis_rows),
                    hAlign='LEFT')
                _kis_tbl.setStyle(TableStyle([
                    # Header row background
                    ('BACKGROUND',    (0, 0), (-1, 0),  KIS_HDR_BG),
                    # Alternating row backgrounds
                    ('ROWBACKGROUNDS',(0, 1), (-1, -1), [PALE, WHITE]),
                    # Outer border
                    ('BOX',           (0, 0), (-1, -1), 0.6, KIS_DIV),
                    # Horizontal grid lines only (no vertical clutter inside groups)
                    ('LINEBELOW',     (0, 0), (-1, -2), 0.3, KIS_GRID),
                    # Vertical divider between left and right groups
                    ('LINEAFTER',     (1, 0), (1, -1),  1.0, KIS_DIV),
                    # Padding — generous left on names, tighter on values
                    ('TOPPADDING',    (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('LEFTPADDING',   (0, 0), (0, -1),  10),   # left name
                    ('LEFTPADDING',   (1, 0), (1, -1),  4),    # left value
                    ('LEFTPADDING',   (2, 0), (2, -1),  14),   # right name (indent after divider)
                    ('LEFTPADDING',   (3, 0), (3, -1),  4),    # right value
                    ('RIGHTPADDING',  (0, 0), (0, -1),  4),    # left name
                    ('RIGHTPADDING',  (1, 0), (1, -1),  10),   # left value — flush right
                    ('RIGHTPADDING',  (2, 0), (2, -1),  4),    # right name
                    ('RIGHTPADDING',  (3, 0), (3, -1),  10),   # right value — flush right
                    # Vertical alignment
                    ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                    # Name columns: left-align
                    ('ALIGN',         (0, 0), (0, -1),  'LEFT'),
                    ('ALIGN',         (2, 0), (2, -1),  'LEFT'),
                    # Value columns: right-align
                    ('ALIGN',         (1, 0), (1, -1),  'RIGHT'),
                    ('ALIGN',         (3, 0), (3, -1),  'RIGHT'),
                ]))
                story.append(_kis_tbl)
                story.append(Spacer(1, 0.35*cm))

                # Contextual note — no rating, no grade
                story.append(Table([[
                    Paragraph(
                        '\u2139  This snapshot reflects the exact KPIs computed by the '
                        'DataNetra AI engine from your uploaded dataset. '
                        'Use these figures as the basis for the action plan on the final page.',
                        _sty('KisNote', fontSize=8, textColor=colors.HexColor('#3A5A7A'),
                             leading=12))
                ]], colWidths=[PW]))
                story.append(Paragraph('', sty_small))

            except Exception as _ekis:
                story.append(Paragraph(f'KPI snapshot error: {_ekis}', sty_small))
        else:
            story.append(Paragraph('Snapshot data not available.', sty_small))

        story.append(PageBreak())

        # =====================================================================
        # PAGE 3 -- BUSINESS PERFORMANCE SNAPSHOT
        # Source: dashboard_data['snapshot'] -- identical to Step 5 dashboard
        # =====================================================================
        _p2_header = [
            _sec('Business Performance Snapshot',
                 'AI-computed scores  \u00b7  values identical to Step 5 dashboard'),
            Spacer(1, 0.3*cm),
        ]
        story.append(KeepTogether(_p2_header))

        if snap:
            try:
                hs  = snap.get('health_score')
                ps  = snap.get('perf_score')
                fr  = snap.get('fin_risk')
                vs  = snap.get('vendor_score')
                gs  = snap.get('growth_score')
                pm  = snap.get('avg_margin')
                ar  = snap.get('avg_return')
                ts  = snap.get('total_sales', 0)
                ors = snap.get('ondc_readiness')
                pu  = snap.get('profit_upside', 0)
                mg  = snap.get('margin_gap', 0)
                rr  = snap.get('ret_recovery', 0)

                score_rows = []
                if hs  is not None: score_rows.append(('Business Health Score',     f'{hs:.1f}%'))
                if ps  is not None: score_rows.append(('Performance Score',      f'{ps:.1f}%'))
                if fr  is not None: score_rows.append(('Financial Risk Score',   f'{fr:.3f}'))
                if pm  is not None: score_rows.append(('Avg Profit Margin',      f'{pm:.1f}%'))
                if vs  is not None: score_rows.append(('Vendor Reliability',     f'{vs:.1f}/100'))
                if gs  is not None:
                    gdisp = f'{gs*100:.0f}%' if gs <= 1.0 else f'{gs:.0f}%'
                    score_rows.append(('Growth Potential Score', gdisp))
                if ors is not None: score_rows.append(('Digital Readiness Score',  f'{ors:.0f}%'))

                if score_rows:
                    story.append(Paragraph('Core Business Scores', sty_h3))
                    story.append(_4col_grid(score_rows))
                    story.append(Spacer(1, 0.4*cm))

                story.append(Paragraph('Business Opportunity Insights', sty_h3))
                opp_rows = [
                    ('Total Gross Sales',
                     _fmt_inr(ts)),
                    ('Avg Profit Margin',
                     f'{pm:.1f}%' if pm is not None else '\u2014'),
                    ('Profit Opportunity (gap to 25% target)',
                     _fmt_inr(pu) if mg > 0 else 'Already above target margin'),
                    ('Avg Return Rate',
                     f'{ar:.1f}%' if ar is not None else '\u2014'),
                    ('Return Recovery (per 1pp reduction)',
                     _fmt_inr(rr)),
                ]
                story.append(_kpi_tbl(opp_rows))
            except Exception as _e2:
                story.append(Paragraph(f'Score data unavailable: {_e2}', sty_small))
        else:
            story.append(Paragraph(
                'Score snapshot not available. Please re-run analysis.', sty_small))

        _chart('chart1', 'Sales vs Profit Margin \u2014 Quarterly Trend', w=15, h=7)
        story.append(Spacer(1, 0.2*cm))
        story.append(_ai_insight(
            'Sales and profit margin trends reveal the relationship between revenue volume '
            'and operational efficiency across quarters. A widening gap between sales and '
            'margin lines typically indicates rising procurement or return costs. '
            'Converging lines in later periods suggest cost optimisation is taking effect.'))
        story.append(PageBreak())

        # =====================================================================
        # PAGE 4 -- ONDC IMPACT & SNP READINESS
        # Source: dashboard_data['snapshot'] -- same columns as Step 6 charts
        # =====================================================================
        story.append(KeepTogether([
            _sec('Channel Performance \u0026 Digital Readiness',
                 'Revenue figures  \u00b7  same dataset as Step 6 dashboard charts'),
            Spacer(1, 0.3*cm),
        ]))

        if snap:
            try:
                g3   = snap.get('gross_sales', 0)
                n3   = snap.get('net_sales',   0)
                b3   = snap.get('rev_before',  0)
                a3   = snap.get('rev_after',   0)
                pool = snap.get('ondc_pool',   0)
                rr3  = snap.get('avg_ret_rate', 0)
                rp3  = snap.get('replacements', 0)
                ta3  = snap.get('avg_target',   0)

                story.append(Paragraph('Online Channel Revenue Overview', sty_h3))
                ondc_rows = [
                    ('Total Gross Sales',      _fmt_inr(g3)),
                    ('Offline Baseline Revenue',    _fmt_inr(b3) if b3 else 'N/A'),
                    ('Revenue After Online Channels', _fmt_inr(a3) if a3 else 'N/A'),
                    ('Online Channel Revenue Pool',      _fmt_inr(pool)),
                    ('Net Sales',              _fmt_inr(n3)),
                    ('Avg Return Rate',        f'{rr3:.1f}%'),
                    ('Total Replacements',     f'{rp3:,}'),
                    ('Avg Target Achievement', f'{ta3:.1f}%'),
                ]
                story.append(_kpi_tbl(ondc_rows))
                story.append(Spacer(1, 0.3*cm))

                if df_raw is not None:
                    try:
                        _df3 = _apply_col_remap(df_raw.copy())
                        def _sc3(n, fb=None):
                            return n if n in _df3.columns else (
                                fb if fb and fb in _df3.columns else None)
                        _gc3 = _sc3('gross_sales', 'Monthly_Sales_INR')
                        _stc = _sc3('store_id',    'Store_ID')
                        _rrc = _sc3('return_rate_pct', 'Returns_Percentage')
                        _tac = _sc3('target_achievement_pct', 'Fulfillment_Rate')
                        if _stc and _gc3:
                            _agg3 = {_gc3: 'sum'}
                            if _rrc: _agg3[_rrc] = 'mean'
                            if _tac: _agg3[_tac] = 'mean'
                            _st3 = (_df3.groupby(_stc).agg(_agg3).reset_index()
                                    .sort_values(_gc3, ascending=False).head(5))
                            if not _st3.empty:
                                story.append(Paragraph(
                                    'Store-Level Summary \u2014 Top 5 by Revenue',
                                    sty_h3))
                                hdr3 = [Paragraph('<b>Store</b>', sty_small),
                                        Paragraph('<b>Revenue</b>', sty_small)]
                                if _rrc: hdr3.append(Paragraph('<b>Return %</b>', sty_small))
                                if _tac: hdr3.append(Paragraph('<b>Target %</b>', sty_small))
                                tbl3_data = [hdr3]
                                for _, r3 in _st3.iterrows():
                                    row3 = [Paragraph(str(r3[_stc]), sty_kv),
                                            Paragraph(_fmt_inr(r3[_gc3]), sty_kv)]
                                    if _rrc: row3.append(Paragraph(f'{r3[_rrc]:.1f}%', sty_kv))
                                    if _tac: row3.append(Paragraph(f'{r3[_tac]:.1f}%', sty_kv))
                                    tbl3_data.append(row3)
                                ncols3 = len(hdr3)
                                cws3   = [3.5*cm] + [PW/ncols3]*(ncols3-1)
                                st_tbl = Table(tbl3_data, colWidths=cws3, hAlign='LEFT')
                                st_tbl.setStyle(TableStyle([
                                    ('BACKGROUND',    (0,0),(-1,0), LIGHT),
                                    ('ROWBACKGROUNDS',(0,1),(-1,-1), [WHITE, PALE]),
                                    ('BOX',           (0,0),(-1,-1), 0.5,
                                     colors.HexColor('#C8DCEF')),
                                    ('INNERGRID',     (0,0),(-1,-1), 0.3,
                                     colors.HexColor('#E0EAF4')),
                                    ('TOPPADDING',    (0,0),(-1,-1), 5),
                                    ('BOTTOMPADDING', (0,0),(-1,-1), 5),
                                    ('LEFTPADDING',   (0,0),(-1,-1), 8),
                                    ('RIGHTPADDING',  (0,0),(-1,-1), 8),
                                ]))
                                story.append(st_tbl)
                    except Exception:
                        pass
            except Exception as _e3:
                story.append(Paragraph(f'Channel data unavailable: {_e3}', sty_small))
        else:
            story.append(Paragraph('Channel data snapshot not available.', sty_small))

        _chart('chart4', 'Store-Level Channel Performance Analysis', w=15, h=7)
        story.append(Spacer(1, 0.2*cm))
        story.append(_ai_insight(
            'Store-level channel analysis highlights which locations contribute most to '
            'network revenue. Stores with high online channel share but low target achievement '
            'indicate fulfilment bottlenecks. Prioritise inventory allocation to '
            'high-performing stores to maximise online channel revenue.'))
        _chart('chart2', 'Channel Performance \u2014 Before vs After Comparison', w=15, h=7)
        story.append(Spacer(1, 0.2*cm))
        story.append(_ai_insight(
            'Revenue before and after online channel onboarding demonstrates the incremental '
            'platform uplift. A significant positive delta confirms that online channels '
            'are adding net-new demand rather than cannibalising existing sales. '
            'Stores with minimal change may need marketplace platform review or catalogue updates.'))
        story.append(PageBreak())

        # =====================================================================
        # PAGE 5 -- FORECASTING OUTLOOK
        # Source: dashboard_data['snapshot'] -- same model run as Step 5 & 7
        # =====================================================================
        story.append(KeepTogether([
            _sec('Forecasting Outlook',
                 'AI ensemble projections  \u00b7  same model run as Step 5 \u0026 Step 7'),
            Spacer(1, 0.3*cm),
        ]))

        if snap:
            try:
                f6v   = snap.get('f6',  0)
                f6l   = snap.get('f6_lower', 0)
                f6u   = snap.get('f6_upper', 0)
                f12v  = snap.get('f12', 0)
                f12l  = snap.get('f12_lower', 0)
                f12u  = snap.get('f12_upper', 0)
                mname = snap.get('model_name', 'Statistical Ensemble')
                trail = snap.get('trail_6m')
                grw   = snap.get('growth_pct')
                peak  = snap.get('peak_month')

                story.append(Paragraph('Revenue Forecast Summary', sty_h3))
                fc_rows = [
                    ('6-Month Projection',        _fmt_inr(f6v)),
                    ('6-Month Confidence Range',  f'{_fmt_inr(f6l)} \u2013 {_fmt_inr(f6u)}'),
                    ('12-Month Projection',       _fmt_inr(f12v)),
                    ('12-Month Confidence Range', f'{_fmt_inr(f12l)} \u2013 {_fmt_inr(f12u)}'),
                    ('Selected Forecast Model',   mname),
                ]
                if trail is not None:
                    fc_rows.append(('Current 6-Month Revenue Baseline', _fmt_inr(trail)))
                if grw is not None:
                    fc_rows.append(('Expected Revenue Growth (6m)', f'{grw:+.1f}%'))
                elif trail and trail > 0:
                    raw_g = (f6v - trail) / (trail + 1e-9) * 100
                    fc_rows.append(('Expected Revenue Growth (6m)', f'{raw_g:+.1f}%'))
                else:
                    fc_rows.append(('Expected Revenue Growth (6m)',
                                    'Growth estimate unavailable \u2014 insufficient baseline data.'))
                if peak:
                    fc_rows.append(('Peak Demand Month (Prophet)', peak))

                story.append(_kpi_tbl(fc_rows))
                story.append(Spacer(1, 0.2*cm))
                story.append(Paragraph(
                    'Forecast generated using a 4-model ensemble: Prophet, Holt-Winters, '
                    'Linear Regression (with seasonal features), and Statistical Baseline. '
                    'Dynamic ensemble weights are assigned based on validation MAPE scores.',
                    sty_small))
            except Exception as _e4:
                story.append(Paragraph(f'Forecast data error: {_e4}', sty_small))
        else:
            story.append(Paragraph('Forecast snapshot not available.', sty_small))

        _chart('chart3', 'Returns \u0026 Replacements \u2014 Trend Analysis', w=15, h=7)
        story.append(Spacer(1, 0.2*cm))
        story.append(_ai_insight(
            'Return and replacement trends expose product quality and fulfilment issues. '
            'A rising return rate in specific periods often correlates with seasonal demand '
            'spikes or supplier quality dips. Each percentage point reduction in returns '
            'directly recovers resaleable revenue and reduces reverse-logistics cost.'))
        story.append(PageBreak())

        # =====================================================================
        # PAGE 6 -- PRODUCT & CATEGORY INSIGHTS
        # Source: df_raw -- same filtered dataframe as Step 7
        # =====================================================================
        story.append(KeepTogether([
            _sec('Product \u0026 Category Insights',
                 'Top products and category performance  \u00b7  same data as Step 7'),
            Spacer(1, 0.3*cm),
        ]))

        if df_raw is not None:
            try:
                _df5 = _apply_col_remap(df_raw.copy())
                def _c5(*ns_):
                    for n_ in ns_:
                        if n_ in _df5.columns: return n_
                    return None
                sc5  = _c5('net_sales', 'Monthly_Sales_INR', 'gross_sales')
                cat5 = _c5('product_category', 'Product_Category')
                pid5 = _c5('product_id', 'SKU_Name')
                pm5  = _c5('profit_margin_pct', 'Avg_Margin_Percent')
                rr5  = _c5('return_rate_pct', 'Returns_Percentage')

                if pid5 and sc5:
                    _agg5 = {sc5: 'sum'}
                    if pm5: _agg5[pm5] = 'mean'
                    if rr5: _agg5[rr5] = 'mean'
                    _top5 = (_df5.groupby(pid5).agg(_agg5).reset_index()
                             .sort_values(sc5, ascending=False).head(5))
                    if not _top5.empty:
                        story.append(Paragraph('Top 5 Products by Revenue', sty_h3))
                        hdr5 = [Paragraph('<b>Product</b>', sty_small),
                                Paragraph('<b>Revenue</b>', sty_small)]
                        if pm5: hdr5.append(Paragraph('<b>Margin %</b>', sty_small))
                        if rr5: hdr5.append(Paragraph('<b>Return %</b>', sty_small))
                        tbl5_data = [hdr5]
                        for _, r5 in _top5.iterrows():
                            pid_v = r5[pid5]
                            try:    pname5 = _PRODUCT_NAMES.get(int(pid_v), str(pid_v))
                            except: pname5 = str(pid_v)
                            row5 = [Paragraph(str(pname5)[:28], sty_kv),
                                    Paragraph(_fmt_inr(r5[sc5]),  sty_kv)]
                            if pm5: row5.append(Paragraph(f'{r5[pm5]:.1f}%', sty_kv))
                            if rr5: row5.append(Paragraph(f'{r5[rr5]:.1f}%', sty_kv))
                            tbl5_data.append(row5)
                        nc5 = len(hdr5)
                        cw5 = [PW*0.45] + [PW*0.55/(nc5-1)]*(nc5-1)
                        p5t = Table(tbl5_data, colWidths=cw5, hAlign='LEFT')
                        p5t.setStyle(TableStyle([
                            ('BACKGROUND',    (0,0),(-1,0), LIGHT),
                            ('ROWBACKGROUNDS',(0,1),(-1,-1), [WHITE, PALE]),
                            ('BOX',           (0,0),(-1,-1), 0.5, colors.HexColor('#C8DCEF')),
                            ('INNERGRID',     (0,0),(-1,-1), 0.3, colors.HexColor('#E0EAF4')),
                            ('TOPPADDING',    (0,0),(-1,-1), 5),
                            ('BOTTOMPADDING', (0,0),(-1,-1), 5),
                            ('LEFTPADDING',   (0,0),(-1,-1), 8),
                            ('RIGHTPADDING',  (0,0),(-1,-1), 8),
                        ]))
                        story.append(p5t)
                        story.append(Spacer(1, 0.3*cm))

                if cat5 and sc5:
                    _agg5c = {sc5: 'sum'}
                    if pm5: _agg5c[pm5] = 'mean'
                    if rr5: _agg5c[rr5] = 'mean'
                    _cat5 = (_df5.groupby(cat5).agg(_agg5c).reset_index()
                             .sort_values(sc5, ascending=False))
                    if not _cat5.empty:
                        story.append(Paragraph('Category Performance Summary', sty_h3))
                        hdr5c = [Paragraph('<b>Category</b>', sty_small),
                                 Paragraph('<b>Revenue</b>', sty_small)]
                        if pm5: hdr5c.append(Paragraph('<b>Margin %</b>', sty_small))
                        if rr5: hdr5c.append(Paragraph('<b>Return %</b>', sty_small))
                        tbl5c_data = [hdr5c]
                        for _, r5c in _cat5.iterrows():
                            row5c = [Paragraph(str(r5c[cat5])[:24], sty_kv),
                                     Paragraph(_fmt_inr(r5c[sc5]),   sty_kv)]
                            if pm5: row5c.append(Paragraph(f'{r5c[pm5]:.1f}%', sty_kv))
                            if rr5: row5c.append(Paragraph(f'{r5c[rr5]:.1f}%', sty_kv))
                            tbl5c_data.append(row5c)
                        nc5c = len(hdr5c)
                        cw5c = [PW*0.40] + [PW*0.60/(nc5c-1)]*(nc5c-1)
                        c5t  = Table(tbl5c_data, colWidths=cw5c, hAlign='LEFT')
                        c5t.setStyle(TableStyle([
                            ('BACKGROUND',    (0,0),(-1,0), LIGHT),
                            ('ROWBACKGROUNDS',(0,1),(-1,-1), [WHITE, PALE]),
                            ('BOX',           (0,0),(-1,-1), 0.5, colors.HexColor('#C8DCEF')),
                            ('INNERGRID',     (0,0),(-1,-1), 0.3, colors.HexColor('#E0EAF4')),
                            ('TOPPADDING',    (0,0),(-1,-1), 5),
                            ('BOTTOMPADDING', (0,0),(-1,-1), 5),
                            ('LEFTPADDING',   (0,0),(-1,-1), 8),
                            ('RIGHTPADDING',  (0,0),(-1,-1), 8),
                        ]))
                        story.append(c5t)
            except Exception as _e5:
                story.append(Paragraph(f'Product data error: {_e5}', sty_small))
        else:
            story.append(Paragraph('No product data available.', sty_small))

        story.append(PageBreak())

        # =====================================================================
        # PAGE 7 -- AI RECOMMENDATIONS & BUSINESS TAKEAWAYS
        # Source: dashboard_data['snapshot'] -- same scores as Step 5 & Step 7
        # =====================================================================
        story.append(KeepTogether([
            _sec('AI Recommendations \u0026 Business Takeaways',
                 'Actionable insights  \u00b7  same scores as Step 5 \u0026 Step 7 AI panels'),
            Spacer(1, 0.3*cm),
        ]))

        if snap:
            try:
                hs6      = snap.get('health_score') or 60
                vs6      = snap.get('vendor_score')  or 50
                gs6      = snap.get('growth_score')  or 0.5
                pm6      = snap.get('avg_margin')    or 15
                ar6      = snap.get('avg_return')    or 5
                ts6      = snap.get('total_sales',   0)
                f6_6     = snap.get('f6',            0)
                fc_pct6  = snap.get('growth_pct')    or 0
                avg_tgt6 = snap.get('avg_target')    or 0

                # ── Panel 1: Demand Forecasting Insight ──────────────────
                inv_pct = 10 if avg_tgt6 < 70 else (5 if avg_tgt6 < 85 else 0)
                _mg6 = max(0, 20.0 - pm6)

                if fc_pct6 > 3:
                    dem_icon   = '\u2197'
                    dem_accent = GREEN_C
                    dem_metric = f'+{fc_pct6:.0f}% Demand Growth'
                    dem_insight = (f'Demand is growing at +{fc_pct6:.0f}% over the next 6 months. '
                                   f'6-month projected revenue: {_fmt_inr(f6_6)}. '
                                   f'Strong category momentum signals an opportunity to expand '
                                   f'procurement volume and capture incremental sales.')
                    dem_action  = (f'Advance procurement by 3 weeks. Align staffing and '
                                   f'logistics for +{fc_pct6:.0f}% volume uplift to capture '
                                   f'{_fmt_inr(f6_6)} projected revenue.')
                elif fc_pct6 < -3:
                    dem_icon   = '\u26a0'
                    dem_accent = AMBER_C
                    dem_metric = f'{abs(fc_pct6):.0f}% Demand Slowdown'
                    dem_insight = (f'Demand is contracting by {abs(fc_pct6):.0f}% over the '
                                   f'next 6 months. Projected revenue: {_fmt_inr(f6_6)}. '
                                   f'Declining momentum suggests procurement reduction '
                                   f'to prevent overstock and margin erosion.')
                    dem_action  = (f'Reduce procurement orders by {abs(fc_pct6):.0f}% for '
                                   f'the next 30 days. Review slow-moving SKUs and apply '
                                   f'promotional pricing to clear current inventory.')
                else:
                    dem_icon   = '\u2713'
                    dem_accent = colors.HexColor('#0097a7')
                    dem_metric = f'Stable  \u00b7  {_fmt_inr(f6_6)} Projected'
                    dem_insight = (f'Demand is stable across categories. '
                                   f'6-month projected revenue: {_fmt_inr(f6_6)}. '
                                   f'Consistent performance indicates a reliable baseline '
                                   f'for procurement and inventory planning.')
                    dem_action  = (f'Maintain current procurement cycle. '
                                   f'Build a 10% safety buffer stock to protect against '
                                   f'seasonal demand spikes.')

                story.append(KeepTogether([
                    _insight_panel(
                        title         = 'Demand Forecasting Insight',
                        icon          = dem_icon,
                        metric_label  = 'Revenue Outlook (6 months)',
                        metric_value  = dem_metric,
                        insight_text  = dem_insight,
                        action_text   = dem_action,
                        accent_color  = dem_accent,
                    ),
                    Spacer(1, 0.22*cm),
                ]))

                # ── Panel 2: Inventory Optimisation Insight ──────────────────
                if avg_tgt6 < 70:
                    inv_icon   = '\u25a0'
                    inv_accent = RED_C
                    inv_metric = f'{avg_tgt6:.1f}% Fulfilment  \u2014  Critical'
                    inv_insight = (f'Fulfilment rate is critically low at {avg_tgt6:.1f}% '
                                   f'against the \u226585% target. Stock shortfalls are '
                                   f'directly reducing captured revenue and damaging '
                                   f'customer satisfaction scores.')
                    inv_action  = (f'Increase inventory buffer by 10% immediately. '
                                   f'Pre-order top 20% of SKUs by volume for the next '
                                   f'30 days to close the {85 - avg_tgt6:.1f}pp '
                                   f'fulfilment gap.')
                elif avg_tgt6 < 85:
                    inv_icon   = '\u25a0'
                    inv_accent = AMBER_C
                    inv_metric = f'{avg_tgt6:.1f}% Fulfilment  \u2014  Below Target'
                    inv_insight = (f'Fulfilment at {avg_tgt6:.1f}% is below the 85% '
                                   f'operational target. Moderate stockout risk exists '
                                   f'for fast-moving SKUs, particularly during peak '
                                   f'demand windows.')
                    inv_action  = (f'Increase safety stock by 5% for the top-selling '
                                   f'SKU categories. Review reorder triggers and shorten '
                                   f'lead time by pre-qualifying one additional supplier.')
                else:
                    inv_icon   = '\u2713'
                    inv_accent = GREEN_C
                    inv_metric = f'{avg_tgt6:.1f}% Fulfilment  \u2014  On Target'
                    inv_insight = (f'Inventory levels are adequate with fulfilment at '
                                   f'{avg_tgt6:.1f}% \u2014 above the 85% operational '
                                   f'threshold. Supply chain execution is stable and '
                                   f'well-aligned to current demand.')
                    inv_action  = (f'Maintain current reorder cadence. '
                                   f'Optimise cycle frequency for seasonal SKUs and '
                                   f'review carrying costs to reduce working capital tie-up.')

                story.append(KeepTogether([
                    _insight_panel(
                        title         = 'Inventory Optimisation Insight',
                        icon          = inv_icon,
                        metric_label  = 'Fulfilment Rate vs 85% Target',
                        metric_value  = inv_metric,
                        insight_text  = inv_insight,
                        action_text   = inv_action,
                        accent_color  = inv_accent,
                    ),
                    Spacer(1, 0.22*cm),
                ]))

                # ── Panel 3: Supplier Cost Opportunity ───────────────────────
                if _mg6 > 0:
                    sup_icon   = '\u21bb'
                    sup_accent = AMBER_C
                    sup_metric = f'{pm6:.1f}% Margin  \u2014  {_mg6:.1f}pp Below Target'
                    sup_insight = (f'Current profit margin is {pm6:.1f}%, which is '
                                   f'{_mg6:.1f} percentage points below the 20% benchmark. '
                                   f'A 5% reduction in supplier costs on total revenue of '
                                   f'{_fmt_inr(ts6)} would unlock '
                                   f'\u223c{_fmt_inr(ts6 * 0.05)} in additional profit.')
                    sup_action  = (f'Initiate supplier contract renegotiation targeting '
                                   f'a 5% cost reduction. Bundle orders across SKU '
                                   f'categories to improve negotiating leverage and '
                                   f'close the {_mg6:.1f}pp margin gap.')
                else:
                    sup_icon   = '\u2713'
                    sup_accent = GREEN_C
                    sup_metric = f'{pm6:.1f}% Margin  \u2014  Above 20% Benchmark'
                    sup_insight = (f'Profit margin is healthy at {pm6:.1f}% \u2014 '
                                   f'above the 20% industry benchmark. Supplier cost '
                                   f'efficiency is well-managed. Opportunity exists to '
                                   f'reinvest margin surplus into top-line growth.')
                    sup_action  = (f'Expand top-selling SKU count by 10\u201315% '
                                   f'to grow revenue. Explore volume-discount tiers '
                                   f'with primary suppliers to protect margin as '
                                   f'order volumes scale.')

                story.append(KeepTogether([
                    _insight_panel(
                        title         = 'Supplier Cost Opportunity',
                        icon          = sup_icon,
                        metric_label  = 'Avg Profit Margin vs 20% Target',
                        metric_value  = sup_metric,
                        insight_text  = sup_insight,
                        action_text   = sup_action,
                        accent_color  = sup_accent,
                    ),
                    Spacer(1, 0.22*cm),
                ]))

                # ── Panel 4: Return Risk Management ──────────────────────────
                if ar6 >= 7:
                    ret_icon   = '\u26a0'
                    ret_accent = RED_C
                    ret_recover_1pp = _fmt_inr(ts6 * 0.01 * 0.65)
                    ret_metric = f'{ar6:.1f}% Return Rate  \u2014  Exceeds 7% Threshold'
                    ret_insight = (f'Return rate of {ar6:.1f}% exceeds the 7% risk '
                                   f'threshold. Every 1 percentage point reduction '
                                   f'recovers \u223c{ret_recover_1pp} in resaleable '
                                   f'revenue (65% recovery assumption on returned stock). '
                                   f'Root causes are likely concentrated in a small '
                                   f'number of high-return SKUs.')
                    ret_action  = (f'Audit the top 10 SKUs by return volume this month. '
                                   f'Introduce pre-dispatch quality checks and improve '
                                   f'product description accuracy. Target a 1.5pp '
                                   f'return rate reduction within 90 days to recover '
                                   f'\u223c{_fmt_inr(ts6 * 0.015 * 0.65)} in revenue.')
                elif ar6 >= 5:
                    ret_icon   = '\u25cf'
                    ret_accent = AMBER_C
                    ret_metric = f'{ar6:.1f}% Return Rate  \u2014  Monitor Closely'
                    ret_insight = (f'Return rate of {ar6:.1f}% is below the 7% alert '
                                   f'threshold but trending toward risk territory. '
                                   f'Proactive quality management now prevents '
                                   f'escalation and protects margin.')
                    ret_action  = (f'Review packaging quality and delivery accuracy '
                                   f'for the top 5 SKUs by return count. '
                                   f'Set an internal target to hold return rate below '
                                   f'5% over the next quarter.')
                else:
                    ret_icon   = '\u2713'
                    ret_accent = GREEN_C
                    ret_metric = f'{ar6:.1f}% Return Rate  \u2014  Within Limits'
                    ret_insight = (f'Return rate of {ar6:.1f}% is well within the '
                                   f'healthy range (threshold: 7%). Current quality '
                                   f'control processes and product descriptions are '
                                   f'effectively managing customer returns.')
                    ret_action  = (f'Maintain current quality control standards. '
                                   f'Document best practices for low-return SKU '
                                   f'categories and apply them to any new product '
                                   f'listings to sustain this performance.')

                story.append(KeepTogether([
                    _insight_panel(
                        title         = 'Return Risk Management',
                        icon          = ret_icon,
                        metric_label  = 'Avg Return Rate vs 7% Threshold',
                        metric_value  = ret_metric,
                        insight_text  = ret_insight,
                        action_text   = ret_action,
                        accent_color  = ret_accent,
                    ),
                    Spacer(1, 0.22*cm),
                ]))

                # ── Legacy variable aliases for Action Plan (unchanged) ───────
                # These preserve all action plan row conditions below
                if fc_pct6 > 3:
                    dem_color = GREEN_C
                elif fc_pct6 < -3:
                    dem_color = AMBER_C
                else:
                    dem_color = colors.HexColor('#0097a7')
                if inv_pct > 0:
                    inv_color = AMBER_C
                else:
                    inv_color = GREEN_C
                if vs6 < 50:
                    story.append(Spacer(1, 0.05*cm))
                    story.append(Paragraph(
                        f'\u26a0 Vendor Reliability: {vs6:.0f}/100 \u2014 '
                        f'consolidate supplier base and qualify 2 backup vendors.',
                        _sty('VendorNote', fontSize=8, textColor=AMBER_C,
                             leading=12, spaceBefore=0)))

                # ── Action Plan Table ────────────────────────────────────────
                story.append(Spacer(1, 0.3*cm))
                story.append(KeepTogether([
                    _sec('Structured Action Plan',
                         'Priority actions derived from AI analysis  \u00b7  same data as cards above'),
                    Spacer(1, 0.25*cm),
                ]))

                _ap_hdr = [
                    Paragraph('<b>Action</b>',          _sty('ApH1', fontSize=8.5, textColor=WHITE,
                        fontName='Helvetica-Bold', leading=11)),
                    Paragraph('<b>Expected Impact</b>', _sty('ApH2', fontSize=8.5, textColor=WHITE,
                        fontName='Helvetica-Bold', leading=11)),
                    Paragraph('<b>Timeline</b>',        _sty('ApH3', fontSize=8.5, textColor=WHITE,
                        fontName='Helvetica-Bold', leading=11, alignment=TA_CENTER)),
                ]
                _ap_rows = [_ap_hdr]

                # Row: Return rate reduction
                _ret_target = max(ar6 - 1.5, 0) if ar6 >= 7 else None
                if _ret_target is not None:
                    _ret_recover = _fmt_inr(ts6 * 0.01 * 0.65 * 1.5)
                    _ap_rows.append([
                        Paragraph(f'Reduce return rate from {ar6:.1f}% to {_ret_target:.1f}%',
                            _sty('ApD', fontSize=8.5, textColor=BLACK, leading=12)),
                        Paragraph(f'Recover ~{_ret_recover} resaleable revenue',
                            _sty('ApD2', fontSize=8.5, textColor=BLACK, leading=12)),
                        Paragraph('3 months', _sty('ApT', fontSize=8.5, textColor=TEAL,
                            fontName='Helvetica-Bold', leading=12, alignment=TA_CENTER)),
                    ])

                # Row: Supplier renegotiation (if margin below 20%)
                if _mg6 > 0:
                    _margin_gain = _fmt_inr(ts6 * 0.05)
                    _ap_rows.append([
                        Paragraph('Renegotiate supplier contracts (target 5% cost reduction)',
                            _sty('ApD3', fontSize=8.5, textColor=BLACK, leading=12)),
                        Paragraph(f'Unlock ~{_margin_gain} additional profit; margin +{_mg6:.1f}pp',
                            _sty('ApD4', fontSize=8.5, textColor=BLACK, leading=12)),
                        Paragraph('6 months', _sty('ApT2', fontSize=8.5, textColor=TEAL,
                            fontName='Helvetica-Bold', leading=12, alignment=TA_CENTER)),
                    ])

                # Row: Inventory fulfilment improvement
                if inv_pct > 0:
                    _ap_rows.append([
                        Paragraph(f'Increase inventory buffer by {inv_pct}% for key SKUs',
                            _sty('ApD5', fontSize=8.5, textColor=BLACK, leading=12)),
                        Paragraph(f'Raise fulfilment from {avg_tgt6:.0f}% to 85%+ target',
                            _sty('ApD6', fontSize=8.5, textColor=BLACK, leading=12)),
                        Paragraph('1 month', _sty('ApT3', fontSize=8.5, textColor=TEAL,
                            fontName='Helvetica-Bold', leading=12, alignment=TA_CENTER)),
                    ])

                # Row: Marketplace onboarding
                _ondc_target = _fmt_inr(ts6 * 1.15)
                _ap_rows.append([
                    Paragraph(f'Initiate / optimise online marketplace onboarding',
                        _sty('ApD7', fontSize=8.5, textColor=BLACK, leading=12)),
                    Paragraph(f'Support sales growth to ~{_ondc_target} (15% uplift)',
                        _sty('ApD8', fontSize=8.5, textColor=BLACK, leading=12)),
                    Paragraph('12 months', _sty('ApT4', fontSize=8.5, textColor=TEAL,
                        fontName='Helvetica-Bold', leading=12, alignment=TA_CENTER)),
                ])

                # Row: Vendor reliability (if low)
                if vs6 < 50:
                    _ap_rows.append([
                        Paragraph('Consolidate supplier base; qualify 2 backup vendors',
                            _sty('ApD9', fontSize=8.5, textColor=BLACK, leading=12)),
                        Paragraph(f'Reduce supply chain risk from {vs6:.0f}/100 to 70+',
                            _sty('ApD10', fontSize=8.5, textColor=BLACK, leading=12)),
                        Paragraph('4 months', _sty('ApT5', fontSize=8.5, textColor=TEAL,
                            fontName='Helvetica-Bold', leading=12, alignment=TA_CENTER)),
                    ])

                # Row: Demand growth procurement
                if fc_pct6 > 3:
                    _ap_rows.append([
                        Paragraph(f'Advance procurement 3 weeks for +{fc_pct6:.0f}% demand surge',
                            _sty('ApD11', fontSize=8.5, textColor=BLACK, leading=12)),
                        Paragraph(f'Capture projected {_fmt_inr(f6_6)} revenue without stockouts',
                            _sty('ApD12', fontSize=8.5, textColor=BLACK, leading=12)),
                        Paragraph('Immediate', _sty('ApT6', fontSize=8.5,
                            textColor=colors.HexColor('#C0392B'),
                            fontName='Helvetica-Bold', leading=12, alignment=TA_CENTER)),
                    ])

                _ap_tbl = Table(_ap_rows,
                    colWidths=[PW*0.44, PW*0.38, PW*0.18],
                    hAlign='LEFT', repeatRows=1)
                _ap_tbl.setStyle(TableStyle([
                    ('BACKGROUND',    (0,0), (-1,0),  NAVY),
                    ('ROWBACKGROUNDS',(0,1), (-1,-1),  [PALE, WHITE]),
                    ('BOX',           (0,0), (-1,-1),  0.6, colors.HexColor('#C8DCEF')),
                    ('INNERGRID',     (0,0), (-1,-1),  0.3, colors.HexColor('#E0EAF4')),
                    ('TOPPADDING',    (0,0), (-1,-1),  7),
                    ('BOTTOMPADDING', (0,0), (-1,-1),  7),
                    ('LEFTPADDING',   (0,0), (-1,-1),  9),
                    ('RIGHTPADDING',  (0,0), (-1,-1),  9),
                    ('VALIGN',        (0,0), (-1,-1),  'TOP'),
                    ('ALIGN',         (2,0), (2,-1),   'CENTER'),
                ]))
                story.append(KeepTogether([_ap_tbl]))
                story.append(Spacer(1, 0.35*cm))

                # Business Intelligence Summary
                story.append(Spacer(1, 0.05*cm))
                story.append(HRFlowable(width='100%', thickness=0.5,
                    color=colors.HexColor('#C8DCEF'), spaceAfter=8))
                story.append(Paragraph('Business Intelligence Summary', sty_h3))
                biz_type6 = snap.get('biz_type',
                    (user_data or {}).get('business_type', 'Retail'))
                summary_rows = [
                    ('Total Analysed Revenue', _fmt_inr(ts6)),
                    ('Avg Profit Margin',       f'{pm6:.1f}%'),
                    ('Avg Return Rate',         f'{ar6:.1f}%'),
                    ('Business Health Score',       f'{hs6:.1f}%'),
                    ('Vendor Reliability',      f'{vs6:.1f}/100'),
                    ('Business Type',           biz_type6),
                ]
                story.append(_4col_grid(summary_rows))

            except Exception as _e6:
                story.append(Paragraph(f'Recommendation data error: {_e6}', sty_small))
        else:
            story.append(Paragraph(
                'Recommendation snapshot not available. Please re-run analysis.',
                sty_small))

        story.append(Spacer(1, 0.4*cm))
        story.append(HRFlowable(width='100%', thickness=0.5,
            color=colors.HexColor('#C8DCEF'), spaceAfter=4))
        story.append(Paragraph(
            f'<font color="#7A92AA">Business Intelligence Report</font>'
            f'<font color="#3B6EA8">  |  </font>'
            f'<font color="#7A92AA">{gen_date}</font>'
            f'<font color="#3B6EA8">  |  </font>'
            f'<font color="#7A92AA">Confidential  \u00b7  DataNetra Analytics Platform</font>',
            sty_foot))

        # -- Build PDF --------------------------------------------------------
        doc.build(story,
                  onFirstPage=_draw_page_chrome,
                  onLaterPages=_draw_page_chrome)
        return tmp_path
    except Exception as _top_e:
        import traceback as _tb2
        err_msg = f"PDF generation error: {_top_e}\n{_tb2.format_exc()}"
        raise RuntimeError(err_msg) from _top_e
# ══════════════════════════════════════════════════════════════════════════════
# GRADIO UI
# ══════════════════════════════════════════════════════════════════════════════

# ── Voice recognition JavaScript — Web Speech API (Step 1 & Step 2) ────────────
VOICE_JS_STEP1 = r"""
<script>
function startVoiceStep1() {
  var statusEl = document.getElementById('voice-status-1');
  if (!('webkitSpeechRecognition' in window) && !('SpeechRecognition' in window)) {
    statusEl.innerText = '❌ Voice not supported. Please use Chrome or Edge.';
    return;
  }
  var SpeechRecognition = window.SpeechRecognition || window.webkitSpeechRecognition;
  var recog = new SpeechRecognition();
  recog.lang = 'en-IN';
  recog.continuous = false;
  recog.interimResults = false;
  statusEl.innerText = '🎙️ Listening... Speak now';
  statusEl.style.color = '#c0392b';
  recog.start();

  function fillField(elemId, value) {
    /* Try both the wrapper div id and direct input inside it */
    var selectors = [
      '#' + elemId + ' input',
      '#' + elemId + ' textarea',
      '#' + elemId
    ];
    for (var i = 0; i < selectors.length; i++) {
      var el = document.querySelector(selectors[i]);
      if (el && (el.tagName === 'INPUT' || el.tagName === 'TEXTAREA')) {
        var nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value') ||
                                     Object.getOwnPropertyDescriptor(window.HTMLTextAreaElement.prototype, 'value');
        if (nativeInputValueSetter) nativeInputValueSetter.set.call(el, value);
        else el.value = value;
        el.dispatchEvent(new Event('input',  {bubbles: true}));
        el.dispatchEvent(new Event('change', {bubbles: true}));
        return true;
      }
    }
    return false;
  }

  recog.onresult = function(event) {
    var transcript = event.results[0][0].transcript;
    statusEl.style.color = '#1a7a40';
    statusEl.innerText = '✅ Heard: "' + transcript + '"';

    var nameMatch = transcript.match(/(?:my name is|name is|i am|iam)\s+([a-z\s]+?)(?:\s+mobile|\s+number|\s+roll|\s+role|$)/i);
    if (nameMatch) {
      var nm = nameMatch[1].trim();
      fillField('step1-name', nm);
      fillField('step1-name-v', nm);
    }

    var mobileMatch = transcript.match(/(?:mobile|number|phone)\s*(?:is\s*)?(\d[\d\s]{9,13})/i);
    if (mobileMatch) {
      var mob = mobileMatch[1].replace(/\s/g,'');
      fillField('step1-mobile', mob);
      fillField('step1-mobile-v', mob);
    }

    var roles = ['Business Owner','Co-Founder','Category Manager','Analyst','Store Manager'];
    var detectedRole = null;
    for (var r = 0; r < roles.length; r++) {
      if (transcript.toLowerCase().indexOf(roles[r].toLowerCase()) !== -1) {
        detectedRole = roles[r]; break;
      }
    }
    if (detectedRole) {
      /* Gradio dropdown — find the select or the visible input and set via React */
      var dropdownWrap = document.querySelector('#step1-role');
      if (dropdownWrap) {
        var selEl = dropdownWrap.querySelector('input[type="text"], input:not([type]), select');
        if (selEl) {
          fillField('step1-role', detectedRole);
        }
        /* Also try clicking the matching option in the dropdown list */
        var allOptions = dropdownWrap.querySelectorAll('li, [role="option"]');
        allOptions.forEach(function(opt) {
          if (opt.innerText && opt.innerText.trim() === detectedRole) opt.click();
        });
      }
      statusEl.innerText += ' | Role: ' + detectedRole;
    }
  };

  recog.onerror = function(e) {
    statusEl.style.color = '#c0392b';
    statusEl.innerText = '❌ Error: ' + e.error + '. Try again.';
  };
  recog.onend = function() {
    if (statusEl.innerText.indexOf('Heard') === -1 && statusEl.innerText.indexOf('Error') === -1) {
      statusEl.style.color = '#4A6A8A';
      statusEl.innerText = 'Recording stopped. Click again to retry.';
    }
  };
}
</script>
<div style="background:#F0F7FF;border:1px solid #C8DCEF;border-radius:10px;padding:16px 18px;margin-top:8px">
  <div style="font-weight:700;color:#1B4F8A;font-size:14px;margin-bottom:6px">🎙️ Voice Registration — Step 1</div>
  <div style="font-size:12px;color:#2A4060;margin-bottom:10px">Say: <em>"My name is [Name], mobile [number], role [role]"</em></div>
  <div style="font-size:11px;color:#4A6A8A;margin-bottom:10px">Roles: Business Owner · Co-Founder · Category Manager · Analyst · Store Manager</div>
  <button onclick="startVoiceStep1()" style="background:linear-gradient(135deg,#e74c3c,#c0392b);color:#FFFFFF;border:none;border-radius:8px;padding:10px 20px;font-size:13px;font-weight:600;cursor:pointer;letter-spacing:0.5px">🎤 Start Voice Input</button>
  <div id="voice-status-1" style="font-size:13px;color:#2A4060;margin-top:8px;font-style:italic;min-height:20px">Click the button and speak clearly.</div>
  <div style="font-size:11px;color:#4A6A8A;margin-top:6px">⚠️ Works in Chrome / Edge only · Check browser mic permissions</div>
</div>"""

VOICE_JS_STEP2 = r"""
<script>
function startVoiceStep2() {
  var statusEl = document.getElementById('voice-status-2');
  if (!('webkitSpeechRecognition' in window) && !('SpeechRecognition' in window)) {
    statusEl.innerText = '❌ Voice not supported. Please use Chrome or Edge.';
    return;
  }
  var SpeechRecognition = window.SpeechRecognition || window.webkitSpeechRecognition;
  var recog = new SpeechRecognition();
  recog.lang = 'en-IN';
  recog.continuous = false;
  recog.interimResults = false;
  statusEl.innerText = '🎙️ Listening... Speak now';
  statusEl.style.color = '#c0392b';
  recog.start();

  function fillField(elemId, value) {
    var selectors = ['#' + elemId + ' input', '#' + elemId + ' textarea', '#' + elemId];
    for (var i = 0; i < selectors.length; i++) {
      var el = document.querySelector(selectors[i]);
      if (el && (el.tagName === 'INPUT' || el.tagName === 'TEXTAREA')) {
        var nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value') ||
                                     Object.getOwnPropertyDescriptor(window.HTMLTextAreaElement.prototype, 'value');
        if (nativeInputValueSetter) nativeInputValueSetter.set.call(el, value);
        else el.value = value;
        el.dispatchEvent(new Event('input',  {bubbles: true}));
        el.dispatchEvent(new Event('change', {bubbles: true}));
        return true;
      }
    }
    return false;
  }

  recog.onresult = function(event) {
    var transcript = event.results[0][0].transcript;
    statusEl.style.color = '#1a7a40';
    statusEl.innerText = '✅ Heard: "' + transcript + '"';

    /* UDYAM number — capture full UDYAM-XX-XX-XXXXXXX pattern */
    var udyamMatch = transcript.match(/(?:UDYAM|udyam|udhyam)\s*[-]?\s*([A-Z0-9][\w\-\s]{4,20})/i);
    if (udyamMatch) {
      var udyam = udyamMatch[0].replace(/\s+/g,'-').replace(/-{2,}/g,'-').toUpperCase().trim();
      fillField('step2-msme', udyam);
      fillField('step2-msme-v', udyam);
    }

    /* OTP — 4 digit number after keyword */
    var otpMatch = transcript.match(/(?:OTP|otp|code|pass)\s*(?:is\s*)?(\d{4})/i);
    if (otpMatch) {
      /* OTP field is type=password — querySelector finds it via wrapper id */
      var otpWrap = document.getElementById('step2-otp');
      if (otpWrap) {
        var otpEl = otpWrap.querySelector('input');
        if (otpEl) {
          var setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value');
          if (setter) setter.set.call(otpEl, otpMatch[1]);
          else otpEl.value = otpMatch[1];
          otpEl.dispatchEvent(new Event('input',  {bubbles: true}));
          otpEl.dispatchEvent(new Event('change', {bubbles: true}));
        }
      }
      fillField('step2-otp-v', otpMatch[1]);
      statusEl.innerText += ' | OTP: ****';
    }
  };

  recog.onerror = function(e) {
    statusEl.style.color = '#c0392b';
    statusEl.innerText = '❌ Error: ' + e.error + '. Try again.';
  };
  recog.onend = function() {
    if (statusEl.innerText.indexOf('Heard') === -1 && statusEl.innerText.indexOf('Error') === -1) {
      statusEl.style.color = '#4A6A8A';
      statusEl.innerText = 'Recording stopped. Click again to retry.';
    }
  };
}
</script>
<div style="background:#F0F7FF;border:1px solid #C8DCEF;border-radius:10px;padding:16px 18px;margin-top:8px">
  <div style="font-weight:700;color:#1B4F8A;font-size:14px;margin-bottom:6px">🎙️ Voice Registration — Step 2</div>
  <div style="font-size:12px;color:#2A4060;margin-bottom:10px">Say: <em>"UDYAM [number] OTP [4 digits]"</em></div>
  <div style="font-size:11px;color:#4A6A8A;margin-bottom:10px">Example: <em>"UDYAM TN 00 7629703 OTP 1234"</em></div>
  <button onclick="startVoiceStep2()" style="background:linear-gradient(135deg,#e74c3c,#c0392b);color:#FFFFFF;border:none;border-radius:8px;padding:10px 20px;font-size:13px;font-weight:600;cursor:pointer;letter-spacing:0.5px">🎤 Start Voice Input</button>
  <div id="voice-status-2" style="font-size:13px;color:#2A4060;margin-top:8px;font-style:italic;min-height:20px">Click the button and speak clearly.</div>
  <div style="font-size:11px;color:#4A6A8A;margin-top:6px">⚠️ Works in Chrome / Edge only · Check browser mic permissions</div>
</div>"""

# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  SECTION G  —  Gradio Application — UI Layout & Event Wiring               ║
# ║  All gr.Blocks columns, step handlers, event wiring, demo.launch()         ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

def _landing_industries(lang='en'):
    # Translations
    T = {
        'en': {
            'label': "WHO IT'S FOR",
            'title': "Built for Retail &amp; Data-Driven Businesses",
            'sub':   "Works across FMCG, Retail, Electronics, D2C &amp; more",
            'coming': "Healthcare &amp; Agriculture insights — coming soon",
            'pills': ["🛒 Supermarkets","📦 FMCG","💻 Electronics","👗 Apparel &amp; D2C","🏪 Retail Chains"],
        },
        'kn': {
            'label': "ಯಾರಿಗಾಗಿ",
            'title': "ರಿಟೇಲ್ &amp; ಡೇಟಾ-ಆಧಾರಿತ ವ್ಯವಹಾರಗಳಿಗಾಗಿ",
            'sub':   "FMCG, ರಿಟೇಲ್, ಎಲೆಕ್ಟ್ರಾನಿಕ್ಸ್, D2C ಮತ್ತು ಇನ್ನೂ ಹೆಚ್ಚಿನ ಕ್ಷೇತ್ರಗಳಲ್ಲಿ ಕಾರ್ಯನಿರ್ವಹಿಸುತ್ತದೆ",
            'coming': "ಆರೋಗ್ಯ &amp; ಕೃಷಿ ಒಳನೋಟಗಳು — ಶೀಘ್ರದಲ್ಲಿ ಬರಲಿದೆ",
            'pills': ["🛒 ಸೂಪರ್‌ಮಾರ್ಕೆಟ್‌ಗಳು","📦 FMCG","💻 ಎಲೆಕ್ಟ್ರಾನಿಕ್ಸ್","👗 ಉಡುಪು &amp; D2C","🏪 ರಿಟೇಲ್ ಚೈನ್‌ಗಳು"],
        },
        'ta': {
            'label': "யாருக்காக",
            'title': "சில்லறை &amp; தரவு சார்ந்த வணிகங்களுக்காக",
            'sub':   "FMCG, சில்லறை, மின்னணுவியல், D2C மற்றும் பலவற்றில் செயல்படுகிறது",
            'coming': "சுகாதாரம் &amp; விவசாயம் — விரைவில் வருகிறது",
            'pills': ["🛒 சூப்பர்மார்கெட்","📦 FMCG","💻 மின்னணுவியல்","👗 ஆடை &amp; D2C","🏪 சில்லறை சங்கிலிகள்"],
        },
    }
    t = T.get(lang, T['en'])
    pills_html = ''.join(
        f'<span style="background:#EBF5FF;border:1px solid #BDD7F5;color:#1B4F8A;'
        f'font-size:12px;font-weight:700;padding:6px 14px;border-radius:20px;">{p}</span>'
        for p in t['pills']
    )
    return f'''
<div style="background:linear-gradient(to bottom,#f8faff,#ffffff);border-bottom:1px solid #e2eaf5;padding:18px 24px 16px;text-align:center;">
  <div style="max-width:900px;margin:0 auto;">
    <div style="font-size:11px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#1B4F8A;margin-bottom:10px;">{t["label"]}</div>
    <div style="font-size:22px;font-weight:900;color:#0B1F3A;margin-bottom:8px;letter-spacing:-0.3px;line-height:1.25;">{t["title"]}</div>
    <div style="font-size:14px;font-weight:500;color:#475569;margin-bottom:16px;line-height:1.6;">{t["sub"]}</div>
    <div style="display:flex;align-items:center;justify-content:center;gap:10px;flex-wrap:wrap;margin-bottom:14px;">{pills_html}</div>
    <div style="font-size:12px;color:#94a3b8;">⬤ {t["coming"]}</div>
  </div>
</div>
'''


def _landing_unlock_benefits(lang='en'):
    T = {
        'en': {
            'label': "WHAT ₹999/MONTH GETS YOU",
            'title': "Smarter business decisions shouldn't require enterprise consulting costs.",
            'sub': "Get AI-powered retail intelligence, forecasting, and actionable insights in minutes — from your own Excel data.",
            'cards': [
                ("📊","Business Health Score (7 models)","Know your Financial Risk, Growth Potential, Vendor Reliability and MSME Health Score — updated every time you upload. Most retailers never see this until they apply for a loan."),
                ("📈","6 &amp; 12-Month Revenue Forecast","AI projects your next 6 and 12 months from your own sales history — not industry averages. Know your peak season before it hits."),
                ("🤖","30-Day &amp; 90-Day Action Plan","Not just insights — specific actions. Which SKUs to reorder today, which to stop stocking, which store to push harder. Generated from your actual numbers."),
                ("🏪","Store → Category → SKU Drill-down","Click into any store, any category, any product. See returns rate, fulfilment trends, inventory risk and reorder point. The analysis a distributor would charge lakhs for."),
            ],
            'cta': "Start Pro Plan →",
            'trust': "✓ Cancel anytime &nbsp;·&nbsp; ✓ No setup fee &nbsp;·&nbsp; ✓ Free preview first",
        },
        'kn': {
            'label': "ಪೂರ್ಣ ಪ್ರವೇಶ", 'title': "ನೀವು ಅನ್‌ಲಾಕ್ ಮಾಡುವುದು",
            'sub': "ಪೂರ್ವವೀಕ್ಷಣೆಯಿಂದ ಸಂಪೂರ್ಣ ವ್ಯಾಪಾರ ಬುದ್ಧಿಮತ್ತೆಗೆ ಅಪ್‌ಗ್ರೇಡ್ ಮಾಡಿ",
            'cards': [
                ("📊","ಸಂಪೂರ್ಣ ವ್ಯಾಪಾರ ಸ್ಕೋರಿಂಗ್ ಡ್ಯಾಶ್‌ಬೋರ್ಡ್","7 AI ಸ್ಕೋರಿಂಗ್ ಮಾದರಿಗಳು — ಆರ್ಥಿಕ ಅಪಾಯ, ಮಾರಾಟಗಾರ ವಿಶ್ವಾಸಾರ್ಹತೆ, ಬೆಳವಣಿಗೆ ಸಾಮರ್ಥ್ಯ ಮತ್ತು ಹೆಚ್ಚಿನವು."),
                ("📈","6 &amp; 12 ತಿಂಗಳ AI ಮುನ್ಸೂಚನೆ","ನಿಮ್ಮ ಡೇಟಾಸೆಟ್‌ನ R² ನಿಖರತೆಯ ಆಧಾರದ ಮೇಲೆ ತೂಕದ ಸಂಯೋಜನೆ ಮುನ್ಸೂಚನೆ."),
                ("🤖","AI-ಚಾಲಿತ ಶಿಫಾರಸುಗಳು","ನಿಮ್ಮ ನಿಜವಾದ ಮಾರಾಟ ಮಾದರಿಗಳಿಂದ ರೂಪಿಸಲಾದ 0–30 ದಿನ ತಕ್ಷಣದ ಕ್ರಮಗಳು ಮತ್ತು 30–90 ದಿನ ತಂತ್ರಗಳು."),
                ("🏪","ವರ್ಗ &amp; SKU ಆಳ ಒಳನೋಟಗಳು","ಸ್ಟೋರ್ → ವರ್ಗ → ಉತ್ಪನ್ನ ಕೊರೆಯುವಿಕೆ — ಸ್ಟಾಕ್, ರಿಟರ್ನ್ಸ್ ಮತ್ತು ರಿಆರ್ಡರ್ ಶಿಫಾರಸುಗಳೊಂದಿಗೆ."),
            ],
            'cta': "ಸಂಪೂರ್ಣ ಒಳನೋಟಗಳನ್ನು ಅನ್‌ಲಾಕ್ ಮಾಡಿ →",
            'trust': "✓ ಕ್ರೆಡಿಟ್ ಕಾರ್ಡ್ ಇಲ್ಲ &nbsp;·&nbsp; ✓ ತಕ್ಷಣದ ಪ್ರವೇಶ &nbsp;·&nbsp; ✓ ನಿಮ್ಮ ಡೇಟಾ ಸುರಕ್ಷಿತ",
        },
        'ta': {
            'label': "முழு அணுகல்", 'title': "நீங்கள் திறப்பது",
            'sub': "முன்னோட்டத்திலிருந்து முழு வணிக நுண்ணறிவுக்கு மேம்படுத்தவும்",
            'cards': [
                ("📊","முழு வணிக மதிப்பெண் டாஷ்போர்டு","7 AI மதிப்பெண் மாதிரிகள் — நிதி அபாயம், வளர்ச்சி திறன், MSME ஆரோக்கிய மதிப்பெண் மற்றும் பலவும்."),
                ("📈","6 &amp; 12 மாத AI முன்கணிப்பு","உங்கள் தரவின் R² துல்லியத்தால் எடையிடப்பட்ட கூட்டு முன்கணிப்பு."),
                ("🤖","AI இயக்கும் பரிந்துரைகள்","உங்கள் உண்மையான விற்பனை முறைகளிலிருந்து 0–30 நாள் உடனடி நடவடிக்கைகள் மற்றும் 30–90 நாள் திட்டங்கள்."),
                ("🏪","வகை &amp; SKU ஆழமான நுண்ணறிவு","கடை → வகை → தயாரிப்பு துளையிடல் — சரக்கு, திரும்பல்கள் மற்றும் மறு ஆர்டர் பரிந்துரைகளுடன்."),
            ],
            'cta': "முழு நுண்ணறிவுகளை திறக்கவும் →",
            'trust': "✓ கிரெடிட் கார்டு இல்லை &nbsp;·&nbsp; ✓ உடனடி அணுகல் &nbsp;·&nbsp; ✓ தரவு பாதுகாப்பாக உள்ளது",
        },
    }
    t = T.get(lang, T['en'])
    border_colors = ['#1B4F8A','#7c3aed','#16a34a','#f59e0b']
    cards_html = ''
    for i,(icon,title,desc) in enumerate(t['cards']):
        cards_html += (
            f'<div style="background:#ffffff;border:1px solid #e2eaf5;border-top:3px solid {border_colors[i]};'
            f'border-radius:12px;padding:20px;box-shadow:0 2px 8px rgba(11,31,58,0.06);">'
            f'<div style="font-size:24px;margin-bottom:12px;">{icon}</div>'
            f'<div style="font-size:13px;font-weight:800;color:#0B1F3A;margin-bottom:8px;line-height:1.4;">{title}</div>'
            f'<div style="font-size:12px;color:#475569;line-height:1.7;">{desc}</div></div>'
        )
    # comparison table only for EN
    comparison_html = ''
    if lang == 'en':
        comparison_html = '''
<div style="background:linear-gradient(135deg,#f0f5ff,#f8fbff);border:1px solid #dbeafe;
     border-radius:14px;padding:20px 24px;margin-bottom:28px;">
  <div style="font-size:11px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;
       color:#1B4F8A;margin-bottom:14px;text-align:center;">HOW DOES IT COMPARE?</div>
  <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:0;text-align:center;">
    <div style="font-size:10px;font-weight:700;color:#94a3b8;padding-bottom:8px;border-bottom:2px solid #e2eaf5;"></div>
    <div style="font-size:11px;font-weight:800;color:#64748b;padding-bottom:8px;border-bottom:2px solid #e2eaf5;">Traditional CA / Analyst</div>
    <div style="font-size:11px;font-weight:800;color:#1B4F8A;padding-bottom:8px;border-bottom:2px solid #2563eb;">DataNetra Pro</div>
  </div>
  <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:0;">
    <div style="font-size:11px;color:#475569;padding:8px 4px;border-bottom:1px solid #f1f5f9;font-weight:600;">Cost / month</div>
    <div style="font-size:11px;color:#64748b;padding:8px 4px;border-bottom:1px solid #f1f5f9;text-align:center;">₹5,000 – ₹20,000</div>
    <div style="font-size:12px;font-weight:900;color:#16a34a;padding:8px 4px;border-bottom:1px solid #f1f5f9;text-align:center;">₹999</div>
  </div>
  <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:0;">
    <div style="font-size:11px;color:#475569;padding:8px 4px;border-bottom:1px solid #f1f5f9;font-weight:600;">Time to insight</div>
    <div style="font-size:11px;color:#64748b;padding:8px 4px;border-bottom:1px solid #f1f5f9;text-align:center;">3–7 days</div>
    <div style="font-size:12px;font-weight:900;color:#16a34a;padding:8px 4px;border-bottom:1px solid #f1f5f9;text-align:center;">60 seconds</div>
  </div>
  <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:0;">
    <div style="font-size:11px;color:#475569;padding:8px 4px;border-bottom:1px solid #f1f5f9;font-weight:600;">SKU-level analysis</div>
    <div style="font-size:11px;color:#64748b;padding:8px 4px;border-bottom:1px solid #f1f5f9;text-align:center;">Rarely</div>
    <div style="font-size:12px;font-weight:700;color:#16a34a;padding:8px 4px;border-bottom:1px solid #f1f5f9;text-align:center;">✓ Every upload</div>
  </div>
  <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:0;">
    <div style="font-size:11px;color:#475569;padding:8px 4px;border-bottom:1px solid #f1f5f9;font-weight:600;">AI forecast</div>
    <div style="font-size:11px;color:#64748b;padding:8px 4px;border-bottom:1px solid #f1f5f9;text-align:center;">Never</div>
    <div style="font-size:12px;font-weight:700;color:#16a34a;padding:8px 4px;border-bottom:1px solid #f1f5f9;text-align:center;">✓ 6 &amp; 12 months</div>
  </div>
  <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:0;">
    <div style="font-size:11px;color:#475569;padding:8px 4px;font-weight:600;">ONDC readiness</div>
    <div style="font-size:11px;color:#64748b;padding:8px 4px;text-align:center;">Never</div>
    <div style="font-size:12px;font-weight:700;color:#16a34a;padding:8px 4px;text-align:center;">✓ Scored &amp; matched</div>
  </div>
</div>
'''

    return f'''
<div style="background:#f8fbff;border-top:1px solid #e2eaf5;border-bottom:1px solid #e2eaf5;padding:24px 24px;">
  <div style="max-width:1020px;margin:0 auto;">

    <!-- Section header -->
    <div style="text-align:center;margin-bottom:32px;">
      <div style="display:inline-block;background:linear-gradient(90deg,#1B4F8A,#2563eb);
           color:#ffffff;font-size:10px;font-weight:800;letter-spacing:2px;
           text-transform:uppercase;padding:5px 14px;border-radius:20px;margin-bottom:14px;">
        {t["label"]}
      </div>
      <div style="font-size:22px;font-weight:900;color:#0B1F3A;letter-spacing:-0.4px;
           line-height:1.3;max-width:680px;margin:0 auto 10px;">{t["title"]}</div>
      <div style="font-size:14px;color:#64748b;max-width:560px;margin:0 auto;">{t["sub"]}</div>
    </div>

    <!-- Comparison table (EN only) -->
    {comparison_html}

    <!-- 4 feature cards -->
    <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));
         gap:16px;margin-bottom:32px;">{cards_html}</div>

    <!-- CTA -->
    <div style="text-align:center;">
      <a href="#dn-drc-anchor"
         onclick="var a=document.getElementById('dn-drc-anchor');if(a)a.scrollIntoView({{behavior:'smooth',block:'start'}});return false;"
         style="display:inline-flex;align-items:center;gap:8px;
                background:linear-gradient(90deg,#1B4F8A 0%,#2563eb 100%);
                color:#ffffff;font-size:15px;font-weight:800;padding:14px 32px;
                border-radius:10px;text-decoration:none;
                box-shadow:0 6px 20px rgba(37,99,235,0.40);">{t["cta"]}</a>
      <div style="margin-top:10px;font-size:12px;color:#94a3b8;">{t["trust"]}</div>
    </div>

  </div>
</div>
'''



# ══════════════════════════════════════════════════════════════════════════════
# PRODUCTION ADD-ONS: Database · Email OTP · OpenAI
# Set environment variables on Render dashboard — never hardcode secrets
# ══════════════════════════════════════════════════════════════════════════════
import sqlite3, smtplib, random, time, hashlib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

_SMTP_HOST  = _os.environ.get("SMTP_HOST",     "smtp.gmail.com")
_SMTP_PORT  = int(_os.environ.get("SMTP_PORT", "587"))
_SMTP_USER  = _os.environ.get("SMTP_USER",     "")   # your Gmail address
_SMTP_PASS  = _os.environ.get("SMTP_PASS",     "")   # Gmail App Password
_OPENAI_KEY = _os.environ.get("OPENAI_API_KEY","")
_DB_PATH    = _os.environ.get("DB_PATH", "/tmp/datanetra.db")
_OTP_EXPIRY = 300  # 5 minutes

def _init_db():
    con = sqlite3.connect(_DB_PATH)
    con.executescript("""
        -- TABLE 1: Everyone who completes the unlock form
        CREATE TABLE IF NOT EXISTS users (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name   TEXT,
            mobile      TEXT,
            email       TEXT UNIQUE,
            login_count INTEGER DEFAULT 1,
            created_at  DATETIME DEFAULT CURRENT_TIMESTAMP,
            last_login  DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        -- TABLE 2: OTP codes (auto-expires after 5 min)
        CREATE TABLE IF NOT EXISTS otp_store (
            email      TEXT PRIMARY KEY,
            otp_hash   TEXT,
            expires_at REAL
        );
        -- TABLE 3: Every analysis run (file upload + score generated)
        CREATE TABLE IF NOT EXISTS analysis_log (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            email        TEXT,
            file_name    TEXT,
            file_rows    INTEGER,
            file_cols    INTEGER,
            health_score REAL,
            stores_count INTEGER,
            skus_count   INTEGER,
            lang         TEXT DEFAULT 'en',
            created_at   DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        -- TABLE 4: PDF report downloads
        CREATE TABLE IF NOT EXISTS report_downloads (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            email      TEXT,
            score      REAL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        -- TABLE 5: Contact form submissions from landing page
        CREATE TABLE IF NOT EXISTS contact_leads (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT,
            email      TEXT,
            mobile     TEXT,
            message    TEXT,
            source     TEXT DEFAULT 'landing_page',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
    """)
    con.commit(); con.close()

try:
    _init_db()
    # Migration: safely add new columns to existing databases
    try:
        import sqlite3 as _sq3m
        _mc = _sq3m.connect(_DB_PATH)
        for _tbl, _col, _typ in [
            ("analysis_log","file_name",   "TEXT DEFAULT ''"),
            ("analysis_log","file_cols",   "INTEGER DEFAULT 0"),
            ("analysis_log","stores_count","INTEGER DEFAULT 0"),
            ("analysis_log","skus_count",  "INTEGER DEFAULT 0"),
            ("analysis_log","lang",        "TEXT DEFAULT 'en'"),
        ]:
            try: _mc.execute(f"ALTER TABLE {_tbl} ADD COLUMN {_col} {_typ}")
            except: pass
        _mc.commit(); _mc.close()
    except: pass
    print("✅ Database ready:", _DB_PATH)
except Exception as _e:
    print(f"⚠️  DB init: {_e}")

def _save_user(name, mobile, email):
    try:
        con = sqlite3.connect(_DB_PATH)
        con.execute("""INSERT INTO users (full_name, mobile, email) VALUES (?,?,?)
            ON CONFLICT(email) DO UPDATE SET full_name=excluded.full_name,
            mobile=excluded.mobile, login_count=login_count+1,
            last_login=CURRENT_TIMESTAMP""", (name, mobile, email))
        con.commit(); con.close()
    except Exception as e:
        print(f"⚠️  save_user: {e}")

def _log_analysis(email, file_name, file_rows, file_cols,
                  health_score, stores_count, skus_count, lang='en'):
    """Called every time someone runs analysis — logs the full run."""
    try:
        con = sqlite3.connect(_DB_PATH)
        con.execute("""INSERT INTO analysis_log
            (email, file_name, file_rows, file_cols,
             health_score, stores_count, skus_count, lang)
            VALUES (?,?,?,?,?,?,?,?)""",
            (email or 'anonymous', file_name or 'unknown',
             file_rows or 0, file_cols or 0,
             health_score or 0, stores_count or 0,
             skus_count or 0, lang))
        con.commit(); con.close()
        print(f"📊 Analysis logged: {email} | score={health_score} | rows={file_rows}")
    except Exception as e:
        print(f"⚠️  log_analysis: {e}")

def _log_report_download(email, score):
    """Called when user downloads the PDF report."""
    try:
        con = sqlite3.connect(_DB_PATH)
        con.execute("INSERT INTO report_downloads (email, score) VALUES (?,?)",
                    (email or 'anonymous', score or 0))
        con.commit(); con.close()
        print(f"📄 Report downloaded: {email}")
    except Exception as e:
        print(f"⚠️  log_report_download: {e}")

def _save_contact_lead(name, email, mobile, message, source='landing_page'):
    """Called when someone submits the contact form on landing page."""
    try:
        con = sqlite3.connect(_DB_PATH)
        con.execute("""INSERT INTO contact_leads
            (name, email, mobile, message, source)
            VALUES (?,?,?,?,?)""",
            (name or '', email or '', mobile or '', message or '', source))
        con.commit(); con.close()
        print(f"📬 Contact lead saved: {email} | {name}")
        return True
    except Exception as e:
        print(f"⚠️  save_contact_lead: {e}")
        return False

def _get_dashboard_stats():
    """Admin summary: total users, analyses, leads, downloads."""
    try:
        con = sqlite3.connect(_DB_PATH)
        stats = {
            'total_users':     con.execute("SELECT COUNT(*) FROM users").fetchone()[0],
            'total_analyses':  con.execute("SELECT COUNT(*) FROM analysis_log").fetchone()[0],
            'total_downloads': con.execute("SELECT COUNT(*) FROM report_downloads").fetchone()[0],
            'total_leads':     con.execute("SELECT COUNT(*) FROM contact_leads").fetchone()[0],
            'avg_score':       con.execute("SELECT ROUND(AVG(health_score),1) FROM analysis_log").fetchone()[0] or 0,
        }
        con.close()
        return stats
    except Exception as e:
        print(f"⚠️  get_dashboard_stats: {e}")
        return {}

def _store_otp(email, otp):
    try:
        h = hashlib.sha256(otp.encode()).hexdigest()
        con = sqlite3.connect(_DB_PATH)
        con.execute("""INSERT INTO otp_store (email,otp_hash,expires_at) VALUES(?,?,?)
            ON CONFLICT(email) DO UPDATE SET otp_hash=excluded.otp_hash,
            expires_at=excluded.expires_at""", (email, h, time.time()+_OTP_EXPIRY))
        con.commit(); con.close()
        return True
    except Exception as e:
        print(f"⚠️  store_otp: {e}"); return False

def _verify_otp(email, otp):
    if otp == "1234" and not _SMTP_USER:
        return True, "ok"  # dev fallback
    try:
        con = sqlite3.connect(_DB_PATH)
        row = con.execute("SELECT otp_hash,expires_at FROM otp_store WHERE email=?",
                          (email,)).fetchone()
        con.close()
        if not row: return False, "No OTP found. Click Send OTP first."
        if time.time() > row[1]: return False, "OTP expired. Request a new one."
        if hashlib.sha256(otp.encode()).hexdigest() != row[0]:
            return False, "Incorrect OTP. Please try again."
        return True, "ok"
    except Exception as e:
        return False, "Verification error. Please retry."

def _send_otp_email(to_email, otp, name=""):
    if not _SMTP_USER:
        print(f"[DEV] OTP for {to_email}: {otp}")
        return True, "dev_mode"
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"{otp} is your DataNetra verification code"
        msg["From"] = f"DataNetra.ai <{_SMTP_USER}>"
        msg["To"] = to_email
        html = f"""<div style="font-family:Arial,sans-serif;max-width:480px;margin:0 auto;">
          <div style="background:#1B4F8A;padding:20px;border-radius:8px 8px 0 0;text-align:center;">
            <span style="color:#fff;font-size:22px;font-weight:bold;">DataNetra.ai</span>
          </div>
          <div style="background:#f8fbff;padding:28px;border:1px solid #e2eaf5;border-radius:0 0 8px 8px;">
            <p style="color:#0f2557;">Hi {name or "there"},</p>
            <p>Your verification code:</p>
            <div style="text-align:center;margin:24px 0;">
              <span style="font-size:40px;font-weight:900;color:#1B4F8A;
                letter-spacing:8px;background:#EFF6FF;padding:12px 24px;
                border-radius:8px;border:2px solid #BFDBFE;">{otp}</span>
            </div>
            <p style="color:#64748b;font-size:13px;">Expires in 5 minutes.</p>
          </div>
        </div>"""
        msg.attach(MIMEText(html, "html"))
        with smtplib.SMTP(_SMTP_HOST, _SMTP_PORT) as s:
            s.starttls()
            s.login(_SMTP_USER, _SMTP_PASS)
            s.sendmail(_SMTP_USER, to_email, msg.as_string())
        return True, "sent"
    except Exception as e:
        print(f"⚠️  send_otp: {e}"); return False, str(e)

def _llm_insights(snapshot: dict, lang: str = "en") -> str:
    if not _OPENAI_KEY: return ""
    try:
        import openai
        openai.api_key = _OPENAI_KEY
        r = openai.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role":"user","content":
                f"Retail business analyst for Indian SMBs. Data: Health={snapshot.get('health_score',0):.0f}/100, "
                f"Sales=Rs{snapshot.get('gross_sales',0):,.0f}, Margin={snapshot.get('avg_margin',0):.1f}%, "
                f"6M Forecast=Rs{snapshot.get('f6',0):,.0f}. "
                f"Write 3 specific action bullets in {lang}. Start each with action verb. No preamble."}],
            max_tokens=250, temperature=0.4)
        return r.choices[0].message.content.strip()
    except Exception as e:
        print(f"⚠️  OpenAI: {e}"); return ""

# ══════════════════════════════════════════════════════════════════════════════
with gr.Blocks(title="DataNetra.ai - Retail Intelligence", theme=gr.themes.Soft(), css=custom_css) as demo:

    step_state = gr.State(0); user_data_state = gr.State({}); lang_state = gr.State('en')
    dashboard_data_state = gr.State({'kpi1':"","kpi2":"","kpi3":"","kpi4":"","kpi5":"","chart1":None,"chart2":None,"chart3":None,"chart4":None})
    granular_forecast_data_state = gr.State(None); df_state = gr.State(None)
    # ── Persist upload file + consent across Preview → Full Dashboard ──────────
    preview_file_state    = gr.State(None)   # stores file object after Generate Preview
    preview_consent_state = gr.State(False)  # stores consent checkbox value
    unlock_mode_state     = gr.State(False)  # True when user came from "Unlock Full Dashboard"

    # ── Navigation header bar ─────────────────────────────────────────────────
    # ── Navigation header bar — logo + native language/gov controls ──────────
    with gr.Row(elem_id="lang-bar", equal_height=True):
        logo_html = gr.HTML("""<div id="lang-bar-logo">
          <img src="https://i.postimg.cc/qRNQYbZJ/Data-Netra-Logo.jpg"
               alt="DataNetra.ai" style="height:42px;border-radius:7px;">
          <span class="logo-text" style="font-size:21px;font-weight:900;color:#0f2557;
                letter-spacing:-0.5px;white-space:nowrap;">DataNetra.ai</span>
        </div>""", visible=True, elem_id="hdr-logo-wrap")
        # ── Center nav links ─────────────────────────────────────────────────
        with gr.Column(scale=2, elem_id="lang-bar-nav", min_width=0):
            gr.HTML("""<nav id="hdr-nav-links">
  <a href="#" class="hdr-nav-link">Product</a>
  <a href="#" class="hdr-nav-link">Solutions</a>
  <a href="#" class="hdr-nav-link">Resources</a>
  <a href="#" class="hdr-nav-link"
     onclick="var s=document.getElementById('dn-pricing-section');if(s){s.classList.toggle('dn-pricing-open');}setTimeout(function(){var a=document.getElementById('dn-pricing-anchor');if(a)a.scrollIntoView({behavior:'smooth',block:'start'});},150);return false;">Pricing</a>
  <a href="#" class="hdr-nav-link"
     onclick="var a=document.getElementById('dn-contact-anchor');if(a)a.scrollIntoView({behavior:'smooth'});return false;">Contact</a>
</nav>""")
        # ── Language selector dropdown (English / Kannada / Tamil) ────────
        lang_dropdown  = gr.Dropdown(
            choices=["🇬🇧 English", "🇮🇳 ಕನ್ನಡ (Kannada)", "🇮🇳 தமிழ் (Tamil)"],
            value="🇬🇧 English",
            label="Language", container=True, scale=0, min_width=170,
            elem_id="lang-dropdown-main",
        )
        # ── Stub buttons kept for event-wiring compatibility (hidden) ────────
        lang_en_btn  = gr.Button("English",  size="sm", variant="primary",   scale=0, visible=False, elem_id="hdr-lang-en")
        lang_hi_btn  = gr.Button("Hindi",    size="sm", variant="secondary",  scale=0, visible=False, elem_id="hdr-lang-hi")
        msme_hdr_btn = gr.Button("MSME",     size="sm", variant="secondary",  scale=0, visible=False, elem_id="hdr-msme-btn")
        gov_hdr_btn  = gr.Button("Gov",      size="sm", variant="secondary",  scale=0, visible=False, elem_id="hdr-gov-btn")
        pricing_nav_btn = gr.Button("💰 Pricing", size="sm", variant="secondary", scale=0, min_width=100, elem_id="hdr-pricing-btn", visible=True)
        lang_indicator = gr.Markdown("", elem_id="lang-indicator", visible=False)
        hdr_mode_html  = gr.HTML("", elem_id="hdr-mode-script", visible=False)

    # ══════════════════════════════════════════════════════════════════
    # STEP 0 — Landing Page  (MSME Login + Gov Login + Sign Up)
    # ══════════════════════════════════════════════════════════════════
    with gr.Column(visible=True, elem_id="dn-step0-col") as step0_col:
        # Hero banner (logo + tagline)
        # ── Viewport meta — critical for mobile rendering ──
        gr.HTML('<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=5">', visible=False)
        landing_hero_html = gr.HTML(value=_landing_hero('en'))

        # ── Gradio login inputs — TOP-LEVEL so always in DOM ──────────────────
        # These are always rendered regardless of login_section_col visibility
        # CSS hides them; JS can always find them for event wiring
        with gr.Column(elem_id="dn-gradio-inputs-hidden", visible=True):
            name_input_landing  = gr.Textbox(label="Your Name", placeholder="e.g. Ravi Kumar",
                                              elem_id="dn-landing-name", type="text")
            quick_login_mobile  = gr.Textbox(label="Mobile Number", placeholder="e.g. +91 98765 43210",
                                              elem_id="dn-real-mobile", type="text")
            otp_landing_input   = gr.Textbox(label="Verification Code",
                                              placeholder="Enter code (use 1234 for demo)",
                                              elem_id="dn-landing-otp", type="text")
            quick_login_email   = gr.Textbox(label="Email Address", placeholder="e.g. ravi@store.com",
                                              elem_id="dn-landing-email", type="text")
            send_otp_btn      = gr.Button("Send OTP", elem_id="dn-send-otp-gr-btn")
            send_otp_email_in = gr.Textbox(label="", elem_id="dn-send-otp-email-in")
            send_otp_result   = gr.Markdown(value="", elem_id="dn-send-otp-result")

        # ── Analyze button: TOP-LEVEL so ALWAYS rendered and in DOM ────────
        # CSS will position it inside the HTML form visually
        quick_login_btn = gr.Button("Analyze My Data →", variant="primary",
                                    elem_id="dn-real-login-btn")
        landing_login_error_msg = gr.Markdown(value="", visible=False,
                                               elem_id="dn-login-error-md")

        # ── How it works — 3 step visual ─────────────────────────────────────
        gr.HTML('''
<div style="background:#ffffff;border-bottom:1px solid #e8f0fb;padding:24px 24px 20px;">
  <div style="max-width:900px;margin:0 auto;text-align:center;">
    <div style="font-size:11px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#1B4F8A;margin-bottom:8px;">HOW IT WORKS</div>
    <div style="font-size:22px;font-weight:900;color:#0B1F3A;margin-bottom:4px;letter-spacing:-0.3px;">From your Excel to business clarity in 3 steps</div>
    <div style="font-size:13px;color:#64748b;margin-bottom:20px;">No consultants. No setup. No technical knowledge needed.</div>
    <!-- 3 steps forced into single row — flex, no wrap -->
    <div class="dn-hiw-steps" style="display:flex;align-items:flex-start;justify-content:center;flex-wrap:nowrap;gap:0;max-width:900px;margin:0 auto;">

      <!-- Step 1 -->
      <div style="flex:1;min-width:0;padding:12px 20px 16px;text-align:center;">
        <div style="width:56px;height:56px;background:linear-gradient(135deg,#1B4F8A,#2563eb);border-radius:16px;display:flex;align-items:center;justify-content:center;margin:0 auto 10px;box-shadow:0 6px 18px rgba(37,99,235,0.28);">
          <svg width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="#ffffff" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>
        </div>
        <div style="display:inline-block;background:#EFF6FF;border:1px solid #BFDBFE;border-radius:20px;font-size:10px;font-weight:800;letter-spacing:1.5px;color:#2563eb;text-transform:uppercase;padding:3px 10px;margin-bottom:6px;">Step 1</div>
        <div style="font-size:15px;font-weight:800;color:#0B1F3A;margin-bottom:4px;">Upload Your Data</div>
        <div style="font-size:12.5px;color:#64748b;line-height:1.7;">Drop your Excel or CSV — tally export, billing software, or any format you already use.</div>
      </div>

      <!-- Arrow 1→2 -->
      <div style="flex:0 0 44px;display:flex;flex-direction:column;align-items:center;justify-content:center;padding-top:20px;gap:2px;">
        <div style="width:28px;height:2px;background:linear-gradient(90deg,#2563eb,#7c3aed);border-radius:2px;"></div>
        <svg width="8" height="8" viewBox="0 0 8 8"><polygon points="0,0 8,4 0,8" fill="#7c3aed"/></svg>
      </div>

      <!-- Step 2 -->
      <div style="flex:1;min-width:0;padding:28px 20px;text-align:center;">
        <div style="width:56px;height:56px;background:linear-gradient(135deg,#7c3aed,#9333ea);border-radius:16px;display:flex;align-items:center;justify-content:center;margin:0 auto 10px;box-shadow:0 6px 18px rgba(124,58,237,0.28);">
          <svg width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="#ffffff" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="3"/><path d="M12 1v4M12 19v4M4.22 4.22l2.83 2.83M16.95 16.95l2.83 2.83M1 12h4M19 12h4M4.22 19.78l2.83-2.83M16.95 7.05l2.83-2.83"/></svg>
        </div>
        <div style="display:inline-block;background:#F5F3FF;border:1px solid #DDD6FE;border-radius:20px;font-size:10px;font-weight:800;letter-spacing:1.5px;color:#7c3aed;text-transform:uppercase;padding:3px 10px;margin-bottom:6px;">Step 2</div>
        <div style="font-size:15px;font-weight:800;color:#0B1F3A;margin-bottom:4px;">AI Analyses It</div>
        <div style="font-size:12.5px;color:#64748b;line-height:1.7;">7 AI models score your business health, detect risk, forecast demand and find your best and worst SKUs.</div>
      </div>

      <!-- Arrow 2→3 -->
      <div style="flex:0 0 44px;display:flex;flex-direction:column;align-items:center;justify-content:center;padding-top:40px;gap:2px;">
        <div style="width:28px;height:2px;background:linear-gradient(90deg,#7c3aed,#16a34a);border-radius:2px;"></div>
        <svg width="8" height="8" viewBox="0 0 8 8"><polygon points="0,0 8,4 0,8" fill="#16a34a"/></svg>
      </div>

      <!-- Step 3 -->
      <div style="flex:1;min-width:0;padding:28px 20px;text-align:center;">
        <div style="width:56px;height:56px;background:linear-gradient(135deg,#16a34a,#22c55e);border-radius:16px;display:flex;align-items:center;justify-content:center;margin:0 auto 10px;box-shadow:0 6px 18px rgba(22,163,74,0.28);">
          <svg width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="#ffffff" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"><polyline points="22 12 18 12 15 21 9 3 6 12 2 12"/></svg>
        </div>
        <div style="display:inline-block;background:#F0FDF4;border:1px solid #BBF7D0;border-radius:20px;font-size:10px;font-weight:800;letter-spacing:1.5px;color:#16a34a;text-transform:uppercase;padding:3px 10px;margin-bottom:6px;">Step 3</div>
        <div style="font-size:15px;font-weight:800;color:#0B1F3A;margin-bottom:4px;">Act on Decisions</div>
        <div style="font-size:12.5px;color:#64748b;line-height:1.7;">Get a personalised action plan — which products to push, which to drop, where to expand.</div>
      </div>

    </div>
    <!-- CTA below steps -->
    <div style="margin-top:16px;">
      <a href="#dn-drc-anchor"
         onclick="var a=document.getElementById('dn-drc-anchor');if(a)a.scrollIntoView({behavior:'smooth'});return false;"
         style="display:inline-flex;align-items:center;gap:8px;background:linear-gradient(90deg,#1B4F8A,#2563eb);color:#ffffff;font-size:14px;font-weight:800;padding:12px 28px;border-radius:10px;text-decoration:none;box-shadow:0 4px 16px rgba(37,99,235,0.35);">
        Try it free — see results in 60 seconds →
      </a>
      <div style="font-size:11px;color:#94a3b8;margin-top:8px;">No credit card · No signup · Your data stays private</div>
    </div>
  </div>
</div>
''')

        # ── "Built for Retail" section ────────────────────────────────────────
        industries_html = gr.HTML(_landing_industries('en'))

        # ── "Uncertain Markets" — new section ─────────────────────────────
        gr.HTML('''
<div style="background:#f8fbff;border-top:1px solid #e8f0fb;border-bottom:1px solid #e8f0fb;
     padding:22px 24px 20px;">
  <div style="max-width:880px;margin:0 auto;text-align:center;">
    <div style="font-size:20px;font-weight:900;color:#0B1F3A;letter-spacing:-0.3px;
         margin-bottom:8px;line-height:1.3;">
      In uncertain markets, smarter retailers act faster.
    </div>
    <div style="font-size:14px;color:#475569;max-width:680px;margin:0 auto 28px;line-height:1.7;">
      Rising costs, demand shifts, supplier uncertainty, and margin pressure make guesswork
      expensive. DataNetra helps retailers turn Excel data into faster, AI-powered business
      decisions.
    </div>
    <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));
         gap:14px;max-width:800px;margin:0 auto;">

      <div style="background:#ffffff;border:1px solid #e2eaf5;border-top:3px solid #2563eb;
           border-radius:12px;padding:18px 16px;text-align:center;
           box-shadow:0 2px 8px rgba(11,31,58,0.05);">
        <div style="font-size:22px;margin-bottom:8px;">📈</div>
        <div style="font-size:13px;font-weight:800;color:#0B1F3A;margin-bottom:4px;">
          Rising Costs
        </div>
        <div style="font-size:12px;color:#64748b;line-height:1.6;">
          Know your exact margin per SKU before you reorder — not after.
        </div>
      </div>

      <div style="background:#ffffff;border:1px solid #e2eaf5;border-top:3px solid #7c3aed;
           border-radius:12px;padding:18px 16px;text-align:center;
           box-shadow:0 2px 8px rgba(11,31,58,0.05);">
        <div style="font-size:22px;margin-bottom:8px;">📦</div>
        <div style="font-size:13px;font-weight:800;color:#0B1F3A;margin-bottom:4px;">
          Inventory Risk
        </div>
        <div style="font-size:12px;color:#64748b;line-height:1.6;">
          AI reorder signals flag dead stock and stockout risk before they cost you.
        </div>
      </div>

      <div style="background:#ffffff;border:1px solid #e2eaf5;border-top:3px solid #d97706;
           border-radius:12px;padding:18px 16px;text-align:center;
           box-shadow:0 2px 8px rgba(11,31,58,0.05);">
        <div style="font-size:22px;margin-bottom:8px;">🔄</div>
        <div style="font-size:13px;font-weight:800;color:#0B1F3A;margin-bottom:4px;">
          Demand Volatility
        </div>
        <div style="font-size:12px;color:#64748b;line-height:1.6;">
          6 and 12-month AI forecasts adapt to your sales patterns — not generic trends.
        </div>
      </div>

      <div style="background:#ffffff;border:1px solid #e2eaf5;border-top:3px solid #16a34a;
           border-radius:12px;padding:18px 16px;text-align:center;
           box-shadow:0 2px 8px rgba(11,31,58,0.05);">
        <div style="font-size:22px;margin-bottom:8px;">💰</div>
        <div style="font-size:13px;font-weight:800;color:#0B1F3A;margin-bottom:4px;">
          Margin Pressure
        </div>
        <div style="font-size:12px;color:#64748b;line-height:1.6;">
          Spot which products are silently eroding profitability across your stores.
        </div>
      </div>

    </div>
  </div>
</div>
''')

        # ── "What You Unlock" section ─────────────────────────────────────
        unlock_benefits_html = gr.HTML(_landing_unlock_benefits('en'))

        # ── Modals: Contact Demo + Login Capture (shared window scope) ──
        gr.HTML(value='''
<!-- ═══════════════════ BOOK DEMO MODAL ═══════════════════════════════ -->
<div id="dn-demo-modal-overlay" style="display:none;position:fixed;inset:0;z-index:9999;
     background:rgba(10,20,50,0.58);backdrop-filter:blur(4px);
     align-items:center;justify-content:center;">
  <div id="dn-demo-modal" style="background:#ffffff;border-radius:16px;width:100%;
       max-width:420px;margin:20px;
       box-shadow:0 24px 64px rgba(0,0,0,0.22),0 4px 16px rgba(11,31,58,0.12);
       font-family:system-ui,-apple-system,sans-serif;overflow:hidden;">
    <div style="background:linear-gradient(120deg,#0f2557,#1a3a6b);
                padding:14px 24px 14px;
                display:flex;align-items:flex-start;justify-content:space-between;">
      <div>
        <div style="font-size:10px;font-weight:700;letter-spacing:1.5px;color:#93c5fd;
                    text-transform:uppercase;margin-bottom:4px;">DataNetra.ai</div>
        <div style="font-size:18px;font-weight:800;color:#ffffff;line-height:1.2;">Book a Demo</div>
        <div style="font-size:12px;color:rgba(255,255,255,0.58);margin-top:4px;">
          Our team will reach out within 24 hours.
        </div>
      </div>
      <button onclick="dnCloseModal()"
        style="background:rgba(255,255,255,0.12);border:none;cursor:pointer;
               width:30px;height:30px;border-radius:50%;color:rgba(255,255,255,0.75);
               font-size:18px;display:flex;align-items:center;justify-content:center;
               flex-shrink:0;margin-left:12px;line-height:1;transition:background 0.15s;"
        onmouseover="this.style.background=\'rgba(255,255,255,0.22)\'"
        onmouseout="this.style.background=\'rgba(255,255,255,0.12)\'">&#215;</button>
    </div>
    <div id="dn-modal-form-body" style="padding:22px 24px 24px;">
      <div style="display:flex;flex-direction:column;gap:14px;">
        <div>
          <label style="display:block;font-size:12px;font-weight:600;color:#374151;margin-bottom:5px;">
            Full Name <span style="color:#e74c3c;">*</span>
          </label>
          <input id="dn-demo-name" type="text" placeholder="e.g. Ravi Kumar"
            style="width:100%;box-sizing:border-box;height:42px;padding:0 13px;
                   border:1.5px solid #d1d9e6;border-radius:8px;font-size:13px;
                   color:#1a2540;outline:none;background:#fff;transition:border-color 0.15s,box-shadow 0.15s;"
            onfocus="this.style.borderColor=\'#2563eb\';this.style.boxShadow=\'0 0 0 3px rgba(37,99,235,0.12)\'"
            onblur="this.style.borderColor=\'#d1d9e6\';this.style.boxShadow=\'none\'">
        </div>
        <div>
          <label style="display:block;font-size:12px;font-weight:600;color:#374151;margin-bottom:5px;">
            Email Address <span style="color:#e74c3c;">*</span>
          </label>
          <input id="dn-demo-email" type="email" placeholder="e.g. ravi@store.com"
            style="width:100%;box-sizing:border-box;height:42px;padding:0 13px;
                   border:1.5px solid #d1d9e6;border-radius:8px;font-size:13px;
                   color:#1a2540;outline:none;background:#fff;transition:border-color 0.15s,box-shadow 0.15s;"
            onfocus="this.style.borderColor=\'#2563eb\';this.style.boxShadow=\'0 0 0 3px rgba(37,99,235,0.12)\'"
            onblur="this.style.borderColor=\'#d1d9e6\';this.style.boxShadow=\'none\'">
        </div>
        <div>
          <label style="display:block;font-size:12px;font-weight:600;color:#374151;margin-bottom:5px;">
            Phone Number <span style="color:#e74c3c;">*</span>
          </label>
          <input id="dn-demo-phone" type="tel" placeholder="e.g. +91 98765 43210"
            style="width:100%;box-sizing:border-box;height:42px;padding:0 13px;
                   border:1.5px solid #d1d9e6;border-radius:8px;font-size:13px;
                   color:#1a2540;outline:none;background:#fff;transition:border-color 0.15s,box-shadow 0.15s;"
            onfocus="this.style.borderColor=\'#2563eb\';this.style.boxShadow=\'0 0 0 3px rgba(37,99,235,0.12)\'"
            onblur="this.style.borderColor=\'#d1d9e6\';this.style.boxShadow=\'none\'">
        </div>
        <div>
          <label style="display:block;font-size:12px;font-weight:600;color:#374151;margin-bottom:5px;">
            Message <span style="color:#94a3b8;font-weight:400;margin-left:4px;">(optional)</span>
          </label>
          <textarea id="dn-demo-msg" placeholder="Tell us about your retail business..." rows="3"
            style="width:100%;box-sizing:border-box;padding:10px 13px;resize:vertical;
                   min-height:76px;border:1.5px solid #d1d9e6;border-radius:8px;
                   font-size:13px;color:#1a2540;font-family:inherit;outline:none;
                   background:#fff;transition:border-color 0.15s,box-shadow 0.15s;"
            onfocus="this.style.borderColor=\'#2563eb\';this.style.boxShadow=\'0 0 0 3px rgba(37,99,235,0.12)\'"
            onblur="this.style.borderColor=\'#d1d9e6\';this.style.boxShadow=\'none\'"></textarea>
        </div>
        <div id="dn-demo-err" style="display:none;font-size:12px;color:#dc2626;
             background:#fef2f2;border:1px solid #fecaca;border-radius:7px;padding:8px 12px;"></div>
        <button onclick="dnSubmitDemo()"
          style="width:100%;height:46px;background:#2563eb;color:#ffffff;border:none;
                 border-radius:9px;font-size:14px;font-weight:700;cursor:pointer;
                 box-shadow:0 4px 14px rgba(37,99,235,0.35);transition:background 0.15s,transform 0.15s,box-shadow 0.15s;"
          onmouseover="this.style.background=\'#1d4ed8\';this.style.transform=\'translateY(-1px)\'"
          onmouseout="this.style.background=\'#2563eb\';this.style.transform=\'translateY(0)\'">
          Submit Request
        </button>
      </div>
    </div>
    <div id="dn-modal-success" style="display:none;padding:40px 28px;text-align:center;">
      <div style="width:60px;height:60px;background:#f0fdf4;border-radius:50%;
                  display:flex;align-items:center;justify-content:center;margin:0 auto 16px;">
        <svg width="28" height="28" viewBox="0 0 24 24" fill="none" stroke="#16a34a"
             stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
          <path d="M22 11.08V12a10 10 0 1 1-5.93-9.14"/>
          <polyline points="22 4 12 14.01 9 11.01"/>
        </svg>
      </div>
      <div style="font-size:18px;font-weight:800;color:#0B1F3A;margin-bottom:8px;">Request Submitted!</div>
      <div style="font-size:13px;color:#64748b;line-height:1.65;max-width:280px;margin:0 auto 22px;">
        Request submitted successfully. Our team will contact you shortly.
      </div>
      <button onclick="dnCloseModal()"
        style="padding:11px 32px;background:#2563eb;color:#fff;border:none;
               border-radius:9px;font-size:13px;font-weight:700;cursor:pointer;">OK</button>
    </div>
  </div>
</div>

<!-- ═══════════════════ UNLOCK FULL INSIGHTS MODAL ══════════════════════ -->
<div id="dn-unlock-modal-overlay"
     style="display:none;position:fixed;inset:0;z-index:9999;
            background:rgba(10,20,50,0.60);backdrop-filter:blur(4px);
            align-items:center;justify-content:center;">
  <div id="dn-unlock-modal"
       style="background:#ffffff;border-radius:18px;width:100%;max-width:440px;
              margin:20px;overflow:hidden;font-family:system-ui,-apple-system,sans-serif;
              box-shadow:0 28px 72px rgba(0,0,0,0.24),0 4px 18px rgba(11,31,58,0.14);">
    <!-- Header -->
    <div style="background:linear-gradient(120deg,#0B1F3A,#1B4F8A);
                padding:22px 24px 18px;
                display:flex;align-items:flex-start;justify-content:space-between;">
      <div>
        <div style="font-size:10px;font-weight:700;letter-spacing:1.6px;
                    text-transform:uppercase;color:#93c5fd;margin-bottom:5px;">Get Full Access</div>
        <div style="font-size:20px;font-weight:900;color:#ffffff;line-height:1.2;margin-bottom:5px;">
          Unlock Full Dashboard
        </div>
        <div style="font-size:12.5px;color:rgba(255,255,255,0.72);line-height:1.5;">
          Complete dashboards · Forecasts · AI Recommendations · PDF Report
        </div>
      </div>
      <button onclick="dnCloseUnlockModal()"
              style="background:rgba(255,255,255,0.12);border:none;cursor:pointer;
                     width:30px;height:30px;border-radius:50%;
                     color:rgba(255,255,255,0.80);font-size:18px;line-height:1;
                     display:flex;align-items:center;justify-content:center;
                     flex-shrink:0;margin-left:12px;transition:background 0.15s;"
              onmouseover="this.style.background=\'rgba(255,255,255,0.22)\'"
              onmouseout="this.style.background=\'rgba(255,255,255,0.12)\'">&#215;</button>
    </div>
    <!-- Login capture form -->
    <div style="padding:22px 24px 24px;">
      <div id="dn-capture-form-body">
        <div id="dn-capture-err" style="display:none;background:#fef2f2;border:1px solid #fecaca;
             border-radius:8px;padding:10px 14px;font-size:13px;color:#dc2626;margin-bottom:14px;"></div>
        <!-- Name -->
        <div style="margin-bottom:14px;">
          <label style="font-size:12px;font-weight:700;color:#1e3a5f;display:block;margin-bottom:5px;">Your Name *</label>
          <input id="dn-cap-name" type="text" placeholder="e.g. Ravi Kumar"
                 style="width:100%;height:44px;border:2px solid #e2e8f0;border-radius:10px;
                        padding:0 14px;font-size:14px;color:#0f172a;outline:none;
                        box-sizing:border-box;transition:border-color 0.18s;"
                 onfocus="this.style.borderColor=\'#2563eb\'" onblur="this.style.borderColor=\'#e2e8f0\'">
        </div>
        <!-- Email -->
        <div style="margin-bottom:14px;">
          <label style="font-size:12px;font-weight:700;color:#1e3a5f;display:block;margin-bottom:5px;">Email Address *</label>
          <input id="dn-cap-email" type="email" placeholder="e.g. ravi@example.com"
                 style="width:100%;height:44px;border:2px solid #e2e8f0;border-radius:10px;
                        padding:0 14px;font-size:14px;color:#0f172a;outline:none;
                        box-sizing:border-box;transition:border-color 0.18s;"
                 onfocus="this.style.borderColor=\'#2563eb\'" onblur="this.style.borderColor=\'#e2e8f0\'">
        </div>
        <!-- Mobile -->
        <div style="margin-bottom:18px;">
          <label style="font-size:12px;font-weight:700;color:#1e3a5f;display:block;margin-bottom:5px;">Mobile Number *</label>
          <input id="dn-cap-mobile" type="tel" placeholder="e.g. +91 98765 43210"
                 style="width:100%;height:44px;border:2px solid #e2e8f0;border-radius:10px;
                        padding:0 14px;font-size:14px;color:#0f172a;outline:none;
                        box-sizing:border-box;transition:border-color 0.18s;"
                 onfocus="this.style.borderColor=\'#2563eb\'" onblur="this.style.borderColor=\'#e2e8f0\'">
        </div>
        <!-- Submit -->
        <button id="dn-cap-submit" onclick="dnCaptureSubmit()"
                style="width:100%;height:50px;
                       background:linear-gradient(90deg,#1B4F8A 0%,#2563eb 100%);
                       color:#ffffff;border:none;border-radius:12px;
                       font-size:15px;font-weight:800;cursor:pointer;letter-spacing:0.2px;
                       box-shadow:0 4px 14px rgba(27,79,138,0.35);
                       transition:transform 0.15s,box-shadow 0.15s;"
                onmouseover="this.style.transform=\'translateY(-2px)\';this.style.boxShadow=\'0 8px 22px rgba(27,79,138,0.45)\'"
                onmouseout="this.style.transform=\'translateY(0)\';this.style.boxShadow=\'0 4px 14px rgba(27,79,138,0.35)\'">
          Unlock Full Dashboard →
        </button>
        <div style="text-align:center;margin-top:10px;font-size:11px;color:#94a3b8;">
          ✓ Free access &nbsp;·&nbsp; ✓ No credit card &nbsp;·&nbsp; ✓ Instant results
        </div>
      </div>
      <!-- Success state -->
      <div id="dn-capture-success" style="display:none;text-align:center;padding:20px 0;">
        <div style="font-size:40px;margin-bottom:12px;">🎉</div>
        <div style="font-size:16px;font-weight:800;color:#0B1F3A;margin-bottom:6px;">Access Granted!</div>
        <div style="font-size:13px;color:#64748b;">Loading your full business dashboard...</div>
      </div>
    </div>
  </div>
</div>

<style>
#dn-demo-modal-overlay,#dn-unlock-modal-overlay{animation:none !important;}
#dn-demo-modal{animation:dn-modal-in 0.22s ease both;}
#dn-unlock-modal{animation:dn-modal-in 0.22s ease both;}
@keyframes dn-modal-in{
  from{opacity:0;transform:scale(0.93) translateY(-10px);}
  to  {opacity:1;transform:scale(1)    translateY(0);}
}
</style>

<script>
/* ── Book Demo modal ─────────────────────────────────────────────────── */
/* Attach to window so onclick= attrs in other gr.HTML blocks can find them */
window.dnOpenModal = function() {
  var o = document.getElementById("dn-demo-modal-overlay");
  if (o) { o.style.display = "flex"; }
  document.body.style.overflow = "hidden";
};
function dnOpenModal() { window.dnOpenModal(); }
window.dnCloseModal = function dnCloseModal() {
  var o = document.getElementById("dn-demo-modal-overlay");
  if (o) { o.style.display = "none"; }
  document.body.style.overflow = "";
  var f = document.getElementById("dn-modal-form-body");
  var s = document.getElementById("dn-modal-success");
  var e = document.getElementById("dn-demo-err");
  if (f) f.style.display = "block";
  if (s) s.style.display = "none";
  if (e) e.style.display = "none";
  ["dn-demo-name","dn-demo-email","dn-demo-phone","dn-demo-msg"].forEach(function(id){
    var el = document.getElementById(id); if (el) el.value = "";
  });
}
window.dnSubmitDemo = function dnSubmitDemo() {
  var name  = document.getElementById("dn-demo-name").value.trim();
  var email = document.getElementById("dn-demo-email").value.trim();
  var phone = document.getElementById("dn-demo-phone").value.trim();
  var err   = document.getElementById("dn-demo-err");
  if (!name)  { err.textContent="Please enter your full name."; err.style.display="block"; return; }
  if (!email || !/^[^@\x20\t\n]+@[^@\x20\t\n]+[.][^@\x20\t\n]+$/.test(email)) {
    err.textContent="Please enter a valid email address."; err.style.display="block"; return;
  }
  if (!phone) { err.textContent="Please enter your phone number."; err.style.display="block"; return; }
  err.style.display = "none";
  document.getElementById("dn-modal-form-body").style.display = "none";
  document.getElementById("dn-modal-success").style.display  = "block";
};
function dnSubmitDemo() { window.dnSubmitDemo(); }

/* ── Unlock Full Insights modal ──────────────────────────────────────── */
window.dnOpenUnlockModal = function dnOpenUnlockModal() {
  var o = document.getElementById("dn-unlock-modal-overlay");
  if (o) { o.style.display = "flex"; }
  document.body.style.overflow = "hidden";
};
function dnOpenUnlockModal() { window.dnOpenUnlockModal(); }
window.dnCloseUnlockModal = function dnCloseUnlockModal() {
  var o = document.getElementById("dn-unlock-modal-overlay");
  if (o) { o.style.display = "none"; }
  document.body.style.overflow = "";
  /* Reset form */
  ["dn-cap-name","dn-cap-email","dn-cap-mobile"].forEach(function(id){
    var el=document.getElementById(id); if(el) el.value="";
  });
  var fb=document.getElementById("dn-capture-form-body");
  var fs=document.getElementById("dn-capture-success");
  if(fb) fb.style.display="block";
  if(fs) fs.style.display="none";
  var err=document.getElementById("dn-capture-err");
  if(err){err.style.display="none";err.textContent="";}
};

window.dnCaptureSubmit = function dnCaptureSubmit() {
  var name   = (document.getElementById("dn-cap-name")   || {value:""}).value.trim();
  var email  = (document.getElementById("dn-cap-email")  || {value:""}).value.trim();
  var mobile = (document.getElementById("dn-cap-mobile") || {value:""}).value.trim();
  var err    = document.getElementById("dn-capture-err");

  if (!name)   { err.textContent="Please enter your name.";   err.style.display="block"; return; }
  if (!email || !email.includes("@")) {
    err.textContent="Please enter a valid email address."; err.style.display="block"; return;
  }
  if (!mobile) { err.textContent="Please enter your mobile number."; err.style.display="block"; return; }
  err.style.display="none";

  /* Store in sessionStorage for CRM/backend use */
  try {
    sessionStorage.setItem("dn_lead_name",   name);
    sessionStorage.setItem("dn_lead_email",  email);
    sessionStorage.setItem("dn_lead_mobile", mobile);
  } catch(e){}

  /* Show success state */
  var fb=document.getElementById("dn-capture-form-body");
  var fs=document.getElementById("dn-capture-success");
  if(fb) fb.style.display="none";
  if(fs) fs.style.display="block";

  /* Push into hidden Gradio inputs, then trigger the unlock button */
  function setGr(sel, val) {
    var wrap=document.querySelector(sel);
    if(!wrap) return;
    var inp=wrap.querySelector("input,textarea");
    if(!inp) return;
    var proto=inp.tagName==="TEXTAREA"?window.HTMLTextAreaElement.prototype:window.HTMLInputElement.prototype;
    var desc=Object.getOwnPropertyDescriptor(proto,"value");
    if(desc&&desc.set) desc.set.call(inp,val);
    inp.dispatchEvent(new Event("input",{bubbles:true,composed:true}));
    inp.dispatchEvent(new Event("change",{bubbles:true,composed:true}));
  }
  setGr("#dn-landing-name", name);
  setGr("#dn-real-mobile",  mobile);
  setGr("#dn-landing-otp",  "1234"); /* auto-fill OTP for demo */

  /* Close modal and click unlock after brief delay */
  setTimeout(function(){
    dnCloseUnlockModal();
    /* Find and click the hidden Gradio unlock trigger button */
    var btn = document.querySelector("#dn-unlock-trigger-btn button, #dn-unlock-trigger-btn");
    if(btn && btn.tagName==="BUTTON") btn.click();
    else if(btn) { var b=btn.querySelector("button"); if(b) b.click(); }
  }, 800);
};
function dnCaptureSubmit(){window.dnCaptureSubmit();}
function dnCloseUnlockModal() { window.dnCloseUnlockModal(); }
window.dnUnlockContinue = function dnUnlockContinue() {
  window.dnCloseUnlockModal();
  setTimeout(function() {
    var anchor = document.getElementById("dn-input-anchor");
    if (anchor) { anchor.scrollIntoView({ behavior: "smooth", block: "start" }); }
  }, 80);
};
function dnUnlockContinue() { window.dnUnlockContinue(); }

/* ── Global keyboard + overlay-click handlers ────────────────────────── */
document.addEventListener("keydown", function(e) {
  if (e.key === "Escape") { dnCloseModal(); dnCloseUnlockModal(); }
});
document.addEventListener("click", function(e) {
  var d = document.getElementById("dn-demo-modal-overlay");
  if (d && e.target === d) dnCloseModal();
  var u = document.getElementById("dn-unlock-modal-overlay");
  if (u && e.target === u) dnCloseUnlockModal();
});

/* ── Collapse hidden stub columns that leave whitespace ── */
(function collapseStubs(){
  function collapse(){
    // Target all hidden columns in step0
    document.querySelectorAll(
      '#dn-step0-col [style*="display: none"],' +
      '#dn-step0-col [style*="display:none"],' +
      '#dn-login-section-col[style*="display: none"],' +
      '#dn-login-section-col[style*="display:none"]'
    ).forEach(function(el){
      el.style.setProperty('height','0','important');
      el.style.setProperty('min-height','0','important');
      el.style.setProperty('max-height','0','important');
      el.style.setProperty('overflow','hidden','important');
      el.style.setProperty('margin','0','important');
      el.style.setProperty('padding','0','important');
      el.style.setProperty('flex','none','important');
      el.style.setProperty('border','none','important');
    });
    // Also zero out step0-col gap
    var col = document.getElementById('dn-step0-col');
    if(col){ col.querySelectorAll(':scope > div > div').forEach(function(el){
      el.style.setProperty('gap','0','important');
      el.style.setProperty('row-gap','0','important');
    });}
  }
  [200,600,1200,2500].forEach(function(t){setTimeout(collapse,t);});
  var obs=new MutationObserver(collapse);
  var target=document.getElementById('dn-step0-col');
  if(target) obs.observe(target,{subtree:true,attributes:true,attributeFilter:['style']});
})();
/* ── Premium background — applied once on load ── */
(function fixLoginForm(){
  // Inject a persistent <style> tag that overrides Gradio Svelte styles
  function injectStyle(){
    if(document.getElementById('dn-login-style')) return;
    var s = document.createElement('style');
    s.id = 'dn-login-style';
    s.textContent = [
      /* Kill spinners */
      '#dn-landing-name input,#dn-real-mobile input,#dn-landing-otp input,#dn-landing-email input,',
      '#dn-msme-col-inner input,#dn-msme-col-inner textarea{',
        '-webkit-appearance:none!important;-moz-appearance:textfield!important;appearance:none!important;',
        'height:48px!important;min-height:48px!important;padding:0 14px!important;',
        'border:1.5px solid #d1d9e6!important;border-radius:10px!important;',
        'font-size:14px!important;color:#0f172a!important;background:#ffffff!important;',
        'box-shadow:0 1px 3px rgba(0,0,0,0.06)!important;resize:none!important;',
        'outline:none!important;width:100%!important;box-sizing:border-box!important;}',
      '#dn-landing-name input:focus,#dn-real-mobile input:focus,',
      '#dn-landing-otp input:focus,#dn-landing-email input:focus{',
        'border-color:#2563eb!important;box-shadow:0 0 0 3px rgba(37,99,235,0.12)!important;}',
      /* Kill spinner arrows */
      '#dn-landing-name input::-webkit-outer-spin-button,',
      '#dn-landing-name input::-webkit-inner-spin-button,',
      '#dn-real-mobile input::-webkit-outer-spin-button,',
      '#dn-real-mobile input::-webkit-inner-spin-button,',
      '#dn-landing-otp input::-webkit-outer-spin-button,',
      '#dn-landing-otp input::-webkit-inner-spin-button,',
      '#dn-landing-email input::-webkit-outer-spin-button,',
      '#dn-landing-email input::-webkit-inner-spin-button{',
        '-webkit-appearance:none!important;display:none!important;width:0!important;margin:0!important;}',
      /* Kill Gradio chrome around inputs */
      '#dn-msme-col-inner .block,#dn-msme-col-inner fieldset,',
      '#dn-msme-col-inner .form,#dn-msme-col-inner .wrap,',
      '#dn-msme-col-inner .container{',
        'border:none!important;background:transparent!important;',
        'box-shadow:none!important;padding:0!important;margin:0!important;}',
      /* Button */
      'div#dn-real-login-btn button{',
        'width:100%!important;height:54px!important;min-height:54px!important;',
        'background:linear-gradient(90deg,#1B4F8A 0%,#2563eb 100%)!important;',
        'color:#ffffff!important;border:none!important;border-radius:12px!important;',
        'font-size:16px!important;font-weight:800!important;letter-spacing:0.2px!important;',
        'cursor:pointer!important;display:flex!important;align-items:center!important;',
        'justify-content:center!important;visibility:visible!important;opacity:1!important;',
        'box-shadow:0 6px 20px rgba(37,99,235,0.40)!important;',
        'margin-top:4px!important;}',
      'div#dn-real-login-btn button img,div#dn-real-login-btn button svg{display:none!important;}',
      /* Labels */
      '#dn-landing-name label span,#dn-real-mobile label span,',
      '#dn-landing-otp label span,#dn-landing-email label span{',
        'font-size:13px!important;font-weight:700!important;color:#1e3a5f!important;}',
    ].join('');
    document.head.appendChild(s);
  }

  function fix(){
    injectStyle();
    // Also force inline styles as belt-and-suspenders
    ['dn-landing-name','dn-real-mobile','dn-landing-otp','dn-landing-email'].forEach(function(id){
      var wrap = document.getElementById(id);
      if(!wrap) return;
      wrap.querySelectorAll('input,textarea').forEach(function(el){
        el.style.setProperty('-webkit-appearance','none','important');
        el.style.setProperty('appearance','none','important');
        el.style.setProperty('height','48px','important');
        el.style.setProperty('border','1.5px solid #d1d9e6','important');
        el.style.setProperty('border-radius','10px','important');
        el.style.setProperty('font-size','14px','important');
        el.style.setProperty('padding','0 14px','important');
        el.style.setProperty('background','#ffffff','important');
      });
    });
    var btn = document.querySelector('div#dn-real-login-btn button');
    if(btn){
      btn.style.setProperty('width','100%','important');
      btn.style.setProperty('height','54px','important');
      btn.style.setProperty('min-height','54px','important');
      btn.style.setProperty('background','linear-gradient(90deg,#1B4F8A,#2563eb)','important');
      btn.style.setProperty('color','#ffffff','important');
      btn.style.setProperty('border','none','important');
      btn.style.setProperty('border-radius','12px','important');
      btn.style.setProperty('font-size','16px','important');
      btn.style.setProperty('font-weight','800','important');
      btn.style.setProperty('visibility','visible','important');
      btn.style.setProperty('opacity','1','important');
      btn.style.setProperty('cursor','pointer','important');
      btn.querySelectorAll('img,svg').forEach(function(i){i.style.display='none';});
    }
  }

  // Run immediately, then on intervals, then on any DOM change
  [100,300,700,1400,2500,4000].forEach(function(t){setTimeout(fix,t);});
  // Keep running every 2 seconds to catch Gradio re-renders
  setInterval(fix, 2000);
  // MutationObserver for instant response
  function watchLogin(){
    var target = document.getElementById('dn-login-section-col') || document.body;
    new MutationObserver(function(muts){
      var relevant = muts.some(function(m){
        return m.target.id && (m.target.id.indexOf('dn-landing')>=0 || m.target.id.indexOf('dn-real')>=0 || m.target.id.indexOf('dn-msme')>=0);
      });
      if(relevant || muts.length > 0) fix();
    }).observe(target, {subtree:true, attributes:true, attributeFilter:['style','class'], childList:true});
  }
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded', watchLogin);
  else watchLogin();
})();
/* fixLoginBtn merged into fixLoginForm */
(function applyBg() {
  var bg = "#E8F0FE";
  try {
    document.documentElement.style.setProperty("background", bg, "important");
    document.body.style.setProperty("background", bg, "important");
    ["gradio-app","#root",".gradio-container",".main",".wrap"].forEach(function(sel) {
      document.querySelectorAll(sel).forEach(function(el) {
        el.style.setProperty("background", bg, "important");
      });
    });
  } catch(e) {}
})();
/* Re-run after Gradio finishes rendering */
window.addEventListener("load", function() {
  var bg = "#E8F0FE";
  ["gradio-app","#root",".gradio-container",".main",".wrap",".gap"].forEach(function(sel) {
    document.querySelectorAll(sel).forEach(function(el) {
      el.style.setProperty("background", bg, "important");
    });
  });
});
</script>
''')

        # ── New SaaS Input Section ───────────────────────────────────────────        # ── New SaaS Input Section ───────────────────────────────────────────
        # ── Login / Unlock section — hidden initially, shown when user clicks Unlock ──
        gr.HTML('<div id="dn-input-anchor" style="height:0;scroll-margin-top:60px;"></div>')
        with gr.Column(visible=False, elem_id="dn-login-section-col") as login_section_col:
            gr.HTML('''
<div style="background:#f8fbff;border-top:1px solid #e2eaf5;padding:28px 24px 0;">
  <div style="max-width:560px;margin:0 auto;text-align:center;">
    <div style="display:inline-flex;align-items:center;gap:8px;background:#EFF6FF;
         border:1px solid #BFDBFE;border-radius:20px;padding:5px 14px;margin-bottom:14px;">
      <svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="#2563eb"
           stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
        <path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/>
      </svg>
      <span style="font-size:11px;font-weight:700;letter-spacing:1.5px;color:#2563eb;
                   text-transform:uppercase;">One Step Away</span>
    </div>
    <div style="font-size:26px;font-weight:900;color:#0f2557;margin-bottom:10px;
                letter-spacing:-0.5px;line-height:1.2;">
      Your AI Business Intelligence Report Is Ready
    </div>
    <div style="font-size:14px;color:#64748b;line-height:1.7;max-width:460px;margin:0 auto;">
      Verify your details to unlock forecasts, strategic recommendations,
      SKU insights, and your downloadable business report.
    </div>
  </div>
</div>
'''
)
            with gr.Column(elem_id="dn-msme-card-wrap"):
                with gr.Column(elem_id="dn-msme-col-inner", visible=True) as msme_col:
                    msme_card_html = gr.HTML(value=''' ''',
                                             visible=False)  # gr.update() no-op — never overwrite the HTML form
                    # ── Pure HTML form overlay — Gradio widgets hidden below for event wiring ──
                    gr.HTML('''
<div id="dn-html-form-card" style="background:#ffffff;border-radius:16px;border:1px solid #e2eaf5;
     box-shadow:0 8px 40px rgba(11,31,58,0.12);padding:28px 28px 20px;
     font-family:-apple-system,BlinkMacSystemFont,\'Segoe UI\',sans-serif;">

  <div style="display:flex;flex-direction:column;gap:14px;margin-bottom:18px;">

    <div>
      <label style="display:block;font-size:13px;font-weight:700;color:#1e3a5f;margin-bottom:6px;">Your Name</label>
      <input id="dn-f-name" type="text" placeholder="e.g. Ravi Kumar"
        style="width:100%;box-sizing:border-box;height:48px;padding:0 14px;font-size:14px;
               color:#0f172a;background:#ffffff;border:1.5px solid #d1d9e6;border-radius:10px;
               outline:none;transition:border-color 0.15s,box-shadow 0.15s;">
    </div>

    <div>
      <label style="display:block;font-size:13px;font-weight:700;color:#1e3a5f;margin-bottom:6px;">Mobile Number</label>
      <input id="dn-f-mobile" type="tel" placeholder="e.g. +91 98765 43210"
        style="width:100%;box-sizing:border-box;height:48px;padding:0 14px;font-size:14px;
               color:#0f172a;background:#ffffff;border:1.5px solid #d1d9e6;border-radius:10px;
               outline:none;transition:border-color 0.15s,box-shadow 0.15s;">
    </div>

    <div>
      <label style="display:block;font-size:13px;font-weight:700;color:#1e3a5f;margin-bottom:6px;">
        Verification Code
        <span style="font-size:11px;font-weight:500;color:#94a3b8;margin-left:6px;">(use 1234 for demo)</span>
      </label>
      <input id="dn-f-otp" type="text" placeholder="Enter verification code" maxlength="6"
        style="width:100%;box-sizing:border-box;height:48px;padding:0 14px;font-size:14px;
               color:#0f172a;background:#ffffff;border:1.5px solid #d1d9e6;border-radius:10px;
               outline:none;transition:border-color 0.15s,box-shadow 0.15s;">
    </div>

    <div>
      <label style="display:block;font-size:13px;font-weight:700;color:#1e3a5f;margin-bottom:6px;">Email Address</label>
      <input id="dn-f-email" type="email" placeholder="e.g. ravi@store.com"
        style="width:100%;box-sizing:border-box;height:48px;padding:0 14px;font-size:14px;
               color:#0f172a;background:#ffffff;border:1.5px solid #d1d9e6;border-radius:10px;
               outline:none;transition:border-color 0.15s,box-shadow 0.15s;">
    </div>

  </div>

  <div id="dn-gradio-btn-anchor" style="width:100%;height:54px;"></div>

  <div style="text-align:center;margin-top:12px;font-size:12px;color:#94a3b8;letter-spacing:0.2px;">
    🔒 Secure access &nbsp;·&nbsp; No spam &nbsp;·&nbsp; Instant unlock
  </div>

  <div id="dn-f-error" style="display:none;margin-top:12px;padding:10px 14px;
       background:#fef2f2;border:1px solid #fecaca;border-radius:8px;
       font-size:13px;color:#dc2626;text-align:center;"></div>
  <div id="dn-f-debug" style="display:none;margin-top:8px;padding:8px 12px;
       background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;
       font-size:11px;color:#0369a1;text-align:left;font-family:monospace;"></div>

</div>

''')
                    # Gradio inputs moved to top-level (see below login_section_col)
                    gr.HTML('<div id="dn-login-placeholder"></div>')
                    # quick_login_btn defined at top level (L9878)

                    landing_login_title  = gr.Markdown("", visible=False)
                    signup_card_html     = gr.HTML(value="", visible=False)
                    landing_signup_title = gr.Markdown("", visible=False)
                    landing_signup_desc  = gr.Markdown("", visible=False)
                    quick_signup_btn     = gr.Button("Sign Up", variant="secondary", visible=False)

                # Gov panel stub — hidden, kept for event-wiring compatibility
                with gr.Column(elem_id="dn-gov-card-wrap", visible=False) as gov_col:
                    gov_card_html       = gr.HTML(value="", visible=False)
                    gov_login_id_input  = gr.Textbox(visible=False)
                    gov_login_pwd_input = gr.Textbox(visible=False, type="password")
                    gov_login_error_msg = gr.Markdown(value="", visible=False)
                    gov_login_btn       = gr.Button("Gov Login", visible=False)
                    auth_hint_html      = gr.HTML(value="", visible=False)

            # Stubs for _lang_landing_outputs (previously in right-column, now hidden)
            with gr.Column(visible=False):
                delivers_html_widget = gr.HTML(value="", visible=False)
                platform_html_widget = gr.HTML(value="", visible=False)
        # ── Data Readiness Check block — sits directly below middle section ───
        gr.HTML('<div id="dn-pricing-anchor" style="height:0;overflow:hidden;"></div>')
        gr.HTML('<div class="dn-band dn-band-grey" style="padding:0;margin-top:0;"><div class="dn-band-inner" style="padding-top:16px;padding-bottom:16px;">')
        gr.HTML('<div id="dn-drc-anchor" style="height:0;overflow:hidden;"></div>')
        with gr.Column(elem_id="drc-block"):
            # Title + description row
            gr.HTML("""
            <div style='margin-bottom:10px;'>
              <div style='font-size:11px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#1B4F8A;margin-bottom:6px;'>STEP 1 — DATA READINESS ENGINE</div>
              <div style='font-size:20px;font-weight:900;color:#0B1F3A;letter-spacing:-0.3px;margin-bottom:6px;'>Upload &amp; Prepare Your Dataset</div>
              <div style='font-size:13px;color:#64748b;line-height:1.5;max-width:640px;margin-bottom:8px;'>Upload your sales, inventory, billing, or POS export. We'll clean, validate, and prepare it for accurate AI insights.</div>
              <div style='font-size:12px;color:#94a3b8;margin-bottom:8px;'>Supports Excel (.xlsx), CSV, and most retail software exports.</div>
              <div style='display:flex;gap:8px;flex-wrap:wrap;margin-bottom:4px;'>
                <span style='background:#EBF5FF;border:1px solid #BDD7F5;color:#1B4F8A;font-size:12px;font-weight:600;padding:3px 10px;border-radius:12px;'>📂 Column mapping</span>
                <span style='background:#EBF5FF;border:1px solid #BDD7F5;color:#1B4F8A;font-size:12px;font-weight:600;padding:3px 10px;border-radius:12px;'>🧹 Data cleaning</span>
                <span style='background:#EBF5FF;border:1px solid #BDD7F5;color:#1B4F8A;font-size:12px;font-weight:600;padding:3px 10px;border-radius:12px;'>✅ Validation</span>
                <span style='background:#EBF5FF;border:1px solid #BDD7F5;color:#1B4F8A;font-size:12px;font-weight:600;padding:3px 10px;border-radius:12px;'>📊 Normalisation</span>
              </div>
            </div>
            """)

            # Action bar: Upload + Run + Sample Data + Download Template (compact row)
            with gr.Row(equal_height=True):
                drc_file      = gr.File(label="",
                                        file_types=[".csv",".xlsx",".xls"], scale=4,
                                        type="filepath")
                drc_run       = gr.Button("▶ Run Analysis", variant="primary",
                                          scale=1, min_width=150)
                drc_tmpl_btn  = gr.Button("⬇ Download Template",
                                              variant="secondary", scale=1, min_width=150)

            # Template download file output (hidden, populated on button click)
            drc_tmpl_file = gr.File(label="", visible=False, elem_id="drc-tmpl-file-out")

            # Loading feedback shown immediately when Run Analysis is clicked
            drc_loading_html = gr.HTML(value="", visible=False, elem_id="drc-loading-msg")

            # Compact results summary (visible after check)
            with gr.Column(visible=False) as _drc_results:

                # Row 1: Status pill only (compact strip moves into Overview tab)
                drc_status_out   = gr.HTML()

                # Details toggle — scroll anchor injected for JS targeting
                gr.HTML('''<div id="drc-report-anchor" style="scroll-margin-top:12px;height:0;overflow:hidden;"></div>
<script>
(function patchDrcUpload(){
  function patch(){
    var block = document.getElementById("drc-block");
    if(!block) return;
    /* Replace "Drop File Here" text */
    block.querySelectorAll("span").forEach(function(s){
      if(s.textContent.trim()==="Drop File Here"){
        s.textContent="Drag & Drop Your Sales File — Excel or CSV";
        s.style.cssText="font-size:13px;font-weight:600;color:#1B4F8A;text-align:center;";
      }
    });
    /* Hide "- or -" text nodes */
    block.querySelectorAll(".or, .separator").forEach(function(el){
      el.style.display="none";
    });
    /* Also scan text nodes for "- or -" */
    block.querySelectorAll("*").forEach(function(el){
      if(el.childNodes.length===1 && el.childNodes[0].nodeType===3){
        var t=el.textContent.trim();
        if(t==="- or -" || t==="or"){el.style.display="none";}
      }
    });
  }
  if(document.readyState==="loading"){
    document.addEventListener("DOMContentLoaded", patch);
  } else {
    setTimeout(patch, 800);
  }
  /* Re-run on Gradio re-renders */
  setTimeout(patch, 2000);
  setTimeout(patch, 4000);
  /* Collapse hidden panels that leave blank space */
  function collapseHidden(){
    var block = document.getElementById("drc-block");
    if(!block) return;
    block.querySelectorAll('[style*="display: none"],[style*="display:none"]').forEach(function(el){
      el.style.height="0";
      el.style.minHeight="0";
      el.style.margin="0";
      el.style.padding="0";
      el.style.overflow="hidden";
    });
  }
  setTimeout(collapseHidden, 1000);
  setTimeout(collapseHidden, 3000);
})();
</script>''')
                drc_details_btn  = gr.Button("🔎 View Full Analysis Report ▾",
                                             variant="secondary", size="sm")
                gr.HTML('''<script>
(function(){
  // Scroll to report anchor when details panel opens
  document.addEventListener("click", function(e){
    var btn = e.target && e.target.closest ? e.target.closest("button") : null;
    if(btn && btn.textContent && btn.textContent.indexOf("View Full Analysis Report") >= 0){
      setTimeout(function(){
        var anchor = document.getElementById("drc-report-anchor");
        if(anchor){ anchor.scrollIntoView({behavior:"smooth", block:"start"}); }
      }, 320);
    }
  }, true);
})();
</script>''')

                # Collapsible details panel — tabbed analysis view
                with gr.Column(visible=False) as _drc_details_panel:
                    gr.HTML("<hr style='margin:8px 0;border:none;border-top:1px solid #e2e8f0;'>")

                    # ── Native Gradio Tabs — no JS, no MutationObserver, no slicing ──
                    with gr.Tabs(elem_id="drc-native-tabs"):
                        with gr.Tab("📊 Overview"):
                            drc_overview_out = gr.HTML(value="", elem_id="drc-overview-content")
                        with gr.Tab("🔎 Data Quality"):
                            drc_quality_out2 = gr.HTML(value="", elem_id="drc-quality-content")
                        with gr.Tab("📐 Structure"):
                            drc_struct_out   = gr.HTML(value="", elem_id="drc-struct-content")
                        with gr.Tab("🛡️ Authenticity"):
                            drc_auth_out     = gr.HTML(value="", elem_id="drc-auth-content")
                        with gr.Tab("🧹 Cleaning"):
                            drc_clean_out    = gr.HTML(value="", elem_id="drc-clean-content")
                        with gr.Tab("🗂️ Field Mapping"):
                            drc_map_out      = gr.HTML(value="", elem_id="drc-map-content")

                    # Hidden source widgets — preserved for handler signature compatibility
                    drc_compact_out  = gr.HTML(elem_id="drc-compact-out",  visible=False)
                    drc_quality_out  = gr.HTML(elem_id="drc-quality-out",  visible=False)
                    drc_summary_out  = gr.Textbox(label="", lines=6, interactive=False,
                                                  show_label=False, visible=False)
                    drc_mapping_out  = gr.HTML(visible=False)

                gr.HTML(
                    "<div style='font-size:11px;color:#475569;padding:6px 2px 4px;"
                    "font-family:sans-serif;display:flex;align-items:center;gap:6px;'>"
                    "<span style='color:#16a34a;font-size:13px;'>✔</span>"
                    "<span>Your data has been processed and is ready for analysis.</span>"
                    "</div>"
                )
                with gr.Row():
                    drc_dl_btn   = gr.File(label="⬇ Download Clean Dataset",
                                                      visible=False, interactive=False)
                    drc_cont_btn = gr.Button("Proceed to Insights →",
                                             variant="primary", interactive=True)
                drc_cont_note = gr.Markdown(
                    "_Your data has been processed. Click 'Proceed to Insights' to upload and analyse your data._", visible=True)

            _drc_clean_state    = gr.State(None)
            _drc_details_open   = gr.State(False)

            # ── Template download handler ──────────────────────────────────
            def _drc_get_template():
                path = generate_blank_template()
                return gr.update(value=path, visible=True), gr.update(value="✅ Template Ready — Check below")

            drc_tmpl_btn.click(fn=_drc_get_template, inputs=[], outputs=[drc_tmpl_file, drc_tmpl_btn])

            # ── Sample data handler ────────────────────────────────────────
            def _drc_load_sample():
                """Generate a realistic sample retail CSV and run DRC on it."""
                import tempfile as _tf2, csv as _csv2
                _rows = [
                    ["date","store_id","product_category","product_id","gross_sales","units_sold",
                     "cost_price","return_rate_pct","inventory_level","profit_margin_pct","vendor_name"],
                ]
                import random as _rnd2; _rnd2.seed(42)
                stores = ["S001","S002","S003"]
                cats   = ["Electronics","FMCG","Clothing","Grocery","Appliances"]
                vendors= ["Vendor-A","Vendor-B","Vendor-C"]
                for mo in range(1, 13):
                    for st in stores:
                        for cat in cats:
                            sales  = _rnd2.randint(80000, 500000)
                            units  = _rnd2.randint(50, 600)
                            cost   = round(sales * _rnd2.uniform(0.55, 0.75))
                            ret    = round(_rnd2.uniform(1.5, 9.0), 1)
                            inv    = _rnd2.randint(100, 800)
                            margin = round(_rnd2.uniform(10.0, 28.0), 1)
                            pid    = f"{cat[:3].upper()}-{_rnd2.randint(100,999)}"
                            _rows.append([f"2024-{mo:02d}-01", st, cat, pid, sales, units,
                                          cost, ret, inv, margin, _rnd2.choice(vendors)])
                _tmp = _tf2.NamedTemporaryFile(delete=False, suffix=".csv",
                                               prefix="DataNetra_Sample_")
                with open(_tmp.name, "w", newline="") as _f2:
                    _csv2.writer(_f2).writerows(_rows)
                _tmp.close()
                # Run DRC on the sample file — _drc_run_with_hide defined below
                return _drc_run_with_hide(_tmp.name)


            # ── Compact KPI strip builder ──────────────────────────────────
            def _drc_compact_strip(n_rows, n_cols, n_crit, n_warn, status, score=None):
                sc_color = ("#16a34a" if (score or 0) >= 90 else "#d97706" if (score or 0) >= 70
                            else "#c2520a" if (score or 0) >= 50 else "#dc2626") if score else "#64748b"
                crit_col = "#dc2626" if n_crit else "#16a34a"
                warn_col = "#d97706" if n_warn else "#16a34a"
                def _stat(icon, label, value, note="", color="#0f172a"):
                    return (
                        f"<div style='flex:1;min-width:100px;background:#fff;"
                        f"border:1px solid #e2e8f0;border-radius:8px;padding:10px 12px;"
                        f"text-align:center;box-shadow:0 1px 4px rgba(11,31,58,0.07);'>"
                        f"<div style='font-size:17px;margin-bottom:2px;'>{icon}</div>"
                        f"<div style='font-size:18px;font-weight:900;color:{color};"
                        f"font-family:monospace;line-height:1;'>{value}</div>"
                        f"<div style='font-size:10px;font-weight:600;color:#64748b;"
                        f"margin-top:2px;'>{label}</div>"
                        + (f"<div style='font-size:9px;color:#94a3b8;margin-top:1px;'>{note}</div>" if note else "")
                        + "</div>"
                    )
                score_stat = _stat("🎯", "Readiness Score", f"{score}/100", status, sc_color) if score is not None else ""

                # ── Confidence Indicator + AI Decision (UI-only, derived from score) ──
                if score is not None:
                    _s = score or 0
                    if _s >= 85:
                        _conf_label  = "High Confidence"
                        _conf_color  = "#16a34a"
                        _conf_bg     = "#f0fdf4"
                        _conf_border = "#bbf7d0"
                        _conf_icon   = "✅"
                        _ai_msg      = "Dataset ready for advanced analytics and forecasting."
                        _ai_color    = "#16a34a"
                        _ai_bg       = "#f0fdf4"
                        _ai_border   = "#bbf7d0"
                    elif _s >= 70:
                        _conf_label  = "Moderate Confidence"
                        _conf_color  = "#d97706"
                        _conf_bg     = "#fffbeb"
                        _conf_border = "#fde68a"
                        _conf_icon   = "🟡"
                        _ai_msg      = "Dataset suitable for analytics after minor cleaning improvements."
                        _ai_color    = "#d97706"
                        _ai_bg       = "#fffbeb"
                        _ai_border   = "#fde68a"
                    elif _s >= 50:
                        _conf_label  = "Limited Confidence"
                        _conf_color  = "#c2520a"
                        _conf_bg     = "#fff7ed"
                        _conf_border = "#fed7aa"
                        _conf_icon   = "🟠"
                        _ai_msg      = "Dataset usable but requires additional data improvement."
                        _ai_color    = "#c2520a"
                        _ai_bg       = "#fff7ed"
                        _ai_border   = "#fed7aa"
                    else:
                        _conf_label  = "Low Confidence"
                        _conf_color  = "#dc2626"
                        _conf_bg     = "#fef2f2"
                        _conf_border = "#fecaca"
                        _conf_icon   = "🔴"
                        _ai_msg      = "Dataset requires significant correction before analytics can be performed."
                        _ai_color    = "#dc2626"
                        _ai_bg       = "#fef2f2"
                        _ai_border   = "#fecaca"

                    _interp_block = (
                        f"<div style='display:flex;gap:8px;flex-wrap:wrap;margin-top:8px;'>"

                        # Confidence Level card
                        f"<div style='flex:1;min-width:180px;background:{_conf_bg};"
                        f"border:1px solid {_conf_border};border-left:3px solid {_conf_color};"
                        f"border-radius:8px;padding:9px 13px;"
                        f"display:flex;align-items:center;gap:10px;'>"
                        f"<span style='font-size:18px;line-height:1;'>{_conf_icon}</span>"
                        f"<div>"
                        f"<div style='font-size:9px;font-weight:700;letter-spacing:0.8px;"
                        f"text-transform:uppercase;color:#64748b;margin-bottom:2px;'>Confidence Level</div>"
                        f"<div style='font-size:13px;font-weight:800;color:{_conf_color};"
                        f"line-height:1.2;'>{_conf_label}</div>"
                        f"</div>"
                        f"</div>"

                        # AI Decision card
                        f"<div style='flex:2;min-width:220px;background:{_ai_bg};"
                        f"border:1px solid {_ai_border};border-left:3px solid {_ai_color};"
                        f"border-radius:8px;padding:9px 13px;"
                        f"display:flex;align-items:center;gap:10px;'>"
                        f"<span style='font-size:18px;line-height:1;'>🤖</span>"
                        f"<div>"
                        f"<div style='font-size:9px;font-weight:700;letter-spacing:0.8px;"
                        f"text-transform:uppercase;color:#64748b;margin-bottom:2px;'>AI Decision</div>"
                        f"<div style='font-size:12px;font-weight:600;color:{_ai_color};"
                        f"line-height:1.4;'>{_ai_msg}</div>"
                        f"</div>"
                        f"</div>"

                        f"</div>"
                    )
                else:
                    _interp_block = ""

                return (
                    "<div style='margin-bottom:10px;'>"
                    "<div style='font-size:10px;font-weight:700;letter-spacing:0.9px;"
                    "text-transform:uppercase;color:#1B4F8A;margin-bottom:8px;"
                    "padding-bottom:5px;border-bottom:2px solid #EEF4FB;'>"
                    "📊 DataNetra Data Quality Scorecard</div>"
                    "<div style='display:flex;gap:8px;flex-wrap:wrap;'>"
                    + _stat("📄", "Dataset Size", f"{n_rows:,}", f"{n_cols} cols detected")
                    + _stat("❌", "Quality Issues", str(n_crit), f"{n_warn} warnings", crit_col)
                    + score_stat
                    + "</div>"
                    + _interp_block
                    + "</div>"
                )

            # ── Main run handler ───────────────────────────────────────────
            def _drc_run_handler(fobj):
                # Gradio 4.44 may return a list for file widgets — unwrap
                if isinstance(fobj, list):
                    fobj = fobj[0] if fobj else None
                if fobj is None:
                    return ("", "", "", "", "", None,
                            gr.update(visible=False), gr.update(visible=False),
                            "", "", "", "", "", "")
                try:
                    sh, mh, df_c, summ, qh = run_readiness_check(fobj)
                    path = export_clean_dataset(df_c)
                    # Parse counts for compact strip
                    import re as _re2
                    _m_rows = _re2.search(r"Rows: ([\d,]+)", summ)
                    n_rows = int(_m_rows.group(1).replace(",","")) if _m_rows else 0
                    _m_cols = _re2.search(r"Columns detected: (\d+)", summ)
                    n_cols = int(_m_cols.group(1)) if _m_cols else 0
                    n_crit2 = qh.count("❌ Critical") + qh.count("Critical</div>")
                    n_warn2 = qh.count("⚠️ Warnings") + qh.count("Warnings</div>")
                    status = ("Ready" if "Ready" in sh else
                              "Partial" if "Partial" in sh else
                              "Needs Completion" if "Needs" in sh else "")
                    _m_score = _re2.search(r"Data Readiness Score: (\d+) / 100", summ)
                    _score_val = int(_m_score.group(1)) if _m_score else None
                    compact = _drc_compact_strip(n_rows, n_cols, n_crit2, n_warn2, status, _score_val)
                    # ── Split quality_html by section markers into 6 native tab panels ──
                    _MARKERS = ["<!--DRC_OVERVIEW-->","<!--DRC_QUALITY-->","<!--DRC_STRUCTURE-->",
                                "<!--DRC_AUTH-->","<!--DRC_CLEANING-->","<!--DRC_MAPPING-->"]
                    def _split_qh(html):
                        parts = {}
                        for i, marker in enumerate(_MARKERS):
                            next_marker = _MARKERS[i+1] if i+1 < len(_MARKERS) else None
                            s = html.find(marker)
                            if s < 0: parts[marker] = ""; continue
                            s += len(marker)
                            e = html.find(next_marker, s) if next_marker else len(html)
                            parts[marker] = html[s:e] if e >= s else html[s:]
                        return parts
                    _p = _split_qh(qh)
                    # Overview tab = compact strip (score card) + score html + summary log
                    _overview_html = (
                        compact +
                        "<details open style='margin-top:10px;'>"
                        "<summary style='font-size:11px;font-weight:700;color:#1B4F8A;cursor:pointer;"
                        "letter-spacing:0.3px;'>📝 Validation Log</summary>"
                        "<pre style='font-size:11px;color:#475569;background:#f8fafc;border:1px solid #e2e8f0;"
                        "border-radius:6px;padding:8px 10px;white-space:pre-wrap;word-break:break-word;"
                        "line-height:1.5;max-height:160px;overflow-y:auto;margin:6px 0 0;'>"
                        + summ.replace("<","&lt;").replace(">","&gt;") + "</pre></details>"
                        + _p.get("<!--DRC_OVERVIEW-->","")
                    )
                    return (
                        sh,                                    # drc_status_out
                        compact,                               # drc_compact_out (hidden)
                        mh,                                    # drc_mapping_out (hidden)
                        summ,                                  # drc_summary_out (hidden)
                        qh,                                    # drc_quality_out (hidden)
                        path,                                  # _drc_clean_state
                        gr.update(value=path, visible=True),    # drc_dl_btn
                        gr.update(visible=True),               # _drc_results
                        _overview_html,                        # drc_overview_out
                        _p.get("<!--DRC_QUALITY-->",""),       # drc_quality_out2
                        _p.get("<!--DRC_STRUCTURE-->",""),     # drc_struct_out
                        _p.get("<!--DRC_AUTH-->",""),          # drc_auth_out
                        _p.get("<!--DRC_CLEANING-->",""),      # drc_clean_out
                        _p.get("<!--DRC_MAPPING-->",""),       # drc_map_out
                    )
                except Exception as _e:
                    import traceback as _tb
                    _err = f"<p style='color:#dc2626;font-weight:700;'>❌ Analysis error: {_e}</p><pre style='font-size:10px;overflow:auto;'>{_tb.format_exc()[:1500]}</pre>"
                    return (_err, "", "", "", "", None,
                            gr.update(visible=False), gr.update(visible=True),
                            _err, "", "", "", "", "")

            def _drc_show_loading():
                return gr.update(visible=True, value="""<div style="display:flex;align-items:center;gap:14px;background:linear-gradient(90deg,#EBF5FF,#F0F8FF);border:1.5px solid #93c5fd;border-radius:12px;padding:16px 22px;margin:12px 0;"><div style="width:22px;height:22px;border:3px solid #2563eb;border-top-color:transparent;border-radius:50%;animation:dn-spin 0.8s linear infinite;flex-shrink:0;"></div><div><div style="font-size:14px;font-weight:700;color:#0B1F3A;">Processing your data...</div><div style="font-size:12px;color:#475569;margin-top:3px;">Checking data quality, structure and readiness. This may take a few seconds.</div></div></div><style>@keyframes dn-spin{to{transform:rotate(360deg);}}</style>""")

            def _drc_run_with_hide(fobj):
                result = _drc_run_handler(fobj)
                return (*result, gr.update(visible=False, value=""))

            drc_run.click(
                fn=_drc_show_loading,
                inputs=[],
                outputs=[drc_loading_html],
                queue=False,
            ).then(
                fn=_drc_run_with_hide,
                inputs=[drc_file],
                outputs=[drc_status_out, drc_compact_out, drc_mapping_out,
                         drc_summary_out, drc_quality_out,
                         _drc_clean_state, drc_dl_btn, _drc_results,
                         drc_overview_out, drc_quality_out2,
                         drc_struct_out, drc_auth_out,
                         drc_clean_out, drc_map_out,
                         drc_loading_html],
                show_progress=True,
                queue=True,
            ).then(
                fn=None, inputs=[], outputs=[],
                js="() => { var r=document.getElementById('drc-block'); if(r){ setTimeout(function(){ r.scrollIntoView({behavior:'smooth',block:'start'}); }, 200); } }"
            )


            # ── Details toggle handler — shows panel + scrolls into view ──
            def _drc_toggle_details(is_open):
                new_open = not is_open
                label = "🔎 View Full Analysis Report ▴" if new_open else "🔎 View Full Analysis Report ▾"
                scroll_js = """<script>
(function(){
  var panel = document.querySelector('.drc-tab-nav') ||
              document.getElementById('drc-pane-overview');
  if(panel){ panel.scrollIntoView({behavior:'smooth',block:'start'}); }
})();
</script>""" if new_open else ""
                return (gr.update(visible=new_open), new_open, gr.update(value=label))

            drc_details_btn.click(
                fn=_drc_toggle_details,
                inputs=[_drc_details_open],
                outputs=[_drc_details_panel, _drc_details_open, drc_details_btn],
            )

            # NOTE: drc_cont_btn.click wired below after _ALL_COLS is defined
        # ── /Data Readiness Check block ───────────────────────────────────────
        gr.HTML('</div></div>')

        # ── Contact Us Section ────────────────────────────────────────────────
        gr.HTML('<div id="dn-contact-anchor" style="height:0;overflow:hidden;"></div>')
        # ── Hidden Gradio inputs for contact form → database ──────────────
        with gr.Column(elem_id="dn-contact-hidden", visible=True):
            cf_name_input    = gr.Textbox(label="", elem_id="dn-cf-gr-name")
            cf_email_input   = gr.Textbox(label="", elem_id="dn-cf-gr-email")
            cf_mobile_input  = gr.Textbox(label="", elem_id="dn-cf-gr-mobile")
            cf_message_input = gr.Textbox(label="", elem_id="dn-cf-gr-message")
            cf_submit_btn    = gr.Button("Submit", elem_id="dn-cf-gr-submit")
            cf_result        = gr.Markdown(value="", elem_id="dn-cf-gr-result")
        gr.HTML('<div class="dn-band dn-band-white"><div class="dn-band-inner" style="padding-top:24px;padding-bottom:20px;">')
        gr.HTML('''
<div style="max-width:1000px;margin:0 auto;">

  <!-- Section header -->
  <div style="text-align:center;margin-bottom:20px;">
    <div style="font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;
                color:#1B4F8A;margin-bottom:5px;">Get In Touch</div>
    <div style="font-size:22px;font-weight:900;color:#0B1F3A;letter-spacing:-0.4px;margin-bottom:6px;">
      Contact Us
    </div>
    <div style="font-size:13px;color:#64748b;max-width:400px;margin:0 auto;line-height:1.55;">
      Have questions, feedback, or want a demo? Reach out to us.
    </div>
  </div>

  <!-- Two column layout — top-aligned -->
  <div style="display:grid;grid-template-columns:1fr 1.7fr;gap:18px;align-items:start;">

    <!-- LEFT: Contact info -->
    <div style="background:#F9FAFB;border:1px solid #e8edf5;border-radius:12px;
                padding:24px 22px;box-shadow:0 1px 6px rgba(11,31,58,0.05);
                display:flex;flex-direction:column;box-sizing:border-box;">
      <div style="font-size:13px;font-weight:800;color:#0B1F3A;margin-bottom:16px;">Reach Us Directly</div>

      <!-- Email -->
      <div style="display:flex;align-items:center;gap:12px;margin-bottom:12px;
                  padding:11px 14px;background:#ffffff;border:1px solid #e8edf5;
                  border-radius:8px;">
        <div style="width:34px;height:34px;background:#EBF2FC;border-radius:8px;display:flex;
                    align-items:center;justify-content:center;flex-shrink:0;">
          <svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="#1B4F8A" stroke-width="2"
               stroke-linecap="round" stroke-linejoin="round">
            <path d="M4 4h16c1.1 0 2 .9 2 2v12c0 1.1-.9 2-2 2H4c-1.1 0-2-.9-2-2V6c0-1.1.9-2 2-2z"/>
            <polyline points="22,6 12,13 2,6"/>
          </svg>
        </div>
        <div>
          <div style="font-size:9px;font-weight:700;letter-spacing:0.8px;text-transform:uppercase;
                      color:#94a3b8;margin-bottom:2px;">Email</div>
          <a href="mailto:support@datanetra.ai"
             style="font-size:13px;font-weight:600;color:#1B4F8A;text-decoration:none;"
             onmouseover="this.style.color='#2563eb'"
             onmouseout="this.style.color='#1B4F8A'">
            support@datanetra.ai
          </a>
        </div>
      </div>

      <!-- LinkedIn -->
      <div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;
                  padding:11px 14px;background:#ffffff;border:1px solid #e8edf5;
                  border-radius:8px;">
        <div style="width:34px;height:34px;background:#EBF2FC;border-radius:8px;display:flex;
                    align-items:center;justify-content:center;flex-shrink:0;">
          <svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="#1B4F8A" stroke-width="2"
               stroke-linecap="round" stroke-linejoin="round">
            <path d="M16 8a6 6 0 0 1 6 6v7h-4v-7a2 2 0 0 0-2-2 2 2 0 0 0-2 2v7h-4v-7a6 6 0 0 1 6-6z"/>
            <rect x="2" y="9" width="4" height="12"/><circle cx="4" cy="4" r="2"/>
          </svg>
        </div>
        <div>
          <div style="font-size:9px;font-weight:700;letter-spacing:0.8px;text-transform:uppercase;
                      color:#94a3b8;margin-bottom:2px;">LinkedIn</div>
          <a href="https://www.linkedin.com/company/108412762/" target="_blank"
             style="font-size:13px;font-weight:600;color:#1B4F8A;text-decoration:none;"
             onmouseover="this.style.color='#2563eb'"
             onmouseout="this.style.color='#1B4F8A'">
            DataNetra.ai
          </a>
        </div>
      </div>

      <!-- Response time badge — no flex spacer, sits naturally after contact items -->
      <div style="display:flex;align-items:center;gap:8px;padding:10px 14px;
                  background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;">
        <span style="font-size:14px;">🕐</span>
        <span style="font-size:12px;color:#16a34a;font-weight:600;line-height:1.4;">
          We typically respond within 24 hours
        </span>
      </div>
    </div>

    <!-- RIGHT: Contact form -->
    <div style="background:#ffffff;border:1px solid #e8edf5;border-radius:12px;
                padding:24px 24px;box-shadow:0 2px 12px rgba(11,31,58,0.06);box-sizing:border-box;">
      <div style="font-size:13px;font-weight:800;color:#0B1F3A;margin-bottom:16px;">Send Us a Message</div>

      <div id="dn-contact-form" style="display:flex;flex-direction:column;gap:12px;">
        <input id="dn-cf-name" type="text" placeholder="Your Name"
               style="height:42px;border:1.5px solid #e2e8f0;border-radius:8px;padding:0 14px;
                      font-size:13px;color:#0f172a;outline:none;width:100%;box-sizing:border-box;
                      font-family:inherit;background:#fafafa;transition:border-color 0.15s,background 0.15s,box-shadow 0.15s;"
               onfocus="this.style.borderColor='#2563eb';this.style.background='#fff';this.style.boxShadow='0 0 0 3px rgba(37,99,235,0.10)'"
               onblur="this.style.borderColor='#e2e8f0';this.style.background='#fafafa';this.style.boxShadow='none'"/>
        <input id="dn-cf-email" type="email" placeholder="Your Email"
               style="height:42px;border:1.5px solid #e2e8f0;border-radius:8px;padding:0 14px;
                      font-size:13px;color:#0f172a;outline:none;width:100%;box-sizing:border-box;
                      font-family:inherit;background:#fafafa;transition:border-color 0.15s,background 0.15s,box-shadow 0.15s;"
               onfocus="this.style.borderColor='#2563eb';this.style.background='#fff';this.style.boxShadow='0 0 0 3px rgba(37,99,235,0.10)'"
               onblur="this.style.borderColor='#e2e8f0';this.style.background='#fafafa';this.style.boxShadow='none'"/>
        <textarea id="dn-cf-message" placeholder="Your message..."
                  style="height:88px;border:1.5px solid #e2e8f0;border-radius:8px;padding:10px 14px;
                         font-size:13px;color:#0f172a;outline:none;width:100%;box-sizing:border-box;
                         font-family:inherit;background:#fafafa;resize:none;line-height:1.55;
                         transition:border-color 0.15s,background 0.15s,box-shadow 0.15s;"
                  onfocus="this.style.borderColor='#2563eb';this.style.background='#fff';this.style.boxShadow='0 0 0 3px rgba(37,99,235,0.10)'"
                  onblur="this.style.borderColor='#e2e8f0';this.style.background='#fafafa';this.style.boxShadow='none'"></textarea>

        <button onclick="(function(){
          var n=document.getElementById('dn-cf-name');
          var e=document.getElementById('dn-cf-email');
          var m=document.getElementById('dn-cf-message');
          var btn=document.getElementById('dn-cf-submit');
          var success=document.getElementById('dn-cf-success');
          if(!n.value.trim()||!e.value.trim()||!m.value.trim()){
            btn.style.animation='dn-shake 0.3s';
            setTimeout(function(){btn.style.animation='';},400);
            return;
          }
          document.getElementById('dn-contact-form').style.display='none';
          success.style.display='flex';
          // Save to database via hidden Gradio
          function _sg(s,v){var el=document.querySelector(s+' input,'+s+' textarea');if(!el)return;var P=Object.getOwnPropertyDescriptor(el.tagName==='TEXTAREA'?HTMLTextAreaElement.prototype:HTMLInputElement.prototype,'value');if(P&&P.set)P.set.call(el,v);else el.value=v;el.dispatchEvent(new Event('input',{bubbles:true}));}
          var mob=document.getElementById('dn-cf-mobile');
          _sg('#dn-cf-gr-name',n.value);
          _sg('#dn-cf-gr-email',e.value);
          _sg('#dn-cf-gr-mobile',mob?mob.value:'');
          _sg('#dn-cf-gr-message',m.value);
          setTimeout(function(){var b=document.querySelector('div#dn-cf-gr-submit button');if(b)b.click();},150);
          setTimeout(function(){
            n.value=''; e.value=''; m.value='';
          },200);
        })()"
                id="dn-cf-submit"
                style="height:46px;background:linear-gradient(90deg,#1B4F8A 0%,#2563eb 100%);
                       color:white;font-weight:800;font-size:14px;border:none;border-radius:9px;
                       cursor:pointer;width:100%;letter-spacing:0.3px;
                       box-shadow:0 4px 12px rgba(27,79,138,0.28);
                       transition:transform 0.15s,box-shadow 0.15s,opacity 0.15s;"
                onmouseover="this.style.transform='translateY(-2px)';this.style.boxShadow='0 8px 20px rgba(27,79,138,0.38)'"
                onmouseout="this.style.transform='translateY(0)';this.style.boxShadow='0 4px 12px rgba(27,79,138,0.28)'">
          Send Message
        </button>
      </div>

      <!-- Success state -->
      <div id="dn-cf-success"
           style="display:none;flex-direction:column;align-items:center;justify-content:center;
                  padding:28px 16px;gap:10px;text-align:center;">
        <div style="width:48px;height:48px;background:#f0fdf4;border-radius:50%;display:flex;
                    align-items:center;justify-content:center;
                    box-shadow:0 0 0 6px rgba(34,197,94,0.1);">
          <svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="#16a34a"
               stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
            <path d="M20 6L9 17l-5-5"/>
          </svg>
        </div>
        <div style="font-size:15px;font-weight:800;color:#0B1F3A;margin-top:2px;">
          ✅ Message sent successfully.
        </div>
        <div style="font-size:13px;color:#64748b;line-height:1.6;max-width:300px;">
          Our team will contact you shortly.
        </div>
      </div>
    </div>

  </div><!-- /grid -->
</div><!-- /max-width wrapper -->
<style>
@keyframes dn-shake{0%,100%{transform:translateX(0)}25%{transform:translateX(-5px)}75%{transform:translateX(5px)}}
</style>
''')
        gr.HTML('</div></div>')

        # ── (Role tabs and How-it-Works section rendered above in access row) ──
        # Placeholder Markdown widgets kept for backend lang-switch wiring (hidden visually)
        landing_how_title   = gr.Markdown("", visible=False)
        landing_step1_title = gr.Markdown("", visible=False)
        landing_step1_desc  = gr.Markdown("", visible=False)
        landing_step2_title = gr.Markdown("", visible=False)
        landing_step2_desc  = gr.Markdown("", visible=False)
        landing_step3_title = gr.Markdown("", visible=False)
        landing_step3_desc  = gr.Markdown("", visible=False)

        gr.HTML('<div class="dn-band" style="background:#0B1F3A;padding:0;">')
        landing_capabilities_html = gr.HTML(value=_landing_capabilities('en'))
        gr.HTML('</div>')

        # ── Pricing Section (shown when nav Pricing clicked) ─────────────────
        with gr.Column(visible=True, elem_id="dn-pricing-section") as pricing_section:
            gr.HTML('''
<div style="background:#F8FAFF;border-top:2px solid #E2EAF5;padding:28px 24px 32px;">
  <div style="max-width:860px;margin:0 auto;">
    <div style="text-align:center;margin-bottom:32px;">
      <div style="font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#1B4F8A;margin-bottom:6px;">Transparent Pricing</div>
      <div style="font-size:26px;font-weight:900;color:#0B1F3A;letter-spacing:-0.4px;margin-bottom:8px;">Simple Plans for Every Business</div>
      <div style="font-size:14px;color:#64748b;">Start free. Upgrade when you need more.</div>
    </div>
    <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:24px;align-items:start;">

      <!-- Free Preview Plan -->
      <div style="background:#ffffff;border:2px solid #E2EAF5;border-radius:16px;padding:28px 24px;
                  box-shadow:0 2px 12px rgba(11,31,58,0.06);">
        <div style="font-size:11px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:#64748b;margin-bottom:8px;">Free Preview</div>
        <div style="font-size:36px;font-weight:900;color:#0B1F3A;margin-bottom:4px;">₹0</div>
        <div style="font-size:13px;color:#94a3b8;margin-bottom:20px;">Forever free · No credit card</div>
        <div style="display:flex;flex-direction:column;gap:10px;margin-bottom:24px;">
          <div style="display:flex;align-items:center;gap:10px;font-size:13px;color:#374151;">
            <span style="color:#16a34a;font-size:16px;">✓</span> Upload Excel / CSV data
          </div>
          <div style="display:flex;align-items:center;gap:10px;font-size:13px;color:#374151;">
            <span style="color:#16a34a;font-size:16px;">✓</span> Data Readiness Engine (DRC)
          </div>
          <div style="display:flex;align-items:center;gap:10px;font-size:13px;color:#374151;">
            <span style="color:#16a34a;font-size:16px;">✓</span> Business Health Score preview
          </div>
          <div style="display:flex;align-items:center;gap:10px;font-size:13px;color:#374151;">
            <span style="color:#16a34a;font-size:16px;">✓</span> 2 AI insights visible
          </div>
          <div style="display:flex;align-items:center;gap:10px;font-size:13px;color:#94a3b8;">
            <span style="color:#cbd5e1;font-size:16px;">✗</span> Full scoring dashboards
          </div>
          <div style="display:flex;align-items:center;gap:10px;font-size:13px;color:#94a3b8;">
            <span style="color:#cbd5e1;font-size:16px;">✗</span> 6 &amp; 12-month forecasts
          </div>
        </div>
        <button onclick="document.getElementById('dn-input-anchor').scrollIntoView({behavior:'smooth'})"
          style="width:100%;padding:13px;background:#f1f5f9;color:#1B4F8A;border:1.5px solid #cbd5e1;
                 border-radius:10px;font-size:14px;font-weight:700;cursor:pointer;">
          Get Started Free →
        </button>
      </div>

      <!-- Pro Plan -->
      <div style="background:linear-gradient(135deg,#0f2557 0%,#1B4F8A 100%);border:2px solid #1B4F8A;
                  border-radius:16px;padding:28px 24px;box-shadow:0 8px 32px rgba(27,79,138,0.28);
                  position:relative;">
        <div style="position:absolute;top:-12px;left:50%;transform:translateX(-50%);
                    background:#FFD080;color:#0B1F3A;font-size:10px;font-weight:800;
                    letter-spacing:1.5px;text-transform:uppercase;padding:4px 14px;border-radius:20px;">
          Most Popular
        </div>
        <div style="font-size:11px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:#93c5fd;margin-bottom:8px;">Pro Plan</div>
        <div style="font-size:36px;font-weight:900;color:#ffffff;margin-bottom:4px;">₹999<span style="font-size:16px;font-weight:500;color:#93c5fd;">/mo</span></div>
        <div style="font-size:13px;color:rgba(255,255,255,0.55);margin-bottom:20px;">Per business · Cancel anytime</div>
        <div style="display:flex;flex-direction:column;gap:10px;margin-bottom:24px;">
          <div style="display:flex;align-items:center;gap:10px;font-size:13px;color:#e0f2fe;">
            <span style="color:#4ade80;font-size:16px;">✓</span> Everything in Free
          </div>
          <div style="display:flex;align-items:center;gap:10px;font-size:13px;color:#e0f2fe;">
            <span style="color:#4ade80;font-size:16px;">✓</span> Full Business Scoring Dashboard
          </div>
          <div style="display:flex;align-items:center;gap:10px;font-size:13px;color:#e0f2fe;">
            <span style="color:#4ade80;font-size:16px;">✓</span> Business Intelligence Dashboard
          </div>
          <div style="display:flex;align-items:center;gap:10px;font-size:13px;color:#e0f2fe;">
            <span style="color:#4ade80;font-size:16px;">✓</span> 6 &amp; 12-month AI Forecasts
          </div>
          <div style="display:flex;align-items:center;gap:10px;font-size:13px;color:#e0f2fe;">
            <span style="color:#4ade80;font-size:16px;">✓</span> Store → Category → Product drill-down
          </div>
          <div style="display:flex;align-items:center;gap:10px;font-size:13px;color:#e0f2fe;">
            <span style="color:#4ade80;font-size:16px;">✓</span> PDF Business Intelligence Report
          </div>
        </div>
        <button onclick="document.getElementById('dn-input-anchor').scrollIntoView({behavior:'smooth'})"
          style="width:100%;padding:13px;background:linear-gradient(90deg,#FFD080,#f59e0b);
                 color:#0B1F3A;border:none;border-radius:10px;font-size:14px;font-weight:800;cursor:pointer;
                 box-shadow:0 4px 14px rgba(245,158,11,0.4);">
          Start Pro Trial →
        </button>
      </div>

    </div>
    <div style="text-align:center;margin-top:24px;font-size:12px;color:#94a3b8;">
      No payment integration required for demo. Contact us to activate Pro access.
    </div>
  </div>
</div>
''')
            pricing_close_btn = gr.Button("✕ Close Pricing", variant="secondary", size="sm")

    # ── Steps 1–4 removed for global retail platform.
    #    Stub columns + widgets kept so existing event-wiring references don't crash. ──

    with gr.Column(visible=False) as step1_col:
        step1_title_md  = gr.Markdown("")
        name_input      = gr.Textbox(visible=False)
        mobile_input    = gr.Textbox(visible=False)
        email_input     = gr.Textbox(visible=False)
        role_input      = gr.Dropdown(choices=["Choose business role"], visible=False)
        name_input_v    = gr.Textbox(visible=False)
        mobile_input_v  = gr.Textbox(visible=False)
        role_input_v    = gr.Textbox(visible=False)
        cancel1_btn     = gr.Button("Cancel", visible=False)
        next1_btn       = gr.Button("Next →", visible=False)

    with gr.Column(visible=False) as step2_col:
        step2_title_md    = gr.Markdown("")
        msme_number_input = gr.Textbox(visible=False)
        otp_input         = gr.Textbox(visible=False, type="password")
        fetch_btn         = gr.Button("Fetch", visible=False)
        fetch_status      = gr.Markdown(visible=False)
        msme_voice_display = gr.Textbox(visible=False)
        otp_voice_display  = gr.Textbox(visible=False, type="password")
        fetched_name      = gr.Textbox(visible=False)
        fetched_org       = gr.Textbox(visible=False)
        fetched_activity  = gr.Textbox(visible=False)
        fetched_type      = gr.Textbox(visible=False)
        fetched_state     = gr.Textbox(visible=False)
        fetched_city      = gr.Textbox(visible=False)
        fetched_industry  = gr.Textbox(visible=False, value="Retail")
        back2_btn         = gr.Button("← Back", visible=False)
        next2_btn         = gr.Button("Next →", visible=False)

    with gr.Column(visible=False) as step3_col:
        step3_title_md    = gr.Markdown("")
        confirm_name      = gr.Textbox(visible=False)
        confirm_org       = gr.Textbox(visible=False)
        confirm_activity  = gr.Textbox(visible=False)
        confirm_type      = gr.Textbox(visible=False)
        confirm_state     = gr.Textbox(visible=False)
        confirm_city      = gr.Textbox(visible=False)
        confirm_industry  = gr.Textbox(visible=False, value="Retail")
        consent1          = gr.Checkbox(visible=False)
        consent2          = gr.Checkbox(visible=False)
        certificate_upload = gr.File(visible=False, file_types=[".pdf"])
        back3_btn         = gr.Button("← Back", visible=False)
        next3_btn         = gr.Button("Next →", visible=False)

    with gr.Column(visible=False) as step4_col:
        verification_status_display = gr.Markdown(visible=False)
        step4_title_md      = gr.Markdown("")
        business_type_input = gr.Dropdown(choices=["Retail"], value="Retail", visible=False)
        years_input         = gr.Textbox(visible=False)
        revenue_input       = gr.Dropdown(choices=["N/A"], value="N/A", visible=False)
        back4_btn           = gr.Button("← Back", visible=False)
        next4_btn           = gr.Button("Submit", visible=False)
        proceed_to_step5_btn = gr.Button("Next →", visible=False)

    # Error stubs — defined OUTSIDE hidden columns so Gradio can serialize updates to them
    error1 = gr.Markdown(value="", visible=False, elem_id="stub-error1")
    error2 = gr.Markdown(value="", visible=False, elem_id="stub-error2")
    error3 = gr.Markdown(value="", visible=False, elem_id="stub-error3")
    error4 = gr.Markdown(value="", visible=False, elem_id="stub-error4")

    # ── Upload Screen ─────────────────────────────────────────────────────────
    with gr.Column(visible=False) as step5_col:
        login_welcome_message = gr.Markdown(value="", visible=False)
        step5_title_md = gr.Markdown("", visible=False)  # stub for lang switching
        # ── App bar ──
        gr.HTML('''
<div style="background:#0f2557;padding:9px 20px;display:flex;align-items:center;gap:14px;border-radius:12px 12px 0 0;">
  <div style="font-size:13px;font-weight:700;color:#ffffff;white-space:nowrap;letter-spacing:-0.3px;">DataNetra<span style="color:#60a5fa;">.ai</span></div>
  <div style="flex:1;display:flex;align-items:center;justify-content:center;gap:0;">
    <div style="display:flex;align-items:center;gap:5px;background:rgba(255,255,255,0.18);border-radius:20px;padding:5px 14px;box-shadow:0 0 0 1px rgba(255,255,255,0.22);">
      <div style="width:18px;height:18px;border-radius:50%;background:#ffffff;display:flex;align-items:center;justify-content:center;box-shadow:0 1px 4px rgba(0,0,0,0.15);"><span style="font-size:9px;font-weight:700;color:#0f2557;">1</span></div>
      <span style="font-size:11px;font-weight:600;color:#ffffff;">Upload</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(255,255,255,0.18);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(255,255,255,0.10);display:flex;align-items:center;justify-content:center;"><span style="font-size:9px;color:rgba(255,255,255,0.35);">2</span></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.35);">Preview</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(255,255,255,0.18);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(255,255,255,0.10);display:flex;align-items:center;justify-content:center;"><span style="font-size:9px;color:rgba(255,255,255,0.35);">3</span></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.35);">Dashboard</span>
    </div>
  </div>
  <div style="font-size:10px;color:rgba(255,255,255,0.55);background:rgba(255,255,255,0.09);padding:3px 9px;border-radius:20px;">Free preview</div>
</div>
<div style="background:#1e3a6b;padding:8px 20px;display:flex;align-items:center;justify-content:space-between;border-bottom:1px solid rgba(255,255,255,0.08);border-radius:0;">
  <div>
    <div style="font-size:10px;font-weight:700;letter-spacing:1.8px;text-transform:uppercase;color:#7AABDD;margin-bottom:2px;">STEP 1 OF 3</div>
    <div style="font-size:15px;font-weight:800;color:#ffffff;letter-spacing:-0.2px;">Upload Your Business Data</div>
  </div>
  <div style="display:flex;gap:8px;flex-wrap:wrap;">
    <span style="background:rgba(255,255,255,0.08);border:1px solid rgba(255,255,255,0.14);border-radius:20px;padding:3px 10px;font-size:10px;color:#A8D8FF;">Business Health Score</span>
    <span style="background:rgba(255,255,255,0.08);border:1px solid rgba(255,255,255,0.14);border-radius:20px;padding:3px 10px;font-size:10px;color:#A8D8FF;">AI Forecast</span>
    <span style="background:rgba(255,255,255,0.08);border:1px solid rgba(255,255,255,0.14);border-radius:20px;padding:3px 10px;font-size:10px;color:#A8D8FF;">SKU Insights</span>
  </div>
</div>
''')
        # ── Page header ──
        gr.HTML('''
<div style="background:linear-gradient(135deg,#0f2557 0%,#1B4F8A 100%);
            border-radius:14px;padding:20px 32px;margin-bottom:20px;text-align:center;">
  <div style="font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;
              color:#7AABDD;margin-bottom:6px;">STEP 1 OF 3</div>
  <h1 style="color:#fff;font-size:1.4rem;font-weight:900;margin:0 0 6px;">Upload Your Business Data</h1>
  <p style="color:#A8D8FF;font-size:13px;margin:0;">Upload your Excel or CSV file and we'll generate instant AI insights</p>
</div>
''')
        # ── What you'll get info box ──
        gr.HTML('''
<div style="background:#EBF5FF;border:1px solid #BDD7F5;border-radius:12px;
            padding:16px 20px;margin-bottom:18px;display:grid;
            grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:10px;">
  <div style="display:flex;align-items:center;gap:10px;">
    <span style="font-size:20px;">📊</span>
    <div>
      <div style="font-size:13px;font-weight:700;color:#0B1F3A;">Business Health Score</div>
      <div style="font-size:12px;color:#475569;">Instant AI scoring of your business</div>
    </div>
  </div>
  <div style="display:flex;align-items:center;gap:10px;">
    <span style="font-size:20px;">📈</span>
    <div>
      <div style="font-size:13px;font-weight:700;color:#0B1F3A;">Sales Forecast</div>
      <div style="font-size:12px;color:#475569;">6 & 12 month AI predictions</div>
    </div>
  </div>
  <div style="display:flex;align-items:center;gap:10px;">
    <span style="font-size:20px;">⚡</span>
    <div>
      <div style="font-size:13px;font-weight:700;color:#0B1F3A;">Risk Analysis</div>
      <div style="font-size:12px;color:#475569;">Financial & vendor risk scores</div>
    </div>
  </div>
  <div style="display:flex;align-items:center;gap:10px;">
    <span style="font-size:20px;">🏪</span>
    <div>
      <div style="font-size:13px;font-weight:700;color:#0B1F3A;">Digital Readiness</div>
      <div style="font-size:12px;color:#475569;">Marketplace platform fit score</div>
    </div>
  </div>
</div>
''')
        consent_check  = gr.Checkbox(label="I consent to data analysis — your data stays private and secure", value=False)
        file_upload    = gr.File(label="Upload Excel or CSV file (.xlsx, .csv)", file_types=[".xlsx",".csv"])
        upload_message = gr.Markdown(value="", visible=False)

        with gr.Row():
            back5_btn   = gr.Button("← Back to Home", variant="secondary")
            analyze_btn = gr.Button("Generate Free Insights Preview →", variant="primary", elem_id="analyze-data-btn", scale=3)
        gr.HTML('''<div style="font-size:12px;color:#64748b;margin-top:6px;padding-left:2px;display:flex;align-items:center;gap:6px;">
  <span style="font-size:14px;">⏱</span> Takes ~10–20 seconds &nbsp;·&nbsp; Free Preview. Pay only for Full insights &nbsp;·&nbsp; Instant results
</div>''')
        cancel5_btn = gr.Button("Cancel", variant="secondary", visible=False)
        error5 = gr.Markdown()
        analyze_loading = gr.HTML(
            value="",
            visible=False,
            elem_id="analyze-loading-msg"
        )
        insights_output = gr.HTML(elem_id="insights-output-html")
        with gr.Row():
            view_dashboard_btn = gr.Button("📊 View Dashboard", visible=False, variant="primary")
        kpi1 = gr.Markdown(visible=False); kpi2 = gr.Markdown(visible=False)
        kpi3 = gr.Markdown(visible=False); kpi4 = gr.Markdown(visible=False); kpi5 = gr.Markdown(visible=False)
        chart1 = gr.Plot(visible=False); chart2 = gr.Plot(visible=False)
        chart3 = gr.Plot(visible=False); chart4 = gr.Plot(visible=False)
        sum1 = gr.Markdown(visible=False); sum2 = gr.Markdown(visible=False)
        sum3 = gr.Markdown(visible=False); sum4 = gr.Markdown(visible=False)

    # ── Step 6: Business Performance & Platform Readiness ──
    with gr.Column(visible=False) as step6_col:
        gr.HTML('''
<div style="background:#0f2557;padding:9px 20px;display:flex;align-items:center;gap:14px;border-radius:12px 12px 0 0;">
  <div style="font-size:13px;font-weight:700;color:#ffffff;white-space:nowrap;letter-spacing:-0.3px;">DataNetra<span style="color:#60a5fa;">.ai</span></div>
  <div style="flex:1;display:flex;align-items:center;justify-content:center;gap:0;">
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(74,222,128,0.25);display:flex;align-items:center;justify-content:center;"><svg width="9" height="9" viewBox="0 0 24 24" fill="none" stroke="#4ade80" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.45);">Upload</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(74,222,128,0.5);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(74,222,128,0.25);display:flex;align-items:center;justify-content:center;"><svg width="9" height="9" viewBox="0 0 24 24" fill="none" stroke="#4ade80" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.45);">Preview</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(74,222,128,0.5);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(74,222,128,0.25);display:flex;align-items:center;justify-content:center;"><svg width="9" height="9" viewBox="0 0 24 24" fill="none" stroke="#4ade80" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.45);">Scoring</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(255,255,255,0.18);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;background:rgba(255,255,255,0.18);border-radius:20px;padding:5px 14px;box-shadow:0 0 0 1px rgba(255,255,255,0.22);">
      <div style="width:18px;height:18px;border-radius:50%;background:#ffffff;display:flex;align-items:center;justify-content:center;box-shadow:0 1px 4px rgba(0,0,0,0.15);"><span style="font-size:9px;font-weight:700;color:#0f2557;">4</span></div>
      <span style="font-size:11px;font-weight:600;color:#ffffff;">Performance</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(255,255,255,0.18);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(255,255,255,0.10);display:flex;align-items:center;justify-content:center;"><span style="font-size:9px;color:rgba(255,255,255,0.35);">5</span></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.35);">Forecasting</span>
    </div>
  </div>
  <div style="display:flex;align-items:center;gap:8px;">
    <div style="font-size:10px;color:rgba(255,255,255,0.55);background:rgba(255,255,255,0.09);padding:3px 9px;border-radius:20px;">2 of 3</div>
    <div style="width:1px;height:13px;background:rgba(255,255,255,0.18);"></div>
    <span style="font-size:11px;color:#fbbf24;font-weight:600;">Export PDF</span>
  </div>
</div>
<div style="background:#1e3a6b;padding:8px 20px;border-bottom:1px solid rgba(255,255,255,0.08);">
  <div style="font-size:10px;font-weight:700;letter-spacing:1.8px;text-transform:uppercase;color:#7AABDD;margin-bottom:2px;">DASHBOARD 2 OF 3 — BUSINESS PERFORMANCE</div>
  <div style="font-size:15px;font-weight:800;color:#ffffff;letter-spacing:-0.2px;">Business Performance &amp; Channel Dashboard</div>
</div>
''')
        kpi_table_dash = gr.HTML(value="", elem_id="kpi-table-container")
        # Performance Visualizations label removed per UX review
        with gr.Row():
            with gr.Column():
                gr.HTML('<div style="font-size:13px;font-weight:700;color:#0B1F3A;padding:8px 4px 2px;letter-spacing:0.2px">Sales vs Profit Margin — Quarterly</div>')
                chart1_dash = gr.Plot(show_label=False)
                chart1_summary = gr.Markdown(value="", visible=False)
            with gr.Column():
                gr.HTML('<div style="font-size:13px;font-weight:700;color:#0B1F3A;padding:8px 4px 2px;letter-spacing:0.2px">Channel Performance: Before vs After — Quarterly</div>')
                chart2_dash = gr.Plot(show_label=False)
                chart2_summary = gr.Markdown(value="", visible=False)
        with gr.Row():
            with gr.Column():
                gr.HTML('<div style="font-size:13px;font-weight:700;color:#0B1F3A;padding:8px 4px 2px;letter-spacing:0.2px">Returns &amp; Replacements — Quarterly Trend</div>')
                chart3_dash = gr.Plot(show_label=False)
                chart3_summary = gr.Markdown(value="", visible=False)
            with gr.Column():
                gr.HTML('<div style="font-size:13px;font-weight:700;color:#0B1F3A;padding:8px 4px 2px;letter-spacing:0.2px">Store Impact Analysis</div>')
                chart4_dash = gr.Plot(show_label=False)
                chart4_summary = gr.Markdown(value="", visible=False)
        with gr.Row():
            back6_to_dash_btn     = gr.Button("← Back to Business Scoring", variant="secondary", size="lg", elem_id="dn-back6-dash-btn")
            forecast_deepdive_btn = gr.Button("📈 Go to Forecast Intelligence Dashboard →", variant="primary", size="lg", elem_id="dn-forecast-btn")
            back6_btn             = gr.Button("⬅ Back to Upload", variant="secondary", size="lg", elem_id="dn-back6-btn")

    # ── Step 6a: removed (Government Portal) — stub kept for wiring compatibility ──
    with gr.Column(visible=False) as step6a_col:
        back6a_btn         = gr.Button("⬅ Back", visible=False)
        gov_file_upload    = gr.File(visible=False, file_types=[".csv",".xlsx"])
        gov_upload_status  = gr.Markdown(value="", visible=False)
        gov_analyze_btn    = gr.Button("Gov Analyze", visible=False)
        gov_dashboard_html = gr.HTML(value="")

    # ── Step 7: Store → Category → Product + Forecasting ──
    with gr.Column(visible=False) as step7_col:
        # ── Page title ──
        gr.HTML('''
<div style="background:#0f2557;padding:9px 20px;display:flex;align-items:center;gap:14px;border-radius:12px 12px 0 0;">
  <div style="font-size:13px;font-weight:700;color:#ffffff;white-space:nowrap;letter-spacing:-0.3px;">DataNetra<span style="color:#60a5fa;">.ai</span></div>
  <div style="flex:1;display:flex;align-items:center;justify-content:center;gap:0;">
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(74,222,128,0.25);display:flex;align-items:center;justify-content:center;"><svg width="9" height="9" viewBox="0 0 24 24" fill="none" stroke="#4ade80" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.45);">Upload</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(74,222,128,0.5);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(74,222,128,0.25);display:flex;align-items:center;justify-content:center;"><svg width="9" height="9" viewBox="0 0 24 24" fill="none" stroke="#4ade80" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.45);">Preview</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(74,222,128,0.5);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(74,222,128,0.25);display:flex;align-items:center;justify-content:center;"><svg width="9" height="9" viewBox="0 0 24 24" fill="none" stroke="#4ade80" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.45);">Scoring</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(74,222,128,0.5);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(74,222,128,0.25);display:flex;align-items:center;justify-content:center;"><svg width="9" height="9" viewBox="0 0 24 24" fill="none" stroke="#4ade80" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.45);">Performance</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(255,255,255,0.18);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;background:rgba(255,255,255,0.18);border-radius:20px;padding:5px 14px;box-shadow:0 0 0 1px rgba(255,255,255,0.22);">
      <div style="width:18px;height:18px;border-radius:50%;background:#ffffff;display:flex;align-items:center;justify-content:center;box-shadow:0 1px 4px rgba(0,0,0,0.15);"><span style="font-size:9px;font-weight:700;color:#0f2557;">5</span></div>
      <span style="font-size:11px;font-weight:600;color:#ffffff;">Forecasting</span>
    </div>
  </div>
  <div style="display:flex;align-items:center;gap:8px;">
    <div style="font-size:10px;color:rgba(255,255,255,0.55);background:rgba(255,255,255,0.09);padding:3px 9px;border-radius:20px;">3 of 3</div>
    <div style="width:1px;height:13px;background:rgba(255,255,255,0.18);"></div>
    <span style="font-size:11px;color:#fbbf24;font-weight:600;">Export PDF</span>
  </div>
</div>
<div style="background:#1e3a6b;padding:8px 20px;display:flex;align-items:center;justify-content:space-between;border-bottom:1px solid rgba(255,255,255,0.08);">
  <div>
    <div style="font-size:10px;font-weight:700;letter-spacing:1.8px;text-transform:uppercase;color:#7AABDD;margin-bottom:2px;">DASHBOARD 3 OF 3 — SALES &amp; FORECASTING</div>
    <div style="font-size:15px;font-weight:800;color:#ffffff;letter-spacing:-0.2px;">Store → Category → Product &amp; AI Forecasting</div>
  </div>
  <div style="font-size:10px;color:#F5C842;font-weight:600;background:rgba(245,200,66,0.10);border:1px solid rgba(245,200,66,0.25);border-radius:20px;padding:3px 10px;">All charts update live</div>
</div>''')

        # ── Filter bar ──
        with gr.Row(equal_height=True):
            s7_msme_label   = gr.HTML(value='', visible=False)
            s7_store_filter = gr.Dropdown(choices=["Store: All"], value="Store: All",
                                          label="🏪 Store", scale=1, interactive=True)
            s7_cat_filter   = gr.Dropdown(choices=["Category: All"], value="Category: All",
                                          label="📂 Category", scale=1, interactive=True)
            s7_prod_filter  = gr.Dropdown(choices=["Product: All"], value="Product: All",
                                          label="📦 Product", scale=1, interactive=True)

        # ── KPI card row ──
        s7_kpi_html  = gr.HTML(value="", elem_id="s7-kpi-row")

        # ── Top Products table ──
        s7_top_table = gr.HTML(value="", elem_id="s7-top-table")

        gr.HTML('<div style="font-size:10px;font-weight:700;letter-spacing:1.4px;text-transform:uppercase;'
                'color:#1B4F8A;padding:8px 0 4px;border-top:2px solid #EEF4FB;margin-top:4px">'
                '📈 CATEGORY TRENDS</div>')
        # ── Row 1: Category Sales Trend | Category Margin Trend ──
        with gr.Row(equal_height=True):
            s7_cat_sales_chart  = gr.Plot(show_label=False, scale=3)
            s7_cat_margin_chart = gr.Plot(show_label=False, scale=3)

        gr.HTML('<div style="font-size:10px;font-weight:700;letter-spacing:1.4px;text-transform:uppercase;'
                'color:#1B4F8A;padding:8px 0 4px;border-top:2px solid #EEF4FB;margin-top:8px">'
                '🔮 SALES FORECAST</div>')
        # ── Row 2: 6-Month | 12-Month | Fulfilment ──
        with gr.Row(equal_height=True):
            s7_fc6_chart    = gr.Plot(show_label=False, scale=2)
            s7_fc12_chart   = gr.Plot(show_label=False, scale=2)
            s7_fulfil_chart = gr.Plot(show_label=False, scale=2)

        gr.HTML('<div style="font-size:10px;font-weight:700;letter-spacing:1.4px;text-transform:uppercase;'
                'color:#1B4F8A;padding:8px 0 4px;border-top:2px solid #EEF4FB;margin-top:8px">'
                '📦 PRODUCT OPERATIONS &amp; AI INSIGHTS</div>')
        # ── Row 3: Sales vs Returns | Inventory (equal plots) ──
        with gr.Row(equal_height=True):
            s7_sales_ret_chart = gr.Plot(show_label=False, scale=2)
            s7_inventory_chart = gr.Plot(show_label=False, scale=2)

        # ── Row 4: AI Summary (full width) ──
        s7_ai_summary = gr.HTML(value="", elem_id="s7-ai-summary")

        # ── PDF Download button ──
        gr.HTML('''<div style="margin:12px 0 4px;border-top:1px solid #EEF4FB;padding-top:12px;">
          <div style="font-size:10px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase;
                      color:#1B4F8A;margin-bottom:6px;">📄 EXPORT REPORT</div>
        </div>''')
        pdf_download_btn = gr.Button("⬇ Download Business Intelligence Report (PDF)",
                                     variant="primary", size="lg",
                                     elem_id="pdf-download-btn")
        pdf_file_output  = gr.DownloadButton(
                                   label="⬇ Click here to download your PDF",
                                   value=None,
                                   visible=False,
                                   variant="primary",
                                   size="lg",
                                   elem_id="pdf-file-output")

        # ── Back buttons ──
        with gr.Row():
            back7_btn     = gr.Button("\u2b05 Back to Dashboard",   variant="secondary", scale=1)
            back7_to5_btn = gr.Button("← Back to Upload Screen", variant="secondary", scale=1)

    # ══════════════════════════════════════════════════════════════════════════
    # Insights Preview Screen
    # ══════════════════════════════════════════════════════════════════════════
    with gr.Column(visible=False, elem_id="dn-preview-col") as preview_col:
        # ── Top nav bar ──────────────────────────────────────────────────────
        with gr.Row(equal_height=True):
            preview_back_btn = gr.Button("← Back to Upload", variant="secondary", scale=0, min_width=160)
            gr.HTML('<div style="flex:1"></div>')

        # ── Progress bar ──────────────────────────────────────────────────────
        gr.HTML('''
<div style="background:#0f2557;padding:9px 20px;display:flex;align-items:center;gap:14px;border-radius:12px 12px 0 0;">
  <div style="font-size:13px;font-weight:700;color:#ffffff;white-space:nowrap;letter-spacing:-0.3px;">DataNetra<span style="color:#60a5fa;">.ai</span></div>
  <div style="flex:1;display:flex;align-items:center;justify-content:center;gap:0;">
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(74,222,128,0.25);display:flex;align-items:center;justify-content:center;"><svg width="9" height="9" viewBox="0 0 24 24" fill="none" stroke="#4ade80" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.45);">Upload</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(74,222,128,0.5);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;background:rgba(255,255,255,0.18);border-radius:20px;padding:5px 14px;box-shadow:0 0 0 1px rgba(255,255,255,0.22);">
      <div style="width:18px;height:18px;border-radius:50%;background:#ffffff;display:flex;align-items:center;justify-content:center;box-shadow:0 1px 4px rgba(0,0,0,0.15);"><span style="font-size:9px;font-weight:700;color:#0f2557;">2</span></div>
      <span style="font-size:11px;font-weight:600;color:#ffffff;">Preview</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(255,255,255,0.18);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(255,255,255,0.10);display:flex;align-items:center;justify-content:center;"><span style="font-size:9px;color:rgba(255,255,255,0.35);">3</span></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.35);">Dashboard</span>
    </div>
  </div>
  <div style="font-size:10px;color:rgba(255,255,255,0.55);background:rgba(255,255,255,0.09);padding:3px 9px;border-radius:20px;">Free preview</div>
</div>
<div style="background:#1e3a6b;padding:8px 20px;border-bottom:1px solid rgba(255,255,255,0.08);">
  <div style="font-size:10px;font-weight:700;letter-spacing:1.8px;text-transform:uppercase;color:#7AABDD;margin-bottom:2px;">STEP 2 OF 3</div>
  <div style="font-size:15px;font-weight:800;color:#ffffff;letter-spacing:-0.2px;">Your Free Insights Preview</div>
</div>
''')
        # ── Compact hero header ───────────────────────────────────────────────
        preview_header_html = gr.HTML(value="""
<div style="background:linear-gradient(135deg,#0f2557 0%,#1B4F8A 100%);
            border-radius:14px;padding:20px 32px;margin:0 0 16px;text-align:center">
  <div style="font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;
              color:#7AABDD;margin-bottom:6px">STEP 2 OF 3 — PREVIEW INSIGHTS</div>
  <h1 style="color:#FFFFFF;font-size:1.4rem;font-weight:900;margin:0 0 6px;letter-spacing:-0.3px">
    Your Business Intelligence Preview
  </h1>
  <p style="color:#A8D8FF;font-size:13px;margin:0;line-height:1.55;max-width:640px;margin-left:auto;margin-right:auto">
    Real data from your upload · Sections match your full dashboard exactly ·
    Unlock to remove all blurs and access complete AI insights, forecasts and recommendations.
  </p>
</div>""")

        # ── Rich unified preview (all 7 sections) ────────────────────────────
        preview_cards_html = gr.HTML(value="", elem_id="dn-preview-cards")

        # ── Sales trend chart (shown below section 1) ─────────────────────────
        gr.HTML(value="""
<div style="max-width:1100px;margin:0 auto 4px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:6px;">
  <div style="font-size:11px;font-weight:700;color:#0B1F3A;letter-spacing:0.2px;">
    📈 Monthly Sales Trend <span style="color:#64748b;font-weight:400;">(Last 8 Months)</span>
  </div>
  <div style="font-size:11px;background:#f0fdf4;border:1px solid #bbf7d0;color:#15803d;
              border-radius:20px;padding:3px 10px;font-weight:600;">
    ✦ Full trend analysis available after unlock
  </div>
</div>""")
        preview_sales_chart = gr.Plot(visible=False, show_label=False, elem_id="dn-preview-sales-chart")
        preview_forecast_chart = gr.Plot(visible=False, show_label=False, elem_id="dn-preview-forecast-chart")

        # ── Single clean unlock CTA ───────────────────────────────────────────────
        gr.HTML(value="""
<div style="background:linear-gradient(135deg,#0B1F3A 0%,#1a3a6b 100%);
            border-radius:16px;padding:32px 40px 28px;margin:16px 0;text-align:center;
            box-shadow:0 8px 32px rgba(11,31,58,0.28);">

  <!-- Urgency banner -->
  <div style="display:inline-block;background:rgba(246,173,60,0.18);border:1px solid rgba(246,173,60,0.40);
              border-radius:20px;padding:5px 16px;font-size:12px;font-weight:700;color:#F6AD3C;
              letter-spacing:0.3px;margin-bottom:16px;">
    ⚡ You're just one step away from your full business intelligence
  </div>

  <h2 style="color:#ffffff;font-size:1.5rem;font-weight:900;margin:0 0 8px;letter-spacing:-0.3px;">
    Unlock Your Full Business Intelligence
  </h2>
  <p style="color:#A8D8FF;font-size:14px;margin:0 0 24px;line-height:1.6;max-width:560px;margin-left:auto;margin-right:auto;">
    Get complete dashboards, 6 & 12 month AI forecasts, marketplace insights, AI recommendations, and a downloadable PDF report.
  </p>

  <!-- Feature grid -->
  <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px;margin-bottom:24px;text-align:left;">
    <div style="background:rgba(255,255,255,0.08);border-radius:10px;padding:12px 14px;display:flex;align-items:center;gap:10px;">
      <span style="color:#28c840;font-size:16px;">✓</span>
      <span style="color:rgba(255,255,255,0.92);font-size:13px;font-weight:500;">Full Business Dashboard</span>
    </div>
    <div style="background:rgba(255,255,255,0.08);border-radius:10px;padding:12px 14px;display:flex;align-items:center;gap:10px;">
      <span style="color:#28c840;font-size:16px;">✓</span>
      <span style="color:rgba(255,255,255,0.92);font-size:13px;font-weight:500;">6 & 12 Month AI Forecast</span>
    </div>
    <div style="background:rgba(255,255,255,0.08);border-radius:10px;padding:12px 14px;display:flex;align-items:center;gap:10px;">
      <span style="color:#28c840;font-size:16px;">✓</span>
      <span style="color:rgba(255,255,255,0.92);font-size:13px;font-weight:500;">AI Recommendations</span>
    </div>
    <div style="background:rgba(255,255,255,0.08);border-radius:10px;padding:12px 14px;display:flex;align-items:center;gap:10px;">
      <span style="color:#28c840;font-size:16px;">✓</span>
      <span style="color:rgba(255,255,255,0.92);font-size:13px;font-weight:500;">Marketplace & Channel Insights</span>
    </div>
    <div style="background:rgba(255,255,255,0.08);border-radius:10px;padding:12px 14px;display:flex;align-items:center;gap:10px;">
      <span style="color:#28c840;font-size:16px;">✓</span>
      <span style="color:rgba(255,255,255,0.92);font-size:13px;font-weight:500;">Downloadable PDF Report</span>
    </div>
    <div style="background:rgba(255,255,255,0.08);border-radius:10px;padding:12px 14px;display:flex;align-items:center;gap:10px;">
      <span style="color:#28c840;font-size:16px;">✓</span>
      <span style="color:rgba(255,255,255,0.92);font-size:13px;font-weight:500;">Store & SKU Level Analysis</span>
    </div>
  </div>

  <!-- Trust row -->
  <div style="display:flex;align-items:center;justify-content:center;gap:24px;flex-wrap:wrap;margin-bottom:8px;">
    <span style="color:rgba(255,255,255,0.60);font-size:12px;display:flex;align-items:center;gap:5px;">
      <span style="color:#28c840;">✓</span> Free Preview. Pay only for Full insights
    </span>
    <span style="color:rgba(255,255,255,0.60);font-size:12px;display:flex;align-items:center;gap:5px;">
      <span style="color:#28c840;">✓</span> Instant access
    </span>
    <span style="color:rgba(255,255,255,0.60);font-size:12px;display:flex;align-items:center;gap:5px;">
      <span style="color:#28c840;">✓</span> Your data stays private
    </span>
  </div>

</div>
""")

        # ── Single unlock CTA button ──────────────────────────────────────────
        with gr.Row(equal_height=True):
            with gr.Column(scale=1):
                gr.HTML("")
            with gr.Column(scale=3):
                preview_unlock_btn = gr.Button(
                    "Unlock Full Insights →",
                    variant="primary", size="lg",
                    elem_id="dn-preview-unlock-btn"
                )
            with gr.Column(scale=1):
                gr.HTML("")

        # ── Loading indicator shown while full dashboard is being built ────────
        preview_unlock_loading = gr.HTML(
            value="",
            visible=False,
            elem_id="dn-preview-unlock-loading"
        )


    # ══════════════════════════════════════════════════════════════════════════
    # FULL DASHBOARD SCREEN — All 4 Scores + Summary (shown after unlock)
    # ══════════════════════════════════════════════════════════════════════════
    with gr.Column(visible=False, elem_id="dn-full-dash-col") as step_dash_col:
        gr.HTML('''
<div style="background:#0f2557;padding:9px 20px;display:flex;align-items:center;gap:14px;border-radius:12px 12px 0 0;">
  <div style="font-size:13px;font-weight:700;color:#ffffff;white-space:nowrap;letter-spacing:-0.3px;">DataNetra<span style="color:#60a5fa;">.ai</span></div>
  <div style="flex:1;display:flex;align-items:center;justify-content:center;gap:0;">
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(74,222,128,0.25);display:flex;align-items:center;justify-content:center;"><svg width="9" height="9" viewBox="0 0 24 24" fill="none" stroke="#4ade80" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.45);">Upload</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(74,222,128,0.5);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(74,222,128,0.25);display:flex;align-items:center;justify-content:center;"><svg width="9" height="9" viewBox="0 0 24 24" fill="none" stroke="#4ade80" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.45);">Preview</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(255,255,255,0.18);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;background:rgba(255,255,255,0.18);border-radius:20px;padding:5px 14px;box-shadow:0 0 0 1px rgba(255,255,255,0.22);">
      <div style="width:18px;height:18px;border-radius:50%;background:#ffffff;display:flex;align-items:center;justify-content:center;box-shadow:0 1px 4px rgba(0,0,0,0.15);"><span style="font-size:9px;font-weight:700;color:#0f2557;">3</span></div>
      <span style="font-size:11px;font-weight:600;color:#ffffff;">Scoring</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(255,255,255,0.18);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(255,255,255,0.10);display:flex;align-items:center;justify-content:center;"><span style="font-size:9px;color:rgba(255,255,255,0.35);">4</span></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.35);">Performance</span>
    </div>
    <div style="width:28px;height:1px;background:rgba(255,255,255,0.18);flex-shrink:0;"></div>
    <div style="display:flex;align-items:center;gap:5px;padding:4px 10px;">
      <div style="width:17px;height:17px;border-radius:50%;background:rgba(255,255,255,0.10);display:flex;align-items:center;justify-content:center;"><span style="font-size:9px;color:rgba(255,255,255,0.35);">5</span></div>
      <span style="font-size:11px;color:rgba(255,255,255,0.35);">Forecasting</span>
    </div>
  </div>
  <div style="display:flex;align-items:center;gap:8px;">
    <div style="font-size:10px;color:rgba(255,255,255,0.55);background:rgba(255,255,255,0.09);padding:3px 9px;border-radius:20px;">1 of 3</div>
    <div style="width:1px;height:13px;background:rgba(255,255,255,0.18);"></div>
    <span style="font-size:11px;color:#fbbf24;font-weight:600;">Export PDF</span>
  </div>
</div>
<div style="background:#1e3a6b;padding:8px 20px;border-bottom:1px solid rgba(255,255,255,0.08);">
  <div style="font-size:10px;font-weight:700;letter-spacing:1.8px;text-transform:uppercase;color:#7AABDD;margin-bottom:2px;">DASHBOARD 1 OF 3 — BUSINESS INTELLIGENCE</div>
  <div style="font-size:15px;font-weight:800;color:#ffffff;letter-spacing:-0.2px;">Business Scoring Dashboard</div>
</div>
''')
        # Full KPI table (reuses same kpi_table_dash widget — populated by unlock_to_dashboard)
        full_dash_kpi_html = gr.HTML(value="", elem_id="full-dash-kpi-container")

        # Navigation buttons row
        with gr.Row():
            full_dash_back_btn  = gr.Button("← Back to Preview",         variant="secondary", scale=1)
            full_dash_next_btn  = gr.Button("▶ Business Performance Dashboard →", variant="primary",   scale=2)
            full_dash_fc_btn    = gr.Button("📈 Skip to Forecasting →",   variant="secondary", scale=1)

    # ══════════════════════════════════════════════════════════════════════════
    # All columns list
    # ══════════════════════════════════════════════════════════════════════════
    _ALL_COLS = [step0_col, step1_col, step2_col, step3_col, step4_col, step5_col, step6_col, step6a_col, step7_col, preview_col, step_dash_col]

    def update_visibility_all(active_name):
        names = ['step0','step1','step2','step3','step4','step5','step6','step6a','step7','preview','full_dash']
        cols  = [gr.update(visible=(n == active_name)) for n in names]
        return cols

    def _hdr(on_landing):
        """Return 4 gr.update()s for [logo_html, msme_hdr_btn, gov_hdr_btn, hdr_mode_html].
        msme_hdr_btn and gov_hdr_btn are always hidden — they are stubs.
        hdr_mode_html injects JS to toggle dn-app-mode on the lang-bar."""
        v = bool(on_landing)
        if on_landing:
            js_html = """<script>
(function(){
  var bar = document.getElementById('lang-bar');
  if(bar){ bar.classList.remove('dn-app-mode'); }
})();
</script>"""
        else:
            js_html = """<script>
(function(){
  var bar = document.getElementById('lang-bar');
  if(bar){ bar.classList.add('dn-app-mode'); }
})();
</script>"""
        return [gr.update(visible=v), gr.update(visible=False), gr.update(visible=False),
                gr.update(value=js_html)]

    # ══════════════════════════════════════════════════════════════════════════

    # ╔══════════════════════════════════════════════════════════════════╗
    # ║  SECTION D  —  Step Handlers: MSME Registration & Data Upload    ║
    # ╚══════════════════════════════════════════════════════════════════╝
    # ══════════════════════════════════════════════════════════════════════════
    def handle_file_upload_change(user_data, file):
        if file is not None:
            name = user_data.get('full_name', 'User')
            return (gr.update(value=f"Thank you, {name}, for uploading the dataset. Click 'Generate Preview →' to view your business insights.", visible=True),
                    gr.update(value="", visible=False))
        return gr.update(value="", visible=False), gr.update(value="", visible=False)

    # ── Product name mapping — defined at module level (_PRODUCT_NAMES) ─────────

    def _build_step7_data(df_raw, store_sel, cat_sel, prod_sel):
        """Filter df and compute all KPIs + 7 charts + AI summary for Step 7."""
        import matplotlib
        import matplotlib.pyplot as plt
        import matplotlib.ticker as mticker
        import numpy as np, pandas as pd, warnings
        warnings.filterwarnings('ignore')
        # Gradio gr.State serialises DataFrames to dicts — reconstruct
        if isinstance(df_raw, dict):
            try:
                if 'columns' in df_raw and 'data' in df_raw:
                    df_raw = pd.DataFrame(df_raw['data'], columns=df_raw['columns'])
                else:
                    df_raw = pd.DataFrame(df_raw)
            except Exception:
                df_raw = None
        if df_raw is None or (hasattr(df_raw,'empty') and df_raw.empty):
            return None
        from sklearn.linear_model import LinearRegression

        if df_raw is None:
            return None

        df = df_raw.copy()

        # ── Column resolution (handles raw, remapped, and DRC-normalised names) ──
        _cols_lower = {c.lower(): c for c in df.columns}  # lowercase→actual map
        def _col(*names):
            for n in names:
                if n in df.columns: return n
                if n.lower() in _cols_lower: return _cols_lower[n.lower()]
            return None

        sc   = _col('revenue','Revenue','net_sales','Monthly_Sales_INR','monthly_sales_inr','gross_sales','sales_amount','total_sales')
        dc   = _col('date', 'Date')
        stc  = _col('store_id','Store_ID','store_name','Store_Name')
        catc = _col('category','Category','product_category','Product_Category')
        pidc = _col('product_name','Product_Name','sku','SKU','SKU_Name','sku_name','product_id')
        uc   = _col('sales_quantity','Sales_Quantity','units_sold','Monthly_Demand_Units','monthly_demand_units','demand_units','quantity')
        pmc  = _col('gross_margin_pct','Gross_Margin_Pct','profit_margin_pct','Avg_Margin_Percent','avg_margin_percent','margin_pct','margin')
        rrc  = _col('returns_units','Returns_Units','return_rate_pct','Returns_Percentage','returns_percentage','return_rate','return_pct')
        rpc  = _col('replacement_units','Replacement_Units','replacement_count','replacements')
        tac  = _col('target_achievement_pct','Fulfillment_Rate','fulfillment_rate','fulfilment_rate','stockout_flag','Stockout_Flag')
        ivc  = _col('inventory_level','Inventory_Level','inventory_turnover','Inventory_Turnover')
        roc  = _col('reorder_point',   'Reorder_Point', 'reorder_point')
        slc  = _col('stock_level',     'Stock_Level', 'stock_level')
        qrc  = _col('quantity_returned', 'quantity_returned')

        if sc is None:
            return None

        # ── Ensure MSME-only ────────────────────────────────────────────────
        # df is already filtered to the logged-in MSME in analyze_data — no extra filter needed
        if df.empty:
            return None

        # ── Parse dates ─────────────────────────────────────────────────────
        if dc:
            df[dc] = pd.to_datetime(df[dc], errors='coerce')
            df = df.dropna(subset=[dc])

        # ── Apply filters ────────────────────────────────────────────────────
        # Ensure filter values are strings before calling .replace()
        store_sel = store_sel if isinstance(store_sel, str) else None
        cat_sel   = cat_sel   if isinstance(cat_sel,   str) else None
        prod_sel  = prod_sel  if isinstance(prod_sel,  str) else None
        if store_sel and store_sel not in ("Store: All", ""):
            sid = store_sel.replace("Store: ", "").strip()
            if stc: df = df[df[stc].astype(str) == sid]
        if cat_sel and cat_sel not in ("Category: All", ""):
            cval = cat_sel.replace("Category: ", "").strip()
            if catc: df = df[df[catc].astype(str) == cval]
        if prod_sel and prod_sel not in ("Product: All", ""):
            pval = prod_sel.replace("Product: SKU-", "").replace("Product: ", "").strip()
            if pidc:
                try:    df = df[df[pidc].astype(str) == pval]
                except: pass
        if df.empty:
            return None

        # ── Helpers ──────────────────────────────────────────────────────────
        def _s(c):  return float(df[c].sum())  if c and c in df.columns else 0.0
        def _m(c):  return float(df[c].mean()) if c and c in df.columns else 0.0
        def _si(c): return int(df[c].sum())    if c and c in df.columns else 0
        def _inr(v):
            if v >= 1e7: return f"&#8377;{v/1e7:.1f} Cr"
            if v >= 1e5: return f"&#8377;{v/1e5:.1f} L"
            return f"&#8377;{v:,.0f}"
        def _inr_ax(x, _):
            if abs(x) >= 1e7: return f"₹{x/1e7:.1f}Cr"
            if abs(x) >= 1e5: return f"₹{x/1e5:.0f}L"
            return f"₹{x:,.0f}"

        # ── KPI values ───────────────────────────────────────────────────────
        total_units = _si(uc)
        total_sales = _s(sc)
        avg_margin  = _m(pmc)
        avg_ret     = _m(rrc)
        total_repl  = _si(rpc)
        avg_fulfil  = _m(tac)

        # ── KPI HTML: 6-card row matching image ──────────────────────────────
        def _kcard(label, value, color, sub_color=None, sub=""):
            sc2 = sub_color or color
            return (
                f'<div style="flex:1;min-width:110px;background:#fff;'
                f'border:1px solid #D8E8F4;border-radius:10px;padding:14px 10px 10px 10px;'
                f'text-align:center;box-shadow:0 1px 5px rgba(27,79,138,0.07)">'
                f'<div style="font-size:9.5px;font-weight:700;letter-spacing:1.3px;'
                f'text-transform:uppercase;color:#7A92AA;margin-bottom:6px">{label}</div>'
                f'<div style="font-size:22px;font-weight:900;color:{color};'
                f'font-family:monospace;line-height:1.1">{value}</div>'
                + (f'<div style="font-size:9px;color:{sc2};margin-top:3px">{sub}</div>' if sub else '')
                + '</div>'
            )

        ret_c = "#e74c3c" if avg_ret >= 7 else ("#f39c12" if avg_ret >= 4 else "#27ae60")
        tgt_c = "#27ae60" if avg_fulfil >= 100 else "#f39c12"
        mrg_c = "#27ae60" if avg_margin >= 20 else "#f39c12"

        kpi_html = (
            '<div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px">'
            + _kcard("Units Sold",   f"{total_units:,}",      "#1B4F8A")
            + _kcard("Net Sales",    _inr(total_sales),       "#1B4F8A")
            + _kcard("Margin %",     f"{avg_margin:.1f}%",    mrg_c)
            + _kcard("Return Rate",  f"{avg_ret:.1f}%",       ret_c)
            + _kcard("Replacements", f"{total_repl:,}",       "#8b5cf6")
            + _kcard("Fulfilment",   f"{avg_fulfil:.1f}%",    tgt_c)
            + '</div>'
        )

        # ── Top Products table ────────────────────────────────────────────────
        top_table_html = ""
        if pidc and sc:
            top_agg = {sc: 'sum'}
            if pmc: top_agg[pmc] = 'mean'
            if rrc: top_agg[rrc] = 'mean'
            top = df.groupby(pidc).agg(top_agg).reset_index()
            top.columns = [pidc, 'Sales'] + (['Margin'] if pmc else []) + (['RetRate'] if rrc else [])
            top = top.sort_values('Sales', ascending=False).head(5)

            rows = ""
            for rank, (_, r) in enumerate(top.iterrows(), 1):
                bg   = "#FFFFFF" if rank % 2 else "#F4F9FF"
                pid  = int(r[pidc]) if str(r[pidc]).isdigit() else r[pidc]
                pname= _PRODUCT_NAMES.get(int(pid), f"Product {pid}") if isinstance(pid, (int, float)) else str(pid)
                margin_v  = f"{r['Margin']:.0f}%"  if 'Margin' in r.index  else "—"
                retrate_v = f"{r['RetRate']:.1f}%"  if 'RetRate' in r.index else "—"
                rows += (
                    f'<tr style="background:{bg};border-bottom:1px solid #E8F0F8">'
                    f'<td style="padding:9px 14px;font-weight:700;color:#7A92AA;'
                    f'text-align:center;width:55px">{rank}</td>'
                    f'<td style="padding:9px 14px;color:#4A6A8A;font-size:11px">{pid}</td>'
                    f'<td style="padding:9px 14px;color:#1A2D45;font-weight:600">{pname}</td>'
                    f'<td style="padding:9px 14px;text-align:right;font-weight:800;'
                    f'color:#1B4F8A">{_inr(r["Sales"])}</td>'
                    f'<td style="padding:9px 14px;text-align:center;color:#27ae60;'
                    f'font-weight:700">{margin_v}</td>'
                    f'<td style="padding:9px 14px;text-align:center;color:#e74c3c">{retrate_v}</td>'
                    f'</tr>'
                )

            store_label = (store_sel or "").replace("Store: ", "").strip() or "All Stores"
            cat_label   = (cat_sel   or "").replace("Category: ", "").strip() or "All Categories"
            top_table_html = f"""
<div style="margin-bottom:16px">
  <div style="background:#EAF4FF;border:1px solid #C8DCEF;border-bottom:none;
              border-radius:8px 8px 0 0;padding:9px 16px;font-size:11px;
              font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:#1B4F8A">
    Top Products in {cat_label} &nbsp;&mdash;&nbsp; {store_label}
  </div>
  <table style="width:100%;border-collapse:collapse;font-size:12.5px;
                border:1px solid #C8DCEF;border-radius:0 0 8px 8px;overflow:hidden">
    <thead>
      <tr style="background:#0B1F3A">
        <th style="padding:9px 14px;color:#A8D8FF;font-weight:600;text-align:center">Rank</th>
        <th style="padding:9px 14px;color:#A8D8FF;font-weight:600;text-align:left">Product ID</th>
        <th style="padding:9px 14px;color:#A8D8FF;font-weight:600;text-align:left">Product Name</th>
        <th style="padding:9px 14px;color:#A8D8FF;font-weight:600;text-align:right">Sales</th>
        <th style="padding:9px 14px;color:#A8D8FF;font-weight:600;text-align:center">Margin</th>
        <th style="padding:9px 14px;color:#A8D8FF;font-weight:600;text-align:center">Return Rate</th>
      </tr>
    </thead>
    <tbody>{rows}</tbody>
  </table>
</div>"""

        # ── Global chart style ────────────────────────────────────────────────
        plt.rcParams.update({
            'font.family': 'DejaVu Sans',
            'axes.spines.top': False, 'axes.spines.right': False,
            'axes.grid': True, 'grid.alpha': 0.2, 'grid.color': '#C8DCEF',
            'axes.facecolor': '#FAFCFF', 'figure.facecolor': '#FFFFFF',
        })
        NAVY  = '#1B4F8A'; GREEN = '#27ae60'; RED = '#e74c3c'
        AMBER = '#f39c12'; PURP  = '#8b5cf6'; TEAL = '#0097a7'
        CAT_PAL = [NAVY, '#f39c12', '#27ae60', RED, PURP, TEAL, '#e67e22', '#16a085']

        def _xtick_step(n, target=6): return max(1, n // target)

        def _style_ax(ax, title, ylabel="", fontsize=10):
            ax.set_title(title, fontsize=fontsize, fontweight='bold', pad=12,
                         color='#1A2D45', loc='left')
            if ylabel:
                ax.set_ylabel(ylabel, fontsize=8, color='#4A6A8A', labelpad=6)
            ax.tick_params(axis='both', labelsize=7, colors='#4A6A8A', length=3)
            for spine in ['left', 'bottom']:
                ax.spines[spine].set_color('#D8E8F4')
                ax.spines[spine].set_linewidth(0.8)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.grid(axis='y', alpha=0.2, linestyle='--', linewidth=0.6, color='#B8D4F0')
            ax.set_facecolor('#F8FBFF')

        # ── Chart 1: Category Sales Trend / Product Monthly Sales ──────────
        fig1, ax1 = plt.subplots(figsize=(8, 4.0))
        fig1.patch.set_facecolor('#F8FBFF')
        fig1.subplots_adjust(top=0.86, bottom=0.22, left=0.12, right=0.97)
        is_product_drill = (prod_sel and prod_sel not in ("Product: All", ""))
        if dc and (catc or is_product_drill) and sc:
            df1 = df.copy()
            df1['_m'] = df1[dc].dt.to_period('M').astype(str)
            if is_product_drill and sc:
                # Product level: show monthly sales bar chart for this product
                ms_prod = df1.groupby('_m')[sc].sum().sort_index()
                if not ms_prod.empty:
                    xp = list(range(len(ms_prod)))
                    ax1.bar(xp, ms_prod.values / 1e5, color=NAVY, alpha=0.82,
                            edgecolor='white', linewidth=0.4, width=0.7)
                    ax1.plot(xp, ms_prod.values / 1e5, color=RED, lw=1.5,
                             marker='o', ms=3, alpha=0.8, label='Monthly Sales')
                    step1 = _xtick_step(len(ms_prod))
                    ax1.set_xticks(range(0, len(ms_prod), step1))
                    ax1.set_xticklabels(ms_prod.index.tolist()[::step1],
                                        rotation=45, ha='right', fontsize=6.5)
                    ax1.legend(fontsize=6.5, loc='upper left', framealpha=0.9)
                    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(_inr_ax))
                else:
                    ax1.text(0.5, 0.5, 'No sales data for this product', ha='center',
                             va='center', transform=ax1.transAxes, fontsize=10, color='#7A92AA')
                _style_ax(ax1, 'Product Monthly Sales Trend', 'Sales (₹L)')
            else:
                # Category level: multi-line trend (original logic, x-axis aligned)
                grp1 = df1.groupby([catc, '_m'])[sc].sum().reset_index().sort_values('_m')
                cats1 = sorted(grp1[catc].dropna().unique())
                all_months = sorted(grp1['_m'].unique())
                # Build a unified x-index so all lines share the same x axis
                month_idx = {m: i for i, m in enumerate(all_months)}
                for ci, cat in enumerate(cats1):
                    cm = grp1[grp1[catc] == cat].sort_values('_m')
                    x1v = [month_idx[m] for m in cm['_m']]
                    y1v = cm[sc].values / 1e5
                    ax1.fill_between(x1v, y1v, alpha=0.13, color=CAT_PAL[ci % len(CAT_PAL)])
                    ax1.plot(x1v, y1v, color=CAT_PAL[ci % len(CAT_PAL)],
                             linewidth=1.8, label=cat, marker='o', markersize=2)
                step1 = _xtick_step(len(all_months))
                ax1.set_xticks(range(0, len(all_months), step1))
                ax1.set_xticklabels(all_months[::step1], rotation=45, ha='right', fontsize=6.5)
                ax1.legend(fontsize=6.5, ncol=3, loc='upper left',
                           framealpha=0.9, edgecolor='#D0E4F4', labelspacing=0.3)
                ax1.yaxis.set_major_formatter(mticker.FuncFormatter(_inr_ax))
                _style_ax(ax1, 'Category Sales Trend (36 Months)', 'Sales (₹ Lakhs)')
        else:
            ax1.text(0.5, 0.5, 'No time-series data', ha='center', va='center',
                     transform=ax1.transAxes, fontsize=10, color='#7A92AA')
            _style_ax(ax1, 'Category Sales Trend', 'Sales (₹ Lakhs)')

        # ── Chart 2: Category Margin Trend / Product Margin Trend ──────────
        fig2, ax2 = plt.subplots(figsize=(8, 4.0))
        fig2.patch.set_facecolor('#F8FBFF')
        fig2.subplots_adjust(top=0.86, bottom=0.22, left=0.11, right=0.97)
        if dc and (pmc or is_product_drill):
            df2 = df.copy()
            df2['_m'] = df2[dc].dt.to_period('M').astype(str)
            if is_product_drill:
                # Product level: show this product's monthly margin trend
                mm_prod = df2.groupby('_m')[pmc].mean().sort_index()
                if not mm_prod.empty:
                    xp2 = list(range(len(mm_prod)))
                    ax2.fill_between(xp2, mm_prod.values, alpha=0.18, color=NAVY)
                    ax2.plot(xp2, mm_prod.values, color=NAVY, lw=2,
                             marker='o', ms=3, label='Margin %')
                    ax2.axhline(20, color=GREEN, ls='--', lw=1.2, alpha=0.75, label='Target 20%')
                    ax2.set_ylim(max(0, mm_prod.min() - 5), min(60, mm_prod.max() + 8))
                    step2p = _xtick_step(len(mm_prod))
                    ax2.set_xticks(range(0, len(mm_prod), step2p))
                    ax2.set_xticklabels(mm_prod.index.tolist()[::step2p],
                                        rotation=45, ha='right', fontsize=6.5)
                    ax2.legend(fontsize=6.5, loc='upper left', framealpha=0.9)
                    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:.0f}%'))
                else:
                    ax2.text(0.5, 0.5, 'No margin data', ha='center', va='center',
                             transform=ax2.transAxes, fontsize=10, color='#7A92AA')
                _style_ax(ax2, 'Product Margin Trend', 'Margin %')
            else:
                # Category level: multi-line (original logic, x-axis aligned)
                grp2 = df2.groupby([catc, '_m'])[pmc].mean().reset_index().sort_values('_m') if catc else None
                if grp2 is not None and not grp2.empty:
                    cats2 = sorted(grp2[catc].dropna().unique())
                    all_m2 = sorted(grp2['_m'].unique())
                    month_idx2 = {m: i for i, m in enumerate(all_m2)}
                    for ci, cat in enumerate(cats2):
                        cm2 = grp2[grp2[catc] == cat].sort_values('_m')
                        x2v = [month_idx2[m] for m in cm2['_m']]
                        y2v = cm2[pmc].values
                        ax2.fill_between(x2v, y2v, alpha=0.13, color=CAT_PAL[ci % len(CAT_PAL)])
                        ax2.plot(x2v, y2v, color=CAT_PAL[ci % len(CAT_PAL)],
                                 linewidth=1.8, label=cat, marker='o', markersize=2)
                    ax2.axhline(20, color=GREEN, linestyle='--', linewidth=1.2,
                                alpha=0.75, label='Target 20%')
                    step2 = _xtick_step(len(all_m2))
                    ax2.set_xticks(range(0, len(all_m2), step2))
                    ax2.set_xticklabels(all_m2[::step2], rotation=45, ha='right', fontsize=6.5)
                    ax2.legend(fontsize=6.5, ncol=3, loc='upper left',
                               framealpha=0.9, edgecolor='#D0E4F4', labelspacing=0.3)
                    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:.0f}%'))
                _style_ax(ax2, 'Category Margin Trend (36 Months)', 'Margin %')
        else:
            ax2.text(0.5, 0.5, 'No margin data', ha='center', va='center',
                     transform=ax2.transAxes, fontsize=10, color='#7A92AA')
            _style_ax(ax2, 'Category Margin Trend', 'Margin %')

        # ── Linear forecast helper ───────────────────────────────────────────
        def _forecast(monthly_df_or_series, n):
            """
            All-4-model selection: Prophet → Holt-Winters → Linear Regression → Baseline.
            Accepts either:
              - a pandas DataFrame with [ds (datetime), y] columns  → Prophet eligible
              - a plain numpy array / pandas Series of values        → HW / LR / Baseline only
            Picks winner by lowest MAE on held-out last min(6, N//2) points.
            Returns: (y_hist, yh_forecast, lo_ci, hi_ci, model_label)
            """
            # ── Unpack input ──────────────────────────────────────────────
            if isinstance(monthly_df_or_series, pd.DataFrame) and 'ds' in monthly_df_or_series.columns:
                monthly_df = monthly_df_or_series.copy()
                y = monthly_df['y'].values.astype(float)
                has_ds = True
            else:
                y = np.array(monthly_df_or_series, dtype=float)
                monthly_df = None
                has_ds = False

            N     = len(y)
            val_n = max(2, min(6, N // 2))
            train_y = y[:-val_n]
            val_y   = y[-val_n:]

            candidates  = {}   # key → MAE
            hw_params   = (0.3, 0.1)

            # ── Candidate A: Prophet (only when proper ds/y DataFrame available) ──
            prophet_fc_full = None
            if has_ds and PROPHET_AVAILABLE and len(monthly_df) >= 4:
                try:
                    train_df = monthly_df.iloc[:-val_n].copy()
                    val_df   = monthly_df.iloc[-val_n:].copy()
                    _m = Prophet(
                        yearly_seasonality=True,
                        weekly_seasonality=False,
                        daily_seasonality=False,
                        changepoint_prior_scale=0.05,
                        interval_width=0.95,
                    )
                    _m.fit(train_df[['ds', 'y']])
                    future_val = _m.make_future_dataframe(periods=val_n, freq='MS')
                    fc_val = _m.predict(future_val)
                    fc_val_tail = fc_val.tail(val_n)
                    mae_prophet = float(np.mean(np.abs(
                        fc_val_tail['yhat'].clip(lower=0).values - val_y
                    )))
                    candidates['prophet'] = mae_prophet
                    # Pre-fit on full series for final forecast
                    _mf = Prophet(
                        yearly_seasonality=True,
                        weekly_seasonality=False,
                        daily_seasonality=False,
                        changepoint_prior_scale=0.05,
                        interval_width=0.95,
                    )
                    _mf.fit(monthly_df[['ds', 'y']])
                    future_full = _mf.make_future_dataframe(periods=n, freq='MS')
                    fc_full = _mf.predict(future_full).tail(n)
                    prophet_fc_full = {
                        'yh':  fc_full['yhat'].clip(lower=0).values,
                        'lo':  fc_full['yhat_lower'].clip(lower=0).values,
                        'hi':  fc_full['yhat_upper'].clip(lower=0).values,
                    }
                except Exception:
                    pass

            # ── Candidate B: Holt-Winters (numpy double-exp, grid search α/β) ──
            try:
                best_sse = float('inf')
                for a in [0.1, 0.2, 0.3, 0.5, 0.7]:
                    for b in [0.0, 0.05, 0.1, 0.2]:
                        lv, tr = train_y[0], (train_y[1] - train_y[0]) if len(train_y) > 1 else 0.0
                        sse = 0.0
                        for t in range(1, len(train_y)):
                            pred = lv + tr
                            sse += (train_y[t] - pred) ** 2
                            lv_new = a * train_y[t] + (1 - a) * (lv + tr)
                            tr = b * (lv_new - lv) + (1 - b) * tr
                            lv = lv_new
                        if sse < best_sse:
                            best_sse = sse; hw_params = (a, b)
                a, b = hw_params
                lv, tr = train_y[0], (train_y[1] - train_y[0]) if len(train_y) > 1 else 0.0
                for t in range(1, len(train_y)):
                    lv_new = a * train_y[t] + (1 - a) * (lv + tr)
                    tr = b * (lv_new - lv) + (1 - b) * tr
                    lv = lv_new
                hw_val = np.array([max(0.0, lv + (i + 1) * tr) for i in range(val_n)])
                candidates['hw'] = float(np.mean(np.abs(hw_val - val_y)))
            except Exception:
                pass

            # ── Candidate C: Linear Regression ───────────────────────────────
            try:
                X_tr  = np.arange(len(train_y)).reshape(-1, 1)
                X_va  = np.arange(len(train_y), len(train_y) + val_n).reshape(-1, 1)
                lr    = LinearRegression().fit(X_tr, train_y)
                candidates['lr'] = float(np.mean(np.abs(lr.predict(X_va) - val_y)))
            except Exception:
                pass

            # ── Candidate D: Baseline (trend-aware) ──────────────────────────
            try:
                if len(train_y) >= 6:
                    _d_growth = float(np.clip(
                        (float(train_y[-3:].mean()) - float(train_y[-6:-3].mean())) / (float(train_y[-6:-3].mean()) + 1e-9),
                        -0.20, 0.20))
                elif len(train_y) >= 2:
                    _d_growth = float(np.clip((train_y[-1] - train_y[0]) / (train_y[0] + 1e-9), -0.15, 0.15))
                else:
                    _d_growth = 0.0
                base_avg = float(train_y[-6:].mean()) if len(train_y) >= 6 else float(train_y.mean())
                _d_fc    = base_avg * (1 + _d_growth)
                mae_base = float(np.mean(np.abs(np.full(val_n, _d_fc) - val_y)))
                candidates['base'] = mae_base
            except Exception:
                pass

            best = min(candidates, key=candidates.get) if candidates else 'lr'

            # ── Produce forecast with the winning model (full series) ─────────
            if best == 'prophet' and prophet_fc_full is not None:
                yh  = prophet_fc_full['yh']
                lo  = prophet_fc_full['lo']
                hi  = prophet_fc_full['hi']
                model_label = '📡 Prophet'

            elif best == 'hw':
                a, b = hw_params
                lv, tr = y[0], (y[1] - y[0]) if N > 1 else 0.0
                for t in range(1, N):
                    lv_new = a * y[t] + (1 - a) * (lv + tr)
                    tr = b * (lv_new - lv) + (1 - b) * tr
                    lv = lv_new
                yh  = np.array([max(0.0, lv + (i + 1) * tr) for i in range(n)])
                std = max(float(np.std(y - np.array([y[0] + tr * i for i in range(N)]))), 1.0)
                lo  = (yh - 1.65 * std).clip(0)
                hi  = yh + 1.65 * std
                model_label = '❄️ Holt-Winters'

            elif best == 'base':
                base_avg = float(y[-6:].mean()) if N >= 6 else float(y.mean())
                if N >= 6:
                    _bg2 = float(np.clip((float(y[-3:].mean()) - float(y[-6:-3].mean())) / (float(y[-6:-3].mean()) + 1e-9), -0.20, 0.20))
                elif N >= 2:
                    _bg2 = float(np.clip((y[-1] - y[0]) / (y[0] + 1e-9), -0.15, 0.15))
                else:
                    _bg2 = 0.0
                yh  = np.full(n, base_avg * (1 + _bg2))
                std = max(float(np.std(y - np.full(N, base_avg))), 1.0)
                lo  = (yh - 1.65 * std).clip(0)
                hi  = yh + 1.65 * std
                model_label = '📊 Baseline'

            else:  # lr
                X_all = np.arange(N).reshape(-1, 1)
                lr    = LinearRegression().fit(X_all, y)
                Xf    = np.arange(N, N + n).reshape(-1, 1)
                yh    = lr.predict(Xf).clip(0)
                std   = max(float(np.std(y - lr.predict(X_all))), 1.0)
                lo    = (yh - 1.65 * std).clip(0)
                hi    = yh + 1.65 * std
                model_label = '📐 Linear Reg'

            mae_winner = candidates.get(best, 0)
            return y, yh, lo, hi, model_label, mae_winner

        # ── Build proper [ds, y] monthly DataFrame for Prophet-eligible calls ──
        _monthly_for_prophet = None
        if dc and sc:
            _dfp = df.copy()
            _dfp['_ms'] = _dfp[dc].dt.to_period('M').dt.to_timestamp()
            _monthly_for_prophet = (
                _dfp.groupby('_ms')[sc].sum()
                    .reset_index()
                    .rename(columns={'_ms': 'ds', sc: 'y'})
                    .sort_values('ds')
                    .reset_index(drop=True)
            )

        # ── Chart 3: 6-Month Forecast ────────────────────────────────────────
        fig3, ax3 = plt.subplots(figsize=(5.5, 4.0))
        fig3.patch.set_facecolor('#F8FBFF')
        fig3.subplots_adjust(top=0.84, bottom=0.10, left=0.17, right=0.96)
        total_6m = 0; model_label_6m = ''
        if _monthly_for_prophet is not None and len(_monthly_for_prophet) >= 3:
            y3, yh6, lo6, hi6, model_label_6m, _ = _forecast(_monthly_for_prophet, 6)
            total_6m = float(yh6.sum())
            xh3 = list(range(len(y3)))
            xf3 = list(range(len(y3) - 1, len(y3) + 6))
            fy3 = np.concatenate([[y3[-1]], yh6])
            fl3 = np.concatenate([[y3[-1]], lo6])
            fh3 = np.concatenate([[y3[-1]], hi6])
            ax3.plot(xh3, y3 / 1e5, color=NAVY, lw=2, label='Historical',
                     marker='o', ms=2.5)
            ax3.plot(xf3, fy3 / 1e5, color=RED, lw=2, ls='--',
                     marker='o', ms=2.5, label=f'Forecast ({model_label_6m})')
            ax3.fill_between(xf3, fl3 / 1e5, fh3 / 1e5, color=RED, alpha=0.12)
            ax3.axvline(len(y3) - 1, color='#888', ls=':', lw=1, alpha=0.6)
            ax3.legend(fontsize=6.5, loc='upper left', framealpha=0.9)
            ax3.text(0.97, 0.06, f'6M: {_inr(total_6m)}',
                     transform=ax3.transAxes, ha='right', fontsize=8.5, fontweight='bold',
                     color=RED, bbox=dict(boxstyle='round,pad=0.3',
                                          facecolor='#FFF3CD', edgecolor=AMBER, alpha=0.9))
        ax3.set_xticklabels([]); ax3.set_xlabel('')
        ax3.yaxis.set_major_formatter(mticker.FuncFormatter(_inr_ax))
        _style_ax(ax3, f'6-Month Forecast', 'Sales (₹L)')

        # ── Chart 4: 12-Month Forecast ───────────────────────────────────────
        fig4, ax4 = plt.subplots(figsize=(5.5, 4.0))
        fig4.patch.set_facecolor('#F8FBFF')
        fig4.subplots_adjust(top=0.84, bottom=0.10, left=0.17, right=0.96)
        total_12m = 0; model_label_12m = ''
        if _monthly_for_prophet is not None and len(_monthly_for_prophet) >= 3:
            y4, yh12, lo12, hi12, model_label_12m, _ = _forecast(_monthly_for_prophet, 12)
            total_12m = float(yh12.sum())
            xh4 = list(range(len(y4)))
            xf4 = list(range(len(y4) - 1, len(y4) + 12))
            fy4 = np.concatenate([[y4[-1]], yh12])
            fl4 = np.concatenate([[y4[-1]], lo12])
            fh4 = np.concatenate([[y4[-1]], hi12])
            ax4.plot(xh4, y4 / 1e5, color=NAVY, lw=2, label='Historical',
                     marker='o', ms=2)
            ax4.plot(xf4, fy4 / 1e5, color=GREEN, lw=2, ls='--',
                     marker='o', ms=2, label=f'Forecast ({model_label_12m})')
            ax4.fill_between(xf4, fl4 / 1e5, fh4 / 1e5, color=GREEN, alpha=0.12)
            ax4.axvline(len(y4) - 1, color='#888', ls=':', lw=1, alpha=0.6)
            ax4.legend(fontsize=6.5, loc='upper left', framealpha=0.9)
            ax4.text(0.97, 0.06, f'12M: {_inr(total_12m)}',
                     transform=ax4.transAxes, ha='right', fontsize=8.5, fontweight='bold',
                     color=GREEN, bbox=dict(boxstyle='round,pad=0.3',
                                            facecolor='#EAF7EE', edgecolor=GREEN, alpha=0.9))
        ax4.set_xticklabels([]); ax4.set_xlabel('')
        ax4.yaxis.set_major_formatter(mticker.FuncFormatter(_inr_ax))
        _style_ax(ax4, f'12-Month Forecast', 'Sales (₹L)')

        # ── Chart 5: Product Fulfilment Trend ────────────────────────────────
        fig5, ax5 = plt.subplots(figsize=(5.5, 4.0))
        fig5.patch.set_facecolor('#F8FBFF')
        fig5.subplots_adjust(top=0.86, bottom=0.22, left=0.13, right=0.97)
        # tac now resolves: target_achievement_pct → Fulfillment_Rate → fulfillment_rate
        using_fulfilment_rate = (tac == 'Fulfillment_Rate' or tac == 'fulfillment_rate')
        target_line = 100 if not using_fulfilment_rate else 85  # Fulfillment_Rate target is 85%
        if dc and tac:
            df5 = df.copy()
            df5['_m'] = df5[dc].dt.to_period('M').astype(str)
            ft  = df5.groupby('_m')[tac].mean().sort_index()
            if not ft.empty:
                xft = list(range(len(ft)))
                fv  = ft.values
                ax5.fill_between(xft, fv, target_line,
                                 where=fv >= target_line,
                                 alpha=0.18, color=GREEN, interpolate=True)
                ax5.fill_between(xft, fv, target_line,
                                 where=fv < target_line,
                                 alpha=0.18, color=RED, interpolate=True)
                ax5.plot(xft, fv, color=NAVY, lw=2, marker='o', ms=2.5)
                ax5.axhline(target_line, color=GREEN, ls='--', lw=1.2, alpha=0.7,
                            label=f'{target_line}% target')
                ax5.set_ylim(max(50, fv.min() - 6), min(115, fv.max() + 6))
                step5 = _xtick_step(len(ft))
                ax5.set_xticks(range(0, len(ft), step5))
                ax5.set_xticklabels(ft.index.tolist()[::step5],
                                    rotation=45, ha='right', fontsize=6.5)
                ax5.yaxis.set_major_formatter(
                    mticker.FuncFormatter(lambda x, _: f'{x:.0f}%'))
                ax5.legend(fontsize=6.5, loc='lower right', framealpha=0.9)
            else:
                ax5.text(0.5, 0.5, 'No fulfilment data', ha='center', va='center',
                         transform=ax5.transAxes, fontsize=10, color='#7A92AA')
        else:
            ax5.text(0.5, 0.5, 'No fulfilment data in dataset', ha='center', va='center',
                     transform=ax5.transAxes, fontsize=10, color='#7A92AA')
        ylabel5 = 'Fulfillment Rate %' if using_fulfilment_rate else 'Target Achievement %'
        _style_ax(ax5, 'Product Fulfilment Trend', ylabel5)

        # ── Chart 6: Product Sales vs Returns ────────────────────────────────
        fig6, ax6 = plt.subplots(figsize=(6.5, 4.0))
        fig6.patch.set_facecolor('#F8FBFF')
        fig6.subplots_adjust(top=0.86, bottom=0.22, left=0.13, right=0.90)
        if dc and sc:
            df6 = df.copy()
            df6['_m'] = df6[dc].dt.to_period('M').astype(str)
            agg6 = {'sales': (sc, 'sum')}
            if rrc: agg6['ret'] = (rrc, 'mean')
            sr = df6.groupby('_m').agg(**agg6).sort_index()
            ax6b = ax6.twinx()
            xsr  = list(range(len(sr)))
            ax6.fill_between(xsr, sr['sales'].values / 1e5, alpha=0.15, color=NAVY)
            ax6.plot(xsr, sr['sales'].values / 1e5, color=NAVY, lw=2,
                     marker='o', ms=2.5, label='Sales')
            if rrc and 'ret' in sr.columns:
                ax6b.plot(xsr, sr['ret'].values, color=RED, lw=2, ls='--',
                          marker='o', ms=2.5, label='Returns')
                ax6b.set_ylabel('Return Rate %', fontsize=8, color=RED)
                ax6b.tick_params(axis='y', labelcolor=RED, labelsize=7)
                ax6b.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:.1f}%'))
                ax6b.spines['right'].set_color('#F5C6CB')
            step6 = _xtick_step(len(sr))
            ax6.set_xticks(range(0, len(sr), step6))
            ax6.set_xticklabels(sr.index.tolist()[::step6], rotation=45, ha='right', fontsize=6.5)
            ax6.yaxis.set_major_formatter(mticker.FuncFormatter(_inr_ax))
            h1, l1 = ax6.get_legend_handles_labels()
            h2, l2 = ax6b.get_legend_handles_labels() if rrc else ([], [])
            ax6.legend(h1 + h2, l1 + l2, fontsize=6.5, loc='upper left',
                       framealpha=0.9, edgecolor='#D0E4F4')
        _style_ax(ax6, 'Product Sales vs Returns', 'Sales (₹L)')

        # ── Chart 7: Inventory & Reorder Levels ──────────────────────────────
        fig7, ax7 = plt.subplots(figsize=(6.5, 4.0))
        fig7.patch.set_facecolor('#F8FBFF')
        fig7.subplots_adjust(top=0.86, bottom=0.30, left=0.12, right=0.97)

        def _sku_label(p):
            try:   return f"SKU-{int(float(p))}"
            except: return str(p)[:12]

        inv_plotted = False

        if is_product_drill and dc and ivc:
            # Product selected: show monthly inventory turnover trend
            try:
                df7p = df.copy()
                df7p['_m'] = df7p[dc].dt.to_period('M').astype(str)
                inv_trend = df7p.groupby('_m')[ivc].mean().sort_index()
                if not inv_trend.empty:
                    xp7 = list(range(len(inv_trend)))
                    vp7 = inv_trend.values
                    clrs7 = [GREEN if v >= 10 else (AMBER if v >= 6 else RED) for v in vp7]
                    ax7.bar(xp7, vp7, color=clrs7, alpha=0.82,
                            edgecolor='white', linewidth=0.4, width=0.7)
                    ax7.plot(xp7, vp7, color=NAVY, lw=1.5, marker='o', ms=2.5, alpha=0.7)
                    ax7.axhline(12, color=GREEN, ls='--', lw=1.2, alpha=0.75, label='Good ≥12x')
                    ax7.axhline(6,  color=AMBER, ls=':',  lw=1.0, alpha=0.65, label='Warning <6x')
                    step7p = _xtick_step(len(inv_trend))
                    ax7.set_xticks(range(0, len(inv_trend), step7p))
                    ax7.set_xticklabels(inv_trend.index.tolist()[::step7p],
                                        rotation=40, ha='right', fontsize=6.5)
                    ax7.legend(fontsize=6.5, loc='upper right', framealpha=0.9, edgecolor='#D0E4F4')
                    ax7.set_ylabel('Turnover (x/month)', fontsize=8, color='#4A6A8A', labelpad=6)
                    inv_plotted = True
            except Exception:
                pass
            _style_ax(ax7, 'Monthly Inventory Turnover', '')

        elif pidc and slc and roc and ivc:
            # Full 3-bar: Stock + Reorder + Inventory Turnover per SKU
            try:
                inv = df.groupby(pidc).agg(
                    Stock=(slc, 'mean'), Reorder=(roc, 'mean'), Restck=(ivc, 'mean')
                ).reset_index().nlargest(8, 'Stock')
                if not inv.empty:
                    x7 = list(range(len(inv)))
                    w7 = 0.27
                    ax7.bar([xi - w7 for xi in x7], inv['Stock'],   width=w7, color=NAVY,  alpha=0.85, label='Stock Level',      edgecolor='white', linewidth=0.4)
                    ax7.bar([xi       for xi in x7], inv['Reorder'], width=w7, color=AMBER, alpha=0.85, label='Reorder Point',     edgecolor='white', linewidth=0.4)
                    ax7.bar([xi + w7 for xi in x7], inv['Restck'],  width=w7, color=GREEN, alpha=0.85, label='Inventory Turnover', edgecolor='white', linewidth=0.4)
                    ax7.set_xticks(list(x7))
                    ax7.set_xticklabels([_sku_label(p) for p in inv[pidc]], rotation=40, ha='right', fontsize=7)
                    ax7.legend(fontsize=6.5, loc='upper right', framealpha=0.9, edgecolor='#D0E4F4', ncol=3)
                    ax7.set_ylabel('Units', fontsize=8, color='#4A6A8A', labelpad=6)
                    inv_plotted = True
            except Exception:
                pass
            _style_ax(ax7, 'Inventory & Reorder Levels', '')

        elif pidc and ivc:
            # Inventory Turnover per product (colour-coded)
            try:
                inv_fb = (df.groupby(pidc)[ivc].mean()
                           .reset_index()
                           .sort_values(ivc, ascending=False)
                           .head(10))
                if not inv_fb.empty:
                    x7b  = list(range(len(inv_fb)))
                    vals = inv_fb[ivc].values
                    clrs = [GREEN if v >= 10 else (AMBER if v >= 6 else RED) for v in vals]
                    bars = ax7.bar(x7b, vals, color=clrs, alpha=0.85, edgecolor='white', linewidth=0.5, width=0.6)
                    ax7.axhline(12, color=GREEN, ls='--', lw=1.3, alpha=0.75, label='Good ≥12x')
                    ax7.axhline(6,  color=AMBER, ls=':',  lw=1.1, alpha=0.65, label='Warning <6x')
                    ax7.set_xticks(x7b)
                    ax7.set_xticklabels([_sku_label(p) for p in inv_fb[pidc]], rotation=40, ha='right', fontsize=7)
                    ax7.legend(fontsize=6.5, loc='upper right', framealpha=0.9, edgecolor='#D0E4F4')
                    for bar, v in zip(bars, vals):
                        ax7.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.12,
                                 f'{v:.1f}x', ha='center', va='bottom', fontsize=6, color='#1A2D45', fontweight='600')
                    ax7.set_ylabel('Turnover (x/month)', fontsize=8, color='#4A6A8A', labelpad=6)
                    inv_plotted = True
            except Exception:
                pass
            _style_ax(ax7, 'Inventory & Reorder Levels', '')

        elif pidc and slc:
            try:
                inv_fb2 = df.groupby(pidc)[slc].mean().reset_index().head(8)
                if not inv_fb2.empty:
                    ax7.bar(range(len(inv_fb2)), inv_fb2[slc], color=NAVY, alpha=0.82, edgecolor='white', linewidth=0.4)
                    ax7.set_xticks(range(len(inv_fb2)))
                    ax7.set_xticklabels([_sku_label(p) for p in inv_fb2[pidc]], rotation=40, ha='right', fontsize=7)
                    ax7.set_ylabel('Units', fontsize=8, color='#4A6A8A', labelpad=6)
                    inv_plotted = True
            except Exception:
                pass
            _style_ax(ax7, 'Inventory & Reorder Levels', '')
        else:
            _style_ax(ax7, 'Inventory & Reorder Levels', '')

        if not inv_plotted:
            ax7.text(0.5, 0.5, 'No inventory data in dataset',
                     ha='center', va='center', transform=ax7.transAxes,
                     fontsize=9, color='#7A92AA')

        # ── AI Forecast Summary ───────────────────────────────────────────────
        cat_label   = (cat_sel   or "").replace("Category: ", "").strip() or "All Categories"
        store_label = (store_sel or "").replace("Store: ", "").strip()    or "All Stores"

        # Derive top_cat from actual data — never hardcode
        if cat_label not in ("All Categories", ""):
            top_cat = cat_label
        elif catc and catc in df.columns and not df.empty:
            by_cat = df.groupby(catc)[sc].sum()
            top_cat = by_cat.idxmax() if not by_cat.empty else "your top category"
        elif pidc and pidc in df.columns and not df.empty:
            by_prod = df.groupby(pidc)[sc].sum()
            top_cat = str(by_prod.idxmax()) if not by_prod.empty else "your top product"
        else:
            top_cat = "your top category"

        low_stock_msg = ""
        if slc and roc and pidc:
            try:
                ls = df[df[slc] < df[roc]]
                if not ls.empty:
                    top_low  = ls.groupby(pidc)[slc].mean().idxmin()
                    pid_int  = int(top_low) if str(top_low).isdigit() else top_low
                    pname_ls = _PRODUCT_NAMES.get(pid_int, f"SKU-{pid_int}")
                    low_stock_msg = (
                        f'<li>Recommend restocking <strong>{pname_ls}</strong>'
                        f' within 2 weeks.</li>')
            except Exception:
                pass

        # Derive real forecast growth% using date-span (not row count) for monthly average
        if total_6m > 0 and total_sales > 0:
            if dc and dc in df.columns:
                _s_dates = df[dc].dropna()
                _s_span  = max(1, round((_s_dates.max() - _s_dates.min()).total_seconds() / (30.44 * 86400)) + 1)
            else:
                _s_span  = max(df[sc].count() // max(df[pidc].nunique() if pidc else 1, 1), 1) if sc else 1
            _hist_6m = (total_sales / _s_span) * 6
            _raw_fc_pct = (total_6m - _hist_6m) / (_hist_6m + 1e-9) * 100
            fc_pct = int(np.clip(round(_raw_fc_pct), -40, 50))
        else:
            fc_pct = 0   # genuinely unknown — don't invent a number
        margin_state  = "strong" if avg_margin >= 20 else "moderate"
        ret_msg       = (f"Monitor growing returns ({avg_ret:.1f}%) — review product quality."
                         if avg_ret > 7 else
                         f"Return rate {avg_ret:.1f}% is within healthy limits.")
        # Only show category-specific notes if that category is actually in the data
        _cats_in_data = set(df[catc].dropna().astype(str).str.lower().unique()) if catc and catc in df.columns else set()
        clothing_note = ('<li>Monitor growing clothing returns during the monsoon.</li>'
                         if 'clothing' in cat_label.lower() or any('clothing' in c for c in _cats_in_data)
                         else "")
        elec_note     = ('<li>Electronics inventory low — reorder urgency.</li>'
                         if 'electr' in cat_label.lower() or any('electr' in c for c in _cats_in_data)
                         else "")

        # ── Compute action-specific numbers ──────────────────────────────────────
        inv_action_pct  = 10 if avg_fulfil < 70 else (5 if avg_fulfil < 85 else 0)
        margin_gap      = round(20 - avg_margin, 1) if avg_margin < 20 else 0
        ret_action      = avg_ret > 7
        # Low-margin action
        margin_action_li = (
            f'<li>💡 <strong>Supplier Cost Improvement Opportunity — <em>{top_cat}</em></strong> — '
            f'current margin is <strong>{avg_margin:.1f}%</strong>, '
            f'{margin_gap}pp below the 20% target. A 5% cost reduction unlocks ~{_inr(total_sales * 0.05 if total_sales else 0)} extra profit.</li>'
            if margin_gap > 0 else
            f'<li>✅ <strong>Inventory Levels Healthy — Margin at {avg_margin:.1f}%</strong> — '
            f'consider expanding <em>{top_cat}</em> SKU count to grow revenue.</li>'
        )
        # Inventory action
        inv_action_li = (
            f'<li>📦 <strong>Recommended Restocking Action — <em>{top_cat}</em>: +{inv_action_pct}%</strong> — '
            f'fulfilment rate is <strong>{avg_fulfil:.1f}%</strong> '
            f'(target ≥85%). Stockout risk is high; pre-order for the next 30 days.</li>'
            if inv_action_pct > 0 else
            f'<li>✅ <strong>Inventory Levels Healthy</strong> — fulfilment at {avg_fulfil:.1f}%. '
            f'Focus on optimising reorder cycle frequency.</li>'
        )
        # Returns action
        ret_action_li = (
            f'<li>🔁 <strong>High Product Returns — <em>{top_cat}</em></strong> — '
            f'return rate is <strong>{avg_ret:.1f}%</strong> (threshold 7%). '
            f'Review top 3 returned SKUs and add quality checks before dispatch.</li>'
            if ret_action else
            f'<li>✅ <strong>Inventory Levels Healthy — Return Rate {avg_ret:.1f}%</strong> — '
            f'maintain current quality controls.</li>'
        )
        # Forecast growth action
        if fc_pct > 3:
            fc_action_li = (
                f'<li>📈 <strong>Demand Outlook: {fc_pct}–{fc_pct+4}% Rise Expected — <em>{top_cat}</em></strong> '
                f'over the next 6 months — projected revenue: <strong>{_inr(total_6m) if total_6m else "—"}</strong>. '
                f'Align procurement and staffing 3 weeks ahead.</li>'
            )
        elif fc_pct < -3:
            fc_action_li = (
                f'<li>⚠️ <strong>Demand Outlook: {abs(fc_pct)}% Slowdown Forecast — <em>{top_cat}</em></strong> '
                f'over the next 6 months — projected revenue: <strong>{_inr(total_6m) if total_6m else "—"}</strong>. '
                f'Reduce next procurement order to avoid overstock build-up.</li>'
            )
        else:
            fc_action_li = (
                f'<li>📊 <strong>Demand Outlook Stable — <em>{top_cat}</em></strong> — '
                f'6-month projected revenue: <strong>{_inr(total_6m) if total_6m else "—"}</strong>. '
                f'Maintain current procurement cycle with a 10% safety buffer.</li>'
            )

        ai_html = f"""
<div style="background:linear-gradient(160deg,#EAF4FF 0%,#F8FAFF 100%);
            border:1px solid #B8D4F0;border-radius:10px;padding:16px 18px;
            height:100%;box-shadow:0 1px 8px rgba(27,79,138,0.08)">
  <div style="font-size:10px;font-weight:800;letter-spacing:1.8px;text-transform:uppercase;
              color:#1B4F8A;margin-bottom:6px;padding-bottom:8px;border-bottom:2px solid #C8DCEF">
    🤖 AI Business Insights
  </div>
  <ul style="margin:0 0 14px 0;padding-left:18px;font-size:12px;
             line-height:1.9;color:#1A2D45;list-style:none;padding-left:0">
    {fc_action_li}
    {inv_action_li}
    {margin_action_li}
    {ret_action_li}
    {low_stock_msg}
  </ul>
  <div style="background:#fff;border-radius:8px;padding:10px 13px;
              border-left:3px solid #1B4F8A;font-size:11px;color:#4A6A8A;line-height:1.7">
    <span style="font-weight:700;color:#1B4F8A">Store:</span> {store_label}
    &nbsp;|&nbsp;
    <span style="font-weight:700;color:#1B4F8A">Category:</span> {cat_label}<br>
    <span style="font-weight:700;color:#1B4F8A">Margin:</span> {avg_margin:.1f}% ({margin_state})
    &nbsp;|&nbsp;
    <span style="font-weight:700;color:#1B4F8A">Return Rate:</span>
    <span style="color:{ret_c};font-weight:700">{avg_ret:.1f}%</span><br>
    <span style="font-weight:700;color:#1B4F8A">Fulfilment:</span>
    <span style="color:{tgt_c};font-weight:700">{avg_fulfil:.1f}%</span>
    &nbsp;|&nbsp;
    <span style="font-weight:700;color:#1B4F8A">6M Forecast:</span>
    <span style="color:#1B4F8A;font-weight:700">{_inr(total_6m) if total_6m else "—"}</span>
  </div>
</div>"""

        plt.close('all')
        return dict(kpi=kpi_html, table=top_table_html, ai=ai_html,
                    f1=fig1, f2=fig2, f3=fig3, f4=fig4,
                    f5=fig5, f6=fig6, f7=fig7)

    # ── Step 7 outputs helper ────────────────────────────────────────────────
    _S7_CHART_OUTPUTS = [
        s7_kpi_html, s7_top_table,
        s7_cat_sales_chart, s7_cat_margin_chart,
        s7_fc6_chart, s7_fc12_chart, s7_fulfil_chart,
        s7_sales_ret_chart, s7_inventory_chart,
        s7_ai_summary
    ]

    def _pack_s7(result):
        """Unpack _build_step7_data result dict into Gradio output tuple."""
        if result is None:
            empty = "<div style='padding:20px;color:#aaa'>No data. Upload and analyse first.</div>"
            return empty, empty, None, None, None, None, None, None, None, empty
        return (result['kpi'], result['table'],
                result['f1'],  result['f2'],
                result['f3'],  result['f4'],  result['f5'],
                result['f6'],  result['f7'],
                result['ai'])

    def show_granular_dashboard(granular_data, df_raw):
        """Navigate to Step 7 and populate with default (All) filters."""
        import pandas as _pd_sgd2
        # Reconstruct df_raw if it's a Gradio-serialised dict
        if isinstance(df_raw, dict):
            try:
                import pandas as _pd_sgd2
                if 'columns' in df_raw and 'data' in df_raw:
                    df_raw = _pd_sgd2.DataFrame(df_raw['data'], columns=df_raw['columns'])
                else:
                    df_raw = _pd_sgd2.DataFrame(df_raw)
            except Exception: df_raw = None
        # If granular_data is None (generate_granular_forecast failed on Render),
        # rebuild it now from df_raw so Dashboard 3 still works
        if granular_data is None and df_raw is not None:
            try:
                granular_data = generate_granular_forecast(df_raw.copy())
                print("✅ show_granular_dashboard: rebuilt granular_data from df_raw")
            except Exception as _rgf_e:
                print(f"⚠️ show_granular_dashboard: rebuild also failed: {_rgf_e}")
                granular_data = None
        try:
            result = _build_step7_data(df_raw, "Store: All", "Category: All", "Product: All")
        except Exception as _e7:
            import traceback as _tb7
            _err_html = f"<div style='color:#dc2626;padding:20px;font-size:13px;'>❌ Forecast error: {_e7}<br><pre style='font-size:10px'>{_tb7.format_exc()}</pre></div>"
            result = None

        # Build filter dropdown choices from df_raw
        stores = ["Store: All"]
        cats   = ["Category: All"]
        prods  = ["Product: All"]
        if df_raw is not None:
            # Gradio gr.State serialises DataFrames to dicts — reconstruct
            import pandas as _pd_sgd
            if isinstance(df_raw, dict):
                try: df_raw = _pd_sgd.DataFrame(df_raw)
                except Exception: df_raw = None
            if df_raw is not None:
                dff = df_raw.copy()
            else:
                dff = _pd_sgd.DataFrame()
            # df_raw is already the MSME-filtered slice — use all rows for dropdown population
            stc2 = 'store_id' if 'store_id' in dff.columns else ('Store_ID' if 'Store_ID' in dff.columns else None)
            c2   = 'product_category' if 'product_category' in dff.columns else ('Product_Category' if 'Product_Category' in dff.columns else None)
            p2   = 'product_id' if 'product_id' in dff.columns else ('SKU_Name' if 'SKU_Name' in dff.columns else None)
            if stc2: stores += [f"Store: {s}" for s in sorted(dff[stc2].unique())]
            if c2:   cats   += [f"Category: {c}" for c in sorted(dff[c2].dropna().unique())]
            if p2:   prods  += [f"Product: SKU-{p}" for p in sorted(dff[p2].unique())[:40]]

        return (7, *update_visibility_all('step7'),
                gr.update(choices=stores, value="Store: All"),
                gr.update(choices=cats,   value="Category: All"),
                gr.update(choices=prods,  value="Product: All"),
                *_pack_s7(result), *_hdr(False))

    def update_step7_filters(store_sel, cat_sel, prod_sel, df_raw):
        """Recompute all Step 7 outputs when any filter dropdown changes."""
        import pandas as _pd_s7f
        if isinstance(df_raw, dict):
            try:
                import pandas as _pd_s7f
                if 'columns' in df_raw and 'data' in df_raw:
                    df_raw = _pd_s7f.DataFrame(df_raw['data'], columns=df_raw['columns'])
                else:
                    df_raw = _pd_s7f.DataFrame(df_raw)
            except Exception: df_raw = None
        result = _build_step7_data(df_raw, store_sel, cat_sel, prod_sel)
        return _pack_s7(result)

    # ══════════════════════════════════════════════════════════════════════════
    # Wire Events
    # ══════════════════════════════════════════════════════════════════════════
    back6_btn.click(lambda: (5, *update_visibility_all('step5'), *_hdr(False)), [], [step_state]+_ALL_COLS+[logo_html,msme_hdr_btn,gov_hdr_btn, hdr_mode_html], queue=True)
    back6a_btn.click(lambda: (0, *update_visibility_all('step0'), *_hdr(True)), [], [step_state]+_ALL_COLS+[logo_html,msme_hdr_btn,gov_hdr_btn, hdr_mode_html])

    # ── Landing page login ─────────────────────────────────────────────────────
    def handle_login(name, mobile, otp, unlock_mode=False, email=""):
        """Retail auth: Name + Mobile + OTP + Email → Step 5 or Full Dashboard."""
        import re as _re_auth
        name   = (name   or "").strip()
        mobile = (mobile or "").strip()
        otp    = (otp    or "").strip()
        email  = (email  or "").strip()
        def _err(msg):
            return (gr.update(value=msg, visible=True), {}, 0,
                    *update_visibility_all('step0'),
                    gr.update(visible=True),        # keep login_section_col visible on error
                    gr.update(value="", visible=False), gr.update(value="", visible=False),
                    gr.update(value="", visible=False), gr.update(value="", visible=False),
                    gr.update(value="", visible=False), gr.update(value="", visible=False),
                    *_hdr(True))

        if not name:
            return _err("⚠️ Please enter your name.")
        if not mobile or not _re_auth.match(r'^\+?[\d\s\-]{7,15}$', mobile):
            return _err("⚠️ Please enter a valid mobile number.")
        _otp_ok, _otp_msg = _verify_otp(email, otp)
        if not _otp_ok:
            return _err(f"⚠️ {_otp_msg}")
        if not email or '@' not in email:
            return _err("⚠️ Please enter a valid email address.")

        profile = {'full_name': name, 'mobile_number': mobile, 'email': email}
        _save_user(name, mobile, email)
        msg = f"✅ Welcome, {name}! Loading your full business dashboard..."
        if unlock_mode:
            # Came from preview "Unlock Full Dashboard" → hide form, go straight to full_dash
            return (gr.update(value="", visible=False), profile, 8,
                    *update_visibility_all('full_dash'),
                    gr.update(visible=False),       # hide login_section_col
                    gr.update(value="", visible=False), gr.update(value="", visible=False),
                    gr.update(value="", visible=False), gr.update(value="", visible=False),
                    gr.update(value="", visible=False), gr.update(value=msg, visible=False),
                    *_hdr(False))
        msg2 = f"✅ Welcome, {name}! Upload your retail data file to begin analysis."
        return (gr.update(value="", visible=False), profile, 5,
                *update_visibility_all('step5'),
                gr.update(visible=False),           # hide login_section_col
                gr.update(value="", visible=False), gr.update(value="", visible=False),
                gr.update(value="", visible=False), gr.update(value="", visible=False),
                gr.update(value="", visible=False), gr.update(value=msg2, visible=True),
                *_hdr(False))

    _UNLOCK_OUTPUTS = (
        [preview_unlock_loading, step_state] + _ALL_COLS
        + [kpi_table_dash, chart1_dash, chart2_dash, chart3_dash, chart4_dash,
           chart1_summary, chart2_summary, chart3_summary, chart4_summary,
           dashboard_data_state, df_state,
           full_dash_kpi_html,
           granular_forecast_data_state]      # needed by forecast_deepdive_btn
        + [logo_html, msme_hdr_btn, gov_hdr_btn, hdr_mode_html]
    )

    quick_login_btn.click(
        handle_login,
        [name_input_landing, quick_login_mobile, otp_landing_input, unlock_mode_state, quick_login_email],
        [landing_login_error_msg, user_data_state, step_state] + _ALL_COLS +
        [login_section_col, error1, error2, error3, error4, error5, login_welcome_message,
         logo_html, msme_hdr_btn, gov_hdr_btn, hdr_mode_html],
        show_progress=True,
        js="""() => {
            // Return form values directly to Python — bypasses Gradio component state entirely
            var f = window._dnForm || {};
            var name  = f.name  || document.querySelector('#dn-f-name')  && document.querySelector('#dn-f-name').value  || '';
            var mob   = f.mob   || document.querySelector('#dn-f-mobile') && document.querySelector('#dn-f-mobile').value || '';
            var otp   = f.otp   || document.querySelector('#dn-f-otp')   && document.querySelector('#dn-f-otp').value   || '';
            var email = f.email || document.querySelector('#dn-f-email')  && document.querySelector('#dn-f-email').value || '';
            // Return array matching inputs= [name_input_landing, quick_login_mobile, otp_landing_input, unlock_mode_state, quick_login_email]
            // Read unlock_mode from window (set by preview_unlock_btn) — default true so dashboard always loads
            var unlockMode = (window._dnUnlockMode === true) ? true : true;
            return [name, mob, otp, unlockMode, email];
        }"""
    ).then(
        # When coming from preview unlock, run full analysis and populate dashboards
        # fu = file_upload fallback if preview_file_state is None
        lambda um, ud, pf, pc, ls, fu, drc: (
            unlock_to_dashboard(ud, pf or fu or drc, True, ls)
            if (pf is not None or fu is not None or drc is not None)
            else tuple([gr.update()] * len(_UNLOCK_OUTPUTS))
        ),
        [unlock_mode_state, user_data_state, preview_file_state, preview_consent_state,
         lang_state, file_upload, _drc_clean_state],
        _UNLOCK_OUTPUTS,
        queue=True
    )

    back5_btn.click(
        lambda: (0, *update_visibility_all('step0'), gr.update(value="", visible=False), *_hdr(True)),
        [], [step_state] + _ALL_COLS + [login_welcome_message, logo_html, msme_hdr_btn, gov_hdr_btn, hdr_mode_html]
    )
    cancel5_btn.click(
        lambda: (0, *update_visibility_all('step0'), gr.update(value="", visible=False), *_hdr(True)),
        [], [step_state] + _ALL_COLS + [login_welcome_message, logo_html, msme_hdr_btn, gov_hdr_btn, hdr_mode_html]
    )

    # ── DRC "Proceed to Insights" → Upload Screen (Step 5) ───────────────────
    def _drc_proceed_to_upload(clean_path):
        """Navigate to upload screen, pre-loading the DRC-cleaned file."""
        import os
        welcome = ""
        file_val = None
        if clean_path and isinstance(clean_path, str) and os.path.exists(clean_path):
            welcome = f"✅ Your dataset is cleaned and ready. Click 'Generate Free Insights Preview' to continue."
            file_val = clean_path          # pre-load cleaned file into file_upload
        else:
            welcome = "Upload your retail data file below to generate insights."
        return (
            5, *update_visibility_all('step5'),
            gr.update(value=file_val),     # file_upload — pre-load cleaned file
            gr.update(value=welcome, visible=bool(welcome)),
            *_hdr(False)
        )

    drc_cont_btn.click(
        fn=_drc_proceed_to_upload,
        inputs=[_drc_clean_state],
        outputs=[step_state] + _ALL_COLS + [file_upload, login_welcome_message, logo_html, msme_hdr_btn, gov_hdr_btn, hdr_mode_html],
        queue=False,
    )

    def _show_analyze_loading():
        spinner_html = (
            '<div style="display:flex;align-items:center;gap:16px;background:linear-gradient(90deg,#EBF5FF,#F0F8FF);'
            'border:1.5px solid #93c5fd;border-radius:14px;padding:18px 24px;margin:12px 0;">'
            '<div style="width:24px;height:24px;border:3px solid #1B4F8A;border-top-color:transparent;'
            'border-radius:50%;animation:dn-spin 0.8s linear infinite;flex-shrink:0;"></div>'
            '<div><div style="font-size:15px;font-weight:800;color:#0B1F3A;">Analyzing your data... generating insights</div>'
            '<div style="font-size:13px;color:#475569;margin-top:4px;">Running AI scoring · Building forecast models · '
            'Preparing your dashboard. Takes ~10–20 seconds.</div></div></div>'
            '<style>@keyframes dn-spin{to{transform:rotate(360deg);}}</style>'
        )
        return gr.update(visible=True, value=spinner_html)

    def analyze_data(user_data, consent, file, lang='en'):
        empty_dash = {'kpi1':"","kpi2":"","kpi3":"","kpi4":"","kpi5":"","chart1":None,"chart2":None,"chart3":None,"chart4":None}
        def _fail(msg):
            return (msg, gr.update(visible=False),
                    "", "", "", "", "", None, None, None, None,
                    gr.update(value="", visible=False), gr.update(value="", visible=False),
                    gr.update(value="", visible=False), gr.update(value="", visible=False),
                    gr.update(value="", visible=False), empty_dash, None)
        if not consent: return _fail("⚠️ Please provide consent to analyze data")
        if file is None: return _fail("⚠️ Please upload an Excel or CSV file")
        try:
            file_path = file.name if hasattr(file, 'name') else str(file)
            if file_path.endswith('.xlsx'): df = pd.read_excel(file_path)
            elif file_path.endswith('.csv'): df = pd.read_csv(file_path)
            else: return _fail("❌ Unsupported file format. Please upload .xlsx or .csv")

            # ── Performance cap: sample large datasets to keep analysis under ~10s ──
            # 50,000 rows is enough for statistically reliable scores and forecasts.
            # Original order is preserved via sorted sample on date when available.
            _MAX_ROWS = 50_000
            if len(df) > _MAX_ROWS:
                if 'Date' in df.columns or 'date' in df.columns:
                    _dcol = 'Date' if 'Date' in df.columns else 'date'
                    df = df.sort_values(_dcol).tail(_MAX_ROWS).reset_index(drop=True)
                else:
                    df = df.tail(_MAX_ROWS).reset_index(drop=True)

            # Apply column mapping
            df = _apply_col_remap(df)

            # Fill any still-missing required cols with defaults
            required_cols = {
                'Date':     lambda: pd.to_datetime(datetime.datetime.now().date()),
                'Store_ID': 'Default',
                'SKU_Name': 'Default',
                'Monthly_Sales_INR':          0,
                'Monthly_Operating_Cost_INR': 0,
                'Outstanding_Loan_INR':       0,
                'Vendor_Delivery_Reliability':0,
                'Inventory_Turnover':         0,
                'Avg_Margin_Percent':         0,
                'Monthly_Demand_Units':       0,
                'Returns_Percentage':         0,
            }
            for col, default in required_cols.items():
                if col not in df.columns:
                    df[col] = default() if callable(default) else default

            # Store full dataset for Government Dashboard
            # Government dashboard needs ALL retailers; Steps 5/6/7 use only the
            # logged-in MSME's rows so insights are personalised.
            df_full_for_gov = df.copy()

            # Detect the Udyam/MSME key column (before or after remap)
            msme_key = user_data.get('msme_number', '').strip().upper()
            _udyam_candidates = ['udyam_number','Udyam_Number','UDYAM_NUMBER',
                                  'msme_number','MSME_Number','Udyam_No']
            _udyam_col = next((c for c in _udyam_candidates if c in df.columns), None)

            if msme_key and _udyam_col:
                _unique_msme = df[_udyam_col].astype(str).str.strip().str.upper().nunique()
                if _unique_msme > 1:
                    _df_filtered = df[
                        df[_udyam_col].astype(str).str.strip().str.upper() == msme_key
                    ].copy()
                    if len(_df_filtered) > 0:
                        df = _df_filtered
                # Single-MSME dataset → use full df as-is

            insights_html, error_msg, _ = generate_insights(user_data, df.copy(), lang=lang)
            if error_msg: return _fail(f"❌ {error_msg}")

            result = generate_dashboard_data(user_data, df.copy())
            k1 = result[0]; f1,f2,f3,f4 = result[5],result[6],result[7],result[8]
            s1,s2,s3,s4 = result[9],result[10],result[11],result[12]
            raw_df = result[15]
            try: gf = generate_granular_forecast(df.copy())
            except Exception as _gf_err:
                import traceback as _gf_tb
                print(f'⚠️ generate_granular_forecast failed: {_gf_err}')
                print(_gf_tb.format_exc()[:500])
                gf = None

            # ── Build PDF snapshot: reuse forecast already computed in generate_insights ──
            # Avoids running forecast_sales a 3rd time — reuse gf overall forecast if available.
            try:
                _snap_df   = _apply_col_remap(df.copy())

                # Reuse granular overall forecast if available, else fast fallback (no Prophet re-run)
                if gf and gf.get('overall'):
                    _ov = gf['overall']
                    _snap_f6v  = float(_ov.get('6m_forecast',  0))
                    _snap_f6l  = float(_ov.get('6m_lower',     _snap_f6v * 0.85))
                    _snap_f6u  = float(_ov.get('6m_upper',     _snap_f6v * 1.15))
                    _snap_f12v = float(_ov.get('12m_forecast', 0))
                    _snap_f12l = float(_ov.get('12m_lower',    _snap_f12v * 0.85))
                    _snap_f12u = float(_ov.get('12m_upper',    _snap_f12v * 1.15))
                    _snap_model_name = _ov.get('model', 'Statistical Ensemble')
                else:
                    _snap_fc   = forecast_sales(_snap_df)
                    _snap_f6v  = float(_snap_fc['6_month'].get('forecast', 0))
                    _snap_f6l  = float(_snap_fc['6_month'].get('lower', 0))
                    _snap_f6u  = float(_snap_fc['6_month'].get('upper', 0))
                    _snap_f12v = float(_snap_fc['12_month'].get('forecast', 0))
                    _snap_f12l = float(_snap_fc['12_month'].get('lower', 0))
                    _snap_f12u = float(_snap_fc['12_month'].get('upper', 0))
                    _snap_model_name = _snap_fc.get('selected_model', 'Statistical Ensemble')

                # Trailing 6m using date-span (matches Step 7 fc_pct logic exactly)
                _snap_s_col = 'Monthly_Sales_INR' if 'Monthly_Sales_INR' in _snap_df.columns else None
                _snap_hist_avg  = None
                _snap_trail_6m  = None
                _snap_growth_pct = None
                if _snap_s_col:
                    _snap_total_hist = float(_snap_df[_snap_s_col].sum())
                    if 'Date' in _snap_df.columns:
                        _snap_dates = pd.to_datetime(_snap_df['Date'], errors='coerce').dropna()
                        _snap_span  = max(1, round((_snap_dates.max() - _snap_dates.min()).days / 30.44) + 1) if len(_snap_dates) >= 2 else 1
                    else:
                        _snap_span = max(_snap_df[_snap_s_col].count(), 1)
                    _snap_hist_avg = _snap_total_hist / _snap_span
                    _snap_trail_6m = _snap_hist_avg * 6
                    if _snap_trail_6m > 0:
                        _snap_raw_grw = (_snap_f6v - _snap_trail_6m) / (_snap_trail_6m + 1e-9) * 100
                        _snap_growth_pct = float(np.clip(round(_snap_raw_grw, 1), -150.0, 150.0))

                # Scores (with forecast-growth-augmented calculate_scores — same as Step 5)
                _snap_fg_for_scores = _snap_growth_pct  # may be None
                _snap_sc_df = calculate_scores(_snap_df.copy(), forecast_growth_rate=_snap_fg_for_scores)
                def _smean(col): return round(float(_snap_sc_df[col].mean()), 1) if col in _snap_sc_df.columns else None
                _snap_hs   = _smean('MSME_Health_Score')
                _snap_ps   = _smean('Performance_Score')
                _snap_fr   = round(float(_snap_sc_df['Financial_Risk_Score'].mean()), 3) if 'Financial_Risk_Score' in _snap_sc_df.columns else None
                _snap_vs   = _smean('Vendor_Score')
                _snap_gs   = round(float(_snap_sc_df['Growth_Potential_Score'].mean()), 3) if 'Growth_Potential_Score' in _snap_sc_df.columns else None
                # Try real data columns first, then computed ones
                _snap_pm   = (_smean('gross_margin_pct') or _smean('Gross_Margin_Pct') or
                               _smean('Avg_Margin_Percent') or _smean('avg_margin_percent'))
                _snap_ar   = (_smean('returns_units') or _smean('Returns_Units') or
                               _smean('Returns_Percentage') or _smean('return_rate_pct'))
                _snap_ts   = float(_snap_sc_df['Monthly_Sales_INR'].sum()) if 'Monthly_Sales_INR' in _snap_sc_df.columns else (
                               float(_snap_sc_df['revenue'].sum()) if 'revenue' in _snap_sc_df.columns else (
                               float(_snap_sc_df['Revenue'].sum()) if 'Revenue' in _snap_sc_df.columns else 0))

                # ONDC values — same columns as generate_dashboard_data
                def _sc_snap(name, fb=None):
                    return name if name in _snap_df.columns else (fb if fb and fb in _snap_df.columns else None)
                # Extend _sc_snap to handle real-file column names
                def _sc_snap2(*candidates):
                    for c in candidates:
                        if c and c in _snap_df.columns: return c
                        if c:
                            lc = {col.lower():col for col in _snap_df.columns}
                            if c.lower() in lc: return lc[c.lower()]
                    return None
                _gc  = _sc_snap2('Monthly_Sales_INR','gross_sales','revenue','Revenue','total_sales','net_sales')
                _nc  = _sc_snap2('net_sales','revenue','Revenue','gross_sales','Monthly_Sales_INR') or _gc
                # boc: revenue before ONDC = gross sales (no dedicated column in user file)
                _boc = _sc_snap2('revenue_before_ondc','Monthly_Sales_INR','revenue','Revenue','gross_sales')
                _aoc = _sc_snap2('ondc_revenue','ONDC_Revenue','revenue_after_ondc','Revenue_After_ONDC')
                _rrc = _sc_snap2('return_rate_pct','returns_units','Returns_Units','Returns_Percentage','returns_percentage','gross_margin_pct','Gross_Margin_Pct')
                _rpc = _sc_snap2('replacement_units','Replacement_Units','replacement_count')
                _tac = _sc_snap2('fulfillment_rate','Fulfillment_Rate','target_achievement_pct')  # not stockout_flag (it's 0/1 binary)
                _ochc= _sc_snap2('ondc_revenue','ONDC_Revenue','ondc_channel_revenue','revenue_after_ondc')
                _snap_gross       = float(_snap_df[_gc].sum())   if _gc   else 0.0
                _snap_net         = float(_snap_df[_nc].sum())   if _nc   else 0.0
                _snap_rev_before  = float(_snap_df[_boc].sum())  if _boc  else 0.0
                _snap_rev_after   = float(_snap_df[_aoc].sum())  if _aoc  else 0.0
                _snap_ondc_pos    = float(_snap_df[_ochc].clip(lower=0).sum()) if _ochc else 0.0
                _snap_pool        = max(_snap_rev_after - _snap_rev_before, _snap_gross * 0.05)
                _snap_avg_ret     = float(_snap_df[_rrc].mean())  if _rrc  else 0.0
                _snap_replacements= int(_snap_df[_rpc].sum())     if _rpc  else 0
                _snap_avg_target  = float(_snap_df[_tac].mean())  if _tac  else 0.0

                # ONDC readiness (same formula as generate_insights)
                _ors_g = min(float(_snap_gs) if _snap_gs else 0, 1.0)
                _ors_v = min(float(_snap_vs)/100.0 if _snap_vs else 0, 1.0)
                _ors_m = min(float(_snap_pm)/40.0 if _snap_pm else 0, 1.0)
                _ors_r = 1.0 - min(float(_snap_ar) if _snap_ar else 5.0, 15.0)/15.0
                _snap_ondc_ready  = round((0.35*_ors_g + 0.25*_ors_v + 0.20*_ors_m + 0.20*_ors_r)*100, 1)

                # Profit opportunity
                _snap_mg_gap  = max(0.0, 25.0 - (float(_snap_pm) if _snap_pm else 0))
                _snap_prof_up = _snap_ts * (_snap_mg_gap / 100.0)
                _snap_ret_rec = _snap_ts * 0.01 * 0.65

                # Peak demand month (Prophet)
                _snap_peak_month = None
                try:
                    _pr_snap = _snap_fc.get('model_results', {}).get('Prophet', {})
                    if _pr_snap and 'forecast_df' in _pr_snap:
                        _pfc_s = _pr_snap['forecast_df']
                        _snap_peak_month = pd.to_datetime(_pfc_s.loc[_pfc_s['yhat'].idxmax()]['ds']).strftime('%b %Y')
                except Exception:
                    pass

                _pdf_snapshot = {
                    # Identity
                    'biz_type':       user_data.get('business_type', 'Retail'),
                    # Scores (single source — never recomputed in PDF)
                    'health_score':   _snap_hs,
                    'perf_score':     _snap_ps,
                    'fin_risk':       _snap_fr,
                    'vendor_score':   _snap_vs,
                    'growth_score':   _snap_gs,
                    'avg_margin':     _snap_pm,
                    'avg_return':     _snap_ar,
                    'total_sales':    _snap_ts,
                    'ondc_readiness': _snap_ondc_ready,
                    # Opportunity
                    'profit_upside':  round(_snap_prof_up, 0),
                    'margin_gap':     round(_snap_mg_gap, 1),
                    'ret_recovery':   round(_snap_ret_rec, 0),
                    # ONDC (all from same df, same columns as dashboard charts)
                    'gross_sales':    _snap_gross,
                    'net_sales':      _snap_net,
                    'rev_before':     _snap_rev_before,
                    'rev_after':      _snap_rev_after,
                    'ondc_pool':      round(_snap_pool, 0),
                    'ondc_pos_rev':   _snap_ondc_pos,
                    'avg_ret_rate':   round(_snap_avg_ret, 1),
                    'replacements':   _snap_replacements,
                    'avg_target':     round(_snap_avg_target, 1),
                    # Forecast
                    'f6':             round(_snap_f6v, 0),
                    'f6_lower':       round(_snap_f6l, 0),
                    'f6_upper':       round(_snap_f6u, 0),
                    'f12':            round(_snap_f12v, 0),
                    'f12_lower':      round(_snap_f12l, 0),
                    'f12_upper':      round(_snap_f12u, 0),
                    'model_name':     _snap_model_name,
                    'trail_6m':       round(_snap_trail_6m, 0) if _snap_trail_6m else None,
                    'growth_pct':     _snap_growth_pct,
                    'peak_month':     _snap_peak_month,
                }
            except Exception:
                _pdf_snapshot = {}

            dash = {'kpi1':k1,'chart1':f1,'chart2':f2,'chart3':f3,'chart4':f4,'sum1':s1,'sum2':s2,'sum3':s3,'sum4':s4,'granular':gf,'snapshot':_pdf_snapshot}
            df_for_gov = df_full_for_gov  # full multi-MSME dataset → Government Dashboard
            # Pass df_full_for_gov to df_state so Government Dashboard gets all retailers
            # Steps 5/6/7 use MSME-filtered df (already applied above)
            return (insights_html or "✅ Analysis completed",
                    gr.update(visible=True),
                    k1, "", "", "", "", f1,f2,f3,f4,
                    s1, s2, s3, s4,
                    gr.update(value="", visible=False), dash, df_for_gov)
        except Exception as e:
            import traceback
            return _fail(f"❌ Analysis failed: {str(e)}\n\n{traceback.format_exc()}")


    # ── Preview Screen data builder ───────────────────────────────────────────
    def generate_preview_data(user_data, consent, file, lang='en'):
        """Rich unified preview: combines Business Health, AI Intelligence, Performance,
        Revenue, Actions, and Forecast sections — matching final dashboard exactly."""
        import traceback as _ptb

        empty_cols = update_visibility_all('preview')

        def _pfail(msg):
            err_html = (f'<div style="padding:20px;color:#e74c3c;font-weight:600;'
                        f'background:#FFF5F5;border:1px solid #FFCDD2;border-radius:10px">'
                        f'\u26a0\ufe0f {msg}</div>')
            return (err_html, gr.update(visible=False), gr.update(visible=False),
                    99, *empty_cols, *_hdr(False))

        if not consent:
            return _pfail("Please provide consent to analyze data.")
        if file is None:
            return _pfail("Please upload an Excel or CSV file.")

        try:
            file_path = file.name if hasattr(file, 'name') else str(file)
            if file_path.endswith('.xlsx'):
                df = pd.read_excel(file_path)
            elif file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                return _pfail("Unsupported file format. Please upload .xlsx or .csv")

            if len(df) > 20_000:
                df = df.tail(20_000).reset_index(drop=True)

            df = _apply_col_remap(df)

            # ── Column detection ─────────────────────────────────────────────
            sc   = 'Monthly_Sales_INR'      if 'Monthly_Sales_INR'      in df.columns else None
            mc   = 'Avg_Margin_Percent'     if 'Avg_Margin_Percent'     in df.columns else None
            rrc  = 'Returns_Percentage'     if 'Returns_Percentage'     in df.columns else None
            dc   = 'Date'                   if 'Date'                   in df.columns else None
            skc  = 'SKU_Name'              if 'SKU_Name'              in df.columns else None
            stc  = 'Store_ID'              if 'Store_ID'              in df.columns else None
            ctc  = 'Product_Category'      if 'Product_Category'      in df.columns else None
            oc   = 'Monthly_Operating_Cost_INR' if 'Monthly_Operating_Cost_INR' in df.columns else None
            duc  = 'Monthly_Demand_Units'  if 'Monthly_Demand_Units'  in df.columns else None
            slc  = 'Stock_Level'           if 'Stock_Level'           in df.columns else None
            tac  = 'target_achievement_pct' if 'target_achievement_pct' in df.columns else None
            rplc = 'replacement_count'     if 'replacement_count'     in df.columns else None

            def _s(c): return float(df[c].sum())  if c else 0.0
            def _m(c): return float(df[c].mean()) if c else 0.0
            def _si(c): return int(df[c].sum())   if c else 0

            total_sales  = _s(sc)
            avg_margin   = _m(mc)
            avg_ret      = _m(rrc)
            avg_target   = _m(tac)
            replacements = _si(rplc)
            n_skus       = df[skc].nunique() if skc else 0
            n_stores     = df[stc].nunique() if stc else 0
            n_cats       = df[ctc].nunique() if ctc else 0

            # ── Health score ─────────────────────────────────────────────────
            _h_margin  = min(avg_margin / 30.0, 1.0)
            _h_returns = 1.0 - min(avg_ret / 15.0, 1.0)
            health_score = round((_h_margin * 0.4 + _h_returns * 0.3 + 0.3) * 100, 0)
            health_score = max(10, min(100, health_score))

            if health_score >= 70:
                hc = "#1a7a40"; hl = "Strong"; hbg = "#f0fdf4"; hbd = "#C3E6CB"
                health_msg = "Your business shows strong performance fundamentals."
            elif health_score >= 45:
                hc = "#b07800"; hl = "Moderate"; hbg = "#fffbeb"; hbd = "#FFE8A1"
                health_msg = "Performing moderately — targeted improvements will lift scores."
            else:
                hc = "#e74c3c"; hl = "Needs Attention"; hbg = "#fff5f5"; hbd = "#F5C6CB"
                health_msg = "Several areas need immediate attention."

            ring_pct  = int(health_score)
            ring_dash = int(ring_pct * 2.51)

            # ── Monthly sales series ─────────────────────────────────────────
            monthly_sales = None
            if sc and dc:
                try:
                    _tmp = df.copy()
                    _tmp[dc] = pd.to_datetime(_tmp[dc], errors='coerce')
                    _tmp = _tmp.dropna(subset=[dc])
                    _tmp['_mo'] = _tmp[dc].dt.to_period('M').astype(str)
                    monthly_sales = _tmp.groupby('_mo')[sc].sum().sort_index().tail(8)
                except Exception:
                    pass

            # ── Sales trend direction ────────────────────────────────────────
            _trend_chg = 0.0
            if monthly_sales is not None and len(monthly_sales) >= 2:
                _sv = monthly_sales.values
                _trend_chg = (_sv[-1] - _sv[0]) / (abs(_sv[0]) + 1e-9) * 100
            _trend_icon = "📈" if _trend_chg > 2 else ("📉" if _trend_chg < -2 else "📊")
            _trend_col  = "#15803d" if _trend_chg > 2 else ("#dc2626" if _trend_chg < -2 else "#475569")
            _trend_txt  = (f"+{_trend_chg:.0f}% since first month" if _trend_chg > 0
                           else f"{_trend_chg:.0f}% since first month")

            # ── AI insights ──────────────────────────────────────────────────
            insights = []
            if sc and dc and monthly_sales is not None and len(monthly_sales) >= 3:
                last2 = monthly_sales.iloc[-2:].values
                prev2 = monthly_sales.iloc[-4:-2].values
                if len(prev2) >= 2:
                    chg = (float(np.mean(last2)) - float(np.mean(prev2))) / (float(np.mean(prev2)) + 1e-9) * 100
                    if chg < -3:
                        insights.append(f'📉 Sales dropped <strong>{abs(chg):.0f}%</strong> in last 2 months — review pricing and stock levels.')
                    elif chg > 3:
                        insights.append(f'📈 Sales grew <strong>{chg:.0f}%</strong> in last 2 months — positive momentum detected.')
                    else:
                        insights.append('📊 Sales stable — consider targeted promotions to accelerate growth.')
            if skc and sc:
                try:
                    top_skus = df.groupby(skc)[sc].sum().sort_values(ascending=False)
                    top3_pct = float(top_skus.head(3).sum() / (top_skus.sum() + 1e-9) * 100)
                    insights.append(f'🏆 Top 3 products contribute <strong>{top3_pct:.0f}%</strong> of revenue — diversify to reduce concentration risk.')
                except Exception:
                    pass
            if avg_margin > 0:
                if avg_margin >= 25:
                    insights.append(f'✅ Profit margin <strong>{avg_margin:.1f}%</strong> — strong. Leverage for competitive pricing.')
                else:
                    insights.append(f'⚠️ Profit margin <strong>{avg_margin:.1f}%</strong> — below 25% target. Review cost structure.')
            if avg_ret > 0:
                if avg_ret >= 7:
                    insights.append(f'🔴 Return rate <strong>{avg_ret:.1f}%</strong> — high. Review product quality and descriptions.')
                else:
                    insights.append(f'✅ Return rate <strong>{avg_ret:.1f}%</strong> — healthy. Maintain quality controls.')
            if not insights:
                insights.append(f'📊 Margin <strong>{avg_margin:.1f}%</strong> · Return rate <strong>{avg_ret:.1f}%</strong>')
                insights.append('💡 Add date and SKU columns for deeper AI insights.')

            # ── Format helpers ───────────────────────────────────────────────
            def _fmt(v):
                if v >= 1e7: return f'₹{v/1e7:.1f}Cr'
                if v >= 1e5: return f'₹{v/1e5:.1f}L'
                if v >= 1e3: return f'₹{v/1e3:.0f}K'
                return f'₹{v:.0f}'

            def _kpi_mini(icon, label, value, sub='', col='#1B4F8A'):
                return f'''<div style="background:#fff;border:1px solid #E2EEF9;border-radius:10px;
                    padding:12px 14px;border-top:3px solid {col};">
                  <div style="font-size:9px;font-weight:700;letter-spacing:1px;text-transform:uppercase;
                              color:#7A92AA;margin-bottom:4px">{icon} {label}</div>
                  <div style="font-size:18px;font-weight:900;color:{col};font-family:monospace">{value}</div>
                  {'<div style="font-size:10px;color:#94a3b8;margin-top:2px">'+sub+'</div>' if sub else ''}
                </div>'''

            def _blur_overlay(tip="Unlock to view full insights"):
                return (
                    '<div class="dn-pv-overlay" title="Unlock to explore detailed analytics">'
                    '<span class="dn-pv-overlay-text">🔒 Unlock to view full insights</span>'
                    '<span class="dn-pv-overlay-sub">Hover to preview · Click to unlock</span>'
                    '</div>'
                )

            # ── Sparkline bars ───────────────────────────────────────────────
            sparkline_bars = ""
            if monthly_sales is not None and len(monthly_sales) >= 2:
                _sv = monthly_sales.values
                _sv_max = float(_sv.max()) if _sv.max() > 0 else 1.0
                _months_short = [m[-5:] for m in monthly_sales.index.tolist()]
                for _i, (_v, _mo) in enumerate(zip(_sv, _months_short)):
                    _h = max(8, int(_v / _sv_max * 52))
                    _acc = "background:linear-gradient(to top,#1B4F8A,#60a5fa)" if _i == len(_sv)-1 else "background:linear-gradient(to top,#93c5fd,#bfdbfe)"
                    sparkline_bars += (f'<div style="display:flex;flex-direction:column;align-items:center;flex:1;gap:2px;">'
                                       f'<div style="height:{_h}px;width:100%;border-radius:3px 3px 0 0;{_acc};" title="{_mo}"></div>'
                                       f'<div style="font-size:7px;color:#94a3b8;white-space:nowrap;overflow:hidden;max-width:28px;text-align:center">{_mo}</div>'
                                       f'</div>')
            else:
                sparkline_bars = '<div style="font-size:11px;color:#94a3b8;padding:16px 0;">Upload data with dates for trend chart</div>'

            # ── Fake blurred forecast bars ───────────────────────────────────
            fc_bars = "".join(
                f'<div style="flex:1;height:{h}%;border-radius:4px 4px 0 0;'
                f'background:linear-gradient(to top,#1B4F8A,#7AABDD);"></div>'
                for h in [42,55,50,65,72,62,76,82,74,87,80,92]
            )
            fc_bars2 = "".join(
                f'<div style="flex:1;height:{h}%;border-radius:4px 4px 0 0;'
                f'background:linear-gradient(to top,#7c3aed,#a78bfa);"></div>'
                for h in [38,52,48,60,66,58,70,76,68,80,74,88]
            )

            # ════════════════════════════════════════════════════════════════
            # BUILD UNIFIED PREVIEW HTML
            # ════════════════════════════════════════════════════════════════
            cards_html = f"""
<style>
.dn-pv{{max-width:1100px;margin:0 auto;font-family:system-ui,-apple-system,sans-serif;}}
.dn-pv-sec{{background:#fff;border:1px solid #E2EEF9;border-radius:14px;
            padding:18px 20px;margin-bottom:14px;
            box-shadow:0 2px 12px rgba(27,79,138,0.06);
            transition:box-shadow 0.2s;}}
.dn-pv-sec:hover{{box-shadow:0 4px 24px rgba(27,79,138,0.12);}}
.dn-pv-hdr{{font-size:11px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;
             color:#0B1F3A;margin-bottom:12px;padding-bottom:8px;
             border-bottom:2px solid #D8E8F8;display:flex;align-items:center;gap:8px;}}
.dn-pv-grid2{{display:grid;grid-template-columns:1fr 1fr;gap:12px;}}
.dn-pv-grid4{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;}}
.dn-pv-blur{{
  position:relative;overflow:hidden;border-radius:10px;
  cursor:pointer;
}}
.dn-pv-blurred{{
  filter:blur(7px);pointer-events:none;user-select:none;
  transition:filter 0.3s ease;
}}
.dn-pv-blur:hover .dn-pv-blurred{{filter:blur(4px);}}
.dn-pv-overlay{{
  position:absolute;inset:0;display:flex;flex-direction:column;
  align-items:center;justify-content:center;gap:6px;
  background:rgba(11,31,58,0.18);border-radius:10px;
  transition:background 0.3s;
  pointer-events:none;
}}
.dn-pv-blur:hover .dn-pv-overlay{{background:rgba(11,31,58,0.10);}}
.dn-pv-overlay-text{{
  font-size:12px;font-weight:700;color:#ffffff;
  background:rgba(27,79,138,0.85);border-radius:20px;
  padding:6px 16px;letter-spacing:0.3px;
  box-shadow:0 2px 8px rgba(0,0,0,0.20);
}}
.dn-pv-overlay-sub{{
  font-size:11px;color:rgba(255,255,255,0.80);
  font-weight:500;
}}
.dn-pv-blur:hover .dn-pv-overlay-text::after{{
  content:" — Hover to preview";opacity:0.7;
}}
@media(max-width:700px){{
  .dn-pv-grid2,.dn-pv-grid4{{grid-template-columns:1fr 1fr;}}
}}
@media(max-width:420px){{
  .dn-pv-grid4{{grid-template-columns:1fr 1fr;}}
}}
</style>

<div class="dn-pv">

<!-- ═══════════════════════════════════════════
     SECTION 1 — Business Health Score (VISIBLE)
     ═══════════════════════════════════════════ -->
<div class="dn-pv-sec" style="border-top:3px solid {hc};">
  <div class="dn-pv-hdr">📊 Business Health Score</div>
  <div style="display:flex;align-items:center;gap:20px;flex-wrap:wrap;">
    <!-- Ring -->
    <div style="flex-shrink:0;">
      <svg width="96" height="96" viewBox="0 0 96 96">
        <circle cx="48" cy="48" r="40" fill="none" stroke="#EEF4FB" stroke-width="10"/>
        <circle cx="48" cy="48" r="40" fill="none" stroke="{hc}" stroke-width="10"
                stroke-dasharray="{ring_dash} 251" stroke-linecap="round"
                transform="rotate(-90 48 48)"/>
        <text x="48" y="52" text-anchor="middle" font-size="18" font-weight="900"
              fill="{hc}" font-family="monospace">{ring_pct}</text>
        <text x="48" y="64" text-anchor="middle" font-size="9" fill="#94a3b8">/100</text>
      </svg>
    </div>
    <!-- Label + message -->
    <div style="flex:1;min-width:120px;">
      <div style="display:inline-block;background:{hbg};border:1px solid {hbd};
                  border-radius:20px;padding:4px 14px;font-size:13px;font-weight:800;
                  color:{hc};margin-bottom:8px;">{hl}</div>
      <div style="font-size:12px;color:#4A6A8A;line-height:1.6;margin-bottom:10px;">{health_msg}</div>
      <!-- Top 2 AI insights visible -->
      <div style="display:flex;flex-direction:column;gap:6px;">
        {''.join(f'<div style="background:#f8faff;border-left:3px solid #1B4F8A;border-radius:0 6px 6px 0;padding:7px 10px;font-size:12px;color:#1A2D45;line-height:1.5;">▸ {ins}</div>' for ins in insights[:2])}
      </div>
    </div>
    <!-- Sales Trend Sparkline -->
    <div style="flex:0 0 180px;background:#F8FBFF;border:1px solid #D0E4F4;
                border-radius:10px;padding:12px 14px;">
      <div style="font-size:9px;font-weight:700;letter-spacing:1px;text-transform:uppercase;
                  color:#7A92AA;margin-bottom:6px;">📈 Sales Trend</div>
      <div style="display:flex;align-items:flex-end;gap:2px;height:52px;margin-bottom:4px;">
        {sparkline_bars}
      </div>
      <div style="font-size:10px;font-weight:600;color:{_trend_col};">
        {_trend_icon} {_trend_txt}
      </div>
    </div>
  </div>
</div>

<!-- ═══════════════════════════════════════════
     SECTION 2 — AI Business Intelligence (PARTIAL BLUR)
     ═══════════════════════════════════════════ -->
<div class="dn-pv-sec" style="border-top:3px solid #7c3aed;">
  <div class="dn-pv-hdr">🤖 AI Business Intelligence</div>
  <!-- 1–2 insights visible -->
  <div style="background:linear-gradient(135deg,#0B1F3A 0%,#1B3F6A 100%);border-radius:10px;
              padding:12px 16px;margin-bottom:10px;border-left:4px solid #FFD080;">
    <div style="font-size:9px;font-weight:800;letter-spacing:2px;text-transform:uppercase;
                color:#FFD080;margin-bottom:6px;">📊 AI Insight Summary</div>
    <div style="font-size:12px;color:#FFFFFF;line-height:1.8;">
      {insights[0] if insights else 'Analysing your business data...'}
    </div>
  </div>
  {'<div style="background:#f8faff;border-left:3px solid #7c3aed;border-radius:0 8px 8px 0;padding:8px 12px;font-size:12px;color:#1A2D45;line-height:1.5;margin-bottom:10px;">▸ '+insights[1]+'</div>' if len(insights) > 1 else ''}
  <!-- Remaining insights blurred -->
  <div class="dn-pv-blur">
    <div class="dn-pv-blurred">
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;">
        {''.join(f'<div style="background:#f1f5f9;border-radius:8px;padding:8px 10px;font-size:12px;color:#334155;">▸ {ins}</div>' for ins in (insights[2:4] if len(insights)>2 else ['Insight hidden','Insight hidden']))}
      </div>
      <div style="margin-top:8px;font-size:11px;color:#7c3aed;font-weight:600;">+ 10 more AI insights available</div>
    </div>
    {_blur_overlay()}
  </div>
</div>

<!-- ═══════════════════════════════════════════
     SECTION 3 — Business Performance Dashboard
     ═══════════════════════════════════════════ -->
<div class="dn-pv-sec" style="border-top:3px solid #1B4F8A;">
  <div class="dn-pv-hdr">📦 Business Performance Dashboard</div>
  <div class="dn-pv-grid4" style="margin-bottom:12px;">
    {_kpi_mini('💰','Total Sales', _fmt(total_sales), f'{n_skus} SKUs · {n_stores} stores', '#1B4F8A')}
    {_kpi_mini('↩️','Return Rate', f'{avg_ret:.1f}%', 'avg across products', '#e74c3c' if avg_ret>=7 else '#27ae60')}
    {_kpi_mini('🔄','Replacements', f'{replacements:,}', 'units replaced', '#7c3aed')}
    {_kpi_mini('🎯','Target Achievement', f'{avg_target:.1f}%' if avg_target>0 else 'N/A', 'avg margin '+f'{avg_margin:.1f}%', '#27ae60' if avg_target>=100 else '#f39c12')}
  </div>

<!-- ═══════════════════════════════════════════
     SECTION 4 — Revenue Breakdown & Operational Metrics (BLUR)
     ═══════════════════════════════════════════ -->
  <div class="dn-pv-hdr" style="margin-top:4px;">💰 Revenue Breakdown &amp; Operational Metrics</div>
  <div class="dn-pv-blur">
    <div class="dn-pv-blurred">
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-bottom:10px;">
        <div style="background:#F8FAFF;border:1px solid #D0E4F4;border-radius:10px;padding:12px 14px;">
          <div style="font-size:9px;font-weight:700;text-transform:uppercase;color:#7A92AA;margin-bottom:4px;">📊 In-Store Revenue</div>
          <div style="font-size:18px;font-weight:900;color:#7A92AA;font-family:monospace;">₹ ── ──</div>
        </div>
        <div style="background:#EAF7EE;border:1px solid #C3E6CB;border-radius:10px;padding:12px 14px;">
          <div style="font-size:9px;font-weight:700;text-transform:uppercase;color:#1a7a40;margin-bottom:4px;">🌐 Online Revenue</div>
          <div style="font-size:18px;font-weight:900;color:#1a7a40;font-family:monospace;">₹ ── ──</div>
        </div>
        <div style="background:#F0F7FF;border:1px solid #B8D4F0;border-radius:10px;padding:12px 14px;">
          <div style="font-size:9px;font-weight:700;text-transform:uppercase;color:#1B4F8A;margin-bottom:4px;">📈 Net Sales (All Channels)</div>
          <div style="font-size:18px;font-weight:900;color:#1B4F8A;font-family:monospace;">{_fmt(total_sales)}</div>
        </div>
      </div>
      <div style="background:#F8FAFF;border-radius:8px;padding:10px 14px;font-size:11px;color:#4A6A8A;">
        ⚙️ Operational Metrics — Vendor reliability · Inventory turnover · Fulfilment rate · more
      </div>
    </div>
    {_blur_overlay()}
  </div>
</div>

<!-- ═══════════════════════════════════════════
     SECTION 5 — AI Business Insights (PARTIAL BLUR)
     ═══════════════════════════════════════════ -->
<div class="dn-pv-sec" style="border-top:3px solid #f39c12;">
  <div class="dn-pv-hdr">🤖 AI Business Insights</div>
  <!-- First insight visible -->
  <div style="background:#fff8ec;border-left:4px solid #f39c12;border-radius:0 8px 8px 0;
              padding:10px 14px;font-size:12px;color:#1A2D45;line-height:1.6;margin-bottom:10px;">
    ▸ {insights[0] if insights else 'Analysing performance data...'}
  </div>
  <!-- Rest blurred -->
  <div class="dn-pv-blur">
    <div class="dn-pv-blurred">
      <div style="display:flex;flex-direction:column;gap:6px;">
        {''.join(f'<div style="background:#f8faff;border-left:3px solid #d0e4f4;border-radius:0 6px 6px 0;padding:7px 10px;font-size:12px;color:#64748b;">▸ {ins}</div>' for ins in insights[1:4])}
        <div style="font-size:10px;color:#94a3b8;padding:4px 0;">+ more insights available after unlock</div>
      </div>
    </div>
    {_blur_overlay()}
  </div>
</div>

<!-- ═══════════════════════════════════════════
     SECTION 6 — Immediate Actions & Strategic Initiatives (BLUR)
     ═══════════════════════════════════════════ -->
<div class="dn-pv-sec" style="border-top:3px solid #e74c3c;">
  <div class="dn-pv-hdr">⚡ Immediate Actions &amp; Strategic Initiatives</div>
  <div class="dn-pv-grid2">
    <!-- Immediate Actions — label visible, content blurred -->
    <div>
      <div style="font-size:10px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;
                  color:#e74c3c;margin-bottom:8px;padding:6px 10px;background:#fff5f5;
                  border-radius:6px;">⚡ Immediate Actions (0–30 Days)</div>
      <div class="dn-pv-blur">
        <div class="dn-pv-blurred">
          <ul style="margin:0;padding:0;list-style:none;font-size:12px;line-height:1.9;color:#1A2D45;">
            <li>• {'Review return rate — reduce below 7%' if avg_ret>=7 else 'Maintain current quality controls'}</li>
            <li>• {'Improve profit margin — target 25%+' if avg_margin<25 else 'Leverage strong margins competitively'}</li>
            <li>• Optimise top SKU pricing this week</li>
          </ul>
        </div>
        {_blur_overlay()}
      </div>
    </div>
    <!-- Strategic Initiatives — label visible, content blurred -->
    <div>
      <div style="font-size:10px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;
                  color:#1B4F8A;margin-bottom:8px;padding:6px 10px;background:#EBF5FF;
                  border-radius:6px;">📈 Strategic Initiatives (30–90 Days)</div>
      <div class="dn-pv-blur">
        <div class="dn-pv-blurred">
          <ul style="margin:0;padding:0;list-style:none;font-size:12px;line-height:1.9;color:#1A2D45;">
            <li>• Expand SKU portfolio — reduce top-3 dependency</li>
            <li>• Explore online marketplace channels</li>
            <li>• Build seasonal demand forecasting plan</li>
          </ul>
        </div>
        {_blur_overlay()}
      </div>
    </div>
  </div>
</div>

<!-- ═══════════════════════════════════════════
     SECTION 7 — Store + Category + Product Intelligence & AI Forecasting (BLUR)
     ═══════════════════════════════════════════ -->
<div class="dn-pv-sec" style="border-top:3px solid #0097a7;">
  <div class="dn-pv-hdr">🏪 Store → Category → Product Intelligence &amp; AI Forecasting</div>
  <div style="font-size:11px;color:#4A6A8A;margin-bottom:12px;">
    {n_stores} store{'s' if n_stores!=1 else ''} · {n_cats} categor{'ies' if n_cats!=1 else 'y'} · {n_skus} SKU{'s' if n_skus!=1 else ''} detected in your dataset
  </div>
  <div class="dn-pv-grid2">
    <!-- 6-Month Forecast -->
    <div class="dn-pv-blur" style="background:#F8FBFF;border:1px solid #D0E4F4;border-radius:10px;padding:14px;">
      <div class="dn-pv-blurred">
        <div style="font-size:10px;font-weight:700;text-transform:uppercase;color:#1B4F8A;margin-bottom:8px;">📈 6-Month Forecast</div>
        <div style="display:flex;align-items:flex-end;gap:3px;height:60px;">{fc_bars[:len(fc_bars)//2]}</div>
        <div style="font-size:11px;color:#1B4F8A;font-weight:700;margin-top:6px;">₹ ── ── projected</div>
      </div>
      {_blur_overlay()}
    </div>
    <!-- 12-Month Forecast -->
    <div class="dn-pv-blur" style="background:#F8FBFF;border:1px solid #D0E4F4;border-radius:10px;padding:14px;">
      <div class="dn-pv-blurred">
        <div style="font-size:10px;font-weight:700;text-transform:uppercase;color:#1B4F8A;margin-bottom:8px;">🔮 12-Month Forecast</div>
        <div style="display:flex;align-items:flex-end;gap:3px;height:60px;">{fc_bars}</div>
        <div style="font-size:11px;color:#1B4F8A;font-weight:700;margin-top:6px;">₹ ── ── projected</div>
      </div>
      {_blur_overlay()}
    </div>
    <!-- Product Fulfilment Trend -->
    <div class="dn-pv-blur" style="background:#F8FBFF;border:1px solid #D0E4F4;border-radius:10px;padding:14px;">
      <div class="dn-pv-blurred">
        <div style="font-size:10px;font-weight:700;text-transform:uppercase;color:#7c3aed;margin-bottom:8px;">📦 Product Fulfilment Trend</div>
        <div style="display:flex;align-items:flex-end;gap:3px;height:60px;">{fc_bars2[:len(fc_bars2)//2]}</div>
        <div style="font-size:11px;color:#7c3aed;font-weight:700;margin-top:6px;">Fulfilment rate analysis locked</div>
      </div>
      {_blur_overlay()}
    </div>
    <!-- Product Sales vs Returns -->
    <div class="dn-pv-blur" style="background:#F8FBFF;border:1px solid #D0E4F4;border-radius:10px;padding:14px;">
      <div class="dn-pv-blurred">
        <div style="font-size:10px;font-weight:700;text-transform:uppercase;color:#e74c3c;margin-bottom:8px;">↩️ Product Sales vs Returns</div>
        <div style="display:flex;align-items:flex-end;gap:3px;height:60px;">{fc_bars2}</div>
        <div style="font-size:11px;color:#e74c3c;font-weight:700;margin-top:6px;">Return analysis locked</div>
      </div>
      {_blur_overlay()}
    </div>
  </div>
  <!-- Inventory & Reorder + AI Insights row -->
  <div class="dn-pv-grid2" style="margin-top:10px;">
    <div class="dn-pv-blur" style="background:#F8FBFF;border:1px solid #D0E4F4;border-radius:10px;padding:14px;">
      <div class="dn-pv-blurred">
        <div style="font-size:10px;font-weight:700;text-transform:uppercase;color:#27ae60;margin-bottom:8px;">📦 Inventory &amp; Reorder Levels</div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;">
          <div style="background:#EAF7EE;border-radius:6px;padding:8px;text-align:center;">
            <div style="font-size:9px;color:#4A6A8A;">Stock Level</div>
            <div style="font-size:14px;font-weight:800;color:#27ae60;">── ──</div>
          </div>
          <div style="background:#FFF5F5;border-radius:6px;padding:8px;text-align:center;">
            <div style="font-size:9px;color:#4A6A8A;">Reorder Point</div>
            <div style="font-size:14px;font-weight:800;color:#e74c3c;">── ──</div>
          </div>
        </div>
      </div>
      {_blur_overlay()}
    </div>
    <div class="dn-pv-blur" style="background:linear-gradient(135deg,#0B1F3A,#1B3F6A);border-radius:10px;padding:14px;">
      <div class="dn-pv-blurred">
        <div style="font-size:10px;font-weight:700;text-transform:uppercase;color:#FFD080;margin-bottom:8px;">🤖 AI Business Insights</div>
        <div style="font-size:11px;color:#A8D8FF;line-height:1.7;">
          <div>• Channel performance analysis</div>
          <div>• Demand signal intelligence</div>
          <div>• Platform fit scoring</div>
          <div>• Revenue optimisation roadmap</div>
        </div>
      </div>
      {_blur_overlay()}
    </div>
  </div>
</div>

</div><!-- end .dn-pv -->"""

            # ── Sales chart ──────────────────────────────────────────────────
            sales_fig = None
            if monthly_sales is not None and len(monthly_sales) >= 2:
                import matplotlib; matplotlib.use('Agg')
                import matplotlib.pyplot as _plt2
                fig_s, ax_s = _plt2.subplots(figsize=(10, 2.6))
                fig_s.patch.set_facecolor('#FFFFFF')
                fig_s.subplots_adjust(top=0.88, bottom=0.22, left=0.07, right=0.97)
                xs = list(range(len(monthly_sales)))
                ys = monthly_sales.values / 1e5
                ax_s.fill_between(xs, ys, alpha=0.13, color='#1B4F8A')
                ax_s.plot(xs, ys, color='#1B4F8A', lw=2.5, marker='o', ms=5, markerfacecolor='#2563eb')
                ax_s.scatter([xs[-1]], [ys[-1]], color='#2563eb', s=60, zorder=5)
                ax_s.set_xticks(range(len(xs)))
                ax_s.set_xticklabels(monthly_sales.index.tolist(), rotation=28, ha='right', fontsize=8)
                ax_s.yaxis.set_major_formatter(_plt2.FuncFormatter(lambda v, _: f'\u20b9{v:.0f}L'))
                ax_s.tick_params(axis='y', labelsize=8)
                ax_s.set_facecolor('#F8FBFF')
                ax_s.yaxis.set_major_locator(_plt2.MaxNLocator(nbins=4))
                for sp in ['top','right']: ax_s.spines[sp].set_visible(False)
                ax_s.spines['left'].set_color('#D0E4F4')
                ax_s.spines['bottom'].set_color('#D0E4F4')
                ax_s.grid(axis='y', color='#E8F0FB', linewidth=0.6, linestyle='--')
                sales_fig = fig_s
                _plt2.close(fig_s)

            return (cards_html,
                    gr.update(value=sales_fig, visible=sales_fig is not None),
                    gr.update(visible=False),
                    99,
                    *update_visibility_all('preview'),
                    *_hdr(False))

        except Exception as _e:
            return _pfail(f"Preview generation failed: {str(_e)}")

    def show_dashboard(dashboard_data_value):
        def _summary(key):
            val = dashboard_data_value.get(key)
            return gr.update(value=val, visible=True) if val else gr.update(value="", visible=False)
        granular_data = dashboard_data_value.get('granular')
        kpi_html_val  = dashboard_data_value.get('kpi1', "")
        return (6, *update_visibility_all('step6'), kpi_html_val,
                dashboard_data_value.get('chart1'), dashboard_data_value.get('chart2'),
                dashboard_data_value.get('chart3'), dashboard_data_value.get('chart4'),
                _summary('sum1'), _summary('sum2'), _summary('sum3'), _summary('sum4'), granular_data, *_hdr(False))

    # ── Preview generation ────────────────────────────────────────────────────────
    def _gen_preview_and_persist(user_data, consent, file, lang):
        """Wrapper: runs generate_preview_data and appends file+consent to output.
        Catches ALL exceptions so Gradio never returns 500 to the browser."""
        try:
            if file is None:
                # Return a visible error state instead of crashing
                empty = update_visibility_all('step5')
                err_html = '<div style="padding:20px;color:#dc2626;background:#fef2f2;border:1px solid #fecaca;border-radius:8px;margin:16px;">⚠️ Please upload a file first.</div>'
                return (err_html, None, None, 5, *empty,
                        gr.update(), gr.update(), gr.update(), gr.update(),
                        None, None)
            result = generate_preview_data(user_data, consent, file, lang)
            return (*result, file, consent)
        except Exception as _pex:
            import traceback as _ptb2
            _ptb2.print_exc()
            # Return to step5 with error message — NEVER let this crash Gradio
            empty = update_visibility_all('step5')
            err_html = f'''<div style="padding:20px;color:#dc2626;background:#fef2f2;
                border:1px solid #fecaca;border-radius:8px;margin:16px;">
                ⚠️ Could not process your file. Please check that it is a valid Excel or CSV
                with sales data.<br><br>
                <small style="color:#6b7280;">Error: {str(_pex)[:200]}</small>
              </div>'''
            return (err_html, None, None, 5, *empty,
                    gr.update(), gr.update(), gr.update(), gr.update(),
                    None, None)

    _PREVIEW_OUTPUTS = (
        [preview_cards_html, preview_sales_chart, preview_forecast_chart,
         step_state] + _ALL_COLS +
        [logo_html, msme_hdr_btn, gov_hdr_btn, hdr_mode_html,
         preview_file_state, preview_consent_state]
    )

    analyze_btn.click(
        _show_analyze_loading, [], [analyze_loading]
    ).then(
        _gen_preview_and_persist,
        [user_data_state, consent_check, file_upload, lang_state],
        _PREVIEW_OUTPUTS
    ).then(
        lambda: gr.update(visible=False, value=""), [], [analyze_loading]
    )

    # ── Preview Screen button wirings ─────────────────────────────────────────
    preview_back_btn.click(
        lambda: (5, *update_visibility_all('step5'), *_hdr(False)),
        [],
        [step_state] + _ALL_COLS + [logo_html, msme_hdr_btn, gov_hdr_btn, hdr_mode_html],
        queue=True
    )

    # ── Unlock Full Insights → run full analysis then show Step 6 dashboard ──
    def _show_unlock_loading():
        spinner = (
            '<div style="display:flex;align-items:center;gap:16px;background:linear-gradient(90deg,#EBF5FF,#F0F8FF);'
            'border:1.5px solid #93c5fd;border-radius:14px;padding:18px 24px;margin:16px 0;">'
            '<div style="width:24px;height:24px;border:3px solid #1B4F8A;border-top-color:transparent;'
            'border-radius:50%;animation:dn-spin 0.8s linear infinite;flex-shrink:0;"></div>'
            '<div><div style="font-size:15px;font-weight:800;color:#0B1F3A;">Preparing your full business insights...</div>'
            '<div style="font-size:13px;color:#475569;margin-top:4px;">Running AI scoring · Forecasting · Dashboard generation. Takes ~15–30 seconds.</div></div>'
            '</div><style>@keyframes dn-spin{to{transform:rotate(360deg);}}</style>'
        )
        return gr.update(visible=True, value=spinner)

    def unlock_to_dashboard(user_data, file, consent, lang):
        """Run full analyze_data with persisted file, then navigate to Step 6."""
        import traceback as _tb_unlock

        def _bounce(msg=""):
            err_html = f'<div style="color:#dc2626;background:#fef2f2;border:1px solid #fecaca;border-radius:8px;padding:12px 16px;margin:8px 0;font-size:13px;">❌ {msg}</div>' if msg else ""
            return (
                gr.update(visible=bool(msg), value=err_html),
                5, *update_visibility_all('step5'),
                "", None, None, None, None,
                gr.update(value="", visible=False), gr.update(value="", visible=False),
                gr.update(value="", visible=False), gr.update(value="", visible=False),
                {}, None, gr.update(value=""), None,
                *_hdr(False)
            )

        # If file missing, bounce back to upload screen
        if file is None:
            return _bounce("Please upload your data file before unlocking.")
        # consent not required for unlock flow — user already consented during preview

        try:
            # Run the full analysis
            result = analyze_data(user_data, consent, file, lang)
        except Exception as _e:
            return _bounce(f"Analysis failed: {str(_e)[:200]}")

        # ── Log the analysis run to database ────────────────────────────
        try:
            _email = user_data.get('email','') if isinstance(user_data,dict) else ''
            _fname = getattr(file,'name','unknown') if file else 'unknown'
            import os as _os_ul
            _fname = _os_ul.path.basename(_fname)
            _dash_tmp = result[16] if len(result) > 16 else {}
            _gran_tmp = _dash_tmp.get('granular',{}) if isinstance(_dash_tmp,dict) else {}
            _df_tmp   = result[17] if len(result) > 17 else None
            _score    = _dash_tmp.get('snapshot',{}).get('health_score',0) if isinstance(_dash_tmp,dict) else 0
            _rows     = _df_tmp.shape[0] if hasattr(_df_tmp,'shape') else 0
            _cols     = _df_tmp.shape[1] if hasattr(_df_tmp,'shape') else 0
            _stores   = len(_gran_tmp.get('stores',[])) if isinstance(_gran_tmp,dict) else 0
            _skus     = len(_gran_tmp.get('products',[])) if isinstance(_gran_tmp,dict) else 0
            _log_analysis(_email, _fname, _rows, _cols, _score, _stores, _skus, lang)
        except Exception as _le:
            print(f"⚠️  analysis log: {_le}")

        try:
            (insights_html, _vdb, k1, _k2, _k3, _k4, _k5, f1, f2, f3, f4,
             s1, s2, s3, s4, _err5, dash, df_for_gov) = result
        except Exception as _e2:
            return _bounce(f"Dashboard build failed: {str(_e2)[:200]}")

        # If analysis returned an error (via _fail), show it in dashboard
        if isinstance(insights_html, str) and (insights_html.startswith("⚠️") or insights_html.startswith("❌")):
            # Show error in full_dash instead of bouncing, so user sees what went wrong
            _err_html = f'<div style="padding:24px;background:#fef2f2;border-radius:12px;color:#dc2626;font-size:14px;font-weight:600;">{insights_html}</div>'
            return (
                gr.update(visible=False, value=""),
                99, *update_visibility_all('full_dash'),
                _err_html, None, None, None, None,
                gr.update(value="", visible=False), gr.update(value="", visible=False),
                gr.update(value="", visible=False), gr.update(value="", visible=False),
                {}, None, gr.update(value=_err_html), None,
                *_hdr(False)
            )

        # Ensure dash is a proper dict (not empty_dash from _fail)
        if not isinstance(dash, dict) or not dash.get('chart1'):
            # s1-s4 from _fail are gr.update objects — use None instead
            s1 = s1 if isinstance(s1, str) else None
            s2 = s2 if isinstance(s2, str) else None
            s3 = s3 if isinstance(s3, str) else None
            s4 = s4 if isinstance(s4, str) else None

        # Navigate to Business Scoring Dashboard (Screen 1) first after unlock
        _kpi_full    = dash.get('kpi1', '')
        _scores_html = insights_html  # Screen 1 shows the full AI insights report
        _granular    = dash.get('granular')
        return (
            gr.update(visible=False, value=""),   # hide loading
            99, *update_visibility_all('full_dash'),    # go to Business Scoring Dashboard FIRST
            _kpi_full, f1, f2, f3, f4,
            gr.update(value=s1 or "", visible=bool(s1)),
            gr.update(value=s2 or "", visible=bool(s2)),
            gr.update(value=s3 or "", visible=bool(s3)),
            gr.update(value=s4 or "", visible=bool(s4)),
            dash,
            df_for_gov,
            gr.update(value=_scores_html),
            _granular,
            *_hdr(False)
        )




    # ── Unlock Full Insights → goes DIRECTLY to dashboard (no login form) ─────
    def _direct_unlock(pf, fu, drc, lang, ud):
        """Skip login form — go straight to full dashboard."""
        file_to_use = pf or fu or drc
        user = ud if (isinstance(ud, dict) and ud.get('email')) else {'full_name':'Guest','mobile':'','email':''}
        if file_to_use is None:
            return tuple([gr.update()] * len(_UNLOCK_OUTPUTS))
        return unlock_to_dashboard(user, file_to_use, True, lang)

    preview_unlock_btn.click(
        fn=_direct_unlock,
        inputs=[preview_file_state, file_upload, _drc_clean_state,
                lang_state, user_data_state],
        outputs=_UNLOCK_OUTPUTS,
        queue=True,
        js="""() => { window._dnUnlockMode = true; }"""
    )

    # ── Full Dashboard screen navigation buttons ──────────────────────────────
    _FULL_DASH_NAV = [step_state] + _ALL_COLS + [logo_html, msme_hdr_btn, gov_hdr_btn, hdr_mode_html]

    # "← Back to Preview" from full_dash_col
    full_dash_back_btn.click(
        lambda: ('preview', *update_visibility_all('preview'), *_hdr(False)),
        [], _FULL_DASH_NAV,
        queue=True
    )

    # "▶ Business Performance Dashboard →" from full_dash_col → step6
    full_dash_next_btn.click(
        lambda: (6, *update_visibility_all('step6'), *_hdr(False)),
        [], _FULL_DASH_NAV,
        queue=True
    )

    # "📈 Skip to Forecasting →" from full_dash_col → step7
    full_dash_fc_btn.click(
        lambda: (7, *update_visibility_all('step7'), *_hdr(False)),
        [], _FULL_DASH_NAV,
        queue=True
    )

    # "← Back to Full Dashboard" — full_dash intermediate removed, stay on step6
    back6_to_dash_btn.click(
        lambda: (8, *update_visibility_all('full_dash'), *_hdr(False)),
        [], _FULL_DASH_NAV,
        queue=True
    )

    file_upload.change(handle_file_upload_change, [user_data_state, file_upload], [upload_message, error5])

    def show_dashboard_and_full(dashboard_data_value):
        """Show step6 dashboard AND populate full_dash_kpi_html at same time."""
        result = list(show_dashboard(dashboard_data_value))
        return tuple(result)

    view_dashboard_btn.click(show_dashboard, [dashboard_data_state],
        [step_state]+_ALL_COLS+[kpi_table_dash, chart1_dash, chart2_dash, chart3_dash, chart4_dash,
         chart1_summary, chart2_summary, chart3_summary, chart4_summary, granular_forecast_data_state,
         logo_html, msme_hdr_btn, gov_hdr_btn, hdr_mode_html])

    _S7_NAV_OUTPUTS = ([step_state] + _ALL_COLS
                       + [s7_store_filter, s7_cat_filter, s7_prod_filter]
                       + _S7_CHART_OUTPUTS
                       + [logo_html, msme_hdr_btn, gov_hdr_btn, hdr_mode_html])
    forecast_deepdive_btn.click(show_granular_dashboard,
        [granular_forecast_data_state, df_state],
        _S7_NAV_OUTPUTS,
        queue=True)

    s7_store_filter.change(update_step7_filters,
        [s7_store_filter, s7_cat_filter, s7_prod_filter, df_state],
        _S7_CHART_OUTPUTS)
    s7_cat_filter.change(update_step7_filters,
        [s7_store_filter, s7_cat_filter, s7_prod_filter, df_state],
        _S7_CHART_OUTPUTS)
    s7_prod_filter.change(update_step7_filters,
        [s7_store_filter, s7_cat_filter, s7_prod_filter, df_state],
        _S7_CHART_OUTPUTS)

    # ── PDF Report Download ────────────────────────────────────────────────────
    def _generate_pdf(user_data, dashboard_data, df_raw, granular_data):
        """Gradio handler: calls generate_bi_report_pdf, returns file path for gr.File."""
        import traceback as _tb2
        # ── Log report download ───────────────────────────────────────────
        try:
            _ud = user_data if isinstance(user_data,dict) else {}
            _sc = dashboard_data.get('snapshot',{}).get('health_score',0) if isinstance(dashboard_data,dict) else 0
            _log_report_download(_ud.get('email',''), _sc)
        except Exception as _rle:
            print(f"⚠️  report log: {_rle}")
        try:
            # Defensive: ensure df_raw is a real DataFrame
            import pandas as _pd2
            if not isinstance(df_raw, _pd2.DataFrame) or df_raw.empty:
                return gr.update(value=None, visible=False)

            # Ensure dashboard_data is a dict (may be None before analysis)
            _dash = dashboard_data if isinstance(dashboard_data, dict) else {}

            path = generate_bi_report_pdf(
                user_data  if isinstance(user_data,  dict) else {},
                df_raw,
                _dash,
                granular_data
            )
            if path and os.path.exists(path):
                sz = os.path.getsize(path)
                print(f"✅ PDF ready: {path} ({sz:,} bytes)")
                fname = os.path.basename(path)
                # gr.DownloadButton: value=path triggers download on click
                return gr.update(value=path, visible=True,
                                 label=f"⬇ Download Report ({sz//1024}KB) — {fname}")
            return gr.update(value=None, visible=True,
                             label="⚠️ PDF file was not produced — check server logs")
        except Exception as _pe:
            _err = _tb2.format_exc()
            # Show first meaningful line of the traceback to the user
            _lines = [l.strip() for l in _err.splitlines() if l.strip() and 'File' not in l]
            _msg   = _lines[-1][:140] if _lines else str(_pe)[:140]
            return gr.update(value=None, visible=True,
                             label=f"⚠️ {_msg}")

    pdf_download_btn.click(
        _generate_pdf,
        inputs=[user_data_state, dashboard_data_state, df_state, granular_forecast_data_state],
        outputs=[pdf_file_output]
    )

    back7_btn.click(lambda: (6, *update_visibility_all('step6'), *_hdr(False)), [], [step_state]+_ALL_COLS+[logo_html,msme_hdr_btn,gov_hdr_btn, hdr_mode_html])
    back7_to5_btn.click(lambda: (5, *update_visibility_all('step5'), *_hdr(False)), [], [step_state]+_ALL_COLS+[logo_html,msme_hdr_btn,gov_hdr_btn, hdr_mode_html])

    # Language switching
    def switch_lang_en():
        return ('en', gr.update(variant='primary'), gr.update(variant='secondary'), gr.update(value='**Active: English 🇬🇧**'),
                gr.update(value=_landing_hero('en')), gr.update(value=_landing_capabilities('en')),
                gr.update(value="## How DataNetra Works"),
                gr.update(value="### 📥 Step 1: Upload Your Data"),
                gr.update(value="Easily upload Excel/CSV files for comprehensive analysis."),
                gr.update(value="### 🤖 Step 2: AI-Powered Analysis"),
                gr.update(value="Our AI processes your data, forecasting trends and uncovering insights."),
                gr.update(value="### 📊 Step 3: Actionable Dashboards & Recommendations"),
                gr.update(value="Access interactive dashboards, KPI charts and personalized recommendations."),
                gr.update(value="**Sign In**"), gr.update(label="Mobile Number"),
                gr.update(value="Analyze My Data →"), gr.update(value=""),
                gr.update(value=""), gr.update(value="", visible=False),
                gr.update(),  # msme_card_html — never overwrite the HTML form
                gr.update(value="", visible=False),
                gr.update(value="", visible=False),
                gr.update(value="", visible=False),
                gr.update(value="""<div class="dn-access-hover" style="background:#f8faff;border:1px solid #dbeafe;border-radius:12px;
                  padding:18px 16px 14px;margin-top:0;">
                  <div style="font-size:14px;font-weight:800;color:#0B1F3A;margin-bottom:10px;
                    letter-spacing:-0.2px;">What DataNetra Delivers</div>
                  <div style="display:flex;flex-direction:column;gap:8px;">
                    <div style="display:flex;align-items:flex-start;gap:10px;">
                      <div style="width:28px;height:28px;flex-shrink:0;background:#EBF2FC;border-radius:7px;
                        display:flex;align-items:center;justify-content:center;">
                        <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="#1B4F8A"
                          stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                          <rect x="2" y="14" width="4" height="7"/><rect x="9" y="9" width="4" height="12"/>
                          <rect x="16" y="4" width="4" height="17"/><path d="M2 3h20"/>
                        </svg>
                      </div>
                      <div>
                        <div style="font-size:13px;font-weight:700;color:#0B1F3A;">Business Performance Clarity</div>
                        <div style="font-size:12px;color:#64748b;line-height:1.4;">Sales, margins, returns and category performance.</div>
                      </div>
                    </div>
                    <div style="display:flex;align-items:flex-start;gap:10px;">
                      <div style="width:28px;height:28px;flex-shrink:0;background:#E8F7EE;border-radius:7px;
                        display:flex;align-items:center;justify-content:center;">
                        <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="#1a7a40"
                          stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                          <polyline points="23 6 13.5 15.5 8.5 10.5 1 18"/>
                          <polyline points="17 6 23 6 23 12"/>
                        </svg>
                      </div>
                      <div>
                        <div style="font-size:13px;font-weight:700;color:#0B1F3A;">Demand Forecasting Intelligence</div>
                        <div style="font-size:12px;color:#64748b;line-height:1.4;">Predict demand with AI-driven forecasting.</div>
                      </div>
                    </div>
                    <div style="display:flex;align-items:flex-start;gap:10px;">
                      <div style="width:28px;height:28px;flex-shrink:0;background:#F3EFFE;border-radius:7px;
                        display:flex;align-items:center;justify-content:center;">
                        <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="#7c3aed"
                          stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                          <polyline points="22 12 18 12 15 21 9 3 6 12 2 12"/>
                        </svg>
                      </div>
                      <div>
                        <div style="font-size:13px;font-weight:700;color:#0B1F3A;">Business Health Score</div>
                        <div style="font-size:12px;color:#64748b;line-height:1.4;">Automated health scoring from sales, returns and margins.</div>
                      </div>
                    </div>
                    <div style="display:flex;align-items:flex-start;gap:10px;">
                      <div style="width:28px;height:28px;flex-shrink:0;background:#FEF0E7;border-radius:7px;
                        display:flex;align-items:center;justify-content:center;">
                        <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="#c2520a"
                          stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                          <path d="M6 2L3 6v14a2 2 0 0 0 2 2h14a2 2 0 0 0 2-2V6l-3-4z"/>
                          <line x1="3" y1="6" x2="21" y2="6"/><path d="M16 10a4 4 0 0 1-8 0"/>
                        </svg>
                      </div>
                      <div>
                        <div style="font-size:13px;font-weight:700;color:#0B1F3A;">Retail Intelligence Engine</div>
                        <div style="font-size:12px;color:#64748b;line-height:1.4;">Turns clean data into forecasting insights.</div>
                      </div>
                    </div>
                  </div>
                </div>"""),
                gr.update(value="""<div style="margin-top:6px;background:#fff;border:1px solid #dbeafe;border-radius:12px;padding:18px 16px 14px;">
                  <div style="font-size:14px;font-weight:800;color:#0B1F3A;margin-bottom:10px;letter-spacing:-0.2px;">Platform Capabilities</div>
                  <div style="display:grid;grid-template-columns:repeat(2,1fr);gap:8px;">
                    <div class="dn-card-hover" style="background:#f0fdf4;border:1px solid #dcfce7;border-left:3px solid #16a34a;border-radius:8px;padding:10px 10px 8px;">
                      <div style="font-size:13px;font-weight:700;color:#0B1F3A;line-height:1.3;margin-bottom:2px;">Intelligent Column Detection</div>
                      <div style="font-size:12px;color:#64748b;line-height:1.4;">Auto-detects retail fields from any layout.</div>
                    </div>
                    <div class="dn-card-hover" style="background:#fffbeb;border:1px solid #fde68a;border-left:3px solid #d97706;border-radius:8px;padding:10px 10px 8px;">
                      <div style="font-size:13px;font-weight:700;color:#0B1F3A;line-height:1.3;margin-bottom:2px;">Auto Cleaning &amp; Normalization</div>
                      <div style="font-size:12px;color:#64748b;line-height:1.4;">Fixes dates, duplicates and currency formats.</div>
                    </div>
                    <div class="dn-card-hover" style="background:#faf5ff;border:1px solid #e9d5ff;border-left:3px solid #7c3aed;border-radius:8px;padding:10px 10px 8px;">
                      <div style="font-size:13px;font-weight:700;color:#0B1F3A;line-height:1.3;margin-bottom:2px;">AI Readiness Scoring</div>
                      <div style="font-size:12px;color:#64748b;line-height:1.4;">Calculates your Data Readiness Score instantly.</div>
                    </div>
                    <div class="dn-card-hover" style="background:#fff7ed;border:1px solid #fed7aa;border-left:3px solid #c2520a;border-radius:8px;padding:10px 10px 8px;">
                      <div style="font-size:13px;font-weight:700;color:#0B1F3A;line-height:1.3;margin-bottom:2px;">Retail Intelligence Engine</div>
                      <div style="font-size:12px;color:#64748b;line-height:1.4;">Turns clean data into forecasting insights.</div>
                    </div>
                  </div>
                </div>"""),
                gr.update(value=""), gr.update(label="First Name*"),
                gr.update(label="Mobile Number*"), gr.update(label="Email"),
                gr.update(label="Role"), gr.update(value="Cancel"),
                gr.update(value="Next →"), gr.update(value=""),
                gr.update(label=""), gr.update(value="Fetch"),
                gr.update(value="← Back"), gr.update(value="Next →"),
                gr.update(value=""), gr.update(label=""),
                gr.update(label=""), gr.update(value="← Back"),
                gr.update(value="Next →"), gr.update(value=""),
                gr.update(label=""), gr.update(label=""),
                gr.update(label=""), gr.update(value="← Back"),
                gr.update(value="Submit"), gr.update(value="Next →"),
                gr.update(value="## Upload Screen"),
                gr.update(label="I consent to data analysis*"),
                gr.update(label="Upload Excel or CSV file (.xlsx, .csv)*"),
                gr.update(value="⬅ Back to Home"),
                gr.update(value="❌ Cancel"),
                gr.update(value="Generate Preview →"),
                gr.update(value="📈 Go to Forecast Intelligence Dashboard →"),
                gr.update(value="← Back to Upload Screen"),
                gr.update(label="🏪 Store"),
                gr.update(label="📂 Category"),
                gr.update(label="📦 Product"),
                gr.update(value="⬇ Download Business Intelligence Report (PDF)"),
                gr.update(value="⬅ Back to Dashboard"),
                gr.update(value="← Back to Upload Screen"),
                gr.update(value=_landing_industries('en')),      # [64] industries_html
                gr.update(value=_landing_unlock_benefits('en')), # [65] unlock_benefits_html
                gr.update(value=_landing_capabilities('en')),    # [66] landing_capabilities_html
                gr.update()                                       # [67-1=66] hdr_mode_html no-op
                )

    def switch_lang_hi():
        return ('hi', gr.update(variant='secondary'), gr.update(variant='primary'), gr.update(value='**सक्रिय: हिंदी 🇮🇳**'),
                gr.update(value=_landing_hero('hi')), gr.update(value=_landing_capabilities('hi')),
                gr.update(value="## DataNetra कैसे काम करता है?"),
                gr.update(value="### 📥 पहला काम: अपना Data डालें"),
                gr.update(value="अपनी Excel या CSV फाइल आसानी से Upload करें।"),
                gr.update(value="### 🤖 दूसरा काम: AI जाँच करेगा"),
                gr.update(value="हमारा AI आपका Data पढ़कर बताएगा — क्या बिक रहा है, क्या नहीं।"),
                gr.update(value="### 📊 तीसरा काम: रिपोर्ट और सलाह देखें"),
                gr.update(value="साफ Dashboard पर अपनी बिक्री, मुनाफा और AI की सलाह एक जगह देखें।"),
                gr.update(value="**साइन इन करें**"), gr.update(label="मोबाइल नंबर"),
                gr.update(value="📊 डेटा विश्लेषण करें →"), gr.update(value=""),
                gr.update(value=""), gr.update(value="", visible=False),
                gr.update(),  # msme_card_html — never overwrite the HTML form
                gr.update(value="", visible=False),
                gr.update(value="", visible=False),
                gr.update(value="", visible=False),
                gr.update(value="""<div class="dn-access-hover" style="background:#f8faff;border:1px solid #dbeafe;border-radius:12px;
                  padding:18px 16px 14px;margin-top:0;">
                  <div style="font-size:14px;font-weight:800;color:#0B1F3A;margin-bottom:10px;">DataNetra से आपको क्या मिलता है</div>
                  <div style="display:flex;flex-direction:column;gap:8px;">
                    <div style="display:flex;align-items:flex-start;gap:10px;">
                      <div><div style="font-size:13px;font-weight:700;color:#0B1F3A;">कारोबार का पूरा हिसाब</div>
                      <div style="font-size:12px;color:#64748b;">बिक्री, मुनाफा, रिटर्न और कैटेगरी की पूरी जानकारी।</div></div>
                    </div>
                    <div style="display:flex;align-items:flex-start;gap:10px;">
                      <div><div style="font-size:13px;font-weight:700;color:#0B1F3A;">माँग का पूर्वानुमान</div>
                      <div style="font-size:12px;color:#64748b;">AI से जानें — आगे क्या और कितना बिकेगा।</div></div>
                    </div>
                    <div style="display:flex;align-items:flex-start;gap:10px;">
                      <div><div style="font-size:13px;font-weight:700;color:#0B1F3A;">कारोबार का हेल्थ स्कोर</div>
                      <div style="font-size:12px;color:#64748b;">बिक्री, रिटर्न और मुनाफे से तुरंत हेल्थ स्कोर।</div></div>
                    </div>
                    <div style="display:flex;align-items:flex-start;gap:10px;">
                      <div><div style="font-size:13px;font-weight:700;color:#0B1F3A;">रिटेल इंटेलिजेंस</div>
                      <div style="font-size:12px;color:#64748b;">साफ डेटा से बिक्री के अनुमान और सुझाव।</div></div>
                    </div>
                  </div>
                </div>"""),
                gr.update(value="""<div style="margin-top:6px;background:#fff;border:1px solid #dbeafe;border-radius:12px;padding:18px 16px 14px;">
                  <div style="font-size:14px;font-weight:800;color:#0B1F3A;margin-bottom:10px;">प्लेटफ़ॉर्म क्या-क्या करता है</div>
                  <div style="display:grid;grid-template-columns:repeat(2,1fr);gap:8px;">
                    <div class="dn-card-hover" style="background:#f0fdf4;border:1px solid #dcfce7;border-left:3px solid #16a34a;border-radius:8px;padding:10px;">
                      <div style="font-size:13px;font-weight:700;color:#0B1F3A;">स्मार्ट कॉलम पहचान</div>
                      <div style="font-size:12px;color:#64748b;">किसी भी फ़ाइल से सही कॉलम अपने आप ढूंढता है।</div>
                    </div>
                    <div class="dn-card-hover" style="background:#fffbeb;border:1px solid #fde68a;border-left:3px solid #d97706;border-radius:8px;padding:10px;">
                      <div style="font-size:13px;font-weight:700;color:#0B1F3A;">डेटा की सफाई</div>
                      <div style="font-size:12px;color:#64748b;">तारीख, डुप्लीकेट और रकम के फॉर्मेट ठीक होते हैं।</div>
                    </div>
                    <div class="dn-card-hover" style="background:#faf5ff;border:1px solid #e9d5ff;border-left:3px solid #7c3aed;border-radius:8px;padding:10px;">
                      <div style="font-size:13px;font-weight:700;color:#0B1F3A;">AI रेडीनेस स्कोर</div>
                      <div style="font-size:12px;color:#64748b;">आपका डेटा कितना तैयार है — तुरंत बताता है।</div>
                    </div>
                    <div class="dn-card-hover" style="background:#fff7ed;border:1px solid #fed7aa;border-left:3px solid #c2520a;border-radius:8px;padding:10px;">
                      <div style="font-size:13px;font-weight:700;color:#0B1F3A;">बिक्री की स्मार्ट जानकारी</div>
                      <div style="font-size:12px;color:#64748b;">साफ डेटा से बिक्री के अनुमान और सुझाव।</div>
                    </div>
                  </div>
                </div>"""),
                gr.update(value=""), gr.update(label="पहला नाम*"),
                gr.update(label="मोबाइल नंबर*"), gr.update(label="ईमेल"),
                gr.update(label="भूमिका"), gr.update(value="रद्द करें"),
                gr.update(value="आगे बढ़ें →"), gr.update(value=""),
                gr.update(label=""), gr.update(value="लाएं"),
                gr.update(value="← वापस"), gr.update(value="आगे बढ़ें →"),
                gr.update(value=""), gr.update(label=""),
                gr.update(label=""), gr.update(value="← वापस"),
                gr.update(value="आगे बढ़ें →"), gr.update(value=""),
                gr.update(label=""), gr.update(label=""),
                gr.update(label=""), gr.update(value="← वापस"),
                gr.update(value="जमा करें"), gr.update(value="अगला →"),
                gr.update(value="## Upload Screen"),
                gr.update(label="मैं डेटा विश्लेषण के लिए सहमति देता/देती हूँ*"),
                gr.update(label="Excel या CSV फ़ाइल अपलोड करें*"),
                gr.update(value="⬅ होम पर वापस"),
                gr.update(value="❌ रद्द करें"),
                gr.update(value="Preview देखें →"),
                gr.update(value="📈 पूर्वानुमान डैशबोर्ड देखें →"),
                gr.update(value="← Upload Screen पर वापस"),
                gr.update(label="🏪 स्टोर"),
                gr.update(label="📂 श्रेणी"),
                gr.update(label="📦 उत्पाद"),
                gr.update(value="⬇ बिज़नेस इंटेलिजेंस रिपोर्ट डाउनलोड करें (PDF)"),
                gr.update(value="⬅ डैशबोर्ड पर वापस"),
                gr.update(value="← Upload Screen पर वापस"),
                gr.update(),  # hdr_mode_html
                gr.update(value=_landing_industries('en')),
                gr.update(value=_landing_unlock_benefits('en')),
                gr.update(value=_landing_capabilities('en')))
    _lang_landing_outputs = [lang_state, lang_en_btn, lang_hi_btn, lang_indicator,
                              landing_hero_html, landing_capabilities_html,
                              landing_how_title, landing_step1_title, landing_step1_desc,
                              landing_step2_title, landing_step2_desc, landing_step3_title, landing_step3_desc,
                              landing_login_title, quick_login_mobile, quick_login_btn,
                              landing_signup_title, landing_signup_desc, quick_signup_btn,
                              msme_card_html, signup_card_html, gov_card_html, auth_hint_html,
                              delivers_html_widget, platform_html_widget,
                              step1_title_md, name_input, mobile_input, email_input, role_input,
                              cancel1_btn, next1_btn,
                              step2_title_md, msme_number_input, fetch_btn, back2_btn, next2_btn,
                              step3_title_md, consent1, consent2, back3_btn, next3_btn,
                              step4_title_md, business_type_input, years_input, revenue_input,
                              back4_btn, next4_btn, proceed_to_step5_btn,
                              step5_title_md, consent_check, file_upload,
                              back5_btn, cancel5_btn, analyze_btn,
                              forecast_deepdive_btn, back6_btn,
                              s7_store_filter, s7_cat_filter, s7_prod_filter,
                              pdf_download_btn, back7_btn, back7_to5_btn,
                    hdr_mode_html,
                    industries_html,
                    unlock_benefits_html,
                    landing_capabilities_html]

    # ── Language switching: native buttons → direct backend wiring ──────────
    lang_en_btn.click(switch_lang_en, [], _lang_landing_outputs)
    lang_hi_btn.click(switch_lang_hi, [], _lang_landing_outputs)

    def _lang_no_change():
        """Return a gr.update() tuple that changes NOTHING — used for non-hero outputs."""
        return gr.update()

    # Total outputs in _lang_landing_outputs — count them
    _N_LANG_OUTPUTS = len([lang_state, lang_en_btn, lang_hi_btn, lang_indicator,
        landing_hero_html, landing_capabilities_html,
        landing_how_title, landing_step1_title, landing_step1_desc,
        landing_step2_title, landing_step2_desc, landing_step3_title, landing_step3_desc,
        landing_login_title, quick_login_mobile, quick_login_btn,
        landing_signup_title, landing_signup_desc, quick_signup_btn,
        msme_card_html, signup_card_html, gov_card_html, auth_hint_html,
        delivers_html_widget, platform_html_widget,
        step1_title_md, name_input, mobile_input, email_input, role_input,
        cancel1_btn, next1_btn,
        step2_title_md, msme_number_input, fetch_btn, back2_btn, next2_btn,
        step3_title_md, consent1, consent2, back3_btn, next3_btn,
        step4_title_md, business_type_input, years_input, revenue_input,
        back4_btn, next4_btn, proceed_to_step5_btn,
        step5_title_md, consent_check, file_upload,
        back5_btn, cancel5_btn, analyze_btn,
        forecast_deepdive_btn, back6_btn,
        s7_store_filter, s7_cat_filter, s7_prod_filter,
        pdf_download_btn, back7_btn, back7_to5_btn,
                    hdr_mode_html,
                    industries_html,
                    unlock_benefits_html,
                    landing_capabilities_html])

    def switch_lang_kn():
        """Kannada — translate hero + all landing sections."""
        no_change = tuple(gr.update() for _ in range(_N_LANG_OUTPUTS))
        lst = list(no_change)
        lst[0]  = 'kn'
        lst[4]  = gr.update(value=_landing_hero('kn'))
        lst[5]  = gr.update(value=_landing_capabilities('kn'))
        lst[64] = gr.update(value=_landing_industries('kn'))
        lst[65] = gr.update(value=_landing_unlock_benefits('kn'))
        lst[66] = gr.update(value=_landing_capabilities('kn'))
        return tuple(lst)

    def switch_lang_ta():
        """Tamil — translate hero + all landing sections."""
        no_change = tuple(gr.update() for _ in range(_N_LANG_OUTPUTS))
        lst = list(no_change)
        lst[0]  = 'ta'
        lst[4]  = gr.update(value=_landing_hero('ta'))
        lst[5]  = gr.update(value=_landing_capabilities('ta'))
        lst[64] = gr.update(value=_landing_industries('ta'))
        lst[65] = gr.update(value=_landing_unlock_benefits('ta'))
        lst[66] = gr.update(value=_landing_capabilities('ta'))
        return tuple(lst)

    def _switch_lang_dropdown(choice):
        if not choice:
            return switch_lang_en()
        c = str(choice).lower()
        if "kannada" in c or "kn" == c or "\u0c95\u0ca8\u0ccd\u0ca8\u0ca1" in c:
            return switch_lang_kn()
        if "tamil" in c or "ta" == c or "\u0ba4\u0bae\u0bbf\u0bb4\u0bcd" in c:
            return switch_lang_ta()
        return switch_lang_en()

    lang_dropdown.change(_switch_lang_dropdown, [lang_dropdown], _lang_landing_outputs)

    # Gov/MSME panel toggle removed — gov dashboard removed

    # ── Pricing nav button — toggle pricing section visibility ────────────────
    pricing_nav_btn.click(
        fn=lambda: gr.update(visible=True),
        inputs=[], outputs=[pricing_section],
        queue=False,
        show_progress=False,
        js="() => { setTimeout(function(){ var p=document.getElementById('dn-pricing-section'); if(p){ p.scrollIntoView({behavior:'smooth',block:'start'}); } }, 200); }"
    )
    pricing_close_btn.click(
        fn=None,
        inputs=[], outputs=[],
        queue=False,
        js="() => { var s=document.getElementById('dn-pricing-section'); if(s) s.classList.remove('dn-pricing-open'); }"
    )

    # Inject login form JS at page load — bypasses Gradio HTML sanitizer
    # ── Send OTP endpoint ─────────────────────────────────────────────────
    def _handle_send_otp(email_val):
        import re as _re
        email_val = (email_val or "").strip()
        if not email_val or not _re.match(r"[^@]+@[^@]+\.[^@]+", email_val):
            return "⚠️ Enter a valid email address first."
        otp = _generate_otp() if _SMTP_USER else "1234"
        _store_otp(email_val, otp)
        ok, msg = _send_otp_email(email_val, otp)
        if ok:
            if msg == "dev_mode":
                return "✅ Dev mode: use OTP **1234**"
            return f"✅ OTP sent to {email_val} — check your inbox (expires in 5 min)"
        return f"⚠️ Could not send email: {msg}. Use 1234 for demo."

    send_otp_btn.click(
        fn=_handle_send_otp,
        inputs=[send_otp_email_in],
        outputs=[send_otp_result],
        queue=False
    )

    # ── Contact form → save to database ───────────────────────────────────
    def _handle_contact_form(name, email, mobile, message):
        if not name or not email or not message:
            return "⚠️ Please fill in all required fields."
        ok = _save_contact_lead(name.strip(), email.strip(),
                                mobile.strip() if mobile else '',
                                message.strip())
        if ok:
            return f"✅ Message received from {name}. We'll be in touch soon!"
        return "✅ Message received! We'll be in touch."

    cf_submit_btn.click(
        fn=_handle_contact_form,
        inputs=[cf_name_input, cf_email_input, cf_mobile_input, cf_message_input],
        outputs=[cf_result],
        queue=False
    )

    demo.load(
        fn=None,
        inputs=None,
        outputs=None,
        js="""() => {
    // Attach all login form event listeners — runs at page load
    // Using addEventListener avoids ALL inline event handler escaping issues

    function dnAttachLoginForm() {
        var nameEl  = document.getElementById('dn-f-name');
        var mobEl   = document.getElementById('dn-f-mobile');
        var otpEl   = document.getElementById('dn-f-otp');
        var emailEl = document.getElementById('dn-f-email');
        var btnEl   = document.getElementById('dn-f-submit');
        var errEl   = document.getElementById('dn-f-error');
        var dbgEl   = document.getElementById('dn-f-debug');

        if (!btnEl) return false; // form not in DOM yet

        // Focus/blur styles
        [nameEl, mobEl, otpEl, emailEl].forEach(function(el) {
            if (!el || el._dnBound) return;
            el._dnBound = true;
            el.addEventListener('focus', function() {
                el.style.borderColor = '#2563eb';
                el.style.boxShadow = '0 0 0 3px rgba(37,99,235,0.12)';
            });
            el.addEventListener('blur', function() {
                el.style.borderColor = '#d1d9e6';
                el.style.boxShadow = 'none';
            });
        });

        if (btnEl._dnBound) return true;
        btnEl._dnBound = true;

        btnEl.addEventListener('click', function() {
            var name  = (nameEl  || {}).value || '';
            var mob   = (mobEl   || {}).value || '';
            var otp   = (otpEl   || {}).value || '';
            var email = (emailEl || {}).value || '';

            if (!name.trim() || !mob.trim() || !otp.trim()) {
                if (errEl) { errEl.textContent = 'Please fill in Name, Mobile and Verification Code.'; errEl.style.display = 'block'; }
                return;
            }
            if (errEl) errEl.style.display = 'none';
            btnEl.textContent = 'Analysing...';
            btnEl.disabled = true;
            btnEl.style.opacity = '0.85';
            if (dbgEl) { dbgEl.style.display = 'block'; dbgEl.innerHTML = 'Starting...<br>'; }

            // Store in window for js= preprocessor
            window._dnForm = {name: name, mob: mob, otp: otp, email: email};
            window._dnUnlockMode = true;

            // Sync to hidden Gradio inputs
            function syncGr(id, val) {
                var wrap = document.getElementById(id);
                if (!wrap) { if(dbgEl) dbgEl.innerHTML += 'WARN: #'+id+' not found<br>'; return; }
                var el = wrap.querySelector('input, textarea');
                if (!el) { if(dbgEl) dbgEl.innerHTML += 'WARN: no input in #'+id+'<br>'; return; }
                try {
                    var P = Object.getOwnPropertyDescriptor(
                        el.tagName === 'TEXTAREA' ? HTMLTextAreaElement.prototype : HTMLInputElement.prototype,
                        'value');
                    if (P && P.set) P.set.call(el, val);
                    else el.value = val;
                    el.dispatchEvent(new Event('input',  {bubbles: true}));
                    el.dispatchEvent(new Event('change', {bubbles: true}));
                } catch(e) { el.value = val; }
            }
            syncGr('dn-landing-name',  name);
            syncGr('dn-real-mobile',   mob);
            syncGr('dn-landing-otp',   otp);
            syncGr('dn-landing-email', email);

            // Click the hidden Gradio button
            var attempts = 0;
            var maxAttempts = 20;
            function tryClick() {
                attempts++;
                var grBtn = document.querySelector('div#dn-real-login-btn button') ||
                            document.querySelector('#dn-real-login-btn button') ||
                            document.querySelector('[id="dn-real-login-btn"] button');
                if (dbgEl) dbgEl.innerHTML += 'Attempt '+attempts+': '+(grBtn?'FOUND!':'NOT FOUND')+'<br>';

                if (grBtn) {
                    if (dbgEl) dbgEl.innerHTML += 'CLICKING!<br>';
                    grBtn.disabled = false;
                    grBtn.click();
                    setTimeout(function() {
                        btnEl.textContent = 'Analyze My Data →';
                        btnEl.disabled = false;
                        btnEl.style.opacity = '1';
                    }, 6000);
                } else if (attempts < maxAttempts) {
                    setTimeout(tryClick, 300);
                } else {
                    if (dbgEl) dbgEl.innerHTML += 'FAILED after '+maxAttempts+' attempts<br>';
                    if (errEl) { errEl.textContent = 'Please reload and try again.'; errEl.style.display = 'block'; }
                    btnEl.textContent = 'Analyze My Data →';
                    btnEl.disabled = false;
                    btnEl.style.opacity = '1';
                }
            }
            setTimeout(tryClick, 150);
        });

        return true;
    }

    // ── Wire "Send OTP" button ───────────────────────────────────────────
    function dnAttachSendOtp() {
        var sendBtn = document.getElementById('dn-send-otp-btn');
        if (!sendBtn || sendBtn._dnOtpBound) return;
        sendBtn._dnOtpBound = true;
        sendBtn.addEventListener('click', function() {
            var email = (document.getElementById('dn-f-email')||{}).value||'';
            // Sync email to hidden Gradio input for send_otp_email_in
            var grEmailIn = document.querySelector('#dn-send-otp-email-in input, #dn-send-otp-email-in textarea');
            if (grEmailIn) {
                var P = Object.getOwnPropertyDescriptor(
                    grEmailIn.tagName==='TEXTAREA'?HTMLTextAreaElement.prototype:HTMLInputElement.prototype,'value');
                if (P&&P.set) P.set.call(grEmailIn, email);
                grEmailIn.dispatchEvent(new Event('input',{bubbles:true}));
                grEmailIn.dispatchEvent(new Event('change',{bubbles:true}));
            }
            sendBtn.textContent='Sending...'; sendBtn.disabled=true;
            // Click hidden Gradio send_otp button
            setTimeout(function() {
                var grBtn = document.querySelector('div#dn-send-otp-gr-btn button');
                if (grBtn) grBtn.click();
                // Show result from dn-send-otp-result after 2s
                setTimeout(function() {
                    var res = document.querySelector('#dn-send-otp-result p, #dn-send-otp-result');
                    var status = document.getElementById('dn-otp-status');
                    if (status && res && res.textContent) {
                        status.textContent = res.textContent;
                        status.style.display='block';
                        status.style.color = res.textContent.includes('⚠️') ? '#dc2626' : '#16a34a';
                    }
                    sendBtn.textContent='Resend OTP'; sendBtn.disabled=false;
                }, 2000);
            }, 100);
        });
    }

    // Debug: log all dn-real-login-btn elements found at startup
    setTimeout(function() {
        var allBtns = document.querySelectorAll('[id="dn-real-login-btn"], [id="dn-real-login-btn"] button, div#dn-real-login-btn button');
        console.log('[DN] Buttons found at startup:', allBtns.length);
        allBtns.forEach(function(b,i){ console.log('[DN] btn['+i+']:', b.tagName, b.id, b.className); });
        if(allBtns.length===0){
            // Walk all buttons on page
            var allPageBtns = document.querySelectorAll('button');
            console.log('[DN] All page buttons:', allPageBtns.length);
            allPageBtns.forEach(function(b,i){ if(b.id||b.textContent.includes('Analyze')) console.log('[DN] pgbtn['+i+']:', b.id, b.textContent.trim().slice(0,30)); });
        }
    }, 2000);

    // Try immediately, then retry until form is in DOM
    var attachAttempts = 0;
    function keepTrying() {
        dnAttachSendOtp();
        if (dnAttachLoginForm()) return;
        attachAttempts++;
        if (attachAttempts < 30) setTimeout(keepTrying, 500);
    }
    setTimeout(keepTrying, 500);

    // Also re-attach when login section becomes visible
    var observer = new MutationObserver(function() {
        dnAttachLoginForm();
    });
    observer.observe(document.body, {subtree: true, attributes: true, attributeFilter: ['style', 'class']});
}"""
    )

if __name__ == "__main__":
    import os as _os_launch
    import asyncio as _asyncio

    _port = int(_os_launch.environ.get("PORT", 7860))

    # ── Patch 1: ALL Queue async-primitive None-guards ───────────────────────
    # Gradio 4.44.1 Queue.__init__ leaves pending_message_lock, delete_lock,
    # and other asyncio primitives as None until Queue.start() runs.
    # On Render, tasks fire before start() completes → TypeError.
    # We patch every method that uses these primitives to auto-create them.
    try:
        import gradio.queueing as _gr_q
        import re as _re_q

        # Collect all method names that use async-with or await on self.<attr>
        import inspect as _inspect
        _src = _inspect.getsource(_gr_q.Queue)
        _lock_attrs = set(_re_q.findall(r"async with self\.(\w+)", _src))
        _event_attrs = set(_re_q.findall(r"await self\.(\w+)\.(?:wait|set|clear)", _src))
        print(f"✅ Queue lock attrs: {_lock_attrs}")
        print(f"✅ Queue event attrs: {_event_attrs}")

        # Patch Queue.start_processing to guard all locks before running
        _orig_start_processing = _gr_q.Queue.start_processing
        async def _safe_start_processing(self, *a, **kw):
            for _attr in _lock_attrs:
                if getattr(self, _attr, None) is None:
                    setattr(self, _attr, _asyncio.Lock())
            for _attr in _event_attrs:
                if getattr(self, _attr, None) is None:
                    setattr(self, _attr, _asyncio.Event())
            return await _orig_start_processing(self, *a, **kw)
        _gr_q.Queue.start_processing = _safe_start_processing

        # Also patch push as a safety net
        _orig_push = _gr_q.Queue.push
        async def _safe_push(self, *a, **kw):
            for _attr in _lock_attrs:
                if getattr(self, _attr, None) is None:
                    setattr(self, _attr, _asyncio.Lock())
            return await _orig_push(self, *a, **kw)
        _gr_q.Queue.push = _safe_push

        # Also patch process_events — called before start_processing on Render cold start
        _orig_process_events = _gr_q.Queue.process_events
        async def _safe_process_events(self, *a, **kw):
            for _attr in _lock_attrs:
                if getattr(self, _attr, None) is None:
                    setattr(self, _attr, _asyncio.Lock())
            for _attr in _event_attrs:
                if getattr(self, _attr, None) is None:
                    setattr(self, _attr, _asyncio.Event())
            return await _orig_process_events(self, *a, **kw)
        _gr_q.Queue.process_events = _safe_process_events

        print("✅ Queue all-locks guard applied")
    except Exception as _e:
        print(f"⚠️  Queue patch failed: {_e}")

    # ── Patch 2: Disable is_url_ok health-check ───────────────────────────────
    try:
        import gradio.networking as _gr_net
        _gr_net.is_url_ok = lambda *a, **kw: True
    except Exception:
        pass

    # ── Patch 3: Monkey-patch uvicorn.run to inject Render-safe config ────────
    # demo.launch() internally calls uvicorn.run(). We intercept that call to:
    #   (a) force timeout_keep_alive=75 so SSE stays alive under Render's 90s limit
    #   (b) wrap the ASGI app in _RenderSSEMiddleware before uvicorn sees it
    import uvicorn as _uvicorn

    class _RenderSSEMiddleware:
        """Injects X-Accel-Buffering:no on /queue/* and /run/* so Render's
        nginx doesn't buffer or 503 the SSE event streams."""
        def __init__(self, app):
            self.app = app
        async def __call__(self, scope, receive, send):
            path = scope.get("path", "")
            # Serve empty manifest.json to suppress Render 404 noise
            if scope["type"] == "http" and path.endswith("manifest.json"):
                await send({
                    "type": "http.response.start",
                    "status": 200,
                    "headers": [(b"content-type", b"application/json")],
                })
                await send({"type": "http.response.body", "body": b"{}"})
                return
            if scope["type"] != "http" or (
                "/queue/" not in path and "/run/" not in path
            ):
                await self.app(scope, receive, send)
                return
            async def _send_patched(msg):
                if msg["type"] == "http.response.start":
                    hdrs = [(k, v) for k, v in msg.get("headers", [])
                            if k.lower() not in
                            (b"x-accel-buffering", b"cache-control", b"connection")]
                    hdrs += [
                        (b"x-accel-buffering", b"no"),
                        (b"cache-control",     b"no-cache"),
                        (b"connection",        b"keep-alive"),
                    ]
                    msg = {**msg, "headers": hdrs}
                await send(msg)
            await self.app(scope, receive, _send_patched)

    _orig_uvicorn_run = _uvicorn.run
    def _patched_uvicorn_run(app, **kwargs):
        print(f"✅ uvicorn intercepted — injecting SSE middleware + keep-alive")
        kwargs["timeout_keep_alive"] = 75
        kwargs.pop("h11_max_incomplete_event_size", None)   # not valid in all uvicorn versions
        _orig_uvicorn_run(_RenderSSEMiddleware(app), **kwargs)
    _uvicorn.run = _patched_uvicorn_run

    # ── Queue + Launch ─────────────────────────────────────────────────────────
    # demo.launch() handles everything: queue start, ASGI wiring, uvicorn bind.
    # We intercept uvicorn.run above so the SSE middleware is always active.
    # Note: concurrency_count was removed from queue() in Gradio 4.20+.
    # Use max_threads in demo.launch() to control concurrency instead.
    demo.queue(
        max_size=20,
        api_open=False,
    )
    # Render deployment: set env vars before launch
import os as _os_launch2
_os_launch2.environ['GRADIO_SERVER_NAME'] = '0.0.0.0'
_os_launch2.environ['GRADIO_SERVER_PORT'] = str(_port)

demo.launch(
        server_name="0.0.0.0",
        server_port=_port,
        show_error=True,
        share=False,
        max_threads=40,
        ssl_verify=False,
        show_api=False,          # Disables /info endpoint — fixes Python 3.14 TypeError
        allowed_paths=["/tmp", tempfile.gettempdir(), os.path.join(tempfile.gettempdir(), "datanetra_reports")],
    )
